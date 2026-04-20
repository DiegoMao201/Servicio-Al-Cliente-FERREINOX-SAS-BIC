import base64
import io
import os
import sys
import unittest
from unittest import mock

from openpyxl import load_workbook


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
BACKEND_DIR = os.path.join(ROOT_DIR, "backend")

sys.path.insert(0, BACKEND_DIR)

os.environ.setdefault("DATABASE_URL", "postgresql://postgres:x@localhost:5432/test")
os.environ.setdefault("OPENAI_API_KEY", "sk-test")

import agent_profiles
import agent_context
import agent_v3
import internal_agent_ops
import main


class InternalAgentProfileTests(unittest.TestCase):
    def test_internal_profile_filters_tools_and_disables_transactional_flows(self):
        with mock.patch.dict(os.environ, {"AGENT_PROFILE": "internal"}, clear=False):
            runtime_config = agent_profiles.get_agent_runtime_config()

        tool_names = [tool["function"]["name"] for tool in runtime_config["tools"]]

        self.assertEqual(runtime_config["profile"], "internal")
        self.assertFalse(runtime_config["enable_order_pipeline"])
        self.assertFalse(runtime_config["enable_quote_pipeline"])
        self.assertFalse(runtime_config["enable_iva_guard"])
        self.assertFalse(runtime_config["force_first_advisory_depth_turn"])
        self.assertEqual(
            tool_names,
            [
                "consultar_inventario",
                "consultar_conocimiento_tecnico",
                "consultar_bi_universal",
                "consultar_ventas_internas",
                "consultar_indicadores_internos",
                "enviar_reporte_interno_correo",
                "buscar_documento_tecnico",
            ],
        )

    def test_internal_profile_blocks_transfer_and_logistics_scope(self):
        with mock.patch.dict(os.environ, {"AGENT_PROFILE": "internal"}, clear=False):
            with mock.patch.object(main, "resolve_internal_session", return_value={"id": 7, "role": "operador", "session_expires_at": "2099-01-01T00:00:00Z"}):
                with mock.patch.object(main, "build_internal_auth_context", return_value={"token": "abc", "role": "operador"}):
                    with mock.patch.object(main, "find_employee_record_by_phone", return_value=None):
                        response = main.handle_internal_whatsapp_message(
                            "crear traslado de viniltex a manizales",
                            {"telefono_e164": "+573001112233"},
                            {"internal_auth": {"token": "abc"}},
                        )

        self.assertIsNotNone(response)
        self.assertEqual(response["intent"], "internal_scope_blocked")
        self.assertIn("No gestiona despachos, reclamos internos ni traslados", response["response_text"])

    def test_internal_advisory_context_does_not_force_square_meters_by_default(self):
        context_text = agent_context.build_turn_context(
            conversation_context={},
            recent_messages=[
                {"direction": "inbound", "contenido": "como pinto una fachada"},
            ],
            user_message="tengo humedad y pintura vieja desgastada",
            internal_auth={"role": "administrador"},
            profile_name="Diego",
        )

        self.assertIn("PROFUNDIZACIÓN DIAGNÓSTICA", context_text)
        self.assertNotIn("¿Cuántos m² tiene el área?", context_text)
        self.assertNotIn("Al final pregunta m² y color para calcular cantidades.", context_text)

    def test_internal_prompt_closes_with_technical_handoff_not_quote_push(self):
        with mock.patch.dict(os.environ, {"AGENT_PROFILE": "internal"}, clear=False):
            runtime_config = agent_profiles.get_agent_runtime_config()

        system_prompt = runtime_config["system_prompt"]

        self.assertIn("No crear cotizaciones.", system_prompt)
        self.assertIn("No preguntes m² por defecto en este canal.", system_prompt)
        self.assertIn("Consultar indicadores internos de ventas, proyeccion, cartera e inventario.", system_prompt)
        self.assertIn("Exportar reportes internos por correo con Excel adjunto", system_prompt)
        self.assertIn("interpreta que pide el consolidado de toda la empresa", system_prompt)
        self.assertIn("consultar_bi_universal", system_prompt)
        self.assertIn("clientes con decrecimiento", system_prompt)
        self.assertIn("clientes para reactivar", system_prompt)
        self.assertIn("productos para impulsar", system_prompt)
        self.assertIn("SU propia cartera y SU propio código de vendedor", system_prompt)
        self.assertIn("Si quieres, te conecto con un asesor comercial para cotizar los productos.", system_prompt)

    def test_universal_bi_plan_detects_open_analytics_intent(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Cuales son los 10 productos que debo impulsar este mes y en que clientes",
            None,
            None,
        )

        self.assertEqual(plan["kind"], "indicator")
        self.assertEqual(plan["tipo_consulta"], "productos_a_impulsar")
        self.assertEqual(plan["limite"], 10)

    def test_universal_bi_plan_detects_semantic_growth_analysis(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Dime el crecimiento por linea este mes vs año anterior",
            None,
            None,
        )

        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "crecimiento")
        self.assertEqual(plan["dimension"], "marca")

    def test_universal_bi_plan_detects_cartera_concentration(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Cual es la concentracion de cartera por cliente",
            None,
            None,
        )

        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "concentracion_cartera")
        self.assertEqual(plan["dimension"], "cliente")

    # ── Regresiones exactas de preguntas fallidas en WhatsApp 2026-04-19 ──

    def test_whatsapp_mix_por_producto_pereira(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Muéstrame el mix por producto en Pereira",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "participacion")
        self.assertEqual(plan["dimension"], "producto")

    def test_whatsapp_crecimiento_por_linea_vs_anio(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Dime el crecimiento por línea este mes vs año anterior",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "crecimiento")
        self.assertEqual(plan["dimension"], "marca")

    def test_whatsapp_caida_frecuencia_vendedores(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Qué vendedores muestran mayor caída de frecuencia",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "caida_frecuencia")

    def test_whatsapp_concentracion_cartera_por_cliente(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Cuál es la concentración de cartera por cliente",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "concentracion_cartera")
        self.assertEqual(plan["dimension"], "cliente")

    def test_whatsapp_oportunidades_por_sede(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Dónde están las mayores oportunidades por sede",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "oportunidades_dimension")
        self.assertEqual(plan["dimension"], "tienda")

    def test_whatsapp_lineas_creciendo_cayendo(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Qué líneas vienen creciendo y cuáles cayendo",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "crecimiento")
        self.assertEqual(plan["dimension"], "marca")

    def test_whatsapp_participacion_por_linea(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "Cuál es la participación por línea este mes",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "participacion")
        self.assertEqual(plan["dimension"], "marca")

    def test_whatsapp_mostradores_creciendo_decreciendo(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "los mostradores vienen creciendo o decreciendo",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "semantic")
        self.assertEqual(plan["analysis"], "crecimiento")
        self.assertEqual(plan["dimension"], "vendedor")
        self.assertEqual(plan["channel"], "mostrador")

    def test_whatsapp_mostradores_ventas(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "cómo van las ventas de mostrador este mes",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "sales")
        self.assertEqual(plan["dimension"], "vendedor")
        self.assertEqual(plan["channel"], "mostrador")

    def test_whatsapp_vendedores_no_channel(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "top 5 vendedores este mes",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "sales")
        self.assertEqual(plan["dimension"], "vendedor")
        self.assertIsNone(plan.get("channel"))

    def test_universal_bi_plan_detects_activate_clients_as_reactivation(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "que clientes debe activar el vendedor hugo nelson",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "indicator")
        self.assertEqual(plan["tipo_consulta"], "clientes_a_reactivar")

    def test_universal_bi_plan_detects_monthly_commercial_plan(self):
        plan = internal_agent_ops._infer_universal_bi_plan(
            "donde tenemos oportunidad y que hacer este mes para vender mas",
            None,
            None,
        )
        self.assertEqual(plan["kind"], "indicator")
        self.assertEqual(plan["tipo_consulta"], "plan_comercial_mensual")

    def test_detect_internal_query_intent_flags_sales_and_projection_questions(self):
        self.assertEqual(
            main.detect_internal_query_intent("Las ventas de Pereira cuáles son y qué proyección tiene para cierre de abril"),
            "consulta_bi",
        )
        self.assertEqual(
            main.detect_internal_query_intent("Cuánto lleva en ventas OLAYA"),
            "consulta_bi",
        )
        self.assertEqual(
            main.detect_internal_query_intent("La empresa cuanto lleva en ventas en abril"),
            "consulta_bi",
        )

    def test_detect_internal_query_intent_does_not_hijack_customer_price_question(self):
        self.assertIsNone(main.detect_internal_query_intent("Cuánto vale el viniltex en Pereira"))

    def test_handle_internal_whatsapp_message_requests_internal_login_for_sales_bi_without_session(self):
        with mock.patch.dict(os.environ, {"AGENT_PROFILE": "internal"}, clear=False):
            with mock.patch.object(main, "find_employee_record_by_phone", return_value=None):
                response = main.handle_internal_whatsapp_message(
                    "La empresa cuanto lleva en ventas en abril",
                    {"telefono_e164": "+573001112233"},
                    {},
                )

        self.assertIsNotNone(response)
        self.assertEqual(response["intent"], "internal_auth_required")
        self.assertIn("Para consultas internas primero debes iniciar sesión", response["response_text"])

    def test_handle_internal_whatsapp_message_keeps_bi_queries_for_internal_llm_after_auth(self):
        with mock.patch.dict(os.environ, {"AGENT_PROFILE": "internal"}, clear=False):
            with mock.patch.object(
                main,
                "resolve_internal_session",
                return_value={"id": 7, "role": "administrador", "session_expires_at": "2099-01-01T00:00:00Z"},
            ):
                with mock.patch.object(main, "build_internal_auth_context", return_value={"token": "abc", "role": "administrador"}):
                    with mock.patch.object(main, "find_employee_record_by_phone", return_value=None):
                        response = main.handle_internal_whatsapp_message(
                            "Cuánto lleva en ventas Pereira",
                            {"telefono_e164": "+573001112233"},
                            {"internal_auth": {"token": "abc"}},
                        )

        self.assertIsNone(response)

    def test_bicomponent_guard_skips_internal_indicator_turns(self):
        assistant_message = mock.Mock()
        assistant_message.content = (
            "Baja rotación: 20 referencias. - INTERGARD 740 stock 2 en Manizales."
        )
        tool_calls_made = [
            {
                "name": "consultar_indicadores_internos",
                "args": {"tipo_consulta": "inventario_baja_rotacion", "almacen": "157", "limite": 20},
                "result": "Baja rotación: 20 referencias y $49,676,072 comprometidos en inventario.",
            }
        ]

        guarded = agent_v3._guardia_bicomponente(
            assistant_message,
            messages=[],
            tool_calls_made=tool_calls_made,
            context={"conversation_id": 148},
            conversation_context={"internal_auth": {"role": "administrador"}},
            m=mock.Mock(),
        )

        self.assertIs(guarded, assistant_message)

    def test_handle_consultar_indicadores_internos_supports_vendedor_nombre(self):
        rows = [
            {
                "cod_cliente": "C001",
                "nombre_cliente": "Pinturas Acme",
                "nom_vendedor": "Hugo Nelson Zapata",
                "ventas_historicas": 1800000,
                "ultima_compra": "2026-03-01",
                "meses_activos": 4,
                "dias_sin_compra": 50,
            }
        ]
        with mock.patch.object(internal_agent_ops, "_resolve_vendor_by_name", return_value="154011") as resolver:
            with mock.patch.object(internal_agent_ops, "_fetch_clients_without_purchase_rows", return_value=(rows, "este mes")) as fetch_rows:
                response = internal_agent_ops.handle_consultar_indicadores_internos(
                    mock.Mock(),
                    {
                        "tipo_consulta": "clientes_a_reactivar",
                        "vendedor_nombre": "Hugo Nelson Zapata",
                        "periodo": "este mes",
                        "limite": 10,
                    },
                    {"internal_auth": {"user_id": 1, "role": "administrador", "employee_context": {}}},
                )

        resolver.assert_called_once()
        fetch_rows.assert_called_once_with(mock.ANY, "este mes", None, "154011", 10)
        self.assertIn("Clientes a reactivar en este mes", response)
        self.assertIn("Pinturas Acme", response)

    def test_handle_consultar_indicadores_internos_resolves_vendedor_name_from_vendedor_codigo_field(self):
        rows = [
            {
                "cod_cliente": "C001",
                "nombre_cliente": "Pinturas Acme",
                "nom_vendedor": "Jerson Ramirez",
                "ventas_historicas": 1800000,
                "ultima_compra": "2026-03-01",
                "meses_activos": 4,
                "dias_sin_compra": 50,
            }
        ]
        with mock.patch.object(internal_agent_ops, "_resolve_vendor_by_name", return_value="154011") as resolver:
            with mock.patch.object(internal_agent_ops, "_fetch_clients_without_purchase_rows", return_value=(rows, "abril 2026")) as fetch_rows:
                response = internal_agent_ops.handle_consultar_indicadores_internos(
                    mock.Mock(),
                    {
                        "tipo_consulta": "clientes_a_reactivar",
                        "vendedor_codigo": "Jerson",
                        "periodo": "abril",
                        "limite": 10,
                    },
                    {"internal_auth": {"user_id": 1, "role": "administrador", "employee_context": {}}},
                )

        resolver.assert_called_once_with(mock.ANY, "Jerson")
        fetch_rows.assert_called_once_with(mock.ANY, "abril", None, "154011", 10)
        self.assertIn("Clientes a reactivar en abril 2026", response)
        self.assertIn("Pinturas Acme", response)

    def test_handle_consultar_bi_universal_resolves_vendor_name_from_args_before_query(self):
        with mock.patch.object(internal_agent_ops, "_resolve_vendor_by_name", return_value="154011") as resolver:
            with mock.patch.object(internal_agent_ops, "handle_consultar_indicadores_internos", return_value="ok") as indicator_handler:
                response = internal_agent_ops.handle_consultar_bi_universal(
                    mock.Mock(),
                    {
                        "pregunta": "que clientes debe activar jerson este mes de abril",
                        "periodo": "abril",
                        "vendedor_codigo": "Jerson",
                    },
                    {"internal_auth": {"user_id": 1, "role": "administrador", "employee_context": {}}},
                )

        resolver.assert_called_once_with(mock.ANY, "Jerson")
        indicator_handler.assert_called_once()
        forwarded_args = indicator_handler.call_args.args[1]
        self.assertEqual(forwarded_args["tipo_consulta"], "clientes_a_reactivar")
        self.assertEqual(forwarded_args["vendedor_codigo"], "154011")
        self.assertEqual(response, "ok")

    def test_handle_consultar_indicadores_internos_builds_monthly_commercial_plan(self):
        snapshot = {
            "period_label": "este mes",
            "neto": 12000000,
            "prev_neto": 15000000,
        }
        reactivation_rows = [
            {
                "cod_cliente": "C001",
                "nombre_cliente": "Pinturas Acme",
                "ventas_historicas": 1800000,
                "dias_sin_compra": 50,
            }
        ]
        decline_rows = [
            {
                "cod_cliente": "C002",
                "nombre_cliente": "Ferrecliente SAS",
                "ventas_actuales": 500000,
                "ventas_previas": 1800000,
                "variacion_absoluta": -1300000,
            }
        ]
        push_rows = [
            {
                "referencia": "1001",
                "descripcion": "VINILTEX 1501",
                "brecha_oportunidad": 900000,
                "clientes_objetivo": "Pinturas Acme, Ferrecliente SAS",
            }
        ]

        with mock.patch.object(internal_agent_ops, "_fetch_sales_total_snapshot", return_value=snapshot):
            with mock.patch.object(internal_agent_ops, "_fetch_clients_without_purchase_rows", return_value=(reactivation_rows, "este mes")):
                with mock.patch.object(internal_agent_ops, "_fetch_client_decline_rows", return_value=(decline_rows, "este mes")):
                    with mock.patch.object(internal_agent_ops, "_fetch_products_to_push_rows", return_value=(push_rows, "este mes")):
                        response = internal_agent_ops.handle_consultar_indicadores_internos(
                            mock.Mock(),
                            {
                                "tipo_consulta": "plan_comercial_mensual",
                                "periodo": "este mes",
                                "limite": 10,
                            },
                            {"internal_auth": {"user_id": 1, "role": "administrador", "employee_context": {}}},
                        )

        self.assertIn("Plan comercial de este mes", response)
        self.assertIn("Pinturas Acme", response)
        self.assertIn("Ferrecliente SAS", response)
        self.assertIn("VINILTEX 1501", response)

    def test_handle_consultar_indicadores_internos_vendedor_defaults_to_own_vendor_scope(self):
        rows = [
            {
                "cod_cliente": "C001",
                "nombre_cliente": "Pinturas Acme",
                "nom_vendedor": "Jerson Atehortua Olarte",
                "ventas_historicas": 1800000,
                "ultima_compra": "2026-03-01",
                "meses_activos": 4,
                "dias_sin_compra": 50,
            }
        ]
        with mock.patch.object(internal_agent_ops, "_fetch_clients_without_purchase_rows", return_value=(rows, "este mes")) as fetch_rows:
            response = internal_agent_ops.handle_consultar_indicadores_internos(
                mock.Mock(),
                {
                    "tipo_consulta": "clientes_a_reactivar",
                    "periodo": "este mes",
                    "limite": 10,
                },
                {
                    "internal_auth": {
                        "user_id": 1,
                        "role": "vendedor",
                        "employee_context": {"codigo_vendedor": "154011", "store_code": "189"},
                    }
                },
            )

        fetch_rows.assert_called_once_with(mock.ANY, "este mes", None, "154011", 10)
        self.assertIn("Clientes a reactivar en este mes", response)
        self.assertIn("Pinturas Acme", response)

    def test_handle_consultar_indicadores_internos_vendedor_projection_uses_vendor_scope(self):
        projection = {
            "ventas_mes_actual": 2500000,
            "dias_transcurridos": 20,
            "proyeccion_cierre_mes": 3750000,
            "variacion_pct": 12.5,
        }
        with mock.patch.object(internal_agent_ops, "_fetch_sales_projection", return_value=projection) as fetch_projection:
            response = internal_agent_ops.handle_consultar_indicadores_internos(
                mock.Mock(),
                {"tipo_consulta": "proyeccion_ventas_mes", "periodo": "este mes"},
                {
                    "internal_auth": {
                        "user_id": 1,
                        "role": "vendedor",
                        "employee_context": {"codigo_vendedor": "154011", "store_code": "189"},
                    }
                },
            )

        fetch_projection.assert_called_once_with(mock.ANY, None, "154011")
        self.assertIn("vendedor 154011", response)

    def test_handle_consultar_indicadores_internos_vendedor_cartera_uses_vendor_scope(self):
        rows = [
            {
                "cod_cliente": "C001",
                "nombre_cliente": "Pinturas Acme",
                "nom_vendedor": "Jerson Atehortua Olarte",
                "balance_total": 900000,
                "balance_31_60": 100000,
                "balance_61_90": 200000,
                "balance_91_plus": 300000,
                "max_dias_vencido": 95,
                "zona": "Pereira",
            }
        ]
        with mock.patch.object(internal_agent_ops, "_fetch_cartera_rows", return_value=rows) as fetch_cartera:
            response = internal_agent_ops.handle_consultar_indicadores_internos(
                mock.Mock(),
                {"tipo_consulta": "cartera_vencida_resumen", "limite": 10},
                {
                    "internal_auth": {
                        "user_id": 1,
                        "role": "vendedor",
                        "employee_context": {"codigo_vendedor": "154011", "store_code": "189"},
                    }
                },
            )

        fetch_cartera.assert_called_once_with(mock.ANY, 10, "154011")
        self.assertIn("Pinturas Acme", response)

    def test_handle_consultar_indicadores_internos_commercial_cargo_ignores_cedula_as_vendor_code(self):
        rows = [
            {
                "cod_cliente": "C001",
                "nombre_cliente": "Pinturas Acme",
                "nom_vendedor": "Jerson Atehortua Olarte",
                "ventas_historicas": 1800000,
                "ultima_compra": "2026-03-01",
                "meses_activos": 4,
                "dias_sin_compra": 50,
            }
        ]
        with mock.patch.object(internal_agent_ops, "_fetch_clients_without_purchase_rows", return_value=(rows, "este mes")) as fetch_rows:
            response = internal_agent_ops.handle_consultar_indicadores_internos(
                mock.Mock(),
                {
                    "tipo_consulta": "clientes_sin_compra_periodo",
                    "vendedor_codigo": "1193084625",
                    "periodo": "este mes",
                    "limite": 10,
                },
                {
                    "internal_auth": {
                        "user_id": 1,
                        "role": "empleado",
                        "employee_context": {
                            "cargo": "Asesor Comercial Externo",
                            "cedula": "1193084625",
                            "codigo_vendedor": "154011",
                            "store_code": "189",
                        },
                    }
                },
            )

        fetch_rows.assert_called_once_with(mock.ANY, "este mes", None, "154011", 10)
        self.assertIn("Clientes sin compra en este mes", response)
        self.assertIn("Pinturas Acme", response)

    def test_handle_consultar_bi_universal_commercial_cargo_uses_own_vendor_scope(self):
        with mock.patch.object(internal_agent_ops, "handle_consultar_indicadores_internos", return_value="ok") as indicator_handler:
            response = internal_agent_ops.handle_consultar_bi_universal(
                mock.Mock(),
                {
                    "pregunta": "Que clientes tengo que visitar este mes ?",
                    "vendedor_codigo": "1193084625",
                },
                {
                    "internal_auth": {
                        "user_id": 1,
                        "role": "empleado",
                        "employee_context": {
                            "cargo": "Asesor Comercial Externo",
                            "cedula": "1193084625",
                            "codigo_vendedor": "154011",
                            "store_code": "189",
                        },
                    }
                },
            )

        indicator_handler.assert_called_once()
        forwarded_args = indicator_handler.call_args.args[1]
        self.assertEqual(forwarded_args["tipo_consulta"], "clientes_a_reactivar")
        self.assertEqual(forwarded_args["vendedor_codigo"], "154011")
        self.assertEqual(response, "ok")

    def test_handle_enviar_reporte_interno_correo_builds_executive_monthly_plan_excel(self):
        snapshot = {
            "period_label": "abril 2026",
            "neto": 12000000,
            "prev_neto": 15000000,
        }
        reactivation_rows = [
            {
                "cod_cliente": "C001",
                "nombre_cliente": "Pinturas Acme",
                "nom_vendedor": "Hugo Nelson Zapata",
                "ventas_historicas": 1800000,
                "ultima_compra": "2026-03-01",
                "meses_activos": 4,
                "dias_sin_compra": 50,
            }
        ]
        decline_rows = [
            {
                "cod_cliente": "C002",
                "nombre_cliente": "Ferrecliente SAS",
                "ventas_actuales": 500000,
                "ventas_previas": 1800000,
                "variacion_absoluta": -1300000,
                "variacion_pct": -0.72,
            }
        ]
        push_rows = [
            {
                "almacen_nombre": "Pereira",
                "referencia": "1001",
                "descripcion": "VINILTEX 1501",
                "stock_total": 20,
                "ventas_actuales": 100000,
                "promedio_base": 1000000,
                "brecha_oportunidad": 900000,
                "clientes_objetivo": "Pinturas Acme, Ferrecliente SAS",
            }
        ]
        send_email = mock.Mock()

        with mock.patch.object(internal_agent_ops, "_fetch_sales_total_snapshot", return_value=snapshot):
            with mock.patch.object(internal_agent_ops, "_fetch_clients_without_purchase_rows", return_value=(reactivation_rows, "abril 2026")):
                with mock.patch.object(internal_agent_ops, "_fetch_client_decline_rows", return_value=(decline_rows, "abril 2026")):
                    with mock.patch.object(internal_agent_ops, "_fetch_products_to_push_rows", return_value=(push_rows, "abril 2026")):
                        response = internal_agent_ops.handle_enviar_reporte_interno_correo(
                            mock.Mock(),
                            {
                                "tipo_reporte": "plan_comercial_mensual",
                                "periodo": "abril 2026",
                                "email_destino": "gerencia@ferreinox.com",
                                "limite": 10,
                            },
                            {
                                "internal_auth": {
                                    "user_id": 1,
                                    "role": "administrador",
                                    "email": "gerencia@ferreinox.com",
                                    "username": "Diego",
                                    "employee_context": {
                                        "full_name": "Diego Garcia",
                                        "cargo": "Gerencia Comercial",
                                        "sede": "Pereira",
                                    },
                                }
                            },
                            lambda title, body: body,
                            send_email,
                        )

        self.assertIn("Reporte enviado a gerencia@ferreinox.com", response)
        attachments = send_email.call_args.kwargs["attachments"]
        workbook_bytes = base64.b64decode(attachments[0]["content"])
        workbook = load_workbook(io.BytesIO(workbook_bytes))

        self.assertIn("Resumen", workbook.sheetnames)
        self.assertIn("Tablero Ejecutivo", workbook.sheetnames)
        self.assertIn("Acciones mes", workbook.sheetnames)
        self.assertIn("Clientes reactivar", workbook.sheetnames)
        self.assertIn("Clientes caida", workbook.sheetnames)
        self.assertIn("Productos impulso", workbook.sheetnames)
        self.assertGreaterEqual(len(workbook["Tablero Ejecutivo"]._charts), 2)
        self.assertEqual(workbook["Acciones mes"]["A2"].value, "Reactivación")


if __name__ == "__main__":
    unittest.main()