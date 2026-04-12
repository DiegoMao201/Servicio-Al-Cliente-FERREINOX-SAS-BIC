import os
import json
import re
import time
import base64
import tomllib
import unicodedata
import io
import uuid
import hmac
import hashlib
import secrets
import asyncio
import threading
from difflib import SequenceMatcher
from datetime import date, timedelta, datetime
from html import escape
from pathlib import Path
from typing import Optional
import logging

import dropbox
import pandas as pd
import requests
from fastapi import Body, FastAPI, Header, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from openai import OpenAI
from openpyxl import Workbook
from pydantic import BaseModel, Field
from sqlalchemy import create_engine, text


logger = logging.getLogger("ferreinox_agent")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(levelname)s: %(message)s")

app = FastAPI(title="CRM Ferreinox Backend", version="2026.3")

# ── Agent V3 (production engine) ──────────────────────────────────────────────
try:
    from agent_v3 import generate_agent_reply_v3
except ImportError:
    from backend.agent_v3 import generate_agent_reply_v3

# ── Color formulas data (from LIBRO DE FORMULAS) ──
_COLOR_FORMULAS: list[dict] = []
_COLOR_FORMULAS_FILE = Path(__file__).resolve().parent.parent / "data" / "color_formulas.json"
if _COLOR_FORMULAS_FILE.exists():
    try:
        _COLOR_FORMULAS = json.loads(_COLOR_FORMULAS_FILE.read_text(encoding="utf-8"))
        logger.info("Loaded %d color formulas from %s", len(_COLOR_FORMULAS), _COLOR_FORMULAS_FILE)
    except Exception as e:
        logger.warning("Failed to load color formulas: %s", e)

# ── International / AkzoNobel product reference (from INTERNATIONAL CODIGOS) ──
_INTERNATIONAL_PRODUCTS: list[dict] = []
_INTERNATIONAL_PRODUCTS_BY_CODE: dict[str, dict] = {}  # code → entry for price enrichment
_INTERNATIONAL_PRODUCTS_FILE = Path(__file__).resolve().parent.parent / "data" / "international_products.json"
if _INTERNATIONAL_PRODUCTS_FILE.exists():
    try:
        _INTERNATIONAL_PRODUCTS = json.loads(_INTERNATIONAL_PRODUCTS_FILE.read_text(encoding="utf-8"))
        logger.info("Loaded %d international product refs from %s", len(_INTERNATIONAL_PRODUCTS), _INTERNATIONAL_PRODUCTS_FILE)
        # Build code→entry index for all product codes (base galón, cat galón, base cuñete, cat cuñete)
        for _ip_entry in _INTERNATIONAL_PRODUCTS:
            for _code_key in ("codigo_base_galon", "codigo_cat_galon", "codigo_cunete",
                              "codigo_cat_cunete", "codigo_galon"):
                _code_val = (_ip_entry.get(_code_key) or "").strip()
                if _code_val:
                    _INTERNATIONAL_PRODUCTS_BY_CODE[_code_val] = _ip_entry
    except Exception as e:
        logger.warning("Failed to load international products: %s", e)


INTERNAL_ROLES = {"empleado", "vendedor", "gerente", "operador", "administrador"}
INTERNAL_SCOPE_TYPES = {"cliente", "vendedor_codigo", "vendedor_nombre", "zona", "almacen"}
INTERNAL_SESSION_TTL_HOURS = int(os.getenv("INTERNAL_AUTH_SESSION_TTL_HOURS", "12"))
INTERNAL_PASSWORD_ITERATIONS = int(os.getenv("INTERNAL_AUTH_PASSWORD_ITERATIONS", "390000"))
INTERNAL_LOGIN_PATTERN = re.compile(r"^\s*login\s+([a-z0-9._-]{3,80})\s+(.+?)\s*$", re.IGNORECASE)
INTERNAL_CEDULA_PATTERN = re.compile(r"\b(\d{6,15})\b")
INTERNAL_LOGOUT_PATTERNS = {
    "logout",
    "salir",
    "cerrar sesion",
    "cerrar sesión",
    "salir interno",
    "logout interno",
}
TRANSFER_REQUEST_MUTATING_ROLES = {"gerente", "operador", "administrador"}
ORDER_DISPATCH_ACTIVE_STATUSES = {"pendiente", "en_transito"}
EMPLOYEE_DIRECTORY_PATH = Path(__file__).resolve().parent.parent / "datos_empleados.xlsx"
EMPLOYEE_DIRECTORY_CACHE_TTL_SECONDS = int(os.getenv("EMPLOYEE_DIRECTORY_CACHE_TTL_SECONDS", "300"))
EMPLOYEE_DIRECTORY_CACHE = {"loaded_at": 0.0, "records": []}
DEFAULT_EMPLOYEE_SEDE_STORE_MAP = {
    "cedi": "155",
    "parque olaya": "189",
    "san antonio": "157",
    "san francisco": "156",
    "opalo": "158",
    "laureles": "238",
    "ferrebox": "439",
    "cerritos": "463",
}
DEFAULT_TRANSFER_DESTINATION_EMAILS = {
    "189": "tiendapintucopereira@ferreinox.co",
    "157": "tiendapintucomanizales@ferreinox.co",
    "158": "tiendapintucodosquebradas@ferreinox.co",
    "156": "tiendapintucoarmenia@ferreinox.co",
    "463": "tiendapintucocerritos@ferreinox.co",
    "238": "tiendapintucolaureles@ferreinox.co",
}
DEFAULT_TRANSFER_CC_EMAILS = ["compras@ferreinox.co"]
CORPORATE_BRAND = {
    "company_name": "FERREINOX S.A.S. BIC",
    "nit": "800.224.617-8",
    "address": "CR 13 19-26, Pereira, Risaralda, Colombia",
    "website": "https://www.ferreinox.co",
    "service_email": "hola@ferreinox.co",
    "pqrs_email": "contacto@ferreinox.co",
    "phone_landline": "(606) 333 0101",
    "phone_mobile": "+57 310 830 5302",
    "phone_mobile_alt": "+57 323 232 8249",
    "brand_dark": "#111827",
    "brand_accent": "#F59E0B",
    "brand_light": "#F9FAFB",
    "brand_border": "#E5E7EB",
}
CORPORATE_LOGO_PATH = Path(__file__).resolve().parent.parent / "LOGO FERREINOX SAS BIC 2024.png"
DEFAULT_FACTURADOR_ROUTING = {
    "155": {"name": "Paula", "phone": "+573102368346"},
    "156": {"name": "Jaime", "phone": "+573165219904"},
    "157": {"name": "Lorena", "phone": "+573136086232"},
    "158": {"name": "M. Paula", "phone": "+573108561506"},
    "189": {"name": "Paula", "phone": "+573102368346"},
    "439": {"name": "Manuel", "phone": "+573209559031"},
    "463": {"name": "Jose Aurelio", "phone": "+573104739586"},
}


class InternalUserScopeInput(BaseModel):
    scope_type: str
    scope_value: str
    scope_label: Optional[str] = None


class InternalBootstrapUserRequest(BaseModel):
    username: str
    password: str
    role: str
    full_name: str
    phone_number: Optional[str] = None
    email: Optional[str] = None
    scopes: list[InternalUserScopeInput] = Field(default_factory=list)


class InternalLoginRequest(BaseModel):
    username: str
    password: str


SECRETS_PATH = Path(__file__).resolve().parent.parent / ".streamlit" / "secrets.toml"
ARTIFACTS_PATH = Path(__file__).resolve().parent.parent / "artifacts"
PRODUCT_NLU_PROMPT_PATH = ARTIFACTS_PATH / "SYSTEM_PROMPT_NLU_EXTRACCION_PRODUCTO.md"
PRODUCT_NLU_PROMPT_FALLBACK = """Eres un extractor NLU para pedidos ferreteros. Devuelve solo JSON válido con las claves cantidad_inferida, presentacion_canonica_inferida, producto_base, color y acabado. Si no sabes un valor, devuelve null. Interpreta 1/5 como cuñete, 1/1 como galon y 1/4 como cuarto. Conserva colores compuestos como verde bronce y no inventes acabados."""
PRODUCT_NLU_PROMPT_CACHE: Optional[str] = None
PRODUCT_NLU_CACHE_MAX_ITEMS = 256
PRODUCT_NLU_CACHE: dict[str, dict] = {}
TECHNICAL_DOC_FOLDER = "/data/FICHAS TÉCNICAS Y HOJAS DE SEGURIDAD"
TECHNICAL_DOC_CACHE_TTL_SECONDS = 600
TECHNICAL_DOC_STOPWORDS = {
    "ficha",
    "fichas",
    "tecnica",
    "tecnicas",
    "tecnico",
    "tecnico",
    "hoja",
    "hojas",
    "seguridad",
    "pdf",
    "envia",
    "enviame",
    "enviamelo",
    "manda",
    "mandame",
    "mandamelo",
    "adjunta",
    "adjuntame",
    "anexa",
    "anexame",
    "sirve",
    "sirva",
    "puedes",
    "puede",
    "podrias",
    "podria",
    "quiero",
    "quieres",
    "necesito",
    "enviar",
    "envies",
    "mandar",
    "mandes",
    "archivo",
    "documento",
    "documentacion",
    "tecnica",
    "tecnico",
    "segun",
    "saber",
    "tienes",
    "tiene",
    "tengo",
    "si",
    "es",
    "que",
    "del",
    "de",
    "me",
    "la",
    "el",
}
TECHNICAL_DOC_CACHE = {"loaded_at": 0.0, "entries": []}


TECHNICAL_ADVISORY_KEYWORDS = [
    "piso",
    "pisos",
    "pintar un piso",
    "pintura para piso",
    "pintura epoxica",
    "pintura epóxica",
    "epoxica",
    "epóxica",
    "cemento",
    "concreto",
    "trafico peatonal",
    "tráfico peatonal",
    "humedad",
    "humedo",
    "húmedo",
    "goteras",
    "gotera",
    "filtracion",
    "filtración",
    "capilaridad",
    "moho",
    "hongo",
    "manchas negras",
    "salitre",
    "se cae la pintura",
    "se descascara",
    "descascaramiento",
    "ampollamiento",
    "corrosion",
    "corrosión",
    "oxido",
    "óxido",
    "barniz",
    "laca",
    "madera",
    "metal",
    "hierro",
    "galvanizado",
    "aluminio",
    "como aplicar",
    "cómo aplicar",
    "como aplico",
    "cómo aplico",
    "como se aplica",
    "cómo se aplica",
    "como pinto",
    "cómo pinto",
    "como pintar",
    "cómo pintar",
    "que rodillo",
    "qué rodillo",
    "que brocha",
    "qué brocha",
    "tiempo de secado",
    "cuanto seca",
    "cuánto seca",
    "cuanto demora",
    "cuánto demora",
    "se puede mezclar",
    "se puede combinar",
    "rendimiento",
    "cuanto rinde",
    "cuánto rinde",
    "manos de pintura",
    "cuantas manos",
    "cuántas manos",
    "preparar la pared",
    "preparar la superficie",
    "lijado",
    "lijar",
    "impermeabilizar",
    "impermeabilizante",
    "diluir",
    "diluyente",
    "que disolvente",
    "qué disolvente",
    "thinner",
    "estuco",
    "estucar",
    "sellar",
    "sellador",
    "para exterior",
    "para interior",
    "anticorrosivo",
    "fondo",
    "imprimante",
    "para madera",
    "para metal",
    "para hierro",
    "para ladrillo",
    "para concreto",
    "diferencia entre",
    "cual es mejor",
    "cuál es mejor",
    "que me recomiendas",
    "qué me recomiendas",
    "que sirve para",
    "qué sirve para",
    "como proteger",
    "cómo proteger",
    "como limpiar",
    "cómo limpiar",
    "techo",
    "fachada",
    "terraza",
    "piscina",
    "tanque",
    "cielo raso",
    "drywall",
    "cubierta",
    "muro",
    "pared",
    "columna",
    "viga",
    "garaje",
    "bodega",
    "parqueadero",
    "sotano",
    "sótano",
    "escalera",
    "barandal",
    "reja",
    "puerta",
    "portón",
    "porton",
    "ventana",
    "tubería",
    "tuberia",
    "cerca",
    "necesito pintar",
    "quiero pintar",
    "voy a pintar",
    "necesito recubrir",
    "quiero proteger",
    "necesito impermeabilizar",
    "que pintura uso",
    "qué pintura uso",
    "que producto necesito",
    "qué producto necesito",
    "que sistema",
    "qué sistema",
    "se pela",
    "se ampolla",
    "se mancha",
    "se deteriora",
    "se agrieta",
    "se fisura",
    "se oxida",
    "pintura vieja",
    "grieta",
    "grietas",
    "fisura",
    "fisuras",
    "me pueden asesorar",
    "asesoria",
    "asesoría",
    "me recomienda",
    "me aconsejas",
    "tengo un problema",
    "necesito solucionar",
    "que le echo",
    "qué le echo",
    "que le aplico",
    "qué le aplico",
    "como soluciono",
    "cómo soluciono",
    "poliuretano",
    "vinilo",
    "acrilico",
    "acrílico",
    "esmalte",
    "latex",
    "látex",
    "epoxico",
    "epóxico",
    "alquidico",
    "alquídico",
    "removedor",
    "removedor de pintura",
    "decapante",
    "quitar pintura",
    "remover pintura",
    "sacar pintura",
    "disco flap",
    "grata",
    "gratas",
    "cepillo metalico",
    "cepillo metálico",
    "cepillo de alambre",
    "abrasivo",
    "abrasivos",
    "con que lijo",
    "con qué lijo",
    "como lijo",
    "cómo lijo",
    "como remuevo",
    "cómo remuevo",
    "tobogan",
    "tobogán",
    "rodadero",
    "resbaladero",
    "juego infantil",
    "parque infantil",
    "pasamanos",
    "baranda",
    "barandas",
]


CLAIM_KEYWORDS = [
    "reclamo",
    "reclamacion",
    "reclamación",
    "garantia",
    "garantía",
    "calidad",
    "no funcion",
    "no funciono",
    "no funcionó",
    "no cubre",
    "no cubrio",
    "no cubrió",
    "defecto",
    "falla",
    "dañado",
    "danado",
    "problema con",
]


QUOTE_KEYWORDS = [
    "cotizacion",
    "cotización",
    "cotizar",
    "presupuesto",
    "propuesta comercial",
]


ORDER_KEYWORDS = [
    "montar pedido",
    "montar un pedido",
    "hacer pedido",
    "hacer un pedido",
    "generar pedido",
    "generar un pedido",
    "realizar pedido",
    "realizar un pedido",
    "orden de compra",
    "confirmar pedido",
    "necesito pedido",
    "necesito un pedido",
    "quiero pedido",
    "quiero un pedido",
    "quiero hacer",
    "necesito hacer",
    "a ser un pedido",
    "aser un pedido",
    "acer un pedido",
    "acer pedido",
    "pedir productos",
    "pedir producto",
    "armar pedido",
    "armar un pedido",
    "pasar pedido",
    "pasar un pedido",
]


NON_PRODUCT_SERVICE_KEYWORDS = [
    "cartera",
    "saldo",
    "deuda",
    "debo",
    "compras",
    "compra",
    "estado de cuenta",
    "factura",
    "facturas",
    "reclamo",
    "garantia",
    "garantía",
    "calidad",
    "cotizacion",
    "cotización",
    "cotizar",
    "pedido",
    "correo",
    "email",
    "ficha tecnica",
    "ficha técnica",
    "hoja de seguridad",
]


PRODUCT_STOPWORDS = {
    "ay",
    "ahi",
    "alli",
    "de",
    "del",
    "la",
    "el",
    "los",
    "las",
    "un",
    "una",
    "unos",
    "unas",
    "para",
    "por",
    "con",
    "sin",
    "que",
    "me",
    "mi",
    "necesito",
    "quiero",
    "cotizar",
    "comprar",
    "compro",
    "compras",
    "hacer",
    "hace",
    "hago",
    "montar",
    "monto",
    "armar",
    "armo",
    "pasar",
    "pasame",
    "pasarme",
    "enviar",
    "enviame",
    "enviarme",
    "mandar",
    "mandame",
    "mandamelo",
    "mandarmelo",
    "pedir",
    "pedido",
    "correo",
    "email",
    "mail",
    "aqui",
    "aca",
    "favor",
    "informacion",
    "información",
    "sobre",
    "tengo",
    "hay",
    "tienen",
    "inventario",
    "stock",
    "en",
    "este",
    "ano",
    "año",
    "cuanto",
    "debo",
    "codigo",
    "cod",
    "ref",
    "refer",
    "es",
    "producto",
    "marca",
    "agregale",
    "agregame",
    "agregalo",
    "agrega",
    "ponle",
    "ponme",
    "ponlo",
    "pon",
    "sumale",
    "sumame",
    "suma",
    "quitale",
    "quitame",
    "quita",
    "otro",
    "otra",
    "otros",
    "otras",
    "optro",
    "optra",
    "nuevo",
    "nueva",
    "nuevos",
    "nuevas",
}


PRESENTATION_ALIASES = {
    "cuñete": ["cunete", "cunetes", "cuenete", "cuenetes", "cuñete", "cuñetes", "caneca", "canecas", "cubeta", "cubetas", "18.93l", "18.93", "1/5", "5gl"],
    "galon": ["galon", "galones", "gal", "3.79l", "3.79", "1/1", "1gl"],
    "cuarto": ["cuarto", "cuartos", "0.95l", "0.95", "1/4"],
}


PRESENTATION_LABELS = {
    "cuñete": ("cuñete", "cuñetes"),
    "galon": ("galón", "galones"),
    "cuarto": ("cuarto", "cuartos"),
}


PRESENTATION_SHORTCUTS = {
    "1": "galon",
    "4": "cuarto",
    "5": "cuñete",
}


PRESENTATION_SIZE_MAP = {
    "18.93": "cuñete",
    "18.93l": "cuñete",
    "3.79": "galon",
    "3.79l": "galon",
    "0.95": "cuarto",
    "0.95l": "cuarto",
}


PORTFOLIO_ALIASES = {
    # ── VINILOS TIPO 1 (Premium) ──
    "viniltex": ["viniltex", "viniltex adv", "viniltex advanced", "vinilico", "vinilo", "vtx", "vinilo premium", "vinilo tipo 1",
                 # Códigos de color Viniltex (cliente dice "1501" para Viniltex Blanco, "1525" para amarillo vivo, etc.)
                 "1501", "viniltex blanco", "viniltex blanco 1501",
                 "1525", "viniltex amarillo", "viniltex amarillo vivo 1525"],
    "vinilico": ["vinilico", "vinilo", "vinilica", "viniloco", "vinilico blanco", "viniltex blanco", "vinilo tipo 1"],
    "viniloco": ["viniloco", "vinilico", "viniltex", "vinilo", "vinilico blanco", "viniltex blanco"],
    "vinil plus": ["vinil plus", "vinilplus", "vinil+", "vinilo plus", "vinilo tipo 1"],
    # ── VINILOS TIPO 2 (Intermedio) ──
    "intervinil": ["intervinil", "inter vinil", "vinilo tipo 2", "vinilo intermedio"],
    "vinil latex": ["vinil latex", "vinil látex", "vinillatex", "vinilo latex", "vinilo tipo 2"],
    "vinilux": ["vinilux", "vinilo tipo 2"],
    # ── VINILOS TIPO 3 (Económico) ──
    "pinturama": ["pinturama", "vinilo tipo 3", "vinilo economico", "vinilo económico"],
    "vinil max": ["vinil max", "vinilmax", "vinil-max", "vinilo tipo 3"],
    "icolatex": ["icolatex", "ico latex", "vinilo tipo 3"],
    # ── ESMALTES ──
    "pintulux": ["pintulux", "pintulux 3en1", "pintulux 3 en 1", "pintulux 3-en-1", "3en1", "3 en 1", "3-en-1", "esmalte pintulux", "esmalte exterior", "esmalte bueno", "esmalte resistente",
                 # Códigos cortos de mostrador (TEU/T-XX → pintulux con ese número de color)
                 "t11", "t-11", "t 11", "teu11", "teu-11", "tu11", "tu-11",
                 "t95", "t-95", "t 95", "teu95", "teu-95", "tu95",
                 "t84", "t-84", "teu84", "t76", "t-76", "teu76",
                 "t80", "t-80", "teu80", "t53", "t-53", "teu53",
                 "t10", "t-10", "teu10", "t89", "t-89", "teu89",
                 "t18", "t-18", "teu18", "t20", "t-20", "teu20",
                 "t26", "t-26", "teu26", "t40", "t-40", "teu40"],
    "domestico": ["domestico", "doméstico", "vinilico", "economico", "económico", "esmalte domestico", "esmalte interior", "esmalte economico", "esmalte económico",
                  # Códigos P-XX de Esmalte Doméstico Pintuco (P-18=amarillo, P-35=azul francés, P-40=azul español, P-50=azul verano, P-153=aluminio)
                  "p18", "p-18", "p35", "p-35", "p40", "p-40", "p50", "p-50", "p153", "p-153"],
    # ── CÓDIGOS CORTOS DE MOSTRADOR ──
    "pintuco": ["pintuco", "viniltex", "p11", "p-11", "p 11"],
    "p11": ["p11", "p-11", "p 11", "pintuco 11", "domestico blanco"],
    "t11": ["t11", "t-11", "t 11", "pintulux 3en1", "pintulux 3 en 1", "3en1 br blanco 11", "br blanco 11"],
    "p53": ["p53", "p-53", "p 53", "verde esmeral", "verde esmer"],
    # ── LÍNEAS ESPECIALIZADAS PINTUCO ──
    "pintucoat": ["pintucoat", "epoxica pintuco", "epoxica", "epóxica", "epóxica pintuco", "epoxy", "epoxi pintuco", "recubrimiento epoxica"],
    "pintura canchas": ["pintura canchas", "pintura para canchas", "pintura de cancha", "canchas"],
    "corrotec": ["corrotec", "anticorrosivo pintuco", "anticorrosivo", "anti corrosivo", "corrotec premium"],
    "pintoxido": ["pintoxido", "desoxidante", "convertidor oxido", "convertidor óxido", "convertidor de oxido"],
    # ── TERINSA ── (anticorrosivos y lacas — códigos de mostrador: 05232=blanco, 05064=gris, 05185=rojo, 05079=verde)
    "terinsa": ["terinsa", "anticorrosivo terinsa", "laca terinsa",
                "05232", "05064", "05185", "05079", "57016", "57048", "57068", "57260", "50232", "50234", "57253"],
    "pintacrom": ["pintacrom", "pinta crom", "anticorrosivo aerosol"],
    "pintulac": ["pintulac", "laca pintuco", "laca", "laca madera"],
    "aerocolor": ["aerocolor", "aerosol pintuco", "aerosol", "spray pintuco", "spray pintura", "pintura spray", "pintura aerosol", "spray"],
    "koraza": ["koraza", "impermeabilizante koraza", "koraza elastomerica", "koraza elastomerico", "pintura fachada", "pintura exterior fachada"],
    "world color": ["world color", "worldcolor", "tinte pintuco", "base tintometrica", "base tintométrica"],
    "imprimante": ["imprimante", "imprimante pintuco", "primer", "fondo", "sellador fondo", "primer pintuco"],
    "pintuco fill": ["pintuco fill", "impermeabilizante pintuco", "fill 7", "fill 12", "pintuco fill 7", "pintuco fill 12"],
    "pintura trafico": ["pintura trafico", "pintura tráfico", "pintutraf", "pintutrafico", "pintura demarcacion", "pintura demarcación", "demarcacion vial", "pintura vial", "trafico", "señalizacion vial"],
    "pintutraf": ["pintutraf", "pintutrafico", "pintura trafico", "pintura tráfico", "pintura demarcacion", "trafico", "señalizacion vial"],
    "microesfera": ["microesfera", "microesferas", "micro esfera", "esferas reflectivas", "esferas vidrio", "esferas trafico"],
    "wash primer": ["wash primer", "washprimer", "primer anticorrosivo", "fondo anticorrosivo"],

    # ── INTERNATIONAL / AKZONOBEL ──
    "interseal": ["interseal", "epoxica international", "epóxica international", "epoxica akzonobel"],
    "interthane": ["interthane", "poliuretano international", "poliuretano akzonobel", "poliuretano"],
    "intergard": ["intergard", "epoxica intergard", "primer intergard"],
    "interfine": ["interfine", "acabado international", "acabado interthane"],
    "interchar": ["interchar", "intumescente", "pintura intumescente", "ignifuga", "ignífuga", "ignifugo"],
    # ── IMPERMEABILIZANTES / SELLADORES ──
    "sika": ["sika", "sika 1", "sika-1", "sikamur", "impermeabilizante sika"],
    "aquablock": ["aquablock", "aqua block", "aquablock ultra", "sello humedad", "sellador humedad", "impermeabilizante humedad", "bloqueador humedad"],
    "sellamur": ["sellamur", "sella mur", "sellador muro", "sellador mur"],
    "siliconite": ["siliconite", "silicon", "silicona impermeabilizante"],
    # ── CERRADURAS / FERRETERÍA ──
    "mega": ["mega", "cerradura mega", "sobreponer"],
    "cerradura": ["cerradura", "cerradur", "chapa", "lock"],
    "derecha": ["derecha", "derecho", "der", "derc"],
    "izquierda": ["izquierda", "izquierdo", "izq"],
    "brocha": ["brocha", "brochas", "pincel", "brochas pintuco"],
    "popular": ["popular", "pop"],
    "goya popular": ["goya popular", "brocha popular", "popular goya", "brocha popu goya", "goya popu"],
    "goya profesional": ["goya profesional", "brocha profesional", "profesional goya", "brocha profe goya", "brocha prof goya", "goya profe", "goya prof"],
    "rodillo": ["rodillo", "rodillos", "felpa", "mini rodillo"],
    "lija": ["lija", "lijas", "papel lija", "lija agua", "lija al agua"],
    # ── ADHESIVOS / PEGANTES (Abracol catalog) ──
    "abracol": ["abracol", "lija abracol", "disco abracol", "fibrodisco abracol"],
    "yale": ["yale", "cerradura yale", "candado yale", "manija yale", "antipanico yale", "barra antipanico"],
    "goya": ["goya", "brocha goya", "rodillo goya", "bandeja goya", "goya popular", "goya profesional"],
    "smith": ["smith", "cinta smith", "tirro smith"],
    "afix": ["afix", "silicona afix", "espuma afix", "epoxi afix"],
    "segurex": ["segurex", "cerradura segurex", "candado segurex", "manija segurex", "cerrojo segurex"],
    "artecola": ["artecola", "pegante artecola", "adhesivo artecola", "cola artecola"],
    "tekbond": ["tekbond", "aerosol tekbond", "adhesivo tekbond", "silicona tekbond", "cinta tekbond", "espuma tekbond"],
    "induma": ["induma", "bisagra induma", "pasador induma", "balde induma", "herraje induma"],
    "norton": ["norton", "disco norton", "lija norton", "disco corte norton", "disco desbaste norton"],
    "carborundum": ["carborundum", "beartex", "disco carborundum", "lija carborundum"],
    "phillips": ["phillips", "cerradura phillips", "manija phillips", "cerrojo phillips"],
    "inafer": ["inafer", "bisagra inafer", "herraje inafer"],
    "delta": ["delta", "pegante delta", "adhesivo delta", "madera delta"],
    # ── CERRADURAS / SEGURIDAD (extendido Abracol catalog) ──
    "candado": ["candado", "candado yale", "candado segurex", "candado dorado"],
    "cerrojo": ["cerrojo", "cerrojo yale", "cerrojo segurex", "cerrojo doble"],
    "manija": ["manija", "manija yale", "manija phillips", "manija segurex", "manija puerta"],
    "antipanico": ["antipanico", "barra antipanico", "antipanico yale", "barra de panico"],
    # ── ABRASIVOS (extendido Abracol catalog) ──
    "fibrodisco": ["fibrodisco", "fibrodiscos", "disco fibra", "disco lija"],
    "disco corte": ["disco corte", "disco de corte", "disco corte norton", "disco corte metal"],
    "disco desbaste": ["disco desbaste", "disco de desbaste", "disco desbaste norton"],
    "copa": ["copa", "grata copa", "copa abrasivo"],
    "lija roja": ["lija roja", "lija al seco", "lija seco"],
    "beartex": ["beartex", "beartex norton", "abrasivo no tejido"],
    # ── CINTAS / ADHESIVOS (extendido Abracol catalog) ──
    "cinta enmascarar": ["cinta enmascarar", "cinta masking", "masking tape", "cinta pintor", "tirro"],
    "cinta teflon": ["cinta teflon", "teflon", "cinta teflón"],
    "silicona": ["silicona", "silicona afix", "silicona tekbond", "sellador silicona"],
    "espuma poliuretano": ["espuma poliuretano", "espuma expansiva", "espuma tekbond", "espuma afix"],
    "pl285": ["pl285", "pl 285", "pegante pl285", "pegante madera"],
    "montana": ["montana", "montana 94", "aerosol montana"],
    # ── DISOLVENTES / COMPLEMENTARIOS ──
    "thinner": ["thinner", "tinner", "diluyente", "disolvente", "solvente"],
    "varsol": ["varsol", "disolvente varsol"],
    "aguarras": ["aguarras", "aguarrás", "trementina"],
    "estuco": ["estuco", "masilla", "estuco plastico", "estuco plástico"],
    # ── BARNICES / LACAS / SD ──
    "barniz sd-1": ["barniz sd-1", "barniz sd1", "sd-1", "sd1", "barniz incoloro sd", "barniz br incoloro sd"],
    "barniz sd-2": ["barniz sd-2", "barniz sd2", "sd-2", "sd2"],
    "barniz sd-3": ["barniz sd-3", "barniz sd3", "sd-3", "sd3"],
    # ── GENÉRICOS QUE DEBEN EXPANDIR ──
    "vinilo": ["vinilo", "vinilico", "viniltex", "domestico", "intervinil", "vinil latex", "pinturama"],
    "esmalte": ["esmalte", "pintulux", "domestico", "esmalte sintetico", "esmalte sintético"],
    "pintura": ["pintura", "viniltex", "pintulux", "koraza", "domestico"],
}

# ── Taxonomía completa del portafolio Ferreinox ───────────────────────────
# Árbol jerárquico de familias, calidades y productos reales.
# Esto lo usa el system prompt para que la IA haga preguntas inteligentes
# y sepa exactamente qué buscar en PostgREST según la necesidad del cliente.
PINTUCO_PRODUCT_TAXONOMY = {
    "vinilos": {
        "descripcion": "Pinturas vinílicas para muros y cielos rasos (base agua)",
        "tipo_1_premium": {
            "descripcion": "Mejor cubrimiento, lavabilidad y rendimiento",
            "marcas": ["Viniltex", "Viniltex Advanced", "Vinil Plus"],
            "uso": "Interior y exterior, paredes donde se necesita durabilidad y lavabilidad",
        },
        "tipo_2_intermedio": {
            "descripcion": "Buen cubrimiento a precio medio",
            "marcas": ["Intervinil", "Vinil Látex", "Vinilux"],
            "uso": "Interior, paredes de tráfico medio, cuartos, bodegas",
        },
        "tipo_3_economico": {
            "descripcion": "Opción económica para obras y alto volumen",
            "marcas": ["Pinturama", "Vinil Max", "Icolatex"],
            "uso": "Interior, cielos rasos, obras masivas, paredes de baja exigencia",
        },
    },
    "esmaltes": {
        "descripcion": "Pinturas a base de solvente para madera, metal y superficies lavables",
        "pintulux_3en1": {
            "descripcion": "Esmalte premium, anticorrosivo, alto brillo o satinado, lavable",
            "marcas": ["Pintulux 3en1"],
            "uso": "Exterior e interior, rejas, puertas, muebles, madera, metal. El mejor esmalte del portafolio.",
        },
        "domestico": {
            "descripcion": "Esmalte económico, buen rendimiento para uso general",
            "marcas": ["Doméstico"],
            "uso": "Interior, puertas, marcos, madera interior. Más económico que Pintulux.",
        },
    },
    "fachada_impermeabilizante": {
        "descripcion": "Pinturas elastoméricas e impermeabilizantes para exteriores",
        "marcas": ["Koraza", "Koraza Elastomérica", "Koraza XP"],
        "uso": "Fachadas, muros exteriores, terrazas, zonas expuestas a lluvia y sol",
    },
    "aerosoles": {
        "descripcion": "Pinturas en spray para retoques, manualidades y aplicaciones rápidas",
        "marcas": ["Aerocolor"],
        "uso": "Retoques, artesanía, metales pequeños, madera, plástico",
    },
    "pisos": {
        "descripcion": "Pinturas especiales para pisos de concreto y cemento",
        "marcas": ["Pintura para Canchas", "Pintucoat", "Intergard 740", "Intergard 2002"],
        "uso": "Pisos de concreto canchas deportivas, senderos peatonales, ciclo rutas (Canchas). Garajes y parqueaderos residenciales (Pintucoat). Pisos industriales de tráfico MEDIO acabado mate (Pintucoat). Pisos industriales acabado brillante resistencia media (Intergard 740). Pisos industriales alta resistencia con cuarzo (Intergard 2002 + Cuarzo ref 5891610).",
    },
    "epoxicas": {
        "descripcion": "Recubrimientos de alto desempeño, dos componentes (resina + catalizador)",
        "marcas": ["Pintucoat", "Interseal", "Intergard", "Intergard 740", "Intergard 2002"],
        "uso": "Pisos industriales tráfico medio acabado mate (Pintucoat), tráfico medio brillante (Intergard 740), tráfico pesado con cuarzo (Intergard 2002 + ref 5891610), ambientes químicos. NO sirven para piscinas ni inmersión en agua.",
    },
    "anticorrosivos": {
        "descripcion": "Fondos y primers protectores contra la oxidación del metal",
        "marcas": ["Corrotec", "Corrotec Premium", "Pintoxido", "Wash Primer", "Intergard"],
        "uso": "Metal oxidado, estructuras metálicas, tuberías, rejas antes de esmalte final",
    },
    "lacas_barnices": {
        "descripcion": "Acabados transparentes o semitransparentes para madera",
        "marcas": ["Barniz Marino", "Barniz SD1"],
        "uso": "Muebles, puertas de madera interior. Para exterior usar Barnex o Wood Stain. NO manejamos Pintulac.",
    },
    "poliuretanos": {
        "descripcion": "Acabados industriales de alta resistencia química y UV",
        "marcas": ["Interthane", "Interfine"],
        "uso": "Acabados finales sobre epóxicas, maquinaria, pisos industriales de alto desempeño",
    },
    "intumescentes": {
        "descripcion": "Pinturas de protección pasiva contra incendios",
        "marcas": ["Interchar"],
        "uso": "Estructuras metálicas que deben cumplir norma de resistencia al fuego",
    },
    "trafico_demarcacion": {
        "descripcion": "Pinturas para señalización vial y demarcación de pisos",
        "marcas": ["Pintura Tráfico"],
        "uso": "Parqueaderos, canchas, bodegas, vías, líneas de seguridad",
    },
}

# ── Mapa de categorías de uso → términos reales de inventario ──────────────
# Cuando el RAG o el usuario pide un producto por su categoría genérica
# (ej. "pintura para piscinas"), este mapa expande a los nombres de marca
# que efectivamente están en el inventario de Ferreinox.
PORTFOLIO_CATEGORY_MAP = {
    # ── Piscinas / tanques de agua ──  (Ferreinox NO maneja pintura para piscinas)
    "piscina": ["__SIN_PRODUCTO_FERREINOX__"],
    "piscinas": ["__SIN_PRODUCTO_FERREINOX__"],
    "tanque": ["__SIN_PRODUCTO_FERREINOX__"],
    "tanque agua": ["__SIN_PRODUCTO_FERREINOX__"],
    "inmersion": ["__SIN_PRODUCTO_FERREINOX__"],
    "inmersion en agua": ["__SIN_PRODUCTO_FERREINOX__"],

    # ── HUMEDAD / FILTRACIONES (muros interiores, sótanos, bases) ──
    "humedad": ["aquablock", "aquablock ultra", "sellamur", "estuco anti humedad", "sika"],
    "humedad interna": ["aquablock", "aquablock ultra", "sellamur", "estuco anti humedad"],
    "humedad freatica": ["aquablock", "aquablock ultra", "sellamur"],
    "humedad ascendente": ["aquablock", "aquablock ultra", "sellamur", "estuco anti humedad"],
    "filtracion": ["aquablock", "aquablock ultra", "sellamur"],
    "filtra agua": ["aquablock", "aquablock ultra", "sellamur"],
    "filtra": ["aquablock", "aquablock ultra", "sellamur"],
    "mojando": ["aquablock", "aquablock ultra", "sellamur"],
    "pared mojada": ["aquablock", "aquablock ultra", "sellamur"],
    "muro mojado": ["aquablock", "aquablock ultra", "sellamur"],
    "salitre": ["aquablock", "aquablock ultra", "estuco anti humedad", "sellamur"],
    "eflorescencia": ["aquablock", "aquablock ultra", "estuco anti humedad"],
    "manchas blancas": ["aquablock", "estuco anti humedad", "sellamur"],
    "pared suda": ["aquablock", "aquablock ultra", "sellamur"],
    "sudando": ["aquablock", "aquablock ultra", "sellamur"],
    "ampollando": ["aquablock", "estuco anti humedad", "sellamur"],
    "ampollas": ["aquablock", "estuco anti humedad", "sellamur"],
    "descascarando interior": ["aquablock", "estuco anti humedad", "sellamur"],
    "presion negativa": ["aquablock", "aquablock ultra"],
    "capilaridad": ["aquablock", "aquablock ultra", "sellamur"],
    "sotano": ["aquablock", "aquablock ultra"],
    "sotano humedo": ["aquablock", "aquablock ultra"],
    "antihumedad": ["aquablock", "aquablock ultra", "sellamur", "estuco anti humedad"],
    "anti humedad": ["aquablock", "aquablock ultra", "sellamur", "estuco anti humedad"],
    "sello humedad": ["aquablock", "aquablock ultra", "sellamur"],
    "bloquear humedad": ["aquablock", "aquablock ultra", "sellamur"],
    "hongos pared": ["aquablock", "aquablock ultra", "viniltex"],
    "moho pared": ["aquablock", "aquablock ultra", "viniltex"],
    "hongos bano": ["aquablock", "viniltex banos y cocinas"],
    "bano humedo": ["aquablock", "viniltex banos y cocinas"],

    # ── FACHADAS / EXTERIORES (muros exteriores, lluvia, sol) ──
    "fachada": ["koraza", "koraza sol y lluvia", "koraza elastomerica"],
    "fachadas": ["koraza", "koraza sol y lluvia", "koraza elastomerica"],
    "frente casa": ["koraza", "koraza sol y lluvia"],
    "exterior casa": ["koraza", "koraza sol y lluvia"],
    "muro exterior": ["koraza", "koraza sol y lluvia"],
    "pared exterior": ["koraza", "koraza sol y lluvia"],
    "pintura exterior": ["koraza", "viniltex", "pintulux"],
    "exterior lluvia": ["koraza", "koraza sol y lluvia"],
    "exterior sol": ["koraza", "koraza sol y lluvia"],
    "aguanta sol": ["koraza", "koraza sol y lluvia"],
    "aguanta lluvia": ["koraza", "koraza sol y lluvia"],
    "pelando exterior": ["koraza", "koraza sol y lluvia"],
    "descascarando fachada": ["koraza", "koraza sol y lluvia"],
    "descascarando exterior": ["koraza", "koraza sol y lluvia"],
    "deterioro fachada": ["koraza", "koraza sol y lluvia"],
    "intemperie": ["koraza", "koraza sol y lluvia", "barnex"],
    "tierra caliente": ["koraza", "koraza sol y lluvia"],
    "clima calido": ["koraza", "koraza sol y lluvia"],
    "cayendo pedazos lluvia": ["koraza", "koraza sol y lluvia"],

    # ── TECHOS / GOTERAS / CUBIERTAS ──
    "techo": ["pintuco fill", "impercoat", "koraza"],
    "techos": ["pintuco fill", "impercoat", "koraza"],
    "goteras": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "gotera": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "goteando": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "terraza": ["pintuco fill", "impercoat", "koraza"],
    "terraza llueve": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "cubierta": ["pintuco fill", "impercoat", "koraza"],
    "plancha": ["pintuco fill", "impercoat"],
    "plancha concreto": ["pintuco fill", "impercoat"],
    "losa": ["pintuco fill", "impercoat"],
    "impermeabilizar techo": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "impermeabilizar terraza": ["pintuco fill", "impercoat"],
    "llueve terraza": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "llueve techo": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "manto": ["pintuco fill", "impercoat", "tela de refuerzo"],
    "fibrocemento": ["pintuco fill", "koraza", "impercoat"],

    # ── IMPERMEABILIZANTES (genérico) ──
    "impermeabilizante": ["pintuco fill", "impercoat", "aquablock", "koraza", "siliconite"],
    "impermeabilizar": ["pintuco fill", "impercoat", "aquablock", "koraza"],

    # ── PISOS ──
    "pintura piso": ["pintucoat", "pintura canchas"],
    "piso": ["pintucoat", "pintura canchas"],
    "pisos": ["pintucoat", "pintura canchas"],
    "piso industrial": ["pintucoat", "intergard 740", "intergard 2002"],
    "piso cemento": ["pintucoat"],
    "piso concreto": ["pintucoat"],
    "piso trafico": ["pintucoat", "intergard 740", "intergard 2002"],
    "piso fabrica": ["intergard 2002", "pintucoat"],
    "piso bodega": ["pintucoat", "intergard 740"],
    "bodega": ["pintucoat", "intergard 740"],
    "garaje": ["pintucoat"],
    "anden": ["pintucoat"],
    "parqueadero": ["pintucoat", "pintura trafico"],
    "montacargas": ["intergard 2002"],
    "trafico pesado": ["intergard 2002"],
    "piso garaje": ["pintucoat"],

    # ── CANCHAS / DEMARCACIÓN ──
    "cancha": ["pintura canchas"],
    "canchas": ["pintura canchas"],
    "microfutbol": ["pintura canchas"],
    "demarcacion": ["pintutraf", "pintura trafico", "demarcacion vial"],
    "trafico": ["pintutraf", "pintura trafico"],
    "senalizacion": ["pintutraf", "pintura trafico"],
    "lineas": ["pintutraf", "pintura trafico"],

    # ── EPÓXICAS ──
    "epoxica": ["pintucoat", "interseal", "intergard", "epoxica"],
    "epoxica industrial": ["pintucoat", "interseal", "intergard"],
    "pintura epoxica": ["pintucoat", "interseal", "intergard"],
    "recubrimiento epoxica": ["pintucoat", "interseal", "intergard"],
    "dos componentes": ["pintucoat", "interseal", "intergard"],

    # ── ANTICORROSIVOS / METAL / ÓXIDO ──
    "anticorrosivo": ["corrotec", "anticorrosivo", "wash primer", "intergard"],
    "metal": ["corrotec", "pintulux", "anticorrosivo", "wash primer"],
    "metalica": ["corrotec", "pintulux", "anticorrosivo"],
    "hierro": ["corrotec", "pintoxido", "anticorrosivo", "pintulux"],
    "acero": ["corrotec", "anticorrosivo", "intergard", "wash primer"],
    "oxidado": ["corrotec", "pintoxido", "anticorrosivo", "wash primer"],
    "oxidada": ["corrotec", "pintoxido", "anticorrosivo", "wash primer"],
    "oxido": ["corrotec", "pintoxido", "anticorrosivo"],
    "comiendo oxido": ["corrotec", "pintoxido", "anticorrosivo"],
    "reja": ["corrotec", "pintoxido", "pintulux", "anticorrosivo"],
    "rejas": ["corrotec", "pintoxido", "pintulux", "anticorrosivo"],
    "reja oxidada": ["corrotec", "pintoxido", "anticorrosivo"],
    "porton": ["corrotec", "pintulux", "anticorrosivo"],
    "puerta hierro": ["corrotec", "pintoxido", "anticorrosivo", "pintulux"],
    "estructura metalica": ["corrotec", "intergard", "anticorrosivo"],
    "tuberia": ["corrotec", "anticorrosivo", "pintulux"],
    "tubo galvanizado": ["wash primer", "corrotec", "anticorrosivo"],
    "galvanizado": ["wash primer", "corrotec"],
    "proteccion metalica": ["corrotec", "pintulux", "anticorrosivo"],

    # ── MADERA ──
    "barniz": ["barniz", "barnex", "barniz marino", "barniz sd1"],
    "madera": ["barnex", "barniz marino", "barniz", "wood stain", "madetec", "pintulux"],
    "madera exterior": ["barnex", "wood stain", "barniz marino"],
    "madera interior": ["barniz marino", "barniz sd1", "domestico", "pintulux"],
    "piso madera": ["poliuretano alto trafico", "barniz marino", "barniz sd1"],
    "piso de madera": ["poliuretano alto trafico", "barniz marino", "barniz sd1"],
    "piso madera interior": ["poliuretano alto trafico", "barniz marino"],
    "vitrificar": ["poliuretano alto trafico"],
    "vitrificar piso": ["poliuretano alto trafico"],
    "resina transparente piso": ["poliuretano alto trafico"],
    "laca piso": ["poliuretano alto trafico"],
    "laca garaje": ["poliuretano alto trafico"],
    "escalera madera": ["poliuretano alto trafico", "barniz marino"],
    "laca": ["barniz marino", "barniz sd1"],
    "deck": ["barnex", "wood stain", "barniz marino"],
    "pergola": ["barnex", "wood stain"],
    "mueble": ["barniz marino", "barniz sd1", "domestico"],
    "puerta madera": ["pintulux", "domestico", "barniz", "barnex"],

    # ── PRODUCTO NOMBRES DIRECTOS (lookup cuando mencionan el producto por nombre) ──
    "viniltex": ["viniltex", "viniltex adv", "acriltex viniltex"],
    "koraza": ["koraza", "koraza sol y lluvia", "koraza elastomerica"],
    "corrotec": ["corrotec", "anticorrosivo", "pintoxido"],
    "pintucoat": ["pintucoat", "epoxica industrial", "anticorrosiva epoxica"],
    "aquablock": ["aquablock", "aquablock ultra", "sellamur"],
    "pintuco fill": ["pintuco fill", "impercoat", "fill 7", "fill 12"],
    "pintulux": ["pintulux", "pintulux 3en1", "domestico"],
    "barnex": ["barnex", "wood stain", "barniz marino"],
    "pintoxido": ["pintoxido", "corrotec", "anticorrosivo"],
    "imprimante": ["imprimante", "fondo", "sellador", "primer"],
    "wash primer": ["wash primer", "imprimante", "corrotec"],
    "impercoat": ["impercoat", "pintuco fill", "tela de refuerzo"],
    "intervinil": ["intervinil", "viniltex", "vinil latex"],
    "pintulac": ["barniz marino", "barniz sd1", "domestico"],
    "catalizador": ["catalizador", "pintucoat", "interseal"],
    "diluir": ["viniltex", "pintulo", "domestico"],
    "diluyente": ["thinner", "agua destilada", "disolvente"],
    "rendimiento": ["pintuco fill", "koraza", "viniltex"],
    "proporcion": ["pintucoat", "interseal", "catalizador"],
    "mezcla": ["pintucoat", "interseal", "catalizador"],
    "preparacion": ["corrotec", "imprimante", "wash primer", "lija"],
    "secado": ["pintucoat", "corrotec", "viniltex"],

    # ── AEROSOLES ──
    "aerosol": ["aerocolor", "montana"],
    "spray": ["aerocolor", "montana"],
    "pintura spray": ["aerocolor", "montana"],
    "retoque": ["aerocolor"],

    # ── INTERIORES (muros, salas, cuartos) ──
    "pintura interior": ["viniltex", "intervinil", "domestico"],
    "interior": ["viniltex", "intervinil"],
    "sala": ["viniltex", "intervinil"],
    "cuarto": ["viniltex", "intervinil"],
    "habitacion": ["viniltex", "intervinil"],
    "alcoba": ["viniltex", "intervinil"],
    "cielo raso": ["viniltex", "pinturama", "pintura cielos"],
    "cocina": ["viniltex", "viniltex banos y cocinas"],
    "bano": ["viniltex", "viniltex banos y cocinas", "aquablock"],
    "lavable": ["viniltex", "viniltex adv"],
    "pintura lavable": ["viniltex", "viniltex adv"],

    # ── VINILOS genéricos ──
    "vinilo": ["viniltex", "intervinil", "pinturama", "vinil latex"],
    "vinilo tipo 1": ["viniltex", "viniltex adv", "vinil plus"],
    "vinilo tipo 2": ["intervinil", "vinil latex", "vinilux"],
    "vinilo tipo 3": ["pinturama", "vinil max", "icolatex"],
    "vinilo economico": ["pinturama", "vinil max", "icolatex"],
    "vinilo bueno": ["viniltex", "viniltex adv", "vinil plus"],
    "vinilo intermedio": ["intervinil", "vinil latex", "vinilux"],
    "vinilo premium": ["viniltex", "viniltex adv", "vinil plus"],
    "vinilo barato": ["pinturama", "vinil max", "icolatex"],

    # ── ESMALTES ──
    "esmalte": ["pintulux", "domestico"],
    "esmalte bueno": ["pintulux", "pintulux 3en1"],
    "esmalte economico": ["domestico"],
    "esmalte interior": ["domestico", "pintulux"],
    "esmalte exterior": ["pintulux", "pintulux 3en1"],
    "esmalte resistente": ["pintulux", "pintulux 3en1"],

    # ── PINTURA genérica ──
    "pintura buena": ["viniltex", "pintulux"],
    "pintura economica": ["pinturama", "vinil max", "domestico"],

    # ── POLIURETANOS ──
    "poliuretano": ["interthane", "poliuretano"],
    "acabado industrial": ["interthane", "interfine", "pintucoat"],

    # ── INTUMESCENTES / FUEGO ──
    "intumescente": ["interchar", "intumescente"],
    "ignifuga": ["interchar", "intumescente"],
    "retardante fuego": ["interchar"],
    "proteccion fuego": ["interchar"],

    # ── LÍNEA INTERNATIONAL/MPY — MANTENIMIENTO INDUSTRIAL ──────────────────
    # Estos keywords fuerzan búsqueda en la Guía de Sistemas de Mantenimiento Industrial
    "international": ["interseal", "interthane", "intergard", "interfine", "interchar"],
    "mpy": ["interseal", "interthane", "intergard", "interfine", "interchar"],
    "akzonobel": ["interseal", "interthane", "intergard", "interfine", "interchar"],
    "mantenimiento industrial": ["interseal", "intergard", "interthane", "interfine"],
    "sistema mantenimiento": ["interseal", "intergard", "interthane"],
    "recubrimiento industrial": ["interseal", "interthane", "intergard", "interfine"],
    "anticorrosivo industrial": ["intergard", "interseal", "interthane"],
    "estructura acero industrial": ["intergard", "interseal", "interthane"],
    "acero estructural": ["intergard", "interseal", "interthane"],
    "planta industrial": ["interseal", "intergard", "interthane"],
    "bodega quimica": ["interseal", "intergard", "interthane"],
    "almacenamiento quimico": ["interseal", "intergard"],
    "ambiente quimico": ["interseal", "intergard", "interthane"],
    "ambientes agresivos": ["interseal", "intergard", "interthane"],
    "iso 12944": ["interseal", "intergard", "interthane"],
    "sspc": ["interseal", "intergard", "interthane"],
    "epimer": ["interseal", "intergard"],
    "epoxy marino": ["interseal", "intergard"],
    "primer industrial": ["intergard", "interseal"],
    "interchar": ["interchar"],
    "interfine": ["interfine", "interthane"],
    "interseal": ["interseal", "intergard"],
    "intergard": ["intergard", "interseal"],
    "interthane": ["interthane", "interfine"],

    # ── SELLADORES / FONDOS ──
    "sellador": ["sellador", "imprimante", "fondo", "primer"],
    "fondo": ["imprimante", "fondo", "sellador", "wash primer"],
    "primer": ["imprimante", "primer", "wash primer"],
    "base": ["imprimante", "fondo", "sellador"],

    # ── ABRASIVOS / LIJAS / PREPARACIÓN DE SUPERFICIE ──
    "lija": ["lija", "lija agua", "lija al agua", "papel lija"],
    "lijas": ["lija", "lija agua", "lija al agua", "papel lija"],
    "lijar": ["lija", "disco flap", "grata"],
    "lijado": ["lija", "disco flap", "grata"],
    "lijo": ["lija", "viniltex", "estuco", "imprimante"],
    "repintar": ["viniltex", "imprimante", "estuco", "lija"],
    "sin pintar": ["wash primer", "imprimante", "fondo"],
    "disco flap": ["disco flap", "disco abrasivo"],
    "disco abrasivo": ["disco flap", "disco abrasivo", "disco corte"],
    "flap": ["disco flap"],
    "grata": ["grata", "grata copa", "grata circular"],
    "gratas": ["grata", "grata copa", "grata circular"],
    "cepillo metalico": ["grata", "grata copa"],
    "cepillo de alambre": ["grata", "grata copa"],
    "abrasivo": ["lija", "disco flap", "grata", "disco abrasivo"],
    "abrasivos": ["lija", "disco flap", "grata", "disco abrasivo"],
    "preparacion superficie": ["lija", "removedor", "grata", "disco flap", "estuco", "masilla"],
    "preparar superficie": ["lija", "removedor", "grata", "disco flap", "estuco", "masilla"],

    # ── REMOVEDORES / DECAPANTES ──
    "removedor": ["removedor", "removedor pintuco", "decapante"],
    "removedor de pintura": ["removedor", "removedor pintuco"],
    "decapante": ["removedor", "removedor pintuco", "decapante"],
    "quitar pintura": ["removedor", "removedor pintuco", "disco flap", "grata"],
    "remover pintura": ["removedor", "removedor pintuco", "disco flap", "grata"],
    "sacar pintura": ["removedor", "removedor pintuco", "disco flap", "grata"],
    "pelar pintura": ["removedor", "removedor pintuco", "disco flap", "grata"],

    # ── JERGA COLOMBIANA / COLOQUIAL (mapear expresiones comunes) ──
    "pelando": ["koraza", "viniltex", "estuco anti humedad"],
    "descascarando": ["koraza", "viniltex", "estuco anti humedad"],
    "cayendo pedazos": ["koraza", "viniltex"],
    "cayendo a pedazos": ["koraza", "viniltex"],
    "aguacero": ["koraza", "koraza sol y lluvia"],
    "polvo blanco": ["aquablock", "estuco anti humedad", "sellamur"],
    "negro hongos": ["aquablock", "viniltex banos y cocinas"],
    "verdoso": ["aquablock", "viniltex"],
    "mohoso": ["aquablock", "viniltex"],
    "moho": ["aquablock", "viniltex"],

    # ── CERRADURAS / SEGURIDAD (Abracol: Yale, Segurex, Phillips) ──
    "cerradura": ["cerradura", "cerradura yale", "cerradura segurex", "cerradura phillips"],
    "candado": ["candado", "candado yale", "candado segurex"],
    "chapa": ["cerradura", "cerradura yale", "cerradura segurex"],
    "manija": ["manija", "manija yale", "manija phillips", "manija segurex"],
    "antipanico": ["antipanico", "barra antipanico", "yale antipanico"],
    "barra antipanico": ["antipanico", "barra antipanico", "yale antipanico"],
    "cerrojo": ["cerrojo", "cerrojo yale", "cerrojo segurex"],
    "pasador": ["pasador", "pasador induma", "pasador puerta"],

    # ── FERRETERÍA / HERRAJES (Abracol: Induma, Inafer) ──
    "bisagra": ["bisagra", "bisagra induma", "bisagra inafer"],
    "herraje": ["herraje", "bisagra", "pasador", "induma", "inafer"],
    "balde": ["balde", "balde induma", "balde construccion"],

    # ── ADHESIVOS / SELLADORES (Abracol: Tekbond, Artecola, Delta, Afix) ──
    "adhesivo": ["adhesivo", "pegante", "tekbond", "artecola", "delta", "afix"],
    "pegante": ["pegante", "adhesivo", "tekbond", "artecola", "delta"],
    "pegante madera": ["pegante madera", "artecola", "delta", "pl285"],
    "cola": ["cola", "pegante", "artecola", "delta"],
    "silicona": ["silicona", "silicona afix", "silicona tekbond", "sellador silicona"],
    "espuma expansiva": ["espuma poliuretano", "espuma tekbond", "espuma afix"],
    "cinta": ["cinta", "cinta enmascarar", "cinta smith", "cinta teflon", "cinta tekbond"],
    "masking": ["cinta enmascarar", "masking tape", "tirro", "cinta smith"],
    "tirro": ["cinta enmascarar", "masking tape", "tirro", "cinta smith"],

    # ── ABRASIVOS DETALLADOS (Abracol: Abracol, Norton, Carborundum) ──
    "lija agua": ["lija agua", "lija al agua", "lija abracol"],
    "lija seco": ["lija roja", "lija al seco", "lija seco abracol"],
    "fibrodisco": ["fibrodisco", "disco fibra", "fibrodiscos abracol"],
    "disco corte": ["disco corte", "disco corte norton", "disco corte metal"],
    "disco desbaste": ["disco desbaste", "disco desbaste norton"],
    "copa": ["copa", "grata copa", "copa abrasivo"],
    "beartex": ["beartex", "abrasivo no tejido", "beartex norton"],

    # ── AEROSOLES DETALLADOS (Abracol: Tekbond) ──
    "aerosol alta temperatura": ["aerosol alta temperatura", "tekbond alta temperatura", "pintura alta temp"],
    "spray metalizado": ["aerosol metalizado", "tekbond metalizado", "spray metalizado"],

    # ── HERRAMIENTAS DE PINTURA (Abracol: Goya, Atlas) ──
    "brocha": ["brocha", "brocha goya", "brocha atlas", "pincel"],
    "rodillo": ["rodillo", "rodillo goya", "rodillo atlas", "felpa", "mini rodillo"],
    "bandeja": ["bandeja", "bandeja goya", "bandeja pintura"],
    "espátula": ["espatula", "espatula goya", "espatula atlas"],
    "herramientas pintura": ["brocha", "rodillo", "bandeja", "espatula", "goya", "atlas"],
}

# ── Árbol de preguntas diagnósticas basadas en conocimiento técnico ──────
# El agente DEBE usar este árbol para formular preguntas inteligentes
# que le permitan identificar el producto correcto antes de buscar en RAG.
# Cada síntoma/problema tiene preguntas clave y la sospecha de producto
# que el agente debe confirmar con las respuestas del cliente.
DIAGNOSTIC_QUESTION_TREE = {
    "humedad_filtracion": {
        "disparadores": ["humedad", "mojado", "mojada", "filtra", "suda", "sudando", "salitre",
                         "manchas blancas", "ampolla", "ampollando", "descascara", "hongos pared",
                         "moho", "verdoso", "negro bano", "hongos bano"],
        "preguntas": [
            "¿La humedad aparece en muros interiores o en una fachada exterior?",
            "¿Viene de abajo (base del muro) o se nota en toda la pared?",
        ],
        "logica_producto": {
            "interior + base del muro": {"producto": "aquablock", "confianza": "alta",
                "razon": "Humedad freática/ascendente en muros interiores → Aquablock Ultra bloquea presión negativa"},
            "interior + toda la pared": {"producto": "aquablock", "confianza": "alta",
                "razon": "Filtración interna en muro → Aquablock sella e impermeabiliza por dentro"},
            "exterior": {"producto": "koraza", "confianza": "alta",
                "razon": "Fachada exterior deteriorada por lluvia/sol → Koraza impermeabilizante"},
            "bano/cocina + hongos": {"producto": "viniltex banos y cocinas", "confianza": "media",
                "razon": "Zona húmeda sin filtración directa → Viniltex Baños y Cocinas (antihongos)"},
        },
    },
    "fachada_exterior": {
        "disparadores": ["fachada", "frente", "exterior", "pelando exterior", "descascarando",
                         "intemperie", "lluvia sol", "deterioro exterior", "muro exterior"],
        "preguntas": [
            "¿Es una fachada completa o solo un muro pequeño?",
            "¿La pintura actual se está pelando o es obra nueva?",
        ],
        "logica_producto": {
            "fachada expuesta lluvia/sol": {"producto": "koraza", "confianza": "alta",
                "razon": "Koraza Sol y Lluvia es la pintura elastomérica #1 para fachadas de Pintuco"},
            "muro exterior pequeño": {"producto": "koraza", "confianza": "alta",
                "razon": "Incluso muros pequeños exteriores van con Koraza por su resistencia a intemperie"},
        },
    },
    "techo_gotera": {
        "disparadores": ["techo", "gotera", "goteras", "goteando", "terraza llueve",
                         "plancha", "losa", "cubierta", "llueve arriba", "filtra techo"],
        "preguntas": [
            "¿Es un techo de concreto (plancha/losa) o de fibrocemento (eternit)?",
            "¿Tiene grietas visibles o solo se humedece?",
        ],
        "logica_producto": {
            "concreto con grietas": {"producto": "pintuco fill", "confianza": "alta",
                "razon": "Pintuco Fill 12 es impermeabilizante para techos con refuerzo de tela para grietas"},
            "concreto sin grietas": {"producto": "pintuco fill", "confianza": "alta",
                "razon": "Pintuco Fill 7 o Fill 12 como sello preventivo, 2-3 manos"},
            "fibrocemento": {"producto": "koraza", "confianza": "media",
                "razon": "Tejas de fibrocemento van con Koraza impermeabilizante o Fill según el caso"},
        },
    },
    "metal_oxido": {
        "disparadores": ["reja", "rejas", "porton", "hierro", "oxidado", "oxidada", "oxido",
                         "comiendo", "tubería", "estructura metal", "acero"],
        "preguntas": [
            "¿Tiene mucho óxido o es solo superficial?",
            "¿Es interior o está expuesto a la intemperie?",
        ],
        "logica_producto": {
            "oxido profundo": {"producto": "pintoxido", "confianza": "alta",
                "razon": "Pintóxido transforma el óxido profundo, luego se aplica Corrotec como anticorrosivo"},
            "oxido superficial": {"producto": "corrotec", "confianza": "alta",
                "razon": "Corrotec como anticorrosivo directo, luego acabado con Pintulux"},
            "metal nuevo/galvanizado": {"producto": "wash primer", "confianza": "alta",
                "razon": "Wash Primer para adherencia, luego anticorrosivo Corrotec + acabado Pintulux"},
        },
    },
    "piso": {
        "disparadores": ["piso", "garaje", "bodega", "parqueadero", "cancha", "anden",
                         "fabrica", "industrial", "montacargas", "trafico", "estibador",
                         "piso concreto", "piso cemento", "piso exterior", "piso interior"],
        "preguntas": [
            "¿El piso es concreto nuevo (obra gris recién fundido) o es un piso viejo/ya pintado? Si es viejo y ya está pintado, ¿qué tipo de pintura tiene?",
            "Si el piso es nuevo, ¿ya cumplió los 28 días de curado del concreto/mortero? (Esto es crítico: la humedad que evapora de la mezcla de cemento puede dañar la pintura si no ha curado completamente).",
            "¿Cuál va a ser el uso del piso? ¿Tráfico de montacargas/estibadores (pesado), solo peatonal, o mixto?",
            "¿Es un piso interior o exterior? (Si es exterior y le da el sol, el sistema de pintura cambia completamente).",
        ],
        "logica_producto": {
            "industrial/pesado + interior": {"producto": "intergard 2002", "confianza": "alta",
                "razon": "Para tráfico PESADO (montacargas, estibadores) se requiere Intergard 2002 (alto volumen de sólidos) + cuarzo ref 5891610 esparcido por broadcasting. "
                         "Sistema: Interseal gris RAL 7038 (imprimante para concreto) → Intergard 2002 + cuarzo ref 5891610 (2-3 manos). "
                         "NOTA: Pintucoat NO resiste montacargas — es de resistencia MEDIA."},
            "industrial/pesado + exterior": {"producto": "intergard 2002", "confianza": "alta",
                "razon": "Intergard 2002 + cuarzo ref 5891610 + acabado Interthane OBLIGATORIO en exterior. "
                         "Sistema exterior: Interseal gris RAL 7038 (imprimante) → Intergard 2002 + cuarzo → Interthane 990 + Cat PHA046 como sello UV. "
                         "NUNCA usar Pintucoat para tráfico pesado."},
            "industrial/medio + interior": {"producto": "pintucoat", "confianza": "alta",
                "razon": "Pintucoat es epóxica bicomponente de resistencia MEDIA, acabado MATE, para tráfico peatonal y carretillas manuales en interior. "
                         "Sistema: Interseal gris RAL 7038 (imprimante) → Pintucoat Comp A + Catalizador 13227 (2-3 manos). Pot life: 6 horas. "
                         "Alternativa brillante: Intergard 740."},
            "industrial/medio + exterior": {"producto": "pintucoat", "confianza": "alta",
                "razon": "Pintucoat + acabado Interthane OBLIGATORIO en exterior. El epóxico ENTIZA (se decolora/chalking) con la exposición UV. "
                         "Sistema exterior: Interseal gris RAL 7038 (imprimante) → Pintucoat Comp A + Cat 13227 → Interthane 990 + Cat PHA046 como sello UV. "
                         "Alternativa brillante: Intergard 740 + Interthane."},
            "industrial/brillante": {"producto": "intergard 740", "confianza": "alta",
                "razon": "Intergard 740 es epóxico bicomponente de resistencia MEDIA con acabado BRILLANTE. "
                         "Ideal cuando el cliente quiere piso industrial con más brillo que el Pintucoat (que es mate). "
                         "Sistema: Interseal gris RAL 7038 (imprimante) → Intergard 740 (2-3 manos)."},
            "residencial/liviano": {"producto": "pintucoat", "confianza": "alta",
                "razon": "Pintucoat: epóxico bicomponente para garajes, andenes, parqueaderos. "
                         "Resistente al tráfico vehicular liviano y medio. Pintura para Canchas es SOLO para canchas deportivas."},
            "cancha deportiva": {"producto": "pintura canchas", "confianza": "alta",
                "razon": "Pintura para Canchas: acabado antideslizante, resistente al desgaste y UV. Ideal para canchas y andenes."},
            "piso viejo ya pintado": {"producto": "pintucoat", "confianza": "media",
                "razon": "Si ya tiene pintura: (1) Identificar tipo de pintura actual (epóxica, acrílica, alquídica). "
                         "(2) Si es compatible, lijar para dar adherencia. Si no es compatible, remover completamente. "
                         "(3) Aplicar sistema nuevo según tipo de tráfico."},
            "concreto sin curar": {"producto": "ninguno", "confianza": "alta",
                "razon": "NUNCA pintar concreto sin curar. La humedad residual del cemento destruye cualquier recubrimiento. "
                         "Esperar mínimo 28 días de curado. Verificar con prueba de humedad (pegar cinta plástica 24h, si condensa hay humedad)."},
        },
    },
    "interior_general": {
        "disparadores": ["pintar cuarto", "pintar sala", "pintar casa", "pintura interior",
                         "cielo raso", "habitacion", "alcoba"],
        "preguntas": [
            "¿Buscas calidad premium (lavable, durable), intermedia o económica?",
        ],
        "logica_producto": {
            "premium": {"producto": "viniltex", "confianza": "alta",
                "razon": "Viniltex Advanced: super lavable, alto cubrimiento, la mejor del portafolio"},
            "intermedia": {"producto": "intervinil", "confianza": "alta",
                "razon": "Intervinil: buena calidad, precio medio, buen cubrimiento"},
            "economica": {"producto": "pinturama", "confianza": "alta",
                "razon": "Pinturama: excelente relación costo/beneficio para grandes áreas"},
        },
    },
    "madera": {
        "disparadores": ["madera", "barniz", "pergola", "deck", "mueble", "puerta madera"],
        "preguntas": [
            "¿La madera está en interior o a la intemperie?",
            "¿Quieres un acabado transparente (se ve la veta) o de color?",
        ],
        "logica_producto": {
            "exterior/intemperie": {"producto": "barnex", "confianza": "alta",
                "razon": "Barnex Extra Protección o Wood Stain: protección UV, lluvia, resiste intemperie"},
            "interior transparente": {"producto": "barniz marino", "confianza": "alta",
                "razon": "Barniz Marino o Barniz SD1: acabado brillante o mate transparente para muebles e interiores. NO manejamos Pintulac."},
            "interior color": {"producto": "pintulux", "confianza": "media",
                "razon": "Pintulux sobre madera interior cuando se quiere color sólido"},
        },
    },
    "preparacion_superficie": {
        "disparadores": ["lijar", "lijado", "lija", "quitar pintura", "remover pintura",
                         "remover", "removedor", "decapar", "decapante", "sacar pintura",
                         "pelar pintura", "preparar superficie", "preparacion superficie",
                         "disco flap", "grata", "cepillo metalico", "cepillo alambre",
                         "abrasivo", "abrasivos", "como lijo", "con que lijo", "como remuevo"],
        "preguntas": [
            "¿Qué superficie necesitas preparar: metal, madera, concreto o una pared ya pintada?",
            "¿Necesitas quitar la pintura vieja completamente o solo lijar para repintar encima?",
        ],
        "logica_producto": {
            "metal oxidado + quitar oxido": {"producto": "disco flap", "confianza": "alta",
                "razon": "Disco flap o grata para remoción mecánica de óxido, luego Pintóxido si queda óxido residual, luego Corrotec anticorrosivo"},
            "metal + quitar pintura vieja": {"producto": "removedor", "confianza": "alta",
                "razon": "Removedor de Pintuco para pintura vieja sobre metal, o disco flap/grata para remoción mecánica rápida"},
            "madera + quitar barniz/pintura": {"producto": "removedor", "confianza": "alta",
                "razon": "Removedor de Pintuco para decapar barniz o pintura vieja de madera, luego lija fina para alisar"},
            "pared pintada + repintar": {"producto": "lija", "confianza": "alta",
                "razon": "Lija al agua grano 150-220 para dar adherencia a la nueva pintura. Si la pintura está en buen estado, solo lijar suave y pintar"},
            "pared descascarada + reparar": {"producto": "estuco", "confianza": "alta",
                "razon": "Raspar todo lo suelto, aplicar estuco para nivelar, lijar con lija 150, sellar con imprimante y luego pintar"},
            "concreto nuevo + preparar": {"producto": "sellador", "confianza": "alta",
                "razon": "Imprimante/sellador para sellar porosidad del concreto nuevo antes de pintar. No se necesita lija en concreto nuevo"},
        },
    },
    "recreativo_especial": {
        "disparadores": ["tobogan", "tobogán", "pasamanos", "juego infantil", "columpio",
                         "playground", "parque infantil", "resbaladero", "resbalador",
                         "rodadero", "tubo metalico", "estructura tubular", "barandas"],
        "preguntas": [
            "¿La estructura es de metal o plástico?",
            "¿Está al aire libre o en un espacio interior/cubierto?",
        ],
        "logica_producto": {
            "metal exterior": {"producto": "corrotec", "confianza": "alta",
                "razon": "Sistema completo: lija/grata para limpiar óxido → Corrotec anticorrosivo → Pintulux acabado de color. Para máxima durabilidad en exteriores."},
            "metal interior": {"producto": "pintulux", "confianza": "alta",
                "razon": "Lija/grata para limpiar → anticorrosivo Corrotec si hay óxido → Pintulux acabado de color"},
            "plastico": {"producto": "aerocolor", "confianza": "media",
                "razon": "Para plásticos, Aerocolor (aerosol) puede dar buenos resultados. Lijar suave para dar adherencia y aplicar."},
        },
    },
    "industrial_mantenimiento": {
        "disparadores": [
            "industrial", "international", "mpy", "interseal", "interthane", "intergard",
            "interfine", "interchar", "estructura de acero", "acero estructural",
            "planta industrial", "bodega quimica", "mantenimiento industrial",
            "sistema de pintura industrial", "recubrimiento industrial",
            "iso 12944", "sspc", "anticorrosivo industrial", "ambiente agresivo",
            "ambientes quimicos", "corrosion industrial", "epoxica industrial",
            "proteccion fuego estructura", "intumescente metal",
        ],
        "preguntas": [
            "¿Qué tipo de estructura o superficie necesitas proteger? (acero estructural, tuberías, tanques, pisos industriales, estructura contra incendio)",
            "¿En qué entorno está expuesta? (interior industrial, exterior corrosivo, ambiente químico, marino, temperatura extrema)",
        ],
        "logica_producto": {
            "acero estructural / entorno corrosivo": {
                "producto": "interseal",
                "confianza": "alta",
                "razon": "Sistema International: Intergard (primer epóxico) + Interseal (body coat epóxico) + Interthane (acabado poliuretano). Norma ISO 12944 / SSPC.",
            },
            "primer / imprimacion industrial": {
                "producto": "intergard",
                "confianza": "alta",
                "razon": "Intergard es el primer epóxico de International/MPY para preparación de superficie de acero. Requiere preparación SSPC SP6 o superior.",
            },
            "acabado industrial / resistencia UV / quimica": {
                "producto": "interthane",
                "confianza": "alta",
                "razon": "Interthane es el poliuretano de acabado de International/MPY. Alta resistencia química, UV y mecánica. Aplicar sobre Interseal o Intergard.",
            },
            "proteccion fuego / incendio": {
                "producto": "interchar",
                "confianza": "alta",
                "razon": "Interchar es el intumescente de International/MPY para protección pasiva contra incendio en estructuras metálicas. Requiere cálculo de espesor por arquitecto.",
            },
            "acabado de alto brillo industrial": {
                "producto": "interfine",
                "confianza": "alta",
                "razon": "Interfine es el acabado de altas prestaciones de International/MPY. Para superficies de alta visibilidad o requerimientos estéticos industriales.",
            },
        },
    },
}

# ── Productos/aplicaciones que Ferreinox NO maneja ────────────────────────
# Solo para casos donde NO existe ningún producto en el portafolio (ni Pintuco ni International).
# Para aplicaciones condicionales (inmersión, agua potable) el agente DEBE consultar el RAG primero.
PORTFOLIO_GAPS = {
    "piscina": "Ferreinox actualmente no maneja en su portafolio una pintura especializada para piscinas con garantía técnica. Te recomiendo comunicarte con uno de nuestros asesores para que te orienten con el fabricante, o consultar en www.ferreinox.co.",
    "piscinas": "Ferreinox actualmente no maneja en su portafolio una pintura especializada para piscinas con garantía técnica. Te recomiendo comunicarte con uno de nuestros asesores para que te orienten con el fabricante, o consultar en www.ferreinox.co.",
    "pintura marina": "No manejamos pintura marina anti-incrustante. Contacta a un asesor para orientación especializada.",
    # NOTA: "tanque agua potable" e "inmersion" SE ELIMINARON del hard-block porque la línea
    # International/AkzoNobel (Interseal 670HS) tiene certificación NSF/ANSI 61 para agua potable
    # con condiciones específicas. El agente debe consultar el RAG (consultar_conocimiento_tecnico)
    # antes de responder esas consultas.
}

# ── Reglas técnicas verificadas por producto ──────────────────────────────
# Estas reglas PREVALECEN sobre el RAG y sobre el conocimiento general.
# Si el RAG o la IA sugieren algo que contradice estas reglas, las reglas ganan.
PRODUCT_TECHNICAL_HARD_RULES = {
    "koraza": {
        "es_para": "Pintura elastomérica para FACHADAS exteriores, muros exteriores expuestos a lluvia y sol, terrazas descubiertas.",
        "no_es_para": "NO es sellador de humedad interna, NO es impermeabilizante de muros con filtración, NO sella grietas con presión de agua. Para humedad interna usa Aquablock o Sellamur.",
    },
    "pintucoat": {
        "es_para": "Recubrimiento epóxico BICOMPONENTE para pisos industriales de tráfico MEDIO (peatonal, carretillas manuales). Acabado MATE. NO resiste tráfico pesado de montacargas — para tráfico pesado usar Intergard 2002 + cuarzo.",
        "no_es_para": "NO es para piscinas, NO es para tanques de agua, NO es para inmersión en agua, NO es para superficies sumergidas. NO es para tráfico pesado de montacargas.",
        "bicomponente": (
            "SIEMPRE requiere catalizador 13227 COMP B. "
            "GALÓN: COMP A 3.44L + catalizador 13227 COMP B 0.37L. "
            "CUÑETE: COMP A 15.14L + catalizador 13227 COMP B 1.89L. Pot-life 6h."
        ),
        "exterior": "En exterior expuesto al sol entiza. Requiere acabado Interthane (poliuretano) encima. Pintulux 3en1 NO es sustituto de poliuretano.",
    },
    "interthane": {
        "es_para": "Poliuretano de acabado BICOMPONENTE, alta resistencia UV y química. Marca International/AkzoNobel. Capa final sobre epóxicos industriales.",
        "no_es_para": "NO es anticorrosivo de primera capa. Se aplica sobre imprimación epóxica curada.",
        "bicomponente": (
            "SIEMPRE requiere catalizador PHA046. "
            "GALÓN: COMP A 3.7L + catalizador PHA046 0.5L. "
            "CUÑETE: COMP A 20L + catalizador PHA046 3.7L."
        ),
    },
    "interseal": {
        "es_para": "Epóxico de alto espesor BICOMPONENTE (body coat), capa intermedia en sistemas industriales International/AkzoNobel.",
        "no_es_para": "NO es acabado final en exterior con exposición UV — requiere Interthane encima.",
        "bicomponente": "Relación y catalizador extraer de ficha técnica International o Guía de Sistemas MPY.",
    },
    "intergard": {
        "es_para": "Primer epóxico BICOMPONENTE, primera capa sobre acero limpio. Protección anticorrosiva base. Marca International/AkzoNobel. La línea Intergard 740 e Intergard 2002 también se usan como acabados de pisos industriales.",
        "no_es_para": "NO es acabado final (excepto Intergard 740 en pisos y Intergard 2002 en pisos con cuarzo).",
        "bicomponente": "Relación y catalizador extraer de ficha técnica International.",
    },
    # ── BLOQUEO ABSOLUTO: PRIMER 50RS / EPOXY PRIMER 50RS / UEA400 ──
    "primer 50rs": {
        "es_para": "Imprimante epóxico para ESTRUCTURAS METÁLICAS únicamente. Protección anticorrosiva sobre acero.",
        "no_es_para": (
            "NUNCA para pisos de concreto. NUNCA para pisos industriales. NUNCA para superficies de cemento. "
            "Si el cliente necesita imprimante para piso de concreto, el correcto es INTERSEAL GRIS RAL 7038. "
            "Primer 50RS es EXCLUSIVO para metal."
        ),
    },
    "epoxy primer 50rs": {
        "es_para": "Imprimante epóxico para ESTRUCTURAS METÁLICAS únicamente.",
        "no_es_para": "NUNCA para pisos de concreto. Para pisos usar INTERSEAL GRIS RAL 7038 como imprimante.",
    },
    # ── IMPRIMANTE CORRECTO PARA PISOS DE CONCRETO ──
    "interseal gris": {
        "es_para": (
            "Imprimante epóxico para PISOS DE CONCRETO. Interseal gris RAL 7038 es el imprimante correcto "
            "que va como primera capa antes de Pintucoat, Intergard 740 o Intergard 2002 en pisos industriales. "
            "Sella la porosidad del concreto y mejora la adherencia del acabado epóxico."
        ),
        "no_es_para": "NO es acabado final. Va como capa de imprimación debajo del sistema de piso.",
    },
    "aquablock": {
        "es_para": "Sellador y bloqueador de humedad para muros interiores con filtración. Disponible en presentación para interiores (Aquablock) y como impermeabilizante (Aquablock Ultra).",
        "no_es_para": "NO es para fachadas (para fachadas usar Koraza). NO es para pisos ni piscinas.",
    },
    "sellamur": {
        "es_para": "Sellador para muros minerales y bases alcalinas donde se necesita uniformar absorción o apoyar sistemas arquitectónicos compatibles.",
        "no_es_para": "NO sustituye un bloqueador de humedad activa ni es acabado final exterior de alta exigencia.",
    },
    "pintuco fill": {
        "es_para": "Impermeabilizante para techos, cubiertas, terrazas. La línea más grande de impermeabilizantes del portafolio (Fill 7, Fill 12).",
        "no_es_para": "NO es para piscinas ni inmersión en agua.",
    },
    "viniltex": {
        "es_para": "Pintura vinílica premium para muros interiores y exteriores. Lavable, buen cubrimiento.",
        "no_es_para": "NO es para pisos, NO es para piscinas, NO es para metal desnudo, NO es para inmersión en agua.",
    },
    "siliconite": {
        "es_para": "Hidrofugante para sustratos minerales vistos como ladrillo, concreto arquitectónico o fachaleta cuando se quiere proteger sin formar película opaca.",
        "no_es_para": "NO es pintura de acabado decorativo ni reemplaza un elastomérico de fachada cuando el objetivo es cubrir y decorar.",
    },
    "espuma de poliuretano": {
        "es_para": "Sellado, relleno y aislamiento de huecos, juntas y pasos de instalaciones en construcción y montaje liviano.",
        "no_es_para": "NO es pintura, NO es acabado decorativo ni reemplaza un impermeabilizante de cubierta o un sistema anticorrosivo.",
    },
    "esmaltes top quality": {
        "es_para": "Acabado decorativo brillante para metal, madera, cemento o asbesto-cemento en usos arquitectónicos o de mantenimiento liviano.",
        "no_es_para": "NO reemplaza un sistema industrial 2K para ambientes químicos, inmersión o alta exigencia anticorrosiva.",
    },
    "wash primer": {
        "es_para": "Imprimante de adherencia para galvanizado, aluminio y metales no ferrosos antes del sistema anticorrosivo o de acabado.",
        "no_es_para": "NO es acabado final y NO sustituye un anticorrosivo ni un sistema de pisos de concreto.",
    },
    "interfine": {
        "es_para": "Acabado industrial de altas prestaciones estéticas y alta retención de color para estructuras metálicas o proyectos donde la apariencia final es crítica.",
        "no_es_para": "NO es primer ni capa anticorrosiva inicial. Debe ir sobre sistema industrial compatible y superficie preparada.",
    },
    "interchar": {
        "es_para": "Recubrimiento intumescente para protección pasiva contra incendio en estructuras metálicas, bajo diseño y espesor calculado.",
        "no_es_para": "NO es esmalte decorativo común ni se debe recomendar sin definir rating de fuego, perfil estructural y espesor requerido.",
    },
    "pintura canchas": {
        "es_para": "Pintura EXCLUSIVA para canchas deportivas, escenarios deportivos, senderos peatonales y ciclo rutas.",
        "no_es_para": "NO es para garajes, NO es para parqueaderos, NO es para bodegas, NO es para tráfico vehicular. Para garajes usar Pintucoat, para industrial Intergard 2002 + cuarzo.",
    },
}


GLOBAL_TECHNICAL_POLICY_RULES = [
    {
        "name": "humedad_interior_negativa",
        "problem_classes": {"humedad_interior_capilaridad", "humedad_interior_general"},
        "priority": "high",
        "required_products": ["Aquablock"],
        "forbidden_products": ["Koraza", "Pintuco Fill"],
        "mandatory_steps": [
            "Retirar acabado soplado, salitre y base floja hasta sustrato sano antes del bloqueador.",
            "Bloquear la humedad primero y solo despues reconstruir el acabado decorativo.",
        ],
        "mandatory_step_signals": ["aquablock", "retirar el acabado", "bloquear la humedad"],
        "rules_text": [
            "La humedad interior por capilaridad o presion negativa no se resuelve con pintura decorativa ni con elastomericos de fachada.",
        ],
    },
    {
        "name": "fachada_alta_exposicion",
        "problem_classes": {"fachada_exterior"},
        "required_products": ["Koraza"],
        "forbidden_products": ["Intervinil", "Pinturama", "vinilos interiores", "Aquablock"],
        "mandatory_steps": [
            "Retirar pintura suelta o base soplada antes del acabado exterior.",
        ],
        "mandatory_step_signals": ["koraza", "pintura suelta"],
        "rules_text": [
            "En fachadas reales expuestas a lluvia y sol el acabado debe ser exterior; no cerrar con vinilos interiores ni con bloqueadores de humedad interior.",
        ],
    },
    {
        "name": "eternit_fibrocemento_exterior",
        "match_any": ["eternit", "fibrocemento", "asbesto", "asbesto cemento"],
        "required_products": ["Sellomax", "Koraza"],
        "required_tools": ["hidrolavadora", "cepillo"],
        "forbidden_products": ["Intervinil", "Pinturama", "vinilos interiores"],
        "forbidden_tools": ["lijas", "rasqueta", "preparacion mecanica"],
        "mandatory_steps": [
            "Preparacion humeda obligatoria; nunca lijar en seco ni rasquetear.",
            "En eternit envejecido o repintado, Sellomax va antes del acabado exterior.",
        ],
        "mandatory_step_signals": ["preparacion humeda", "sellomax", "koraza"],
        "rules_text": [
            "El fibrocemento exterior se maneja con control de polvo y sistema exterior; no con vinilos interiores.",
        ],
    },
    {
        "name": "ladrillo_a_la_vista",
        "match_any": ["ladrillo a la vista", "ladrillo", "fachaleta", "mamposteria a la vista"],
        "required_products": ["Construcleaner Limpiador Desengrasante", "Siliconite 7"],
        "forbidden_products": ["Koraza", "acido muriatico"],
        "mandatory_steps": [
            "Limpiar el ladrillo con limpiador adecuado antes de hidrofugar.",
            "Conservar la apariencia del sustrato; proteger sin formar pelicula opaca.",
        ],
        "mandatory_step_signals": ["construcleaner", "siliconite"],
        "rules_text": [
            "El ladrillo a la vista se limpia e hidrofuga; no se debe deteriorar con acidos fuertes ni taparlo con pintura elastomerica si se quiere conservar la textura.",
        ],
    },
    {
        "name": "metal_pintado_alquidico",
        "problem_classes": {"metal_pintado_alquidico"},
        "forbidden_products": ["Interseal 670", "Interseal", "Intergard", "Interthane 990", "Pintucoat"],
        "mandatory_steps": [
            "Remocion total hasta metal desnudo antes de migrar a epoxicos o poliuretanos.",
        ],
        "mandatory_step_signals": ["metal desnudo", "remocion total"],
        "rules_text": [
            "No migrar directo de anticorrosivo alquidico viejo o esmalte sintetico a un sistema industrial 2K.",
        ],
    },
    {
        "name": "arquitectonico_sobre_base_agua",
        "match_any": ["base agua", "vinilo", "viniltex", "intervinil", "pinturama", "acrilico", "muros de casa"],
        "forbidden_products": ["Interthane 990", "Interseal", "Intergard", "Pintucoat"],
        "mandatory_steps": [
            "Mantener compatibilidad de familia: agua con agua sobre sistemas arquitectonicos existentes.",
        ],
        "mandatory_step_signals": ["misma familia", "agua con agua"],
        "rules_text": [
            "Sobre pintura base agua o sistema arquitectonico existente no ofrecer sistemas industriales bicomponentes como solucion directa.",
        ],
    },
    {
        "name": "piso_industrial_trafico_pesado",
        "problem_classes": {"piso_industrial"},
        "match_any": ["montacargas", "estibador", "llantas duras", "trafico pesado", "trafico industrial pesado"],
        "required_products": ["Interseal gris RAL 7038", "Intergard 2002", "Arena de Cuarzo ref 5891610"],
        "forbidden_products": ["Pintucoat", "Primer 50RS", "Epoxy Primer 50RS"],
        "mandatory_steps": [
            "Preparacion mecanica y desengrase profundo antes del sistema epoxico.",
            "Confirmar m2, estado del concreto y tipo de trafico antes de cerrar sistema o cantidades.",
        ],
        "mandatory_step_signals": ["intergard 2002", "cuarzo", "preparacion mecanica"],
        "rules_text": [
            "El trafico pesado de montacargas o estibadores no se resuelve con Pintucoat y tampoco con imprimantes para metal.",
        ],
    },
    {
        "name": "piso_industrial_trafico_medio",
        "problem_classes": {"piso_industrial"},
        "match_any": ["trafico medio", "peatonal", "garaje", "parqueadero", "carretilla manual", "residencial"],
        "required_products": ["Interseal gris RAL 7038", "Pintucoat"],
        "forbidden_products": ["Primer 50RS", "Epoxy Primer 50RS"],
        "mandatory_steps": [
            "Confirmar si el piso es nuevo o ya pintado antes de definir compatibilidad.",
        ],
        "mandatory_step_signals": ["pintucoat", "interseal gris"],
        "rules_text": [
            "Para concreto de trafico medio, el imprimante correcto es Interseal gris RAL 7038; Primer 50RS es solo para metal.",
        ],
    },
    {
        "name": "piso_exterior_uv",
        "problem_classes": {"piso_industrial"},
        "match_any": ["exterior", "sol", "uv", "intemperie"],
        "required_products": ["Interthane 990 + Catalizador"],
        "mandatory_steps": [
            "Todo sistema epoxico de piso exterior expuesto al sol debe cerrarse con poliuretano UV.",
        ],
        "mandatory_step_signals": ["interthane"],
        "rules_text": [
            "Los epoxicos expuestos al sol entizan; no pueden quedar como acabado final exterior.",
        ],
    },
    {
        "name": "concreto_sin_curado",
        "problem_classes": {"piso_industrial"},
        "match_any": ["recien fundido", "reci en fundido", "recien", "recien vaciado", "obra gris", "concreto nuevo", "sin curar"],
        "priority": "high",
        "mandatory_steps": [
            "Esperar minimo 28 dias de curado y validar humedad antes de pintar.",
        ],
        "mandatory_step_signals": ["28 dias", "curado"],
        "rules_text": [
            "No se debe pintar concreto fresco porque la humedad residual destruye el recubrimiento.",
        ],
    },
    {
        "name": "madera_exterior",
        "problem_classes": {"madera"},
        "match_any": ["exterior", "intemperie", "deck", "pergola", "pergola", "fachada"],
        "required_products": ["Barnex", "Wood Stain"],
        "forbidden_products": ["Poliuretano Alto Trafico 1550/1551"],
        "mandatory_steps": [
            "En madera exterior usar sistema con proteccion UV y no un poliuretano transparente de piso interior.",
        ],
        "mandatory_step_signals": ["barnex", "wood stain"],
        "rules_text": [
            "La madera exterior necesita proteccion UV y flexibilidad; el poliuretano de alto trafico interior no sirve para intemperie.",
        ],
    },
    {
        "name": "madera_interior_alto_trafico",
        "problem_classes": {"madera"},
        "match_any": ["piso", "escalera", "vitrificar", "laca piso", "interior"],
        "required_products": ["Poliuretano Alto Trafico 1550/1551"],
        "forbidden_products": ["Barnex", "Pintulac", "barniz arquitectonico"],
        "mandatory_steps": [
            "Mezclar A+B y respetar el lijado fino entre manos en el sistema poliuretano interior.",
        ],
        "mandatory_step_signals": ["poliuretano alto trafico", "a+b"],
        "rules_text": [
            "Cuando el cliente quiere vitrificado resistente en piso o escalera interior, el sistema correcto es el poliuretano bicomponente 1550/1551.",
        ],
    },
    {
        "name": "techo_concreto_grietas",
        "match_all": ["techo", "concreto"],
        "match_any": ["grieta", "grietas", "fisura", "fisuras", "terraza", "losa"],
        "required_products": ["Pintuco Fill"],
        "forbidden_products": ["Koraza", "Viniltex", "Intervinil", "Pinturama"],
        "mandatory_steps": [
            "En techos de concreto con grietas tratar la impermeabilizacion como sistema de cubierta, no como pintura decorativa.",
            "Definir si requiere refuerzo de tela o tratamiento de fisuras antes del acabado final.",
        ],
        "mandatory_step_signals": ["pintuco fill", "grietas"],
        "rules_text": [
            "La losa o terraza con fisuras debe entrar por la ruta de impermeabilizacion de cubiertas y no por pintura de fachada.",
        ],
    },
    {
        "name": "bano_cocina_antihongos",
        "match_any": ["bano", "baño", "cocina", "hongos", "moho", "zona humeda", "zona húmeda"],
        "exclude_any": ["capilaridad", "salitre", "presion negativa", "presión negativa", "jardinera"],
        "required_products": ["Viniltex Baños y Cocinas"],
        "forbidden_products": ["Koraza", "Pintucoat", "Interseal", "Intergard", "Interthane 990"],
        "mandatory_steps": [
            "Separar condensacion o hongos superficiales de una humedad estructural real antes de definir el sistema.",
        ],
        "mandatory_step_signals": ["viniltex baños y cocinas"],
        "rules_text": [
            "En baños y cocinas con hongos superficiales sin presión de agua activa, el acabado debe ser antihongos y lavable, no un industrial 2K.",
        ],
    },
    {
        "name": "interior_koraza_redirect",
        "match_any": ["muro interior", "sala", "alcoba", "pasillo cerrado", "habitacion", "habitación", "interior"],
        "match_all_non_negated": ["koraza"],
        "required_products": ["Viniltex Advanced"],
        "forbidden_products": ["Koraza"],
        "mandatory_steps": [
            "Si el cliente pide Koraza para interior cerrado, reconducir a un vinilo premium compatible con ese uso.",
        ],
        "mandatory_step_signals": ["viniltex advanced"],
        "rules_text": [
            "Koraza es una ruta de intemperie y no debe quedar como recomendacion principal en interiores cerrados convencionales.",
        ],
    },
    {
        "name": "cancha_sendero_peatonal",
        "match_any": ["cancha", "ciclo ruta", "cicloruta", "sendero peatonal", "escenario deportivo"],
        "required_products": ["Pintura Canchas"],
        "forbidden_products": ["Pintucoat", "Intergard 2002", "Intergard 740"],
        "mandatory_steps": [
            "No mezclar la ruta deportiva con pisos industriales de montacargas o bodegas.",
        ],
        "mandatory_step_signals": ["pintura canchas"],
        "rules_text": [
            "La cancha o sendero peatonal debe resolverse con sistema deportivo o peatonal, no con epoxicos industriales de bodega.",
        ],
    },
    {
        "name": "metal_nuevo_galvanizado",
        "match_any": ["galvanizado", "galvanizada", "aluminio", "metal nuevo", "lamina zinc", "lámina zinc"],
        "required_products": ["Wash Primer"],
        "forbidden_products": ["Interseal gris RAL 7038", "Pintucoat"],
        "mandatory_steps": [
            "En galvanizado o metal no ferroso primero resolver adherencia con wash primer antes del anticorrosivo o acabado.",
        ],
        "mandatory_step_signals": ["wash primer"],
        "rules_text": [
            "El galvanizado no se trata igual que el acero negro oxidado ni como un piso de concreto.",
        ],
    },
    {
        "name": "metal_oxidado_mantenimiento",
        "match_any": ["reja oxidada", "metal oxidado", "oxido superficial", "óxido superficial", "corrosion superficial", "corrosión superficial", "oxidada"],
        "required_products": ["Pintóxido", "Corrotec"],
        "forbidden_products": ["Viniltex", "Koraza", "Pintucoat"],
        "mandatory_steps": [
            "Separar oxido superficial de corrosion profunda antes de definir transformador o remocion mecanica intensiva.",
        ],
        "mandatory_step_signals": ["pintoxido", "corrotec"],
        "rules_text": [
            "El metal oxidado de mantenimiento liviano debe entrar por desoxidante y anticorrosivo, no por pinturas para muro ni sistemas de piso.",
        ],
    },
    {
        "name": "metal_oxidado_preparacion_incorrecta",
        "match_any": ["reja oxidada", "metal oxidado", "oxido superficial", "óxido superficial", "oxidada"],
        "match_any_non_negated": ["agua y jabon", "agua con jabon", "lavar con agua", "lavarlo con agua", "agua jabonosa", "jabon", "jabón", "con agua antes del anticorrosivo", "lavar la reja oxidada con agua"],
        "priority": "high",
        "required_tools": ["grata", "lija"],
        "forbidden_tools": ["agua y jabon"],
        "mandatory_steps": [
            "En metal oxidado no usar agua y jabón como preparación principal; retirar óxido y cascarilla con grata o lija antes del convertidor o anticorrosivo.",
            "Aplicar el sistema solo sobre metal seco y con el óxido flojo removido.",
        ],
        "mandatory_step_signals": ["grata", "lija", "metal seco"],
        "rules_text": [
            "El lavado con agua y jabón no sustituye la preparación mecánica del metal oxidado y puede empeorar la condición si deja humedad retenida.",
        ],
    },
    {
        "name": "espuma_poliuretano_sellado",
        "match_any": ["espuma de poliuretano", "sellar huecos", "rellenar huecos", "aislamiento termico", "aislamiento térmico", "espuma expansiva"],
        "required_products": ["Espuma de Poliuretano"],
        "forbidden_products": ["Koraza", "Viniltex", "Pintuco Fill", "Interseal"],
        "mandatory_steps": [
            "Usar la espuma como sistema de sellado o relleno sobre superficie limpia, no como pintura o acabado decorativo.",
        ],
        "mandatory_step_signals": ["espuma de poliuretano"],
        "rules_text": [
            "La espuma expansiva es una solucion de sellado y aislamiento, no una pintura ni un impermeabilizante de acabado.",
        ],
    },
    {
        "name": "esmalte_decorativo_mantenimiento",
        "match_any": ["esmalte top quality", "esmalte brillante", "acabado brillante", "mantenimiento liviano", "metal decorativo", "madera decorativa"],
        "exclude_any": ["ambiente quimico", "ambiente químico", "interseal", "interthane", "intergard", "inmersion", "inmersión"],
        "required_products": ["Esmaltes Top Quality"],
        "forbidden_products": ["Interseal", "Intergard", "Interthane 990"],
        "mandatory_steps": [
            "Tratar el caso como acabado decorativo o mantenimiento liviano, no como sistema industrial 2K.",
        ],
        "mandatory_step_signals": ["esmaltes top quality"],
        "rules_text": [
            "Cuando la necesidad es decorativa con brillo y exposición convencional, el esmalte arquitectónico es mejor ruta que un sistema industrial completo.",
        ],
    },
    {
        "name": "inmersion_agua_potable_condicional",
        "match_any": ["inmersion", "inmersión", "sumergido", "sumergida", "tanque agua potable", "agua potable", "nsf", "ansi 61"],
        "priority": "critical",
        "forbidden_products": ["Pintucoat", "Viniltex", "Koraza", "Pintulux 3 en 1"],
        "mandatory_steps": [
            "Validar ficha tecnica, certificacion aplicable y preparacion Sa 2.5 o SSPC-SP10 antes de recomendar un sistema de inmersion o agua potable.",
            "Si se trata de agua potable, confirmar la condicion NSF/ANSI 61 y el volumen del tanque antes de cerrar el sistema.",
        ],
        "mandatory_step_signals": ["agua potable", "sa 2.5", "nsf"],
        "rules_text": [
            "Las consultas de inmersion o agua potable requieren ruta tecnica condicionada; no se deben resolver con pinturas arquitectonicas ni con Pintucoat.",
        ],
    },
    {
        "name": "proteccion_pasiva_incendio",
        "match_any": ["interchar", "intumescente", "proteccion fuego", "protección fuego", "incendio", "ignifugo", "ignífugo"],
        "priority": "critical",
        "required_products": ["Interchar"],
        "forbidden_products": ["Koraza", "Viniltex", "Pintulux 3 en 1"],
        "mandatory_steps": [
            "Definir rating de fuego, perfil estructural y espesor requerido antes de recomendar el sistema intumescente.",
            "No tratar la proteccion pasiva contra incendio como una pintura decorativa comun.",
        ],
        "mandatory_step_signals": ["interchar", "espesor requerido"],
        "rules_text": [
            "El intumescente exige calculo y compatibilidad de sistema; no es una respuesta genérica de pintura para metal.",
        ],
    },
    {
        "name": "acabado_industrial_alta_estetica",
        "match_any": ["interfine", "alto brillo industrial", "alta estetica", "alta estética", "retencion de color", "retención de color"],
        "required_products": ["Interfine"],
        "forbidden_products": ["Corrotec", "Pintulux 3 en 1"],
        "mandatory_steps": [
            "Usar Interfine solo como acabado de altas prestaciones sobre sistema industrial compatible, no como primer.",
        ],
        "mandatory_step_signals": ["interfine"],
        "rules_text": [
            "Cuando el requerimiento es estético industrial de alto nivel, la conversación debe ir por Interfine o poliuretano industrial compatible, no por esmaltes domésticos.",
        ],
    },
    {
        "name": "ambiente_quimico_industrial",
        "match_any": ["ambiente quimico", "ambiente químico", "quimicos", "químicos", "planta industrial", "corrosion industrial", "corrosión industrial", "iso 12944", "sspc"],
        "priority": "high",
        "required_products": ["Intergard", "Interseal", "Interthane 990 + Catalizador"],
        "forbidden_products": ["Corrotec", "Pintulux 3 en 1", "Viniltex", "Koraza"],
        "mandatory_steps": [
            "Resolver preparacion de superficie y ambiente de exposición antes de cerrar un sistema industrial anticorrosivo.",
            "No degradar una consulta industrial severa a soluciones arquitectonicas o esmaltes domesticos.",
        ],
        "mandatory_step_signals": ["intergard", "interseal", "interthane"],
        "rules_text": [
            "Ambiente químico o anticorrosión industrial severa exige sistema industrial completo y no productos de mantenimiento liviano.",
        ],
    },
    {
        "name": "concreto_sin_curado_acido_incorrecto",
        "problem_classes": {"piso_industrial"},
        "match_any": ["recien fundido", "reci en fundido", "recien vaciado", "concreto nuevo", "sin curar", "obra gris"],
        "match_any_non_negated": ["acido muriatico", "ácido muriático", "echar acido muriatico", "echar ácido muriático"],
        "priority": "high",
        "forbidden_tools": ["acido muriatico"],
        "mandatory_steps": [
            "No usar ácido muriático para forzar curado ni preparación temprana del concreto recién fundido.",
            "Esperar el curado mínimo, validar humedad y luego definir el tratamiento de superficie correcto.",
        ],
        "mandatory_step_signals": ["28 dias", "curado"],
        "rules_text": [
            "El ácido muriático no acelera el curado del concreto fresco y puede comprometer la superficie antes del sistema de recubrimiento.",
        ],
    },
]


def _mention_is_negated_in_query(normalized_query: str, start_index: int) -> bool:
    window = normalized_query[max(0, start_index - 45):start_index]
    negation_cues = [
        " no ", " nunca ", " evita ", " evitar ", " prohibido ", " jamas ", " jamás ",
        " no quiero ", " no usar ", " no voy a usar ", " no pienso usar ", " no aplicar ",
    ]
    return any(cue in f" {window} " for cue in negation_cues)


def _query_matches_token(normalized_query: str, token: str, allow_negated: bool = True) -> bool:
    padded_query = f" {normalized_query} "
    normalized_token = normalize_text_value(token)
    if not normalized_token:
        return False

    search_candidates = [f" {normalized_token} ", normalized_token]
    for search_token in search_candidates:
        start = padded_query.find(search_token)
        while start != -1:
            if allow_negated or not _mention_is_negated_in_query(padded_query, start):
                return True
            start = padded_query.find(search_token, start + len(search_token))
    return False


def _query_matches_all_tokens(normalized_query: str, tokens: list[str], allow_negated: bool = True) -> bool:
    return all(_query_matches_token(normalized_query, token, allow_negated=allow_negated) for token in tokens or [])


def _query_matches_any_token(normalized_query: str, tokens: list[str], allow_negated: bool = True) -> bool:
    return any(_query_matches_token(normalized_query, token, allow_negated=allow_negated) for token in tokens or [])


def _matches_global_policy_rule(rule: dict, normalized_query: str, diagnosis: dict) -> bool:
    problem_class = diagnosis.get("problem_class")
    problem_classes = rule.get("problem_classes") or set()
    if problem_classes and problem_class not in problem_classes:
        return False
    if rule.get("match_all") and not _query_matches_all_tokens(normalized_query, rule.get("match_all") or []):
        return False
    if rule.get("match_all_non_negated") and not _query_matches_all_tokens(normalized_query, rule.get("match_all_non_negated") or [], allow_negated=False):
        return False
    if rule.get("match_any") and not _query_matches_any_token(normalized_query, rule.get("match_any") or []):
        return False
    if rule.get("match_any_non_negated") and not _query_matches_any_token(normalized_query, rule.get("match_any_non_negated") or [], allow_negated=False):
        return False
    if rule.get("exclude_any") and _query_matches_any_token(normalized_query, rule.get("exclude_any") or []):
        return False
    return bool(problem_classes or rule.get("match_all") or rule.get("match_all_non_negated") or rule.get("match_any") or rule.get("match_any_non_negated"))


def _infer_problem_class_from_rag_query(question: str, product: str = "") -> Optional[str]:
    normalized = normalize_text_value(f"{question} {product}")
    if not normalized:
        return None

    if any(token in normalized for token in ["eternit", "fibrocemento", "asbesto", "asbesto cemento"]):
        return "eternit_fibrocemento"

    if any(token in normalized for token in ["ladrillo a la vista", "ladrillo", "fachaleta", "mamposteria a la vista"]):
        return "ladrillo_vista"

    if any(token in normalized for token in ["madera", "barnex", "wood stain", "barniz", "vitrificar", "escalera madera", "deck", "pergola", "pérgola"]):
        return "madera"

    if any(token in normalized for token in ["metal", "reja", "porton", "acero", "hierro"]) and any(
        token in normalized for token in ["alquidico", "alquidico", "esmalte sintetico", "pintura de aceite", "anticorrosivo viejo", "ya pintado", "pintada"]
    ):
        return "metal_pintado_alquidico"

    if any(token in normalized for token in ["humedad", "salitre", "capilaridad", "presion negativa", "presión negativa", "filtracion", "filtración", "moho", "pared mojada", "muro mojado"]):
        if any(token in normalized for token in ["interior", "muro", "pared", "base del muro", "viene del piso", "jardinera", "sube del piso"]):
            if any(token in normalized for token in ["base del muro", "viene del piso", "jardinera", "capilaridad", "presion negativa", "presión negativa"]):
                return "humedad_interior_capilaridad"
            return "humedad_interior_general"
        if any(token in normalized for token in ["fachada", "exterior", "lluvia", "sol y lluvia"]):
            return "fachada_exterior"

    if any(token in normalized for token in ["piso", "garaje", "montacargas", "trafico", "tráfico", "estibador", "intergard 2002", "pintucoat", "intergard 740", "epoxico", "epóxico", "epoxi", "concreto nuevo", "recien fundido", "reci en fundido", "recien vaciado", "sin curar", "obra gris"]):
        return "piso_industrial"

    if any(token in normalized for token in ["fachada", "muro exterior", "exterior", "intemperie", "lluvia sol"]):
        return "fachada_exterior"

    if any(token in normalized for token in ["oxido", "óxido", "reja", "metal", "corrosion", "corrosión", "corrotec", "pintoxido"]):
        return "metal_oxidado"

    return None


def _estimate_problem_class_confidence(problem_class: Optional[str], question: str, product: str, best_similarity: float) -> str:
    if not problem_class:
        return "baja"

    normalized = normalize_text_value(f"{question} {product}")
    signal_count = 0
    signal_map = {
        "eternit_fibrocemento": ["eternit", "fibrocemento", "asbesto", "sellomax", "koraza"],
        "ladrillo_vista": ["ladrillo", "siliconite", "construcleaner"],
        "metal_pintado_alquidico": ["metal", "reja", "alquidico", "esmalte", "anticorrosivo viejo"],
        "humedad_interior_capilaridad": ["humedad", "salitre", "interior", "muro", "viene del piso", "jardinera", "capilaridad"],
        "humedad_interior_general": ["humedad", "salitre", "interior", "muro", "pared"],
        "fachada_exterior": ["fachada", "exterior", "intemperie", "lluvia"],
        "metal_oxidado": ["metal", "reja", "oxido", "óxido", "corrotec"],
        "piso_industrial": ["piso", "concreto", "trafico", "tráfico", "montacargas"],
        "madera": ["madera", "barnex", "barniz", "wood stain"],
    }
    for token in signal_map.get(problem_class, []):
        if token in normalized:
            signal_count += 1

    if best_similarity >= 0.82 or signal_count >= 5:
        return "alta"
    if best_similarity >= 0.68 or signal_count >= 3:
        return "media"
    return "baja"


def _build_structured_diagnosis(question: str, product: str, best_similarity: float) -> dict:
    problem_class = _infer_problem_class_from_rag_query(question, product)
    confidence = _estimate_problem_class_confidence(problem_class, question, product, best_similarity)
    normalized = normalize_text_value(f"{question} {product}")

    probable_cause = None
    if problem_class == "eternit_fibrocemento":
        probable_cause = "fibrocemento exterior sensible a polvo y envejecimiento"
    elif problem_class == "ladrillo_vista":
        probable_cause = "sustrato mineral expuesto que debe protegerse sin ocultar la textura"
    elif problem_class == "metal_pintado_alquidico":
        probable_cause = "incompatibilidad quimica entre base alquidica vieja y sistema industrial 2K"
    elif problem_class == "humedad_interior_capilaridad":
        probable_cause = "capilaridad/presión negativa"
    elif problem_class == "humedad_interior_general":
        probable_cause = "humedad interior por definir"
    elif problem_class == "fachada_exterior":
        probable_cause = "intemperismo/filtración exterior"
    elif problem_class == "metal_oxidado":
        probable_cause = "corrosión por exposición"
    elif problem_class == "piso_industrial":
        probable_cause = "desgaste mecánico / requerimiento por tráfico"

    required_validations_map = {
        "eternit_fibrocemento": [
            "Confirmar si el fibrocemento es exterior y si ya esta pintado o envejecido.",
            "Validar si hay polvo de asbesto o deterioro que obligue a preparacion humeda.",
            "Solicitar m2 reales antes de cotizar.",
        ],
        "ladrillo_vista": [
            "Confirmar si el cliente quiere conservar la apariencia natural del ladrillo.",
            "Validar si requiere solo limpieza o limpieza mas hidrofugacion.",
            "Solicitar m2 reales antes de cotizar.",
        ],
        "metal_pintado_alquidico": [
            "Confirmar si la base actual es esmalte sintetico, anticorrosivo alquidico o pintura de aceite.",
            "Validar si aceptan remocion total hasta metal desnudo.",
            "Solicitar m2 o dimensiones antes de cotizar.",
        ],
        "humedad_interior_capilaridad": [
            "Confirmar origen de humedad desde base del muro/piso/jardinera.",
            "Validar estado del revoque o base soplada.",
            "Solicitar m² reales antes de cotizar.",
        ],
        "humedad_interior_general": [
            "Confirmar causa: base del muro, arriba, lateral o temporada.",
            "Validar estado de la base/revoque.",
            "Solicitar m² reales antes de cotizar.",
        ],
        "fachada_exterior": [
            "Confirmar si es exterior real y nivel de deterioro.",
            "Solicitar m² reales antes de cotizar.",
        ],
        "metal_oxidado": [
            "Confirmar grado de oxidación.",
            "Confirmar si es interior o exterior.",
            "Solicitar m² o dimensiones antes de cotizar.",
        ],
        "piso_industrial": [
            "Confirmar si es concreto nuevo o viejo/ya pintado.",
            "Confirmar curado de 28 días si es nuevo.",
            "Confirmar tipo de tráfico y si es interior/exterior.",
            "Solicitar m² reales antes de cotizar.",
        ],
        "madera": [
            "Confirmar si es interior o exterior.",
            "Confirmar si quiere acabado transparente o color sólido.",
            "Solicitar área o dimensiones antes de cotizar.",
        ],
    }

    observed_signals = []
    for token in ["humedad", "salitre", "capilaridad", "jardinera", "muro", "interior", "fachada", "exterior", "oxido", "óxido", "reja", "piso", "montacargas", "madera", "barniz", "eternit", "fibrocemento", "ladrillo", "alquidico"]:
        if token in normalized:
            observed_signals.append(token)

    return {
        "problem_class": problem_class,
        "confidence": confidence,
        "probable_cause": probable_cause,
        "pricing_ready": False if problem_class in {"eternit_fibrocemento", "ladrillo_vista", "metal_pintado_alquidico", "humedad_interior_capilaridad", "humedad_interior_general", "fachada_exterior", "metal_oxidado", "piso_industrial", "madera"} else True,
        "required_validations": required_validations_map.get(problem_class, []),
        "observed_signals": observed_signals,
    }


def _build_structured_technical_guide(question: str, product: str, diagnosis: dict, expert_notes: list[dict], best_similarity: float) -> dict:
    problem_class = diagnosis.get("problem_class")
    normalized = normalize_text_value(f"{question} {product}")

    guide = {
        "problem_class": problem_class,
        "source_confidence": diagnosis.get("confidence") or _estimate_problem_class_confidence(problem_class, question, product, best_similarity),
        "preparation_steps": [],
        "base_or_primer": [],
        "intermediate_steps": [],
        "finish_options": [],
        "diluents_or_adjusters": [],
        "tools": [],
        "required_questions": diagnosis.get("required_validations") or [],
        "forbidden_products_or_shortcuts": [],
        "commercial_alternatives": [],
        "pricing_gate": "m2_required" if problem_class in {"eternit_fibrocemento", "ladrillo_vista", "metal_pintado_alquidico", "humedad_interior_capilaridad", "humedad_interior_general", "fachada_exterior", "metal_oxidado", "piso_industrial", "madera"} else "none",
        "hard_rules_applied": [],
    }

    if problem_class == "eternit_fibrocemento":
        guide["preparation_steps"] = [
            "Preparacion humeda con hidrolavadora, jabon, hipoclorito y cepillo; nunca lijar en seco ni rasquetear.",
            "Retirar solo material flojo sin generar polvo.",
        ]
        guide["base_or_primer"] = ["Sellomax antes del acabado si el eternit ya esta pintado o envejecido."]
        guide["finish_options"] = [
            {"producto": "Koraza", "rol": "acabado exterior", "nivel": "premium"},
        ]
        guide["tools"] = ["Hidrolavadora", "Cepillo", "Escoba de cerdas duras", "Brocha", "Rodillo"]
        guide["forbidden_products_or_shortcuts"] = [
            "Intervinil, Pinturama o vinilos interiores como acabado exterior.",
            "Lijado en seco, rasqueteo o preparacion mecanica que genere polvo.",
        ]
        guide["hard_rules_applied"] = [
            "Fibrocemento exterior: preparacion humeda obligatoria + Sellomax + acabado exterior.",
        ]
    elif problem_class == "ladrillo_vista":
        guide["preparation_steps"] = [
            "Limpieza tecnica del ladrillo antes de protegerlo.",
        ]
        guide["base_or_primer"] = ["Construcleaner Limpiador Desengrasante como limpieza previa."]
        guide["finish_options"] = [
            {"producto": "Siliconite 7", "rol": "hidrofugante", "nivel": "premium"},
        ]
        guide["tools"] = ["Cepillo", "Brocha", "Rodillo segun absorcion"]
        guide["forbidden_products_or_shortcuts"] = [
            "Acido muriatico para limpieza.",
            "Koraza si el objetivo es conservar el ladrillo a la vista.",
        ]
    elif problem_class == "metal_pintado_alquidico":
        guide["preparation_steps"] = [
            "Remocion total hasta metal desnudo antes de migrar a sistema epoxico o poliuretano.",
        ]
        guide["base_or_primer"] = ["Corrotec si se mantiene sistema alquidico. Wash Primer o sistema epoxico solo despues de remocion total."]
        guide["finish_options"] = [
            {"producto": "Pintulux 3 en 1", "rol": "acabado compatible si se mantiene familia alquidica", "nivel": "estándar"},
        ]
        guide["tools"] = ["Disco flap", "Grata", "Lija Abracol", "Brocha Goya Profesional"]
        guide["forbidden_products_or_shortcuts"] = [
            "Aplicar epoxicos o poliuretanos directamente sobre esmalte sintetico o anticorrosivo alquidico viejo.",
        ]
    elif problem_class == "humedad_interior_capilaridad":
        guide["preparation_steps"] = [
            "Remover por completo pintura soplada/descascarada y salitre hasta base sana.",
            "Si el revoque está quemado o meteorizado, reemplazarlo antes del sistema nuevo.",
        ]
        guide["base_or_primer"] = ["Aquablock Ultra - 2 manos con brocha para cargar producto."]
        guide["intermediate_steps"] = ["Estuco Acrílico después del Aquablock para nivelar. NUNCA antes."]
        guide["finish_options"] = [
            {"producto": "Viniltex Advanced", "rol": "acabado final", "nivel": "premium"},
            {"producto": "Intervinil", "rol": "acabado final", "nivel": "intermedio"},
            {"producto": "Pinturama", "rol": "acabado final", "nivel": "económico"},
        ]
        guide["tools"] = ["Brocha Goya Profesional", "Rodillo", "Lija / raspado para preparación"]
        guide["forbidden_products_or_shortcuts"] = [
            "Koraza como imprimante o acabado interior.",
            "Pintuco Fill como solución principal para capilaridad interior desde la base del muro.",
            "Cotizar por galones sugeridos por el cliente sin metraje.",
        ]
        guide["hard_rules_applied"] = [
            PRODUCT_TECHNICAL_HARD_RULES["koraza"]["no_es_para"],
            PRODUCT_TECHNICAL_HARD_RULES["aquablock"]["es_para"],
        ]
    elif problem_class == "humedad_interior_general":
        guide["preparation_steps"] = [
            "Diagnosticar causa de humedad antes de pintar.",
            "Remover base dañada y salitre donde aplique.",
        ]
        guide["base_or_primer"] = ["Aquablock / Aquablock Ultra según presión negativa y severidad."]
        guide["intermediate_steps"] = ["Estuco Acrílico si se requiere nivelación después del bloqueador de humedad."]
        guide["finish_options"] = [
            {"producto": "Viniltex Advanced", "rol": "acabado final", "nivel": "premium"},
            {"producto": "Intervinil", "rol": "acabado final", "nivel": "intermedio"},
            {"producto": "Pinturama", "rol": "acabado final", "nivel": "económico"},
        ]
        guide["tools"] = ["Brocha", "Rodillo", "Lija / raspado para preparación"]
        guide["forbidden_products_or_shortcuts"] = ["Koraza como sellador de humedad interior."]
        guide["hard_rules_applied"] = [PRODUCT_TECHNICAL_HARD_RULES["koraza"]["no_es_para"]]
    elif problem_class == "fachada_exterior":
        guide["preparation_steps"] = ["Remover pintura suelta o base soplada antes de repintar."]
        guide["finish_options"] = [
            {"producto": "Koraza", "rol": "acabado exterior", "nivel": "premium"},
            {"producto": "Viniltex", "rol": "acabado exterior cuando la exposicion es moderada y el sistema es arquitectonico compatible", "nivel": "intermedio"},
        ]
        guide["tools"] = ["Lija Abracol", "Brocha Goya Profesional", "Rodillo"]
        guide["forbidden_products_or_shortcuts"] = [
            "Intervinil o Pinturama como acabado en fachadas de alta exposicion.",
            "Aquablock como acabado exterior.",
        ]
    elif problem_class == "metal_oxidado":
        guide["preparation_steps"] = ["Preparación mecánica con lija, disco flap o grata según el grado de óxido."]
        guide["base_or_primer"] = ["Pintóxido si hay óxido profundo.", "Corrotec o Corrotec Premium como anticorrosivo."]
        guide["finish_options"] = [{"producto": "Pintulux 3 en 1", "rol": "acabado final", "nivel": "estándar"}]
        guide["tools"] = ["Disco flap", "Grata", "Brocha Goya Profesional", "Lija Abracol"]
    elif problem_class == "piso_industrial":
        guide["preparation_steps"] = ["Confirmar estado del piso y preparación mecánica adecuada."]
        guide["base_or_primer"] = ["Interseal gris RAL 7038 para concreto cuando aplique."]
        guide["finish_options"] = [
            {"producto": "Pintucoat", "rol": "acabado para tráfico medio", "nivel": "medio"},
            {"producto": "Intergard 740", "rol": "acabado brillante tráfico medio", "nivel": "medio"},
            {"producto": "Intergard 2002 + cuarzo", "rol": "sistema tráfico pesado", "nivel": "alto desempeño"},
        ]
        guide["forbidden_products_or_shortcuts"] = ["No cotizar sin m² ni sin protocolo diagnóstico del piso."]
    elif problem_class == "madera":
        guide["preparation_steps"] = ["Diagnosticar si es interior/exterior y si quiere transparente o color sólido."]
        guide["finish_options"] = [
            {"producto": "Barnex", "rol": "acabado exterior transparente", "nivel": "premium"},
            {"producto": "Wood Stain", "rol": "acabado exterior transparente", "nivel": "intermedio"},
            {"producto": "Esmalte Doméstico", "rol": "acabado color sólido", "nivel": "económico"},
            {"producto": "Pintulux Máxima Protección", "rol": "acabado color sólido", "nivel": "premium"},
        ]
        guide["tools"] = ["Brocha Goya Profesional", "Lijas Abracol 80-100 y 220-320", "Removedor Pintuco"]

    if "interthane" in normalized:
        guide["diluents_or_adjusters"].append("UFA151 como ajustador/diluyente del sistema de poliuretano cuando aplique.")
        guide["hard_rules_applied"].append(PRODUCT_TECHNICAL_HARD_RULES["interthane"]["bicomponente"])
    if "pintucoat" in normalized:
        guide["diluents_or_adjusters"].append("Thinner Epóxico Pintuco cuando el sistema lo requiera.")
        guide["hard_rules_applied"].append(PRODUCT_TECHNICAL_HARD_RULES["pintucoat"]["bicomponente"])

    if expert_notes:
        guide["expert_overrides"] = [
            {
                "tipo": note.get("tipo"),
                "recomendar": note.get("producto_recomendado"),
                "evitar": note.get("producto_desestimado"),
                "nota": note.get("nota_comercial"),
            }
            for note in expert_notes[:5]
        ]
    else:
        guide["expert_overrides"] = []

    return guide


def _split_policy_items(raw_value: Optional[str]) -> list[str]:
    if not raw_value:
        return []
    cleaned = raw_value.replace("\n", ",")
    chunks = re.split(r"[;,]|\s+\+\s+|\s+y\s+", cleaned, flags=re.IGNORECASE)
    results = []
    for chunk in chunks:
        value = (chunk or "").strip(" .:-")
        normalized = normalize_text_value(value)
        if len(normalized) < 3:
            continue
        if value not in results:
            results.append(value)
    return results


def _is_tool_policy_item(item: str) -> bool:
    normalized = normalize_text_value(item)
    tool_tokens = {
        "hidrolavadora", "escoba", "cepillo", "brocha", "rodillo", "lija", "lijas",
        "rasqueta", "espatula", "espátula", "grata", "disco flap", "pulidora",
        "pistola", "airless", "thinner", "solvente", "jabón", "jabon", "hipoclorito",
    }
    return any(token in normalized for token in tool_tokens)


def _extract_forbidden_note_items(note_text: str) -> list[str]:
    items = []
    patterns = [
        r"nunca\s+(?:recomendar|usar|aplicar|incluir|listar ni incluir)\s+(.+?)(?:\.|$)",
        r"prohibido\s+(?:usar|recomendar|incluir)\s+(.+?)(?:\.|$)",
        r"evitar\s+(.+?)(?:\.|$)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, note_text or "", flags=re.IGNORECASE):
            for item in _split_policy_items(match.group(1)):
                if item not in items:
                    items.append(item)
    return items


def _build_hard_policies_for_context(question: str, product: str, diagnosis: dict, guide: dict, expert_notes: list[dict]) -> dict:
    policies = {
        "problem_class": diagnosis.get("problem_class"),
        "required_products": [],
        "forbidden_products": [],
        "required_tools": [],
        "forbidden_tools": [],
        "mandatory_steps": [],
        "mandatory_step_signals": [],
        "rules_text": [],
        "policy_names": [],
        "critical_policy_names": [],
        "high_priority_policy_names": [],
        "dominant_policy_names": [],
        "highest_priority_level": "none",
    }

    def _append_unique(bucket: str, value: str):
        cleaned = (value or "").strip()
        if not cleaned:
            return
        if cleaned not in policies[bucket]:
            policies[bucket].append(cleaned)

    for step in (guide.get("preparation_steps") or []):
        _append_unique("mandatory_steps", step)

    for step in (guide.get("preparation_steps") or []):
        lowered_step = normalize_text_value(step)
        for candidate in ["preparacion humeda", "sellomax", "koraza", "aquablock", "retirar el acabado", "metal desnudo", "intergard 2002", "cuarzo", "interthane", "28 dias", "curado", "construcleaner", "siliconite", "barnex", "wood stain", "poliuretano alto trafico"]:
            if candidate in lowered_step:
                _append_unique("mandatory_step_signals", candidate)

    for note in expert_notes or []:
        note_text = (note.get("nota_comercial") or "").strip()
        if note_text:
            _append_unique("rules_text", note_text)

        for item in _split_policy_items(note.get("producto_recomendado")):
            bucket = "required_tools" if _is_tool_policy_item(item) else "required_products"
            _append_unique(bucket, item)

        explicit_avoid_items = _split_policy_items(note.get("producto_desestimado"))
        note_avoid_items = _extract_forbidden_note_items(note_text)
        for item in explicit_avoid_items + note_avoid_items:
            bucket = "forbidden_tools" if _is_tool_policy_item(item) else "forbidden_products"
            _append_unique(bucket, item)

        normalized_note = normalize_text_value(note_text)
        for candidate in ["preparacion humeda", "sellomax", "koraza", "aquablock", "metal desnudo", "intergard 2002", "cuarzo", "interthane", "28 dias", "curado", "construcleaner", "siliconite", "barnex", "wood stain", "poliuretano alto trafico", "misma familia", "agua con agua"]:
            if candidate in normalized_note:
                _append_unique("mandatory_step_signals", candidate)

    for forbidden in (guide.get("forbidden_products_or_shortcuts") or []):
        _append_unique("rules_text", forbidden)
        for item in _split_policy_items(forbidden):
            bucket = "forbidden_tools" if _is_tool_policy_item(item) else "forbidden_products"
            _append_unique(bucket, item)

    normalized_query = normalize_text_value(f"{question} {product}")
    for rule in GLOBAL_TECHNICAL_POLICY_RULES:
        if not _matches_global_policy_rule(rule, normalized_query, diagnosis):
            continue
        rule_name = rule.get("name") or "regla_contextual"
        _append_unique("policy_names", rule_name)
        priority = normalize_text_value(rule.get("priority") or "normal")
        if priority == "critical":
            _append_unique("critical_policy_names", rule_name)
        elif priority == "high":
            _append_unique("high_priority_policy_names", rule_name)
        for value in rule.get("required_products") or []:
            _append_unique("required_products", value)
        for value in rule.get("forbidden_products") or []:
            _append_unique("forbidden_products", value)
        for value in rule.get("required_tools") or []:
            _append_unique("required_tools", value)
        for value in rule.get("forbidden_tools") or []:
            _append_unique("forbidden_tools", value)
        for value in rule.get("mandatory_steps") or []:
            _append_unique("mandatory_steps", value)
        for value in rule.get("mandatory_step_signals") or []:
            _append_unique("mandatory_step_signals", value)
        for value in rule.get("rules_text") or []:
            _append_unique("rules_text", value)

    if any(token in normalized_query for token in ["eternit", "fibrocemento", "asbesto"]):
        _append_unique("mandatory_steps", "Preparación húmeda obligatoria; nunca lijar en seco ni rasquetear.")
        _append_unique("mandatory_step_signals", "preparacion humeda")

    if policies["critical_policy_names"]:
        policies["dominant_policy_names"] = list(policies["critical_policy_names"])
        policies["highest_priority_level"] = "critical"
    elif policies["high_priority_policy_names"]:
        policies["dominant_policy_names"] = list(policies["high_priority_policy_names"])
        policies["highest_priority_level"] = "high"
    elif policies["policy_names"]:
        policies["dominant_policy_names"] = [policies["policy_names"][0]]
        policies["highest_priority_level"] = "normal"

    return policies

# ── Catálogo verificado de productos bicomponentes ────────────────────────
# Fuente de verdad interna: los catalizadores y proporciones aquí registrados
# PREVALECEN sobre cualquier respuesta del RAG o memoria del LLM.
# Si el RAG no confirma la relación, el agente DEBE citar este catálogo.
BICOMPONENT_CATALOG: dict[str, dict] = {
    # ─ Pintucoat (Pintuco) ─────────────────────────────────────────────────
    # galón COMP A (3.44L, ref 516 o 517) → catalizador 13227 COMP B 0.37L (1/8)
    # cuñete COMP A (15.14L) → catalizador 13227 COMP B 1.89L
    "pintucoat": {
        "tipo_sistema": "epoxica_dos_componentes",
        "componente_a_descripcion": "Pintucoat 516 o 517 COMP A (base de color)",
        "componente_b_codigo": "13227",
        "componente_b_descripcion": "Pintucoat COMP B catalizador",
        "proporcion_galon": "COMP A 3.44L + catalizador 13227 COMP B 0.37L (1/8 de galón)",
        "proporcion_cunete": "COMP A 15.14L + catalizador 13227 COMP B 1.89L",
        "pot_life_horas": 6,
        "restriccion_exterior": (
            "Pintucoat es epóxico y ENTIZA (se decolora) en exteriores expuestos al sol. "
            "En exterior REQUIERE capa de acabado con poliuretano (Interthane). "
            "NUNCA ofrecer Pintulux 3en1 como acabado sobre Pintucoat: Pintulux es esmalte alquídico, "
            "no es poliuretano y no da la resistencia UV requerida."
        ),
        "acabado_exterior_obligatorio": "interthane",
        "resistencia": "media",
        "acabado": "mate",
        "uso_piso": True,
        "nota_resistencia": (
            "Pintucoat es de resistencia MEDIA. NO resiste tráfico pesado de montacargas/estibadores. "
            "Para tráfico pesado recomendar Intergard 2002 + cuarzo (ref 5891610). "
            "Alternativa brillante de resistencia media: Intergard 740."
        ),
    },
    # ─ Interthane (International / AkzoNobel) ─────────────────────────────
    # galón (3.7L, PHA120 o PHA130) → catalizador PHA046 0.5L
    # cuñete (20L) → catalizador PHA046 3.7L
    "interthane": {
        "tipo_sistema": "poliuretano_dos_componentes",
        "componente_a_descripcion": "Interthane 990 COMP A (color, ej. PHA120 o PHA130)",
        "componente_b_codigo": "PHA046",
        "componente_b_descripcion": "Interthane 990 PHA046 catalizador (hardener)",
        "proporcion_galon": "COMP A 3.7L + catalizador PHA046 0.5L",
        "proporcion_cunete": "COMP A 20L + catalizador PHA046 3.7L",
        "nota": "Verificar relación exacta en ficha técnica según número de lote y temperatura.",
    },
    # ─ Interseal (International / AkzoNobel) ──────────────────────────────
    "interseal": {
        "tipo_sistema": "epoxica_dos_componentes",
        "componente_a_descripcion": "Interseal COMP A",
        "componente_b_descripcion": "Interseal COMP B catalizador — consultar ficha técnica Internacional",
        "nota": "Relación de mezcla y código de catalizador deben extraerse de la ficha técnica International o la Guía de Sistemas.",
        "aplicacion_condicional_agua_potable": (
            "Interseal 670HS tiene certificación NSF/ANSI 61 para agua potable en tanques > 100 gal (378.5L). "
            "Condiciones obligatorias: (1) preparación Sa 2.5 / SSPC-SP10, (2) colores específicos certificados "
            "(verificar lote con distribuidor), (3) respetar tiempo de curado completo antes de servicio. "
            "Alternativa de mayor desempeño en inmersión permanente: línea Interline (100% sólidos, sin solventes)."
        ),
    },
    # ─ Intergard (International / AkzoNobel) ── GENÉRICO ────────────────────
    "intergard": {
        "tipo_sistema": "epoxica_dos_componentes",
        "componente_a_descripcion": "Intergard COMP A (primer epóxico)",
        "componente_b_descripcion": "Intergard COMP B catalizador — consultar ficha técnica International",
        "nota": "Relación de mezcla y código de catalizador deben extraerse de la ficha técnica International o la Guía de Sistemas.",
    },
    # ─ Intergard 740 (International / AkzoNobel) ── PISOS TRÁFICO MEDIO ACABADO BRILLANTE ──
    "intergard 740": {
        "tipo_sistema": "epoxica_dos_componentes",
        "componente_a_descripcion": "Intergard 740 COMP A (acabado brillante)",
        "componente_b_descripcion": "Intergard 740 COMP B catalizador — consultar ficha técnica International",
        "nota": "Epóxico para pisos de tráfico MEDIO con acabado BRILLANTE. Alternativa al Pintucoat cuando el cliente quiere más brillo.",
        "resistencia": "media",
        "acabado": "brillante",
        "uso_piso": True,
    },
    # ─ Intergard 2002 (International / AkzoNobel) ── SOBRE PEDIDO — ESCALAR A ASESOR ──
    "intergard 2002": {
        "tipo_sistema": "epoxica_dos_componentes",
        "componente_a_descripcion": "Intergard 2002 COMP A (alto volumen de sólidos)",
        "componente_b_descripcion": "Intergard 2002 COMP B catalizador — consultar ficha técnica International",
        "sobre_pedido": True,
        "nota": (
            "⚠️ PRODUCTO SOBRE PEDIDO — NO cotizar precio, NO buscar inventario. "
            "Intergard 2002 es un sistema especializado para pisos de tráfico PESADO (montacargas, estibadores) "
            "que requiere asesoría técnica personalizada. ESCALAR al Asesor Técnico Comercial. "
            "Sistema referencial: Interseal gris RAL 7038 (imprimante) → Intergard 2002 + cuarzo ref 5891610 → sello opcional."
        ),
        "resistencia": "alta (con cuarzo)",
        "acabado": "mate/satinado",
        "uso_piso": True,
        "cuarzo_ref": "5891610",
    },
    # ─ Interfine (International / AkzoNobel) ──────────────────────────────
    "interfine": {
        "tipo_sistema": "poliuretano_dos_componentes",
        "componente_a_descripcion": "Interfine 979 COMP A",
        "componente_b_descripcion": "Interfine COMP B catalizador — consultar ficha técnica International",
        "nota": "Relación de mezcla y código de catalizador deben extraerse de la ficha técnica International.",
    },
}

# Alias rápidos para buscar si un producto cae en BICOMPONENT_CATALOG
_BICOMPONENT_KEYWORDS: frozenset[str] = frozenset(BICOMPONENT_CATALOG.keys()) | frozenset([
    "pintucoat 516", "pintucoat 517", "pintucoat plus",
    "interthane 990", "interseal 670", "intergard 475",
    "dos componentes", "bicomponente", "comp a", "comp b",
    "pha046", "pha120", "pha130",
    "catalizador 13227",
    "1550", "1551", "poliuretano alto trafico", "pisos trafic alt",
    "vitrificar", "vitrificar piso",
])


def get_bicomponent_info(product_name_or_query: str) -> dict | None:
    """Return bicomponent catalog entry if the query matches a known 2-component product."""
    q = normalize_text_value(product_name_or_query)
    for key, info in BICOMPONENT_CATALOG.items():
        if key in q:
            return {"producto_base": key, **info}
    return None


DIRECTION_ALIASES = {
    "derecha": ["derecha", "derecho", "der", "derc"],
    "izquierda": ["izquierda", "izquierdo", "izq"],
}


STORE_CODE_LABELS = {
    "155": "CEDI",
    "156": "San Francisco - Armenia",
    "157": "San Antonio - Manizales",
    "158": "Ópalo - Dosquebradas",
    "189": "Parque Olaya - Pereira",
    "238": "Laureles",
    "439": "FerreBOX - Pereira",
    "463": "Cerritos",
}


STORE_ALIASES = {
    "cedi": ["155", "cedi", "centro de distribucion", "centro de distribución", "pereira cedi"],
    "armenia": ["156", "armenia", "tienda armenia", "san francisco"],
    "manizales": ["157", "manizales", "tienda manizales", "san antonio"],
    "opalo": ["158", "opalo", "ópalo", "tienda opalo", "tienda ópalo", "dosquebradas"],
    "pereira": ["189", "pereira", "tienda pereira", "parque olaya", "olaya"],
    "laures": ["238", "laures", "laureles", "tienda laures", "tienda laureles"],
    "cerritos": ["463", "cerritos", "tienda cerritos"],
    "ferrebox": ["439", "ferrebox", "ferre box", "tienda ferrebox"],
}


BRAND_ALIASES = {
    # Marcas madre → sub-marcas que pertenecen a ese fabricante
    "viniltex": ["viniltex", "viniltex adv", "viniltex advanced", "vtx", "vinil plus"],
    "domestico": ["domestico", "doméstico", "blanca economica", "blanca económica", "vinilo barato", "p11", "p-11", "p 11", "esmalte domestico"],
    "pintulux": ["pintulux", "pintulux 3en1", "pintulux 3 en 1", "t11", "t-11", "t 11", "tu11", "tu-11", "tu 11", "teu11", "teu-11", "teu 11", "esmalte pintulux", "esmalte exterior"],
    "koraza": ["koraza", "koraza elastomerica", "koraza elastomerico", "koraza xp"],
    "pintuco": ["pintuco", "viniltex", "domestico", "doméstico", "pintulux", "koraza", "aerocolor", "pintucoat",
                "pintura canchas", "corrotec", "pintulac", "pinturama", "intervinil", "vinil latex", "vinil max",
                "icolatex", "vinilux", "vinil plus", "world color", "pintuco fill"],
    "pintucoat": ["pintucoat", "epoxica pintuco", "epoxi pintuco", "recubrimiento epoxica"],
    "pintura canchas": ["pintura canchas", "pintura pisos pintuco", "pintura pisos"],
    "corrotec": ["corrotec", "corrotec premium", "anticorrosivo pintuco"],
    "aerocolor": ["aerocolor", "aerosol pintuco", "spray pintuco"],
    "pintulac": ["pintulac", "laca pintuco", "laca"],
    "intervinil": ["intervinil", "inter vinil"],
    "vinil latex": ["vinil latex", "vinil látex", "vinillatex"],
    "vinilux": ["vinilux"],
    "pinturama": ["pinturama"],
    "vinil max": ["vinil max", "vinilmax"],
    "icolatex": ["icolatex", "ico latex"],
    "vinil plus": ["vinil plus", "vinilplus"],
    "international": ["international", "interseal", "interthane", "intergard", "interfine", "interchar"],
    "akzonobel": ["akzonobel", "international", "interseal", "interthane", "intergard"],
    "interseal": ["interseal", "epoxica international"],
    "interthane": ["interthane", "poliuretano international"],
    "intergard": ["intergard", "primer intergard"],
    "interchar": ["interchar", "intumescente international"],
    "abracol": ["abracol"],
    "yale": ["yale"],
    "goya": ["goya"],
    "smith": ["smith"],
    "afix": ["afix"],
    "segurex": ["segurex"],
    "artecola": ["artecola", "pl285", "pl 285"],
    "montana": ["montana", "montana 94"],
    "sika": ["sika", "sikaflex", "sikaguard", "sikalastic"],
}


MONTH_ALIASES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


PURCHASE_LINE_FILTER = """
COALESCE(valor_venta_neto, 0) > 0
AND COALESCE(unidades_vendidas_netas, 0) > 0
AND COALESCE(nombre_articulo, '') <> ''
AND COALESCE(nombre_articulo, '') NOT ILIKE '%TOTAL%'
AND COALESCE(nombre_articulo, '') NOT ILIKE '%NOTA CREDITO%'
"""


def get_postgrest_url():
    return os.getenv("PGRST_URL", "http://localhost:3000").rstrip("/")


def _read_streamlit_secret_value(*keys: str) -> Optional[str]:
    secrets_path = Path(__file__).resolve().parent.parent / ".streamlit" / "secrets.toml"
    if not secrets_path.exists() or not keys:
        return None
    try:
        raw_text = secrets_path.read_text(encoding="utf-8")
    except Exception:
        return None

    last_key = re.escape(keys[-1])
    quoted_match = re.search(rf"(?mi)^\s*{last_key}\s*=\s*\"([^\"]+)\"\s*$", raw_text)
    if quoted_match:
        return quoted_match.group(1).strip()

    bare_match = re.search(rf"(?mi)^\s*{last_key}\s*=\s*([^#\r\n]+)", raw_text)
    if bare_match:
        return bare_match.group(1).strip().strip('"').strip("'")

    try:
        parsed = tomllib.loads(raw_text)
    except Exception:
        return None

    current = parsed
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
        if current is None:
            return None
    if isinstance(current, str):
        return current.strip()
    return None


def get_database_url():
    database_url = (
        os.getenv("DATABASE_URL")
        or os.getenv("POSTGRES_DB_URI")
        or _read_streamlit_secret_value("DATABASE_URL")
        or _read_streamlit_secret_value("postgres", "db_uri")
    )
    if not database_url:
        raise RuntimeError("No se encontró DATABASE_URL o POSTGRES_DB_URI para el backend.")
    return database_url


def get_whatsapp_verify_token():
    return os.getenv("WHATSAPP_VERIFY_TOKEN", "ferreinox-verify-token")


def get_openai_api_key():
    return os.getenv("OPENAI_API_KEY") or _read_streamlit_secret_value("openai", "api_key")


def get_openai_model():
    return os.getenv("OPENAI_MODEL", "gpt-4o-mini")


def get_whatsapp_access_token():
    token = os.getenv("WHATSAPP_ACCESS_TOKEN")
    if not token:
        raise RuntimeError("No se encontró WHATSAPP_ACCESS_TOKEN para enviar mensajes.")
    return token


def get_whatsapp_phone_number_id():
    phone_number_id = os.getenv("WHATSAPP_PHONE_NUMBER_ID")
    if not phone_number_id:
        raise RuntimeError("No se encontró WHATSAPP_PHONE_NUMBER_ID para enviar mensajes.")
    return phone_number_id


def get_openai_client():
    api_key = get_openai_api_key()
    if not api_key:
        raise RuntimeError("No se encontró OPENAI_API_KEY para generar respuestas del agente.")
    return OpenAI(api_key=api_key)


def get_product_nlu_system_prompt():
    global PRODUCT_NLU_PROMPT_CACHE
    if PRODUCT_NLU_PROMPT_CACHE is not None:
        return PRODUCT_NLU_PROMPT_CACHE

    try:
        PRODUCT_NLU_PROMPT_CACHE = PRODUCT_NLU_PROMPT_PATH.read_text(encoding="utf-8").strip()
    except Exception:
        PRODUCT_NLU_PROMPT_CACHE = PRODUCT_NLU_PROMPT_FALLBACK

    if not PRODUCT_NLU_PROMPT_CACHE:
        PRODUCT_NLU_PROMPT_CACHE = PRODUCT_NLU_PROMPT_FALLBACK
    return PRODUCT_NLU_PROMPT_CACHE


def normalize_nullable_phrase(raw_value):
    normalized = normalize_text_value(raw_value)
    if normalized in {"", "null", "none", "ninguno", "ninguna", "n/a", "na", "sin", "no aplica"}:
        return None
    return normalized


def canonicalize_presentation_value(raw_value):
    normalized = normalize_text_value(raw_value)
    if not normalized:
        return None
    for canonical_value, aliases in PRESENTATION_ALIASES.items():
        alias_values = {normalize_text_value(alias) for alias in aliases}
        alias_values.add(normalize_text_value(canonical_value))
        if normalized in alias_values:
            return canonical_value
    return None


def merge_unique_terms(*term_groups):
    merged_terms = []
    seen_terms = set()
    for group in term_groups:
        if not group:
            continue
        if isinstance(group, str):
            candidates = [group]
        else:
            candidates = list(group)
        for value in candidates:
            normalized = normalize_text_value(value)
            if not normalized or normalized in seen_terms:
                continue
            seen_terms.add(normalized)
            merged_terms.append(normalized)
    return merged_terms


def tokenize_search_phrase(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return []
    return [
        token
        for token in re.findall(r"[a-z0-9.-]+", normalized)
        if len(token) >= 2 and token not in PRODUCT_STOPWORDS and not is_store_alias_term(token)
    ]


# ── Abreviaciones de mostrador que los clientes usan sin el término completo ──────────────────
# Pares (regex, reemplazo) aplicados sobre el texto normalizado antes de cualquier búsqueda.
# Esto permite que "brocha profe" encuentre "BROCHA PROFESIONAL", etc.
FERRETERIA_WORD_EXPANSIONS: list[tuple[str, str]] = [
    (r"\bprofe\b", "profesional"),
    (r"\bpopu\b", "popular"),
    (r"\bprof\.?(?=\s|$)", "profesional"),   # "prof." o "prof " al final de palabra
    (r"\bbarni\b", "barniz"),
    (r"\besmal\b", "esmalte"),
    # ── Fracciones de pulgada pegadas (ERP: "11/2" = 1½", "21/2" = 2½") ──
    (r"\b([1-4])1/2\b", r"\1 1/2"),   # "11/2" → "1 1/2", "21/2" → "2 1/2", etc.
    (r"\bvini\b", "viniltex"),
    # ── Diminutivos y jerga coloquial ferretera ──
    (r"\bbrochitas?\b", "brocha"),
    (r"\bpinceles?\b", "brocha"),
    (r"\btarritos?\b", "cuarto"),
    (r"\bcuñeticos?\b", "cuñete"),
    (r"\btarros?\s+pequeños?\b", "cuarto"),
    (r"\btarros?\s+grandes?\b", "cuñete"),
    # ── Nombres coloquiales → marca/producto ──
    (r"\bpintura\s+lavable\b", "viniltex"),
    (r"\bsatinado\b", "acriltex"),
    (r"\bacabado\s+satinado\b", "acriltex"),
    (r"\bpintura\s+satinada\b", "acriltex"),
]

# ── Bidirectional search-term variants ──────────────────────────────────────
# When a term appears in a search query, these variants are ALSO searched via ILIKE.
# This handles the DB having "PROF." when user says "profesional" and vice-versa,
# as well as common ERP truncations.
_SEARCH_TERM_VARIANTS: dict[str, list[str]] = {
    "profesional": ["prof", "profes", "profesio"],
    "prof":        ["profesional", "profes"],
    "popular":     ["popu", "popul"],
    "popu":        ["popular", "popul"],
    "barniz":      ["barni", "barn"],
    "esmalte":     ["esmal"],
    "brillante":   ["brill", "br"],
    "mate":        ["mat"],
    "satinado":    ["sat", "satin"],
    "incoloro":    ["incol"],
    "galones":     ["gal", "galon"],
    "galon":       ["gal", "galones"],
    "cuartos":     ["cuarto", "cto"],
    "cuarto":      ["cuartos", "cto"],
    "pulgadas":    ["pulg", "pulgada"],
    "pulgada":     ["pulg", "pulgadas"],
    "acrilico":    ["acril"],
    "anticorrosivo": ["anticorr", "anticorrosi"],
    "impermeabilizante": ["imperm", "impermeab"],
}


def expand_search_terms_with_variants(query_terms: list[str]) -> list[str]:
    """Given query terms, expand them with DB abbreviation variants for broader matching."""
    expanded = list(query_terms)
    for term in query_terms:
        low = term.lower()
        variants = _SEARCH_TERM_VARIANTS.get(low, [])
        for v in variants:
            upper_v = v.upper()
            if upper_v not in expanded:
                expanded.append(upper_v)
    return expanded


def expand_ferreteria_text(text_value: Optional[str]) -> str:
    """Expande abreviaciones de ferretería antes de normalizar para búsqueda."""
    normalized = normalize_text_value(text_value) or ""
    for pattern, replacement in FERRETERIA_WORD_EXPANSIONS:
        normalized = re.sub(pattern, replacement, normalized)
    return normalized


def translate_customer_jargon(raw_query: str) -> str:
    """Traduce jerga coloquial del cliente a términos de catálogo antes de consultar inventario.

    Esta función se ejecuta ANTES de apply_deterministic_product_alias_rules y realiza
    traducciones simples de texto que el modelo gpt-4o-mini no necesita hacer.
    """
    if not raw_query:
        return raw_query
    text = raw_query.strip()
    # Aplicar expansiones de ferretería (diminutivos, abreviaciones, saturado→acriltex, etc.)
    text = expand_ferreteria_text(text)
    # Diminutivos genéricos residuales (-itas, -itos → base)
    text = re.sub(r'(\w{4,}?)itas?\b', r'\1a', text)
    text = re.sub(r'(\w{4,}?)itos?\b', r'\1o', text)
    return text


# ── Códigos TEU/T-XX de Pintulux 3en1 → colores canónicos ──────────────────────────────────────
# La tabla articulos_maestro tiene estas claves en descripcion_adicional (p.ej. T-95 → negro 95).
# Aquí las mapeamos a términos de búsqueda canonizados para que el agente encuentre el producto correcto
# aunque el cliente use el código corto.
PINTULUX_TEU_COLOR_CODES: dict[str, dict] = {
    # Brillantes
    r"\b(?:t(?:eu?|u)?)-?\s*11\b":  {"color": "blanco",           "code": "11",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*95\b":  {"color": "negro",            "code": "95",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*84\b":  {"color": "gris plata",       "code": "84",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*76\b":  {"color": "caoba",            "code": "76",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*80\b":  {"color": "verde bronce",     "code": "80",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*53\b":  {"color": "verde esmeralda",  "code": "53",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*18\b":  {"color": "amarillo",         "code": "18",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*20\b":  {"color": "naranja",          "code": "20",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*26\b":  {"color": "rojo bermellon",   "code": "26",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*40\b":  {"color": "azul espanol",     "code": "40",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*12\b":  {"color": "crema",            "code": "12",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*5\b":   {"color": "marfil",           "code": "5",   "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*29\b":  {"color": "rojo",             "code": "29",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*35\b":  {"color": "azul claro",       "code": "35",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*38\b":  {"color": "azul mediano",     "code": "38",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*44\b":  {"color": "verde turquesa",   "code": "44",  "acabado": "brillante"},
    r"\b(?:t(?:eu?|u)?)-?\s*47\b":  {"color": "verde maquina",    "code": "47",  "acabado": "brillante"},
    # Mate
    r"\b(?:t(?:eu?|u)?)-?\s*10\b":  {"color": "blanco mate",      "code": "10",  "acabado": "mate"},
    r"\b(?:t(?:eu?|u)?)-?\s*89\b":  {"color": "negro mate",       "code": "89",  "acabado": "mate"},
}


def apply_deterministic_product_alias_rules(text_value: Optional[str], prepared_request: dict):
    # Expandir abreviaciones antes de normalizar
    normalized = expand_ferreteria_text(text_value)
    alias_rules = [
        {
            "pattern": r"\b(blanca economica|blanca economica|la economica|vinilo barato|p11|p-11|p 11)\b",
            "canonical_product": "domestico blanco",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "blanco"],
            "color_filters": ["blanco"],
        },
        # ── Códigos P-XX de Esmalte Doméstico Pintuco ──
        {
            "pattern": r"\b[Pp]-?18\b",
            "canonical_product": "domestico amarillo p18",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "amarillo", "P18"],
            "color_filters": ["amarillo"],
        },
        {
            "pattern": r"\b[Pp]-?35\b",
            "canonical_product": "domestico azul frances p35",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "azul", "frances", "P35"],
            "color_filters": ["azul frances"],
        },
        {
            "pattern": r"\b[Pp]-?40\b",
            "canonical_product": "domestico azul espanol p40",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "azul", "espanol", "P40"],
            "color_filters": ["azul espanol"],
        },
        {
            "pattern": r"\b[Pp]-?50\b",
            "canonical_product": "domestico azul verano p50",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "azul", "verano", "P50"],
            "color_filters": ["azul verano"],
        },
        {
            "pattern": r"\b[Pp]-?153\b",
            "canonical_product": "domestico aluminio p153",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "aluminio", "P153"],
            "color_filters": ["aluminio"],
        },
        # Código genérico P-XX (sin match específico) → esmalte domestico
        {
            "pattern": r"\b[Pp]-?\d{1,3}(?!\d)\b",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "esmalte"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*11\b",
            "canonical_product": "pintulux blanco 11",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "blanco 11", "3en1 br blanco 11"],
            "color_filters": ["blanco"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*95\b",
            "canonical_product": "pintulux negro 95",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "negro 95", "3en1 br negro 95"],
            "color_filters": ["negro"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*10\b",
            "canonical_product": "pintulux blanco mate 10",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "mat blanco 10", "3en1 mat blanco 10"],
            "color_filters": ["blanco mate"],
            "finish_filters": ["mate"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*89\b",
            "canonical_product": "pintulux negro mate 89",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "mat negro 89", "3en1 mat negro"],
            "color_filters": ["negro mate"],
            "finish_filters": ["mate"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*84\b",
            "canonical_product": "pintulux gris plata 84",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "gris plata 84", "br gris"],
            "color_filters": ["gris plata"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*76\b",
            "canonical_product": "pintulux caoba 76",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "caoba 76", "br caoba 76"],
            "color_filters": ["caoba"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*80\b",
            "canonical_product": "pintulux verde bronce 80",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "verde bronce 80", "br verde bronce"],
            "color_filters": ["verde bronce"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*53\b",
            "canonical_product": "pintulux verde esmeralda 53",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "verde esmeralda 53", "br verde esmer"],
            "color_filters": ["verde esmeralda"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*18\b",
            "canonical_product": "pintulux amarillo 18",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "amarillo 18", "br amarillo 18"],
            "color_filters": ["amarillo"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*20\b",
            "canonical_product": "pintulux naranja 20",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "naranja 20", "br naranja 20"],
            "color_filters": ["naranja"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*26\b",
            "canonical_product": "pintulux rojo bermellon 26",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "rojo bermellon 26", "br rojo"],
            "color_filters": ["rojo bermellon"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*40\b",
            "canonical_product": "pintulux azul espanol 40",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "azul espanol 40", "br azul"],
            "color_filters": ["azul espanol"],
        },
        {
            "pattern": r"\b(?:t(?:eu?|u)?)-?\s*12\b",
            "canonical_product": "pintulux crema 12",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "crema 12", "br crema 12"],
            "color_filters": ["crema"],
        },
        {
            "pattern": r"\b(p53|p-53|p 53)\b",
            "canonical_product": "domestico verde esmeralda",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico", "verde esmeralda"],
            "color_filters": ["verde esmeralda"],
        },
        {
            "pattern": r"\b(pl285|pl 285)\b",
            "canonical_product": "pegante pl285 madera",
            "brand_filters": ["artecola"],
            "core_terms": ["pl285", "madera"],
        },
        # ── Barniz SD (SD-1, SD-2, SD-3) ──
        {
            "pattern": r"\bsd[\s-]*1\b",
            "canonical_product": "barniz sd-1",
            "brand_filters": ["pintuco"],
            "core_terms": ["barniz", "incoloro", "sd-1"],
        },
        {
            "pattern": r"\bsd[\s-]*2\b",
            "canonical_product": "barniz sd-2",
            "brand_filters": ["pintuco"],
            "core_terms": ["barniz", "sd-2"],
        },
        {
            "pattern": r"\bsd[\s-]*3\b",
            "canonical_product": "barniz sd-3",
            "brand_filters": ["pintuco"],
            "core_terms": ["barniz", "sd-3"],
        },
        {
            "pattern": r"\b(pintulux)\b.*\bverde\s+bronce\b|\bverde\s+bronce\b.*\b(pintulux)\b",
            "canonical_product": "pintulux verde bronce",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux", "verde bronce"],
            "color_filters": ["verde bronce"],
        },
        {
            "pattern": r"\b(candado\s+yale)\b",
            "canonical_product": "candado yale",
            "brand_filters": ["yale"],
            "core_terms": ["candado", "yale"],
        },
        # ── Vinilo por tipo de calidad ──
        {
            "pattern": r"\bvinilo\s+tipo\s*1\b|\bvinilo\s+premium\b|\bvinilo\s+bueno\b|\bvinilo\s+lavable\b",
            "brand_filters": ["viniltex", "pintuco"],
            "core_terms": ["viniltex"],
        },
        {
            "pattern": r"\bvinilo\s+tipo\s*2\b|\bvinilo\s+intermedio\b",
            "brand_filters": ["intervinil", "pintuco"],
            "core_terms": ["intervinil"],
        },
        {
            "pattern": r"\bvinilo\s+tipo\s*3\b|\bvinilo\s+econom\w*\b|\bvinilo\s+barato\b|\bvinilo\s+de\s+obra\b",
            "brand_filters": ["pinturama", "pintuco"],
            "core_terms": ["pinturama"],
        },
        # ── Esmalte por calidad ──
        {
            "pattern": r"\besmalte\s+buen[oa]?\b|\besmalte\s+resistente\b|\besmalte\s+exterior\b",
            "brand_filters": ["pintulux", "pintuco"],
            "core_terms": ["pintulux"],
        },
        {
            "pattern": r"\besmalte\s+econom\w*\b|\besmalte\s+barat[oa]?\b|\besmalte\s+interior\b",
            "brand_filters": ["domestico", "pintuco"],
            "core_terms": ["domestico"],
        },
        # ── Aerosoles ──
        {
            "pattern": r"\b(aerosol|spray|pintura\s+spray|pintura\s+en\s+spray)\b",
            "brand_filters": ["aerocolor", "pintuco"],
            "core_terms": ["aerocolor"],
        },
        # ── Epóxicas ──
        {
            "pattern": r"\b(epox\w+|epoxi\w*)\b",
            "brand_filters": ["pintucoat", "pintuco"],
            "core_terms": ["pintucoat"],
        },
        # ── Anticorrosivos ──
        {
            "pattern": r"\b(anticorrosi\w+|anti\s*corrosi\w+)\b",
            "brand_filters": ["corrotec", "anticorrosivo", "pintuco"],
            "core_terms": ["corrotec"],
        },
        # ── Pintura para piso ──
        {
            "pattern": r"\bpintura\s+(para\s+)?piso\w*\b|\bpintar\s+(el\s+)?piso\b",
            "brand_filters": ["pintura canchas", "pintucoat", "pintuco"],
            "core_terms": ["pintura canchas"],
        },
        # ── Pintura de fachada ──
        {
            "pattern": r"\bpintura\s+(para\s+|de\s+)?fachada\w*\b|\bpintar\s+(la\s+)?fachada\b",
            "brand_filters": ["koraza", "pintuco"],
            "core_terms": ["koraza"],
        },
        # ── Impermeabilizante ──
        {
            "pattern": r"\bimpermeabiliz\w+\b",
            "brand_filters": ["koraza", "pintuco"],
            "core_terms": ["koraza"],
        },
        # ── Laca / barniz ──
        {
            "pattern": r"\b(laca|barniz)\b",
            "brand_filters": ["pintulac", "pintuco"],
            "core_terms": ["pintulac"],
        },
        # ── Pintura de tráfico / demarcación vial ──
        {
            "pattern": r"\b(pintutraf\w*|pintutrafico|pintura\s+(?:de\s+)?traf\w*|trafico\s+vial|demarcaci[oó]n\s+vial|señalizaci[oó]n\s+vial)\b",
            "canonical_product": "pintutraf",
            "brand_filters": ["pintuco"],
            "core_terms": ["pintutraf"],
        },
        # ── Microesferas (complemento de pintura tráfico) ──
        {
            "pattern": r"\b(microesfera\w*|micro\s+esfera\w*|esferas?\s+(?:reflectiva|vidrio|trafico))\b",
            "canonical_product": "microesferas",
            "core_terms": ["microesfera", "microesferas"],
        },
        # ── Poliuretano → Interthane (International) ──
        {
            "pattern": r"\bpoliuretano\b",
            "brand_filters": ["interthane", "international"],
            "core_terms": ["interthane"],
        },
        # ── Intumescente / ignífuga → Interchar ──
        {
            "pattern": r"\b(intumescente|ignifug\w+|proteccion\s+(?:al\s+)?fuego)\b",
            "brand_filters": ["interchar", "international"],
            "core_terms": ["interchar"],
        },
        # ── Removedor / decapante ──
        {
            "pattern": r"\b(removedor|decapante|quitar\s+pintura)\b",
            "canonical_product": "removedor",
            "brand_filters": ["pintuco"],
            "core_terms": ["removedor"],
        },
        # ── Humedad / filtración → Aquablock ──
        {
            "pattern": r"\b(humedad\s+interna|filtraci[oó]n|sellador\s+humedad)\b",
            "brand_filters": ["aquablock", "pintuco"],
            "core_terms": ["aquablock"],
        },
    ]

    for rule in alias_rules:
        if not re.search(rule["pattern"], normalized):
            continue
        if rule.get("canonical_product") and not prepared_request.get("canonical_product"):
            prepared_request["canonical_product"] = rule["canonical_product"]
        prepared_request["brand_filters"] = merge_unique_terms(prepared_request.get("brand_filters"), rule.get("brand_filters"))
        prepared_request["core_terms"] = merge_unique_terms(prepared_request.get("core_terms"), rule.get("core_terms"))
        prepared_request["color_filters"] = merge_unique_terms(prepared_request.get("color_filters"), rule.get("color_filters"))
        if rule.get("finish_filters"):
            prepared_request["finish_filters"] = merge_unique_terms(prepared_request.get("finish_filters"), rule.get("finish_filters"))

    if "verde bronce" in normalized:
        prepared_request["color_filters"] = merge_unique_terms(prepared_request.get("color_filters"), ["verde bronce"])
    elif "blanco puro" in normalized:
        prepared_request["color_filters"] = merge_unique_terms(prepared_request.get("color_filters"), ["blanco puro"])
    elif "transparente" in normalized:
        prepared_request["color_filters"] = merge_unique_terms(prepared_request.get("color_filters"), ["transparente"])

    if "mate" in normalized:
        prepared_request["finish_filters"] = merge_unique_terms(prepared_request.get("finish_filters"), ["mate"])
    elif "brillante" in normalized:
        prepared_request["finish_filters"] = merge_unique_terms(prepared_request.get("finish_filters"), ["brillante"])

    yale_size_match = re.search(r"\bcandado\s+yale\b.*?\b(30|40|50|60|70)\s*mm\b", normalized)
    if yale_size_match:
        yale_terms = ["candado", "yale", yale_size_match.group(1)]
        if yale_size_match.group(1) in {"30", "40", "50", "60"}:
            yale_terms.append("italiano")
        prepared_request["canonical_product"] = prepared_request.get("canonical_product") or "candado yale"
        prepared_request["brand_filters"] = merge_unique_terms(prepared_request.get("brand_filters"), ["yale"])
        prepared_request["core_terms"] = merge_unique_terms(prepared_request.get("core_terms"), yale_terms)

    prepared_request["search_terms"] = expand_product_terms(
        merge_unique_terms(
            prepared_request.get("search_terms"),
            prepared_request.get("core_terms"),
            prepared_request.get("color_filters"),
            prepared_request.get("finish_filters"),
        )
    )[:14]
    return prepared_request


def should_attempt_product_nlu(text_value: Optional[str], base_request: Optional[dict]):
    normalized = normalize_text_value(text_value)
    request = base_request or {}
    if not normalized or len(normalized) < 3:
        return False
    if is_technical_document_message(text_value) or is_technical_advisory_message(text_value):
        return False
    if has_non_product_business_signal(text_value) and not (
        request.get("product_codes")
        or request.get("brand_filters")
        or request.get("requested_unit")
        or request.get("requested_quantity")
    ):
        return False

    meaningful_terms = [term for term in (request.get("core_terms") or []) if not is_store_alias_term(term)]
    if request.get("product_codes"):
        return True
    if request.get("brand_filters") or request.get("requested_unit") or request.get("requested_quantity"):
        return True
    if len(meaningful_terms) >= 2:
        return True
    return bool(re.search(r"\b\d+\s*/\s*(1|4|5)\b", normalized))


def extract_json_object(raw_text: str):
    if not raw_text:
        raise ValueError("La respuesta del modelo llegó vacía.")
    raw_text = raw_text.strip()
    try:
        return json.loads(raw_text)
    except json.JSONDecodeError:
        start = raw_text.find("{")
        end = raw_text.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(raw_text[start : end + 1])
        raise


def extract_product_entities_with_llm(text_value: Optional[str], base_request: Optional[dict] = None):
    normalized = normalize_text_value(text_value)
    if not should_attempt_product_nlu(text_value, base_request):
        return {}
    if not get_openai_api_key():
        return {}
    if normalized in PRODUCT_NLU_CACHE:
        return dict(PRODUCT_NLU_CACHE[normalized])

    try:
        response = get_openai_client().responses.create(
            model=get_openai_model(),
            input=[
                {"role": "system", "content": get_product_nlu_system_prompt()},
                {"role": "user", "content": text_value or ""},
            ],
            temperature=0,
            text={"format": {"type": "json_object"}},
        )
        parsed = extract_json_object(response.output_text)
    except Exception:
        return {}

    nlu_payload = {
        "cantidad_inferida": parse_numeric_value(parsed.get("cantidad_inferida")),
        "presentacion_canonica_inferida": canonicalize_presentation_value(parsed.get("presentacion_canonica_inferida")),
        "producto_base": normalize_nullable_phrase(parsed.get("producto_base")),
        "color": normalize_nullable_phrase(parsed.get("color")),
        "acabado": normalize_nullable_phrase(parsed.get("acabado")),
    }
    if not any(nlu_payload.values()):
        return {}

    if len(PRODUCT_NLU_CACHE) >= PRODUCT_NLU_CACHE_MAX_ITEMS:
        PRODUCT_NLU_CACHE.pop(next(iter(PRODUCT_NLU_CACHE)))
    PRODUCT_NLU_CACHE[normalized] = dict(nlu_payload)
    return nlu_payload


def prepare_product_request_for_search(text_value: Optional[str], product_request: Optional[dict] = None):
    prepared_request = dict(product_request or extract_product_request(text_value))
    if prepared_request.get("nlu_processed"):
        return prepared_request

    prepared_request.setdefault("color_filters", [])
    prepared_request.setdefault("finish_filters", [])
    prepared_request = apply_deterministic_product_alias_rules(text_value, prepared_request)

    nlu_payload = extract_product_entities_with_llm(text_value, prepared_request)
    prepared_request["nlu_processed"] = True
    prepared_request["nlu_extraction"] = nlu_payload or None

    if not nlu_payload:
        return prepared_request

    canonical_product = nlu_payload.get("producto_base")
    canonical_color = nlu_payload.get("color")
    canonical_finish = nlu_payload.get("acabado")
    canonical_presentation = nlu_payload.get("presentacion_canonica_inferida")
    inferred_quantity = nlu_payload.get("cantidad_inferida")

    if inferred_quantity is not None and prepared_request.get("requested_quantity") is None:
        prepared_request["requested_quantity"] = inferred_quantity
    if canonical_presentation and not prepared_request.get("requested_unit"):
        prepared_request["requested_unit"] = canonical_presentation

    if canonical_product:
        prepared_request["canonical_product"] = canonical_product
    if canonical_color:
        prepared_request["color_filters"] = merge_unique_terms(prepared_request.get("color_filters"), [canonical_color], tokenize_search_phrase(canonical_color))
    if canonical_finish:
        prepared_request["finish_filters"] = merge_unique_terms(prepared_request.get("finish_filters"), [canonical_finish], tokenize_search_phrase(canonical_finish))

    merged_core_terms = merge_unique_terms(
        prepared_request.get("core_terms"),
        [canonical_product] if canonical_product else [],
        tokenize_search_phrase(canonical_product),
        prepared_request.get("color_filters"),
        prepared_request.get("finish_filters"),
    )
    prepared_request["core_terms"] = merged_core_terms[:10]

    merged_search_terms = merge_unique_terms(
        prepared_request.get("search_terms"),
        merged_core_terms,
        [canonical_presentation] if canonical_presentation else [],
    )
    prepared_request["search_terms"] = expand_product_terms(merged_search_terms)[:14]

    derived_brand_filters = extract_brand_filters(
        " ".join(
            value
            for value in [canonical_product, canonical_color, canonical_finish]
            if value
        )
    )
    prepared_request["brand_filters"] = merge_unique_terms(prepared_request.get("brand_filters"), derived_brand_filters)
    return prepared_request


_db_engine_singleton = None
_db_engine_url = None

def get_db_engine():
    global _db_engine_singleton, _db_engine_url
    url = get_database_url()
    if _db_engine_singleton is None or _db_engine_url != url:
        _db_engine_singleton = create_engine(url, pool_size=5, max_overflow=10, pool_recycle=300, pool_pre_ping=True)
        _db_engine_url = url
    return _db_engine_singleton


def safe_json_dumps(value):
    return json.dumps(value, ensure_ascii=False, default=str)


def normalize_internal_username(username: Optional[str]):
    normalized = (username or "").strip().lower()
    if not re.fullmatch(r"[a-z0-9._-]{3,80}", normalized):
        raise HTTPException(status_code=400, detail="El usuario interno debe usar solo letras, números, punto, guion o guion bajo.")
    return normalized


def normalize_phone_e164(phone_number: Optional[str]):
    digits = "".join(character for character in str(phone_number or "") if character.isdigit())
    if not digits:
        return None
    if digits.startswith("57") and len(digits) >= 12:
        return f"+{digits}"
    if len(digits) == 10:
        return f"+57{digits}"
    return f"+{digits}"


def normalize_employee_header(value: Optional[str]):
    return normalize_text_value(value).replace("�", "")


def parse_employee_document(value: Optional[str]):
    digits = "".join(character for character in str(value or "") if character.isdigit())
    return digits or None


def parse_employee_phone(value: Optional[str]):
    return normalize_phone_e164(value)


def load_employee_sede_store_map():
    configured = {}
    raw_json = os.getenv("EMPLOYEE_SEDE_STORE_MAP_JSON")
    if raw_json:
        try:
            configured = json.loads(raw_json)
        except Exception as exc:
            raise RuntimeError(f"EMPLOYEE_SEDE_STORE_MAP_JSON inválido: {exc}")
    merged = {**DEFAULT_EMPLOYEE_SEDE_STORE_MAP}
    for raw_key, raw_value in (configured or {}).items():
        key = normalize_text_value(raw_key)
        store_code = normalize_store_code(str(raw_value)) or str(raw_value).strip()
        if key and store_code:
            merged[key] = store_code
    return merged


def resolve_store_code_from_employee_sede(sede_value: Optional[str]):
    normalized_sede = normalize_text_value(sede_value)
    if not normalized_sede:
        return None
    mapped_code = load_employee_sede_store_map().get(normalized_sede)
    if mapped_code:
        return normalize_store_code(mapped_code) or mapped_code
    return normalize_store_code(normalized_sede)


def derive_internal_role_from_employee(record: dict):
    cargo = normalize_text_value(record.get("cargo"))
    if not cargo:
        return "empleado"
    if "administr" in cargo or "director" in cargo or ("lider" in cargo and "compras" in cargo):
        return "administrador"
    if "gerente" in cargo:
        return "gerente"
    if any(token in cargo for token in ["vendedor", "asesor comercial", "representante de ventas", "mostrador"]):
        return "vendedor"
    if any(token in cargo for token in ["lider", "líder", "facturacion", "facturación", "logistica", "logística", "inventario", "cartera", "tesoreria", "tesorería", "compras"]):
        return "operador"
    return "empleado"


def employee_has_advanced_access(record: dict):
    role = derive_internal_role_from_employee(record)
    return role in {"vendedor", "gerente", "operador", "administrador"}


def employee_can_manage_transfers(record: dict):
    cargo = normalize_text_value(record.get("cargo"))
    return any(
        token in cargo
        for token in [
            "lider de tienda",
            "líder de tienda",
            "lider logistica",
            "líder logística",
            "lider de inventario",
            "líder de inventario",
            "facturacion y despachos",
            "facturación y despachos",
            "auxiliar de mostrador y facturacion",
            "auxiliar de mostrador y facturación",
            "auxiliar logistico",
            "auxiliar logístico",
            "gestor logistico",
            "gestor logístico",
        ]
    )


def load_employee_directory(force_refresh: bool = False):
    cache_age = time.time() - float(EMPLOYEE_DIRECTORY_CACHE.get("loaded_at") or 0)
    if not force_refresh and EMPLOYEE_DIRECTORY_CACHE.get("records") and cache_age < EMPLOYEE_DIRECTORY_CACHE_TTL_SECONDS:
        return EMPLOYEE_DIRECTORY_CACHE["records"]

    if not EMPLOYEE_DIRECTORY_PATH.exists():
        EMPLOYEE_DIRECTORY_CACHE.update({"loaded_at": time.time(), "records": []})
        return []

    dataframe = pd.read_excel(EMPLOYEE_DIRECTORY_PATH)
    normalized_columns = {normalize_employee_header(column): column for column in dataframe.columns}

    def get_column(*aliases):
        for alias in aliases:
            column = normalized_columns.get(normalize_employee_header(alias))
            if column:
                return column
        return None

    full_name_column = get_column("nombre completo", "nombre", "nombre empleado")
    document_column = get_column("cedula", "cédula")
    sede_column = get_column("sede")
    cargo_column = get_column("cargo")
    phone_column = get_column("telefono", "teléfono")
    email_column = get_column("correo electronico", "correo electrónico", "correo")
    # ERP vendor code: allows precise matching in vw_ventas_netas.codigo_vendedor
    codigo_vendedor_column = get_column("codigo_vendedor", "código_vendedor", "cod_vendedor", "codigovendedor", "erp_code", "vendor_code")

    if not full_name_column or not document_column:
        raise RuntimeError(f"datos_empleados.xlsx no tiene columnas mínimas esperadas. Columnas detectadas: {list(dataframe.columns)}")

    records = []
    for _, row in dataframe.iterrows():
        full_name = str(row.get(full_name_column) or "").strip()
        cedula = parse_employee_document(row.get(document_column))
        if not full_name or not cedula:
            continue
        sede = str(row.get(sede_column) or "").strip() if sede_column else ""
        cargo = str(row.get(cargo_column) or "").strip() if cargo_column else ""
        phone_e164 = parse_employee_phone(row.get(phone_column)) if phone_column else None
        email = str(row.get(email_column) or "").strip().lower() if email_column else None
        role = derive_internal_role_from_employee({"cargo": cargo})
        store_code = resolve_store_code_from_employee_sede(sede)
        raw_cod_vend = row.get(codigo_vendedor_column) if codigo_vendedor_column else None
        codigo_vendedor = str(int(float(raw_cod_vend))).strip() if raw_cod_vend and str(raw_cod_vend).strip() not in ("", "nan", "None") else None
        records.append(
            {
                "full_name": full_name,
                "cedula": cedula,
                "sede": sede,
                "cargo": cargo,
                "phone_e164": phone_e164,
                "email": email,
                "store_code": store_code,
                "role": role,
                "advanced_access": employee_has_advanced_access({"cargo": cargo}),
                "is_facturador": employee_can_manage_transfers({"cargo": cargo}),
                "codigo_vendedor": codigo_vendedor,
            }
        )

    EMPLOYEE_DIRECTORY_CACHE.update({"loaded_at": time.time(), "records": records})
    return records


def find_employee_record_by_phone(phone_e164: Optional[str]):
    normalized_phone = normalize_phone_e164(phone_e164)
    if not normalized_phone:
        return None
    for record in load_employee_directory():
        if record.get("phone_e164") == normalized_phone:
            return record
    fallback_record = fetch_internal_employee_record_by_phone(normalized_phone)
    if fallback_record:
        return fallback_record
    return None


def find_employee_record_by_cedula(document_value: Optional[str]):
    normalized_document = parse_employee_document(document_value)
    if not normalized_document:
        return None
    for record in load_employee_directory():
        if record.get("cedula") == normalized_document:
            return record
    fallback_record = fetch_internal_employee_record_by_cedula(normalized_document)
    if fallback_record:
        return fallback_record
    return None


def build_employee_username(record: dict):
    return f"cedula.{record.get('cedula')}"


def extract_internal_cedula_candidate(content: Optional[str]):
    raw_content = (content or "").strip()
    if not raw_content:
        return None

    normalized = normalize_text_value(raw_content)
    # If message has conversational/order/quote keywords, restrict to labeled cédula only
    _RESTRICT_KEYWORDS = [
        "login ", "pedido", "cotizacion", "cotización", "traslado",
        "galon", "galón", "cunete", "cuñete",
        # Customer identity context (name + cédula for quotes)
        "nombre", "señor", "señora", "angela", "maria", "contreras",
    ]
    # Also restrict if message has both alpha words AND digits (likely name + cédula)
    _has_alpha_words = bool(re.search(r"[a-záéíóúñü]{3,}", normalized))
    _has_digits = bool(re.search(r"\d{6,}", raw_content))
    _is_mixed_content = _has_alpha_words and _has_digits and len(raw_content) > 20
    if _is_mixed_content or any(fragment in normalized for fragment in _RESTRICT_KEYWORDS):
        labeled_match = re.search(r"(?:cedula|cédula|documento|doc)\s*(?:es|:)?\s*([0-9][0-9.\s-]{5,})", raw_content, re.IGNORECASE)
        if labeled_match:
            labeled_digits = parse_employee_document(labeled_match.group(1))
            if labeled_digits and 6 <= len(labeled_digits) <= 15:
                return labeled_digits
        compact_digits = parse_employee_document(raw_content)
        if compact_digits and 6 <= len(compact_digits) <= 15 and re.fullmatch(r"[0-9.\s-]+", raw_content):
            return compact_digits
        return None

    match = INTERNAL_CEDULA_PATTERN.search(raw_content)
    if match:
        return match.group(1)
    digits = parse_employee_document(raw_content)
    if digits and 6 <= len(digits) <= 15:
        return digits
    return None


def build_employee_record_from_internal_user_row(user_row: Optional[dict]):
    if not user_row:
        return None
    metadata = dict(user_row.get("metadata") or {})
    cedula = parse_employee_document(metadata.get("cedula"))
    if not cedula:
        username_match = re.fullmatch(r"cedula\.(\d{6,15})", str(user_row.get("username") or ""))
        cedula = username_match.group(1) if username_match else None
    if not cedula:
        return None
    role_value = user_row.get("role") or derive_internal_role_from_employee({"cargo": metadata.get("cargo")})
    return {
        "full_name": user_row.get("full_name"),
        "cedula": cedula,
        "sede": metadata.get("sede"),
        "cargo": metadata.get("cargo"),
        "phone_e164": normalize_phone_e164(user_row.get("phone_e164") or metadata.get("telefono")),
        "email": (user_row.get("email") or metadata.get("email") or "").strip().lower() or None,
        "store_code": normalize_store_code(metadata.get("store_code")),
        "role": role_value,
        "advanced_access": bool(metadata.get("advanced_access", role_value in {"vendedor", "gerente", "operador", "administrador"})),
        "is_facturador": bool(metadata.get("is_facturador", role_value == "administrador")),
        "auth_source": metadata.get("auth_source") or "agent_user",
        "codigo_vendedor": metadata.get("codigo_vendedor") or None,
    }


def fetch_internal_employee_record_by_phone(phone_e164: Optional[str]):
    normalized_phone = normalize_phone_e164(phone_e164)
    if not normalized_phone:
        return None
    ensure_internal_auth_tables()
    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT id, username, full_name, role, phone_e164, email, metadata
                FROM public.agent_user
                WHERE is_active = true
                  AND phone_e164 = :phone_e164
                ORDER BY updated_at DESC
                LIMIT 1
                """
            ),
            {"phone_e164": normalized_phone},
        ).mappings().one_or_none()
    return build_employee_record_from_internal_user_row(dict(row) if row else None)


def fetch_internal_employee_record_by_cedula(document_value: Optional[str]):
    normalized_document = parse_employee_document(document_value)
    if not normalized_document:
        return None
    ensure_internal_auth_tables()
    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT id, username, full_name, role, phone_e164, email, metadata
                FROM public.agent_user
                WHERE is_active = true
                  AND (
                    metadata ->> 'cedula' = :cedula
                    OR username = :username
                  )
                ORDER BY updated_at DESC
                LIMIT 1
                """
            ),
            {"cedula": normalized_document, "username": f"cedula.{normalized_document}"},
        ).mappings().one_or_none()
    return build_employee_record_from_internal_user_row(dict(row) if row else None)


def hash_password_with_salt(password: str, salt_hex: str):
    return hashlib.pbkdf2_hmac(
        "sha256",
        (password or "").encode("utf-8"),
        bytes.fromhex(salt_hex),
        INTERNAL_PASSWORD_ITERATIONS,
    ).hex()


def build_password_credentials(password: str):
    if not password or len(password) < 8:
        raise HTTPException(status_code=400, detail="La contraseña interna debe tener al menos 8 caracteres.")
    salt_hex = secrets.token_hex(16)
    return salt_hex, hash_password_with_salt(password, salt_hex)


def verify_password_hash(password: str, salt_hex: str, expected_hash: str):
    calculated_hash = hash_password_with_salt(password, salt_hex)
    return hmac.compare_digest(calculated_hash, expected_hash)


def hash_session_token(raw_token: str):
    return hashlib.sha256((raw_token or "").encode("utf-8")).hexdigest()


def ensure_internal_auth_tables():
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS public.agent_user (
                    id bigserial PRIMARY KEY,
                    username varchar(80) NOT NULL,
                    full_name varchar(180) NOT NULL,
                    role varchar(30) NOT NULL,
                    password_salt varchar(128) NOT NULL,
                    password_hash varchar(256) NOT NULL,
                    phone_e164 varchar(30),
                    email varchar(180),
                    is_active boolean NOT NULL DEFAULT true,
                    metadata jsonb NOT NULL DEFAULT '{}'::jsonb,
                    created_at timestamptz NOT NULL DEFAULT now(),
                    updated_at timestamptz NOT NULL DEFAULT now(),
                    CONSTRAINT uq_agent_user_username UNIQUE (username),
                    CONSTRAINT uq_agent_user_phone UNIQUE (phone_e164)
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS public.agent_user_scope (
                    id bigserial PRIMARY KEY,
                    user_id bigint NOT NULL REFERENCES public.agent_user(id) ON DELETE CASCADE,
                    scope_type varchar(40) NOT NULL,
                    scope_value varchar(180) NOT NULL,
                    scope_label varchar(180),
                    metadata jsonb NOT NULL DEFAULT '{}'::jsonb,
                    created_at timestamptz NOT NULL DEFAULT now(),
                    updated_at timestamptz NOT NULL DEFAULT now(),
                    CONSTRAINT uq_agent_user_scope UNIQUE (user_id, scope_type, scope_value)
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS public.agent_user_session (
                    id bigserial PRIMARY KEY,
                    user_id bigint NOT NULL REFERENCES public.agent_user(id) ON DELETE CASCADE,
                    token_hash varchar(128) NOT NULL,
                    channel varchar(30) NOT NULL DEFAULT 'api',
                    contact_id bigint REFERENCES public.whatsapp_contacto(id) ON DELETE SET NULL,
                    phone_e164 varchar(30),
                    expires_at timestamptz NOT NULL,
                    last_used_at timestamptz,
                    revoked_at timestamptz,
                    metadata jsonb NOT NULL DEFAULT '{}'::jsonb,
                    created_at timestamptz NOT NULL DEFAULT now(),
                    updated_at timestamptz NOT NULL DEFAULT now(),
                    CONSTRAINT uq_agent_user_session_token UNIQUE (token_hash)
                )
                """
            )
        )
        connection.execute(text("ALTER TABLE public.agent_user DROP CONSTRAINT IF EXISTS chk_agent_user_role"))
        connection.execute(
            text(
                """
                ALTER TABLE public.agent_user
                ADD CONSTRAINT chk_agent_user_role
                CHECK (role IN ('empleado', 'vendedor', 'gerente', 'operador', 'administrador'))
                """
            )
        )
        connection.execute(text("ALTER TABLE public.agent_user_scope DROP CONSTRAINT IF EXISTS chk_agent_user_scope_type"))
        connection.execute(
            text(
                """
                ALTER TABLE public.agent_user_scope
                ADD CONSTRAINT chk_agent_user_scope_type
                CHECK (scope_type IN ('cliente', 'vendedor_codigo', 'vendedor_nombre', 'zona', 'almacen'))
                """
            )
        )
        connection.execute(text("ALTER TABLE public.agent_user_session DROP CONSTRAINT IF EXISTS chk_agent_user_session_channel"))
        connection.execute(
            text(
                """
                ALTER TABLE public.agent_user_session
                ADD CONSTRAINT chk_agent_user_session_channel
                CHECK (channel IN ('api', 'whatsapp'))
                """
            )
        )


def fetch_internal_user_scopes(user_id: int):
    ensure_internal_auth_tables()
    engine = get_db_engine()
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                """
                SELECT scope_type, scope_value, scope_label
                FROM public.agent_user_scope
                WHERE user_id = :user_id
                ORDER BY scope_type, scope_value
                """
            ),
            {"user_id": user_id},
        ).mappings().all()
    return [dict(row) for row in rows]


def fetch_internal_user_by_username(username: str):
    ensure_internal_auth_tables()
    normalized_username = normalize_internal_username(username)
    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT id, username, full_name, role, password_salt, password_hash,
                       phone_e164, email, is_active, metadata
                FROM public.agent_user
                WHERE username = :username
                LIMIT 1
                """
            ),
            {"username": normalized_username},
        ).mappings().one_or_none()
    if not row:
        return None
    user_payload = dict(row)
    user_payload["scopes"] = fetch_internal_user_scopes(user_payload["id"])
    return user_payload


def fetch_internal_user_by_id(user_id: int):
    ensure_internal_auth_tables()
    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT id, username, full_name, role, phone_e164, email, is_active, metadata
                FROM public.agent_user
                WHERE id = :user_id
                LIMIT 1
                """
            ),
            {"user_id": user_id},
        ).mappings().one_or_none()
    if not row:
        return None
    user_payload = dict(row)
    user_payload["scopes"] = fetch_internal_user_scopes(user_payload["id"])
    return user_payload


def sync_internal_user_from_employee_record(record: dict):
    ensure_internal_auth_tables()
    username = build_employee_username(record)
    metadata = {
        "auth_source": "datos_empleados.xlsx",
        "cedula": record.get("cedula"),
        "sede": record.get("sede"),
        "cargo": record.get("cargo"),
        "store_code": record.get("store_code"),
        "advanced_access": bool(record.get("advanced_access")),
        "is_facturador": bool(record.get("is_facturador")),
        # codigo_vendedor: ERP vendor code (e.g. "154011") from datos_empleados.xlsx column "codigo_vendedor"
        # Enables precise nom_vendedor matching in vw_ventas_netas without relying on ILIKE name patterns
        "codigo_vendedor": record.get("codigo_vendedor") or None,
    }

    engine = get_db_engine()
    with engine.begin() as connection:
        existing_row = connection.execute(
            text(
                """
                SELECT id
                FROM public.agent_user
                WHERE username = :username
                LIMIT 1
                """
            ),
            {"username": username},
        ).mappings().one_or_none()

        if existing_row:
            user_id = existing_row["id"]
            connection.execute(
                text(
                    """
                    UPDATE public.agent_user
                    SET full_name = :full_name,
                        role = :role,
                        phone_e164 = :phone_e164,
                        email = :email,
                        metadata = CAST(:metadata AS jsonb),
                        is_active = true,
                        updated_at = now()
                    WHERE id = :user_id
                    """
                ),
                {
                    "user_id": user_id,
                    "full_name": record.get("full_name"),
                    "role": record.get("role") or "empleado",
                    "phone_e164": record.get("phone_e164"),
                    "email": record.get("email"),
                    "metadata": safe_json_dumps(metadata),
                },
            )
        else:
            salt_hex, password_hash = build_password_credentials(f"ExcelAuth-{record.get('cedula')}-Fx")
            user_id = connection.execute(
                text(
                    """
                    INSERT INTO public.agent_user (
                        username, full_name, role, password_salt, password_hash, phone_e164, email, is_active, metadata, created_at, updated_at
                    )
                    VALUES (
                        :username, :full_name, :role, :password_salt, :password_hash, :phone_e164, :email, true, CAST(:metadata AS jsonb), now(), now()
                    )
                    RETURNING id
                    """
                ),
                {
                    "username": username,
                    "full_name": record.get("full_name"),
                    "role": record.get("role") or "empleado",
                    "password_salt": salt_hex,
                    "password_hash": password_hash,
                    "phone_e164": record.get("phone_e164"),
                    "email": record.get("email"),
                    "metadata": safe_json_dumps(metadata),
                },
            ).scalar_one()

        connection.execute(text("DELETE FROM public.agent_user_scope WHERE user_id = :user_id"), {"user_id": user_id})
        if record.get("store_code"):
            connection.execute(
                text(
                    """
                    INSERT INTO public.agent_user_scope (
                        user_id, scope_type, scope_value, scope_label, created_at, updated_at
                    ) VALUES (
                        :user_id, 'almacen', :scope_value, :scope_label, now(), now()
                    )
                    """
                ),
                {
                    "user_id": user_id,
                    "scope_value": record.get("store_code"),
                    "scope_label": record.get("sede") or record.get("store_code"),
                },
            )

    return fetch_internal_user_by_id(user_id)


def build_internal_auth_context(user_payload: dict, token: str, expires_at: Optional[str]):
    metadata = dict(user_payload.get("metadata") or {})
    return {
        "token": token,
        "user_id": user_payload.get("id"),
        "username": user_payload.get("username"),
        "role": user_payload.get("role"),
        "expires_at": expires_at,
        "employee_context": {
            "full_name": user_payload.get("full_name"),
            "cedula": metadata.get("cedula"),
            "sede": metadata.get("sede"),
            "cargo": metadata.get("cargo"),
            "telefono": user_payload.get("phone_e164"),
            "store_code": metadata.get("store_code"),
            "advanced_access": metadata.get("advanced_access", False),
            "is_facturador": metadata.get("is_facturador", False),
            # ERP vendor code if available from employees Excel ("154011" style)
            "codigo_vendedor": metadata.get("codigo_vendedor") or None,
        },
    }


def internal_user_has_advanced_access(user_payload: dict):
    if (user_payload or {}).get("role") in {"vendedor", "gerente", "operador", "administrador"}:
        return True
    metadata = dict((user_payload or {}).get("metadata") or {})
    return bool(metadata.get("advanced_access"))


def internal_user_can_manage_transfers(user_payload: dict):
    if (user_payload or {}).get("role") == "administrador":
        return True
    metadata = dict((user_payload or {}).get("metadata") or {})
    return bool(metadata.get("is_facturador"))


def upsert_internal_user(user_request: InternalBootstrapUserRequest):
    ensure_internal_auth_tables()
    role = (user_request.role or "").strip().lower()
    if role not in INTERNAL_ROLES:
        raise HTTPException(status_code=400, detail="Rol interno inválido.")

    normalized_username = normalize_internal_username(user_request.username)
    phone_e164 = normalize_phone_e164(user_request.phone_number)
    salt_hex, password_hash = build_password_credentials(user_request.password)
    normalized_scopes = []
    for scope in user_request.scopes:
        scope_type = (scope.scope_type or "").strip().lower()
        if scope_type not in INTERNAL_SCOPE_TYPES:
            raise HTTPException(status_code=400, detail=f"Tipo de alcance no válido: {scope.scope_type}")
        scope_value = normalize_text_value(scope.scope_value)
        if not scope_value:
            raise HTTPException(status_code=400, detail="Todos los alcances deben tener un valor válido.")
        normalized_scopes.append(
            {
                "scope_type": scope_type,
                "scope_value": scope_value,
                "scope_label": scope.scope_label,
            }
        )

    engine = get_db_engine()
    with engine.begin() as connection:
        existing_row = connection.execute(
            text(
                """
                SELECT id
                FROM public.agent_user
                WHERE username = :username
                LIMIT 1
                """
            ),
            {"username": normalized_username},
        ).mappings().one_or_none()

        if existing_row:
            user_id = existing_row["id"]
            connection.execute(
                text(
                    """
                    UPDATE public.agent_user
                    SET full_name = :full_name,
                        role = :role,
                        password_salt = :password_salt,
                        password_hash = :password_hash,
                        phone_e164 = :phone_e164,
                        email = :email,
                        is_active = true,
                        updated_at = now()
                    WHERE id = :user_id
                    """
                ),
                {
                    "user_id": user_id,
                    "full_name": user_request.full_name,
                    "role": role,
                    "password_salt": salt_hex,
                    "password_hash": password_hash,
                    "phone_e164": phone_e164,
                    "email": user_request.email,
                },
            )
            connection.execute(text("DELETE FROM public.agent_user_scope WHERE user_id = :user_id"), {"user_id": user_id})
        else:
            user_id = connection.execute(
                text(
                    """
                    INSERT INTO public.agent_user (
                        username, full_name, role, password_salt, password_hash, phone_e164, email, is_active, created_at, updated_at
                    )
                    VALUES (
                        :username, :full_name, :role, :password_salt, :password_hash, :phone_e164, :email, true, now(), now()
                    )
                    RETURNING id
                    """
                ),
                {
                    "username": normalized_username,
                    "full_name": user_request.full_name,
                    "role": role,
                    "password_salt": salt_hex,
                    "password_hash": password_hash,
                    "phone_e164": phone_e164,
                    "email": user_request.email,
                },
            ).scalar_one()

        for scope in normalized_scopes:
            connection.execute(
                text(
                    """
                    INSERT INTO public.agent_user_scope (
                        user_id, scope_type, scope_value, scope_label, created_at, updated_at
                    ) VALUES (
                        :user_id, :scope_type, :scope_value, :scope_label, now(), now()
                    )
                    """
                ),
                {
                    "user_id": user_id,
                    "scope_type": scope["scope_type"],
                    "scope_value": scope["scope_value"],
                    "scope_label": scope["scope_label"],
                },
            )

    return fetch_internal_user_by_id(user_id)


def create_internal_session(user_payload: dict, channel: str = "api", contact_id: Optional[int] = None, phone_e164: Optional[str] = None):
    ensure_internal_auth_tables()
    raw_token = secrets.token_urlsafe(32)
    token_hash = hash_session_token(raw_token)
    expires_at = datetime.utcnow() + timedelta(hours=INTERNAL_SESSION_TTL_HOURS)
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO public.agent_user_session (
                    user_id, token_hash, channel, contact_id, phone_e164, expires_at,
                    last_used_at, created_at, updated_at
                ) VALUES (
                    :user_id, :token_hash, :channel, :contact_id, :phone_e164, :expires_at,
                    now(), now(), now()
                )
                """
            ),
            {
                "user_id": user_payload["id"],
                "token_hash": token_hash,
                "channel": channel,
                "contact_id": contact_id,
                "phone_e164": phone_e164,
                "expires_at": expires_at,
            },
        )
    return {"token": raw_token, "expires_at": expires_at.isoformat()}


def resolve_internal_session(raw_token: Optional[str]):
    if not raw_token:
        return None
    ensure_internal_auth_tables()
    token_hash = hash_session_token(raw_token)
    engine = get_db_engine()
    with engine.begin() as connection:
        row = connection.execute(
            text(
                """
                SELECT s.user_id, s.expires_at, s.channel, s.phone_e164, s.contact_id,
                      u.username, u.full_name, u.role, u.email, u.phone_e164 AS user_phone, u.is_active, u.metadata
                FROM public.agent_user_session s
                JOIN public.agent_user u ON u.id = s.user_id
                WHERE s.token_hash = :token_hash
                  AND s.revoked_at IS NULL
                  AND s.expires_at > now()
                LIMIT 1
                """
            ),
            {"token_hash": token_hash},
        ).mappings().one_or_none()
        if not row or not row["is_active"]:
            return None
        connection.execute(
            text(
                """
                UPDATE public.agent_user_session
                SET last_used_at = now(),
                    updated_at = now()
                WHERE token_hash = :token_hash
                """
            ),
            {"token_hash": token_hash},
        )
    user_payload = {
        "id": row["user_id"],
        "username": row["username"],
        "full_name": row["full_name"],
        "role": row["role"],
        "email": row["email"],
        "phone_e164": row["user_phone"],
        "metadata": row.get("metadata") or {},
        "session_expires_at": row["expires_at"].isoformat() if row.get("expires_at") else None,
        "scopes": fetch_internal_user_scopes(row["user_id"]),
    }
    return user_payload


def revoke_internal_session(raw_token: Optional[str]):
    if not raw_token:
        return
    ensure_internal_auth_tables()
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE public.agent_user_session
                SET revoked_at = now(),
                    updated_at = now()
                WHERE token_hash = :token_hash
                  AND revoked_at IS NULL
                """
            ),
            {"token_hash": hash_session_token(raw_token)},
        )


def authenticate_internal_user(username: str, password: str, phone_number: Optional[str] = None):
    user_payload = fetch_internal_user_by_username(username)
    if not user_payload or not user_payload.get("is_active"):
        return None
    if not verify_password_hash(password, user_payload["password_salt"], user_payload["password_hash"]):
        return None
    registered_phone = normalize_phone_e164(user_payload.get("phone_e164"))
    incoming_phone = normalize_phone_e164(phone_number)
    if registered_phone and incoming_phone and registered_phone != incoming_phone:
        raise HTTPException(status_code=403, detail="Este usuario interno solo puede autenticarse desde su número registrado.")
    return user_payload


def extract_bearer_token(authorization: Optional[str]):
    if not authorization:
        return None
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        return None
    return token.strip()


def require_internal_user(authorization: Optional[str]):
    token = extract_bearer_token(authorization)
    user_payload = resolve_internal_session(token)
    if not user_payload:
        raise HTTPException(status_code=401, detail="Sesión interna inválida o vencida.")
    return user_payload


def fetch_customer_lookup_row(cliente_codigo: str):
    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT cliente_codigo, nombre_cliente, nit, numero_documento, telefono1, telefono2, email,
                       vendedor, vendedor_codigo, zona, ultima_compra, ventas_netas_total,
                       saldo_cartera, max_dias_vencido, documentos_vencidos
                FROM public.vw_agente_clientes_lookup
                WHERE cliente_codigo = :cliente_codigo
                LIMIT 1
                """
            ),
            {"cliente_codigo": cliente_codigo},
        ).mappings().one_or_none()
    return dict(row) if row else None


def search_customer_lookup_rows(search_text: str, limit: int = 5):
    query_text = (search_text or "").strip()
    if not query_text:
        return []
    normalized = normalize_text_value(query_text)
    digits = re.sub(r"\D", "", query_text)
    compact = normalize_reference_value(query_text)
    engine = get_db_engine()
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                """
                SELECT cliente_codigo, nombre_cliente, nit, numero_documento, telefono1, telefono2, email,
                       vendedor, vendedor_codigo, zona, ultima_compra, ventas_netas_total,
                       saldo_cartera, max_dias_vencido, documentos_vencidos
                FROM public.vw_agente_clientes_lookup
                WHERE (
                    :digits <> '' AND (
                        regexp_replace(COALESCE(nit, ''), '[^0-9]', '', 'g') LIKE :digits_pattern
                        OR regexp_replace(COALESCE(numero_documento, ''), '[^0-9]', '', 'g') LIKE :digits_pattern
                        OR regexp_replace(COALESCE(cliente_codigo, ''), '[^0-9]', '', 'g') = :digits
                    )
                )
                OR (
                    :compact IS NOT NULL AND regexp_replace(lower(COALESCE(cliente_codigo, '')), '[^a-z0-9]', '', 'g') = :compact
                )
                OR (
                    :normalized <> '' AND search_blob ILIKE :name_pattern
                )
                ORDER BY
                    CASE
                        WHEN regexp_replace(COALESCE(cliente_codigo, ''), '[^0-9]', '', 'g') = :digits THEN 0
                        WHEN regexp_replace(COALESCE(nit, ''), '[^0-9]', '', 'g') = :digits THEN 1
                        WHEN search_blob ILIKE :name_pattern THEN 2
                        ELSE 3
                    END,
                    max_dias_vencido DESC NULLS LAST,
                    ventas_netas_total DESC NULLS LAST,
                    nombre_cliente ASC
                LIMIT :limit
                """
            ),
            {
                "digits": digits,
                "digits_pattern": f"{digits}%" if digits else "",
                "compact": compact,
                "normalized": normalized,
                "name_pattern": f"%{normalized}%" if normalized else "",
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows]


def internal_user_can_access_customer(user_payload: dict, customer_row: Optional[dict]):
    if not customer_row:
        return False
    role = (user_payload or {}).get("role")
    if role in {"administrador", "gerente", "operador"}:
        return True
    if role != "vendedor":
        return False

    scopes = (user_payload or {}).get("scopes") or []
    if not scopes:
        return False

    customer_code = normalize_text_value(customer_row.get("cliente_codigo"))
    vendor_name = normalize_text_value(customer_row.get("vendedor"))
    vendor_code = normalize_text_value(customer_row.get("vendedor_codigo"))
    zone_name = normalize_text_value(customer_row.get("zona"))

    for scope in scopes:
        scope_type = normalize_text_value(scope.get("scope_type"))
        scope_value = normalize_text_value(scope.get("scope_value"))
        if scope_type == "cliente" and scope_value == customer_code:
            return True
        if scope_type == "vendedor_nombre" and scope_value and vendor_name and scope_value == vendor_name:
            return True
        if scope_type == "vendedor_codigo" and scope_value and vendor_code and scope_value == vendor_code:
            return True
        if scope_type == "zona" and scope_value and zone_name and scope_value == zone_name:
            return True
    return False


def format_internal_customer_summary(customer_row: dict, purchase_summary: Optional[dict] = None):
    customer_name = customer_row.get("nombre_cliente") or customer_row.get("cliente_codigo") or "Cliente"
    customer_code = customer_row.get("cliente_codigo") or "sin código"
    vendor_name = customer_row.get("vendedor") or "sin vendedor asignado"
    zone_name = customer_row.get("zona") or "sin zona"
    balance = format_currency(customer_row.get("saldo_cartera"))
    overdue_docs = customer_row.get("documentos_vencidos") or 0
    overdue_days = customer_row.get("max_dias_vencido") or 0
    last_purchase = customer_row.get("ultima_compra")
    last_purchase_text = str(last_purchase) if last_purchase else "sin compra reciente"
    response_lines = [
        f"{customer_name} ({customer_code})",
        f"Vendedor: {vendor_name}. Zona: {zone_name}.",
        f"Cartera: {balance}. Vencidos: {overdue_docs} doc(s), {overdue_days} día(s).",
        f"Última compra: {last_purchase_text}.",
    ]
    if purchase_summary and purchase_summary.get("top_products"):
        top_product = purchase_summary["top_products"][0]
        response_lines.append(
            f"Top compra reciente: {top_product.get('nombre_articulo') or top_product.get('descripcion') or 'producto'}.")
    return "\n".join(response_lines)


def build_internal_login_reply(content: str, context: dict, conversation_context: dict):
    match = INTERNAL_LOGIN_PATTERN.match(content or "")
    if match:
        username = match.group(1)
        password = match.group(2)
        user_payload = authenticate_internal_user(username, password, context.get("telefono_e164"))
        if not user_payload:
            return {
                "response_text": "No pude autenticarte con ese usuario y contraseña. Revisa el acceso interno.",
                "intent": "internal_auth_login_failed",
                "context_updates": {},
            }
        session_payload = create_internal_session(
            user_payload,
            channel="whatsapp",
            contact_id=context.get("contact_id"),
            phone_e164=context.get("telefono_e164"),
        )
        return {
            "response_text": (
                f"Acceso interno activo para {user_payload.get('full_name')} como {user_payload.get('role')}. "
                "Ya puedes pedir cartera, compras y contexto de tus clientes."
            ),
            "intent": "internal_auth_login",
            "context_updates": {
                "awaiting_internal_auth_cedula": None,
                "internal_auth": build_internal_auth_context(user_payload, session_payload["token"], session_payload["expires_at"]),
            },
        }

    employee_by_phone = find_employee_record_by_phone(context.get("telefono_e164"))
    awaiting_cedula = bool((conversation_context or {}).get("awaiting_internal_auth_cedula"))
    cedula = extract_internal_cedula_candidate(content)
    employee_by_cedula = find_employee_record_by_cedula(cedula) if cedula else None

    # ── Escape hatch: user says they're not an employee → clear auth state ──
    _content_norm = normalize_text_value(content)
    _CANCEL_PHRASES = [
        "no soy colaborador", "no soy empleado", "no soy trabajador",
        "no soy usuario interno", "no soy interno", "no soy de la empresa",
        "soy cliente", "no trabajo ahi", "no trabajo aqui", "no trabajo alla",
        "cancelar", "no quiero login", "salir", "dejame como cliente",
    ]
    _auth_cancelled = bool((conversation_context or {}).get("_auth_cancelled"))
    if any(phrase in _content_norm for phrase in _CANCEL_PHRASES):
        return {
            "response_text": "¡Entendido! Te atiendo como cliente. ¿En qué te puedo ayudar hoy?",
            "intent": "internal_auth_cancelled",
            "context_updates": {"awaiting_internal_auth_cedula": None, "_auth_cancelled": True},
        }

    if not employee_by_phone and not awaiting_cedula and not employee_by_cedula:
        return None
    # If user previously cancelled auth, don't re-enter login flow
    if _auth_cancelled and not awaiting_cedula and not employee_by_cedula:
        return None
    # ── Anti-falso-login: si el teléfono NO es de empleado y NO estamos en flujo auth,
    #    solo proceder si el mensaje es EXCLUSIVAMENTE una cédula (no mezcla con nombre/frase). ──
    if not employee_by_phone and not awaiting_cedula and employee_by_cedula:
        _stripped_msg = (content or "").strip()
        _is_pure_number = bool(re.fullmatch(r"[0-9.\s-]+", _stripped_msg))
        _is_labeled_cedula_only = bool(re.fullmatch(
            r"(?:mi\s+)?(?:cedula|cédula|documento)\s*(?:es|:)?\s*[0-9.\s-]+\s*",
            _stripped_msg, re.IGNORECASE,
        ))
        if not _is_pure_number and not _is_labeled_cedula_only:
            return None  # Message has other content (name, quote request, etc.) — not a login attempt

    if not cedula:
        return {
            "response_text": "Para continuar como colaborador Ferreinox necesito validar tu acceso. Envíame tu número de cédula.",
            "intent": "internal_auth_request_cedula",
            "context_updates": {"awaiting_internal_auth_cedula": True},
        }

    employee_record = employee_by_cedula
    if not employee_record:
        return {
            "response_text": "No encontré esa cédula en la base de colaboradores. Revisa el número o valida con administración.",
            "intent": "internal_auth_login_failed",
            "context_updates": {"awaiting_internal_auth_cedula": True},
        }

    incoming_phone = normalize_phone_e164(context.get("telefono_e164"))
    registered_phone = employee_record.get("phone_e164")
    if registered_phone and incoming_phone and registered_phone != incoming_phone:
        return {
            "response_text": (
                "La cédula existe, pero el número de WhatsApp no coincide con el registrado. "
                "Si eres colaborador, escribe desde tu número autorizado. "
                "Si eres cliente, escribe *'soy cliente'* y con gusto te atiendo."
            ),
            "intent": "internal_auth_phone_mismatch",
            "context_updates": {"awaiting_internal_auth_cedula": True},
        }

    if employee_by_phone and employee_by_phone.get("cedula") != employee_record.get("cedula"):
        return {
            "response_text": "La cédula enviada no coincide con el colaborador asociado a este número de WhatsApp.",
            "intent": "internal_auth_phone_mismatch",
            "context_updates": {"awaiting_internal_auth_cedula": True},
        }

    user_payload = sync_internal_user_from_employee_record(employee_record)
    session_payload = create_internal_session(
        user_payload,
        channel="whatsapp",
        contact_id=context.get("contact_id"),
        phone_e164=context.get("telefono_e164"),
    )
    role_label = user_payload.get("role") or "empleado"
    return {
        "response_text": (
            f"Acceso interno activo para {employee_record.get('full_name')}. "
            f"Cargo: {employee_record.get('cargo') or 'Sin cargo'} | Sede: {employee_record.get('sede') or 'Sin sede'} | Perfil: {role_label}."
        ),
        "intent": "internal_auth_login",
        "context_updates": {
            "awaiting_internal_auth_cedula": None,
            "internal_auth": build_internal_auth_context(user_payload, session_payload["token"], session_payload["expires_at"]),
        },
    }


def build_internal_logout_reply(conversation_context: dict):
    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    if not internal_auth.get("token"):
        return None
    revoke_internal_session(internal_auth.get("token"))
    return {
        "response_text": "La sesión interna quedó cerrada. Si necesitas entrar de nuevo, envíame login usuario clave.",
        "intent": "internal_auth_logout",
        "context_updates": {"internal_auth": None, "awaiting_internal_auth_cedula": None, "internal_transfer_flow": None},
    }


def extract_internal_customer_candidate(text_value: Optional[str]):
    if not text_value:
        return None
    normalized = normalize_text_value(text_value)
    cleaned = normalized
    for fragment in [
        "dame",
        "ver",
        "mostrar",
        "muestreme",
        "muéstrame",
        "consulta",
        "consultar",
        "cartera",
        "compras",
        "compra",
        "cliente",
        "contexto",
        "resumen",
        "del",
        "de",
        "por favor",
        "nit",
        "cedula",
        "cédula",
        "codigo",
        "código",
        "cod",
        "nombre",
        "cuanto",
        "cuánto",
        "compro",
        "compró",
        "suma",
        "suma lo",
        "sumar",
        "total",
        "vencido",
        "vencidos",
        "vencida",
        "vencidas",
        "deuda",
        "debo",
        "debido",
        "ano",
        "año",
        "en",
        "el",
    ]:
        cleaned = re.sub(rf"\b{re.escape(fragment)}\b", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        return None
    candidate = extract_identity_lookup_candidate(cleaned, {}, allow_unprompted=True)
    if candidate:
        return candidate
    digits = re.findall(r"\d{6,15}", text_value or "")
    if digits:
        return {"type": "document", "value": digits[0]}
    tokens = [token for token in cleaned.split() if len(token) >= 3]
    if len(tokens) >= 2:
        return {"type": "name", "value": " ".join(tokens[:6])}
    return None


def resolve_internal_customer_context(identity_candidate: Optional[dict], phone_number: Optional[str] = None):
    if not identity_candidate:
        return None
    candidate_type = identity_candidate.get("type")
    candidate_value = identity_candidate.get("value")
    if candidate_type == "document":
        return find_cliente_contexto_by_document(candidate_value)
    if candidate_type == "customer_code":
        return find_cliente_contexto_by_customer_code(candidate_value)
    if candidate_type == "name":
        return find_cliente_contexto_by_name(candidate_value)
    if candidate_type == "phone":
        return find_cliente_contexto_by_phone(phone_number or candidate_value)
    return None


def _is_clear_customer_product_message(normalized: str) -> bool:
    """Detecta si el mensaje es claramente una consulta de cliente externo sobre productos.
    Esto previene que frases como 'que tiene la pintura vieja' activen el flujo interno."""
    _CUSTOMER_PRODUCT_SIGNALS = [
        "galon", "galones", "cunete", "cuñete", "cuarto", "litro",
        "koraza", "viniltex", "pintulux", "pintucoat", "interthane", "intergard",
        "interseal", "barnex", "esmalte", "anticorrosivo", "corrotec", "pintoxido",
        "pintóxido", "aquablock", "impermeabilizante", "lija", "brocha", "rodillo",
        "pinturama", "intervinil", "sellomax", "eternit", "fibrocemento",
        "pintar", "pintura", "barniz", "sellador", "estuco",
        "humedad", "salitre", "oxido", "óxido", "fachada", "pared", "muro",
        "piso", "techo", "madera", "metal", "hierro", "rejas", "exterior", "interior",
        "obra nueva", "repintura", "pelado", "soplado", "descascarado",
        "me cotizas", "me cotiza", "cotizacion", "cotización", "precio", "precios",
        "cuanto vale", "cuánto vale", "cuanto cuesta", "cuánto cuesta",
        "necesito", "quiero comprar", "metros cuadrados", "m2", "m²",
    ]
    _CUSTOMER_OBJECTION_SIGNALS = [
        "no compro", "no comprar", "no me sirve", "no es para exterior", "no es exterior",
        "para exterior", "para interior", "eso no", "ese no", "esa no",
    ]
    if any(signal in normalized for signal in _CUSTOMER_OBJECTION_SIGNALS):
        return True
    return sum(1 for signal in _CUSTOMER_PRODUCT_SIGNALS if signal in normalized) >= 2


def detect_internal_query_intent(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return None
    # ── Escape temprano: si el mensaje es claramente una consulta de cliente → no es intent interno ──
    if _is_clear_customer_product_message(normalized):
        return None
    if any(fragment in normalized for fragment in ["reclamos pendientes", "pendientes de reclamos", "crm pendientes de reclamos", "reclamos crm", "garantias pendientes", "garantías pendientes", "casos pendientes de reclamo"]):
        return "consulta_reclamos_pendientes"
    if any(fragment in normalized for fragment in ["carro va", "carro para", "van para", "va para", "traemos algo", "traemos mercancia", "traemos mercancía", "mercancia para", "mercancía para", "en ruta"]):
        return "consulta_ruta_mercancia"
    if any(fragment in normalized for fragment in ["pendientes despacho", "pendiente despacho", "pendientes por despachar", "por despachar", "pedidos pendientes", "despachos pendientes"]):
        return "consulta_despachos"
    if any(fragment in normalized for fragment in ["pedir algo a", "mandar algo a", "llevar algo a", "mover algo a", "pasar algo a"]):
        return "consulta_traslados"
    if any(fragment in normalized for fragment in ["genera traslado", "genere traslado", "crear traslado", "crea traslado", "solicita traslado", "traslada", "trasladar"]):
        return "crear_traslado"
    # "que tiene" y "que hay en" solo son internas si NO es contexto de producto/superficie
    # (ej. "que tiene la pintura vieja" es cliente, "que tiene Manizales" es interno)
    _store_context_traslado_signals = ["comparar tiendas", "comparar inventario", "cumplir pedidos", "cumplir pedido", "traslados sugeridos", "sugerencia de traslado"]
    _store_name_signals = ["pereira", "manizales", "armenia", "dosquebradas", "cerritos", "laureles", "tienda", "sede", "bodega principal"]
    _has_store_context = any(s in normalized for s in _store_name_signals)
    if any(fragment in normalized for fragment in _store_context_traslado_signals):
        return "consulta_traslados"
    if _has_store_context and any(fragment in normalized for fragment in ["que hay en", "qué hay en", "que tiene", "qué tiene", "no tenga", "no tiene"]):
        return "consulta_traslados"
    if any(fragment in normalized for fragment in ["cartera", "saldo", "vencid"]):
        return "consulta_cartera"
    _internal_purchase_signals = [
        "compras de", "historial de compras", "ultima compra", "última compra",
        "ultimo pedido", "último pedido", "que compro", "qué compró", "que compro el cliente",
        "qué compró el cliente", "que compro este cliente", "qué compró este cliente",
    ]
    if any(fragment in normalized for fragment in _internal_purchase_signals):
        return "consulta_compras"
    if any(fragment in normalized for fragment in ["contexto", "resumen", "perfil", "todo del cliente", "info del cliente"]):
        return "consulta_contexto"
    return None


def build_pending_dispatches_response(store_code: Optional[str] = None):
    rows = fetch_pending_dispatches(store_code)
    if not rows:
        store_label = get_store_short_label(store_code)
        if store_label:
            return f"No veo pedidos pendientes por despachar para {store_label}."
        return "No veo pedidos pendientes por despachar en este momento."

    response_lines = ["Pedidos pendientes de facturación/despacho:"]
    for row in rows[:6]:
        contacto = row.get("contacto_nombre") or "cliente sin nombre"
        destino = row.get("destination_store_name") or "sede pendiente"
        archivo = row.get("export_filename") or f"pedido_{row.get('order_id')}"
        estado = (row.get("status") or "pendiente").replace("_", " ")
        response_lines.append(
            f"- Pedido {row.get('order_id')} | {destino} | {contacto} | archivo {archivo} | estado {estado}"
        )
    return "\n".join(response_lines)


def build_internal_route_merchandise_response(content: str, internal_user: dict):
    store_mentions = extract_store_mentions_in_order(content)
    if not store_mentions:
        return "Dime la sede destino del carro o de la ruta y te digo qué pedidos, faltantes o traslados veo para esa ciudad.", None

    destination_store_code = store_mentions[-1]
    destination_label = get_store_short_label(destination_store_code) or STORE_CODE_LABELS.get(destination_store_code) or destination_store_code
    internal_metadata = dict((internal_user or {}).get("metadata") or {})
    preferred_origin_store_code = normalize_store_code(internal_metadata.get("store_code"))
    dispatches = fetch_pending_dispatches(destination_store_code, limit=6)
    flow_payload = build_internal_transfer_flow_payload(destination_store_code, preferred_origin_store_code)

    response_lines = [f"Ruta operativa hacia {destination_label}:"]
    if dispatches:
        response_lines.append(f"Veo {len(dispatches)} despacho(s) pendientes o en tránsito:")
        for row in dispatches[:4]:
            status_label = str(row.get("status") or "pendiente").replace("_", " ")
            contacto = row.get("contacto_nombre") or "cliente sin nombre"
            archivo = row.get("export_filename") or f"pedido_{row.get('order_id')}"
            response_lines.append(
                f"- Pedido {row.get('order_id')} | {contacto} | archivo {archivo} | estado {status_label}"
            )
    else:
        response_lines.append(f"No veo despachos pendientes o en tránsito hacia {destination_label} en este momento.")

    suggestions = flow_payload.get("suggestions") or []
    unresolved = flow_payload.get("unresolved_shortages") or []
    if suggestions:
        response_lines.append("Además puedo cubrir faltantes con estos traslados sugeridos:")
        response_lines.append(summarize_transfer_candidates(suggestions, max_items=4))
    if unresolved:
        response_lines.append("Y estos faltantes tocaría escalarlos a compras:")
        response_lines.append(summarize_transfer_candidates(unresolved, max_items=4))
    if suggestions or unresolved:
        response_lines.append("Si quieres, responde `confirmar traslado`, `compras` o `cancelar` y sigo la guía operativa contigo.")

    return "\n".join(response_lines), (flow_payload if (suggestions or unresolved) else None)


def fetch_pending_agent_tasks(task_types: list[str], limit: int = 6, destination_store_code: Optional[str] = None):
    cleaned_task_types = [str(task_type).strip() for task_type in (task_types or []) if str(task_type).strip()]
    if not cleaned_task_types:
        return []

    where_clauses = ["estado IN ('pendiente', 'abierta', 'en_proceso')"]
    params = {"limit": max(1, int(limit or 6))}
    type_placeholders = []
    for index, task_type in enumerate(cleaned_task_types):
        param_name = f"task_type_{index}"
        params[param_name] = task_type
        type_placeholders.append(f":{param_name}")
    where_clauses.append(f"tipo_tarea IN ({', '.join(type_placeholders)})")

    store_code = normalize_store_code(destination_store_code)
    if store_code:
        where_clauses.append("COALESCE(detalle->>'destination_store_code', '') = :destination_store_code")
        params["destination_store_code"] = store_code

    engine = get_db_engine()
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                f"""
                SELECT
                    id,
                    tipo_tarea,
                    prioridad,
                    estado,
                    resumen,
                    detalle,
                    created_at,
                    updated_at
                FROM public.agent_task
                WHERE {' AND '.join(where_clauses)}
                ORDER BY updated_at DESC NULLS LAST, created_at DESC NULLS LAST, id DESC
                LIMIT :limit
                """
            ),
            params,
        ).mappings().all()
    return [dict(row) for row in rows]


def build_internal_procurement_pending_response(content: str, internal_user: dict):
    store_mentions = extract_store_mentions_in_order(content)
    if not store_mentions:
        return "Dime la sede que quieres revisar para compras pendientes, por ejemplo Pereira o Manizales.", None

    destination_store_code = store_mentions[-1]
    destination_label = get_store_short_label(destination_store_code) or STORE_CODE_LABELS.get(destination_store_code) or destination_store_code
    internal_metadata = dict((internal_user or {}).get("metadata") or {})
    preferred_origin_store_code = normalize_store_code(internal_metadata.get("store_code"))
    flow_payload = build_internal_transfer_flow_payload(destination_store_code, preferred_origin_store_code)
    unresolved = flow_payload.get("unresolved_shortages") or []
    suggestions = flow_payload.get("suggestions") or []
    pending_tasks = fetch_pending_agent_tasks(["abastecimiento_compras"], limit=5, destination_store_code=destination_store_code)

    response_lines = [f"Compras pendientes para {destination_label}:"]
    if unresolved:
        response_lines.append("Estos faltantes no tienen origen interno útil y tocaría gestionarlos por compras:")
        response_lines.append(summarize_transfer_candidates(unresolved, max_items=5))
    else:
        response_lines.append(f"No veo faltantes activos sin origen interno para {destination_label} en este momento.")

    if suggestions:
        response_lines.append("Antes de comprar, sí veo estos traslados posibles para cubrir parte de la necesidad:")
        response_lines.append(summarize_transfer_candidates(suggestions, max_items=4))

    if pending_tasks:
        response_lines.append("Además ya hay seguimientos de compras abiertos:")
        for task in pending_tasks[:4]:
            detail = dict(task.get("detalle") or {})
            shortages = detail.get("shortages") or []
            note_suffix = f" | {len(shortages)} faltante(s)" if shortages else ""
            response_lines.append(
                f"- Caso {task.get('id')} | {task.get('resumen') or 'abastecimiento pendiente'} | prioridad {task.get('prioridad') or 'media'}{note_suffix}"
            )

    if unresolved or suggestions:
        response_lines.append("Si quieres, responde `confirmar traslado` para mover lo transferible o `compras` para escalar faltantes sin origen.")

    return "\n".join(response_lines), (flow_payload if (unresolved or suggestions) else None)


def build_internal_pending_claims_response(limit: int = 6):
    tasks = fetch_pending_agent_tasks(["reclamo_calidad"], limit=limit)
    if not tasks:
        return "No veo reclamos pendientes abiertos en el CRM en este momento."

    response_lines = ["Reclamos pendientes en CRM:"]
    for task in tasks[:limit]:
        detail = dict(task.get("detalle") or {})
        cliente_label = detail.get("cliente") or detail.get("nombre_cliente") or "cliente sin nombre"
        product_label = detail.get("product_label") or detail.get("producto_reclamado") or "producto pendiente"
        store_name = detail.get("store_name") or "sede no indicada"
        response_lines.append(
            f"- Caso {task.get('id')} | {cliente_label} | {product_label} | {store_name} | prioridad {task.get('prioridad') or 'media'}"
        )
    return "\n".join(response_lines)


def build_transfer_suggestions_response(content: str, internal_user: dict):
    store_mentions = extract_store_mentions_in_order(content)
    normalized = normalize_text_value(content)
    internal_metadata = dict((internal_user or {}).get("metadata") or {})
    internal_store_code = normalize_store_code(internal_metadata.get("store_code"))
    if len(store_mentions) >= 2:
        source_store_code, destination_store_code = store_mentions[0], store_mentions[1]
    elif len(store_mentions) == 1:
        source_store_code, destination_store_code = internal_store_code, store_mentions[0]
    else:
        source_store_code, destination_store_code = internal_store_code, None

    if destination_store_code and any(fragment in normalized for fragment in ["cumplir pedidos", "cumplir pedido", "pendiente", "despacho"]):
        suggestions = build_transfer_suggestions_for_pending_dispatches(destination_store_code, source_store_code)
        if not suggestions:
            destination_label = get_store_short_label(destination_store_code) or STORE_CODE_LABELS.get(destination_store_code) or "la sede consultada"
            return f"No encontré faltantes pendientes con traslado útil hacia {destination_label}.", []
        response_lines = ["Sugerencias de traslado para cubrir pedidos pendientes:"]
        for suggestion in suggestions[:6]:
            response_lines.append(
                f"- Pedido {suggestion['order_id']} | {suggestion['reference']} | {suggestion['origin_store_name']} -> {suggestion['destination_store_name']} | sugerido {format_quantity(suggestion['suggested_qty'])} und"
            )
        return "\n".join(response_lines), suggestions

    if source_store_code and destination_store_code:
        gaps = fetch_inventory_gap_between_stores(source_store_code, destination_store_code)
        if not gaps:
            source_label = get_store_short_label(source_store_code) or STORE_CODE_LABELS.get(source_store_code) or source_store_code
            destination_label = get_store_short_label(destination_store_code) or STORE_CODE_LABELS.get(destination_store_code) or destination_store_code
            return f"No veo referencias con stock en {source_label} y quiebre total en {destination_label}.", []
        response_lines = [
            f"Referencias que sí están en {get_store_short_label(source_store_code) or STORE_CODE_LABELS.get(source_store_code)} y no están en {get_store_short_label(destination_store_code) or STORE_CODE_LABELS.get(destination_store_code)}:"
        ]
        for gap in gaps[:8]:
            response_lines.append(
                f"- {gap['referencia']} | {gap.get('descripcion') or 'sin descripción'} | stock origen {format_quantity(gap['stock_origen'])}"
            )
        return "\n".join(response_lines), []

    suggestions = build_transfer_suggestions_for_pending_dispatches(destination_store_code, source_store_code)
    if suggestions:
        response_lines = ["Traslados sugeridos sobre pedidos pendientes:"]
        for suggestion in suggestions[:6]:
            response_lines.append(
                f"- Pedido {suggestion['order_id']} | {suggestion['reference']} | {suggestion['origin_store_name']} -> {suggestion['destination_store_name']} | sugerido {format_quantity(suggestion['suggested_qty'])} und"
            )
        return "\n".join(response_lines), suggestions

    return "Dime la tienda destino o compárame dos tiendas, por ejemplo: 'qué hay en Manizales que Pereira no tenga' o 'traslada a Pereira para cumplir pedidos'.", []


def handle_pending_internal_transfer_flow(content: str, context: dict, conversation_context: dict, internal_user: dict, internal_auth: dict):
    flow_payload = dict((conversation_context or {}).get("internal_transfer_flow") or {})
    if not flow_payload:
        return None

    # ── ENSEÑANZA tiene prioridad sobre el flujo de traslados ──
    # Si un experto dice ENSEÑAR mientras hay un transfer flow activo,
    # liberamos el flujo y dejamos pasar al LLM para registrar conocimiento.
    _TEACHING_ESCAPE_SIGNALS = [
        "ensenar", "ensenanza", "anota esto", "guarda esto", "aprender esto",
        "enseñar", "enseñanza",
        "la recomendacion esta mal", "la recomendación está mal",
        "no es asi", "no es así", "eso esta mal", "eso está mal",
        "te equivocaste", "la respuesta esta mal", "la respuesta está mal",
    ]
    _norm_teach = normalize_text_value(content)
    if any(signal in _norm_teach for signal in _TEACHING_ESCAPE_SIGNALS):
        # Clear the transfer flow and let the message pass to the LLM
        if isinstance(conversation_context, dict):
            conversation_context["internal_transfer_flow"] = None
        return None

    normalized = normalize_text_value(content)
    if normalized in {"cancelar", "cancelar traslado", "salir", "no"}:
        return {
            "response_text": "Listo, cierro esta guía operativa de traslados y abastecimiento.",
            "intent": "internal_transfer_flow_cancelled",
            "context_updates": {
                "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                "internal_transfer_flow": None,
            },
        }

    if flow_payload.get("step") == "awaiting_destination":
        store_mentions = extract_store_mentions_in_order(content)
        destination_store_code = store_mentions[-1] if store_mentions else None
        if not destination_store_code:
            return {
                "response_text": "Todavía me falta la sede destino. Dímela como ciudad o tienda, por ejemplo: Manizales, Armenia, Laureles o Cerritos.",
                "intent": "internal_transfer_flow_destination_missing",
                "context_updates": {
                    "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                    "internal_transfer_flow": flow_payload,
                },
            }
        refreshed_payload = build_internal_transfer_flow_payload(destination_store_code, flow_payload.get("origin_store_code"))
        return {
            "response_text": build_internal_transfer_guidance_text(refreshed_payload),
            "intent": "internal_transfer_flow_ready",
            "context_updates": {
                "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                "internal_transfer_flow": refreshed_payload,
            },
        }

    selected_indexes = parse_internal_selection_indexes(content, len(flow_payload.get("suggestions") or []))
    suggestions = flow_payload.get("suggestions") or []
    selected_suggestions = [suggestions[index - 1] for index in selected_indexes] if selected_indexes else suggestions
    wants_transfer = any(token in normalized for token in ["confirmar", "crear", "traslado", "si", "sí", "listo"])
    wants_procurement = any(token in normalized for token in ["compras", "abastecimiento", "escalar", "comprar"])

    if wants_transfer and selected_suggestions:
        created_requests = create_transfer_request_records(selected_suggestions[:5], internal_user, notes=content)
        if not created_requests:
            return {
                "response_text": "Las opciones seleccionadas no pertenecen a tu sede de origen o ya no aplican. Puedo recalcular si me dices nuevamente la sede destino.",
                "intent": "internal_transfer_origin_mismatch",
                "context_updates": {
                    "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                    "internal_transfer_flow": None,
                },
            }
        notification_result = notify_transfer_requests_by_email(created_requests, internal_user, notes=content)
        upsert_agent_task(
            context["conversation_id"],
            context.get("cliente_id"),
            "traslado_interno",
            "Solicitud interna de traslado generada por guía conversacional",
            {"solicitudes": created_requests, "mensaje": content, "notificacion": notification_result},
            "alta",
        )
        response_lines = ["Traslados creados desde la guía operativa:"]
        response_lines.append(summarize_transfer_candidates(created_requests))
        if notification_result.get("sent"):
            for row in notification_result["sent"]:
                response_lines.append(f"Correo enviado a {row['to_email']} con copia a {', '.join(row.get('cc_emails') or [])}.")
        if notification_result.get("errors"):
            response_lines.append("Errores de notificación: " + "; ".join(notification_result["errors"]))
        return {
            "response_text": "\n".join(response_lines),
            "intent": "crear_traslado",
            "context_updates": {
                "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                "internal_transfer_flow": None,
            },
        }

    unresolved = flow_payload.get("unresolved_shortages") or []
    if wants_procurement and unresolved:
        followup = create_procurement_followup(context, internal_user, unresolved[:8], notes=content)
        response_lines = ["Faltantes escalados a compras para abastecimiento:", summarize_transfer_candidates(unresolved)]
        notification = followup.get("notification") or {}
        if notification.get("sent"):
            response_lines.append(f"Correo enviado a {notification.get('to_email')} con adjunto {notification.get('attachment')}.")
        if notification.get("errors"):
            response_lines.append("Errores de notificación: " + "; ".join(notification.get("errors") or []))
        return {
            "response_text": "\n".join([line for line in response_lines if line]),
            "intent": "internal_procurement_escalated",
            "context_updates": {
                "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                "internal_transfer_flow": None,
            },
        }

    # ── Loop-escape: user signals they already acted or is dismissing the flow ──
    done_phrases = [
        "ya", "ya lo hice", "ya confirme", "ya fue", "ya quedo", "ya di la orden",
        "ya esta", "ok", "entendido", "gracias", "de acuerdo", "listo ya",
        "ya confirme", "ya se hizo", "ya se creo",
    ]
    is_done_signal = any(phrase in normalized for phrase in done_phrases)
    if is_done_signal and not wants_transfer and not wants_procurement:
        return {
            "response_text": "Perfecto, queda en curso. Avísame si necesitas algo más con traslados o abastecimiento.",
            "intent": "internal_transfer_flow_acknowledged",
            "context_updates": {
                "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                "internal_transfer_flow": None,
            },
        }

    # ── Auto-clear after 3 consecutive unrecognized messages (prevent infinite loop) ──
    turn_count = (flow_payload.get("_unrecognized_turns") or 0) + 1
    if turn_count >= 3:
        return {
            "response_text": "Cerrando la guía operativa de traslados. Si necesitas retomar, menciona 'faltantes', 'traslado' o 'abastecimiento'.",
            "intent": "internal_transfer_flow_timeout",
            "context_updates": {
                "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
                "internal_transfer_flow": None,
            },
        }
    flow_payload["_unrecognized_turns"] = turn_count

    return {
        "response_text": build_internal_transfer_guidance_text(flow_payload),
        "intent": "internal_transfer_flow_pending",
        "context_updates": {
            "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
            "internal_transfer_flow": flow_payload,
        },
    }


def handle_internal_whatsapp_message(content: Optional[str], context: dict, conversation_context: dict):
    if not content:
        return None

    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    active_internal_user = resolve_internal_session(internal_auth.get("token")) if internal_auth.get("token") else None

    if active_internal_user:
        internal_auth = build_internal_auth_context(
            active_internal_user,
            internal_auth.get("token"),
            active_internal_user.get("session_expires_at"),
        )
        if isinstance(conversation_context, dict):
            conversation_context["internal_auth"] = internal_auth
            conversation_context["awaiting_internal_auth_cedula"] = None
    elif internal_auth.get("token"):
        return {
            "response_text": "La sesión interna venció. Vuelve a ingresar con login usuario clave o envíame tu cédula nuevamente.",
            "intent": "internal_auth_expired",
            "context_updates": {"internal_auth": None, "awaiting_internal_auth_cedula": None, "internal_transfer_flow": None},
        }

    normalized = normalize_text_value(content)
    if not active_internal_user:
        login_reply = build_internal_login_reply(content, context, conversation_context)
        if login_reply:
            return login_reply

    if normalized in INTERNAL_LOGOUT_PATTERNS:
        return build_internal_logout_reply(conversation_context)

    employee_by_phone = find_employee_record_by_phone(context.get("telefono_e164"))
    _auth_was_cancelled = bool((conversation_context or {}).get("_auth_cancelled"))

    # ── Escape temprano: si el mensaje es claramente de cliente externo (productos, m², precios),
    #    NO bloquear con flujo de empleado aunque el teléfono esté en la base de empleados.
    _norm_for_customer_check = normalize_text_value(content)
    if _is_clear_customer_product_message(_norm_for_customer_check):
        return None

    if employee_by_phone and not internal_auth.get("token") and not _auth_was_cancelled:
        return {
            "response_text": "Antes de seguir necesito validar tu acceso interno. Envíame tu cédula para activar la sesión de colaborador.",
            "intent": "internal_auth_request_cedula",
            "context_updates": {"awaiting_internal_auth_cedula": True},
        }

    if not active_internal_user:
        if detect_internal_query_intent(content):
            return {
                "response_text": "Para consultas internas primero debes iniciar sesión. Si eres colaborador, envíame tu cédula. Si eres usuario técnico legado, puedes escribir login usuario clave.",
                "intent": "internal_auth_required",
                "context_updates": {},
            }
        return None

    internal_user = active_internal_user

    pending_transfer_flow_reply = handle_pending_internal_transfer_flow(content, context, conversation_context, internal_user, internal_auth)
    if pending_transfer_flow_reply:
        return pending_transfer_flow_reply

    # ── Si hay un reclamo activo en contexto o el mensaje es de reclamo, dejar que el LLM lo maneje ──
    _claim_case = (conversation_context or {}).get("claim_case") or {}
    if _claim_case.get("active") and not _claim_case.get("submitted"):
        return None  # reclamo en curso → pasa al LLM

    _norm_claim = normalize_text_value(content)
    _CLAIM_SIGNALS = [
        "reclam", "queja", "garantia", "garantía", "devolucion", "devolución",
        "producto malo", "producto dañado", "no sirve", "no funciona",
        "se peló", "se pela", "no cubre", "se agrietó", "se cuarteó",
        "grumo", "mal olor", "olor raro", "vencido", "caducado",
        "defecto", "filtra", "filtrando",
    ]
    if any(signal in _norm_claim for signal in _CLAIM_SIGNALS):
        return None  # señal de reclamo → pasa al LLM con flujo de 5 fases

    # ── ENSEÑANZA / CORRECCIÓN de Pablo u otros expertos → dejar que el LLM lo maneje ──
    # Debe evaluarse ANTES de detect_internal_query_intent para evitar que palabras
    # como "compra", "traslado" dentro de una corrección técnica secuestren el flujo.
    _TEACHING_SIGNALS = [
        "ensenar", "ensenanza", "anota esto", "guarda esto", "aprender esto",
        "enseñar", "enseñanza",
        # Señales de corrección implícita (Pablo corrigiendo sin palabra clave)
        "la recomendacion esta mal", "la recomendación está mal",
        "no es asi", "no es así", "eso esta mal", "eso está mal",
        "te equivocaste", "mucho cuidado con", "ojo que", "no confundas",
        "el proceso no esta bien", "el proceso no está bien",
        "primero se debe", "el orden correcto es", "lo que hay que hacer es",
        "en realidad se usa", "ese producto no sirve para",
        "para este caso es mejor", "la respuesta esta mal", "la respuesta está mal",
        "debes responder", "debes preguntar", "primero el estado",
    ]
    _norm_for_teach = normalize_text_value(content)
    if any(signal in _norm_for_teach for signal in _TEACHING_SIGNALS):
        return None  # enseñanza/corrección experta → pasa al LLM para registrar_conocimiento_experto

    # ── Detección de asesoría técnica en curso (diagnóstico/recomendación de producto) ──
    # Si el mensaje contiene señales de consulta técnica de producto/superficie,
    # dejar que el LLM maneje el flujo diagnóstico completo en vez de rutear a intents internos.
    _TECHNICAL_ADVISORY_SIGNALS = [
        "pintar un piso", "pintar piso", "pintar una pared", "pintar pared",
        "pintar un techo", "pintar techo", "pintar fachada", "pintar madera",
        "pintar metal", "pintar hierro", "humedad", "impermeabilizar",
        "que me recomiendas", "qué me recomiendas", "alternativas para pintar",
        "sistema para", "necesito pintar", "quiero pintar",
        "que producto uso", "qué producto uso", "como protejo", "cómo protejo",
    ]
    if any(signal in _norm_for_teach for signal in _TECHNICAL_ADVISORY_SIGNALS):
        return None  # consulta técnica → pasa al LLM con diagnóstico + RAG

    intent = detect_internal_query_intent(content)
    if not intent:
        return None

    if intent == "consulta_despachos":
        if not internal_user_has_advanced_access(internal_user):
            return {
                "response_text": "Tu perfil actual puede crear pedidos, pero no consultar despachos ni operación interna completa.",
                "intent": "internal_access_denied",
                "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
            }
        store_mentions = extract_store_mentions_in_order(content)
        store_code = store_mentions[0] if store_mentions else None
        return {
            "response_text": build_pending_dispatches_response(store_code),
            "intent": intent,
            "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
        }

    if intent == "consulta_ruta_mercancia":
        if not internal_user_has_advanced_access(internal_user):
            return {
                "response_text": "Tu perfil actual puede crear pedidos, pero no consultar la operación logística interna completa.",
                "intent": "internal_access_denied",
                "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
            }
        response_text, flow_payload = build_internal_route_merchandise_response(content, internal_user)
        context_updates = {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))}
        if flow_payload:
            context_updates["internal_transfer_flow"] = flow_payload
        return {
            "response_text": response_text,
            "intent": intent,
            "context_updates": context_updates,
        }

    if intent == "consulta_reclamos_pendientes":
        if not internal_user_has_advanced_access(internal_user):
            return {
                "response_text": "Tu perfil actual puede crear pedidos, pero no consultar la bandeja interna de reclamos.",
                "intent": "internal_access_denied",
                "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
            }
        return {
            "response_text": build_internal_pending_claims_response(),
            "intent": intent,
            "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
        }

    if intent in {"consulta_traslados", "crear_traslado"}:
        if not internal_user_has_advanced_access(internal_user):
            return {
                "response_text": "Tu perfil actual puede tomar pedidos, pero no consultar ni gestionar traslados internos.",
                "intent": "internal_access_denied",
                "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
            }
        response_text, suggestions = build_transfer_suggestions_response(content, internal_user)
        context_updates = {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))}
        store_mentions = extract_store_mentions_in_order(content)
        requested_destination = store_mentions[-1] if store_mentions else None
        internal_metadata = dict((internal_user or {}).get("metadata") or {})
        default_origin_store = normalize_store_code(internal_metadata.get("store_code"))
        if not requested_destination:
            context_updates["internal_transfer_flow"] = {
                "step": "awaiting_destination",
                "origin_store_code": default_origin_store,
            }
            return {
                "response_text": "¿Para qué sede quieres revisar faltantes o generar traslado? Si no me dices origen, tomaré como origen tu sede autenticada.",
                "intent": "internal_transfer_flow_destination_missing",
                "context_updates": context_updates,
            }

        flow_payload = build_internal_transfer_flow_payload(requested_destination, default_origin_store)
        context_updates["internal_transfer_flow"] = flow_payload
        if intent == "crear_traslado":
            if not internal_user_can_manage_transfers(internal_user):
                return {
                    "response_text": "Solo el facturador de la sede origen puede crear solicitudes de traslado. Debes tener cargo de líder de tienda, líder logístico/inventario o auxiliar de facturación.",
                    "intent": "internal_transfer_forbidden",
                    "context_updates": context_updates,
                }
            if not flow_payload.get("suggestions") and not flow_payload.get("unresolved_shortages"):
                return {
                    "response_text": response_text,
                    "intent": intent,
                    "context_updates": context_updates,
                }
            return {
                "response_text": build_internal_transfer_guidance_text(flow_payload),
                "intent": "internal_transfer_flow_ready",
                "context_updates": context_updates,
            }
        return {
            "response_text": build_internal_transfer_guidance_text(flow_payload) if (flow_payload.get("suggestions") or flow_payload.get("unresolved_shortages")) else response_text,
            "intent": intent,
            "context_updates": context_updates,
        }

    if intent == "consulta_compras":
        candidate = extract_internal_customer_candidate(content)
        store_mentions = extract_store_mentions_in_order(content)
        normalized_content = normalize_text_value(content)
        has_explicit_customer_reference = bool(re.search(r"\b\d{6,15}\b", content or "")) or bool(
            re.search(r"\b(cliente|nit|cedula|cédula|codigo cliente|código cliente)\b", normalized_content)
        )
        if store_mentions and not has_explicit_customer_reference:
            if not internal_user_has_advanced_access(internal_user):
                return {
                    "response_text": "Tu perfil actual puede crear pedidos, pero no consultar compras y faltantes internos.",
                    "intent": "internal_access_denied",
                    "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
                }
            response_text, flow_payload = build_internal_procurement_pending_response(content, internal_user)
            context_updates = {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))}
            if flow_payload:
                context_updates["internal_transfer_flow"] = flow_payload
            return {
                "response_text": response_text,
                "intent": "consulta_compras_pendientes",
                "context_updates": context_updates,
            }

    candidate = extract_internal_customer_candidate(content)
    last_cliente_codigo = (conversation_context or {}).get("internal_last_cliente_codigo")
    cliente_contexto = resolve_internal_customer_context(candidate, context.get("telefono_e164")) if candidate else None
    if (not cliente_contexto or not cliente_contexto.get("cliente_codigo")) and last_cliente_codigo and not candidate:
        cliente_contexto = fetch_customer_lookup_row(last_cliente_codigo) or {"cliente_codigo": last_cliente_codigo}

    if not cliente_contexto or not cliente_contexto.get("cliente_codigo"):
        if not candidate:
            return {
                "response_text": "Dime el NIT, la cédula, el código o el nombre del cliente y te muestro la información.",
                "intent": "internal_customer_missing",
                "context_updates": {},
            }
        return {
            "response_text": "No encontré ese cliente con la información enviada. Prueba con NIT, código cliente o nombre completo.",
            "intent": "internal_customer_not_found",
            "context_updates": {},
        }

    if not internal_user_has_advanced_access(internal_user):
        return {
            "response_text": "Tu perfil actual puede crear pedidos, pero no consultar cartera, compras ni CRM interno completo.",
            "intent": "internal_access_denied",
            "context_updates": {"internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at"))},
        }

    customer_row = fetch_customer_lookup_row(cliente_contexto.get("cliente_codigo")) or cliente_contexto
    if not internal_user_can_access_customer(internal_user, customer_row):
        return {
            "response_text": "No tienes permisos para consultar ese cliente con tu rol actual.",
            "intent": "internal_access_denied",
            "context_updates": {},
        }

    if intent == "consulta_cartera":
        cartera_query = extract_cartera_query(content)
        overdue_info = fetch_overdue_documents(cliente_contexto.get("cliente_codigo"))
        totals = (overdue_info or {}).get("totals") or {}
        if cartera_query.get("wants_overdue_only"):
            response_text = (
                f"{customer_row.get('nombre_cliente') or cliente_contexto.get('cliente_codigo')}\n"
                f"Cartera vencida: {format_currency(totals.get('saldo_vencido'))}.\n"
                f"Documentos vencidos: {totals.get('documentos_vencidos', customer_row.get('documentos_vencidos') or 0)} | "
                f"máximo atraso: {totals.get('max_dias_vencido', customer_row.get('max_dias_vencido') or 0)} día(s)."
            )
        else:
            response_text = (
                f"{customer_row.get('nombre_cliente') or cliente_contexto.get('cliente_codigo')}\n"
                f"Saldo cartera: {format_currency(customer_row.get('saldo_cartera'))}.\n"
                f"Vencidos: {totals.get('documentos_vencidos', customer_row.get('documentos_vencidos') or 0)} documento(s), "
                f"máximo {totals.get('max_dias_vencido', customer_row.get('max_dias_vencido') or 0)} día(s)."
            )
    elif intent == "consulta_compras":
        purchase_query = extract_purchase_query(content)
        if purchase_query.get("wants_last_purchase"):
            purchase_summary = fetch_latest_purchase_detail(cliente_contexto.get("cliente_codigo"))
            if not purchase_summary:
                response_text = f"No encontré compras recientes para {customer_row.get('nombre_cliente') or cliente_contexto.get('cliente_codigo')}."
            else:
                totals = purchase_summary.get("totals") or {}
                top_products = ", ".join(
                    (item.get("nombre_articulo") or "")
                    for item in (purchase_summary.get("products") or [])[:3]
                    if item.get("nombre_articulo")
                ) or "sin detalle"
                response_text = (
                    f"{customer_row.get('nombre_cliente') or cliente_contexto.get('cliente_codigo')}\n"
                    f"Última compra: {purchase_summary.get('fecha_venta') or 'sin fecha'} por {format_currency(totals.get('valor_total'))}.\n"
                    f"Productos: {top_products}"
                )
        else:
            purchases = fetch_purchase_summary(
                cliente_contexto.get("cliente_codigo"),
                purchase_query.get("start_date"),
                purchase_query.get("end_date"),
            )
            response_text = build_purchase_summary_response_text(
                customer_row.get("nombre_cliente") or cliente_contexto.get("cliente_codigo"),
                purchase_query,
                purchases,
            )
    else:
        purchase_summary = fetch_latest_purchase_detail(cliente_contexto.get("cliente_codigo"))
        response_text = format_internal_customer_summary(customer_row, purchase_summary)

    return {
        "response_text": response_text,
        "intent": intent,
        "context_updates": {
            "internal_auth": build_internal_auth_context(internal_user, internal_auth.get("token"), internal_user.get("session_expires_at")),
            "internal_last_cliente_codigo": cliente_contexto.get("cliente_codigo"),
        },
    }


def load_local_secrets():
    if not SECRETS_PATH.exists():
        return {}
    return tomllib.loads(SECRETS_PATH.read_text(encoding="utf-8"))


def get_dropbox_ventas_config():
    env_config = {
        "app_key": os.getenv("DROPBOX_VENTAS_APP_KEY"),
        "app_secret": os.getenv("DROPBOX_VENTAS_APP_SECRET"),
        "refresh_token": os.getenv("DROPBOX_VENTAS_REFRESH_TOKEN"),
        "folder": os.getenv("DROPBOX_VENTAS_FOLDER") or "/data",
    }
    if env_config["app_key"] and env_config["app_secret"] and env_config["refresh_token"]:
        return env_config

    secrets = load_local_secrets()
    config = secrets.get("dropbox_ventas") or {}
    if config.get("app_key") and config.get("app_secret") and config.get("refresh_token"):
        return config
    raise RuntimeError("No se encontró configuración válida para Dropbox Ventas.")


def get_dropbox_ventas_client():
    config = get_dropbox_ventas_config()
    return dropbox.Dropbox(
        oauth2_refresh_token=config["refresh_token"],
        app_key=config["app_key"],
        app_secret=config["app_secret"],
    )


def get_sendgrid_config():
    env_config = {
        "api_key": os.getenv("SENDGRID_API_KEY"),
        "from_email": os.getenv("SENDGRID_FROM_EMAIL"),
        "from_name": os.getenv("SENDGRID_FROM_NAME") or "Ferreinox S.A.S. BIC",
        "reclamos_to_email": os.getenv("SENDGRID_RECLAMOS_TO_EMAIL") or os.getenv("SENDGRID_QUALITY_TO_EMAIL"),
        "ventas_to_email": os.getenv("SENDGRID_VENTAS_TO_EMAIL"),
        "contabilidad_to_email": os.getenv("SENDGRID_CONTABILIDAD_TO_EMAIL"),
    }
    if env_config["api_key"] and env_config["from_email"]:
        return env_config

    secrets = load_local_secrets()
    config = secrets.get("sendgrid") or {}
    if config.get("api_key") and config.get("from_email"):
        return {
            "api_key": config.get("api_key"),
            "from_email": config.get("from_email"),
            "from_name": config.get("from_name") or "Ferreinox S.A.S. BIC",
            "reclamos_to_email": config.get("reclamos_to_email") or config.get("quality_to_email") or config.get("from_email"),
            "ventas_to_email": config.get("ventas_to_email"),
            "contabilidad_to_email": config.get("contabilidad_to_email"),
        }
    return None


def send_sendgrid_email(
    to_email: str,
    subject: str,
    html_content: str,
    text_content: str,
    reply_to: Optional[str] = None,
    cc_emails: Optional[list[str]] = None,
    attachments: Optional[list[dict]] = None,
):
    config = get_sendgrid_config()
    if not config:
        raise RuntimeError("SendGrid no está configurado.")

    personalization = {"to": [{"email": to_email}], "subject": subject}
    cc_payload = [{"email": email} for email in (cc_emails or []) if email]
    if cc_payload:
        personalization["cc"] = cc_payload

    payload = {
        "personalizations": [personalization],
        "from": {
            "email": config["from_email"],
            "name": config.get("from_name") or "Ferreinox S.A.S. BIC",
        },
        "content": [
            {"type": "text/plain", "value": text_content},
            {"type": "text/html", "value": html_content},
        ],
    }
    if reply_to:
        payload["reply_to"] = {"email": reply_to}
    if attachments:
        payload["attachments"] = attachments

    response = requests.post(
        "https://api.sendgrid.com/v3/mail/send",
        headers={
            "Authorization": f"Bearer {config['api_key']}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=20,
    )
    if response.status_code >= 400:
        try:
            error_payload = response.json()
        except Exception:
            error_payload = {"raw": response.text}
        raise RuntimeError(f"SendGrid devolvió {response.status_code}: {safe_json_dumps(error_payload)}")
    return True


def run_background_io(task_name: str, target, *args, **kwargs):
    def _runner():
        try:
            target(*args, **kwargs)
        except Exception as exc:
            logger.warning("Background task %s failed: %s", task_name, exc)

    thread = threading.Thread(target=_runner, name=f"bg-{task_name}", daemon=True)
    thread.start()
    return thread


_PROCESSING_WATCHDOGS: dict[str, dict] = {}
_PROCESSING_WATCHDOG_LOCK = threading.Lock()
PROCESSING_FOLLOWUP_SECONDS = int(os.getenv("WA_PROCESSING_FOLLOWUP_SECONDS", "120"))


def _send_processing_status_message(context: dict, body: str, status_tag: str):
    try:
        outbound_payload = send_whatsapp_text_message(context["telefono_e164"], body)
        provider_message_id = None
        if outbound_payload.get("messages"):
            provider_message_id = outbound_payload["messages"][0].get("id")
        store_outbound_message(
            context["conversation_id"],
            provider_message_id,
            "text",
            body,
            outbound_payload,
            intent_detectado=status_tag,
        )
    except Exception as exc:
        logger.warning("Failed to send processing status message (%s): %s", status_tag, exc)


def should_send_processing_ack(content: Optional[str], conversation_context: Optional[dict]) -> bool:
    normalized = normalize_text_value(content or "")
    draft = dict((conversation_context or {}).get("commercial_draft") or {})
    draft_has_items = bool(draft.get("items"))
    long_flow_tokens = [
        "cotizacion", "cotizar", "pedido", "pdf", "confirmar", "confirmo",
        "procede", "proceder", "generar", "genera", "envia", "enviar",
        "manda", "cerrar pedido", "cerrar cotizacion",
    ]
    if draft_has_items and any(token in normalized for token in long_flow_tokens):
        return True
    explicit_close_tokens = [
        "genera la cotizacion", "genera cotizacion", "envia la cotizacion", "manda la cotizacion",
        "genera el pedido", "genera pedido", "confirmar pedido", "confirmar cotizacion",
        "envia el pdf", "manda el pdf", "procede con el pedido", "procede con la cotizacion",
    ]
    return any(token in normalized for token in explicit_close_tokens)


def start_processing_watchdog(context: dict, initial_message: Optional[str] = None):
    key = f"{context.get('conversation_id')}:{context.get('telefono_e164')}"
    stop_event = threading.Event()

    with _PROCESSING_WATCHDOG_LOCK:
        previous = _PROCESSING_WATCHDOGS.pop(key, None)
        if previous:
            previous["stop_event"].set()
        _PROCESSING_WATCHDOGS[key] = {"stop_event": stop_event}

    if initial_message:
        _send_processing_status_message(context, initial_message, "processing_ack")

    def _runner():
        followup_count = 0
        while not stop_event.wait(PROCESSING_FOLLOWUP_SECONDS):
            followup_count += 1
            if followup_count == 1:
                body = "⏳ Un momento por favor, sigo procesando tu solicitud y casi termino."
            else:
                body = "⏳ Gracias por la espera. Sigo trabajando en tu solicitud para entregártela completa."
            _send_processing_status_message(context, body, "processing_followup")

        with _PROCESSING_WATCHDOG_LOCK:
            current = _PROCESSING_WATCHDOGS.get(key)
            if current and current.get("stop_event") is stop_event:
                _PROCESSING_WATCHDOGS.pop(key, None)

    thread = threading.Thread(target=_runner, name=f"processing-watchdog-{context.get('conversation_id')}", daemon=True)
    thread.start()
    return key


def stop_processing_watchdog(watchdog_key: Optional[str]):
    if not watchdog_key:
        return
    with _PROCESSING_WATCHDOG_LOCK:
        current = _PROCESSING_WATCHDOGS.pop(watchdog_key, None)
    if current:
        current["stop_event"].set()


def normalize_store_code(store_value: Optional[str]):
    normalized = normalize_text_value(store_value)
    if not normalized:
        return None
    if normalized.isdigit() and normalized in STORE_CODE_LABELS:
        return normalized
    for aliases in STORE_ALIASES.values():
        store_code = next((candidate for candidate in aliases if candidate.isdigit()), None)
        if not store_code:
            continue
        for alias in aliases:
            if normalize_text_value(alias) == normalized:
                return store_code
    return None


def get_store_short_label(store_code: Optional[str]):
    store_code = normalize_store_code(store_code)
    if not store_code:
        return None
    label = STORE_CODE_LABELS.get(store_code) or store_code
    return re.sub(r"^Tienda\s+", "", label, flags=re.IGNORECASE).strip()


def extract_store_stock_from_summary(stock_summary: Optional[str], store_code: Optional[str]):
    normalized_code = normalize_store_code(store_code)
    if not stock_summary or not normalized_code:
        return None

    match_tokens = set()
    for alias in STORE_ALIASES.get(next((key for key, aliases in STORE_ALIASES.items() if normalized_code in aliases), ""), []):
        alias_normalized = normalize_text_value(alias)
        if alias_normalized and not alias_normalized.isdigit():
            match_tokens.add(alias_normalized)
    label_normalized = normalize_text_value(STORE_CODE_LABELS.get(normalized_code) or "")
    if label_normalized:
        match_tokens.add(label_normalized)

    for fragment in str(stock_summary).split(";"):
        left_value, _, right_value = fragment.partition(":")
        left_normalized = normalize_text_value(left_value)
        if any(token and token in left_normalized for token in match_tokens):
            return parse_numeric_value(right_value)
    return None


def filter_previous_product_context(conversation_context: Optional[dict], product_request: Optional[dict]):
    previous_rows = list((conversation_context or {}).get("last_product_context") or [])
    if not previous_rows:
        return []

    filtered_rows = previous_rows
    requested_unit = (product_request or {}).get("requested_unit")
    if requested_unit:
        unit_rows = [row for row in filtered_rows if infer_product_presentation_from_row(row) == requested_unit]
        if unit_rows:
            filtered_rows = unit_rows

    requested_store_codes = (product_request or {}).get("store_filters") or []
    if len(requested_store_codes) == 1:
        requested_store_code = requested_store_codes[0]
        visible_rows = []
        for row in filtered_rows:
            store_stock = extract_store_stock_from_summary(row.get("stock_por_tienda"), requested_store_code)
            row_copy = dict(row)
            row_copy["stock_en_tienda_solicitada"] = store_stock
            row_copy["visibilidad_tienda_exacta"] = store_stock is not None
            visible_rows.append(row_copy)
        filtered_rows = visible_rows

    return filtered_rows[:10]


def extract_store_mentions_in_order(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return []

    found_matches = []
    for aliases in STORE_ALIASES.values():
        store_code = next((candidate for candidate in aliases if candidate.isdigit()), None)
        if not store_code:
            continue
        best_pos = None
        for alias in aliases:
            alias_normalized = normalize_text_value(alias)
            if not alias_normalized:
                continue
            match = re.search(rf"\b{re.escape(alias_normalized)}\b", normalized)
            if match and (best_pos is None or match.start() < best_pos):
                best_pos = match.start()
        if best_pos is not None:
            found_matches.append((best_pos, store_code))

    ordered_codes = []
    seen_codes = set()
    for _, store_code in sorted(found_matches, key=lambda item: item[0]):
        if store_code not in seen_codes:
            seen_codes.add(store_code)
            ordered_codes.append(store_code)
    return ordered_codes


def get_dropbox_icg_orders_folder():
    config = get_dropbox_ventas_config()
    base_folder = (config.get("folder") or "/data").rstrip("/") or "/data"
    subfolder = (os.getenv("DROPBOX_VENTAS_PEDIDOS_ICG_SUBFOLDER") or "PedidosICG").strip("/")
    if not subfolder:
        return base_folder
    if base_folder.endswith(f"/{subfolder}"):
        return base_folder
    return f"{base_folder}/{subfolder}"


def sanitize_filename_segment(raw_value: Optional[str], fallback: str):
    raw_text = str(raw_value or "").strip()
    sanitized = "".join(character if character.isalnum() or character in {"_", "-", " "} else "_" for character in raw_text)
    sanitized = re.sub(r"\s+", "_", sanitized).strip("_")
    return sanitized or fallback


def get_employee_facturador_candidates_for_store(store_code: Optional[str]):
    store_code = normalize_store_code(store_code)
    if not store_code:
        return []

    def rank_candidate(record: dict):
        cargo = normalize_text_value(record.get("cargo"))
        if "lider de tienda" in cargo or "líder de tienda" in cargo:
            return 0
        if "lider logistica" in cargo or "líder logística" in cargo or "lider de inventario" in cargo or "líder de inventario" in cargo:
            return 1
        if "facturacion" in cargo or "facturación" in cargo:
            return 2
        return 3

    candidates = []
    for record in load_employee_directory():
        if record.get("store_code") != store_code or not record.get("is_facturador"):
            continue
        if not record.get("email") and not record.get("phone_e164"):
            continue
        candidates.append(record)
    return sorted(candidates, key=rank_candidate)


def get_transfer_destination_email_map():
    configured = {}
    raw_json = os.getenv("TRANSFER_DESTINATION_EMAILS_JSON")
    if raw_json:
        try:
            configured = json.loads(raw_json)
        except Exception as exc:
            raise RuntimeError(f"TRANSFER_DESTINATION_EMAILS_JSON inválido: {exc}")
    normalized_map = {**DEFAULT_TRANSFER_DESTINATION_EMAILS}
    for raw_key, raw_value in (configured or {}).items():
        store_code = normalize_store_code(raw_key)
        email = str(raw_value or "").strip().lower()
        if store_code and email:
            normalized_map[store_code] = email
    return normalized_map


def get_transfer_cc_emails():
    raw_csv = os.getenv("TRANSFER_NOTIFICATION_CC_EMAILS")
    if raw_csv:
        return [email.strip().lower() for email in raw_csv.split(",") if email.strip()]
    return list(DEFAULT_TRANSFER_CC_EMAILS)


def build_brand_email_shell(title: str, body_html: str):
    dark = CORPORATE_BRAND["brand_dark"]
    accent = CORPORATE_BRAND["brand_accent"]
    light = CORPORATE_BRAND["brand_light"]
    border = CORPORATE_BRAND["brand_border"]
    return (
        "<div style='margin:0;padding:24px;background:#f3f4f6;font-family:Segoe UI,Arial,sans-serif;color:#111827;'>"
        f"<div style='max-width:760px;margin:0 auto;background:#ffffff;border:1px solid {border};border-radius:18px;overflow:hidden;'>"
        f"<div style='background:{dark};padding:28px 32px;color:#ffffff;'>"
        f"<div style='font-size:12px;letter-spacing:1.2px;text-transform:uppercase;opacity:.8;'>Ferreinox SAS BIC</div>"
        f"<div style='font-size:30px;font-weight:700;margin-top:6px;'>{escape(title)}</div>"
        f"<div style='margin-top:10px;font-size:13px;color:#d1d5db;'>NIT {escape(CORPORATE_BRAND['nit'])} | {escape(CORPORATE_BRAND['address'])}</div>"
        "</div>"
        f"<div style='padding:28px 32px;background:{light};'>{body_html}</div>"
        f"<div style='padding:22px 32px;background:#ffffff;border-top:1px solid {border};font-size:12px;color:#6b7280;'>"
        f"<strong style='color:#111827;'>{escape(CORPORATE_BRAND['company_name'])}</strong><br>"
        f"Sitio web: <a href='{escape(CORPORATE_BRAND['website'])}' style='color:{accent};'>{escape(CORPORATE_BRAND['website'])}</a><br>"
        f"Correo: {escape(CORPORATE_BRAND['service_email'])} | Tel: {escape(CORPORATE_BRAND['phone_landline'])} | Cel: {escape(CORPORATE_BRAND['phone_mobile'])}"
        "</div></div></div>"
    )


def build_transfer_request_excel_bytes(request_rows: list[dict]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Traslado"
    sheet.append(["Referencia", "Descripción del producto", "Cantidad"])
    for row in request_rows:
        sheet.append([
            row.get("reference") or row.get("referencia") or "",
            row.get("description") or row.get("descripcion") or "",
            row.get("suggested_qty") or row.get("quantity_requested") or "",
        ])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def notify_transfer_requests_by_email(request_rows: list[dict], requested_by_user: dict, notes: Optional[str] = None):
    if not request_rows:
        return {"sent": [], "errors": []}

    grouped_rows = {}
    for row in request_rows:
        destination_store_code = normalize_store_code(row.get("destination_store_code"))
        if not destination_store_code:
            continue
        grouped_rows.setdefault(destination_store_code, []).append(row)

    results = {"sent": [], "errors": []}
    cc_emails = get_transfer_cc_emails()
    requested_by_name = requested_by_user.get("full_name") or "Colaborador Ferreinox"
    requested_by_metadata = dict(requested_by_user.get("metadata") or {})
    requested_by_sede = requested_by_metadata.get("sede") or requested_by_metadata.get("store_code") or "Sede origen"
    requested_by_cargo = requested_by_metadata.get("cargo") or requested_by_user.get("role") or "Perfil interno"
    email_map = get_transfer_destination_email_map()

    for destination_store_code, rows in grouped_rows.items():
        to_email = email_map.get(destination_store_code)
        if not to_email:
            results["errors"].append(f"No existe correo configurado para la sede destino {destination_store_code}.")
            continue

        destination_label = rows[0].get("destination_store_name") or STORE_CODE_LABELS.get(destination_store_code) or destination_store_code
        origin_label = rows[0].get("origin_store_name") or requested_by_sede
        attachment_name = f"traslado_{sanitize_filename_segment(origin_label, 'Origen')}_{sanitize_filename_segment(destination_label, 'Destino')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        attachment_bytes = build_transfer_request_excel_bytes(rows)
        body_html = (
            "<p style='margin:0 0 14px 0;font-size:15px;'>Se generó una nueva solicitud interna de traslado desde WhatsApp.</p>"
            f"<div style='background:#ffffff;border:1px solid {CORPORATE_BRAND['brand_border']};border-radius:14px;padding:18px 20px;margin-bottom:18px;'>"
            f"<p style='margin:0 0 8px 0;'><strong>Solicitante:</strong> {escape(requested_by_name)}</p>"
            f"<p style='margin:0 0 8px 0;'><strong>Cargo:</strong> {escape(str(requested_by_cargo))}</p>"
            f"<p style='margin:0 0 8px 0;'><strong>Sede origen:</strong> {escape(str(origin_label))}</p>"
            f"<p style='margin:0 0 8px 0;'><strong>Sede destino:</strong> {escape(str(destination_label))}</p>"
            f"<p style='margin:0;'><strong>Referencias solicitadas:</strong> {len(rows)}</p>"
            "</div>"
            "<table style='width:100%;border-collapse:collapse;background:#ffffff;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;'>"
            "<thead><tr style='background:#111827;color:#ffffff;'>"
            "<th style='padding:12px;text-align:left;'>Referencia</th>"
            "<th style='padding:12px;text-align:left;'>Descripción</th>"
            "<th style='padding:12px;text-align:center;'>Cantidad</th>"
            "</tr></thead><tbody>"
            + "".join(
                f"<tr><td style='padding:10px 12px;border-top:1px solid #e5e7eb;'>{escape(str(row.get('reference') or row.get('referencia') or ''))}</td>"
                f"<td style='padding:10px 12px;border-top:1px solid #e5e7eb;'>{escape(str(row.get('description') or row.get('descripcion') or ''))}</td>"
                f"<td style='padding:10px 12px;border-top:1px solid #e5e7eb;text-align:center;'>{escape(str(format_quantity(row.get('suggested_qty') or row.get('quantity_requested') or 0)))}</td></tr>"
                for row in rows
            )
            + "</tbody></table>"
            + (f"<p style='margin-top:18px;'><strong>Observaciones:</strong> {escape(notes)}</p>" if notes else "")
            + "<p style='margin-top:18px;'>Se adjunta archivo Excel de control con el detalle exacto del traslado.</p>"
        )
        html_content = build_brand_email_shell("Solicitud de traslado entre sedes", body_html)
        text_content = (
            f"Solicitud de traslado Ferreinox\n"
            f"Solicitante: {requested_by_name}\n"
            f"Cargo: {requested_by_cargo}\n"
            f"Origen: {origin_label}\n"
            f"Destino: {destination_label}\n"
            f"Referencias: {len(rows)}\n"
            f"Observaciones: {notes or 'Sin observaciones'}"
        )
        try:
            send_sendgrid_email(
                to_email,
                f"Solicitud de traslado | {origin_label} -> {destination_label}",
                html_content,
                text_content,
                cc_emails=cc_emails,
                attachments=[
                    {
                        "content": base64.b64encode(attachment_bytes).decode("ascii"),
                        "filename": attachment_name,
                        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "disposition": "attachment",
                    }
                ],
            )
            results["sent"].append(
                {
                    "destination_store_code": destination_store_code,
                    "destination_store_name": destination_label,
                    "to_email": to_email,
                    "cc_emails": cc_emails,
                    "attachment": attachment_name,
                    "count": len(rows),
                }
            )
        except Exception as exc:
            results["errors"].append(f"{destination_label}: {exc}")

    return results


def notify_procurement_request_by_email(shortages: list[dict], requested_by_user: dict, notes: Optional[str] = None):
    if not shortages:
        return {"sent": False, "errors": []}

    to_email = "compras@ferreinox.co"
    requested_by_name = requested_by_user.get("full_name") or "Colaborador Ferreinox"
    requested_by_metadata = dict(requested_by_user.get("metadata") or {})
    requested_by_sede = requested_by_metadata.get("sede") or requested_by_metadata.get("store_code") or "Sede origen"
    destination_label = shortages[0].get("destination_store_name") or "Sede destino"
    attachment_name = f"abastecimiento_{sanitize_filename_segment(destination_label, 'Destino')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    attachment_bytes = build_transfer_request_excel_bytes(shortages)
    body_html = (
        "<p style='margin:0 0 14px 0;font-size:15px;'>Se detectaron faltantes sin origen interno disponible para cumplir pedidos pendientes.</p>"
        f"<div style='background:#ffffff;border:1px solid {CORPORATE_BRAND['brand_border']};border-radius:14px;padding:18px 20px;margin-bottom:18px;'>"
        f"<p style='margin:0 0 8px 0;'><strong>Solicitante:</strong> {escape(requested_by_name)}</p>"
        f"<p style='margin:0 0 8px 0;'><strong>Sede origen:</strong> {escape(str(requested_by_sede))}</p>"
        f"<p style='margin:0 0 8px 0;'><strong>Sede que necesita abastecimiento:</strong> {escape(str(destination_label))}</p>"
        f"<p style='margin:0;'><strong>Referencias sin origen útil:</strong> {len(shortages)}</p>"
        "</div>"
        "<table style='width:100%;border-collapse:collapse;background:#ffffff;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;'>"
        "<thead><tr style='background:#111827;color:#ffffff;'>"
        "<th style='padding:12px;text-align:left;'>Referencia</th>"
        "<th style='padding:12px;text-align:left;'>Descripción</th>"
        "<th style='padding:12px;text-align:center;'>Faltante</th>"
        "</tr></thead><tbody>"
        + "".join(
            f"<tr><td style='padding:10px 12px;border-top:1px solid #e5e7eb;'>{escape(str(row.get('reference') or ''))}</td>"
            f"<td style='padding:10px 12px;border-top:1px solid #e5e7eb;'>{escape(str(row.get('description') or ''))}</td>"
            f"<td style='padding:10px 12px;border-top:1px solid #e5e7eb;text-align:center;'>{escape(str(format_quantity(row.get('shortage_qty') or row.get('required_qty') or 0)))}</td></tr>"
            for row in shortages
        )
        + "</tbody></table>"
        + (f"<p style='margin-top:18px;'><strong>Observaciones:</strong> {escape(notes)}</p>" if notes else "")
        + "<p style='margin-top:18px;'>Se adjunta Excel con el detalle para abastecimiento o compra.</p>"
    )
    html_content = build_brand_email_shell("Solicitud de abastecimiento / compras", body_html)
    text_content = (
        f"Solicitud de abastecimiento Ferreinox\n"
        f"Solicitante: {requested_by_name}\n"
        f"Sede origen: {requested_by_sede}\n"
        f"Sede destino: {destination_label}\n"
        f"Referencias: {len(shortages)}\n"
        f"Observaciones: {notes or 'Sin observaciones'}"
    )
    try:
        send_sendgrid_email(
            to_email,
            f"Abastecimiento requerido | {destination_label}",
            html_content,
            text_content,
            cc_emails=get_transfer_cc_emails(),
            attachments=[
                {
                    "content": base64.b64encode(attachment_bytes).decode("ascii"),
                    "filename": attachment_name,
                    "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "disposition": "attachment",
                }
            ],
        )
        return {"sent": True, "to_email": to_email, "attachment": attachment_name, "errors": []}
    except Exception as exc:
        return {"sent": False, "errors": [str(exc)]}


def get_facturador_routing_config():
    configured_map = None
    raw_json = os.getenv("FACTURADOR_ROUTING_JSON")
    if raw_json:
        try:
            configured_map = json.loads(raw_json)
        except json.JSONDecodeError:
            configured_map = None
    if configured_map is None:
        secrets = load_local_secrets()
        configured_map = secrets.get("facturadores") or secrets.get("facturador_routing") or {}

    sendgrid_config = get_sendgrid_config() or {}
    fallback_email = sendgrid_config.get("ventas_to_email") or sendgrid_config.get("from_email")
    normalized_map = {}
    source_map = configured_map or DEFAULT_FACTURADOR_ROUTING
    for raw_key, raw_entry in source_map.items():
        store_code = normalize_store_code(raw_key)
        if not store_code:
            continue
        entry = raw_entry or {}
        normalized_map[store_code] = {
            "name": entry.get("name") or entry.get("nombre") or DEFAULT_FACTURADOR_ROUTING.get(store_code, {}).get("name") or "Facturación",
            "email": (entry.get("email") or entry.get("correo") or fallback_email or "").strip() or None,
            "phone": normalize_phone_e164(entry.get("phone") or entry.get("telefono") or entry.get("whatsapp") or DEFAULT_FACTURADOR_ROUTING.get(store_code, {}).get("phone")),
        }
    if fallback_email:
        for store_code, fallback_entry in DEFAULT_FACTURADOR_ROUTING.items():
            normalized_map.setdefault(
                store_code,
                {
                    "name": fallback_entry.get("name") or "Facturación",
                    "email": fallback_email,
                    "phone": normalize_phone_e164(fallback_entry.get("phone")),
                },
            )
    return normalized_map


def get_facturador_route_for_store(store_code: Optional[str]):
    store_code = normalize_store_code(store_code)
    if not store_code:
        return None
    employee_candidates = get_employee_facturador_candidates_for_store(store_code)
    if employee_candidates:
        top_candidate = employee_candidates[0]
        return {
            "name": top_candidate.get("full_name") or top_candidate.get("cargo") or "Facturación",
            "email": top_candidate.get("email"),
            "phone": top_candidate.get("phone_e164"),
            "source": "datos_empleados.xlsx",
            "cargo": top_candidate.get("cargo"),
            "sede": top_candidate.get("sede"),
        }
    return get_facturador_routing_config().get(store_code)


def upload_bytes_to_dropbox(content_bytes: bytes, dropbox_path: str):
    dbx = get_dropbox_ventas_client()
    normalized_path = "/" + str(dropbox_path or "").lstrip("/")
    metadata = dbx.files_upload(content_bytes, normalized_path, mode=dropbox.files.WriteMode("overwrite"))
    return metadata.path_display or normalized_path


def build_icg_excel_rows_from_draft(commercial_draft: dict):
    rows = []
    for item in commercial_draft.get("items") or []:
        if item.get("status") != "matched":
            continue
        matched_product = item.get("matched_product") or {}
        product_request = item.get("product_request") or {}
        referencia = matched_product.get("referencia") or matched_product.get("codigo_articulo") or matched_product.get("producto_codigo")
        if not referencia:
            continue
        rows.append(
            {
                "REFERENCIA": referencia,
                "CANTIDAD": parse_numeric_value(product_request.get("requested_quantity")) or 1,
                "PRECIO": parse_numeric_value(matched_product.get("precio") or matched_product.get("precio_venta") or item.get("precio_unitario")) or 0,
                "DESCUENTO": parse_numeric_value(matched_product.get("descuento_pct") or item.get("discount_pct")) or 0,
                "DESCRIPCION": matched_product.get("descripcion") or matched_product.get("nombre_articulo") or item.get("original_text") or referencia,
            }
        )
    return rows


def generate_icg_order_excel_bytes(commercial_draft: dict):
    excel_rows = build_icg_excel_rows_from_draft(commercial_draft)
    if not excel_rows:
        raise RuntimeError("No hay líneas confirmadas para exportar a ICG.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PedidoICG"
    worksheet.append(["REFERENCIA", "CANTIDAD", "PRECIO", "DESCUENTO"])
    for row in excel_rows:
        worksheet.append([row["REFERENCIA"], row["CANTIDAD"], row["PRECIO"], row["DESCUENTO"]])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), excel_rows


def build_icg_order_filename(cliente_label: Optional[str], store_code: Optional[str], order_id: int):
    today_label = datetime.now().strftime("%Y-%m-%d")
    cliente_safe = sanitize_filename_segment(cliente_label, "SinCliente")
    tienda_safe = sanitize_filename_segment(get_store_short_label(store_code) or STORE_CODE_LABELS.get(store_code) or store_code or "SinTienda", "SinTienda")
    return f"pedido_{cliente_safe}_{today_label}_{tienda_safe}_{order_id}.xlsx"


def mark_agent_order_status(order_id: int, status: str, metadata_update: Optional[dict] = None, numero_externo: Optional[str] = None):
    if not order_id:
        return
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE public.agent_order
                SET estado = :status,
                    numero_externo = COALESCE(:numero_externo, numero_externo),
                    submitted_at = COALESCE(submitted_at, now()),
                    metadata = COALESCE(metadata, '{}'::jsonb) || CAST(:metadata_update AS jsonb),
                    updated_at = now()
                WHERE id = :order_id
                """
            ),
            {
                "order_id": order_id,
                "status": status,
                "numero_externo": numero_externo,
                "metadata_update": safe_json_dumps(metadata_update or {}),
            },
        )


def upsert_order_dispatch_record(
    order_id: int,
    conversation_id: int,
    contact_id: Optional[int],
    cliente_id: Optional[int],
    destination_store_code: Optional[str],
    destination_store_name: Optional[str],
    export_filename: str,
    dropbox_folder: str,
    dropbox_path: str,
    facturador_route: Optional[dict],
    exported_by_user_id: Optional[int],
    observations: Optional[str],
    metadata: Optional[dict] = None,
    status: str = "pendiente",
    email_sent: bool = False,
    whatsapp_sent: bool = False,
):
    engine = get_db_engine()
    with engine.begin() as connection:
        dispatch_id = connection.execute(
            text(
                """
                INSERT INTO public.agent_order_dispatch (
                    order_id,
                    conversation_id,
                    contacto_id,
                    cliente_id,
                    destination_store_code,
                    destination_store_name,
                    exported_by_user_id,
                    facturador_name,
                    facturador_email,
                    facturador_phone,
                    export_filename,
                    dropbox_folder,
                    dropbox_path,
                    status,
                    observations,
                    metadata,
                    exported_at,
                    notified_email_at,
                    notified_whatsapp_at,
                    created_at,
                    updated_at
                )
                VALUES (
                    :order_id,
                    :conversation_id,
                    :contact_id,
                    :cliente_id,
                    :destination_store_code,
                    :destination_store_name,
                    :exported_by_user_id,
                    :facturador_name,
                    :facturador_email,
                    :facturador_phone,
                    :export_filename,
                    :dropbox_folder,
                    :dropbox_path,
                    :status,
                    :observations,
                    CAST(:metadata AS jsonb),
                    now(),
                    CASE WHEN :email_sent THEN now() ELSE NULL END,
                    CASE WHEN :whatsapp_sent THEN now() ELSE NULL END,
                    now(),
                    now()
                )
                ON CONFLICT (order_id)
                DO UPDATE SET
                    conversation_id = EXCLUDED.conversation_id,
                    contacto_id = EXCLUDED.contacto_id,
                    cliente_id = EXCLUDED.cliente_id,
                    destination_store_code = EXCLUDED.destination_store_code,
                    destination_store_name = EXCLUDED.destination_store_name,
                    exported_by_user_id = EXCLUDED.exported_by_user_id,
                    facturador_name = EXCLUDED.facturador_name,
                    facturador_email = EXCLUDED.facturador_email,
                    facturador_phone = EXCLUDED.facturador_phone,
                    export_filename = EXCLUDED.export_filename,
                    dropbox_folder = EXCLUDED.dropbox_folder,
                    dropbox_path = EXCLUDED.dropbox_path,
                    status = EXCLUDED.status,
                    observations = EXCLUDED.observations,
                    metadata = COALESCE(public.agent_order_dispatch.metadata, '{}'::jsonb) || EXCLUDED.metadata,
                    exported_at = now(),
                    notified_email_at = CASE WHEN :email_sent THEN now() ELSE public.agent_order_dispatch.notified_email_at END,
                    notified_whatsapp_at = CASE WHEN :whatsapp_sent THEN now() ELSE public.agent_order_dispatch.notified_whatsapp_at END,
                    updated_at = now()
                RETURNING id
                """
            ),
            {
                "order_id": order_id,
                "conversation_id": conversation_id,
                "contact_id": contact_id,
                "cliente_id": cliente_id,
                "destination_store_code": destination_store_code,
                "destination_store_name": destination_store_name,
                "exported_by_user_id": exported_by_user_id,
                "facturador_name": (facturador_route or {}).get("name"),
                "facturador_email": (facturador_route or {}).get("email"),
                "facturador_phone": (facturador_route or {}).get("phone"),
                "export_filename": export_filename,
                "dropbox_folder": dropbox_folder,
                "dropbox_path": dropbox_path,
                "status": status,
                "observations": observations,
                "metadata": safe_json_dumps(metadata or {}),
                "email_sent": email_sent,
                "whatsapp_sent": whatsapp_sent,
            },
        ).scalar_one()
    return dispatch_id


def notify_facturador_about_order(
    order_id: int,
    cliente_label: str,
    store_code: Optional[str],
    file_name: str,
    dropbox_path: str,
    facturador_route: Optional[dict],
    observations: Optional[str],
):
    route = facturador_route or {}
    store_label = get_store_short_label(store_code) or STORE_CODE_LABELS.get(store_code) or "sede sin definir"
    facturador_name = route.get("name") or "equipo de facturación"
    subject = f"Ferreinox | Pedido ICG listo {file_name}"
    text_content = (
        f"Hola {facturador_name}.\n\n"
        f"Ya quedó exportado un pedido para {store_label}.\n"
        f"Cliente: {cliente_label}\n"
        f"Pedido interno: {order_id}\n"
        f"Archivo: {file_name}\n"
        f"Ruta Dropbox: {dropbox_path}\n"
        f"Observación: {observations or 'Sin observación'}\n\n"
        "Por favor valida la carpeta compartida y continúa la facturación."
    )
    html_content = (
        "<div style='font-family:Segoe UI,Arial,sans-serif;background:#f4f6f8;padding:24px;color:#111827'>"
        "<div style='max-width:760px;margin:0 auto;background:#ffffff;border:1px solid #e5e7eb;border-radius:18px;padding:28px'>"
        f"<h1 style='margin-top:0'>Pedido listo para facturación</h1>"
        f"<p><strong>Facturador:</strong> {escape(facturador_name)}</p>"
        f"<p><strong>Tienda destino:</strong> {escape(store_label)}</p>"
        f"<p><strong>Cliente:</strong> {escape(cliente_label)}</p>"
        f"<p><strong>Pedido interno:</strong> {order_id}</p>"
        f"<p><strong>Archivo ICG:</strong> {escape(file_name)}</p>"
        f"<p><strong>Ruta Dropbox:</strong> {escape(dropbox_path)}</p>"
        f"<p><strong>Observación:</strong> {escape(observations or 'Sin observación')}</p>"
        "<p>Revisa la carpeta compartida y continúa el proceso en ICG.</p>"
        "</div></div>"
    )

    email_sent = False
    whatsapp_sent = False
    errors = []
    if route.get("email"):
        try:
            send_sendgrid_email(route["email"], subject, html_content, text_content)
            email_sent = True
        except Exception as exc:
            errors.append(f"email: {exc}")
    if route.get("phone"):
        whatsapp_body = (
            f"Hola {facturador_name}, ya quedó exportado el pedido {order_id} para {store_label}. "
            f"Cliente: {cliente_label}. Archivo: {file_name}. Ruta: {dropbox_path}."
        )
        try:
            send_whatsapp_text_message(route["phone"], whatsapp_body)
            whatsapp_sent = True
        except Exception as exc:
            errors.append(f"whatsapp: {exc}")
    return {
        "email_sent": email_sent,
        "whatsapp_sent": whatsapp_sent,
        "email": route.get("email"),
        "phone": route.get("phone"),
        "errors": errors,
    }


def export_confirmed_order_to_icg(
    order_id: int,
    context: dict,
    commercial_draft: dict,
    cliente_contexto: Optional[dict],
    internal_user: Optional[dict],
):
    store_filters = commercial_draft.get("store_filters") or []
    store_code = normalize_store_code(store_filters[0]) if store_filters else None
    if not store_code:
        raise RuntimeError("El pedido no tiene tienda o ciudad definida para exportar a ICG.")

    file_name = build_icg_order_filename(
        (cliente_contexto or {}).get("nombre_cliente") or context.get("nombre_visible"),
        store_code,
        order_id,
    )
    excel_bytes, excel_rows = generate_icg_order_excel_bytes(commercial_draft)
    dropbox_folder = get_dropbox_icg_orders_folder()
    dropbox_path = upload_bytes_to_dropbox(excel_bytes, f"{dropbox_folder}/{file_name}")
    facturador_route = get_facturador_route_for_store(store_code)
    observations = commercial_draft.get("facturador_notes") or commercial_draft.get("observaciones") or commercial_draft.get("nombre_despacho")
    notification_result = notify_facturador_about_order(
        order_id,
        (cliente_contexto or {}).get("nombre_cliente") or context.get("nombre_visible") or "Cliente Ferreinox",
        store_code,
        file_name,
        dropbox_path,
        facturador_route,
        observations,
    )
    dispatch_id = upsert_order_dispatch_record(
        order_id=order_id,
        conversation_id=context["conversation_id"],
        contact_id=context.get("contact_id"),
        cliente_id=context.get("cliente_id"),
        destination_store_code=store_code,
        destination_store_name=STORE_CODE_LABELS.get(store_code),
        export_filename=file_name,
        dropbox_folder=dropbox_folder,
        dropbox_path=dropbox_path,
        facturador_route=facturador_route,
        exported_by_user_id=(internal_user or {}).get("id"),
        observations=observations,
        metadata={"excel_rows": excel_rows, "notificacion": notification_result},
        status="pendiente",
        email_sent=notification_result.get("email_sent", False),
        whatsapp_sent=notification_result.get("whatsapp_sent", False),
    )
    mark_agent_order_status(
        order_id,
        "enviado_erp",
        metadata_update={
            "dispatch_id": dispatch_id,
            "dropbox_path": dropbox_path,
            "facturador": facturador_route,
            "icg_filename": file_name,
        },
        numero_externo=f"PED-{order_id}",
    )
    return {
        "dispatch_id": dispatch_id,
        "dropbox_path": dropbox_path,
        "dropbox_folder": dropbox_folder,
        "file_name": file_name,
        "facturador": facturador_route,
        "notification": notification_result,
    }


def list_technical_document_entries(force_refresh: bool = False):
    cache_age = time.time() - float(TECHNICAL_DOC_CACHE.get("loaded_at") or 0)
    if not force_refresh and TECHNICAL_DOC_CACHE.get("entries") and cache_age < TECHNICAL_DOC_CACHE_TTL_SECONDS:
        return TECHNICAL_DOC_CACHE["entries"]

    dbx = get_dropbox_ventas_client()
    entries = []
    result = dbx.files_list_folder(TECHNICAL_DOC_FOLDER, recursive=True)
    while True:
        entries.extend(
            entry for entry in result.entries
            if isinstance(entry, dropbox.files.FileMetadata) and entry.name.lower().endswith(".pdf")
        )
        if not result.has_more:
            break
        result = dbx.files_list_folder_continue(result.cursor)

    TECHNICAL_DOC_CACHE["loaded_at"] = time.time()
    TECHNICAL_DOC_CACHE["entries"] = entries
    return entries


def is_technical_document_message(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    return any(
        phrase in normalized
        for phrase in [
            "ficha tecnica",
            "ficha técnicas",
            "ficha tecnica",
            "hoja de seguridad",
            "hoja seguridad",
            "fds",
            "msds",
        ]
    )


def extract_technical_document_request(text_value: Optional[str], product_request: Optional[dict] = None, conversation_context: Optional[dict] = None):
    normalized = normalize_text_value(text_value)
    request = product_request or extract_product_request(text_value)
    previous_document_request = (conversation_context or {}).get("last_document_request") or {}
    previous_request = (conversation_context or {}).get("last_product_request") or {}

    def collect_terms(source_terms: list[str]):
        collected_terms = []
        for term in source_terms:
            normalized_term = normalize_text_value(term)
            if (
                normalized_term
                and normalized_term not in TECHNICAL_DOC_STOPWORDS
                and normalized_term not in PRODUCT_STOPWORDS
                and not is_store_alias_term(normalized_term)
                and len(normalized_term) >= 3
                and normalized_term not in collected_terms
            ):
                collected_terms.append(normalized_term)
        return collected_terms

    current_terms = collect_terms(request.get("core_terms") or [])
    previous_document_terms = collect_terms(previous_document_request.get("terms") or [])
    previous_product_terms = collect_terms(previous_request.get("core_terms") or [])

    if current_terms:
        terms = current_terms
    elif previous_document_terms:
        terms = previous_document_terms
    else:
        terms = previous_product_terms

    wants_safety_sheet = any(keyword in normalized for keyword in ["hoja de seguridad", "hoja seguridad", "seguridad", "fds", "msds"])
    wants_technical_sheet = any(keyword in normalized for keyword in ["ficha tecnica", "ficha técnica", "ficha", "tecnica", "técnica"])

    return {
        "query": text_value or "",
        "terms": terms[:8],
        "wants_safety_sheet": wants_safety_sheet,
        "wants_technical_sheet": wants_technical_sheet or not wants_safety_sheet,
    }


def search_technical_documents(document_request: dict):
    terms = document_request.get("terms") or []
    if not terms:
        return []

    ranked_documents = []
    for entry in list_technical_document_entries():
        path_value = normalize_text_value(entry.path_lower or entry.name)
        name_value = normalize_text_value(entry.name)
        exact_hits = sum(1 for term in terms if term in path_value)
        if exact_hits == 0 and not any(sequence_similarity(term, name_value) >= 0.74 for term in terms):
            continue

        safety_score = 0
        if document_request.get("wants_safety_sheet"):
            safety_score = 1 if any(token in path_value for token in ["hoja", "seguridad", "fds", "msds"]) else 0
        technical_score = 0
        if document_request.get("wants_technical_sheet"):
            technical_score = 1 if not any(token in path_value for token in ["fds", "msds"]) else 0

        ranked_documents.append(
            {
                "name": entry.name,
                "path_lower": entry.path_lower,
                "exact_hits": exact_hits,
                "safety_score": safety_score,
                "technical_score": technical_score,
                "fuzzy_score": round(max(sequence_similarity(term, name_value) for term in terms), 4),
            }
        )

    ranked_documents.sort(
        key=lambda item: (
            item.get("safety_score") or 0,
            item.get("technical_score") or 0,
            item.get("exact_hits") or 0,
            item.get("fuzzy_score") or 0,
            len(item.get("name") or ""),
        ),
        reverse=True,
    )
    return ranked_documents[:6]


def resolve_technical_document_choice(text_value: Optional[str], document_options: list[dict]):
    normalized = normalize_text_value(text_value)
    if not normalized or not document_options:
        return None

    ordinal_map = {
        "1": 0,
        "uno": 0,
        "primera": 0,
        "primer": 0,
        "primero": 0,
        "2": 1,
        "dos": 1,
        "segunda": 1,
        "segundo": 1,
        "3": 2,
        "tres": 2,
        "tercera": 2,
        "tercero": 2,
        "4": 3,
        "cuatro": 3,
        "cuarta": 3,
        "cuarto": 3,
    }
    if normalized in ordinal_map and ordinal_map[normalized] < len(document_options):
        return document_options[ordinal_map[normalized]]

    for ordinal_text, option_index in ordinal_map.items():
        if option_index >= len(document_options):
            continue
        if re.fullmatch(rf"(?:la|el|opcion|archivo)?\s*{re.escape(ordinal_text)}", normalized):
            return document_options[option_index]

    for option in document_options:
        option_name = normalize_text_value(option.get("name"))
        if option_name and (option_name in normalized or normalized in option_name):
            return option
    return None


def get_dropbox_temporary_link(file_path: str):
    dbx = get_dropbox_ventas_client()
    return dbx.files_get_temporary_link(file_path).link


def normalize_text_value(text_value: Optional[str]):
    if not text_value:
        return ""
    normalized = unicodedata.normalize("NFKD", text_value)
    normalized = "".join(character for character in normalized if not unicodedata.combining(character))
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9./+-]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def normalize_reference_value(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return ""
    return re.sub(r"[^a-z0-9]+", "", normalized)


def parse_numeric_value(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float)):
        return float(raw_value)

    cleaned = str(raw_value).strip()
    if not cleaned:
        return None

    cleaned = re.sub(r"[^0-9,.-]", "", cleaned)
    if not cleaned:
        return None

    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        return float(cleaned)
    except ValueError:
        return None


def sequence_similarity(left_value: Optional[str], right_value: Optional[str]):
    left_normalized = normalize_text_value(left_value)
    right_normalized = normalize_text_value(right_value)
    if not left_normalized or not right_normalized:
        return 0.0
    return SequenceMatcher(None, left_normalized, right_normalized).ratio()


# ══════════════════════════════════════════════════════════════════════════════
# SMART MATCHER ENGINE — Pattern-based product search (replaces manual aliases)
# ══════════════════════════════════════════════════════════════════════════════

# ── Spanish Phonetic Key ──────────────────────────────────────────────────────
# Adapted Soundex for Spanish: normalizes common misspellings/phonetic variants
# so "rodiyo"≈"rodillo", "liha"≈"lija", "brosha"≈"brocha", "barnís"≈"barniz"
_SPANISH_PHONETIC_MAP: list[tuple[str, str]] = [
    # Vowels collapse (keep first, absorb rest)
    # Handled specially below
    # Consonant equivalences in Latin American Spanish
    (r"ll", "y"),       # ll → y (yeísmo)
    (r"rr", "r"),       # rr → r
    (r"cc", "c"),       # cc → c
    (r"ss", "s"),       # ss → s
    (r"nn", "n"),       # nn → n
    (r"qu", "k"),       # qu → k
    (r"ch", "X"),       # ch → X (placeholder)
    (r"sh", "X"),       # sh → X (brosha→broXa)
    (r"ck", "k"),       # ck → k
    (r"ph", "f"),       # ph → f
    (r"v", "b"),        # v → b (labial merge)
    (r"z", "s"),        # z → s (seseo)
    (r"ce", "se"),      # ce → se
    (r"ci", "si"),      # ci → si
    (r"ge", "je"),      # ge → je
    (r"gi", "ji"),      # gi → ji
    (r"gü", "w"),       # gü → w
    (r"gu(?=[ei])", "g"),  # gue/gui → ge/gi
    (r"h", ""),         # h is silent (liha→lia→lja)
    (r"j", "j"),        # j stays j
    (r"x", "ks"),       # x → ks
    (r"w", "u"),        # w → u
    (r"ñ", "ny"),       # ñ → ny
    (r"y$", "i"),       # final y → i
]

def spanish_phonetic_key(text_value: Optional[str]) -> str:
    """Generate a Spanish phonetic key for fuzzy matching.
    Maps common misspellings/phonetic variants to the same key."""
    normalized = normalize_text_value(text_value)
    if not normalized:
        return ""
    # Apply phonetic transformations
    result = normalized
    for pattern, replacement in _SPANISH_PHONETIC_MAP:
        result = re.sub(pattern, replacement, result)
    # Collapse consecutive duplicate consonants
    result = re.sub(r"([bcdfgjklmnpqrstvxyz])\1+", r"\1", result)
    # Collapse consecutive vowels to single vowel
    result = re.sub(r"([aeiou])\1+", r"\1", result)
    return result


def spanish_phonetic_similarity(query: str, candidate: str) -> float:
    """Compare two strings using Spanish phonetic keys + character overlap."""
    key_q = spanish_phonetic_key(query)
    key_c = spanish_phonetic_key(candidate)
    if not key_q or not key_c:
        return 0.0
    return SequenceMatcher(None, key_q, key_c).ratio()


# ── Smart Brand Anchor ────────────────────────────────────────────────────────
# Leader brands that should anchor search when detected in query
_LEADER_BRANDS_PRIORITY: list[tuple[str, list[str]]] = [
    ("pintuco", ["pintuco", "viniltex", "koraza", "pintulux", "domestico", "aerocolor",
                 "pintucoat", "corrotec", "pintulac", "intervinil", "pintuco fill",
                 "aquablock", "sellamur", "siliconite", "pintura canchas", "barnex",
                 "wood stain", "estucomast", "epoxipoliamida", "vinilux", "vinil max",
                 "vinil plus", "icolatex"]),
    ("abracol", ["abracol", "disco abracol", "lija abracol"]),
    ("goya", ["goya", "brocha goya", "rodillo goya"]),
    ("yale", ["yale", "cerradura yale", "candado yale"]),
    ("smith", ["smith", "cinta smith"]),
    ("international", ["international", "interseal", "interthane", "intergard", "interchar", "interfine"]),
    ("sika", ["sika", "sikaflex", "sikaguard", "sikalastic"]),
    ("norton", ["norton", "disco norton"]),
    ("afix", ["afix"]),
    ("artecola", ["artecola"]),
    ("montana", ["montana", "spray montana"]),
]

def detect_brand_anchor(query_text: str) -> Optional[str]:
    """Detect if a leader brand is mentioned in the query and return its canonical name."""
    normalized = normalize_text_value(query_text)
    if not normalized:
        return None
    for brand_name, triggers in _LEADER_BRANDS_PRIORITY:
        for trigger in triggers:
            if trigger in normalized:
                return brand_name
    return None


# ── Smart Scoring System (0.0 – 1.0) ─────────────────────────────────────────
_KIT_PROMO_KEYWORDS = frozenset(["KIT ", "PAGUE ", "PAGU ", "NO INV", "GRATIS", "GTIS", "LLEVE", "OBSEQUIO", "REGALO"])
_GENERIC_BRAND_CODES = frozenset(["0", "", "NaN"])

def smart_score_product(
    candidate: dict,
    query_text: str,
    product_request: dict,
    rotation_cache: Optional[dict] = None,
) -> float:
    """Unified scoring: rotation(0.4) + text_match(0.3) + stock(0.2) + penalties.
    Returns a float in the range [-1.5, 1.0]."""
    score = 0.0
    desc = candidate.get("descripcion") or candidate.get("nombre_articulo") or ""
    desc_upper = desc.upper()
    candidate_code = str(candidate.get("producto_codigo") or candidate.get("referencia") or candidate.get("codigo_articulo") or "")
    candidate_brand = str(candidate.get("marca") or candidate.get("marca_producto") or "")

    # ── +0.4: Rotation score (historical sales velocity) ──
    if rotation_cache and candidate_code in rotation_cache:
        score += 0.4 * rotation_cache[candidate_code]

    # ── +0.3: Text similarity (phonetic + trigram-style character overlap) ──
    normalized_query = normalize_text_value(query_text)
    # Build a rich candidate text from ALL available fields
    candidate_fields = [
        desc,
        candidate.get("descripcion_ebs") or "",
        candidate.get("familia_clasificacion") or "",
        candidate.get("cat_producto") or "",
        candidate.get("marca_clasificacion") or "",
        candidate.get("aplicacion_clasificacion") or "",
    ]
    rich_candidate_text = " ".join(f for f in candidate_fields if f and f != "NaN")
    normalized_candidate = normalize_text_value(rich_candidate_text)

    if normalized_query and normalized_candidate:
        # Character-level similarity (SequenceMatcher)
        char_sim = SequenceMatcher(None, normalized_query, normalized_candidate).ratio()
        # Phonetic similarity
        phonetic_sim = spanish_phonetic_similarity(normalized_query, rich_candidate_text)
        # Term overlap: what fraction of query terms appear in candidate?
        query_tokens = set(normalized_query.split())
        candidate_tokens = set(normalized_candidate.split())
        if query_tokens:
            term_overlap = len(query_tokens & candidate_tokens) / len(query_tokens)
        else:
            term_overlap = 0.0
        # Weighted blend: term overlap is most important, then phonetic, then raw char sim
        text_score = 0.5 * term_overlap + 0.3 * phonetic_sim + 0.2 * char_sim
        score += 0.3 * text_score

    # ── +0.2: Stock availability ──
    stock = parse_numeric_value(candidate.get("stock_total")) or 0
    if stock > 0:
        score += 0.2

    # ── -0.5: Kit/Promo penalty (unless user explicitly asks for a kit) ──
    is_kit_requested = any(kw in (normalized_query or "") for kw in ["kit", "combo", "promo", "pague"])
    if not is_kit_requested and any(kw in desc_upper for kw in _KIT_PROMO_KEYWORDS):
        score -= 0.5

    # ── -1.0: Generic/no-brand penalty when a branded version exists ──
    if candidate_brand in _GENERIC_BRAND_CODES:
        brand_anchor = detect_brand_anchor(query_text)
        if brand_anchor:
            score -= 1.0

    # ── Brand anchor bonus: +0.1 if matches detected brand ──
    brand_anchor = detect_brand_anchor(query_text)
    if brand_anchor:
        brand_fields = normalize_text_value(
            f"{candidate.get('marca_clasificacion') or ''} {desc} {candidate.get('familia_clasificacion') or ''}"
        )
        if brand_anchor in brand_fields or any(
            alias in brand_fields
            for alias in (BRAND_ALIASES.get(brand_anchor) or [])
        ):
            score += 0.1

    return round(score, 4)


# ── Global rotation cache with TTL ────────────────────────────────────────────
_rotation_cache_data: dict = {}
_rotation_cache_ts: float = 0.0
_ROTATION_CACHE_TTL = 300  # 5 minutes


def fetch_rotation_cache(connection) -> dict:
    """Load the rotation scores with 5-minute in-memory cache."""
    global _rotation_cache_data, _rotation_cache_ts
    now = time.time()
    if _rotation_cache_data and (now - _rotation_cache_ts) < _ROTATION_CACHE_TTL:
        return _rotation_cache_data
    try:
        rows = connection.execute(
            text("SELECT producto_codigo, rotation_score FROM mv_product_rotation")
        ).mappings().all()
        _rotation_cache_data = {str(row["producto_codigo"]): float(row["rotation_score"]) for row in rows}
        _rotation_cache_ts = now
        return _rotation_cache_data
        return {str(row["producto_codigo"]): float(row["rotation_score"]) for row in rows}
    except Exception:
        return {}


# ── Fuzzy Multi-Column Search (Trigram + Phonetic) ────────────────────────────
def fetch_smart_product_rows(
    connection,
    query_text: str,
    query_terms: list[str],
    product_request: dict,
    store_filters: list[str],
    limit: int = 30,
) -> list[dict]:
    """Multi-column fuzzy search using pg_trgm similarity + ILIKE.
    Searches across: descripcion, descripcion_ebs, referencia, familia_clasificacion,
    cat_producto, marca_clasificacion, aplicacion_clasificacion, search_blob."""
    if not query_terms:
        return []

    params: dict = {}
    ilike_filters: list[str] = []
    score_parts: list[str] = []

    # Standard ILIKE term matching + bidirectional abbreviation variants combined
    # Each term + its variants produce a SINGLE 1-point score (not additive)
    var_idx = 200
    for idx, term in enumerate(query_terms[:6]):
        params[f"pat_{idx}"] = f"%{term}%"
        compact = normalize_reference_value(term)
        params[f"cpt_{idx}"] = f"%{compact}%"
        ilike_filters.append(f"search_blob ILIKE :pat_{idx}")
        if compact:
            ilike_filters.append(f"search_compact LIKE :cpt_{idx}")

        # Collect variant ILIKE conditions for the WHERE clause
        variant_conditions = []
        variants = _SEARCH_TERM_VARIANTS.get(term.lower(), [])
        for variant in variants[:3]:
            vk = f"var_{var_idx}"
            var_idx += 1
            params[vk] = f"%{variant.upper()}%"
            ilike_filters.append(f"search_blob ILIKE :{vk}")
            variant_conditions.append(f"search_blob ILIKE :{vk}")

        # Combined score: 1 point if original OR compact OR any variant matches
        all_conditions = [f"search_blob ILIKE :pat_{idx}"]
        if compact:
            all_conditions.append(f"search_compact LIKE :cpt_{idx}")
        all_conditions.extend(variant_conditions)
        score_parts.append(
            f"CASE WHEN {' OR '.join(all_conditions)} THEN 1 ELSE 0 END"
        )

    # Phonetic expansion: generate phonetic variants and search them too
    phonetic_query = spanish_phonetic_key(query_text)
    if phonetic_query and len(phonetic_query) >= 4:
        # Split into phonetic tokens and search each
        phonetic_tokens = [t for t in phonetic_query.split() if len(t) >= 3]
        for pidx, ptok in enumerate(phonetic_tokens[:4]):
            pk = f"phon_{pidx}"
            params[pk] = f"%{ptok}%"
            ilike_filters.append(f"search_blob ILIKE :{pk}")
            score_parts.append(f"CASE WHEN search_blob ILIKE :{pk} THEN 1 ELSE 0 END")

    # Abbreviation prefix matching (existing approach)
    abbrev_idx = 100
    for i in range(len(query_terms[:6]) - 1):
        concat_compact = normalize_reference_value(query_terms[i]) + normalize_reference_value(query_terms[i + 1])
        if len(concat_compact) < 8:
            continue
        for trim in range(0, min(len(concat_compact) - 6, 5)):
            prefix = concat_compact[:len(concat_compact) - trim]
            pk = f"abr_{abbrev_idx}"
            abbrev_idx += 1
            params[pk] = f"%{prefix}%"
            ilike_filters.append(f"search_compact LIKE :{pk}")
            score_parts.append(f"CASE WHEN search_compact LIKE :{pk} THEN 1 ELSE 0 END")

    # Trigram query kept for potential Python-side use, but NOT computed in SQL (expensive)
    trgm_score = "0"  # Disabled in SQL for speed; smart_score handles fuzzy ranking in Python

    # Numeric pattern → prioritize by referencia/producto_codigo exact match
    numeric_pattern_bonus = []
    for term in query_terms[:3]:
        if re.fullmatch(r"\d{4,}", term):
            npk = f"numex_{len(numeric_pattern_bonus)}"
            params[npk] = term
            numeric_pattern_bonus.append(
                f"CASE WHEN referencia = :{npk} THEN 50 ELSE 0 END"
            )

    all_score_parts = score_parts + numeric_pattern_bonus
    match_score_sql = " + ".join(all_score_parts) if all_score_parts else "0"

    where_clause = f"({' OR '.join(ilike_filters)})"

    if store_filters:
        return _fetch_smart_from_store(connection, where_clause, params, match_score_sql, store_filters, limit)

    rows = connection.execute(
        text(
            f"""
            SELECT p.producto_codigo, p.referencia, p.descripcion, p.marca, p.departamentos, p.stock_total, p.costo_promedio_und, p.stock_por_tienda,
                   p.linea_clasificacion, p.marca_clasificacion, p.familia_clasificacion, p.aplicacion_clasificacion, p.cat_producto, p.descripcion_ebs, p.tipo_articulo,
                   p.nombre_comercial_abracol, p.familia_abracol, p.descripcion_larga_abracol, p.portafolio_abracol,
                   ({match_score_sql}) AS match_score,
                   COALESCE(rot.rotation_score, 0) AS rotation_score
            FROM mv_productos p
            LEFT JOIN mv_product_rotation rot ON rot.producto_codigo = p.producto_codigo
            WHERE {where_clause}
            ORDER BY ({match_score_sql}) DESC, COALESCE(rot.rotation_score, 0) DESC, stock_total DESC NULLS LAST
            LIMIT {int(limit)}
            """
        ),
        params,
    ).mappings().all()
    return [dict(r) for r in rows]


def _fetch_smart_from_store(connection, where_clause, params, match_score_sql, store_filters, limit):
    """Store-filtered variant of smart search."""
    store_sql = []
    for si, sc in enumerate(store_filters):
        params[f"store_{si}"] = sc
        store_sql.append(f"cod_almacen = :store_{si}")

    # Build inner WHERE: use pat_ and var_ ILIKE filters for broader matching
    pat_keys = sorted([k for k in params if k.startswith("pat_") or k.startswith("var_")])
    inner_filters = [f"search_blob ILIKE :{k}" for k in pat_keys]
    if not inner_filters:
        inner_filters = ["TRUE"]
    inner_where = f"({' OR '.join(inner_filters)}) AND ({' OR '.join(store_sql)})"

    rows = connection.execute(
        text(
            f"""
            SELECT referencia, descripcion, marca, departamentos, stock_total, costo_promedio_und, stock_por_tienda,
                   linea_clasificacion, marca_clasificacion, familia_clasificacion, aplicacion_clasificacion, cat_producto, descripcion_ebs, tipo_articulo,
                   ({match_score_sql}) AS match_score,
                   COALESCE(rot.rotation_score, 0) AS rotation_score
            FROM (
                SELECT
                    referencia,
                    descripcion,
                    marca,
                    STRING_AGG(DISTINCT departamento, ', ' ORDER BY departamento) AS departamentos,
                    COALESCE(SUM(stock_disponible), 0) AS stock_total,
                    AVG(costo_promedio_und) AS costo_promedio_und,
                    STRING_AGG(
                        almacen_nombre || ': ' || COALESCE(stock_disponible::text, '0'),
                        '; ' ORDER BY almacen_nombre
                    ) FILTER (WHERE COALESCE(stock_disponible, 0) > 0) AS stock_por_tienda,
                    MAX(search_blob) AS search_blob,
                    public.fn_keep_alnum(MAX(descripcion) || ' ' || MAX(referencia) || ' ' || MAX(marca)) AS search_compact,
                    MAX(referencia_normalizada) AS referencia_normalizada,
                    MAX(linea_clasificacion) AS linea_clasificacion,
                    MAX(marca_clasificacion) AS marca_clasificacion,
                    MAX(familia_clasificacion) AS familia_clasificacion,
                    MAX(aplicacion_clasificacion) AS aplicacion_clasificacion,
                    MAX(cat_producto) AS cat_producto,
                    MAX(descripcion_ebs) AS descripcion_ebs,
                    MAX(tipo_articulo) AS tipo_articulo
                FROM public.vw_inventario_agente
                WHERE {inner_where}
                GROUP BY referencia, descripcion, marca
            ) inventory
            LEFT JOIN mv_product_rotation rot ON rot.producto_codigo = inventory.referencia
            ORDER BY ({match_score_sql}) DESC, COALESCE(rot.rotation_score, 0) DESC, stock_total DESC NULLS LAST
            LIMIT {int(limit)}
            """
        ),
        params,
    ).mappings().all()
    return [dict(r) for r in rows]


def translate_product_to_commercial(description: Optional[str], presentation: Optional[str] = None, brand: Optional[str] = None):
    """Convert raw DB descriptions like 'PQ VINILTEX ADV MAT BLANCO 1501 18.93L' to commercial language."""
    if not description:
        return "producto"
    raw = str(description).strip()
    # Remove common prefixes
    for prefix in ["PQ ", "IQ ", "EQ ", "SQ ", "MEG "]:
        if raw.upper().startswith(prefix):
            raw = raw[len(prefix):].strip()
    # Remove size suffixes like 18.93L, 3.79L, 0.95L, 0.22L
    cleaned = re.sub(r"\s+\d+\.\d+L$", "", raw, flags=re.IGNORECASE)
    # Remove trailing reference codes like " 1501", " 12286", but only at end
    cleaned = re.sub(r"\s+\d{3,6}$", "", cleaned)
    # Clean up double-quoted inches
    cleaned = cleaned.replace('""', '"').replace('"', '"')
    # Title case
    words = cleaned.split()
    titled_words = []
    skip_words = {"BR", "MAT", "ADV", "SAT", "SB", "CRE", "DEEP"}
    for w in words:
        if w.upper() in skip_words:
            continue
        titled_words.append(w.capitalize())
    commercial_name = " ".join(titled_words).strip()
    if not commercial_name:
        commercial_name = raw.title()
    # Add presentation label
    pres_label = ""
    if presentation:
        pres_map = {"cuñete": "en cuñete", "galon": "en galón", "cuarto": "en cuarto"}
        pres_label = pres_map.get(presentation, f"en {presentation}")
    if pres_label:
        commercial_name = f"{commercial_name} {pres_label}"
    return commercial_name


def get_exact_product_description(product_row: Optional[dict]):
    raw_description = (product_row or {}).get("descripcion") or (product_row or {}).get("nombre_articulo") or "producto"
    return re.sub(r"\s+", " ", str(raw_description).strip())


def build_product_audit_label(product_row: Optional[dict]):
    row = product_row or {}
    reference_value = row.get("referencia") or row.get("codigo_articulo") or row.get("codigo") or row.get("producto_codigo") or "sin referencia"
    return f"[{reference_value}] - {get_exact_product_description(row)}"


def has_meaningful_product_anchor(product_request: Optional[dict]):
    request = product_request or {}
    if request.get("product_codes") or request.get("brand_filters"):
        return True
    return bool(get_specific_product_terms(request))


def build_followup_inventory_request(text_value: Optional[str], product_request: Optional[dict], conversation_context: Optional[dict]):
    request = dict(product_request or {})
    previous_request = dict((conversation_context or {}).get("last_product_request") or {})
    if not previous_request or has_meaningful_product_anchor(request):
        return request

    has_followup_filter = any(
        request.get(field_name)
        for field_name in [
            "requested_unit",
            "store_filters",
            "direction_filters",
            "size_filters",
            "color_filters",
            "finish_filters",
        ]
    )
    if not has_followup_filter:
        return request

    # Short follow-ups like 'el de galon en ferrebox' should inherit the last confirmed product anchor.
    merged_request = dict(previous_request)
    for field_name in [
        "requested_quantity",
        "requested_unit",
        "quantity_expression",
        "store_filters",
        "direction_filters",
        "size_filters",
        "color_filters",
        "finish_filters",
        "brand_filters",
    ]:
        current_value = request.get(field_name)
        if current_value:
            if isinstance(current_value, list):
                merged_request[field_name] = merge_unique_terms(merged_request.get(field_name), current_value)
            else:
                merged_request[field_name] = current_value
    merged_request["original_query"] = text_value or request.get("original_query") or previous_request.get("original_query") or ""
    merged_request["followup_from_previous_product"] = True
    return merged_request


def is_technical_advisory_message(text_value: Optional[str]):
    """Detect if the client is asking for product application advice, not a product search."""
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    if any(keyword in normalized for keyword in TECHNICAL_ADVISORY_KEYWORDS):
        return True
    advisory_patterns = [
        r"\b(necesito|quiero|voy a|puedo|como|cómo)\b.{0,25}\b(pintar|recubrir|impermeabilizar|sellar|proteger|aplicar|tratar|barnizar|lacar|esmaltar|lijar|resanar)\b",
        r"\b(que|qué)\s+(pintura|producto|sistema|recubrimiento|impermeabilizante|sellador|esmalte|barniz)\b.{0,25}\b(para|uso|necesito|sirve|recomienda)\b",
        r"\b(tengo|tiene|hay)\s+(un\s+)?(problema|daño|deterioro|fisura|grieta|filtración|filtracion)\b",
        r"\b(se\s+)(pela|descascara|ampolla|mancha|deteriora|cae|sale|agrieta|fisura|oxida)\b",
        r"\b(me\s+)?(recomienda|aconseja|sugiere|asesora)\b.{0,30}\b(para|sobre|con)\b",
    ]
    for pattern in advisory_patterns:
        if re.search(pattern, normalized):
            return True
    return False


def has_keyword_or_similar(text_value: Optional[str], keywords: list[str], threshold: float = 0.84):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    if any(keyword in normalized for keyword in keywords):
        return True
    tokens = re.findall(r"[a-z0-9.-]+", normalized)
    for token in tokens:
        for keyword in keywords:
            if SequenceMatcher(None, token, normalize_text_value(keyword)).ratio() >= threshold:
                return True
    return False


def get_presentation_label(unit_value: Optional[str], quantity_value: Optional[float] = None):
    if not unit_value:
        return ""
    singular_label, plural_label = PRESENTATION_LABELS.get(unit_value, (unit_value, f"{unit_value}s"))
    if quantity_value is not None and float(quantity_value) == 1:
        return singular_label
    return plural_label


def should_store_learning_phrase(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    if is_product_code_message(normalized):
        return False
    if normalized in {"si", "sí", "esa", "ese", "la primera", "la segunda", "la tercera", "opcion 1", "opcion 2", "opcion 3"}:
        return False
    return len(normalized) >= 4


def contains_product_correction_signal(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    return any(
        phrase in normalized
        for phrase in [
            "no era",
            "me corrijo",
            "corrijo",
            "quise decir",
            "mas bien",
            "más bien",
            "no ese",
            "no esa",
            "cambialo",
            "cámbialo",
            "cambio la referencia",
            "la referencia correcta",
        ]
    )


def should_merge_previous_learning_phrase(current_request: Optional[dict], previous_request: Optional[dict]):
    if not current_request or not previous_request:
        return False
    if contains_product_correction_signal(current_request.get("original_query")):
        return False
    current_codes = {normalize_reference_value(value) for value in (current_request.get("product_codes") or []) if value}
    previous_codes = {normalize_reference_value(value) for value in (previous_request.get("product_codes") or []) if value}
    if current_codes and previous_codes and current_codes != previous_codes:
        return False
    current_terms = set(get_specific_product_terms(current_request) or current_request.get("core_terms") or [])
    previous_terms = set(get_specific_product_terms(previous_request) or previous_request.get("core_terms") or [])
    if current_terms and previous_terms and not (current_terms & previous_terms):
        return False
    return True


def resolve_confirmed_learning_product_row(description_asociada: str, conversation_context: Optional[dict]):
    clarification_options = list((conversation_context or {}).get("clarification_options") or [])
    selected_option = resolve_product_clarification_choice(description_asociada, clarification_options)
    if selected_option:
        return {
            "referencia": selected_option.get("reference") or selected_option.get("referencia"),
            "descripcion": selected_option.get("name") or selected_option.get("descripcion"),
            "marca": selected_option.get("brand") or selected_option.get("marca"),
            "presentacion_canonica": selected_option.get("presentation") or selected_option.get("presentacion_canonica"),
        }

    learning_request = prepare_product_request_for_search(description_asociada)
    product_rows = lookup_product_context(description_asociada, learning_request)
    reliable_rows = select_reliable_learning_rows(learning_request, product_rows)
    if len(reliable_rows) == 1:
        return reliable_rows[0]

    explicit_codes = extract_product_codes(description_asociada)
    if explicit_codes and product_rows:
        normalized_codes = {normalize_reference_value(code) for code in explicit_codes}
        for row in product_rows:
            row_code = normalize_reference_value(row.get("referencia") or row.get("codigo_articulo"))
            if row_code and row_code in normalized_codes:
                return row
    return None


def build_learning_phrase_candidates(product_request: Optional[dict]):
    if not product_request:
        return []

    generic_terms = set(PRODUCT_STOPWORDS)
    for alias_group in PRESENTATION_ALIASES.values():
        generic_terms.update(normalize_text_value(alias) for alias in alias_group)
    for alias_group in STORE_ALIASES.values():
        generic_terms.update(normalize_text_value(alias) for alias in alias_group)

    phrases = []

    def add_phrase(raw_phrase: Optional[str]):
        normalized_phrase = normalize_text_value(raw_phrase)
        if not normalized_phrase or not should_store_learning_phrase(normalized_phrase):
            return
        if normalized_phrase not in phrases:
            phrases.append(normalized_phrase)

    add_phrase(product_request.get("original_query"))
    add_phrase(" ".join(product_request.get("core_terms") or []))

    specific_terms = []
    for term in product_request.get("core_terms") or []:
        normalized_term = normalize_text_value(term)
        if not normalized_term or normalized_term in generic_terms:
            continue
        if normalized_term not in specific_terms:
            specific_terms.append(normalized_term)

    if specific_terms:
        add_phrase(" ".join(specific_terms))
        if product_request.get("requested_unit"):
            add_phrase(" ".join(specific_terms + [product_request.get("requested_unit")]))
        if product_request.get("brand_filters"):
            for brand_name in product_request.get("brand_filters")[:2]:
                add_phrase(" ".join(specific_terms + [brand_name]))

    return phrases[:8]


def select_reliable_learning_rows(product_request: Optional[dict], product_context: list[dict]):
    if not product_request or not product_context:
        return []

    if product_request.get("product_codes"):
        return product_context[:1]

    if len(product_context) == 1:
        return product_context[:1]

    top_row = product_context[0]
    second_row = product_context[1] if len(product_context) > 1 else None
    top_specific = top_row.get("specific_score") or 0
    top_match = top_row.get("match_score") or 0
    top_brand = top_row.get("brand_score") or 0
    top_size = top_row.get("size_score") or 0
    second_specific = second_row.get("specific_score") or 0 if second_row else 0
    second_match = second_row.get("match_score") or 0 if second_row else 0

    if top_size > 0 and top_match >= 2:
        return [top_row]
    if top_brand > 0 and top_match >= 2:
        return [top_row]
    if top_specific >= 2 and (top_specific > second_specific or top_match > second_match):
        return [top_row]
    if top_match >= 3 and top_match > second_match:
        return [top_row]
    return []


def is_learned_reference_relevant(product_request: Optional[dict], learned_row: dict):
    if not product_request:
        return False

    description_text = normalize_text_value(learned_row.get("canonical_description"))
    brand_text = normalize_text_value(learned_row.get("canonical_brand"))
    combined_text = f"{description_text} {brand_text}".strip()
    specific_terms = get_specific_product_terms(product_request)
    if specific_terms and not any(term in combined_text for term in specific_terms):
        return False

    requested_unit = product_request.get("requested_unit")
    learned_presentation = normalize_text_value(learned_row.get("canonical_presentation"))
    if requested_unit and learned_presentation and requested_unit != learned_presentation:
        return False

    brand_filters = product_request.get("brand_filters") or []
    if brand_filters:
        matches_brand = False
        for brand_name in brand_filters:
            if brand_name in combined_text:
                matches_brand = True
                break
            for alias in BRAND_ALIASES.get(brand_name, []):
                if normalize_text_value(alias) in combined_text:
                    matches_brand = True
                    break
            if matches_brand:
                break
        if not matches_brand:
            return False

    return True


_product_learning_table_ensured = False

def ensure_product_learning_table():
    global _product_learning_table_ensured
    if _product_learning_table_ensured:
        return
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS public.agent_product_learning (
                    id bigserial PRIMARY KEY,
                    normalized_phrase text NOT NULL,
                    raw_phrase text NOT NULL,
                    canonical_reference text NOT NULL,
                    canonical_description text,
                    canonical_brand text,
                    canonical_presentation text,
                    source_conversation_id bigint REFERENCES public.agent_conversation(id) ON DELETE SET NULL,
                    source_message text,
                    confidence numeric(5,4) NOT NULL DEFAULT 0.7500,
                    usage_count integer NOT NULL DEFAULT 1,
                    created_at timestamptz NOT NULL DEFAULT now(),
                    updated_at timestamptz NOT NULL DEFAULT now(),
                    CONSTRAINT uq_agent_product_learning UNIQUE (normalized_phrase, canonical_reference)
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS idx_agent_product_learning_phrase
                ON public.agent_product_learning(normalized_phrase)
                """
            )
        )
    _product_learning_table_ensured = True


_product_companion_table_ensured = False

def ensure_product_companion_table():
    global _product_companion_table_ensured
    if _product_companion_table_ensured:
        return
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS public.agent_product_companion (
                    id bigserial PRIMARY KEY,
                    producto_referencia text NOT NULL,
                    producto_descripcion text,
                    companion_referencia text NOT NULL,
                    companion_descripcion text,
                    tipo_relacion varchar(60) NOT NULL,
                    proporcion text,
                    notas text,
                    source_conversation_id bigint REFERENCES public.agent_conversation(id) ON DELETE SET NULL,
                    confidence numeric(5,4) NOT NULL DEFAULT 0.9500,
                    activo boolean NOT NULL DEFAULT true,
                    created_at timestamptz NOT NULL DEFAULT now(),
                    updated_at timestamptz NOT NULL DEFAULT now(),
                    CONSTRAINT uq_agent_product_companion UNIQUE (producto_referencia, companion_referencia, tipo_relacion)
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS idx_agent_product_companion_ref
                ON public.agent_product_companion(producto_referencia)
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS idx_agent_product_companion_companion
                ON public.agent_product_companion(companion_referencia)
                """
            )
        )
    _product_companion_table_ensured = True


# ---------------------------------------------------------------------------
# WhatsApp Media Download & Document Processing for Expert Training
# ---------------------------------------------------------------------------

def download_whatsapp_media(media_id: str) -> tuple[bytes, str]:
    """Download media bytes from WhatsApp Cloud API. Returns (bytes, mime_type)."""
    # Step 1: Get media URL
    url_resp = requests.get(
        f"https://graph.facebook.com/v22.0/{media_id}",
        headers={"Authorization": f"Bearer {get_whatsapp_access_token()}"},
        timeout=15,
    )
    if url_resp.status_code >= 400:
        raise RuntimeError(f"WhatsApp media metadata error {url_resp.status_code}: {url_resp.text[:300]}")
    media_info = url_resp.json()
    media_url = media_info.get("url")
    mime_type = media_info.get("mime_type", "application/octet-stream")
    if not media_url:
        raise RuntimeError(f"No URL in media response: {media_info}")

    # Step 2: Download actual bytes
    dl_resp = requests.get(
        media_url,
        headers={"Authorization": f"Bearer {get_whatsapp_access_token()}"},
        timeout=60,
    )
    if dl_resp.status_code >= 400:
        raise RuntimeError(f"WhatsApp media download error {dl_resp.status_code}")
    return dl_resp.content, mime_type


def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    """Extract text from PDF using PyMuPDF (mirrors ingest_technical_sheets.py)."""
    import fitz
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    for page in doc:
        page_parts = []
        try:
            tables = page.find_tables()
            if tables and tables.tables:
                for table in tables:
                    table_data = table.extract()
                    if table_data:
                        formatted_rows = []
                        for row in table_data:
                            clean_cells = [str(cell).strip() if cell else "" for cell in row]
                            if len(clean_cells) == 2 and clean_cells[0] and clean_cells[1]:
                                formatted_rows.append(f"{clean_cells[0]}: {clean_cells[1]}")
                            elif any(c for c in clean_cells):
                                formatted_rows.append(" | ".join(c for c in clean_cells if c))
                        if formatted_rows:
                            page_parts.append("\n".join(formatted_rows))
        except Exception:
            pass
        text_content = page.get_text("text")
        if text_content and text_content.strip():
            if page_parts:
                blocks = page.get_text("blocks")
                non_table_text = []
                for block in blocks:
                    if block[6] == 0:
                        block_text = block[4].strip()
                        if block_text:
                            non_table_text.append(block_text)
                if non_table_text:
                    page_parts.insert(0, "\n".join(non_table_text))
                pages.append("\n\n".join(page_parts))
            else:
                pages.append(text_content.strip())
    doc.close()
    return "\n\n".join(pages)


def extract_text_from_excel_bytes(excel_bytes: bytes) -> str:
    """Extract text from Excel file (all sheets) using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows_text.append(" | ".join(c for c in cells if c))
        if rows_text:
            parts.append(f"[HOJA: {sheet_name}]\n" + "\n".join(rows_text))
    wb.close()
    return "\n\n".join(parts)


def extract_text_from_image_bytes(image_bytes: bytes, mime_type: str) -> str:
    """Use OpenAI GPT-4o vision to extract text/content from an image."""
    try:
        b64 = base64.b64encode(image_bytes).decode("utf-8")
        client = get_openai_client()
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": (
                                "Extrae TODO el texto visible en esta imagen. "
                                "Si es una ficha técnica, tabla de datos, cálculo o documento técnico, "
                                "transcríbelo completo preservando la estructura (tablas, listas, secciones). "
                                "Si es una foto de un producto o superficie, describe exactamente lo que ves. "
                                "Responde SOLO con el contenido extraído, sin comentarios adicionales."
                            ),
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{mime_type};base64,{b64}"},
                        },
                    ],
                }
            ],
            max_tokens=4000,
            temperature=0,
        )
        return response.choices[0].message.content.strip()
    except Exception as exc:
        logger.warning("Image text extraction failed: %s", exc)
        return f"[Error extrayendo texto de imagen: {exc}]"


def extract_text_from_media(media_bytes: bytes, mime_type: str, filename: str = "") -> tuple[str, str]:
    """Extract text from media based on MIME type. Returns (text, doc_type)."""
    mime_lower = (mime_type or "").lower()
    fname_lower = (filename or "").lower()

    if "pdf" in mime_lower or fname_lower.endswith(".pdf"):
        return extract_text_from_pdf_bytes(media_bytes), "pdf"
    elif any(x in mime_lower for x in ["spreadsheet", "excel", "xlsx", "xls"]) or \
         fname_lower.endswith((".xlsx", ".xls")):
        return extract_text_from_excel_bytes(media_bytes), "excel"
    elif any(x in mime_lower for x in ["image/", "png", "jpeg", "jpg", "webp", "gif"]):
        return extract_text_from_image_bytes(media_bytes, mime_type), "image"
    elif "text" in mime_lower or fname_lower.endswith((".txt", ".csv", ".md")):
        try:
            return media_bytes.decode("utf-8"), "text"
        except UnicodeDecodeError:
            return media_bytes.decode("latin-1"), "text"
    else:
        # Try as text first, fall back to binary description
        try:
            text = media_bytes.decode("utf-8")
            if text.strip():
                return text, "text"
        except UnicodeDecodeError:
            pass
        return f"[Archivo binario no soportado: {mime_type}, {len(media_bytes)} bytes]", "unsupported"


def _chunk_expert_document(text: str, filename: str, max_chars: int = 2000, overlap: int = 300) -> list[str]:
    """Chunk an expert-uploaded document for RAG storage."""
    if not text or not text.strip():
        return []
    # Section-aware splitting (same logic as ingest_technical_sheets.py)
    section_header_re = [
        r"^[A-ZÁÉÍÓÚÑ\s/]{5,60}$",
        r"^(?:\d+[\.\)]\s*)?[A-ZÁÉÍÓÚÑ][\w\s/]+:?\s*$",
    ]

    def is_header(line):
        s = line.strip()
        if not s or len(s) < 4 or len(s) > 80:
            return False
        return any(re.match(p, s) for p in section_header_re)

    lines = text.split("\n")
    sections = []
    cur_hdr, cur_body = "GENERAL", []
    for line in lines:
        if is_header(line) and line.strip():
            if cur_body:
                body = "\n".join(cur_body).strip()
                if body:
                    sections.append((cur_hdr, body))
            cur_hdr = line.strip()
            cur_body = []
        else:
            cur_body.append(line)
    if cur_body:
        body = "\n".join(cur_body).strip()
        if body:
            sections.append((cur_hdr, body))
    if not sections:
        sections = [("GENERAL", text)]

    product_name = re.sub(r"\.\w+$", "", filename).strip()
    context_header = f"[DOCUMENTO EXPERTO: {product_name}] [FUENTE: Subido por asesor Ferreinox]"

    chunks = []
    for section_header, section_body in sections:
        prefix = f"{context_header}\n[SECCIÓN: {section_header}]\n\n"
        available = max_chars - len(prefix)
        if len(section_body) <= available:
            chunks.append(f"{prefix}{section_body}")
        else:
            start = 0
            while start < len(section_body):
                end = start + available
                if end < len(section_body):
                    brk = section_body.rfind("\n\n", start + available // 2, end)
                    if brk == -1:
                        brk = section_body.rfind(". ", start + available // 2, end)
                    if brk > start:
                        end = brk + 1
                sub = section_body[start:end].strip()
                if sub:
                    chunks.append(f"{prefix}{sub}")
                start = end - overlap if end < len(section_body) else len(section_body)
    return chunks


def ingest_expert_document_to_rag(
    extracted_text: str,
    filename: str,
    expert_cedula: str,
    conversation_id: int | None = None,
    marca: str | None = None,
) -> dict:
    """Chunk, embed and store an expert-uploaded document into the RAG system."""
    chunks = _chunk_expert_document(extracted_text, filename)
    if not chunks:
        return {"ingested": False, "reason": "No se pudo extraer texto útil del documento."}

    # Generate embeddings
    try:
        client = get_openai_client()
        all_embeddings = []
        for i in range(0, len(chunks), 50):
            batch = chunks[i:i + 50]
            resp = client.embeddings.create(
                model="text-embedding-3-small",
                input=batch,
                dimensions=1536,
            )
            batch_emb = [item.embedding for item in sorted(resp.data, key=lambda x: x.index)]
            all_embeddings.extend(batch_emb)
    except Exception as exc:
        return {"ingested": False, "reason": f"Error generando embeddings: {exc}"}

    # Store in agent_technical_doc_chunk
    doc_path_lower = f"expert_upload/{expert_cedula}/{filename.lower()}"
    doc_filename = filename

    # Infer brand from content/filename
    if not marca:
        combined = (filename + " " + extracted_text[:500]).lower()
        brand_patterns = [
            "pintuco", "viniltex", "koraza", "pintulux", "domestico", "doméstico",
            "abracol", "yale", "goya", "mega", "international",
            "interseal", "intergard", "interchar", "interzone", "interthane",
        ]
        for bp in brand_patterns:
            if bp in combined:
                marca = bp.capitalize()
                break

    familia = re.sub(r"\.\w+$", "", filename).strip()

    try:
        engine = get_db_engine()
        raw_conn = engine.raw_connection()
        try:
            cur = raw_conn.cursor()
            # Delete previous chunks for this exact document (re-upload overwrites)
            cur.execute(
                "DELETE FROM public.agent_technical_doc_chunk WHERE doc_path_lower = %s",
                (doc_path_lower,),
            )
            for idx, (chunk_text, embedding) in enumerate(zip(chunks, all_embeddings)):
                embedding_str = "[" + ",".join(str(v) for v in embedding) + "]"
                metadata_json = json.dumps({
                    "uploaded_by_expert": expert_cedula,
                    "conversation_id": conversation_id,
                    "upload_date": datetime.utcnow().isoformat(),
                    "source": "whatsapp_expert_upload",
                }, ensure_ascii=False)
                cur.execute(
                    """
                    INSERT INTO public.agent_technical_doc_chunk
                        (doc_filename, doc_path_lower, chunk_index, chunk_text,
                         marca, familia_producto, tipo_documento, metadata,
                         embedding, token_count)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::vector, %s)
                    ON CONFLICT (doc_path_lower, chunk_index) DO UPDATE SET
                        chunk_text = EXCLUDED.chunk_text,
                        marca = EXCLUDED.marca,
                        familia_producto = EXCLUDED.familia_producto,
                        metadata = EXCLUDED.metadata,
                        embedding = EXCLUDED.embedding,
                        token_count = EXCLUDED.token_count,
                        ingested_at = now()
                    """,
                    (
                        doc_filename, doc_path_lower, idx, chunk_text,
                        marca, familia, "ficha_tecnica_experto",
                        metadata_json, embedding_str, len(chunk_text) // 4,
                    ),
                )
            raw_conn.commit()
        finally:
            raw_conn.close()
    except Exception as exc:
        return {"ingested": False, "reason": f"Error almacenando en base de datos: {exc}"}

    return {
        "ingested": True,
        "chunks_count": len(chunks),
        "doc_path": doc_path_lower,
        "filename": filename,
        "marca_detectada": marca,
    }


# ─── Alertas de superficie extensibles via DB (Rec 3) ────────────────────────

_surface_alerts_cache: Optional[list] = None
_surface_alerts_cache_ts: float = 0
_SURFACE_ALERTS_CACHE_TTL = 300  # 5 min


def ensure_surface_alerts_table():
    """Create agent_surface_alerts table for extensible surface warning rules."""
    engine = get_db_engine()
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS public.agent_surface_alerts (
                id bigserial PRIMARY KEY,
                surfaces text[] NOT NULL,
                conditions text[],
                alert_text text NOT NULL,
                severity text NOT NULL DEFAULT 'critica',
                activo boolean NOT NULL DEFAULT true,
                created_by text,
                created_at timestamptz NOT NULL DEFAULT now()
            )
        """))
        # Seed from hardcoded defaults if table is empty
        count = conn.execute(text("SELECT count(*) FROM public.agent_surface_alerts")).scalar()
        if count == 0:
            _seed_surface_alerts_defaults(conn)


def _seed_surface_alerts_defaults(conn):
    """Seed agent_surface_alerts with the 6 original hardcoded rules."""
    defaults = [
        (
            ["concreto", "piso", "piso industrial", "piso vehicular"],
            ["superficie nueva", "sin pintar"],
            "🚨 ALERTA CRÍTICA DE SUPERFICIE: El concreto nuevo exige MÍNIMO 28 DÍAS de curado "
            "antes de aplicar cualquier recubrimiento. ANTES de recomendar productos, DEBES "
            "informar esto al cliente y preguntar: '¿Hace cuánto fue vaciado el concreto?' "
            "Si tiene menos de 28 días → NO recomiendes aplicar. La humedad residual del "
            "concreto causará FALLA por ampollamiento, descascaramiento y pérdida de adherencia.",
            "critica",
        ),
        (
            ["metal", "metal/inmersión", "reja"],
            ["óxido"],
            "🚨 ALERTA CRÍTICA DE SUPERFICIE: Metal con óxido requiere preparación mecánica "
            "OBLIGATORIA antes de cualquier recubrimiento. DEBES preguntar el GRADO de oxidación "
            "(leve/moderado/severo) y recomendar lija, disco flap o grata según corresponda. "
            "Sin preparación correcta, CUALQUIER anticorrosivo fallará por falta de adherencia.",
            "critica",
        ),
        (
            ["interior húmedo"],
            ["humedad", "filtración", "goteras", "moho/hongos"],
            "🚨 ALERTA CRÍTICA DE SUPERFICIE: Problema de humedad detectado. ANTES de recomendar "
            "pintura, DEBES diagnosticar la CAUSA de la humedad (capilaridad, filtración, "
            "condensación). Si la fuente no se elimina, cualquier recubrimiento fallará. "
            "Pregunta: '¿La humedad viene de adentro del muro, de arriba, o aparece por temporada?'",
            "critica",
        ),
        (
            ["fachada", "exterior"],
            ["pintura descascarando", "pintura soplada"],
            "🚨 ALERTA CRÍTICA DE SUPERFICIE: Fachada con pintura en mal estado. ANTES de "
            "recomendar repintura, DEBES indicar que se necesita REMOVER la pintura suelta "
            "completamente (raspar, lijar, hidrolavado). Aplicar sobre pintura soplada causa "
            "falla inmediata. Pregunta al cliente cómo piensa preparar la superficie.",
            "critica",
        ),
        (
            ["madera", "madera exterior", "madera/metal"],
            ["superficie nueva", "sin pintar"],
            "⚠️ ALERTA DE SUPERFICIE: Madera nueva requiere verificar contenido de humedad "
            "(máximo 18%) antes de aplicar recubrimiento. PREGUNTA al cliente si la madera es "
            "nueva/seca o si estuvo expuesta a lluvia. Madera húmeda → dejar secar primero.",
            "advertencia",
        ),
        (
            ["piso deportivo"],
            None,
            "⚠️ ALERTA DE SUPERFICIE: Pisos deportivos (canchas) requieren productos "
            "específicos con resistencia a abrasión y tráfico. NUNCA recomendar Pintucoat, "
            "Interseal ni recubrimientos industriales. El producto correcto es Pintura para "
            "Canchas de Pintuco. Si es concreto nuevo → aplica regla de 28 días de curado.",
            "advertencia",
        ),
    ]
    for surfaces, conditions, alert_text, severity in defaults:
        conn.execute(text("""
            INSERT INTO public.agent_surface_alerts (surfaces, conditions, alert_text, severity, created_by)
            VALUES (:surfaces, :conditions, :alert_text, :severity, :created_by)
        """), {
            "surfaces": surfaces,
            "conditions": conditions,
            "alert_text": alert_text,
            "severity": severity,
            "created_by": "SYSTEM_SEED",
        })
    logger.info("Seeded %d default surface alerts", len(defaults))


def fetch_surface_alerts_from_db() -> list[dict]:
    """Load active surface alerts from DB with in-memory cache (5 min TTL)."""
    global _surface_alerts_cache, _surface_alerts_cache_ts
    import time as _time
    now = _time.time()
    if _surface_alerts_cache is not None and (now - _surface_alerts_cache_ts) < _SURFACE_ALERTS_CACHE_TTL:
        return _surface_alerts_cache
    try:
        engine = get_db_engine()
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT surfaces, conditions, alert_text
                FROM public.agent_surface_alerts
                WHERE activo = true
                ORDER BY id
            """)).fetchall()
        alerts = []
        for row in rows:
            alerts.append({
                "surfaces": row[0] or [],
                "conditions": row[1] or [None],
                "alert": row[2],
            })
        _surface_alerts_cache = alerts
        _surface_alerts_cache_ts = now
        return alerts
    except Exception as exc:
        logger.debug("fetch_surface_alerts_from_db error: %s", exc)
        # Return hardcoded fallback if DB fails
        return []


def invalidate_surface_alerts_cache():
    """Force refresh of surface alerts cache."""
    global _surface_alerts_cache, _surface_alerts_cache_ts
    _surface_alerts_cache = None
    _surface_alerts_cache_ts = 0


def ensure_expert_knowledge_table():
    """Create agent_expert_knowledge table for commercial reinforcement notes."""
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS public.agent_expert_knowledge (
                    id bigserial PRIMARY KEY,
                    cedula_experto text NOT NULL,
                    nombre_experto text NOT NULL DEFAULT 'PABLO CESAR MAFLA BANOL',
                    contexto_tags text NOT NULL,
                    producto_recomendado text,
                    producto_desestimado text,
                    nota_comercial text NOT NULL,
                    tipo text NOT NULL DEFAULT 'preferencia',
                    activo boolean NOT NULL DEFAULT true,
                    conversation_id bigint REFERENCES public.agent_conversation(id) ON DELETE SET NULL,
                    created_at timestamptz NOT NULL DEFAULT now(),
                    updated_at timestamptz NOT NULL DEFAULT now()
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS idx_agent_expert_knowledge_tags
                ON public.agent_expert_knowledge USING gin(to_tsvector('spanish', contexto_tags || ' ' || nota_comercial))
                """
            )
        )
    # ── Ensure embedding column for semantic search (Solution 3) ──
    _ensure_expert_knowledge_embedding_column()
    # ── Ensure extensible surface alerts table (Rec 3) ──
    try:
        ensure_surface_alerts_table()
    except Exception as _sa_exc:
        logger.debug("ensure_surface_alerts_table error: %s", _sa_exc)
    # Run Phase 19 seeds after table is ensured
    seed_expert_knowledge_phase19()
    seed_expert_knowledge_polyurethane_1550()


_expert_embedding_column_ensured = False


def _ensure_expert_knowledge_embedding_column():
    """Add vector embedding column to agent_expert_knowledge if missing.
    This enables semantic search for expert directives in agent_context.py."""
    global _expert_embedding_column_ensured
    if _expert_embedding_column_ensured:
        return
    try:
        engine = get_db_engine()
        with engine.begin() as conn:
            # Check if column exists
            result = conn.execute(text(
                """SELECT column_name FROM information_schema.columns 
                   WHERE table_name = 'agent_expert_knowledge' AND column_name = 'embedding'"""
            )).fetchone()
            if not result:
                conn.execute(text(
                    "ALTER TABLE public.agent_expert_knowledge ADD COLUMN embedding vector(1536)"
                ))
                conn.execute(text(
                    """CREATE INDEX IF NOT EXISTS idx_agent_expert_knowledge_embedding
                       ON public.agent_expert_knowledge
                       USING hnsw (embedding vector_cosine_ops)
                       WITH (m = 16, ef_construction = 64)"""
                ))
                logger.info("Added embedding column + HNSW index to agent_expert_knowledge")
        _expert_embedding_column_ensured = True
        # Backfill embeddings for existing rows
        _backfill_expert_knowledge_embeddings()
    except Exception as exc:
        logger.debug("_ensure_expert_knowledge_embedding_column error: %s", exc)
        _expert_embedding_column_ensured = True  # Don't retry on failure


def _backfill_expert_knowledge_embeddings():
    """Generate embeddings for expert knowledge rows that don't have one yet."""
    try:
        engine = get_db_engine()
        with engine.connect() as conn:
            rows = conn.execute(text(
                """SELECT id, contexto_tags, nota_comercial, producto_recomendado, producto_desestimado
                   FROM public.agent_expert_knowledge
                   WHERE activo = true AND embedding IS NULL"""
            )).mappings().all()
        
        if not rows:
            return
        
        client = get_openai_client()
        logger.info("Backfilling %d expert knowledge embeddings...", len(rows))
        
        for row in rows:
            # Build embedding text from all searchable fields
            embed_text = " ".join(filter(None, [
                row.get("contexto_tags"),
                row.get("nota_comercial"),
                row.get("producto_recomendado"),
                row.get("producto_desestimado"),
            ]))
            if not embed_text.strip():
                continue
            try:
                resp = client.embeddings.create(
                    model="text-embedding-3-small",
                    input=embed_text.strip()[:500],
                    dimensions=1536,
                )
                embedding = resp.data[0].embedding
                embedding_literal = "[" + ",".join(str(v) for v in embedding) + "]"
                
                with engine.begin() as conn:
                    conn.execute(
                        text("UPDATE public.agent_expert_knowledge SET embedding = :emb::vector WHERE id = :id"),
                        {"emb": embedding_literal, "id": row["id"]},
                    )
            except Exception as exc:
                logger.debug("Backfill embedding error for id=%s: %s", row.get("id"), exc)
                continue
        
        logger.info("Expert knowledge embedding backfill complete")
    except Exception as exc:
        logger.debug("_backfill_expert_knowledge_embeddings error: %s", exc)


def seed_expert_knowledge_phase19():
    """Seed critical expert knowledge from Diego García (Phase 19 corrections)."""
    try:
        engine = get_db_engine()
        seeds = [
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso, pintucoat, resistencia, tráfico medio, acabado mate",
                "rec": "Pintucoat", "des": None,
                "nota": "Pintucoat es de resistencia MEDIA, acabado MATE. NO resiste tráfico pesado de montacargas. Para tráfico pesado usar Intergard 2002 + cuarzo.",
                "tipo": "correccion",
            },
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso, intergard 740, resistencia media, acabado brillante",
                "rec": "Intergard 740", "des": None,
                "nota": "Intergard 740: epóxico para pisos de resistencia MEDIA con acabado BRILLANTE. Alternativa al Pintucoat cuando el cliente quiere más brillo.",
                "tipo": "preferencia",
            },
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso, intergard 2002, cuarzo, tráfico pesado, montacargas, alta resistencia",
                "rec": "Intergard 2002", "des": "Pintucoat",
                "nota": "Intergard 2002 es de alto volumen de sólidos. Con cuarzo (ref 5891610) esparcido por broadcasting se obtiene alta resistencia para montacargas y estibadores. Sistema: Interseal gris RAL 7038 → Intergard 2002 + cuarzo.",
                "tipo": "preferencia",
            },
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso, primer 50rs, epoxy primer 50rs, imprimante, concreto, metal",
                "rec": "Interseal gris RAL 7038", "des": "Primer 50RS",
                "nota": "PRIMER 50RS es EXCLUSIVAMENTE para estructuras metálicas. NUNCA para pisos de concreto. Para pisos de concreto el imprimante correcto es Interseal gris RAL 7038.",
                "tipo": "contraindicacion",
            },
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso, imprimante, concreto, interseal gris, ral 7038",
                "rec": "Interseal gris RAL 7038", "des": None,
                "nota": "Para pisos de concreto, el imprimante correcto es Interseal gris RAL 7038. Sella la porosidad del concreto y mejora adherencia del sistema epóxico.",
                "tipo": "uso_especifico",
            },
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso, cuarzo, ref 5891610, broadcasting, antideslizante",
                "rec": "Cuarzo ref 5891610", "des": None,
                "nota": "El cuarzo ref 5891610 se aplica por broadcasting (esparcido) sobre el Intergard 2002 fresco. Aumenta la resistencia mecánica y proporciona acabado antideslizante.",
                "tipo": "uso_especifico",
            },
        ]
        with engine.begin() as conn:
            for s in seeds:
                # Only insert if not already seeded (check by nota_comercial)
                existing = conn.execute(
                    text("SELECT id FROM public.agent_expert_knowledge WHERE nota_comercial = :nota AND cedula_experto = :cedula LIMIT 1"),
                    {"nota": s["nota"], "cedula": s["cedula"]},
                ).fetchone()
                if existing:
                    continue
                conn.execute(
                    text(
                        """
                        INSERT INTO public.agent_expert_knowledge
                            (cedula_experto, nombre_experto, contexto_tags, producto_recomendado,
                             producto_desestimado, nota_comercial, tipo)
                        VALUES (:cedula, :nombre, :tags, :rec, :des, :nota, :tipo)
                        """
                    ),
                    s,
                )
        logger.info("Phase 19 expert knowledge seeds applied successfully.")
    except Exception as exc:
        logger.warning("Could not seed Phase 19 expert knowledge: %s", exc)


def seed_expert_knowledge_polyurethane_1550():
    """Seed expert knowledge for Poliuretano Alto Tráfico 1550/1551 system (wood floors)."""
    try:
        engine = get_db_engine()
        seeds = [
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso madera, poliuretano alto trafico, 1550, 1551, vitrificar, barnizar piso, resina transparente, laca piso, escalera madera, interior",
                "rec": "Poliuretano Alto Tráfico 1550/1551", "des": "Barnex, Pintulac, barniz arquitectónico",
                "nota": (
                    "SISTEMA POLIURETANO TRANSPARENTE PARA PISOS (1550/1551): SOLO INTERIOR. "
                    "Bicomponente relación 1:1. Comp A = ref 1550 (MH PISOS TRAFIC ALT A BR INCO 3.79L), "
                    "Comp B/Cat = ref 1551 (MH PISOS TRAFICO ALT B CAT BR 3.79L). "
                    "Para pisos de madera, escaleras, cemento con acabado transparente resistente. "
                    "JERGA: 'vitrificar', 'barnizar piso cemento', 'resina transparente', 'laca garaje'. "
                    "PROHIBIDO en exteriores (entiza y amarillea). "
                    "Aplicación: 1) superficie limpia seca, 2) primera mano A+B, 3) secar 2-3h, "
                    "4) lijar grano 320/400, 5) limpiar y segunda mano. "
                    "Ofrecer cuando el cliente quiere algo MÁS resistente/fino que un barniz común."
                ),
                "tipo": "uso_especifico",
            },
            {
                "cedula": "1088266407", "nombre": "DIEGO MAURICIO GARCIA RENGIFO",
                "tags": "piso madera exterior, barniz piso exterior, poliuretano exterior",
                "rec": "Barnex, Wood Stain", "des": "Poliuretano Alto Tráfico 1550",
                "nota": (
                    "PROHIBIDO recomendar Poliuretano Alto Tráfico 1550/1551 en exteriores. "
                    "El poliuretano transparente entiza y amarillea con exposición UV. "
                    "Para madera exterior siempre ofrecer Barnex + Wood Stain."
                ),
                "tipo": "contraindicacion",
            },
        ]
        with engine.begin() as conn:
            for s in seeds:
                existing = conn.execute(
                    text("SELECT id FROM public.agent_expert_knowledge WHERE nota_comercial = :nota AND cedula_experto = :cedula LIMIT 1"),
                    {"nota": s["nota"], "cedula": s["cedula"]},
                ).fetchone()
                if existing:
                    continue
                conn.execute(
                    text(
                        """
                        INSERT INTO public.agent_expert_knowledge
                            (cedula_experto, nombre_experto, contexto_tags, producto_recomendado,
                             producto_desestimado, nota_comercial, tipo)
                        VALUES (:cedula, :nombre, :tags, :rec, :des, :nota, :tipo)
                        """
                    ),
                    s,
                )
        logger.info("Polyurethane 1550/1551 expert knowledge seeds applied successfully.")
    except Exception as exc:
        logger.warning("Could not seed Polyurethane 1550 expert knowledge: %s", exc)


def fetch_expert_knowledge(query: str, limit: int = 8) -> list[dict]:
    """Fetch commercial expert knowledge matching the query context.

    Uses an in-memory cache (refreshed every 120s) to avoid DB round-trips
    on every tool call.  With ~50 rows this is negligible memory.
    """
    if not query:
        return []
    try:
        normalized = normalize_text_value(query)
        raw_terms = re.findall(r"[a-z0-9áéíóúñ]+", normalized)
        stop_terms = {
            "para", "con", "sin", "por", "que", "como", "sobre", "entre", "desde",
            "hasta", "este", "esta", "estos", "estas", "solo", "necesito", "quiero",
            "techo", "techos", "pintar", "pintado", "exterior", "interior", "anos",
            "ano", "hace", "viejo", "vieja", "nuevo", "nueva", "usar", "aplicar",
            "producto", "productos", "sistema", "recomendar", "recomendacion",
        }
        terms = []
        for term in raw_terms:
            if len(term) < 3 or term in stop_terms or term in terms:
                continue
            terms.append(term)
        if not terms:
            terms = [t for t in raw_terms if len(t) >= 2][:10]
        if not terms:
            return []

        all_rows = _get_expert_knowledge_cache()

        scored = []
        seen_keys = set()
        anchor_terms = [
            t for t in terms
            if len(t) >= 6 or t in {"eternit", "fibrocemento", "asbesto", "sellomax", "koraza", "intervinil"}
        ]
        for row in all_rows:
            context_text = normalize_text_value(row.get("contexto_tags") or "")
            note_text = normalize_text_value(row.get("nota_comercial") or "")
            recommended_text = normalize_text_value(row.get("producto_recomendado") or "")
            rejected_text = normalize_text_value(row.get("producto_desestimado") or "")
            searchable = (
                context_text
                + " " + note_text
                + " " + recommended_text
                + " " + rejected_text
            )
            matched_terms = [t for t in terms if t in searchable]
            if not matched_terms:
                continue

            score = 0.0
            context_hits = 0
            for term in matched_terms:
                score += 1.0
                if term in context_text:
                    score += 2.0
                    context_hits += 1
                elif term in note_text:
                    score += 1.0
                elif term in recommended_text or term in rejected_text:
                    score += 0.4
                if len(term) >= 7:
                    score += 0.35

            anchor_context_hits = sum(1 for term in anchor_terms if term in context_text)
            if anchor_context_hits:
                score += 2.5 * anchor_context_hits
            elif anchor_terms:
                score -= 1.5

            if row.get("tipo") == "alerta_superficie" and context_hits:
                score += 1.5
            if row.get("tipo") == "evitar" and any(term in rejected_text for term in matched_terms):
                score += 0.75

            dedupe_key = (
                row.get("tipo") or "",
                context_text,
                note_text,
                recommended_text,
                rejected_text,
            )
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)

            if score >= 2.0:
                scored.append((score, len(matched_terms), row))
        scored.sort(key=lambda item: (-item[0], -item[1], -(item[2].get("_ts") or 0)))
        return [row for _, _, row in scored[:limit]]
    except Exception as exc:
        logger.debug("fetch_expert_knowledge error: %s", exc)
        return []


# ── In-memory cache for expert knowledge (tiny table, avoids DB round-trips) ──
_expert_knowledge_cache: list[dict] = []
_expert_knowledge_cache_ts: float = 0.0
_EXPERT_CACHE_TTL = 120  # seconds


def _get_expert_knowledge_cache() -> list[dict]:
    global _expert_knowledge_cache, _expert_knowledge_cache_ts
    import time as _time
    now = _time.time()
    if _expert_knowledge_cache and (now - _expert_knowledge_cache_ts) < _EXPERT_CACHE_TTL:
        return _expert_knowledge_cache
    try:
        engine = get_db_engine()
        with engine.connect() as connection:
            rows = connection.execute(
                text(
                    """SELECT id, contexto_tags, producto_recomendado, producto_desestimado,
                              nota_comercial, tipo, nombre_experto, created_at
                       FROM public.agent_expert_knowledge
                       WHERE activo = true
                       ORDER BY created_at DESC"""
                )
            ).mappings().all()
            _expert_knowledge_cache = [
                {**dict(r), "_ts": r["created_at"].timestamp() if r.get("created_at") else 0}
                for r in rows
            ]
            _expert_knowledge_cache_ts = now
    except Exception as exc:
        logger.debug("_get_expert_knowledge_cache refresh error: %s", exc)
    return _expert_knowledge_cache


def invalidate_expert_knowledge_cache():
    """Call after inserting new expert knowledge to force refresh on next read."""
    global _expert_knowledge_cache_ts
    _expert_knowledge_cache_ts = 0.0


# ── Color formula lookup (from LIBRO DE FORMULAS data) ──
def lookup_color_base(color_name: str, producto: str = "") -> list[dict]:
    """Find which BASE (Pastel/Tint/Deep/Accent) a color name requires.
    Returns matching entries from the color formulas catalog."""
    if not color_name or not _COLOR_FORMULAS:
        return []
    color_lower = color_name.lower().strip()
    producto_lower = producto.lower().strip() if producto else ""
    results = []
    for entry in _COLOR_FORMULAS:
        name_match = color_lower in entry["nombre"].lower()
        code_match = color_lower == entry["codigo"].lower()
        if not (name_match or code_match):
            continue
        if producto_lower and producto_lower not in entry["producto"].lower():
            continue
        results.append(entry)
    # Deduplicate by (codigo, base) keeping first
    seen = set()
    unique = []
    for r in results:
        key = (r["codigo"], r["base"])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique[:10]


# ── International product reference lookup ──
def lookup_international_product(producto: str, base: str = "", ral: str = "") -> list[dict]:
    """Find International/AkzoNobel product references with prices and catalyst info."""
    if not producto or not _INTERNATIONAL_PRODUCTS:
        return []
    producto_lower = producto.lower().strip()
    base_lower = base.lower().strip() if base else ""
    ral_str = str(ral).strip() if ral else ""
    results = []
    for entry in _INTERNATIONAL_PRODUCTS:
        prod_name = (entry.get("producto") or "").lower()
        linea = (entry.get("linea") or "").lower()
        if producto_lower not in prod_name and producto_lower not in linea:
            continue
        if base_lower and base_lower not in (entry.get("base") or "").lower():
            continue
        if ral_str and ral_str != (entry.get("ral") or "").strip():
            continue
        results.append(entry)
    return results[:20]


def fetch_product_price(referencia: str) -> Optional[dict]:
    """Look up price for a product by its referencia code from agent_precios.
    Uses pvp_sap for Pintuco/MPY brands, pvp_franquicia for complementary brands (Goya, Yale, Abracol, etc.)."""
    if not referencia:
        return None
    try:
        engine = get_db_engine()
        with engine.connect() as conn:
            row = conn.execute(
                text("""
                    SELECT referencia, descripcion, marca, cat_producto, aplicacion,
                           pvp_sap, pvp_franquicia,
                           COALESCE(NULLIF(pvp_sap, 0), NULLIF(pvp_franquicia, 0)) AS precio_mejor
                    FROM public.agent_precios
                    WHERE referencia = :ref AND (pvp_sap > 0 OR pvp_franquicia > 0)
                    LIMIT 1
                """),
                {"ref": str(referencia).strip()},
            ).mappings().first()
            if row:
                return dict(row)
    except Exception as exc:
        logger.debug("fetch_product_price error: %s", exc)
    return None


def fetch_client_by_nif_or_codigo(criterio: str) -> Optional[dict]:
    """Look up client info from agent_clientes by NIF (cedula/NIT) or by codigo."""
    if not criterio:
        return None
    clean = str(criterio).strip().replace(".", "").replace("-", "")
    try:
        engine = get_db_engine()
        with engine.connect() as conn:
            # Try by NIF first
            row = conn.execute(
                text("""
                    SELECT codigo, nombre, nif, direccion, telefono, ciudad, categoria,
                           email, persona_contacto, segmento, negocio, razon_social, riesgo_concedido
                    FROM public.agent_clientes
                    WHERE REPLACE(REPLACE(nif, '.', ''), '-', '') = :nif
                    LIMIT 1
                """),
                {"nif": clean},
            ).mappings().first()
            if row:
                return dict(row)
            # Fallback: try by codigo
            if clean.isdigit():
                row = conn.execute(
                    text("""
                        SELECT codigo, nombre, nif, direccion, telefono, ciudad, categoria,
                               email, persona_contacto, segmento, negocio, razon_social, riesgo_concedido
                        FROM public.agent_clientes
                        WHERE codigo = :cod
                        LIMIT 1
                    """),
                    {"cod": int(clean)},
                ).mappings().first()
                if row:
                    return dict(row)
    except Exception as exc:
        logger.debug("fetch_client_by_nif_or_codigo error: %s", exc)
    return None


def fetch_product_companions(referencia: str) -> list[dict]:
    """Fetch all active companion/complementary products for a given reference."""
    if not referencia:
        return []
    try:
        engine = get_db_engine()
        with engine.connect() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT c.companion_referencia, c.companion_descripcion, c.tipo_relacion,
                           c.proporcion, c.notas, c.confidence,
                           p.stock_total, p.descripcion AS descripcion_inventario
                    FROM public.agent_product_companion c
                    LEFT JOIN mv_productos p ON p.referencia = c.companion_referencia
                    WHERE c.producto_referencia = :ref AND c.activo = true
                    ORDER BY c.tipo_relacion, c.confidence DESC
                    """
                ),
                {"ref": str(referencia)},
            ).mappings().all()
            return [dict(r) for r in rows]
    except Exception:
        return []


# ---------------------------------------------------------------------------
# RAG: búsqueda semántica en fichas técnicas vectorizadas
# ---------------------------------------------------------------------------

def _generate_query_embedding(query_text: str) -> list[float] | None:
    """Generate embedding vector for a search query using OpenAI."""
    try:
        client = get_openai_client()
        response = client.embeddings.create(
            model="text-embedding-3-small",
            input=query_text.strip(),
            dimensions=1536,
        )
        return response.data[0].embedding
    except Exception:
        return None


PORTFOLIO_SEGMENT_ALIASES = {
    "recubrimientos": "recubrimientos_pinturas",
    "pinturas": "recubrimientos_pinturas",
    "recubrimientos_pinturas": "recubrimientos_pinturas",
    "auxiliares": "auxiliares_aplicacion",
    "auxiliares_aplicacion": "auxiliares_aplicacion",
    "diluyentes": "auxiliares_aplicacion",
    "thinners": "auxiliares_aplicacion",
    "ferreteria": "herrajes_seguridad",
    "herrajes": "herrajes_seguridad",
    "seguridad": "herrajes_seguridad",
    "herrajes_seguridad": "herrajes_seguridad",
    "herramientas": "herramientas_accesorios",
    "accesorios": "herramientas_accesorios",
    "herramientas_accesorios": "herramientas_accesorios",
}
PORTFOLIO_SEGMENT_QUERY_HINTS = {
    "auxiliares_aplicacion": [
        "ajustador", "thinner", "xilol", "varsol", "solvente", "diluyente", "catalizador",
        "endurecedor", "limpieza", "desengrase", "removedor",
    ],
    "herrajes_seguridad": [
        "cerradura", "cerraduras", "candado", "candados", "bisagra", "bisagras", "cerrojo",
        "picaporte", "manija", "pomo", "cierrapuerta", "barra antipanico", "barra antipánico",
        "llave", "cilindro", "yale",
    ],
    "herramientas_accesorios": [
        "brocha", "brochas", "rodillo", "rodillos", "lija", "lijas", "disco flap", "grata",
        "espatula", "espátula", "llana", "pistola", "felpa", "abrasiv",
    ],
    "recubrimientos_pinturas": [
        "pintura", "esmalte", "vinilo", "barniz", "laca", "sellador", "estuco", "impermeabil",
        "anticorros", "epox", "epóx", "poliuret", "corrotec", "aquablock", "koraza", "viniltex",
        "interseal", "intergard", "interthane", "interchar", "interzone", "pintulux", "primer",
        "imprimante", "fondo", "fachada", "humedad", "madera", "piso",
    ],
}


def _normalize_portfolio_segment(value: str | None) -> str | None:
    normalized = normalize_text_value(value or "")
    if not normalized:
        return None
    return PORTFOLIO_SEGMENT_ALIASES.get(normalized)


def _infer_portfolio_segments_for_query(pregunta: str, producto: str = "", explicit_segment: str | None = None) -> list[str]:
    normalized_explicit = _normalize_portfolio_segment(explicit_segment)
    if normalized_explicit:
        return [normalized_explicit]

    combined = normalize_text_value(f"{producto} {pregunta}")
    if not combined:
        return []

    detected = []
    for segment, tokens in PORTFOLIO_SEGMENT_QUERY_HINTS.items():
        if any(token in combined for token in tokens):
            detected.append(segment)

    if len(detected) > 1 and "recubrimientos_pinturas" in detected and "auxiliares_aplicacion" in detected and producto:
        product_hint = normalize_text_value(producto)
        if any(token in product_hint for token in PORTFOLIO_SEGMENT_QUERY_HINTS["auxiliares_aplicacion"]):
            return ["auxiliares_aplicacion"]
        return [segment for segment in detected if segment != "auxiliares_aplicacion"]
    return detected


def search_technical_chunks(query: str, top_k: int = 5, marca_filter: str | None = None,
                            segment_filters: list[str] | None = None) -> list[dict]:
    """Semantic search over vectorized technical sheet chunks using pgvector cosine distance."""
    embedding = _generate_query_embedding(query)
    if not embedding:
        return []

    embedding_literal = "[" + ",".join(str(v) for v in embedding) + "]"

    where_clauses = [
        "tipo_documento IN ('ficha_tecnica', 'ficha_tecnica_experto')",
        "COALESCE(metadata ->> 'document_scope', 'primary') = 'primary'",
        "COALESCE(metadata ->> 'quality_tier', 'primary') <> 'rejected'",
    ]
    params: list = [embedding_literal]
    if marca_filter:
        where_clauses.append("LOWER(marca) = LOWER(%s)")
        params.append(marca_filter)
    if segment_filters:
        where_clauses.append("COALESCE(metadata ->> 'portfolio_segment', 'portafolio_general') = ANY(%s)")
        params.append(segment_filters)
    params.extend([embedding_literal, top_k])

    try:
        engine = get_db_engine()
        raw_conn = engine.raw_connection()
        try:
            cur = raw_conn.cursor()
            cur.execute(
                f"""
                  SELECT doc_filename, doc_path_lower, chunk_index, chunk_text,
                      metadata,
                       marca, familia_producto, tipo_documento,
                       1 - (embedding <=> %s::vector) AS similarity
                                FROM public.agent_technical_doc_chunk
                                    WHERE {' AND '.join(where_clauses)}
                ORDER BY embedding <=> %s::vector
                LIMIT %s
                """,
                params,
            )
            columns = [desc[0] for desc in cur.description]
            rows = [dict(zip(columns, row)) for row in cur.fetchall()]
            return rows
        finally:
            raw_conn.close()
    except Exception:
        return []


def search_supporting_technical_guides(query: str, top_k: int = 3, marca_filter: str | None = None,
                                      segment_filters: list[str] | None = None) -> list[dict]:
    embedding = _generate_query_embedding(query)
    if not embedding:
        return []

    embedding_literal = "[" + ",".join(str(v) for v in embedding) + "]"
    where_clauses = [
        "tipo_documento = 'guia_solucion'",
        "COALESCE(metadata ->> 'document_scope', 'guide') = 'guide'",
    ]
    params: list = [embedding_literal]
    if marca_filter:
        where_clauses.append("LOWER(marca) = LOWER(%s)")
        params.append(marca_filter)
    if segment_filters:
        where_clauses.append("COALESCE(metadata ->> 'portfolio_segment', 'portafolio_general') = ANY(%s)")
        params.append(segment_filters)
    params.extend([embedding_literal, top_k])

    try:
        engine = get_db_engine()
        raw_conn = engine.raw_connection()
        try:
            cur = raw_conn.cursor()
            cur.execute(
                f"""
                  SELECT doc_filename, doc_path_lower, chunk_index, chunk_text,
                      metadata,
                       marca, familia_producto, tipo_documento,
                       1 - (embedding <=> %s::vector) AS similarity
                                FROM public.agent_technical_doc_chunk
                                    WHERE {' AND '.join(where_clauses)}
                ORDER BY embedding <=> %s::vector
                LIMIT %s
                """,
                params,
            )
            columns = [desc[0] for desc in cur.description]
            return [dict(zip(columns, row)) for row in cur.fetchall()]
        finally:
            raw_conn.close()
    except Exception:
        return []


def fetch_technical_profiles(canonical_families: list[str], source_files: list[str] | None = None,
                             limit: int = 3, segment_filters: list[str] | None = None) -> list[dict]:
    families = [family for family in canonical_families if family]
    files = [name for name in (source_files or []) if name]
    if not families and not files:
        return []

    try:
        engine = get_db_engine()
        raw_conn = engine.raw_connection()
        try:
            cur = raw_conn.cursor()
            clauses = []
            params: list = []
            if families:
                clauses.append("canonical_family = ANY(%s)")
                params.append(families)
            if files:
                clauses.append("source_doc_filename = ANY(%s)")
                params.append(files)
            segment_clause = ""
            if segment_filters:
                segment_clause = "AND COALESCE(profile_json -> 'product_identity' ->> 'portfolio_segment', 'portafolio_general') = ANY(%s)"
                params.append(segment_filters)
            cur.execute(
                f"""
                SELECT canonical_family, source_doc_filename, source_doc_path_lower,
                       marca, tipo_documento, completeness_score, extraction_status, profile_json
                FROM public.agent_technical_profile
                WHERE extraction_status = 'ready'
                  AND ({' OR '.join(clauses)})
                  {segment_clause}
                ORDER BY completeness_score DESC, canonical_family
                LIMIT %s
                """,
                [*params, limit],
            )
            columns = [desc[0] for desc in cur.description]
            rows = []
            for row in cur.fetchall():
                item = dict(zip(columns, row))
                profile_json = item.get("profile_json")
                if isinstance(profile_json, str):
                    try:
                        item["profile_json"] = json.loads(profile_json)
                    except Exception:
                        item["profile_json"] = None
                rows.append(item)
            return rows
        finally:
            raw_conn.close()
    except Exception:
        return []


def build_rag_context(chunks: list[dict], max_chunks: int = 4) -> str:
    """Build a textual context from RAG chunks for injection into the agent prompt.

    Skips FDS/HDS (safety data sheet) chunks entirely — they contain chemical
    hazard classifications and transport regulations that add noise and zero
    value for product recommendation.  Only FT (ficha técnica) content is
    useful for advising customers.
    """
    if not chunks:
        return ""
    parts = []
    seen_files = set()
    seen_signatures = set()
    for chunk in chunks[:max_chunks + 4]:  # read more to compensate for FDS skips
        if len(parts) >= max_chunks:
            break
        similarity = chunk.get("similarity", 0)
        if similarity < 0.25:
            continue
        filename = chunk.get("doc_filename", "desconocido")
        # Skip FDS/HDS safety data sheets — no recommendation value
        fn_upper = (filename or "").upper()
        if fn_upper.startswith("FDS") or fn_upper.startswith("HDS"):
            continue
        text_content = (chunk.get("chunk_text") or "").strip()
        if not text_content:
            continue
        metadata = chunk.get("metadata") or {}
        canonical_family = metadata.get("canonical_family") or chunk.get("familia_producto") or filename
        section_match = re.search(r"\[SECCIÓN:\s*([^\]]+)\]", text_content)
        section_name = (section_match.group(1).strip().lower() if section_match else "general")
        signature = f"{canonical_family}|{section_name}"
        if signature in seen_signatures:
            continue
        header = f"[Fuente: {filename}]"
        parts.append(f"{header}\n{text_content}")
        seen_files.add(filename)
        seen_signatures.add(signature)
    if not parts:
        return ""
    return "\n\n---\n\n".join(parts)


def learn_product_resolution(conversation_id: Optional[int], product_request: Optional[dict], product_context: list[dict], conversation_context: Optional[dict] = None):
    if not product_request or not product_context:
        return

    reliable_rows = select_reliable_learning_rows(product_request, product_context)
    if not reliable_rows:
        return

    # --- Anti-tambor filter: never learn absurd presentations ---
    BANNED_LEARNING_TOKENS = ["tambor", "50 galones", "55 galones", "200 litros"]
    original_query_lower = (product_request.get("original_query") or "").lower()
    filtered_rows = []
    for row in reliable_rows:
        desc_lower = ((row.get("descripcion") or row.get("nombre_articulo")) or "").lower()
        if any(token in desc_lower for token in BANNED_LEARNING_TOKENS):
            if not any(token in original_query_lower for token in BANNED_LEARNING_TOKENS):
                continue
        filtered_rows.append(row)
    reliable_rows = filtered_rows
    if not reliable_rows:
        return

    phrases = []
    previous_product_request = (conversation_context or {}).get("last_product_request") or {}
    phrase_groups = [build_learning_phrase_candidates(product_request)]
    if should_merge_previous_learning_phrase(product_request, previous_product_request):
        phrase_groups.append(build_learning_phrase_candidates(previous_product_request))
    for candidate_phrase in [phrase for group in phrase_groups for phrase in group]:
        if candidate_phrase not in phrases:
            phrases.append(candidate_phrase)

    if not phrases:
        return

    ensure_product_learning_table()
    engine = get_db_engine()
    with engine.begin() as connection:
        for phrase in phrases[:6]:
            for row in reliable_rows:
                reference_value = row.get("referencia") or row.get("codigo_articulo")
                if not reference_value:
                    continue
                canonical_presentation = None
                description_value = normalize_text_value(row.get("descripcion") or row.get("nombre_articulo"))
                for size_token, unit_name in PRESENTATION_SIZE_MAP.items():
                    if size_token in description_value:
                        canonical_presentation = unit_name
                        break
                connection.execute(
                    text(
                        """
                        INSERT INTO public.agent_product_learning (
                            normalized_phrase,
                            raw_phrase,
                            canonical_reference,
                            canonical_description,
                            canonical_brand,
                            canonical_presentation,
                            source_conversation_id,
                            source_message,
                            confidence,
                            usage_count,
                            created_at,
                            updated_at
                        )
                        VALUES (
                            :normalized_phrase,
                            :raw_phrase,
                            :canonical_reference,
                            :canonical_description,
                            :canonical_brand,
                            :canonical_presentation,
                            :source_conversation_id,
                            :source_message,
                            :confidence,
                            1,
                            now(),
                            now()
                        )
                        ON CONFLICT (normalized_phrase, canonical_reference)
                        DO UPDATE SET
                            canonical_description = COALESCE(EXCLUDED.canonical_description, public.agent_product_learning.canonical_description),
                            canonical_brand = COALESCE(EXCLUDED.canonical_brand, public.agent_product_learning.canonical_brand),
                            canonical_presentation = COALESCE(EXCLUDED.canonical_presentation, public.agent_product_learning.canonical_presentation),
                            source_conversation_id = COALESCE(EXCLUDED.source_conversation_id, public.agent_product_learning.source_conversation_id),
                            source_message = COALESCE(EXCLUDED.source_message, public.agent_product_learning.source_message),
                            confidence = GREATEST(public.agent_product_learning.confidence, EXCLUDED.confidence),
                            usage_count = public.agent_product_learning.usage_count + 1,
                            updated_at = now()
                        """
                    ),
                    {
                        "normalized_phrase": phrase,
                        "raw_phrase": phrase,
                        "canonical_reference": str(reference_value),
                        "canonical_description": row.get("descripcion") or row.get("nombre_articulo"),
                        "canonical_brand": row.get("marca") or row.get("marca_producto"),
                        "canonical_presentation": canonical_presentation,
                        "source_conversation_id": conversation_id,
                        "source_message": product_request.get("original_query"),
                        "confidence": 0.95 if product_request.get("product_codes") else 0.82,
                    },
                )


def fetch_learned_product_references(product_request: Optional[dict]):
    if not product_request:
        return []

    phrases = build_learning_phrase_candidates(product_request)

    # Also search learning table by product codes (P-53, 17174, etc.)
    for code in (product_request.get("product_codes") or []):
        normalized_code = normalize_text_value(str(code))
        if normalized_code and normalized_code not in phrases:
            phrases.insert(0, normalized_code)

    if not phrases:
        return []

    ensure_product_learning_table()
    engine = get_db_engine()
    learned_rows = []
    with engine.connect() as connection:
        for index, phrase in enumerate(phrases[:4]):
            row_set = connection.execute(
                text(
                    """
                    SELECT canonical_reference, canonical_description, canonical_brand, canonical_presentation,
                           MAX(confidence) AS confidence, SUM(usage_count) AS total_hits
                    FROM public.agent_product_learning
                    WHERE normalized_phrase = :normalized_phrase
                    GROUP BY canonical_reference, canonical_description, canonical_brand, canonical_presentation
                    ORDER BY MAX(confidence) DESC, SUM(usage_count) DESC
                    LIMIT 5
                    """
                ),
                {"normalized_phrase": phrase},
            ).mappings().all()
            learned_rows.extend(row for row in row_set if is_learned_reference_relevant(product_request, row))

    ordered_references = []
    seen_references = set()
    for row in learned_rows:
        reference_value = row.get("canonical_reference")
        if reference_value and reference_value not in seen_references:
            seen_references.add(reference_value)
            ordered_references.append(reference_value)
    return ordered_references[:5]


def extract_product_codes(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return []

    excluded_codes = {"3en1", "p11", "t11", "p53", "1gl", "5gl"}
    codes = []
    seen_codes = set()

    raw_candidates = re.findall(r"\b[a-z]?\d[a-z0-9-]{1,14}\b|\b\d{4,10}\b|\b[a-z0-9-]{4,16}\b", normalized)
    for prefix, suffix in re.findall(r"\b(sku|ref|referencia|cod|codigo)\s*[-/]?\s*(\d{3,10}[a-z0-9-]{0,8})\b", normalized):
        raw_candidates.append(f"{prefix}{suffix}")

    for raw_code in raw_candidates:
        cleaned_code = normalize_reference_value(raw_code)
        if len(cleaned_code) < 3 or cleaned_code in seen_codes or cleaned_code in excluded_codes:
            continue
        has_letters = bool(re.search(r"[a-z]", cleaned_code))
        has_digits = bool(re.search(r"\d", cleaned_code))
        if not re.fullmatch(r"\d{4,10}", cleaned_code) and not (has_letters and has_digits):
            continue
        if re.fullmatch(r"\d{1,3}mm", cleaned_code):
            continue
        seen_codes.add(cleaned_code)
        codes.append(cleaned_code)
    return codes


def is_product_code_message(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    if re.fullmatch(r"[a-z]?\d[a-z0-9-]{1,14}", normalized):
        return True
    return bool(re.fullmatch(r"\d{4,10}", normalized))


def extract_store_filters(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return []

    matched_codes = []
    seen_codes = set()
    for store_aliases in STORE_ALIASES.values():
        for alias in store_aliases:
            alias_normalized = normalize_text_value(alias)
            if not alias_normalized:
                continue
            if alias_normalized.isdigit():
                matched = bool(re.search(rf"\b{re.escape(alias_normalized)}\b", normalized))
            else:
                matched = bool(re.search(rf"\b{re.escape(alias_normalized)}\b", normalized))
            if matched:
                code = next((candidate for candidate in store_aliases if candidate.isdigit()), None)
                if code and code not in seen_codes:
                    seen_codes.add(code)
                    matched_codes.append(code)
                break
    return matched_codes


def is_store_alias_term(term_value: Optional[str]):
    normalized = normalize_text_value(term_value)
    if not normalized:
        return False
    for store_aliases in STORE_ALIASES.values():
        for alias in store_aliases:
            if normalize_text_value(alias) == normalized:
                return True
    return False


def extract_brand_filters(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return []

    matched_brands = []
    for brand_name, aliases in BRAND_ALIASES.items():
        for alias in aliases:
            alias_normalized = normalize_text_value(alias)
            if alias_normalized and re.search(rf"\b{re.escape(alias_normalized)}\b", normalized):
                if brand_name not in matched_brands:
                    matched_brands.append(brand_name)
                break
    return matched_brands


def extract_direction_filters(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return []

    matched_directions = []
    for direction_name, aliases in DIRECTION_ALIASES.items():
        for alias in aliases:
            alias_normalized = normalize_text_value(alias)
            if alias_normalized and re.search(rf"\b{re.escape(alias_normalized)}\b", normalized):
                matched_directions.append(direction_name)
                break
    return matched_directions


def infer_product_presentation_from_row(product_row: dict):
    explicit_presentation = canonicalize_presentation_value(product_row.get("presentacion_canonica"))
    if explicit_presentation:
        return explicit_presentation
    description_value = normalize_text_value(product_row.get("descripcion") or product_row.get("nombre_articulo"))
    for canonical_value, aliases in PRESENTATION_ALIASES.items():
        if any(normalize_text_value(alias) and normalize_text_value(alias) in description_value for alias in aliases):
            return canonical_value
    for size_token, unit_name in PRESENTATION_SIZE_MAP.items():
        if size_token in description_value:
            return unit_name
    return None


def extract_size_filters(text_value: Optional[str]):
    if not text_value:
        return []

    # Normalize fractional sizes: "11/2" → "1 1/2", "21/2" → "2 1/2"
    normalized_text = re.sub(r'\b([1-4])1/2', r'\1 1/2', text_value)

    normalized_sizes = []
    seen_sizes = set()
    # Match sizes followed by " or pulgada
    for raw_match in re.findall(r"\b(\d+(?:\s+\d/\d)?)(?=\s*(?:\"|''|pulgadas?|pulg))", normalized_text, flags=re.IGNORECASE):
        size_value = re.sub(r"\s+", " ", raw_match.strip())
        if size_value and size_value not in seen_sizes:
            seen_sizes.add(size_value)
            normalized_sizes.append(size_value)
    # Also match bare fractions like "1/2", "1 1/2", "2 1/2" even without trailing quote
    for raw_match in re.findall(r"\b(\d+\s+\d/\d)\b", normalized_text):
        size_value = re.sub(r"\s+", " ", raw_match.strip())
        if size_value and size_value not in seen_sizes:
            seen_sizes.add(size_value)
            normalized_sizes.append(size_value)
    return normalized_sizes


def infer_product_size_from_row(product_row: dict):
    raw_description = str(product_row.get("descripcion") or product_row.get("nombre_articulo") or "")
    # Normalize ERP fractional sizes: "11/2" → "1 1/2", "21/2" → "2 1/2"
    norm_desc = re.sub(r'\b([1-4])1/2\b', r'\1 1/2', raw_description)
    size_match = re.search(r"\b(\d+(?:/\d+)?(?:\s+\d/\d+)?)(?=\")", norm_desc)
    if size_match:
        return re.sub(r"\s+", " ", size_match.group(1).strip())

    normalized_description = normalize_text_value(norm_desc)
    for size_value in ["2 1/2", "1 1/2", "3 1/2", "4 1/2", "1/2"]:
        if size_value in normalized_description:
            return size_value
    standalone_match = re.search(r"\b(\d+(?:\s+\d/\d)?)\b", normalized_description)
    if standalone_match and any(keyword in normalized_description for keyword in ["brocha", "rodillo", "cerradura", "bisagra", "pasador", "portacandado"]):
        return standalone_match.group(1)
    return None


def infer_product_direction_from_row(product_row: dict):
    description_value = normalize_text_value(product_row.get("descripcion") or product_row.get("nombre_articulo"))
    for direction_name, aliases in DIRECTION_ALIASES.items():
        if any(normalize_text_value(alias) in description_value for alias in aliases):
            return direction_name
    return None


def infer_product_color_from_row(product_row: dict):
    color_value = normalize_nullable_phrase(product_row.get("color_detectado") or product_row.get("color_raiz"))
    if color_value:
        return color_value
    description_value = normalize_text_value(product_row.get("descripcion") or product_row.get("nombre_articulo"))
    for compound_color in ["verde bronce", "blanco puro", "rojo fiesta", "verde esmeralda"]:
        if compound_color in description_value:
            return compound_color
    for candidate_color in ["blanco", "negro", "gris", "rojo", "verde", "azul", "amarillo", "naranja", "marfil", "crema", "bronce", "transparente"]:
        if re.search(rf"\b{re.escape(candidate_color)}\b", description_value):
            return candidate_color
    return None


def infer_product_finish_from_row(product_row: dict):
    finish_value = normalize_nullable_phrase(product_row.get("acabado_detectado"))
    if finish_value:
        return finish_value
    description_value = normalize_text_value(product_row.get("descripcion") or product_row.get("nombre_articulo"))
    for candidate_finish in ["mate", "brillante", "satinado", "semibrillante", "semimate", "texturizado"]:
        if re.search(rf"\b{re.escape(candidate_finish)}\b", description_value):
            return candidate_finish
    return None


def row_matches_requested_colors(product_row: dict, requested_colors: list[str]):
    if not requested_colors:
        return False

    inferred_color = infer_product_color_from_row(product_row)
    description_value = normalize_text_value(product_row.get("descripcion") or product_row.get("nombre_articulo"))
    description_tokens = tokenize_search_phrase(description_value)
    for color_value in requested_colors:
        normalized_color = normalize_text_value(color_value)
        if not normalized_color:
            continue
        if inferred_color == normalized_color or normalized_color in description_value:
            return True
        color_tokens = tokenize_search_phrase(normalized_color)
        if color_tokens and all(
            any(
                desc_token.startswith(color_token[:4]) or color_token.startswith(desc_token[:4])
                for desc_token in description_tokens
            )
            for color_token in color_tokens
        ):
            return True
    return False


def infer_product_brand_from_row(product_row: dict):
    brand_text = normalize_text_value(product_row.get("marca") or product_row.get("marca_producto") or "")
    if re.fullmatch(r"\d+", brand_text or ""):
        brand_text = ""
    description_value = normalize_text_value(product_row.get("descripcion") or product_row.get("nombre_articulo"))
    combined_value = f"{brand_text} {description_value}".strip()
    for brand_name, aliases in BRAND_ALIASES.items():
        if any(alias and normalize_text_value(alias) in combined_value for alias in aliases):
            return brand_name
    return brand_text or None


def summarize_product_option(product_row: dict):
    reference_value = product_row.get("referencia") or product_row.get("codigo_articulo") or "sin referencia"
    description_value = product_row.get("descripcion") or product_row.get("nombre_articulo") or reference_value
    stock_value = product_row.get("stock_total") if product_row.get("stock_total") is not None else product_row.get("stock")
    presentation_value = infer_product_presentation_from_row(product_row)
    brand_value = infer_product_brand_from_row(product_row)
    department_value = product_row.get("departamentos") or product_row.get("categoria_producto")
    summary_parts = [description_value]
    if presentation_value:
        summary_parts.append(get_presentation_label(presentation_value, 1))
    if brand_value:
        summary_parts.append(str(brand_value).upper())
    if department_value and str(department_value).strip().upper() != "NULL":
        summary_parts.append(str(department_value))
    if stock_value is not None:
        summary_parts.append(f"stock {format_quantity(stock_value)}")
    return f"{reference_value}: {' | '.join(summary_parts)}"


def get_product_variant_signature(product_row: Optional[dict]):
    row = product_row or {}
    family_value = normalize_text_value(row.get("producto_padre_busqueda") or row.get("familia_consulta"))
    if family_value:
        return family_value

    raw_description = get_exact_product_description(row)
    cleaned_description = re.sub(r"^\s*(?:PQ|IQ|EQ|SQ|MEG)\s+", "", raw_description, flags=re.IGNORECASE)
    cleaned_description = re.sub(r"\s+\d+(?:\.\d+)?L\b", "", cleaned_description, flags=re.IGNORECASE)
    cleaned_description = re.sub(r"\s+", " ", cleaned_description).strip()
    return normalize_text_value(cleaned_description)


def get_product_variant_label(product_row: Optional[dict]):
    row = product_row or {}
    family_value = (row.get("producto_padre_busqueda") or row.get("familia_consulta") or "").strip()
    if family_value:
        return re.sub(r"\s+", " ", family_value)

    raw_description = get_exact_product_description(row)
    cleaned_description = re.sub(r"^\s*(?:PQ|IQ|EQ|SQ|MEG)\s+", "", raw_description, flags=re.IGNORECASE)
    cleaned_description = re.sub(r"\s+\d+(?:\.\d+)?L\b", "", cleaned_description, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", cleaned_description).strip()


def should_ask_product_clarification(product_request: Optional[dict], product_context: list[dict]):
    if not product_request or not product_context:
        return False

    top_candidates = product_context[:4]
    unique_references = {row.get("referencia") or row.get("codigo_articulo") for row in top_candidates if row.get("referencia") or row.get("codigo_articulo")}
    if len(unique_references) < 2:
        return False

    variant_signatures = {get_product_variant_signature(row) for row in top_candidates}
    variant_signatures.discard(None)
    variant_signatures.discard("")

    if product_request.get("product_codes") and len(variant_signatures) >= 2:
        return True

    presentation_values = {infer_product_presentation_from_row(row) for row in top_candidates}
    brand_values = {infer_product_brand_from_row(row) for row in top_candidates}
    direction_values = {infer_product_direction_from_row(row) for row in top_candidates}
    color_values = {infer_product_color_from_row(row) for row in top_candidates}
    finish_values = {infer_product_finish_from_row(row) for row in top_candidates}
    presentation_values.discard(None)
    brand_values.discard(None)
    direction_values.discard(None)
    color_values.discard(None)
    finish_values.discard(None)

    if not product_request.get("requested_unit") and len(presentation_values) >= 2:
        return True
    if len(brand_values) >= 2:
        return True
    if not (product_request.get("direction_filters") or []) and len(direction_values) >= 2:
        return True
    if not (product_request.get("color_filters") or []) and len(color_values) >= 2:
        return True
    if not (product_request.get("finish_filters") or []) and len(finish_values) >= 2:
        return True
    return False


def build_best_product_clarification_question(product_request: Optional[dict], product_context: list[dict]):
    top_candidates = product_context[:4]
    for row in top_candidates:
        question = (row.get("pregunta_desambiguacion") or "").strip()
        if question:
            return question

    variant_labels = []
    seen_variant_labels = set()
    for row in top_candidates:
        variant_label = get_product_variant_label(row)
        variant_signature = get_product_variant_signature(row)
        if not variant_label or not variant_signature or variant_signature in seen_variant_labels:
            continue
        seen_variant_labels.add(variant_signature)
        variant_labels.append(variant_label)

    presentation_values = {infer_product_presentation_from_row(row) for row in top_candidates}
    brand_values = {infer_product_brand_from_row(row) for row in top_candidates}
    color_values = {infer_product_color_from_row(row) for row in top_candidates}
    finish_values = {infer_product_finish_from_row(row) for row in top_candidates}
    presentation_values.discard(None)
    brand_values.discard(None)
    color_values.discard(None)
    finish_values.discard(None)

    if (product_request or {}).get("product_codes") and len(variant_labels) >= 2:
        options_text = ", ".join(variant_labels[:3])
        return f"Encontré ese código con varias descripciones exactas: {options_text}. ¿Cuál necesitas exactamente?"

    if not (product_request or {}).get("requested_unit") and len(presentation_values) >= 2:
        return "¿Lo necesitas en cuarto, galón o cuñete?"
    if not (product_request or {}).get("brand_filters") and len(brand_values) >= 2:
        brand_options = ", ".join(sorted(str(value).title() for value in brand_values))
        return f"Tengo opciones de {brand_options}. ¿Cuál marca buscas?"
    if not (product_request or {}).get("finish_filters") and len(finish_values) >= 2:
        finish_options = " y ".join(sorted(str(value) for value in finish_values))
        return f"Tengo esa línea en acabado {finish_options}. ¿Cuál necesitas?"
    if not (product_request or {}).get("color_filters") and len(color_values) >= 2:
        return "Tengo varias opciones de color en esa línea. ¿Cuál color necesitas exactamente?"
    return "Tengo varias opciones cercanas. ¿Cuál necesitas exactamente?"


def filter_rows_by_requested_presentation(product_rows: list[dict], product_request: Optional[dict]):
    if not product_request or not product_request.get("requested_unit"):
        return product_rows
    exact_rows = [row for row in product_rows if infer_product_presentation_from_row(row) == product_request.get("requested_unit")]
    return exact_rows or product_rows


def filter_rows_by_requested_size(product_rows: list[dict], product_request: Optional[dict]):
    requested_sizes = (product_request or {}).get("size_filters") or []
    if not requested_sizes:
        return product_rows
    exact_rows = [row for row in product_rows if infer_product_size_from_row(row) in requested_sizes]
    return exact_rows or product_rows


def resolve_product_clarification_choice(text_value: Optional[str], clarification_options: list[dict]):
    normalized = normalize_text_value(text_value)
    if not normalized or not clarification_options:
        return None

    ordinal_map = {
        "1": 0,
        "uno": 0,
        "primera": 0,
        "primer": 0,
        "primero": 0,
        "2": 1,
        "dos": 1,
        "segunda": 1,
        "segundo": 1,
        "3": 2,
        "tres": 2,
        "tercera": 2,
        "tercero": 2,
        "4": 3,
        "cuatro": 3,
        "cuarta": 3,
        "cuarto": 3,
    }
    if normalized in ordinal_map and ordinal_map[normalized] < len(clarification_options):
        return clarification_options[ordinal_map[normalized]]

    for option in clarification_options:
        reference_value = normalize_reference_value(option.get("referencia") or option.get("codigo_articulo"))
        if reference_value and reference_value in normalize_reference_value(normalized):
            return option
    return None


def expand_product_terms(search_terms: list[str]):
    expanded_terms = []
    seen_terms = set()

    def add_term(raw_term: Optional[str]):
        normalized_term = normalize_text_value(raw_term)
        if not normalized_term or normalized_term in seen_terms:
            return
        seen_terms.add(normalized_term)
        expanded_terms.append(normalized_term)

    for term in search_terms:
        add_term(term)
        normalized_key = normalize_reference_value(term)
        if normalized_key in PORTFOLIO_ALIASES:
            for alias_term in PORTFOLIO_ALIASES[normalized_key]:
                add_term(alias_term)
        elif len(normalized_key) >= 4:
            best_ratio = 0.0
            best_aliases = None
            for alias_key, aliases in PORTFOLIO_ALIASES.items():
                if len(alias_key) < 4:
                    continue
                ratio = SequenceMatcher(None, normalized_key, alias_key).ratio()
                if ratio >= 0.75 and ratio > best_ratio:
                    best_ratio = ratio
                    best_aliases = aliases
            if best_aliases:
                for alias_term in best_aliases:
                    add_term(alias_term)

    return expanded_terms


def get_specific_product_terms(product_request: Optional[dict]):
    if not product_request:
        return []

    generic_terms = set()
    for aliases in PRESENTATION_ALIASES.values():
        generic_terms.update(normalize_text_value(alias) for alias in aliases)

    specific_terms = []
    for term in product_request.get("core_terms") or []:
        normalized_term = normalize_text_value(term)
        if not normalized_term or normalized_term in generic_terms or normalized_term in PRODUCT_STOPWORDS:
            continue
        if normalized_term not in specific_terms:
            specific_terms.append(normalized_term)
        normalized_key = normalize_reference_value(term)
        if normalized_key in PORTFOLIO_ALIASES:
            for alias_term in PORTFOLIO_ALIASES[normalized_key]:
                normalized_alias = normalize_text_value(alias_term)
                if (
                    normalized_alias
                    and normalized_alias not in generic_terms
                    and normalized_alias not in PRODUCT_STOPWORDS
                    and len(normalized_alias) >= 4
                    and normalized_alias not in specific_terms
                ):
                    specific_terms.append(normalized_alias)
    return specific_terms[:5]


def normalize_phone(phone_number: Optional[str]):
    if not phone_number:
        return None
    digits = "".join(character for character in phone_number if character.isdigit())
    if not digits:
        return None
    return digits if digits.startswith("+") else f"+{digits}"


def ensure_contact_and_conversation(phone_number: str, profile_name: Optional[str]):
    engine = get_db_engine()
    normalized_phone = normalize_phone(phone_number)
    if not normalized_phone:
        raise RuntimeError("No fue posible normalizar el teléfono recibido.")

    with engine.begin() as connection:
        contact_row = connection.execute(
            text(
                """
                INSERT INTO public.whatsapp_contacto (telefono_e164, nombre_visible, ultima_interaccion_at, updated_at)
                VALUES (:telefono_e164, :nombre_visible, now(), now())
                ON CONFLICT (telefono_e164)
                DO UPDATE SET
                    nombre_visible = COALESCE(EXCLUDED.nombre_visible, public.whatsapp_contacto.nombre_visible),
                    ultima_interaccion_at = now(),
                    updated_at = now()
                RETURNING id, cliente_id, telefono_e164, nombre_visible
                """
            ),
            {"telefono_e164": normalized_phone, "nombre_visible": profile_name},
        ).mappings().one()

        conversation_row = connection.execute(
            text(
                """
                SELECT id, cliente_id
                FROM public.agent_conversation
                WHERE contacto_id = :contacto_id AND estado IN ('abierta', 'pendiente')
                ORDER BY updated_at DESC
                LIMIT 1
                """
            ),
            {"contacto_id": contact_row["id"]},
        ).mappings().one_or_none()

        if conversation_row is None:
            conversation_row = connection.execute(
                text(
                    """
                    INSERT INTO public.agent_conversation (contacto_id, cliente_id, canal, estado, started_at, last_message_at, updated_at)
                    VALUES (:contacto_id, :cliente_id, 'whatsapp', 'abierta', now(), now(), now())
                    RETURNING id, cliente_id
                    """
                ),
                {"contacto_id": contact_row["id"], "cliente_id": contact_row["cliente_id"]},
            ).mappings().one()
        else:
            connection.execute(
                text(
                    """
                    UPDATE public.agent_conversation
                    SET last_message_at = now(), updated_at = now()
                    WHERE id = :conversation_id
                    """
                ),
                {"conversation_id": conversation_row["id"]},
            )

    return {
        "contact_id": contact_row["id"],
        "cliente_id": contact_row["cliente_id"],
        "conversation_id": conversation_row["id"],
        "telefono_e164": contact_row["telefono_e164"],
        "nombre_visible": contact_row["nombre_visible"],
    }


def store_inbound_message(
    conversation_id: int,
    provider_message_id: Optional[str],
    message_type: str,
    content: Optional[str],
    payload: dict,
):
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO public.agent_message (
                    conversation_id,
                    provider_message_id,
                    direction,
                    message_type,
                    contenido,
                    payload,
                    estado,
                    created_at
                )
                VALUES (
                    :conversation_id,
                    :provider_message_id,
                    'inbound',
                    :message_type,
                    :contenido,
                    CAST(:payload AS jsonb),
                    'recibido',
                    now()
                )
                ON CONFLICT DO NOTHING
                """
            ),
            {
                "conversation_id": conversation_id,
                "provider_message_id": provider_message_id,
                "message_type": message_type,
                "contenido": content,
                "payload": safe_json_dumps(payload),
            },
        )


def inbound_message_already_processed(provider_message_id: Optional[str]):
    if not provider_message_id:
        return False

    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT id
                FROM public.agent_message
                WHERE provider_message_id = :provider_message_id
                  AND direction = 'inbound'
                LIMIT 1
                """
            ),
            {"provider_message_id": provider_message_id},
        ).mappings().one_or_none()
    return row is not None


def store_outbound_message(
    conversation_id: int,
    provider_message_id: Optional[str],
    message_type: str,
    content: Optional[str],
    payload: dict,
    intent_detectado: Optional[str] = None,
):
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO public.agent_message (
                    conversation_id,
                    provider_message_id,
                    direction,
                    message_type,
                    intent_detectado,
                    contenido,
                    payload,
                    estado,
                    created_at
                )
                VALUES (
                    :conversation_id,
                    :provider_message_id,
                    'outbound',
                    :message_type,
                    :intent_detectado,
                    :contenido,
                    CAST(:payload AS jsonb),
                    'respondido',
                    now()
                )
                """
            ),
            {
                "conversation_id": conversation_id,
                "provider_message_id": provider_message_id,
                "message_type": message_type,
                "intent_detectado": intent_detectado,
                "contenido": content,
                "payload": safe_json_dumps(payload),
            },
        )


def load_recent_conversation_messages(conversation_id: int, limit: int = 12):
    engine = get_db_engine()
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                """
                SELECT direction, message_type, contenido, created_at
                FROM public.agent_message
                WHERE conversation_id = :conversation_id
                ORDER BY created_at DESC
                LIMIT :limit
                """
            ),
            {"conversation_id": conversation_id, "limit": limit},
        ).mappings().all()
    return list(reversed(rows))


def get_conversation_snapshot(conversation_id: int):
    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT id, cliente_id, resumen, contexto, last_message_at
                FROM public.agent_conversation
                WHERE id = :conversation_id
                """
            ),
            {"conversation_id": conversation_id},
        ).mappings().one()
    return row


def find_cliente_contexto_by_phone(phone_number: str):
    normalized_digits = "".join(character for character in phone_number if character.isdigit())
    if normalized_digits.startswith("57"):
        normalized_digits = normalized_digits[2:]

    if not normalized_digits:
        return None

    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT cod_cliente, nombre_cliente, telefono1, telefono2, email
                FROM public.vw_estado_cartera
                WHERE regexp_replace(COALESCE(telefono1, ''), '[^0-9]', '', 'g') LIKE :phone_pattern
                   OR regexp_replace(COALESCE(telefono2, ''), '[^0-9]', '', 'g') LIKE :phone_pattern
                ORDER BY dias_vencido DESC NULLS LAST, fecha_documento DESC NULLS LAST
                LIMIT 1
                """
            ),
            {"phone_pattern": f"%{normalized_digits}"},
        ).mappings().one_or_none()

        if row is None:
            row = connection.execute(
                text(
                    """
                    SELECT codigo AS cod_cliente, nombre_legal AS nombre_cliente, telefono AS telefono1, celular AS telefono2, email
                    FROM public.cliente
                    WHERE regexp_replace(COALESCE(telefono, ''), '[^0-9]', '', 'g') LIKE :phone_pattern
                       OR regexp_replace(COALESCE(celular, ''), '[^0-9]', '', 'g') LIKE :phone_pattern
                    LIMIT 1
                    """
                ),
                {"phone_pattern": f"%{normalized_digits}"},
            ).mappings().one_or_none()

    if not row or not row["cod_cliente"]:
        return None

    try:
        return get_cliente_contexto(row["cod_cliente"])
    except HTTPException:
        return {
            "cliente_codigo": row["cod_cliente"],
            "nombre_cliente": row["nombre_cliente"],
        }


def find_cliente_contexto_by_document(document_number: str):
    normalized_document = re.sub(r"\D", "", document_number or "")
    if not normalized_document:
        return None

    engine = get_db_engine()
    with engine.connect() as connection:
        # Exact match first
        row = connection.execute(
            text(
                """
                SELECT cod_cliente, nombre_cliente, nit
                FROM public.vw_estado_cartera
                WHERE regexp_replace(COALESCE(nit, ''), '[^0-9]', '', 'g') = :document_number
                LIMIT 1
                """
            ),
            {"document_number": normalized_document},
        ).mappings().one_or_none()

        # Prefix match for NITs with verification digit (e.g. 1088266407 matches 10882664078)
        if row is None:
            row = connection.execute(
                text(
                    """
                    SELECT cod_cliente, nombre_cliente, nit
                    FROM public.vw_estado_cartera
                    WHERE regexp_replace(COALESCE(nit, ''), '[^0-9]', '', 'g') LIKE :document_prefix
                    ORDER BY fecha_documento DESC NULLS LAST
                    LIMIT 1
                    """
                ),
                {"document_prefix": f"{normalized_document}%"},
            ).mappings().one_or_none()

        if row is None:
            row = connection.execute(
                text(
                    """
                    SELECT codigo AS cod_cliente, nombre_legal AS nombre_cliente, numero_documento AS nit
                    FROM public.cliente
                    WHERE regexp_replace(COALESCE(numero_documento, ''), '[^0-9]', '', 'g') = :document_number
                       OR regexp_replace(COALESCE(numero_documento, ''), '[^0-9]', '', 'g') LIKE :document_prefix
                    LIMIT 1
                    """
                ),
                {"document_number": normalized_document, "document_prefix": f"{normalized_document}%"},
            ).mappings().one_or_none()

        # Also search in cod_cliente (some systems use cédula as client code)
        if row is None:
            row = connection.execute(
                text(
                    """
                    SELECT cod_cliente, nombre_cliente, nit
                    FROM public.vw_estado_cartera
                    WHERE regexp_replace(COALESCE(cod_cliente::text, ''), '[^0-9]', '', 'g') = :document_number
                    LIMIT 1
                    """
                ),
                {"document_number": normalized_document},
            ).mappings().one_or_none()

    if not row or not row["cod_cliente"]:
        return None

    try:
        contexto = get_cliente_contexto(row["cod_cliente"])
    except HTTPException:
        contexto = {
            "cliente_codigo": row["cod_cliente"],
            "nombre_cliente": row["nombre_cliente"],
        }

    contexto["verified_document"] = normalized_document
    return contexto


def find_cliente_contexto_in_sales(customer_code: Optional[str] = None, name_value: Optional[str] = None):
    engine = get_db_engine()

    if customer_code:
        normalized_code = normalize_reference_value(customer_code)
        if not normalized_code:
            return None
        with engine.connect() as connection:
            row = connection.execute(
                text(
                    """
                    SELECT cliente_id AS cod_cliente, nombre_cliente
                    FROM public.raw_ventas_detalle
                    WHERE regexp_replace(lower(COALESCE(cliente_id, '')), '[^a-z0-9]', '', 'g') = :customer_code
                    GROUP BY 1, 2
                    ORDER BY COUNT(*) DESC, nombre_cliente ASC
                    LIMIT 1
                    """
                ),
                {"customer_code": normalized_code},
            ).mappings().one_or_none()
        if not row:
            return None
        return {
            "cliente_codigo": row["cod_cliente"],
            "nombre_cliente": row["nombre_cliente"],
            "verified_cliente_codigo": row["cod_cliente"],
            "verified_source": "raw_sales",
        }

    normalized_name = normalize_text_value(name_value)
    tokens = [token for token in normalized_name.split() if len(token) >= 3]
    if len(tokens) < 2:
        return None

    primary_pattern = f"%{tokens[0]}%"
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                """
                SELECT cliente_id AS cod_cliente, nombre_cliente
                FROM public.raw_ventas_detalle
                WHERE regexp_replace(lower(COALESCE(nombre_cliente, '')), '[^a-z0-9 ]', '', 'g') ILIKE :pattern
                GROUP BY 1, 2
                ORDER BY COUNT(*) DESC, nombre_cliente ASC
                LIMIT 20
                """
            ),
            {"pattern": primary_pattern},
        ).mappings().all()

    candidates = []
    for row in rows:
        candidate_name = normalize_text_value(row.get("nombre_cliente"))
        if not candidate_name:
            continue
        token_hits = sum(1 for token in tokens if token in candidate_name)
        similarity = sequence_similarity(normalized_name, candidate_name)
        exact_phrase = 1 if normalized_name in candidate_name or candidate_name in normalized_name else 0
        if token_hits < max(2, len(tokens) - 1) and similarity < 0.82:
            continue
        candidates.append(
            {
                "cod_cliente": row["cod_cliente"],
                "nombre_cliente": row.get("nombre_cliente"),
                "score": token_hits * 2 + exact_phrase * 3 + similarity,
            }
        )

    if not candidates:
        return None

    best_match = sorted(candidates, key=lambda item: item["score"], reverse=True)[0]
    return {
        "cliente_codigo": best_match["cod_cliente"],
        "nombre_cliente": best_match.get("nombre_cliente"),
        "verified_cliente_codigo": best_match["cod_cliente"],
        "verified_source": "raw_sales",
    }


def find_cliente_contexto_by_customer_code(customer_code: str):
    normalized_code = normalize_reference_value(customer_code)
    if not normalized_code:
        return None

    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                """
                SELECT cod_cliente, nombre_cliente, nit
                FROM public.vw_estado_cartera
                WHERE regexp_replace(lower(COALESCE(cod_cliente::text, '')), '[^a-z0-9]', '', 'g') = :customer_code
                LIMIT 1
                """
            ),
            {"customer_code": normalized_code},
        ).mappings().one_or_none()

        if row is None:
            row = connection.execute(
                text(
                    """
                    SELECT codigo AS cod_cliente, nombre_legal AS nombre_cliente, numero_documento AS nit
                    FROM public.cliente
                    WHERE regexp_replace(lower(COALESCE(codigo::text, '')), '[^a-z0-9]', '', 'g') = :customer_code
                    LIMIT 1
                    """
                ),
                {"customer_code": normalized_code},
            ).mappings().one_or_none()

    if not row or not row["cod_cliente"]:
        return find_cliente_contexto_in_sales(customer_code=normalized_code)

    try:
        contexto = get_cliente_contexto(row["cod_cliente"])
    except HTTPException:
        contexto = {
            "cliente_codigo": row["cod_cliente"],
            "nombre_cliente": row["nombre_cliente"],
        }

    contexto["verified_cliente_codigo"] = row["cod_cliente"]
    return contexto


def find_cliente_contexto_by_name(name_value: str):
    normalized_name = normalize_text_value(name_value)
    tokens = [token for token in normalized_name.split() if len(token) >= 3]
    if len(tokens) < 2:
        return None

    primary_pattern = f"%{tokens[0]}%"
    engine = get_db_engine()
    candidates = []
    with engine.connect() as connection:
        cartera_rows = connection.execute(
            text(
                """
                SELECT cod_cliente, nombre_cliente, nit
                FROM public.vw_estado_cartera
                WHERE regexp_replace(lower(COALESCE(nombre_cliente, '')), '[^a-z0-9 ]', '', 'g') ILIKE :pattern
                ORDER BY fecha_documento DESC NULLS LAST
                LIMIT 20
                """
            ),
            {"pattern": primary_pattern},
        ).mappings().all()

        client_rows = connection.execute(
            text(
                """
                SELECT codigo AS cod_cliente, nombre_legal AS nombre_cliente, numero_documento AS nit
                FROM public.cliente
                WHERE regexp_replace(lower(COALESCE(nombre_legal, '')), '[^a-z0-9 ]', '', 'g') ILIKE :pattern
                LIMIT 20
                """
            ),
            {"pattern": primary_pattern},
        ).mappings().all()

    seen_codes = set()
    for row in list(cartera_rows) + list(client_rows):
        customer_code = row.get("cod_cliente")
        if not customer_code or customer_code in seen_codes:
            continue
        seen_codes.add(customer_code)

        candidate_name = normalize_text_value(row.get("nombre_cliente"))
        if not candidate_name:
            continue

        token_hits = sum(1 for token in tokens if token in candidate_name)
        similarity = sequence_similarity(normalized_name, candidate_name)
        exact_phrase = 1 if normalized_name in candidate_name or candidate_name in normalized_name else 0
        if token_hits < max(2, len(tokens) - 1) and similarity < 0.82:
            continue

        candidates.append(
            {
                "cod_cliente": customer_code,
                "nombre_cliente": row.get("nombre_cliente"),
                "score": token_hits * 2 + exact_phrase * 3 + similarity,
            }
        )

    if not candidates:
        return find_cliente_contexto_in_sales(name_value=name_value)

    best_match = sorted(candidates, key=lambda item: item["score"], reverse=True)[0]
    try:
        contexto = get_cliente_contexto(best_match["cod_cliente"])
    except HTTPException:
        contexto = {
            "cliente_codigo": best_match["cod_cliente"],
            "nombre_cliente": best_match.get("nombre_cliente"),
        }

    contexto["verified_cliente_codigo"] = best_match["cod_cliente"]
    return contexto


def update_contact_cliente(contact_id: int, cliente_codigo: Optional[str]):
    if not cliente_codigo:
        return

    engine = get_db_engine()
    with engine.begin() as connection:
        cliente_row = connection.execute(
            text(
                """
                SELECT id
                FROM public.cliente
                WHERE codigo = :codigo
                LIMIT 1
                """
            ),
            {"codigo": cliente_codigo},
        ).mappings().one_or_none()

        if not cliente_row:
            return

        connection.execute(
            text(
                """
                UPDATE public.whatsapp_contacto
                SET cliente_id = :cliente_id, updated_at = now()
                WHERE id = :contact_id
                """
            ),
            {"cliente_id": cliente_row["id"], "contact_id": contact_id},
        )

        connection.execute(
            text(
                """
                UPDATE public.agent_conversation
                SET cliente_id = :cliente_id, updated_at = now()
                WHERE contacto_id = :contact_id AND estado IN ('abierta', 'pendiente')
                """
            ),
            {"cliente_id": cliente_row["id"], "contact_id": contact_id},
        )

    return cliente_row["id"]


def update_conversation_context(conversation_id: int, context_updates: dict, summary: Optional[str] = None):
    engine = get_db_engine()
    with engine.begin() as connection:
        existing_row = connection.execute(
            text(
                """
                SELECT contexto, resumen
                FROM public.agent_conversation
                WHERE id = :conversation_id
                """
            ),
            {"conversation_id": conversation_id},
        ).mappings().one()

        merged_context = dict(existing_row["contexto"] or {})
        merged_context.update(context_updates or {})
        connection.execute(
            text(
                """
                UPDATE public.agent_conversation
                SET resumen = :summary,
                    contexto = CAST(:context_payload AS jsonb),
                    updated_at = now(),
                    last_message_at = now()
                WHERE id = :conversation_id
                """
            ),
            {
                "summary": summary if summary is not None else existing_row["resumen"],
                "context_payload": safe_json_dumps(merged_context),
                "conversation_id": conversation_id,
            },
        )


def close_conversation(conversation_id: int, context_updates: dict, summary: Optional[str] = None, final_status: str = "gestionado"):
    engine = get_db_engine()
    with engine.begin() as connection:
        existing_row = connection.execute(
            text(
                """
                SELECT contexto, resumen
                FROM public.agent_conversation
                WHERE id = :conversation_id
                """
            ),
            {"conversation_id": conversation_id},
        ).mappings().one()

        merged_context = dict(existing_row["contexto"] or {})
        merged_context.update(context_updates or {})
        merged_context["final_status"] = final_status

        connection.execute(
            text(
                """
                UPDATE public.agent_conversation
                SET estado = 'cerrada',
                    resumen = :summary,
                    contexto = CAST(:context_payload AS jsonb),
                    updated_at = now(),
                    last_message_at = now()
                WHERE id = :conversation_id
                """
            ),
            {
                "summary": summary if summary is not None else existing_row["resumen"],
                "context_payload": safe_json_dumps(merged_context),
                "conversation_id": conversation_id,
            },
        )


# ── Detección de despedida del cliente ──
_FAREWELL_PATTERNS = re.compile(
    r"^\s*(?:(?:muchas\s+)?gracias|chao|cha[uo]|adios|adiós|bye|hasta\s+luego|"
    r"nos\s+vemos|listo\s+gracias|ok\s+gracias|bueno\s+gracias|"
    r"no\s+más\s+(?:por\s+ahora|gracias)|eso\s+(?:es\s+todo|era\s+todo)|"
    r"ya\s+(?:no\s+más|es\s+todo|era\s+todo)|buen(?:o|a)\s+(?:gracias|noche|tarde|día))\s*[.!]*\s*$",
    re.IGNORECASE,
)


def is_simple_greeting(user_message: str) -> bool:
    """Detect simple greetings/farewells that don't need RAG enforcement."""
    msg = (user_message or "").strip().lower()
    if len(msg) < 3:
        return True
    _GREETING_WORDS = {
        "hola", "buenas", "buenos días", "buenos dias", "buenas tardes",
        "buenas noches", "hey", "ey", "hi", "hello", "qué tal", "que tal",
        "cómo estás", "como estas", "gracias", "muchas gracias", "ok", "listo",
        "dale", "perfecto", "sí", "si", "no", "vale", "chao", "adiós", "adios",
        "hasta luego", "bye", "nos vemos",
    }
    # If the entire message is a greeting/short response
    if msg in _GREETING_WORDS or msg.rstrip(".,!?¡¿ ") in _GREETING_WORDS:
        return True
    # Very short messages (< 15 chars) that are likely confirmations
    if len(msg) < 15 and not any(w in msg for w in ["pintar", "pintura", "pint", "fachada", "piso", "techo", "reja", "madera", "humedad", "moho", "goteras", "óxido", "oxido"]):
        return True
    return False


def detect_farewell(user_message: str) -> bool:
    return bool(_FAREWELL_PATTERNS.match(user_message.strip()))


# ── Puntuación de confianza de la respuesta del agente ──
_LOW_CONFIDENCE_SIGNALS = [
    "no encontr", "no tengo", "no tenemos", "no puedo", "no manejo",
    "no dispongo", "no sé", "no se", "no conozco",
    "no tengo información", "no tengo informacion",
    "un asesor te contactará", "escalar", "no aplica",
    "no fue posible", "error al", "fallo al",
]
_MEDIUM_CONFIDENCE_SIGNALS = [
    "te recomiendo comunicarte", "consultar con", "validarlo con",
    "verificar con", "confirmar con logística", "confirmar con el fabricante",
    "no manejamos",
]


def score_agent_confidence(response_text: str, tool_calls: list, intent: str) -> dict:
    resp_lower = (response_text or "").lower()
    score = 1.0
    signals = []

    # Penalizar señales de baja confianza (-0.30 cada una)
    for signal in _LOW_CONFIDENCE_SIGNALS:
        if signal in resp_lower:
            score -= 0.30
            signals.append(f"low:{signal}")
    # Penalizar señales de confianza media (-0.15 cada una)
    for signal in _MEDIUM_CONFIDENCE_SIGNALS:
        if signal in resp_lower:
            score -= 0.15
            signals.append(f"med:{signal}")

    # Bonificar uso de herramientas (respuestas basadas en datos reales)
    tool_names = [tc.get("name", "") for tc in (tool_calls or [])]
    if "consultar_conocimiento_tecnico" in tool_names:
        score += 0.15
        signals.append("rag_used")
    if "consultar_inventario" in tool_names:
        score += 0.10
        signals.append("inventory_used")
    if "buscar_documento_tecnico" in tool_names:
        score += 0.10
        signals.append("doc_used")

    # Penalizar respuestas triviales (< 10 chars) o muy cortas (< 30 chars)
    if len(response_text or "") < 10:
        score -= 0.55
        signals.append("trivial_response")
    elif len(response_text or "") < 30:
        score -= 0.30
        signals.append("very_short_response")

    # Penalizar cuando hay señales LOW y no se usaron herramientas
    if not tool_names and any(s.startswith("low:") for s in signals):
        score -= 0.25
        signals.append("no_tools_with_issues")

    # Penalizar si el intent es consulta general y hay herramientas no usadas
    if intent == "consulta_general" and not tool_names:
        if any(kw in resp_lower for kw in ["pintura", "pintar", "humedad", "corrosión", "óxido", "techo", "fachada"]):
            score -= 0.10
            signals.append("technical_without_tools")

    score = max(0.0, min(1.0, score))
    level = "alta" if score >= 0.75 else ("media" if score >= 0.50 else "baja")
    return {"score": round(score, 2), "level": level, "signals": signals}


# ── Alerta inteligente para conversaciones problemáticas ──
def evaluate_and_create_alert(
    conversation_id: int,
    cliente_id: Optional[int],
    user_message: str,
    ai_result: dict,
    confidence: dict,
):
    alert_type = None
    priority = "media"
    detail = {
        "mensaje_cliente": user_message[:300],
        "respuesta_agente": (ai_result.get("response_text") or "")[:300],
        "confianza": confidence,
        "intent": ai_result.get("intent"),
    }

    if confidence["level"] == "baja":
        alert_type = "respuesta_baja_confianza"
        priority = "alta"
        detail["razon"] = "El agente respondió con baja confianza. Revisar si la respuesta es correcta."
    elif any(s.startswith("low:") for s in confidence.get("signals", [])):
        # Tiene señales de problema aunque la confianza total no sea baja
        low_signals = [s for s in confidence["signals"] if s.startswith("low:")]
        if len(low_signals) >= 2:
            alert_type = "multiples_señales_problema"
            detail["razon"] = f"Múltiples señales de respuesta inadecuada: {low_signals}"

    if alert_type:
        try:
            upsert_agent_task(
                conversation_id,
                cliente_id,
                alert_type,
                f"Alerta: {alert_type} — revisar respuesta del agente",
                detail,
                priority,
            )
            logger.warning(
                "ALERTA agente [%s] conv=%d confianza=%.2f señales=%s",
                alert_type, conversation_id, confidence["score"], confidence["signals"],
            )
        except Exception as exc:
            logger.error("No se pudo crear alerta: %s", exc)


def upsert_agent_task(conversation_id: int, cliente_id: Optional[int], task_type: str, summary: str, detail: dict, priority: str):
    engine = get_db_engine()
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO public.agent_task (
                    conversation_id,
                    cliente_id,
                    tipo_tarea,
                    prioridad,
                    estado,
                    resumen,
                    detalle,
                    created_at,
                    updated_at
                )
                VALUES (
                    :conversation_id,
                    :cliente_id,
                    :task_type,
                    :priority,
                    'pendiente',
                    :summary,
                    CAST(:detail AS jsonb),
                    now(),
                    now()
                )
                """
            ),
            {
                "conversation_id": conversation_id,
                "cliente_id": cliente_id,
                "task_type": task_type,
                "priority": priority,
                "summary": summary,
                "detail": safe_json_dumps(detail),
            },
        )


def upsert_commercial_draft(
    intent: str,
    conversation_id: int,
    contact_id: Optional[int],
    cliente_id: Optional[int],
    commercial_draft: dict,
):
    if intent not in {"pedido", "cotizacion"}:
        return None

    matched_items = [item for item in (commercial_draft.get("items") or []) if item.get("status") == "matched" and item.get("matched_product")]
    store_filters = commercial_draft.get("store_filters") or []
    store_code = store_filters[0] if store_filters else None
    store_name = STORE_CODE_LABELS.get(store_code) if store_code else None
    summary = f"Borrador de {'pedido' if intent == 'pedido' else 'cotización'} con {len(commercial_draft.get('items') or [])} líneas conversacionales"
    header_table = "agent_order" if intent == "pedido" else "agent_quote"
    line_table = "agent_order_line" if intent == "pedido" else "agent_quote_line"
    foreign_key = "order_id" if intent == "pedido" else "quote_id"
    draft_id = commercial_draft.get("draft_id")

    engine = get_db_engine()
    with engine.begin() as connection:
        if draft_id:
            connection.execute(
                text(
                    f"""
                    UPDATE public.{header_table}
                    SET contacto_id = :contact_id,
                        cliente_id = :cliente_id,
                        almacen_codigo = :store_code,
                        almacen_nombre = :store_name,
                        resumen = :summary,
                        metadata = CAST(:metadata AS jsonb),
                        updated_at = now()
                    WHERE id = :draft_id
                    """
                ),
                {
                    "draft_id": draft_id,
                    "contact_id": contact_id,
                    "cliente_id": cliente_id,
                    "store_code": store_code,
                    "store_name": store_name,
                    "summary": summary,
                    "metadata": safe_json_dumps(commercial_draft),
                },
            )
        else:
            draft_id = connection.execute(
                text(
                    f"""
                    INSERT INTO public.{header_table} (
                        conversation_id,
                        contacto_id,
                        cliente_id,
                        almacen_codigo,
                        almacen_nombre,
                        resumen,
                        metadata,
                        created_at,
                        updated_at
                    )
                    VALUES (
                        :conversation_id,
                        :contact_id,
                        :cliente_id,
                        :store_code,
                        :store_name,
                        :summary,
                        CAST(:metadata AS jsonb),
                        now(),
                        now()
                    )
                    RETURNING id
                    """
                ),
                {
                    "conversation_id": conversation_id,
                    "contact_id": contact_id,
                    "cliente_id": cliente_id,
                    "store_code": store_code,
                    "store_name": store_name,
                    "summary": summary,
                    "metadata": safe_json_dumps(commercial_draft),
                },
            ).scalar_one()

        connection.execute(text(f"DELETE FROM public.{line_table} WHERE {foreign_key} = :draft_id"), {"draft_id": draft_id})

        for line_number, item in enumerate(matched_items, start=1):
            product = item.get("matched_product") or {}
            product_request = item.get("product_request") or {}
            quantity_value = parse_numeric_value(product_request.get("requested_quantity")) or 1
            stock_value = parse_numeric_value(product.get("stock_total") if product.get("stock_total") is not None else product.get("stock"))
            connection.execute(
                text(
                    f"""
                    INSERT INTO public.{line_table} (
                        {foreign_key},
                        line_number,
                        producto_codigo,
                        referencia,
                        descripcion,
                        marca,
                        presentacion,
                        almacen_codigo,
                        almacen_nombre,
                        cantidad,
                        stock_confirmado,
                        metadata,
                        created_at,
                        updated_at
                    )
                    VALUES (
                        :draft_id,
                        :line_number,
                        :producto_codigo,
                        :referencia,
                        :descripcion,
                        :marca,
                        :presentacion,
                        :almacen_codigo,
                        :almacen_nombre,
                        :cantidad,
                        :stock_confirmado,
                        CAST(:metadata AS jsonb),
                        now(),
                        now()
                    )
                    """
                ),
                {
                    "draft_id": draft_id,
                    "line_number": line_number,
                    "producto_codigo": product.get("producto_codigo") or product.get("codigo_articulo"),
                    "referencia": product.get("referencia") or product.get("codigo_articulo"),
                    "descripcion": product.get("descripcion") or product.get("nombre_articulo") or item.get("original_text"),
                    "marca": infer_product_brand_from_row(product),
                    "presentacion": infer_product_presentation_from_row(product),
                    "almacen_codigo": store_code,
                    "almacen_nombre": store_name,
                    "cantidad": quantity_value,
                    "stock_confirmado": stock_value,
                    "metadata": safe_json_dumps(item),
                },
            )

    return draft_id


def fetch_pending_dispatches(destination_store_code: Optional[str] = None, limit: int = 8):
    where_sql = "WHERE d.status IN ('pendiente', 'en_transito')"
    params = {"limit": limit}
    if destination_store_code:
        where_sql += " AND d.destination_store_code = :destination_store_code"
        params["destination_store_code"] = normalize_store_code(destination_store_code)

    engine = get_db_engine()
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                f"""
                SELECT
                    d.id,
                    d.order_id,
                    d.status,
                    d.destination_store_code,
                    d.destination_store_name,
                    d.facturador_name,
                    d.facturador_email,
                    d.export_filename,
                    d.dropbox_path,
                    d.exported_at,
                    o.numero_externo,
                    o.resumen,
                    wc.nombre_visible AS contacto_nombre
                FROM public.agent_order_dispatch d
                LEFT JOIN public.agent_order o ON o.id = d.order_id
                LEFT JOIN public.whatsapp_contacto wc ON wc.id = d.contacto_id
                {where_sql}
                ORDER BY d.exported_at DESC NULLS LAST, d.id DESC
                LIMIT :limit
                """
            ),
            params,
        ).mappings().all()
    return [dict(row) for row in rows]


def fetch_pending_dispatch_shortages(destination_store_code: Optional[str] = None, limit: int = 20):
    where_sql = "WHERE d.status IN ('pendiente', 'en_transito')"
    params = {"limit": limit}
    normalized_destination = normalize_store_code(destination_store_code)
    if normalized_destination:
        where_sql += " AND d.destination_store_code = :destination_store_code"
        params["destination_store_code"] = normalized_destination

    engine = get_db_engine()
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                f"""
                WITH pending_lines AS (
                    SELECT
                        d.id AS dispatch_id,
                        d.order_id,
                        d.destination_store_code,
                        d.destination_store_name,
                        ol.referencia,
                        MAX(ol.descripcion) AS descripcion,
                        SUM(COALESCE(ol.cantidad, 0)) AS required_qty
                    FROM public.agent_order_dispatch d
                    JOIN public.agent_order_line ol ON ol.order_id = d.order_id
                    {where_sql}
                    GROUP BY 1, 2, 3, 4, 5
                ),
                destination_stock AS (
                    SELECT
                        referencia,
                        cod_almacen,
                        COALESCE(SUM(stock_disponible), 0) AS stock_destino
                    FROM public.vw_inventario_agente
                    GROUP BY 1, 2
                )
                SELECT
                    p.dispatch_id,
                    p.order_id,
                    p.destination_store_code,
                    p.destination_store_name,
                    p.referencia,
                    p.descripcion,
                    p.required_qty,
                    COALESCE(ds.stock_destino, 0) AS stock_destino,
                    GREATEST(p.required_qty - COALESCE(ds.stock_destino, 0), 0) AS shortage_qty
                FROM pending_lines p
                LEFT JOIN destination_stock ds
                    ON ds.referencia = p.referencia
                   AND ds.cod_almacen = p.destination_store_code
                WHERE GREATEST(p.required_qty - COALESCE(ds.stock_destino, 0), 0) > 0
                ORDER BY shortage_qty DESC, p.required_qty DESC, p.referencia ASC
                LIMIT :limit
                """
            ),
            params,
        ).mappings().all()
    return [dict(row) for row in rows]


def fetch_best_origin_stock_for_reference(referencia: str, destination_store_code: str, preferred_origin_store_code: Optional[str] = None):
    where_sql = "WHERE referencia = :referencia AND cod_almacen <> :destination_store_code AND COALESCE(stock_disponible, 0) > 0"
    params = {"referencia": referencia, "destination_store_code": normalize_store_code(destination_store_code)}
    origin_code = normalize_store_code(preferred_origin_store_code)
    if origin_code:
        where_sql += " AND cod_almacen = :origin_store_code"
        params["origin_store_code"] = origin_code

    engine = get_db_engine()
    with engine.connect() as connection:
        row = connection.execute(
            text(
                f"""
                SELECT
                    cod_almacen AS origin_store_code,
                    almacen_nombre AS origin_store_name,
                    referencia,
                    MAX(descripcion) AS descripcion,
                    COALESCE(SUM(stock_disponible), 0) AS stock_origen
                FROM public.vw_inventario_agente
                {where_sql}
                GROUP BY 1, 2, 3
                ORDER BY stock_origen DESC, origin_store_name ASC
                LIMIT 1
                """
            ),
            params,
        ).mappings().first()
    return dict(row) if row else None


def build_transfer_suggestions_for_pending_dispatches(destination_store_code: Optional[str], preferred_origin_store_code: Optional[str] = None, limit: int = 6):
    suggestions = []
    for shortage in fetch_pending_dispatch_shortages(destination_store_code, limit=limit * 2):
        origin_stock = fetch_best_origin_stock_for_reference(
            shortage.get("referencia"),
            shortage.get("destination_store_code"),
            preferred_origin_store_code,
        )
        if not origin_stock:
            continue
        available_qty = parse_numeric_value(origin_stock.get("stock_origen")) or 0
        shortage_qty = parse_numeric_value(shortage.get("shortage_qty")) or 0
        suggested_qty = min(available_qty, shortage_qty)
        if suggested_qty <= 0:
            continue
        suggestions.append(
            {
                "dispatch_id": shortage.get("dispatch_id"),
                "order_id": shortage.get("order_id"),
                "reference": shortage.get("referencia"),
                "description": shortage.get("descripcion"),
                "destination_store_code": shortage.get("destination_store_code"),
                "destination_store_name": shortage.get("destination_store_name"),
                "destination_stock": shortage.get("stock_destino"),
                "required_qty": shortage.get("required_qty"),
                "shortage_qty": shortage_qty,
                "origin_store_code": origin_stock.get("origin_store_code"),
                "origin_store_name": origin_stock.get("origin_store_name"),
                "origin_stock": available_qty,
                "suggested_qty": suggested_qty,
            }
        )
        if len(suggestions) >= limit:
            break
    return suggestions


def fetch_shortages_without_internal_origin(destination_store_code: Optional[str], preferred_origin_store_code: Optional[str] = None, limit: int = 6):
    unresolved = []
    for shortage in fetch_pending_dispatch_shortages(destination_store_code, limit=limit * 3):
        origin_stock = fetch_best_origin_stock_for_reference(
            shortage.get("referencia"),
            shortage.get("destination_store_code"),
            preferred_origin_store_code,
        )
        if origin_stock:
            continue
        unresolved.append(
            {
                "dispatch_id": shortage.get("dispatch_id"),
                "order_id": shortage.get("order_id"),
                "reference": shortage.get("referencia"),
                "description": shortage.get("descripcion"),
                "destination_store_code": shortage.get("destination_store_code"),
                "destination_store_name": shortage.get("destination_store_name"),
                "destination_stock": shortage.get("stock_destino"),
                "required_qty": shortage.get("required_qty"),
                "shortage_qty": shortage.get("shortage_qty"),
            }
        )
        if len(unresolved) >= limit:
            break
    return unresolved


def parse_internal_selection_indexes(content: Optional[str], max_items: int):
    digits = [int(value) for value in re.findall(r"\d+", content or "")]
    indexes = []
    for value in digits:
        if 1 <= value <= max_items and value not in indexes:
            indexes.append(value)
    return indexes


def summarize_transfer_candidates(items: list[dict], max_items: int = 5):
    lines = []
    for idx, item in enumerate(items[:max_items], start=1):
        origin_label = item.get("origin_store_name") or item.get("source_store_name") or "sin origen"
        destination_label = item.get("destination_store_name") or "sin destino"
        qty_value = item.get("suggested_qty") if item.get("suggested_qty") is not None else item.get("shortage_qty")
        lines.append(
            f"{idx}. Pedido {item.get('order_id')} | {item.get('reference')} | {origin_label} -> {destination_label} | {format_quantity(qty_value)} und"
        )
    return "\n".join(lines)


def build_internal_transfer_guidance_text(flow_payload: dict):
    destination_label = flow_payload.get("destination_store_name") or "la sede destino"
    suggestions = flow_payload.get("suggestions") or []
    unresolved = flow_payload.get("unresolved_shortages") or []
    sections = [f"Revisión operativa para {destination_label}:"]
    if suggestions:
        sections.append("Traslados sugeridos listos para crear:")
        sections.append(summarize_transfer_candidates(suggestions))
    if unresolved:
        sections.append("Faltantes sin origen interno útil:")
        sections.append(summarize_transfer_candidates(unresolved))
    sections.append(
        "Responde `confirmar traslado` para crear todos los traslados sugeridos, `1 y 3` para crear solo algunas opciones, `compras` para escalar faltantes sin origen, o `cancelar` para salir."
    )
    return "\n".join(section for section in sections if section)


def build_internal_transfer_flow_payload(destination_store_code: Optional[str], preferred_origin_store_code: Optional[str] = None):
    destination_store_code = normalize_store_code(destination_store_code)
    preferred_origin_store_code = normalize_store_code(preferred_origin_store_code)
    suggestions = build_transfer_suggestions_for_pending_dispatches(destination_store_code, preferred_origin_store_code)
    unresolved = fetch_shortages_without_internal_origin(destination_store_code, preferred_origin_store_code)
    return {
        "destination_store_code": destination_store_code,
        "destination_store_name": get_store_short_label(destination_store_code) or STORE_CODE_LABELS.get(destination_store_code) or destination_store_code,
        "origin_store_code": preferred_origin_store_code,
        "origin_store_name": get_store_short_label(preferred_origin_store_code) or STORE_CODE_LABELS.get(preferred_origin_store_code) or preferred_origin_store_code,
        "suggestions": suggestions,
        "unresolved_shortages": unresolved,
        "step": "awaiting_confirmation",
    }


def create_procurement_followup(context: dict, requested_by_user: dict, shortages: list[dict], notes: Optional[str] = None):
    if not shortages:
        return {"created": False, "task_type": None, "notification": None}
    destination_label = shortages[0].get("destination_store_name") or "sede destino"
    detail = {
        "requested_by": requested_by_user,
        "destination_store_code": shortages[0].get("destination_store_code"),
        "destination_store_name": destination_label,
        "shortages": shortages,
        "notes": notes,
    }
    upsert_agent_task(
        context["conversation_id"],
        context.get("cliente_id"),
        "abastecimiento_compras",
        f"Faltantes sin origen interno para {destination_label}",
        detail,
        "alta",
    )
    notification = notify_procurement_request_by_email(shortages, requested_by_user, notes=notes)
    return {"created": True, "task_type": "abastecimiento_compras", "notification": notification}


def fetch_inventory_gap_between_stores(source_store_code: str, destination_store_code: str, limit: int = 10):
    source_store_code = normalize_store_code(source_store_code)
    destination_store_code = normalize_store_code(destination_store_code)
    if not source_store_code or not destination_store_code:
        return []

    engine = get_db_engine()
    with engine.connect() as connection:
        rows = connection.execute(
            text(
                """
                SELECT
                    referencia,
                    MAX(descripcion) AS descripcion,
                    MAX(marca) AS marca,
                    COALESCE(SUM(CASE WHEN cod_almacen = :source_store_code THEN stock_disponible ELSE 0 END), 0) AS stock_origen,
                    COALESCE(SUM(CASE WHEN cod_almacen = :destination_store_code THEN stock_disponible ELSE 0 END), 0) AS stock_destino
                FROM public.vw_inventario_agente
                WHERE cod_almacen = :source_store_code OR cod_almacen = :destination_store_code
                GROUP BY referencia
                HAVING COALESCE(SUM(CASE WHEN cod_almacen = :source_store_code THEN stock_disponible ELSE 0 END), 0) > 0
                   AND COALESCE(SUM(CASE WHEN cod_almacen = :destination_store_code THEN stock_disponible ELSE 0 END), 0) <= 0
                ORDER BY stock_origen DESC, referencia ASC
                LIMIT :limit
                """
            ),
            {
                "source_store_code": source_store_code,
                "destination_store_code": destination_store_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows]


def create_transfer_request_records(suggestions: list[dict], requested_by_user: dict, notes: Optional[str] = None):
    if not suggestions:
        return []
    requester_metadata = dict((requested_by_user or {}).get("metadata") or {})
    requester_store_code = normalize_store_code(requester_metadata.get("store_code"))
    engine = get_db_engine()
    created_rows = []
    with engine.begin() as connection:
        for suggestion in suggestions:
            origin_store_code = normalize_store_code(suggestion.get("origin_store_code"))
            if requester_store_code and origin_store_code and requester_store_code != origin_store_code:
                continue
            transfer_id = connection.execute(
                text(
                    """
                    INSERT INTO public.agent_transfer_request (
                        order_dispatch_id,
                        order_id,
                        requested_by_user_id,
                        requested_via,
                        source_store_code,
                        source_store_name,
                        destination_store_code,
                        destination_store_name,
                        referencia,
                        descripcion,
                        quantity_requested,
                        status,
                        summary,
                        notes,
                        metadata,
                        created_at,
                        updated_at
                    )
                    VALUES (
                        :order_dispatch_id,
                        :order_id,
                        :requested_by_user_id,
                        'whatsapp_interno',
                        :source_store_code,
                        :source_store_name,
                        :destination_store_code,
                        :destination_store_name,
                        :referencia,
                        :descripcion,
                        :quantity_requested,
                        'pendiente',
                        :summary,
                        :notes,
                        CAST(:metadata AS jsonb),
                        now(),
                        now()
                    )
                    RETURNING id
                    """
                ),
                {
                    "order_dispatch_id": suggestion.get("dispatch_id"),
                    "order_id": suggestion.get("order_id"),
                    "requested_by_user_id": requested_by_user.get("id"),
                    "source_store_code": origin_store_code,
                    "source_store_name": suggestion.get("origin_store_name"),
                    "destination_store_code": suggestion.get("destination_store_code"),
                    "destination_store_name": suggestion.get("destination_store_name"),
                    "referencia": suggestion.get("reference"),
                    "descripcion": suggestion.get("description"),
                    "quantity_requested": suggestion.get("suggested_qty"),
                    "summary": f"Traslado {suggestion.get('reference')} {suggestion.get('origin_store_name')} -> {suggestion.get('destination_store_name')}",
                    "notes": notes,
                    "metadata": safe_json_dumps(suggestion),
                },
            ).scalar_one()
            created_rows.append({"id": transfer_id, **suggestion})
    return created_rows


def extract_document_candidate(text_value: Optional[str]):
    if not text_value:
        return None
    matches = re.findall(r"\b\d{6,15}\b", text_value)
    return matches[0] if matches else None


def extract_email_address(text_value: Optional[str]):
    if not text_value:
        return None
    match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text_value, flags=re.IGNORECASE)
    return match.group(0).strip() if match else None


def extract_delivery_channel(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return None
    if extract_email_address(text_value) or re.search(r"\b(correo|email|mail)\b", normalized):
        return "email"
    if any(phrase in normalized for phrase in ["whatsapp", "wpp", "chat", "por aqui", "por aquí", "aca", "acá"]):
        return "chat"
    return None


def summarize_commercial_item(item: dict):
    product_request = item.get("product_request") or {}
    matched_product = item.get("matched_product") or {}
    raw_description = matched_product.get("descripcion") or matched_product.get("nombre_articulo") or item.get("original_text") or "producto"
    presentation = infer_product_presentation_from_row(matched_product) if matched_product else None
    brand = infer_product_brand_from_row(matched_product) if matched_product else None
    commercial_name = translate_product_to_commercial(raw_description, presentation, brand)
    requested_quantity = parse_numeric_value(product_request.get("requested_quantity")) or 1
    requested_unit = product_request.get("requested_unit")
    if requested_unit:
        quantity_label = f"{format_quantity(requested_quantity)} {get_presentation_label(requested_unit, requested_quantity)} de "
    elif requested_quantity > 1:
        quantity_label = f"{format_quantity(requested_quantity)} unidades de "
    else:
        quantity_label = ""
    return f"{quantity_label}{commercial_name}".strip()


def summarize_commercial_items(items: list[dict]):
    labels = [summarize_commercial_item(item) for item in items[:6] if item.get("status") == "matched"]
    if not labels:
        return ""
    if len(labels) == 1:
        return labels[0]
    if len(labels) == 2:
        return f"{labels[0]} y {labels[1]}"
    return ", ".join(labels[:-1]) + f" y {labels[-1]}"


def extract_identity_lookup_candidate(text_value: Optional[str], conversation_context: Optional[dict], allow_unprompted: bool = False):
    if not text_value:
        return None

    context = conversation_context or {}
    verification_flow_active = bool(
        context.get("awaiting_verification")
        or context.get("pending_intent") in {"consulta_cartera", "consulta_compras"}
    )
    if not verification_flow_active and not allow_unprompted:
        return None

    normalized_text = normalize_text_value(text_value)
    if not normalized_text:
        return None

    numeric_matches = re.findall(r"\b\d{4,15}\b", normalized_text)
    remaining_text = re.sub(r"\b\d{4,15}\b", " ", normalized_text)
    remaining_tokens = [token for token in remaining_text.split() if token]
    allowed_code_tokens = {"mi", "codigo", "cod", "cliente", "es", "el", "del", "de"}
    if numeric_matches and len(numeric_matches) == 1 and all(token in allowed_code_tokens for token in remaining_tokens):
        return {"type": "numeric_lookup", "value": numeric_matches[0]}

    if not verification_flow_active:
        return None

    # During verification, ignore messages that are clearly questions or commands, not names
    if any(character.isalpha() for character in text_value):
        # Skip common question patterns that aren't name lookups
        question_patterns = [
            r"\b(como|cómo|donde|dónde|cuando|cuándo|cual|cuál|que|qué|por que|por qué|puedo|pueden|hay|tiene)\b",
            r"\b(pagar|enviar|comprar|hacer|necesito|quiero|ayuda|informacion|información)\b",
        ]
        for qp in question_patterns:
            if re.search(qp, normalized_text):
                return None

        product_request = extract_product_request(text_value)
        tokens = [token for token in normalized_text.split() if len(token) >= 3]
        candidate_intent = detect_business_intent(text_value)
        strong_product_signal = bool(
            product_request.get("product_codes")
            or product_request.get("requested_unit")
            or product_request.get("requested_quantity")
            or product_request.get("store_filters")
            or product_request.get("brand_filters")
        )
        if candidate_intent not in {"consulta_general", "consulta_cartera", "consulta_compras"}:
            return None
        if 2 <= len(tokens) <= 6 and not strong_product_signal and candidate_intent == "consulta_general":
            return {"type": "name_lookup", "value": text_value.strip()}
        if candidate_intent == "consulta_general" and not looks_like_product_query(text_value, product_request) and 2 <= len(tokens) <= 6:
            return {"type": "name_lookup", "value": text_value.strip()}

    return None


def resolve_identity_candidate(identity_candidate: Optional[dict], phone_number: Optional[str] = None):
    if not identity_candidate:
        return None, None

    candidate_type = identity_candidate.get("type")
    candidate_value = identity_candidate.get("value")
    verified_context = None
    verified_by = None

    if candidate_type == "numeric_lookup":
        verified_context = find_cliente_contexto_by_document(candidate_value)
        if verified_context:
            verified_by = "document"
        else:
            verified_context = find_cliente_contexto_by_customer_code(candidate_value)
            if verified_context:
                verified_by = "customer_code"
    elif candidate_type == "name_lookup":
        verified_context = find_cliente_contexto_by_name(candidate_value)
        if verified_context:
            verified_by = "name"

    if not verified_context and phone_number:
        verified_context = find_cliente_contexto_by_phone(phone_number)
        if verified_context:
            verified_by = "phone"

    return verified_context, verified_by


def build_identity_not_found_reply(identity_candidate: Optional[dict]):
    candidate_value = (identity_candidate or {}).get("value") or "ese dato"
    candidate_type = (identity_candidate or {}).get("type")
    if candidate_type == "name_lookup":
        return (
            f"No me aparece {candidate_value} por acá, ¿de pronto está a nombre de otra persona o empresa? "
            "Envíame la cédula o NIT y con eso te busco."
        )
    return (
        f"No me aparece {candidate_value} en el sistema. "
        "¿De pronto es otro número? Prueba con tu cédula, NIT o código de cliente."
    )


def build_commercial_customer_snapshot(customer_context: Optional[dict]):
    if not customer_context:
        return None
    return {
        "cliente_codigo": customer_context.get("cliente_codigo") or customer_context.get("verified_cliente_codigo") or customer_context.get("cod_cliente"),
        "nombre_cliente": customer_context.get("nombre_cliente"),
        "nit": customer_context.get("nit") or customer_context.get("numero_documento") or customer_context.get("documento"),
        "email": customer_context.get("email"),
        "telefono1": customer_context.get("telefono1"),
        "telefono2": customer_context.get("telefono2"),
        "vendedor": customer_context.get("vendedor"),
        "vendedor_codigo": customer_context.get("vendedor_codigo"),
        "zona": customer_context.get("zona"),
    }


def resolve_commercial_customer_context(customer_input: Optional[str]):
    raw_value = (customer_input or "").strip()
    if not raw_value:
        return None, None

    identity_candidate = extract_identity_lookup_candidate(
        raw_value,
        {"awaiting_verification": True, "pending_intent": "consulta_cartera"},
        allow_unprompted=True,
    )
    if not identity_candidate:
        normalized = normalize_text_value(raw_value)
        tokens = [token for token in normalized.split() if len(token) >= 2]
        if 2 <= len(tokens) <= 8:
            identity_candidate = {"type": "name_lookup", "value": raw_value}

    verified_context, verified_by = resolve_identity_candidate(identity_candidate)
    if not verified_context:
        return None, verified_by
    return build_commercial_customer_snapshot(verified_context), verified_by


def trim_commercial_customer_candidate(raw_value: Optional[str]):
    text_value = (raw_value or "").strip()
    if not text_value:
        return ""
    quantity_words_pattern = "|".join(QUANTITY_WORD_MAP.keys())
    split_match = re.search(
        r"\s+(?=(?:\d+\s*/\s*(?:1|4|5)|\d+|" + quantity_words_pattern + r")\s+(?:cunetes?|cuñetes?|galones?|galon|cuartos?|canecas?|cubetas?|rodillos?|brochas?|bochas?|lijas?|cintas?|bultos?|kilos?|metros?|rollos?|tubos?|tarros?|cajas?|paquetes?|unidades?|cerraduras?|candados?|chapas?|selladores?|silicones?|llaves?|bisagras?|manijas?|laminas?|láminas?|tejas?|perfiles?|angulos?|ángulos?|baldes?|de)\b)",
        text_value,
        flags=re.IGNORECASE,
    )
    if split_match:
        text_value = text_value[:split_match.start()].strip(" ,.;:-")
    return text_value.strip()


def is_identity_verification_message(text_value: Optional[str], conversation_context: Optional[dict]):
    return extract_identity_lookup_candidate(text_value, conversation_context) is not None


def is_sensitive_intent_message(text_value: Optional[str]):
    if not text_value:
        return False
    lowered = normalize_text_value(text_value)
    sensitive_keywords = [
        "cartera",
        "saldo",
        "debo",
        "deuda",
        "cupo",
        "credito",
        "vencid",
        "factura",
        "facturas",
        "pago",
        "pagos",
        "estado de cuenta",
        "ventas",
        "compras",
        "recaudo",
    ]
    return any(keyword in lowered for keyword in sensitive_keywords) or has_keyword_or_similar(lowered, ["factura", "facturas", "vencida", "vencidas"])


def has_non_product_business_signal(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    return any(keyword in normalized for keyword in NON_PRODUCT_SERVICE_KEYWORDS)


def is_product_intent_message(text_value: Optional[str]):
    if not text_value:
        return False
    lowered = normalize_text_value(text_value)
    product_keywords = [
        "producto",
        "productos",
        "referencia",
        "inventario",
        "stock",
        "marca",
        "articulo",
        "precio",
        "viniltex",
        "vinilico",
        "vinilux",
        "pintulux",
        "3en1",
        "cerradura",
        "cerradur",
        "brocha",
        "rodillo",
        "bisagra",
        "pasador",
        "domestico",
        "pintuco",
        "abracol",
        "yale",
        "goya",
        "galon",
        "galones",
        "cuarto",
        "cuartos",
        "cunete",
        "cunetes",
        "cuñete",
        "cuñetes",
        "caneca",
        "canecas",
        "cubeta",
        "cubetas",
        "p-11",
        "p11",
        "t-11",
        "t11",
    ]
    has_keyword = any(keyword in lowered for keyword in product_keywords)
    quantity_format = bool(re.search(r"\b\d+(?:[.,]\d+)?\s*(galones?|galon|cuartos?|cunetes?|cuñetes?|canecas?|cubetas?)\b", lowered))
    shorthand_format = bool(re.search(r"\b\d+\s*/\s*\d+\b", lowered))
    code_with_stock = bool(re.search(r"\b\d{4,10}\b", lowered)) and any(kw in lowered for kw in ["stock", "inventario", "hay", "precio", "cuanto", "producto"])
    fuzzy_keyword = has_keyword_or_similar(lowered, ["pintulux", "cerradura", "brocha", "rodillo", "domestico", "vinilico", "viniltex", "pintulux", "goya", "p11", "t11", "mega"], threshold=0.78)
    return has_keyword or quantity_format or shorthand_format or code_with_stock or fuzzy_keyword


def is_greeting_message(text_value: Optional[str]):
    lowered = normalize_text_value(text_value)
    if not lowered:
        return False
    exact_greetings = {"hola", "buen dia", "buenos dias", "buenas tardes", "buenas noches", "hello", "hi", "hey"}
    if lowered in exact_greetings:
        return True
    greeting_candidates = ["hola", "buen dia", "buenos dias", "buenas tardes", "buenas noches", "hello", "hi", "hey"]
    if len(lowered.split()) <= 4 and max(sequence_similarity(lowered, candidate) for candidate in greeting_candidates) >= 0.82:
        return True
    tokens = lowered.split()
    if tokens and max(sequence_similarity(tokens[0], candidate) for candidate in ["hola", "hello", "hi", "hey"]) >= 0.8:
        remaining_text = " ".join(tokens[1:]).strip()
        if not remaining_text:
            return True
        if max(sequence_similarity(remaining_text, candidate) for candidate in ["buen dia", "buenos dias", "buenas tardes", "buenas noches"]) >= 0.72:
            return True
    return bool(re.match(
        r"^(hola|hey|buenas?|buenos?\s+dias?|buenas?\s+tardes?|buenas?\s+noches?)"
        r"(\s+(como estas|como esta|que tal|buen dia|buenos dias|buenas tardes|buenas noches))?"
        r"[.!?,\s]*$",
        lowered,
    ))


def is_thanks_or_closing_message(text_value: Optional[str]):
    lowered = normalize_text_value(text_value)
    if not lowered:
        return False

    gratitude_patterns = [
        r"^(gracias|muchas gracias|mil gracias|genial gracias|super gracias|perfecto gracias)[.!?,\s]*$",
        r"^(genial|perfecto|listo|excelente|super|buenisimo|buenisima)(\s+(muchas\s+)?gracias)?[.!?,\s]*$",
        r"^(ok|okay|vale|dale|entendido|comprendido)(\s+(muchas\s+)?gracias)?[.!?,\s]*$",
        r"^(quedo atento|quedo atenta|te aviso|te escribo luego|eso era|nada mas|nada mas gracias)[.!?,\s]*$",
    ]
    return any(re.match(pattern, lowered) for pattern in gratitude_patterns)


def is_new_order_request(text_value: Optional[str]):
    """Detect phrases like 'otro pedido', 'nuevo pedido', 'necesito otro pedido', 'otra cotización'."""
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    return bool(re.search(r"\b(otr[oa]s?|nuev[oa]s?)\s+(pedido|cotizaci[oó]n|orden|lista)\b", normalized))


def resolve_option_selections(text_value: Optional[str], existing_items: list[dict]):
    """Parse option selection patterns like '2a', '4b', 'la 1a y la 3b' from the message."""
    normalized = normalize_text_value(text_value)
    if not normalized or not existing_items:
        return {}
    selections = re.findall(r"(?:^|\s)(\d)\s*([a-c])\b", normalized)
    if not selections:
        selections = re.findall(r"(?:la|el)\s+(\d)\s*([a-c])\b", normalized)
    updates = {}
    for item_idx_str, letter in selections:
        item_idx = int(item_idx_str) - 1
        alt_idx = ord(letter) - 97
        if 0 <= item_idx < len(existing_items):
            item = existing_items[item_idx]
            alternatives = item.get("alternatives") or []
            if 0 <= alt_idx < len(alternatives):
                updates[item_idx] = alternatives[alt_idx]
    return updates


def is_affirmative_message(text_value: Optional[str]):
    lowered = normalize_text_value(text_value)
    if not lowered:
        return False
    return bool(re.match(
        r"^(si|sí|eso es|asi es|as[ií] esta|correcto|exacto|dale|listo|de una|h[aá]gale|ok|okay|perfecto|confirmado)[.!?,\s]*$",
        lowered,
    ))


def is_negative_message(text_value: Optional[str]):
    lowered = normalize_text_value(text_value)
    if not lowered:
        return False
    return bool(re.match(r"^(no|nop|negativo|ya no|ya no mas|ya no m[aá]s)[.!?,\s]*$", lowered))


def has_active_commercial_flow(conversation_context: Optional[dict]):
    context = conversation_context or {}
    draft = dict(context.get("commercial_draft") or {})
    active_intent = draft.get("intent") or context.get("last_direct_intent") or context.get("intent")
    if active_intent not in {"pedido", "cotizacion"}:
        return False
    if draft.get("items"):
        return True
    return bool(active_intent in {"pedido", "cotizacion"} and not draft.get("internal_notified"))


def build_conversation_closing_reply(profile_name: Optional[str]):
    return "¡Con gusto! Quedo por aquí para lo que necesites 👋"


def is_nudge_or_followup(text_value: Optional[str]):
    """Detects short follow-up nudges like '?', 'y?', 'entonces?', 'hola?'."""
    lowered = normalize_text_value(text_value)
    if not lowered:
        return False
    return bool(re.match(r"^[?!¿¡.\s]+$", lowered) or re.match(
        r"^(y|entonces|que paso|qué pasó|que pasa|que hay|hola|ey|oye|bueno|listo|y entonces|y que|y qué|dale|alo|aló|hey)[?!¿¡.\s]*$",
        lowered,
    ))


def build_nudge_reply(conversation_context: Optional[dict]):
    """Build a reply for nudge messages based on the active flow."""
    context = conversation_context or {}
    active_intent = context.get("last_direct_intent") or context.get("intent")
    claim_case = context.get("claim_case") or {}
    commercial_draft = context.get("commercial_draft") or {}

    if claim_case.get("active") and not claim_case.get("submitted"):
        step = claim_case.get("step")
        if step == "awaiting_product":
            return "Disculpa la demora. ¿Me cuentas qué producto es el del reclamo?"
        elif step == "awaiting_detail":
            return "Sigo acá. ¿Me cuentas qué pasó con el producto?"
        elif step == "awaiting_evidence":
            return "Estoy pendiente. ¿Tienes alguna foto o número de lote para el caso?"
        elif step == "awaiting_email":
            return "Solo me falta tu correo para enviarte la constancia del caso. ¿Me lo regalas?"
        return "Sigo acá pendiente, ¿en qué íbamos?"

    if active_intent in {"pedido", "cotizacion"} or commercial_draft.get("intent"):
        items = commercial_draft.get("items") or []
        if items:
            matched = sum(1 for i in items if i.get("status") == "matched")
            pending = len(items) - matched
            if pending > 0:
                return f"Ya tengo {matched} producto(s) listos y me faltan {pending} por precisar. ¿Me confirmas esos que quedaron pendientes?"
            if not commercial_draft.get("store_filters"):
                return "Ya tengo los productos listos. ¿En qué tienda o ciudad los necesitas?"
            return "Ya tengo todo listo. ¿Te confirmo el pedido por aquí o te lo mando al correo?"
        label = "cotización" if active_intent == "cotizacion" else "pedido"
        return f"Claro, seguimos con el {label}. ¿Qué productos necesitas?"

    if context.get("awaiting_verification"):
        return "Sigo esperando tu número de cédula o NIT para poder revisarte esa info 🔒"

    return None


def should_continue_claim_flow(conversation_context: Optional[dict], detected_intent: Optional[str], text_value: Optional[str]):
    claim_case = dict((conversation_context or {}).get("claim_case") or {})
    if not claim_case.get("active") or claim_case.get("submitted"):
        return False

    normalized = normalize_text_value(text_value)
    if not normalized:
        return False

    if extract_email_address(text_value):
        return True
    if detected_intent in {"pedido", "cotizacion", "consulta_cartera", "consulta_compras", "consulta_documentacion"}:
        return False
    if any(keyword in normalized for keyword in ["inventario", "stock", "precio", "cotizar", "cotizacion", "pedido", "cartera", "compras"]):
        return False
    return True


QUANTITY_WORD_MAP = {
    "un": 1, "una": 1, "uno": 1,
    "dos": 2, "tres": 3, "cuatro": 4, "cinco": 5,
    "seis": 6, "siete": 7, "ocho": 8, "nueve": 9, "diez": 10,
    "once": 11, "doce": 12, "quince": 15, "veinte": 20,
    "medio": 0.5, "media": 0.5,
}


def should_continue_commercial_flow(conversation_context: Optional[dict], detected_intent: Optional[str], text_value: Optional[str]):
    context = conversation_context or {}
    active_intent = context.get("last_direct_intent")
    if not active_intent and context.get("intent") in {"pedido", "cotizacion"}:
        active_intent = context.get("intent")
    commercial_draft = dict(context.get("commercial_draft") or {})
    if active_intent not in {"pedido", "cotizacion"}:
        return False
    if commercial_draft and commercial_draft.get("intent") not in {None, active_intent}:
        return False
    if detected_intent in {"consulta_cartera", "consulta_compras", "consulta_documentacion", "reclamo_servicio"}:
        return False

    normalized = normalize_text_value(text_value)
    if not normalized:
        return False

    if extract_store_filters(text_value) or extract_email_address(text_value):
        return True
    if is_affirmative_message(text_value) or is_negative_message(text_value):
        return True
    if is_new_order_request(text_value):
        return True
    if re.search(r"\b\d\s*[a-c]\b", normalized):
        return True
    if re.search(r"\b(confirma|confirmame|conf[ií]rmame|conf[ií]rmalo|sep[aá]ralo|separalo|env[ií]ame|enviame|m[aá]ndame|mandame|eso ser[ií]a todo|eso es|as[ií] est[aá])\b", normalized):
        return True
    if re.search(r"\b(agregale|agregame|ponle|ponme|sumale|quitale|quitame)\b", normalized):
        return True
    if is_product_intent_message(text_value):
        return True
    if len(split_commercial_line_items(text_value)) >= 2:
        return True
    if bool(re.search(r"\b\d+\s*/\s*(1|4|5)\b", normalized)):
        return True

    quantity_words_pattern = "|".join(QUANTITY_WORD_MAP.keys())
    if re.search(r"(?:\d+|(?:" + quantity_words_pattern + r"))\s+\w+", normalized):
        extracted = extract_product_request(text_value)
        if extracted.get("core_terms") and (extracted.get("requested_quantity") or extracted.get("requested_unit")):
            return True

    # If there are ambiguous items in draft, any text with product terms could be a clarification
    if commercial_draft.get("items"):
        has_ambiguous = any(item.get("status") == "ambiguous" for item in commercial_draft["items"])
        if has_ambiguous:
            extracted = extract_product_request(text_value)
            if extracted.get("core_terms") or extracted.get("product_codes"):
                return True

    return False


def split_commercial_line_items(text_value: Optional[str]):
    if not text_value:
        return []

    filler_fragments = {
        "necesito",
        "tambien necesito",
        "también necesito",
        "tambien",
        "también",
        "ademas necesito",
        "además necesito",
        "ademas",
        "además",
        "y",
        "y tambien necesito",
        "y también necesito",
        "agregale",
        "agregame",
        "agregalo",
        "agrega",
        "ponle",
        "ponme",
        "ponlo",
        "sumale",
        "sumame",
        "quitale",
        "quitame",
    }

    def clean_split_candidates(candidates: list[str]):
        cleaned = []
        carry_prefix = ""
        for segment in candidates:
            stripped_segment = segment.strip()
            normalized_segment = normalize_text_value(stripped_segment)
            if not normalized_segment:
                continue
            if normalized_segment in filler_fragments:
                carry_prefix = f"{carry_prefix} {stripped_segment}".strip()
                continue
            if carry_prefix:
                stripped_segment = f"{carry_prefix} {stripped_segment}".strip()
                carry_prefix = ""
            stripped_segment = re.sub(
                r"\s+(?:y|tambien|también|ademas|además|tambien necesito|también necesito|ademas necesito|además necesito)\s*$",
                "",
                stripped_segment,
                flags=re.IGNORECASE,
            ).strip()
            cleaned.append(stripped_segment)
        return cleaned

    quantity_words_pattern = "|".join(QUANTITY_WORD_MAP.keys())
    quantity_start = re.search(
        r"(?:\b\d+\s*/\s*(?:1|4|5)\b|\b\d+\b|\b(?:" + quantity_words_pattern + r")\b)\s+(?:cunetes?|cuñetes?|galones?|galon|cuartos?|canecas?|cubetas?|rodillos?|brochas?|bochas?|lijas?|cintas?|bultos?|kilos?|metros?|rollos?|tubos?|tarros?|cajas?|paquetes?|unidades?|cerraduras?|candados?|chapas?|selladores?|silicones?|llaves?|bisagras?|manijas?|laminas?|láminas?|tejas?|perfiles?|angulos?|ángulos?|baldes?|de)\b",
        text_value,
        flags=re.IGNORECASE,
    )
    text_for_items = text_value[quantity_start.start():].strip() if quantity_start else text_value

    prepared_text = re.sub(r"[;|]+", "\n", text_for_items)
    raw_lines = [line.strip(" -*•\t") for line in prepared_text.splitlines()]
    lines = [line for line in raw_lines if line]
    if len(lines) >= 2:
        return lines

    normalized = normalize_text_value(text_for_items)
    if re.search(r"\b\d+\s*/\s*(1|4|5)\b", normalized):
        split_candidates = [segment.strip() for segment in re.split(r"\s{2,}|,(?=\s*\d)|(?<=\")\s+(?=\d)", text_for_items) if segment.strip()]
        if len(split_candidates) >= 2:
            return split_candidates

    qty_boundary = re.compile(
        r"(?<=\S)\s+(?=(?:\d+\s*/\s*(?:1|4|5)|\d+|" + quantity_words_pattern + r")\s+(?:cunetes?|cuñetes?|galones?|galon|cuartos?|canecas?|cubetas?|rodillos?|brochas?|bochas?|lijas?|cintas?|bultos?|kilos?|metros?|rollos?|tubos?|tarros?|cajas?|paquetes?|unidades?|cerraduras?|candados?|chapas?|selladores?|silicones?|llaves?|bisagras?|manijas?|laminas?|láminas?|tejas?|perfiles?|angulos?|ángulos?|baldes?|de)\b)",
        re.IGNORECASE,
    )
    split_candidates = [segment.strip() for segment in qty_boundary.split(text_for_items) if segment.strip()]
    if len(split_candidates) >= 2:
        cleaned_candidates = clean_split_candidates(split_candidates)
        return cleaned_candidates or split_candidates

    comma_split = [segment.strip() for segment in re.split(r",\s*", text_for_items) if segment.strip()]
    if len(comma_split) >= 2:
        product_like = sum(1 for seg in comma_split if re.search(r"\d|" + quantity_words_pattern, seg.lower()))
        if product_like >= 2:
            return comma_split

    y_split = [segment.strip() for segment in re.split(r"\by\b", text_for_items, flags=re.IGNORECASE) if segment.strip()]
    if len(y_split) >= 2:
        product_like = sum(1 for seg in y_split if re.search(r"\d|" + quantity_words_pattern, seg.lower()))
        if product_like >= 2:
            return y_split

    return lines if lines else [text_for_items.strip()]


def merge_store_filters(product_request: dict, inherited_store_filters: list[str]):
    merged_request = dict(product_request or {})
    if inherited_store_filters and not merged_request.get("store_filters"):
        merged_request["store_filters"] = list(inherited_store_filters)
    return merged_request


def describe_commercial_item_need(item: dict):
    request = item.get("product_request") or {}
    if request.get("product_codes"):
        return request["product_codes"][0]
    if request.get("core_terms"):
        return " ".join(request.get("core_terms")[:4])
    return item.get("original_text") or "ese producto"


def build_commercial_item_result(raw_line: str, inherited_store_filters: list[str], mode: str):
    product_request = merge_store_filters(prepare_product_request_for_search(raw_line), inherited_store_filters)
    product_rows = lookup_product_context(raw_line, product_request)
    requested_store_codes = product_request.get("store_filters") or []
    requested_store_label = STORE_CODE_LABELS.get(requested_store_codes[0]) if len(requested_store_codes) == 1 else None

    item_result = {
        "original_text": raw_line,
        "product_request": product_request,
        "matches": product_rows,
        "status": "missing",
        "message": "",
        "matched_product": None,
        "alternatives": [],
    }

    if not product_rows:
        item_result["message"] = f"{describe_commercial_item_need(item_result)}: necesito la referencia exacta o la presentación para ubicarlo."
        return item_result

    # Build alternatives list from all returned product rows
    seen_refs = set()
    for row in product_rows[:5]:
        ref = row.get("referencia") or row.get("codigo_articulo") or ""
        if ref in seen_refs:
            continue
        seen_refs.add(ref)
        alt_commercial_name = translate_product_to_commercial(
            row.get("descripcion") or row.get("nombre_articulo"),
            infer_product_presentation_from_row(row),
            infer_product_brand_from_row(row),
        )
        alt_stock = parse_numeric_value(row.get("stock_total") if row.get("stock_total") is not None else row.get("stock")) or 0
        item_result["alternatives"].append({
            "commercial_name": alt_commercial_name,
            "referencia": ref,
            "stock_total": alt_stock,
            "row": dict(row),
        })

    if should_ask_product_clarification(product_request, product_rows):
        item_result["status"] = "ambiguous"
        options_text = "\n".join(
            f"   {chr(97 + idx)}) {alt['commercial_name']}"
            for idx, alt in enumerate(item_result["alternatives"][:4])
        )
        item_result["message"] = f"{describe_commercial_item_need(item_result)} — opciones:\n{options_text}"
        return item_result

    top_row = dict(product_rows[0])
    item_result["status"] = "matched"
    item_result["matched_product"] = top_row
    raw_description = top_row.get("descripcion") or top_row.get("nombre_articulo") or "producto"
    top_presentation = infer_product_presentation_from_row(top_row)
    top_brand = infer_product_brand_from_row(top_row)
    commercial_name = translate_product_to_commercial(raw_description, top_presentation, top_brand)
    stock_value = parse_numeric_value(top_row.get("stock_total") if top_row.get("stock_total") is not None else top_row.get("stock")) or 0
    requested_quantity = parse_numeric_value(product_request.get("requested_quantity"))

    if requested_store_label:
        if stock_value <= 0:
            item_result["message"] = f"❌ {commercial_name}: no disponible en {requested_store_label} en este momento."
        else:
            item_result["message"] = f"✅ {commercial_name}: disponible en {requested_store_label}"
            if requested_quantity:
                availability = ", te alcanza" if stock_value >= requested_quantity else ", pero no alcanza para toda la cantidad"
                item_result["message"] += availability
            item_result["message"] += "."
    else:
        if stock_value <= 0:
            item_result["message"] = f"❌ {commercial_name}: agotado en este momento."
        else:
            item_result["message"] = f"✅ {commercial_name}: disponible"
            if requested_quantity:
                availability = ", sí alcanza" if stock_value >= requested_quantity else ", pero no para toda la cantidad"
                item_result["message"] += availability
            item_result["message"] += "."

    return item_result


def format_draft_conversational(resolved_items: list[dict], store_label: Optional[str] = None):
    """Format the commercial draft as natural conversational text instead of numbered menus."""
    if not resolved_items:
        return "", False

    matched_labels = []
    ambiguous_parts = []
    missing_parts = []
    needs_input = False

    for item in resolved_items:
        pr = item.get("product_request") or {}
        qty = parse_numeric_value(pr.get("requested_quantity"))
        unit = pr.get("requested_unit")

        if item["status"] == "matched":
            mp = item.get("matched_product") or {}
            raw_desc = mp.get("descripcion") or mp.get("nombre_articulo") or "producto"
            pres = infer_product_presentation_from_row(mp) if mp else None
            brand = infer_product_brand_from_row(mp) if mp else None
            commercial_name = translate_product_to_commercial(raw_desc, pres, brand)
            if qty and unit:
                matched_labels.append(f"{format_quantity(qty)} {get_presentation_label(unit, qty)} de {commercial_name}")
            elif qty and qty > 1:
                matched_labels.append(f"{format_quantity(qty)} {commercial_name}")
            else:
                matched_labels.append(commercial_name)

        elif item["status"] == "ambiguous":
            needs_input = True
            orig = (item.get("original_text") or "").strip()
            alts = item.get("alternatives") or []
            alt_names = [a["commercial_name"] for a in alts[:4]]
            if len(alt_names) == 1:
                ambiguous_parts.append(f"Del *{orig}* tengo el {alt_names[0]}, ¿te sirve?")
            elif len(alt_names) == 2:
                ambiguous_parts.append(f"Del *{orig}* tengo el {alt_names[0]} y el {alt_names[1]}. ¿Cuál manejas?")
            else:
                options_text = ", ".join(alt_names[:-1]) + f" y {alt_names[-1]}"
                ambiguous_parts.append(f"Del *{orig}* tengo {options_text}. ¿Cuál necesitas?")

        elif item["status"] == "missing":
            needs_input = True
            orig = (item.get("original_text") or "").strip()
            missing_parts.append(orig)

    parts = []
    if matched_labels:
        if len(matched_labels) == 1:
            parts.append(f"✅ Te anoto {matched_labels[0]}.")
        elif len(matched_labels) == 2:
            parts.append(f"✅ Te anoto {matched_labels[0]} y {matched_labels[1]}.")
        else:
            items_text = ", ".join(matched_labels[:-1]) + f" y {matched_labels[-1]}"
            parts.append(f"✅ Te anoto {items_text}.")

    for amb in ambiguous_parts:
        parts.append(amb)

    if missing_parts:
        if len(missing_parts) == 1:
            parts.append(f"❌ No ubiqué *{missing_parts[0]}*, ¿me pasas la referencia o el código exacto?")
        else:
            items_text = " ni ".join(f"*{m}*" for m in missing_parts)
            parts.append(f"❌ No ubiqué {items_text}, ¿me pasas las referencias?")

    return "\n\n".join(parts), needs_input


def try_resolve_ambiguous_with_clarification(raw_line: str, existing_items: list[dict], inherited_store_filters: list[str], mode: str):
    """Try to match a clarification message to an existing ambiguous item and resolve it.

    Returns the index of the matched item, or None if no match found.
    """
    new_request = extract_product_request(raw_line)
    new_terms = set(new_request.get("core_terms") or [])
    new_codes = set(new_request.get("product_codes") or [])
    if not new_terms and not new_codes:
        return None

    best_idx = None
    best_score = 0

    for idx, item in enumerate(existing_items):
        if item.get("status") != "ambiguous":
            continue
        original_terms = set((item.get("product_request") or {}).get("core_terms") or [])
        # Check term overlap with original request
        overlap = len(new_terms & original_terms)
        # Also check if clarification matches any alternative's reference or name
        for alt in (item.get("alternatives") or []):
            alt_ref = normalize_text_value(alt.get("referencia") or "")
            alt_name = normalize_text_value(alt.get("commercial_name") or "")
            for term in new_terms:
                if term in alt_name:
                    overlap += 1
                    break
            for code in new_codes:
                if code == alt_ref or code in alt_ref:
                    overlap += 2
                    break
        if overlap > best_score:
            best_score = overlap
            best_idx = idx

    if best_idx is not None and best_score >= 1:
        return best_idx
    return None


def build_commercial_flow_reply(intent: str, profile_name: Optional[str], user_message: Optional[str], conversation_context: Optional[dict]):
    context = conversation_context or {}
    existing_draft = dict(context.get("commercial_draft") or {})
    last_intent = context.get("last_direct_intent")
    normalized_message = normalize_text_value(user_message)
    incoming_store_filters = extract_store_filters(user_message)
    incoming_email = extract_email_address(user_message)
    incoming_delivery_channel = extract_delivery_channel(user_message)
    inherited_store_filters = incoming_store_filters or existing_draft.get("store_filters") or []
    current_lines = split_commercial_line_items(user_message)
    has_existing_items = bool(existing_draft.get("items"))
    has_contextual_followup = bool(incoming_store_filters or incoming_email or incoming_delivery_channel)
    is_affirmative_followup = is_affirmative_message(user_message)
    is_negative_followup = is_negative_message(user_message)
    wants_order_confirmation = bool(re.search(r"\b(confirma|confirmame|conf[ií]rmame|conf[ií]rmalo|sep[aá]ralo|separalo|env[ií]ame|enviame|mandame|m[aá]ndame)\b", normalized_message))

    # ── Detect new order request ("otro pedido", "nuevo pedido") ──
    if is_new_order_request(user_message) and has_existing_items:
        summary_label = "cotización" if intent == "cotizacion" else "pedido"
        return {
            "tono": "consultivo",
            "intent": intent,
            "priority": "alta" if intent == "pedido" else "media",
            "summary": f"Nuevo {summary_label}",
            "response_text": (
                f"¡Dale! Arrancamos con un nuevo {summary_label}. "
                "Pásame los productos y la tienda o ciudad de entrega."
            ),
            "should_create_task": True,
            "task_type": intent,
            "task_summary": f"Nuevo {summary_label} iniciado por WhatsApp",
            "task_detail": {"mensaje": user_message, "mode": intent},
            "conversation_context_updates": {
                "commercial_draft": {
                    "intent": intent,
                    "store_filters": [],
                    "delivery_channel": None,
                    "contact_email": None,
                    "items": [],
                    "items_confirmed": False,
                }
            },
        }

    # ── Process option selections (e.g., "2a y 4b") ──
    option_updates = resolve_option_selections(user_message, existing_draft.get("items") or [])
    if option_updates and has_existing_items:
        draft_items_for_update = list(existing_draft.get("items") or [])
        for opt_idx, alt in option_updates.items():
            if 0 <= opt_idx < len(draft_items_for_update):
                old_item = draft_items_for_update[opt_idx]
                new_row = alt["row"]
                old_item["status"] = "matched"
                old_item["matched_product"] = new_row
                raw_desc = new_row.get("descripcion") or new_row.get("nombre_articulo") or "producto"
                pres = infer_product_presentation_from_row(new_row)
                brand = infer_product_brand_from_row(new_row)
                commercial_name = translate_product_to_commercial(raw_desc, pres, brand)
                old_item["message"] = f"{commercial_name}: ✅ seleccionado."
        existing_draft["items"] = draft_items_for_update
        has_existing_items = True
        current_lines = []

    items_confirmed = bool(existing_draft.get("items_confirmed"))

    has_explicit_items = any(
        extract_product_request(line).get("product_codes")
        or extract_product_request(line).get("core_terms")
        or extract_product_request(line).get("requested_unit")
        for line in current_lines
    )

    # Set items_confirmed when user explicitly confirms
    if is_affirmative_followup or wants_order_confirmation:
        items_confirmed = True

    if has_existing_items and (incoming_email or incoming_delivery_channel) and not incoming_store_filters:
        current_lines = []
        has_explicit_items = False

    if has_existing_items and (is_affirmative_followup or is_negative_followup or wants_order_confirmation) and not has_explicit_items and not has_contextual_followup and not option_updates:
        current_lines = []

    if not has_explicit_items and not has_existing_items and not has_contextual_followup and not option_updates:
        summary_label = "cotización" if intent == "cotizacion" else "pedido"
        intro_label = "la cotización" if intent == "cotizacion" else "el pedido"
        return {
            "tono": "consultivo",
            "intent": intent,
            "priority": "alta" if intent == "pedido" else "media",
            "summary": f"Inicio de {summary_label}",
            "response_text": (
                f"Con mucho gusto te ayudo con {intro_label}. "
                "Pásame las referencias o productos como los manejas normalmente y, si ya sabes la tienda o ciudad de entrega, me la dejas de una vez."
            ),
            "should_create_task": last_intent != intent,
            "task_type": intent,
            "task_summary": f"Solicitud de {summary_label} iniciada por WhatsApp",
            "task_detail": {"mensaje": user_message, "mode": intent},
            "conversation_context_updates": {
                "commercial_draft": {
                    "intent": intent,
                    "store_filters": inherited_store_filters,
                    "delivery_channel": incoming_delivery_channel,
                    "contact_email": incoming_email,
                    "items": existing_draft.get("items") or [],
                    "items_confirmed": False,
                }
            },
        }

    draft_items = []
    if existing_draft.get("intent") == intent:
        draft_items = list(existing_draft.get("items") or [])

    # When new explicit items are added, reset confirmation
    if has_explicit_items:
        items_confirmed = False

    if draft_items and incoming_store_filters and not has_explicit_items:
        current_lines = [item.get("original_text") for item in draft_items if item.get("original_text")]
        draft_items = []
    elif draft_items and not has_explicit_items and has_contextual_followup:
        current_lines = []

    if draft_items and len(current_lines) == 1 and (incoming_store_filters or extract_product_request(user_message).get("product_codes")):
        base_lines = [item.get("original_text") for item in draft_items if item.get("original_text")]
        if base_lines:
            refined_lines = list(base_lines)
            incoming_codes = extract_product_request(user_message).get("product_codes") or []
            if incoming_codes:
                refined_lines[0] = f"{refined_lines[0]} {user_message}".strip()
            current_lines = refined_lines
            draft_items = []

    if draft_items and len(current_lines) == 1 and extract_product_request(user_message).get("product_codes"):
        unresolved_index = next((index for index, item in enumerate(draft_items) if item.get("status") != "matched"), None)
        if unresolved_index is not None:
            base_text = draft_items[unresolved_index].get("original_text") or ""
            current_lines = [f"{base_text} {user_message}".strip()]
            draft_items = [item for index, item in enumerate(draft_items) if index != unresolved_index]

    resolved_items = list(draft_items)
    for raw_line in current_lines:
        if not raw_line:
            continue
        raw_request = extract_product_request(raw_line)
        if not (
            raw_request.get("product_codes")
            or raw_request.get("core_terms")
            or raw_request.get("requested_unit")
            or is_product_intent_message(raw_line)
        ):
            continue
        # Try to resolve an existing ambiguous item with this clarification
        matched_idx = try_resolve_ambiguous_with_clarification(raw_line, resolved_items, inherited_store_filters, intent)
        if matched_idx is not None:
            original_text = resolved_items[matched_idx].get("original_text")
            resolved_items[matched_idx] = build_commercial_item_result(raw_line, inherited_store_filters, intent)
            resolved_items[matched_idx]["original_text"] = original_text or raw_line
        else:
            resolved_items.append(build_commercial_item_result(raw_line, inherited_store_filters, intent))

    matched_items = [item for item in resolved_items if item.get("status") == "matched"]
    ambiguous_items = [item for item in resolved_items if item.get("status") == "ambiguous"]
    missing_items = [item for item in resolved_items if item.get("status") == "missing"]
    store_label = STORE_CODE_LABELS.get(inherited_store_filters[0]) if len(inherited_store_filters) == 1 else None
    delivery_channel = incoming_delivery_channel or existing_draft.get("delivery_channel")
    contact_email = incoming_email or existing_draft.get("contact_email")
    internal_notified = bool(existing_draft.get("internal_notified"))
    customer_email_sent = bool(existing_draft.get("customer_email_sent"))
    destinatario = existing_draft.get("destinatario") or ""
    customer_identity_input = (existing_draft.get("customer_identity_input") or "").strip()
    customer_context = dict(existing_draft.get("customer_context") or {})
    customer_resolution_status = existing_draft.get("customer_resolution_status")

    # Extract destinatario from message like "a nombre de Juan Pérez"
    nombre_match = re.search(r"\ba\s+nombre\s+de\s+(.+?)(?:\s*[.,;]|$)", normalized_message)
    if nombre_match:
        customer_identity_input = trim_commercial_customer_candidate(nombre_match.group(1))
        resolved_customer_context, _ = resolve_commercial_customer_context(customer_identity_input)
        if resolved_customer_context:
            customer_context = resolved_customer_context
            customer_resolution_status = "resolved"
            destinatario = resolved_customer_context.get("nombre_cliente") or customer_identity_input.title()
        else:
            customer_context = {}
            customer_resolution_status = "unresolved"
            destinatario = customer_identity_input.title()
    elif customer_identity_input and not customer_context:
        resolved_customer_context, _ = resolve_commercial_customer_context(customer_identity_input)
        if resolved_customer_context:
            customer_context = resolved_customer_context
            customer_resolution_status = "resolved"
            destinatario = resolved_customer_context.get("nombre_cliente") or destinatario or customer_identity_input.title()
        elif customer_resolution_status != "resolved":
            customer_resolution_status = "unresolved"

    if customer_context and not destinatario:
        destinatario = customer_context.get("nombre_cliente") or ""

    compact_summary = summarize_commercial_items(matched_items)
    has_store = bool(inherited_store_filters)
    all_items_resolved = bool(matched_items) and not ambiguous_items and not missing_items
    requires_customer_resolution = bool(customer_identity_input) and customer_resolution_status != "resolved"
    ready_to_close = all_items_resolved and has_store and items_confirmed and not requires_customer_resolution

    if ready_to_close and not incoming_delivery_channel and wants_order_confirmation:
        delivery_channel = existing_draft.get("delivery_channel")

    if not ready_to_close:
        # ── Format response conversationally ──
        list_text, has_options = format_draft_conversational(resolved_items, store_label)

        closing_parts = []
        if not has_store:
            closing_parts.append("¿En qué tienda o ciudad lo necesitas?")
        if requires_customer_resolution:
            closing_parts.append("Antes de cerrarlo necesito validar a nombre de qué cliente va. Envíame NIT, código o nombre completo para no cruzar la facturación.")
        if all_items_resolved and has_store and not items_confirmed:
            request_label = "cotización" if intent == "cotizacion" else "pedido"
            closing_parts.append(f"¿Te confirmo el {request_label}? ¿A nombre de quién va el despacho?")
        elif all_items_resolved and not has_store:
            pass  # Already asked for store
        elif not has_options and not missing_items:
            closing_parts.append("Apenas me confirmes te armo el consolidado completo.")

        response_text = list_text
        if closing_parts:
            response_text += "\n\n" + " ".join(closing_parts)

        draft_state = {
            "intent": intent,
            "store_filters": inherited_store_filters,
            "delivery_channel": delivery_channel,
            "contact_email": contact_email,
            "internal_notified": internal_notified,
            "customer_email_sent": customer_email_sent,
            "items_confirmed": items_confirmed,
            "destinatario": destinatario,
            "customer_identity_input": customer_identity_input or None,
            "customer_context": customer_context or None,
            "customer_resolution_status": customer_resolution_status,
            "items": resolved_items,
        }

        return {
            "tono": "consultivo",
            "intent": intent,
            "priority": "alta" if intent == "pedido" else "media",
            "summary": "Consolidado comercial multiproducto",
            "response_text": response_text.strip(),
            "should_create_task": last_intent != intent,
            "task_type": intent,
            "task_summary": "Seguimiento a solicitud comercial por WhatsApp",
            "task_detail": {"items": resolved_items, "store_filters": inherited_store_filters, "mode": intent},
            "conversation_context_updates": {"commercial_draft": draft_state},
            "commercial_draft": draft_state,
        }

    # ── Ready to close ──
    request_label = "cotización" if intent == "cotizacion" else "pedido"
    destination_label = store_label or "la sede indicada"
    display_customer_label = (customer_context or {}).get("nombre_cliente") or destinatario
    destinatario_label = f" a nombre de {display_customer_label}" if display_customer_label else ""
    if not delivery_channel:
        response_text = (
            f"¡Listo! Ya te dejé montado el {request_label} para {destination_label}{destinatario_label} con {compact_summary}. "
            "¿Te lo confirmo por aquí o prefieres que te envíe un PDF al correo?"
        )
    elif delivery_channel == "email" and not contact_email:
        response_text = (
            f"¡Listo! Ya tengo el {request_label} para {destination_label}{destinatario_label} con {compact_summary}. "
            "Regálame tu correo y te mando el PDF con todo el detalle."
        )
    elif delivery_channel == "email":
        response_text = (
            f"¡Listo! Ya te dejé montado el {request_label} para {destination_label}{destinatario_label} con {compact_summary}. "
            f"Te va a llegar al correo {contact_email} un PDF con el detalle de las referencias y cantidades."
        )
    else:
        response_text = (
            f"¡Listo! Ya te dejé montado el {request_label} para {destination_label}{destinatario_label} con {compact_summary}. "
            "Te envío el PDF por aquí mismo para que lo tengas de referencia 📄"
        )

    final_confirmation_ready = ready_to_close and (delivery_channel == "chat" or (delivery_channel == "email" and bool(contact_email)))
    should_notify_internal = final_confirmation_ready and not internal_notified
    should_send_customer_email = final_confirmation_ready and delivery_channel == "email" and bool(contact_email) and not customer_email_sent
    draft_state = {
        "intent": intent,
        "store_filters": inherited_store_filters,
        "delivery_channel": delivery_channel,
        "contact_email": contact_email,
        "ready_to_close": ready_to_close,
        "internal_notified": internal_notified or should_notify_internal,
        "customer_email_sent": customer_email_sent or should_send_customer_email,
        "items_confirmed": items_confirmed,
        "destinatario": destinatario,
        "customer_identity_input": customer_identity_input or None,
        "customer_context": customer_context or None,
        "customer_resolution_status": customer_resolution_status,
        "items": resolved_items,
    }

    return {
        "tono": "consultivo",
        "intent": intent,
        "priority": "alta" if intent == "pedido" else "media",
        "summary": "Consolidado comercial multiproducto",
        "response_text": response_text,
        "should_create_task": last_intent != intent or should_notify_internal,
        "task_type": intent,
        "task_summary": f"Solicitud de {request_label} lista para seguimiento",
        "task_detail": {"items": resolved_items, "store_filters": inherited_store_filters, "mode": intent, "delivery_channel": delivery_channel, "contact_email": contact_email, "destinatario": destinatario, "customer_context": customer_context or None},
        "conversation_context_updates": {"commercial_draft": draft_state},
        "commercial_draft": draft_state,
        "email_route": "ventas" if should_notify_internal else None,
        "email_detail": draft_state if should_notify_internal else None,
        "commercial_customer_email_confirmation": draft_state if should_send_customer_email else None,
    }


def build_technical_document_reply(profile_name: Optional[str], document_request: dict, document_options: list[dict]):
    requested_label = "hoja de seguridad" if document_request.get("wants_safety_sheet") else "ficha técnica"
    if not document_options:
        return {
            "tono": "informativo",
            "intent": "consulta_documentacion",
            "priority": "media",
            "summary": "Consulta de documentación sin coincidencia clara",
            "response_text": (
                f"Revisé la carpeta de {requested_label} y no encontré una coincidencia clara con ese nombre. "
                "Si quieres, dime la referencia, la línea o el nombre comercial y te muestro las opciones más cercanas."
            ),
            "document_options": [],
            "awaiting_document_choice": False,
        }

    option_lines = [f"{index}. {row['name']}" for index, row in enumerate(document_options[:4], start=1)]
    intro = f"Esto fue lo que encontré en {requested_label}:"
    outro = "Respóndeme con el número o con el nombre del archivo y te lo envío por WhatsApp."
    return {
        "tono": "consultivo",
        "intent": "consulta_documentacion",
        "priority": "media",
        "summary": "Consulta de documentación técnica",
        "response_text": intro + "\n" + "\n".join(option_lines) + "\n" + outro,
        "document_options": document_options[:4],
        "awaiting_document_choice": True,
    }


def find_technical_document_entry_by_name(filename: Optional[str]):
    if not filename:
        return None
    normalized_target = normalize_text_value(filename)
    for entry in list_technical_document_entries():
        if normalize_text_value(entry.get("name")) == normalized_target:
            return entry
    return None


def infer_technical_problem_category(text_value: Optional[str], existing_category: Optional[str] = None):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return existing_category or "general"
    if any(token in normalized for token in ["piso", "pisos", "cemento", "concreto", "pintura para piso", "epoxica", "epóxica"]):
        return "piso"
    if any(token in normalized for token in ["humedad", "gotera", "goteras", "filtracion", "filtración", "capilaridad", "moho", "salitre", "descascar", "manchas negras", "barranco"]):
        return "humedad"
    if any(token in normalized for token in ["madera", "barniz", "laca", "lasur", "protector madera"]):
        return "madera"
    if any(token in normalized for token in ["metal", "hierro", "galvanizado", "aluminio", "corrosion", "corrosión", "oxido", "óxido", "anticorrosivo"]):
        return "metal"
    return existing_category or "general"


def _merge_unique_text_values(existing_values: Optional[list[str]], new_values: list[str]) -> list[str]:
    merged = list(existing_values or [])
    for value in new_values:
        if value and value not in merged:
            merged.append(value)
    return merged


def extract_technical_advisory_case(text_value: Optional[str], conversation_context: Optional[dict]):
    case = dict((conversation_context or {}).get("technical_advisory_case") or {})
    normalized = normalize_text_value(text_value)
    category = infer_technical_problem_category(text_value, case.get("category"))

    # Track conversation history for RAG search context
    history = list(case.get("conversation_history") or [])
    if text_value and text_value.strip():
        history.append(text_value.strip())
    history = history[-10:]

    # Track diagnostic turns
    diagnostic_turns = case.get("diagnostic_turns", 0)
    if case.get("stage") == "diagnosing":
        diagnostic_turns += 1

    case.update({
        "active": True,
        "category": category,
        "last_user_message": text_value or "",
        "conversation_history": history,
        "diagnostic_turns": diagnostic_turns,
    })

    # --- Category-specific field extraction (enriches RAG search) ---
    if category == "humedad":
        if any(token in normalized for token in ["barranco", "terreno", "talud", "contencion", "contención"]):
            case["source_context"] = "muro contra terreno o barranco"
            case["probable_pressure"] = "presion_negativa"
        elif any(token in normalized for token in ["tuberia", "tubería", "tubo", "fuga"]):
            case["source_context"] = "posible tuberia o fuga interna"
        elif any(token in normalized for token in ["fachada", "lluvia", "exterior", "afuera"]):
            case["source_context"] = "posible filtracion desde fachada o exterior"

        if any(token in normalized for token in ["interior", "adentro", "dentro de la casa", "casa"]):
            case["wall_location"] = "interior"
        elif any(token in normalized for token in ["exterior", "fachada", "afuera"]):
            case["wall_location"] = "exterior"

        if "obra negra" in normalized:
            case["surface_state"] = "obra negra"
        elif any(token in normalized for token in ["pintada", "pintado", "pintura"]):
            case["surface_state"] = "pintada"
        elif any(token in normalized for token in ["estuco", "estucada", "estucado"]):
            case["surface_state"] = "estucada"

        symptoms = []
        if any(token in normalized for token in ["descascar", "se cae la pintura", "pintura caida", "pintura caída"]):
            symptoms.append("descascaramiento")
        if any(token in normalized for token in ["manchas negras", "negra", "negras", "moho", "hongo"]):
            symptoms.append("manchas negras o moho")
        if any(token in normalized for token in ["humeda", "húmeda", "humedo", "húmedo"]):
            symptoms.append("superficie humeda")
        if any(token in normalized for token in ["salitre", "polvillo blanco", "blanquea"]):
            symptoms.append("salitre")
        case["symptoms"] = _merge_unique_text_values(case.get("symptoms"), symptoms)

    elif category == "piso":
        if any(token in normalized for token in ["interior", "adentro", "bajo techo"]):
            case["floor_location"] = "interior"
        elif any(token in normalized for token in ["exterior", "afuera", "intemperie"]):
            case["floor_location"] = "exterior"

        if any(token in normalized for token in ["cemento", "concreto", "mortero"]):
            case["floor_material"] = "cemento o concreto"
        elif any(token in normalized for token in ["ceramica", "cerámica", "baldosa", "porcelanato"]):
            case["floor_material"] = "ceramica o baldosa"
        elif any(token in normalized for token in ["madera"]):
            case["floor_material"] = "madera"

        if any(token in normalized for token in ["alto trafico", "alto tráfico", "montacarga", "vehiculo", "vehículo", "carro"]):
            case["traffic_level"] = "alto trafico"
        elif any(token in normalized for token in ["peatonal", "residencial", "casa", "habitacion", "habitación"]):
            case["traffic_level"] = "trafico peatonal o residencial"

    elif category == "madera":
        if any(token in normalized for token in ["intemperie", "exterior", "sol", "lluvia"]):
            case["exposure"] = "intemperie"
        elif any(token in normalized for token in ["interior", "bajo techo", "adentro"]):
            case["exposure"] = "bajo techo"

        if any(token in normalized for token in ["barniz", "laca", "pintada", "pintado", "tiene recubrimiento", "ya tiene"]):
            case["previous_coating"] = "con recubrimiento previo"
        elif any(token in normalized for token in ["virgen", "sin pintar", "sin barniz", "madera nueva"]):
            case["previous_coating"] = "sin recubrimiento previo"

    elif category == "metal":
        if any(token in normalized for token in ["hierro", "ferroso", "acero al carbon", "acero al carbón"]):
            case["metal_type"] = "metal ferroso"
        elif "galvanizado" in normalized:
            case["metal_type"] = "galvanizado"
        elif "aluminio" in normalized:
            case["metal_type"] = "aluminio"

        if any(token in normalized for token in ["marino", "mar", "costa", "playa"]):
            case["environment"] = "marino"
        elif any(token in normalized for token in ["industrial", "quimico", "químico", "planta"]):
            case["environment"] = "industrial"
        elif any(token in normalized for token in ["urbano", "ciudad", "residencial"]):
            case["environment"] = "urbano"

    # --- Universal readiness check ---
    # Known categories: ready when key fields are filled
    category_fields_ready = False
    if category == "humedad":
        category_fields_ready = bool(case.get("source_context") and case.get("surface_state") and case.get("symptoms"))
    elif category == "piso":
        category_fields_ready = bool(case.get("floor_location") and case.get("floor_material"))
    elif category == "madera":
        category_fields_ready = bool(case.get("exposure") and case.get("previous_coating"))
    elif category == "metal":
        category_fields_ready = bool(case.get("metal_type") and case.get("environment"))

    # Ready if category fields are complete OR at least 1 diagnostic exchange done
    case["ready"] = category_fields_ready or diagnostic_turns >= 1

    return case


def build_technical_diagnostic_questions(technical_case: dict) -> list[str]:
    category = technical_case.get("category")
    questions: list[str] = []

    # Fast path for known categories
    if category == "humedad":
        if not technical_case.get("source_context"):
            questions.append("¿Esa pared te da contra terreno, barranco, fachada o crees que viene de una tubería?")
        if not technical_case.get("surface_state"):
            questions.append("¿La pared está pintada, estucada o en obra negra?")
        if not technical_case.get("symptoms"):
            questions.append("¿Qué síntoma ves más claro: se descascara, sale moho, blanquea o solo se siente húmeda?")
        if not technical_case.get("wall_location"):
            questions.append("¿Eso te está pasando por la cara interior del muro o por la exterior?")
    elif category == "piso":
        if not technical_case.get("floor_location"):
            questions.append("¿Ese piso es interior o exterior?")
        if not technical_case.get("floor_material"):
            questions.append("¿El piso es de cemento, concreto, cerámica o madera?")
        if technical_case.get("floor_location") and technical_case.get("floor_material") and not technical_case.get("traffic_level"):
            questions.append("¿Ese piso va a tener tráfico peatonal residencial o un uso más pesado?")
    elif category == "madera":
        if not technical_case.get("exposure"):
            questions.append("¿Esa madera va a quedar a la intemperie o bajo techo?")
        if not technical_case.get("previous_coating"):
            questions.append("¿La madera ya tiene barniz/laca encima o está virgen?")
    elif category == "metal":
        if not technical_case.get("metal_type"):
            questions.append("¿Ese metal es hierro/acero, galvanizado o aluminio?")
        if not technical_case.get("environment"):
            questions.append("¿Va a trabajar en ambiente urbano, industrial o marino?")

    if questions:
        return questions[:2]

    # LLM-powered diagnostic questions for general/unknown categories
    conversation_history = technical_case.get("conversation_history") or []
    context_text = "\n".join(f"- {msg}" for msg in conversation_history) if conversation_history else technical_case.get("last_user_message", "")
    try:
        client = get_openai_client()
        response = client.chat.completions.create(
            model=get_openai_model(),
            temperature=0,
            max_tokens=200,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Eres un asesor técnico experto en pinturas, recubrimientos, impermeabilizantes, selladores, esmaltes, barnices, "
                        "anticorrosivos y productos para construcción y mantenimiento de superficies. "
                        "Tu tarea: generar EXACTAMENTE 2 preguntas diagnósticas cortas y concretas que te ayuden a recomendar el producto "
                        "o sistema correcto para el problema del cliente. "
                        "Las preguntas deben ser específicas al problema descrito, en tono coloquial colombiano directo. "
                        "Responde SOLO con las 2 preguntas, una por línea, sin numeración, viñetas ni explicación adicional."
                    ),
                },
                {
                    "role": "user",
                    "content": f"El cliente dice:\n{context_text}\n\nGenera 2 preguntas diagnósticas:",
                },
            ],
        )
        raw = (response.choices[0].message.content or "").strip()
        llm_questions = [q.strip().lstrip("0123456789.-) ●•→ ") for q in raw.split("\n") if q.strip() and "?" in q]
        if llm_questions:
            return llm_questions[:2]
    except Exception:
        pass

    return ["¿Qué material o superficie vas a tratar exactamente?", "¿Eso es interior, exterior o a la intemperie?"]


def build_technical_search_query(technical_case: dict, user_message: Optional[str] = None) -> str:
    category = technical_case.get("category") or "general"
    parts: list[str] = []

    # Always include conversation history for richer semantic search
    conversation_history = technical_case.get("conversation_history") or []
    history_text = " ".join(conversation_history[-5:]) if conversation_history else ""

    if category == "humedad":
        parts.extend([
            "humedad en muro",
            technical_case.get("source_context") or "",
            "presion negativa" if technical_case.get("probable_pressure") == "presion_negativa" else "",
            technical_case.get("wall_location") or "",
            technical_case.get("surface_state") or "",
            " ".join(technical_case.get("symptoms") or []),
        ])
    elif category == "piso":
        parts.extend([
            "sistema pintura para piso",
            technical_case.get("floor_location") or "",
            technical_case.get("floor_material") or "",
            technical_case.get("traffic_level") or "",
        ])
    elif category == "madera":
        parts.extend([
            "sistema para madera",
            technical_case.get("exposure") or "",
            technical_case.get("previous_coating") or "",
        ])
    elif category == "metal":
        parts.extend([
            "sistema anticorrosivo",
            technical_case.get("metal_type") or "",
            technical_case.get("environment") or "",
        ])
    else:
        # General: use full conversation context as semantic query
        parts.append(history_text or user_message or technical_case.get("last_user_message") or "")

    query = " ".join(part for part in parts if part).strip()
    # Enrich known categories with conversation context too
    if category != "general" and history_text:
        query = f"{query} {history_text}".strip()
    return query or (user_message or technical_case.get("last_user_message") or "")


def extract_area_square_meters(text_value: Optional[str]) -> Optional[float]:
    if not text_value:
        return None
    normalized = normalize_text_value(text_value).replace(",", ".")
    match = re.search(r"(\d+(?:\.\d+)?)\s*(m2|m\^2|mts2|mts cuadrados|metros cuadrados)", normalized)
    if match:
        try:
            return float(match.group(1))
        except Exception:
            return None
    return None


def is_coverage_followup_question(text_value: Optional[str]) -> bool:
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    return any(
        phrase in normalized
        for phrase in [
            "cuanto necesito",
            "cuantos necesito",
            "cuánto necesito",
            "cuántos necesito",
            "para cubrir",
            "para 10 m",
            "cuanto me alcanza",
            "cuánto me alcanza",
        ]
    )


def _derive_portfolio_candidates_from_question(question: str) -> list[str]:
    """Derive product candidates from a question using PORTFOLIO_CATEGORY_MAP.

    This ensures the right products are ALWAYS suggested as candidates even
    when RAG returns irrelevant chunks (e.g. safety data sheets instead of
    technical sheets).  The products come from the curated portfolio map.
    """
    if not question:
        return []
    q_norm = normalize_text_value(question)
    candidates: list[str] = []
    seen: set[str] = set()
    # Full-phrase matching against category keys
    for category_key, brand_terms in PORTFOLIO_CATEGORY_MAP.items():
        if category_key in q_norm or q_norm in category_key:
            for bt in brand_terms:
                if bt != "__SIN_PRODUCTO_FERREINOX__" and bt not in seen:
                    seen.add(bt)
                    candidates.append(bt)
    # Word-level matching
    _SKIP_WORDS = {"para", "como", "esto", "esta", "esos", "esas", "unos", "unas",
                   "tiene", "cada", "todo", "toda", "estos", "estas", "necesito",
                   "quiero", "pintar", "casa", "esta"}
    for word in q_norm.split():
        if len(word) < 4 or word in _SKIP_WORDS:
            continue
        if word in PORTFOLIO_CATEGORY_MAP:
            for bt in PORTFOLIO_CATEGORY_MAP[word]:
                if bt != "__SIN_PRODUCTO_FERREINOX__" and bt not in seen:
                    seen.add(bt)
                    candidates.append(bt)
    return candidates


def extract_candidate_products_from_rag_context(
    rag_context: str,
    source_file: Optional[str] = None,
    original_question: str = "",
) -> list[str]:
    candidates: list[str] = []
    # A) Extract explicitly tagged products from RAG chunks
    # Skip FDS/HDS (safety data sheets) — they match broadly but are not
    # product recommendations. Only keep FT (ficha técnica) product tags.
    for match in re.finditer(r"\[PRODUCTO:\s*([^\]]+)\]", rag_context or "", flags=re.IGNORECASE):
        candidate = match.group(1).strip()
        if candidate and candidate not in candidates:
            candidate_upper = candidate.upper()
            if candidate_upper.startswith("FDS") or candidate_upper.startswith("HDS"):
                continue
            candidates.append(candidate)
    # B) Extract brand/product names mentioned in the RAG text that match known portfolio
    if rag_context:
        rag_lower = normalize_text_value(rag_context)
        _KNOWN_PRODUCT_NAMES = [
            # Pintuco líneas principales
            "pintucoat", "pintura canchas", "corrotec", "pintulac", "aerocolor", "koraza",
            "viniltex", "pintulux", "domestico", "pinturama", "intervinil", "vinil latex",
            "vinilux", "vinil max", "icolatex", "vinil plus", "pintacrom",
            "pintuco fill", "world color", "wash primer", "imprimante",
            "pintoxido", "pintura trafico", "barniz marino", "barnex", "wood stain",
            "estuco anti humedad", "impercoat", "tela de refuerzo",
            "pintura cielos", "pintuobra", "aislante", "emulsion asfaltica",
            "viniltex banos y cocinas", "viniltex advanced", "viniltex ultralavable",
            "madetec", "construmastic", "pintulac nitro",
            # International / AkzoNobel
            "interseal", "interthane", "intergard", "interfine", "interchar",
            # Impermeabilizantes / selladores
            "aquablock", "aquablock ultra", "sellamur", "siliconite", "sika",
            "koraza elastomerica", "koraza xp", "koraza sol y lluvia",
        ]
        for product_name in _KNOWN_PRODUCT_NAMES:
            if product_name in rag_lower and product_name not in candidates:
                candidates.append(product_name)
    # C) Inject portfolio-derived candidates from the original question
    # This ensures correct products appear even when RAG returns wrong chunks
    if original_question:
        portfolio_candidates = _derive_portfolio_candidates_from_question(original_question)
        for pc in portfolio_candidates:
            if pc not in candidates:
                candidates.append(pc)
    if source_file:
        normalized_file = re.sub(r"\.pdf$", "", source_file, flags=re.IGNORECASE).strip()
        normalized_file = re.sub(r"\s*\(.*?\)\s*", " ", normalized_file).strip()
        if normalized_file and normalized_file not in candidates:
            candidates.insert(0, normalized_file)
    return candidates[:12]


def _expand_terms_with_portfolio_knowledge(terms: list[str]) -> list[str]:
    """Expand generic product/category terms into brand-specific inventory search terms.

    Uses PORTFOLIO_CATEGORY_MAP and PORTFOLIO_ALIASES to translate generic
    use-case language (e.g. 'pintura para piscinas') into real brand names
    that exist in the Ferreinox inventory (e.g. 'pintucoat', 'koraza').
    """
    expanded: list[str] = []
    seen: set[str] = set()
    for term in terms:
        norm = normalize_text_value(term)
        if not norm:
            continue
        if norm not in seen:
            seen.add(norm)
            expanded.append(norm)
        # Check PORTFOLIO_CATEGORY_MAP for category-based expansion
        for category_key, brand_terms in PORTFOLIO_CATEGORY_MAP.items():
            if category_key in norm or norm in category_key:
                for bt in brand_terms:
                    if bt == "__SIN_PRODUCTO_FERREINOX__":
                        continue
                    if bt not in seen:
                        seen.add(bt)
                        expanded.append(bt)
        # Check individual words against category map
        for word in norm.split():
            if len(word) < 4:
                continue
            if word in PORTFOLIO_CATEGORY_MAP:
                for bt in PORTFOLIO_CATEGORY_MAP[word]:
                    if bt == "__SIN_PRODUCTO_FERREINOX__":
                        continue
                    if bt not in seen:
                        seen.add(bt)
                        expanded.append(bt)
        # Check PORTFOLIO_ALIASES for alias expansion
        norm_ref = normalize_reference_value(term)
        if norm_ref in PORTFOLIO_ALIASES:
            for alias in PORTFOLIO_ALIASES[norm_ref]:
                alias_norm = normalize_text_value(alias)
                if alias_norm and alias_norm not in seen:
                    seen.add(alias_norm)
                    expanded.append(alias_norm)
    return expanded


def lookup_inventory_candidates_from_terms(terms: list[str], conversation_context: Optional[dict]) -> list[dict]:
    seen_codes = set()
    resolved: list[dict] = []
    local_context = dict(conversation_context or {})

    # First pass: search with original terms
    for term in terms:
        if not term:
            continue
        rows = lookup_product_context(term, prepare_product_request_for_search(term))
        for row in rows[:2]:
            code = row.get("codigo_articulo") or row.get("referencia") or row.get("codigo")
            if not code or code in seen_codes:
                continue
            seen_codes.add(code)
            resolved.append(
                {
                    "codigo": code,
                    "descripcion": get_exact_product_description(row),
                    "etiqueta_auditable": build_product_audit_label(row),
                    "marca": row.get("marca") or row.get("marca_producto"),
                    "presentacion": infer_product_presentation_from_row(row),
                    "stock_total": parse_numeric_value(row.get("stock_total")),
                    "precio": row.get("precio_venta"),
                    "productos_complementarios": [
                        {
                            "referencia": c.get("companion_referencia"),
                            "descripcion": c.get("companion_descripcion") or c.get("descripcion_inventario"),
                            "tipo": c.get("tipo_relacion"),
                            "proporcion": c.get("proporcion"),
                        }
                        for c in fetch_product_companions(code)
                    ],
                }
            )
            local_context["last_product_query"] = term
        if len(resolved) >= 4:
            break

    # Second pass: if first pass found nothing, expand terms using portfolio knowledge
    if not resolved:
        expanded_terms = _expand_terms_with_portfolio_knowledge(terms)
        # Remove terms already tried in first pass
        original_normalized = {normalize_text_value(t) for t in terms if t}
        new_terms = [t for t in expanded_terms if t not in original_normalized]
        for term in new_terms:
            if not term:
                continue
            rows = lookup_product_context(term, prepare_product_request_for_search(term))
            for row in rows[:2]:
                code = row.get("codigo_articulo") or row.get("referencia") or row.get("codigo")
                if not code or code in seen_codes:
                    continue
                seen_codes.add(code)
                resolved.append(
                    {
                        "codigo": code,
                        "descripcion": get_exact_product_description(row),
                        "etiqueta_auditable": build_product_audit_label(row),
                        "marca": row.get("marca") or row.get("marca_producto"),
                        "presentacion": infer_product_presentation_from_row(row),
                        "stock_total": parse_numeric_value(row.get("stock_total")),
                        "precio": row.get("precio_venta"),
                        "productos_complementarios": [
                            {
                                "referencia": c.get("companion_referencia"),
                                "descripcion": c.get("companion_descripcion") or c.get("descripcion_inventario"),
                                "tipo": c.get("tipo_relacion"),
                                "proporcion": c.get("proporcion"),
                            }
                            for c in fetch_product_companions(code)
                        ],
                    }
                )
                local_context["last_product_query"] = term
            if len(resolved) >= 4:
                break

    return resolved[:4]


def format_inventory_product_block(products: list[dict]) -> str:
    if not products:
        return ""
    lines = ["Vea, los productos que necesitas son estos:"]
    for product in products[:4]:
        lines.append(f"- ✅ {product.get('etiqueta_auditable')}: Disponible")
        companions = product.get("productos_complementarios") or []
        for companion in companions[:2]:
            label = companion.get("descripcion") or companion.get("referencia")
            companion_type = companion.get("tipo") or "complemento"
            if label:
                extra = f" ({companion.get('proporcion')})" if companion.get("proporcion") else ""
                lines.append(f"- Complementario {companion_type}: {label}{extra}")
    return "\n".join(lines)


def generate_grounded_technical_sales_reply(
    technical_case: dict,
    rag_context: str,
    user_message: Optional[str],
    inventory_products: list[dict],
    area_m2: Optional[float] = None,
) -> str:
    client = get_openai_client()
    case_summary = safe_json_dumps(technical_case)
    inventory_summary = safe_json_dumps(inventory_products)
    conversation_history = technical_case.get("conversation_history") or []
    history_text = "\n".join(f"- {msg}" for msg in conversation_history) if conversation_history else ""
    area_text = f"Área a cubrir: {area_m2} m2" if area_m2 else "Área a cubrir: no especificada"
    response = client.chat.completions.create(
        model=get_openai_model(),
        temperature=0.1,
        messages=[
            {
                "role": "system",
                "content": (
                    "Eres el asesor técnico y comercial senior de Ferreinox, experto en pinturas, recubrimientos, "
                    "impermeabilizantes, selladores, esmaltes, barnices, anticorrosivos y todo producto para "
                    "construcción y mantenimiento de superficies. "
                    "Responde SOLO con base en el contexto recuperado de fichas técnicas y el inventario suministrado. "
                    "Objetivo: recomendar el sistema o producto adecuado según el diagnóstico del cliente y cerrar la venta con productos reales del portafolio. "
                    "REGLAS: "
                    "1) No inventes rendimientos, número de galones, pasos, manos, catalizadores o productos que no estén en el contexto recuperado. "
                    "2) Si el rendimiento exacto no aparece, dilo de frente y NO calcules cantidades. "
                    "3) Si sí hay respaldo suficiente, explica el sistema o solución en lenguaje claro y luego termina con una sección final que empiece EXACTAMENTE con: 'Vea, los productos que necesitas son estos:'. "
                    "4) En esa sección final solo puedes listar productos del inventario suministrado. "
                    "5) PROHIBIDO mandar al cliente a otra parte, sugerir buscar fuera del portafolio, o decir que no tienes el producto sin verificar el inventario. "
                    "6) Usa tono colombiano, directo, útil y comercial. Máximo 8 líneas antes de la lista de productos."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Caso diagnosticado: {case_summary}\n"
                    f"Historial de conversación del caso:\n{history_text}\n"
                    f"Mensaje actual del cliente: {user_message or technical_case.get('last_user_message') or ''}\n"
                    f"{area_text}\n"
                    f"Inventario candidato real: {inventory_summary}\n\n"
                    f"Contexto recuperado de fichas/FDS:\n{rag_context}"
                ),
            },
        ],
    )
    return (response.choices[0].message.content or "").strip()


def generate_grounded_technical_reply(technical_case: dict, rag_context: str, user_message: Optional[str] = None) -> str:
    client = get_openai_client()
    case_summary = safe_json_dumps(technical_case)
    response = client.chat.completions.create(
        model=get_openai_model(),
        temperature=0.1,
        messages=[
            {
                "role": "system",
                "content": (
                    "Eres el asesor técnico senior de Ferreinox. Responde SOLO con base en el contexto recuperado. "
                    "PROHIBIDO inventar productos, rendimientos, pasos o proporciones que no estén en el texto. "
                    "No menciones inventario ni listas genéricas. Si falta un dato exacto, dilo claramente. "
                    "Redacta máximo 6 líneas, en tono colombiano, útil y comercial. "
                    "Si el contexto menciona un sistema o producto concreto, explícalo. Si no lo menciona, no lo inventes. "
                    "Cierra preguntando si quiere que le cotices el sistema correcto."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Caso diagnosticado: {case_summary}\n"
                    f"Mensaje actual del cliente: {user_message or technical_case.get('last_user_message') or ''}\n\n"
                    f"Contexto recuperado de fichas/FDS:\n{rag_context}"
                ),
            },
        ],
    )
    return (response.choices[0].message.content or "").strip()


def should_continue_technical_advisory_flow(conversation_context: Optional[dict], detected_intent: Optional[str], text_value: Optional[str]):
    technical_case = dict((conversation_context or {}).get("technical_advisory_case") or {})
    if not technical_case.get("active"):
        return False
    # Only exit if the user clearly switches to a different structured intent
    if detected_intent in {"pedido", "cotizacion", "consulta_cartera", "consulta_compras", "consulta_documentacion", "reclamo_servicio"}:
        return False
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    return True


def build_technical_advisory_flow_reply(profile_name: Optional[str], user_message: Optional[str], conversation_context: Optional[dict]):
    technical_case = extract_technical_advisory_case(user_message, conversation_context)
    area_m2 = extract_area_square_meters(user_message)
    if area_m2:
        technical_case["area_m2"] = area_m2
    if not technical_case.get("ready"):
        questions = build_technical_diagnostic_questions(technical_case)
        category = technical_case.get("category")
        intro_map = {
            "humedad": "Claro. Para no mandarte a comprar algo que no te sirva, necesito cerrar bien el diagnóstico.",
            "madera": "Claro. Para recomendarte el sistema correcto para esa madera, primero cierro dos datos clave.",
            "metal": "Claro. Para llevarte al sistema anticorrosivo correcto, primero necesito ubicar bien el metal y el ambiente.",
            "piso": "Claro. Para recomendarte el sistema correcto para ese piso, necesito cerrar un par de datos.",
        }
        response_text = intro_map.get(category, "Claro, te asesoro. Para recomendarte el producto correcto y no adivinar, necesito cerrar un par de datos clave.")
        if questions:
            response_text += " " + " ".join(questions)
        technical_case["stage"] = "diagnosing"
        return {
            "response_text": response_text,
            "intent": "asesoria_tecnica",
            "context_updates": {"technical_advisory_case": technical_case},
        }

    search_query = build_technical_search_query(technical_case, user_message)
    if is_coverage_followup_question(user_message) and technical_case.get("source_file"):
        search_query = f"{technical_case.get('source_file')}: rendimiento cobertura numero de manos consumo {user_message or ''}".strip()
    chunks = search_technical_chunks(search_query, top_k=6)
    rag_context = build_rag_context(chunks, max_chunks=4)
    source_file = next((chunk.get("doc_filename") for chunk in chunks if chunk.get("similarity", 0) >= 0.25 and chunk.get("doc_filename")), None)
    if not source_file:
        source_file = technical_case.get("source_file")
    technical_case["search_query"] = search_query
    technical_case["source_file"] = source_file

    if not rag_context:
        technical_case["stage"] = "diagnosed_without_rag"
        response_text = (
            "Con lo que me cuentas ya tengo mejor ubicado el caso, pero no te voy a mandar a comprar un producto cualquiera sin una ficha que lo respalde. "
            "En la base técnica no me salió un sistema suficientemente claro para este diagnóstico exacto, así que prefiero no alucinar ni improvisarte una recomendación. "
            "Si quieres, seguimos afinando el caso con el uso exacto y la línea que buscas para aterrizarlo bien dentro del portafolio."
        )
        return {
            "response_text": response_text,
            "intent": "asesoria_tecnica",
            "context_updates": {"technical_advisory_case": technical_case},
        }

    candidate_products = extract_candidate_products_from_rag_context(rag_context, source_file)
    inventory_products = lookup_inventory_candidates_from_terms(candidate_products, conversation_context)
    technical_case["candidate_products"] = candidate_products
    technical_case["inventory_products"] = inventory_products

    try:
        response_text = generate_grounded_technical_sales_reply(
            technical_case,
            rag_context,
            user_message,
            inventory_products,
            area_m2=technical_case.get("area_m2"),
        )
    except Exception:
        response_text = (
            "Ya tengo el diagnóstico y encontré respaldo técnico, pero prefiero no resumírtelo mal en este momento. "
            "Te ayudo a validarlo con la ficha base correcta para no hacerte comprar algo que no te sirva."
        )

    if inventory_products and "Vea, los productos que necesitas son estos:" not in response_text:
        inventory_block = format_inventory_product_block(inventory_products)
        response_text = f"{response_text}\n\n{inventory_block}".strip()

    technical_case["stage"] = "recommended"
    return {
        "response_text": response_text,
        "intent": "asesoria_tecnica",
        "context_updates": {"technical_advisory_case": technical_case},
        "technical_source_filename": source_file,
    }


def detect_business_intent(text_value: Optional[str]):
    if not text_value:
        return "consulta_general"

    lowered = normalize_text_value(text_value)
    if is_technical_document_message(text_value):
        return "consulta_documentacion"
    if any(keyword in lowered for keyword in CLAIM_KEYWORDS):
        return "reclamo_servicio"
    if any(keyword in lowered for keyword in QUOTE_KEYWORDS):
        return "cotizacion"
    if any(keyword in lowered for keyword in ORDER_KEYWORDS):
        return "pedido"
    if re.search(r"\b(necesito|quiero|quisiera|me gustaria|podria|puedo)\b.*\bpedido\b", lowered):
        return "pedido"
    if is_technical_advisory_message(text_value):
        return "asesoria_tecnica"
    if any(keyword in lowered for keyword in ["cartera", "saldo", "deuda", "debo", "vencid", "estado de cuenta", "cupo", "credito", "cuanto debo", "cuánto debo", "documentos"]):
        return "consulta_cartera"
    if has_keyword_or_similar(text_value, ["factura", "facturas", "vencida", "vencidas"]):
        return "consulta_cartera"
    if any(
        keyword in lowered
        for keyword in [
            "ultima compra",
            "última compra",
            "ultimo pedido",
            "último pedido",
            "compra",
            "compro",
            "compró",
            "que he comprado",
            "qué he comprado",
            "que productos compre",
            "qué productos compré",
            "que compre ese dia",
            "qué compré ese día",
            "ese pedido",
            "esa compra",
            "he comprado",
            "comprado",
            "compras",
            "historial de compras",
            "este ano",
            "este año",
            "ultimo ano",
            "ultimo año",
            "último año",
            "ultimos 12 meses",
            "últimos 12 meses",
            "ventas",
        ]
    ):
        return "consulta_compras"
    if is_product_intent_message(text_value):
        return "consulta_productos"
    return "consulta_general"


def format_currency(value):
    try:
        number = float(value or 0)
    except Exception:
        number = 0.0
    return f"${number:,.0f}".replace(",", ".")


def format_quantity(value):
    number = parse_numeric_value(value)
    if number is None:
        return str(value)
    return f"{int(number)}" if float(number).is_integer() else f"{number:g}"


def format_days(value):
    total_days = int(parse_numeric_value(value) or 0)
    return f"{total_days} día" if total_days == 1 else f"{total_days} días"


def format_stock_by_store(stock_by_store: Optional[str]):
    if not stock_by_store:
        return stock_by_store
    return re.sub(
        r":\s*(-?\d+(?:\.\d+)?)",
        lambda match: f": {format_quantity(match.group(1))}",
        str(stock_by_store),
    )


def extract_product_request(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return {
            "core_terms": [],
            "search_terms": [],
            "requested_quantity": None,
            "requested_unit": None,
            "quantity_expression": None,
            "product_codes": [],
            "brand_filters": [],
            "direction_filters": [],
            "size_filters": [],
            "store_filters": [],
            "original_query": "",
        }

    requested_quantity = None
    requested_unit = None
    quantity_expression = None

    quantity_match = re.search(r"(?<![a-z0-9-])(\d+(?:[.,]\d+)?)\s*(galones?|galon|cuartos?|cunetes?|cuñetes?|canecas?|cubetas?)\b", normalized)
    if quantity_match:
        requested_quantity = parse_numeric_value(quantity_match.group(1))
        raw_unit = quantity_match.group(2)
        if raw_unit in PRESENTATION_ALIASES["galon"]:
            requested_unit = "galon"
        elif raw_unit in PRESENTATION_ALIASES["cuñete"]:
            requested_unit = "cuñete"
        elif raw_unit in PRESENTATION_ALIASES["cuarto"]:
            requested_unit = "cuarto"

    quantity_match_reversed = re.search(r"\b(galones?|galon|cuartos?|cunetes?|cuñetes?|canecas?|cubetas?)\s*(\d+(?:[.,]\d+)?)\b", normalized)
    if quantity_match_reversed and requested_quantity is None:
        raw_unit = quantity_match_reversed.group(1)
        requested_quantity = parse_numeric_value(quantity_match_reversed.group(2))
        if raw_unit in PRESENTATION_ALIASES["galon"]:
            requested_unit = "galon"
        elif raw_unit in PRESENTATION_ALIASES["cuñete"]:
            requested_unit = "cuñete"
        elif raw_unit in PRESENTATION_ALIASES["cuarto"]:
            requested_unit = "cuarto"

    shorthand_match = re.search(r"\b(\d+(?:[.,]\d+)?)\s*/\s*(1|4|5)\b", normalized)
    if shorthand_match:
        quantity_expression = f"{shorthand_match.group(1)}/{shorthand_match.group(2)}"
        if requested_quantity is None:
            requested_quantity = parse_numeric_value(shorthand_match.group(1))
        if requested_unit is None:
            requested_unit = PRESENTATION_SHORTCUTS.get(shorthand_match.group(2))

    if requested_unit is None:
        for size_token, unit_name in PRESENTATION_SIZE_MAP.items():
            if re.search(rf"\b{re.escape(size_token)}\b", normalized):
                requested_unit = unit_name
                break

    tokens = [token for token in re.findall(r"[a-z0-9.-]+", normalized) if len(token) >= 2]
    search_terms = []
    for token in tokens:
        if token in PRODUCT_STOPWORDS:
            continue
        if is_store_alias_term(token):
            continue
        if re.fullmatch(r"\d+(?:[./]\d+)?", token):
            continue
        search_terms.append(token)

    if requested_unit in PRESENTATION_ALIASES:
        search_terms.extend(PRESENTATION_ALIASES[requested_unit])

    if requested_unit is None:
        for candidate_unit, aliases in PRESENTATION_ALIASES.items():
            if any(term in search_terms for term in aliases):
                requested_unit = candidate_unit
                search_terms.extend(aliases)
                break

    product_codes = extract_product_codes(text_value)
    if requested_quantity and requested_quantity >= 1000 and product_codes:
        requested_quantity = None
        quantity_expression = None

    core_terms = []
    seen_core_terms = set()
    for term in search_terms:
        normalized_term = normalize_text_value(term)
        if not normalized_term or normalized_term in seen_core_terms:
            continue
        seen_core_terms.add(normalized_term)
        core_terms.append(normalized_term)

    deduped_terms = expand_product_terms(core_terms)

    return {
        "core_terms": core_terms[:8],
        "search_terms": deduped_terms[:8],
        "requested_quantity": requested_quantity,
        "requested_unit": requested_unit,
        "quantity_expression": quantity_expression,
        "product_codes": product_codes,
        "brand_filters": extract_brand_filters(text_value),
        "direction_filters": extract_direction_filters(text_value),
        "size_filters": extract_size_filters(text_value),
        "store_filters": extract_store_filters(text_value),
        "original_query": text_value or "",
    }


def month_date_range(year: int, month: int):
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    return start_date, end_date


def clamp_purchase_end_date(end_date: date, today: Optional[date] = None):
    today = today or date.today()
    return min(end_date, today)


def format_purchase_period_label(month_numbers: list[int], year_value: int):
    unique_months = []
    for month_number in month_numbers:
        if month_number not in unique_months:
            unique_months.append(month_number)
    month_names = []
    for month_number in unique_months:
        month_name = next((name for name, value in MONTH_ALIASES.items() if value == month_number and len(name) > 3), None)
        if month_name:
            month_names.append(month_name)
    if not month_names:
        return f"{year_value}"
    if len(month_names) == 1:
        return f"{month_names[0]} de {year_value}"
    return f"{month_names[0]} a {month_names[-1]} de {year_value}"


def extract_purchase_query(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    today = date.today()
    result = {
        "start_date": None,
        "end_date": None,
        "label": "los ultimos 12 meses",
        "wants_last_purchase": any(keyword in normalized for keyword in ["ultima compra", "última compra", "ultimo pedido", "último pedido"]),
        "wants_products": any(keyword in normalized for keyword in ["producto", "productos", "que compre", "que productos", "qué compré", "qué productos"]),
        "has_time_filter": False,
    }

    exact_match = re.search(r"\b(\d{1,2})\s+de\s+([a-záéíóú]+)\s+de\s+(\d{4}|este ano|este año)\b", normalized)
    if exact_match:
        day_value = int(exact_match.group(1))
        month_value = MONTH_ALIASES.get(exact_match.group(2))
        year_token = exact_match.group(3)
        year_value = today.year if year_token in {"este ano", "este año"} else int(year_token)
        if month_value:
            exact_date = date(year_value, month_value, day_value)
            result.update(
                {
                    "start_date": exact_date,
                    "end_date": exact_date,
                    "label": exact_date.isoformat(),
                    "wants_products": True,
                    "has_time_filter": True,
                }
            )
            return result

    year_match = re.search(r"\b(?:en\s+el\s+)?(?:ano|año)\s+(\d{4}|este ano|este año)\b", normalized)
    explicit_year = None
    if year_match:
        year_token = year_match.group(1)
        explicit_year = today.year if year_token in {"este ano", "este año"} else int(year_token)

    month_mentions = []
    for month_name, month_value in MONTH_ALIASES.items():
        if month_name in normalized:
            month_mentions.append((normalized.find(month_name), month_value))

    if month_mentions:
        month_mentions.sort(key=lambda item: item[0])
        ordered_months = [item[1] for item in month_mentions]
        year_value = explicit_year or today.year
        start_month = min(ordered_months)
        end_month = max(ordered_months)
        start_date, _ = month_date_range(year_value, start_month)
        _, end_date = month_date_range(year_value, end_month)
        if year_value == today.year:
            end_date = clamp_purchase_end_date(end_date, today)
        result.update(
            {
                "start_date": start_date,
                "end_date": end_date,
                "label": format_purchase_period_label(ordered_months, year_value),
                "has_time_filter": True,
            }
        )
        return result

    if explicit_year is not None:
        start_date = date(explicit_year, 1, 1)
        end_date = date(explicit_year, 12, 31)
        if explicit_year == today.year:
            end_date = today
        result.update(
            {
                "start_date": start_date,
                "end_date": end_date,
                "label": f"{explicit_year}",
                "has_time_filter": True,
            }
        )
        return result

    for month_name, month_value in MONTH_ALIASES.items():
        if month_name in normalized:
            year_value = today.year
            year_match = re.search(rf"{month_name}\s+de\s+(\d{{4}}|este ano|este año)", normalized)
            if year_match:
                year_token = year_match.group(1)
                year_value = today.year if year_token in {"este ano", "este año"} else int(year_token)
            start_date, end_date = month_date_range(year_value, month_value)
            if year_value == today.year:
                end_date = clamp_purchase_end_date(end_date, today)
            result.update(
                {
                    "start_date": start_date,
                    "end_date": end_date,
                    "label": f"{month_name} de {year_value}",
                    "has_time_filter": True,
                }
            )
            return result

    return result


def extract_cartera_query(text_value: Optional[str]):
    normalized = normalize_text_value(text_value)
    return {
        "wants_overdue_only": "vencid" in normalized or has_keyword_or_similar(normalized, ["vencida", "vencidas", "vencido", "vencidos"]),
        "wants_invoice_list": any(keyword in normalized for keyword in ["cuales", "cuáles", "que facturas", "qué facturas", "documentos"]) or has_keyword_or_similar(normalized, ["factura", "facturas", "facrura", "facruras"]),
    }


def has_temporal_reference(text_value: Optional[str]):
    purchase_query = extract_purchase_query(text_value)
    return bool(purchase_query.get("has_time_filter"))


def looks_like_product_query(text_value: Optional[str], product_request: Optional[dict]):
    if is_product_intent_message(text_value):
        return True
    request = product_request or extract_product_request(text_value)
    if has_non_product_business_signal(text_value) and not (
        request.get("product_codes")
        or request.get("brand_filters")
        or request.get("requested_unit")
        or request.get("requested_quantity")
        or request.get("store_filters")
    ):
        return False
    if request.get("product_codes"):
        return True
    if request.get("brand_filters") or request.get("requested_unit") or request.get("size_filters"):
        return True
    meaningful_terms = [term for term in (request.get("core_terms") or []) if not is_store_alias_term(term)]
    return len(meaningful_terms) >= 2


def detect_context_switch(conversation_context: Optional[dict], detected_intent: Optional[str], identity_verification_message: bool):
    context = conversation_context or {}
    previous_intent = context.get("last_direct_intent") or context.get("intent")
    if identity_verification_message or context.get("awaiting_verification"):
        return False
    if not previous_intent or not detected_intent or detected_intent == "consulta_general":
        return False
    if previous_intent == detected_intent:
        return False
    tracked_intents = {
        "consulta_productos",
        "consulta_documentacion",
        "asesoria_tecnica",
        "consulta_cartera",
        "consulta_compras",
        "reclamo_servicio",
        "cotizacion",
        "pedido",
    }
    return previous_intent in tracked_intents and detected_intent in tracked_intents


def summarize_claim_product(product_request: Optional[dict], conversation_context: Optional[dict]):
    request = product_request or {}
    previous_request = (conversation_context or {}).get("last_product_request") or {}
    search_terms = list(request.get("core_terms") or request.get("search_terms") or [])
    if not search_terms:
        search_terms = list(previous_request.get("core_terms") or previous_request.get("search_terms") or [])
    claim_noise = {
        "hacer",
        "pintura",
        "tenia",
        "tenía",
        "cubrimiento",
        "bajo",
        "su",
        "alcanzo",
        "alcanzó",
        "no",
        "bien",
        "cubrio",
        "cubrió",
        "funciono",
        "funcionó",
        "reclamo",
        "problema",
        "falla",
        "montar",
        "poner",
        "quiero",
        "necesito",
        "caso",
        "ayuda",
        "cunete",
        "cunetes",
        "cuñete",
        "cuñetes",
    }
    filtered_terms = []
    for term in search_terms:
        normalized_term = normalize_text_value(term)
        if (
            normalized_term
            and normalized_term not in PRODUCT_STOPWORDS
            and normalized_term not in NON_PRODUCT_SERVICE_KEYWORDS
            and normalized_term not in claim_noise
            and normalized_term not in filtered_terms
        ):
            filtered_terms.append(normalized_term)
    if not filtered_terms:
        return None
    product_label = " ".join(filtered_terms[:4])
    quantity_expression = request.get("quantity_expression") or previous_request.get("quantity_expression")
    if quantity_expression and quantity_expression not in product_label:
        product_label = f"{product_label} {quantity_expression}".strip()
    return product_label


def is_weak_claim_product_label(product_label: Optional[str]):
    normalized = normalize_text_value(product_label)
    return normalized in {"", "hacer", "reclamo", "problema", "caso", "ayuda"}


def extract_claim_case_details(text_value: Optional[str], conversation_context: Optional[dict], product_request: Optional[dict]):
    existing_case = dict((conversation_context or {}).get("claim_case") or {})
    normalized = normalize_text_value(text_value)
    raw_text = (text_value or "").strip()
    email_address = extract_email_address(text_value) or existing_case.get("contact_email")
    evidence_note = existing_case.get("evidence_note")
    current_step = existing_case.get("step") or "awaiting_product"
    notes = list(existing_case.get("notes") or [])
    if raw_text and raw_text not in notes and not is_greeting_message(text_value):
        notes.append(raw_text[:600])

    detected_product_label = summarize_claim_product(product_request, conversation_context)
    product_label = existing_case.get("product_label")
    if detected_product_label and (not product_label or is_weak_claim_product_label(product_label)):
        product_label = detected_product_label
    issue_summary = existing_case.get("issue_summary")
    generic_openers = {
        "necesito montar un reclamo",
        "necesito hacer un reclamo",
        "quiero montar un reclamo",
        "quiero hacer un reclamo",
        "quiero poner un reclamo",
        "necesito poner un reclamo",
        "tengo un reclamo",
        "montar un reclamo",
        "hacer un reclamo",
        "quiero abrir un reclamo",
        "necesito abrir un reclamo",
    }
    if raw_text and normalized not in generic_openers and not extract_email_address(text_value):
        has_claim_signal = any(keyword in normalized for keyword in CLAIM_KEYWORDS)
        if current_step in {"awaiting_product", "awaiting_detail"} and (has_claim_signal or existing_case.get("active")):
            issue_summary = raw_text[:600]
        elif current_step == "awaiting_evidence":
            evidence_note = raw_text[:600]

    if not evidence_note and re.search(r"\b(lote|foto|fotos|adjunto|adjunta|imagen|imagenes)\b", normalized):
        evidence_note = raw_text[:600]

    store_name = existing_case.get("store_name")
    store_filters = (product_request or {}).get("store_filters") or []
    if store_filters:
        store_name = STORE_CODE_LABELS.get(store_filters[0]) or store_filters[0]

    missing_fields = []
    if not product_label:
        missing_fields.append("producto")
    if not issue_summary:
        missing_fields.append("detalle")
    if not evidence_note:
        missing_fields.append("evidencia")
    if not email_address:
        missing_fields.append("correo")

    if missing_fields:
        if "producto" in missing_fields:
            next_step = "awaiting_product"
        elif "detalle" in missing_fields:
            next_step = "awaiting_detail"
        elif "evidencia" in missing_fields:
            next_step = "awaiting_evidence"
        else:
            next_step = "awaiting_email"
    else:
        next_step = "ready_to_submit"

    severity = existing_case.get("severity") or (
        "critica" if any(keyword in normalized for keyword in ["no funciono", "no funcionó", "dañado", "danado", "garantia", "garantía"]) else "alta"
    )

    return {
        **existing_case,
        "active": True,
        "product_label": product_label,
        "issue_summary": issue_summary,
        "evidence_note": evidence_note,
        "contact_email": email_address,
        "store_name": store_name,
        "notes": notes[-8:],
        "severity": severity,
        "step": next_step,
        "missing_fields": missing_fields,
        "ready_to_submit": not missing_fields,
    }


def build_claim_reply(profile_name: Optional[str], claim_case: dict, cliente_contexto: Optional[dict]):
    if claim_case.get("submitted"):
        return {
            "tono": "empatico",
            "intent": "reclamo_servicio",
            "priority": claim_case.get("severity") or "alta",
            "summary": f"Seguimiento a reclamo de {claim_case.get('product_label') or 'cliente'}",
            "response_text": (
                "Tu caso ya quedó radicado y sigue en seguimiento. "
                "Si quieres, todavía puedo agregar más detalle, fotos, lote o la tienda donde ocurrió para que el área técnica lo reciba mejor documentado."
            ),
            "should_create_task": False,
            "task_type": "reclamo_calidad",
            "task_summary": "Seguimiento a reclamo existente",
            "task_detail": claim_case,
            "conversation_context_updates": {"claim_case": claim_case},
        }
    if claim_case.get("ready_to_submit"):
        cliente_label = None
        if cliente_contexto:
            cliente_label = cliente_contexto.get("nombre_cliente") or cliente_contexto.get("cliente_codigo")
        response_text = (
            f"Perfecto. Ya dejé radicado el caso de {claim_case.get('product_label')}. "
            "Lo voy a escalar con el área técnica y en unos minutos te llegará al correo la constancia con el detalle para que tengas seguimiento."
        )
        return {
            "tono": "empatico",
            "intent": "reclamo_servicio",
            "priority": claim_case.get("severity") or "alta",
            "summary": f"Reclamo radicado de {claim_case.get('product_label')}",
            "response_text": response_text,
            "should_create_task": True,
            "task_type": "reclamo_calidad",
            "task_summary": f"Reclamo de calidad o funcionamiento: {claim_case.get('product_label')}",
            "task_detail": {**claim_case, "cliente": cliente_label},
            "conversation_context_updates": {"claim_case": {**claim_case, "submitted": True, "active": False, "step": "submitted"}},
            "email_route": "reclamos",
            "email_detail": {**claim_case, "cliente": cliente_label},
            "customer_email_confirmation": {**claim_case, "cliente": cliente_label},
        }

    missing_fields = claim_case.get("missing_fields") or []
    if "producto" in missing_fields:
        response_text = "Claro que sí, lamento el inconveniente. Cuéntame, ¿con qué producto tuviste el problema y qué pasó exactamente?"
    elif "detalle" in missing_fields:
        response_text = (
            f"Entiendo, el caso va sobre {claim_case.get('product_label')}. "
            "Cuéntame qué pasó exactamente para dejarlo bien sustentado, por ejemplo si no cubrió, cambió el tono o presentó alguna falla."
        )
    elif "evidencia" in missing_fields:
        response_text = "Entiendo. ¿De casualidad tienes el número de lote o alguna foto que me puedas compartir?"
    else:
        response_text = "Perfecto. Por favor regálame un correo electrónico para enviarte el número de radicado y hacerle seguimiento."

    return {
        "tono": "empatico",
        "intent": "reclamo_servicio",
        "priority": claim_case.get("severity") or "alta",
        "summary": "Toma de datos para reclamo",
        "response_text": response_text,
        "should_create_task": False,
        "task_type": "reclamo_calidad",
        "task_summary": "Toma inicial de reclamo",
        "task_detail": claim_case,
        "conversation_context_updates": {"claim_case": claim_case},
    }


def build_operational_email_payload(intent: str, profile_name: Optional[str], cliente_contexto: Optional[dict], detail: dict, recent_messages: list[dict]):
    config = get_sendgrid_config()
    if not config:
        return None

    route_map = {
        "reclamos": config.get("reclamos_to_email") or config.get("from_email"),
        "ventas": config.get("ventas_to_email") or config.get("from_email"),
        "contabilidad": config.get("contabilidad_to_email") or config.get("from_email"),
    }
    to_email = route_map.get(intent)
    if not to_email:
        return None

    cliente_label = (cliente_contexto or {}).get("nombre_cliente") or profile_name or "Cliente Ferreinox"
    cliente_codigo = (cliente_contexto or {}).get("cliente_codigo") or "sin_codigo"
    transcript_rows = []
    for row in recent_messages[-8:]:
        direction = "Cliente" if row.get("direction") == "inbound" else "Agente"
        contenido = (row.get("contenido") or "").strip()
        if contenido:
            transcript_rows.append((direction, contenido[:1200]))

    transcript_html = "".join(
        f"<tr><td style='padding:8px;border-bottom:1px solid #e5e7eb;font-weight:600'>{escape(direction)}</td><td style='padding:8px;border-bottom:1px solid #e5e7eb'>{escape(contenido)}</td></tr>"
        for direction, contenido in transcript_rows
    ) or "<tr><td colspan='2' style='padding:8px'>Sin historial disponible.</td></tr>"
    transcript_text = "\n".join(f"{direction}: {contenido}" for direction, contenido in transcript_rows) or "Sin historial disponible."

    if intent == "ventas":
        request_label = "Pedido" if detail.get("intent") == "pedido" else "Cotización"
        store_filters = detail.get("store_filters") or []
        store_name = STORE_CODE_LABELS.get(store_filters[0]) if len(store_filters) == 1 else (", ".join(STORE_CODE_LABELS.get(code, code) for code in store_filters) if store_filters else "Pendiente")
        items_html = "".join(
            f"<li style='margin:0 0 8px 0'>{escape(summarize_commercial_item(item))}</li>"
            for item in (detail.get("items") or [])
            if item.get("status") == "matched"
        ) or "<li>Sin líneas confirmadas.</li>"
        items_text = "\n".join(
            f"- {summarize_commercial_item(item)}"
            for item in (detail.get("items") or [])
            if item.get("status") == "matched"
        ) or "- Sin líneas confirmadas."
        subject = f"Ferreinox CRM | {request_label} cliente {cliente_label}"
        html_content = (
            "<div style='font-family:Segoe UI,Arial,sans-serif;color:#111827;background:#f3f4f6;padding:24px'>"
            "<div style='max-width:900px;margin:0 auto;background:#ffffff;border-radius:18px;overflow:hidden;border:1px solid #e5e7eb'>"
            "<div style='background:#111827;color:#ffffff;padding:24px 28px'>"
            f"<h1 style='margin:0;font-size:24px'>{request_label} preparado desde CRM Ferreinox</h1>"
            "<p style='margin:8px 0 0 0;color:#d1d5db'>Solicitud comercial consolidada desde el agente conversacional.</p>"
            "</div>"
            "<div style='padding:28px'>"
            f"<p><strong>Cliente:</strong> {escape(cliente_label)}</p>"
            f"<p><strong>Código cliente:</strong> {escape(str(cliente_codigo))}</p>"
            f"<p><strong>Tienda/Ciudad:</strong> {escape(store_name)}</p>"
            f"<p><strong>Canal solicitado:</strong> {escape(detail.get('delivery_channel') or 'chat')}</p>"
            f"<p><strong>Correo cliente:</strong> {escape(detail.get('contact_email') or (cliente_contexto or {}).get('email') or 'Pendiente')}</p>"
            "<h2 style='margin-top:28px;font-size:18px'>Líneas consolidadas</h2>"
            f"<ul style='padding-left:20px'>{items_html}</ul>"
            "<p style='margin-top:20px;color:#4b5563'>Esta solicitud quedó lista para revisión comercial. Por ahora no incluye precios automáticos desde PostgREST.</p>"
            "<h2 style='margin-top:28px;font-size:18px'>Historial reciente</h2>"
            "<table style='width:100%;border-collapse:collapse;font-size:14px'>"
            f"{transcript_html}"
            "</table>"
            "</div></div></div>"
        )
        text_content = (
            f"{request_label} preparado desde CRM Ferreinox\n\n"
            f"Cliente: {cliente_label}\n"
            f"Código cliente: {cliente_codigo}\n"
            f"Tienda/Ciudad: {store_name}\n"
            f"Canal solicitado: {detail.get('delivery_channel') or 'chat'}\n"
            f"Correo cliente: {detail.get('contact_email') or (cliente_contexto or {}).get('email') or 'Pendiente'}\n\n"
            f"Líneas consolidadas:\n{items_text}\n\n"
            "Esta solicitud quedó lista para revisión comercial. Por ahora no incluye precios automáticos desde PostgREST.\n\n"
            f"Historial reciente:\n{transcript_text}"
        )
        return {"to_email": to_email, "subject": subject, "html_content": html_content, "text_content": text_content}

    diagnostico_previo = detail.get("diagnostico_previo") or "Sin diagnóstico técnico registrado"
    evidencia_nota = detail.get("evidence_note") or "Pendiente de recibir"
    correo_cliente = detail.get("contact_email") or (cliente_contexto or {}).get("email") or "No proporcionado"
    caso_ref = detail.get("case_reference") or "Pendiente"

    subject = f"Ferreinox CRM | Reclamo radicado {caso_ref} | {detail.get('product_label') or 'sin producto'}"
    html_content = (
        "<div style='font-family:Segoe UI,Arial,sans-serif;color:#111827;background:#f3f4f6;padding:24px'>"
        "<div style='max-width:900px;margin:0 auto;background:#ffffff;border-radius:18px;overflow:hidden;border:1px solid #e5e7eb'>"
        "<div style='background:#111827;color:#ffffff;padding:24px 28px'>"
        f"<h1 style='margin:0;font-size:24px'>Caso radicado {escape(caso_ref)}</h1>"
        "<p style='margin:8px 0 0 0;color:#d1d5db'>Reclamo de calidad o funcionamiento — Resumen ejecutivo generado por el agente.</p>"
        "</div>"
        "<div style='padding:28px'>"
        "<div style='background:#f9fafb;border:1px solid #e5e7eb;border-radius:16px;padding:20px;margin-bottom:20px'>"
        "<h2 style='margin:0 0 12px 0;font-size:16px;color:#374151'>Datos del cliente</h2>"
        f"<p style='margin:0 0 6px 0'><strong>Cliente:</strong> {escape(cliente_label)}</p>"
        f"<p style='margin:0 0 6px 0'><strong>Código cliente:</strong> {escape(str(cliente_codigo))}</p>"
        f"<p style='margin:0 0 6px 0'><strong>Tienda/Ciudad:</strong> {escape(detail.get('store_name') or 'Pendiente')}</p>"
        f"<p style='margin:0'><strong>Correo cliente:</strong> {escape(correo_cliente)}</p>"
        "</div>"
        "<div style='background:#fef2f2;border:1px solid #fecaca;border-radius:16px;padding:20px;margin-bottom:20px'>"
        "<h2 style='margin:0 0 12px 0;font-size:16px;color:#991b1b'>Detalle del reclamo</h2>"
        f"<p style='margin:0 0 6px 0'><strong>Producto reportado:</strong> {escape(detail.get('product_label') or 'Pendiente')}</p>"
        f"<p style='margin:0 0 6px 0'><strong>Problema descrito:</strong> {escape(detail.get('issue_summary') or 'Pendiente de ampliar')}</p>"
        f"<p style='margin:0'><strong>Evidencia:</strong> {escape(evidencia_nota)}</p>"
        "</div>"
        "<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:16px;padding:20px;margin-bottom:20px'>"
        "<h2 style='margin:0 0 12px 0;font-size:16px;color:#1e40af'>Diagnóstico técnico del agente</h2>"
        f"<p style='margin:0;white-space:pre-wrap'>{escape(diagnostico_previo)}</p>"
        "</div>"
        "<p style='margin:20px 0 0 0;color:#6b7280;font-size:13px'>Este resumen fue generado automáticamente por el agente CRM Ferreinox. "
        "No incluye la conversación completa por privacidad — solo los datos relevantes para gestionar el caso.</p>"
        "</div></div></div>"
    )
    text_content = (
        f"CASO RADICADO: {caso_ref}\n"
        f"{'=' * 40}\n\n"
        f"DATOS DEL CLIENTE:\n"
        f"  Cliente: {cliente_label}\n"
        f"  Código: {cliente_codigo}\n"
        f"  Tienda/Ciudad: {detail.get('store_name') or 'Pendiente'}\n"
        f"  Correo: {correo_cliente}\n\n"
        f"DETALLE DEL RECLAMO:\n"
        f"  Producto: {detail.get('product_label') or 'Pendiente'}\n"
        f"  Problema: {detail.get('issue_summary') or 'Pendiente de ampliar'}\n"
        f"  Evidencia: {evidencia_nota}\n\n"
        f"DIAGNÓSTICO TÉCNICO DEL AGENTE:\n"
        f"  {diagnostico_previo}\n\n"
        f"Resumen generado por el agente CRM Ferreinox."
    )
    return {"to_email": to_email, "subject": subject, "html_content": html_content, "text_content": text_content}


def build_customer_claim_confirmation_email(conversation_id: int, profile_name: Optional[str], cliente_contexto: Optional[dict], detail: dict):
    to_email = detail.get("contact_email")
    if not to_email:
        return None

    cliente_label = (cliente_contexto or {}).get("nombre_cliente") or detail.get("cliente") or profile_name or "Cliente Ferreinox"
    cliente_codigo = (cliente_contexto or {}).get("cliente_codigo") or "sin_codigo"
    case_reference = detail.get("case_reference") or f"CRM-{conversation_id}"
    product_label = detail.get("product_label") or "Producto pendiente"
    issue_summary = detail.get("issue_summary") or "Pendiente de ampliar"
    evidence_note = detail.get("evidence_note") or "Pendiente de recibir"
    store_name = detail.get("store_name") or "Pendiente"

    diagnostico_previo = detail.get("diagnostico_previo") or "Pendiente de revisión por el equipo técnico"

    subject = f"Ferreinox | Solicitud radicada {case_reference}"
    html_content = (
        "<div style='font-family:Segoe UI,Arial,sans-serif;background:#f4f6f8;padding:32px;color:#111827'>"
        "<div style='max-width:760px;margin:0 auto;background:#ffffff;border:1px solid #e5e7eb;border-radius:22px;overflow:hidden'>"
        "<div style='background:#111827;padding:28px 32px;color:#ffffff'>"
        "<div style='font-size:12px;letter-spacing:0.18em;text-transform:uppercase;color:#d1d5db'>Ferreinox S.A.S. BIC</div>"
        "<h1 style='margin:10px 0 0 0;font-size:28px;line-height:1.2'>Tu solicitud ya quedó radicada</h1>"
        f"<p style='margin:10px 0 0 0;color:#d1d5db'>Radicado {escape(case_reference)} | Área técnica y servicio</p>"
        "</div>"
        "<div style='padding:32px'>"
        f"<p style='margin:0 0 18px 0'>Hola, {escape(str(cliente_label))}. Ya registramos tu solicitud y nuestro equipo hará seguimiento con esta información:</p>"
        "<div style='background:#f9fafb;border:1px solid #e5e7eb;border-radius:16px;padding:20px;margin-bottom:16px'>"
        f"<p style='margin:0 0 10px 0'><strong>Producto reportado:</strong> {escape(str(product_label))}</p>"
        f"<p style='margin:0 0 10px 0'><strong>Detalle del caso:</strong> {escape(str(issue_summary))}</p>"
        f"<p style='margin:0'><strong>Evidencia recibida:</strong> {escape(str(evidence_note))}</p>"
        "</div>"
        "<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:16px;padding:20px;margin-bottom:16px'>"
        "<p style='margin:0 0 8px 0'><strong>Diagnóstico técnico preliminar:</strong></p>"
        f"<p style='margin:0;color:#1e40af'>{escape(diagnostico_previo)}</p>"
        "</div>"
        "<p style='margin:16px 0 0 0'>Si necesitas ampliar el caso, responde a este correo o escríbenos por WhatsApp y lo anexamos al mismo radicado.</p>"
        "<p style='margin:16px 0 0 0'>Gracias por confiar en Ferreinox.</p>"
        "</div>"
        "</div>"
        "</div>"
    )
    text_content = (
        f"Tu solicitud ya quedó radicada en Ferreinox.\n\n"
        f"Radicado: {case_reference}\n"
        f"Producto reportado: {product_label}\n"
        f"Detalle del caso: {issue_summary}\n"
        f"Evidencia recibida: {evidence_note}\n\n"
        f"Diagnóstico técnico preliminar:\n{diagnostico_previo}\n\n"
        "Si necesitas ampliar el caso, responde este correo o escríbenos por WhatsApp y lo anexamos al mismo radicado."
    )
    return {"to_email": to_email, "subject": subject, "html_content": html_content, "text_content": text_content}


def build_customer_commercial_confirmation_email(conversation_id: int, profile_name: Optional[str], cliente_contexto: Optional[dict], detail: dict):
    to_email = detail.get("contact_email") or (cliente_contexto or {}).get("email")
    if not to_email:
        return None

    request_label = "pedido" if detail.get("intent") == "pedido" else "cotización"
    request_label_title = "Pedido" if detail.get("intent") == "pedido" else "Cotización"
    cliente_label = (cliente_contexto or {}).get("nombre_cliente") or profile_name or "Cliente Ferreinox"
    cliente_codigo = (cliente_contexto or {}).get("cliente_codigo") or "sin_codigo"
    case_reference = f"CRM-{conversation_id}"
    store_filters = detail.get("store_filters") or []
    store_name = STORE_CODE_LABELS.get(store_filters[0]) if len(store_filters) == 1 else (", ".join(STORE_CODE_LABELS.get(code, code) for code in store_filters) if store_filters else "Pendiente")
    items_html = "".join(
        f"<li style='margin:0 0 8px 0'>{escape(summarize_commercial_item(item))}</li>"
        for item in (detail.get("items") or [])
        if item.get("status") == "matched"
    ) or "<li>Sin líneas confirmadas.</li>"
    items_text = "\n".join(
        f"- {summarize_commercial_item(item)}"
        for item in (detail.get("items") or [])
        if item.get("status") == "matched"
    ) or "- Sin líneas confirmadas."

    subject = f"Ferreinox | {request_label_title} preparado {case_reference}"
    html_content = (
        "<div style='font-family:Segoe UI,Arial,sans-serif;background:#f4f6f8;padding:32px;color:#111827'>"
        "<div style='max-width:760px;margin:0 auto;background:#ffffff;border:1px solid #e5e7eb;border-radius:22px;overflow:hidden'>"
        "<div style='background:#111827;padding:28px 32px;color:#ffffff'>"
        "<div style='font-size:12px;letter-spacing:0.18em;text-transform:uppercase;color:#d1d5db'>Ferreinox S.A.S. BIC</div>"
        f"<h1 style='margin:10px 0 0 0;font-size:28px;line-height:1.2'>{request_label_title} preparado</h1>"
        f"<p style='margin:10px 0 0 0;color:#d1d5db'>Solicitud {case_reference}</p>"
        "</div>"
        "<div style='padding:32px'>"
        f"<p style='margin:0 0 18px 0'>Te comparto el resumen de la solicitud de {request_label} que dejamos lista para seguimiento comercial.</p>"
        "<div style='background:#f9fafb;border:1px solid #e5e7eb;border-radius:16px;padding:20px'>"
        f"<p style='margin:0 0 10px 0'><strong>Cliente:</strong> {escape(str(cliente_label))}</p>"
        f"<p style='margin:0 0 10px 0'><strong>Código cliente:</strong> {escape(str(cliente_codigo))}</p>"
        f"<p style='margin:0 0 10px 0'><strong>Tienda o ciudad:</strong> {escape(store_name)}</p>"
        f"<p style='margin:0 0 10px 0'><strong>Canal solicitado:</strong> {escape(detail.get('delivery_channel') or 'chat')}</p>"
        "<div style='margin-top:16px'><strong>Líneas solicitadas:</strong><ul style='margin:10px 0 0 0;padding-left:20px'>"
        f"{items_html}"
        "</ul></div>"
        "</div>"
        "<p style='margin:22px 0 0 0'>Nuestro equipo comercial revisará esta solicitud y continuará el proceso contigo. Por ahora este resumen no incluye precios automáticos.</p>"
        "<p style='margin:22px 0 0 0'>Gracias por confiar en Ferreinox.</p>"
        "</div>"
        "</div>"
        "</div>"
    )
    text_content = (
        f"Ferreinox | {request_label_title} preparado\n\n"
        f"Solicitud: {case_reference}\n"
        f"Cliente: {cliente_label}\n"
        f"Código cliente: {cliente_codigo}\n"
        f"Tienda o ciudad: {store_name}\n"
        f"Canal solicitado: {detail.get('delivery_channel') or 'chat'}\n\n"
        f"Líneas solicitadas:\n{items_text}\n\n"
        "Nuestro equipo comercial revisará esta solicitud y continuará el proceso contigo. Por ahora este resumen no incluye precios automáticos."
    )
    return {"to_email": to_email, "subject": subject, "html_content": html_content, "text_content": text_content}


def is_purchase_followup_message(text_value: Optional[str], conversation_context: Optional[dict]):
    normalized = normalize_text_value(text_value)
    if not normalized:
        return False
    if (conversation_context or {}).get("last_direct_intent") != "consulta_compras":
        return False

    followup_phrases = [
        "ese dia",
        "ese pedido",
        "esa compra",
        "esa fecha",
        "que productos compre",
        "que productos compre ese dia",
        "que compre ese dia",
        "que compre ese pedido",
        "productos compre",
        "productos comprados",
    ]
    return any(phrase in normalized for phrase in followup_phrases)


PDF_STORAGE: dict[str, dict] = {}

COMMERCIAL_COLOR_KEYWORDS: tuple[str, ...] = (
    "blanco", "gris", "negro", "rojo", "verde", "azul", "amarillo", "beige",
    "crema", "marfil", "cafe", "café", "ocre", "plata", "aluminio",
)


def _extract_confirmed_item_summary(item: dict) -> dict:
    matched_product = item.get("matched_product") or {}
    raw_description = (
        matched_product.get("descripcion")
        or matched_product.get("nombre_articulo")
        or item.get("descripcion_comercial")
        or item.get("original_text")
        or "Producto"
    )
    reference = (
        matched_product.get("referencia")
        or matched_product.get("codigo_articulo")
        or item.get("referencia")
        or ""
    )
    product_request = item.get("product_request") or {}
    requested_unit = (
        product_request.get("requested_unit")
        or item.get("unidad_medida")
        or infer_product_presentation_from_row(matched_product)
        or "unidad"
    )
    requested_quantity = product_request.get("requested_quantity")
    if requested_quantity is None:
        requested_quantity = item.get("cantidad")
    return {
        "raw_description": str(raw_description),
        "normalized_text": normalize_text_value(raw_description),
        "reference": str(reference).strip(),
        "requested_unit": requested_unit,
        "requested_quantity": parse_numeric_value(requested_quantity),
    }


def _build_confirmed_item_from_row(
    row: dict,
    requested_quantity,
    requested_unit: Optional[str] = None,
    *,
    source: str = "manual",
    relationship_type: Optional[str] = None,
    auto_note: Optional[str] = None,
):
    matched_product = dict(row or {})
    reference = (
        matched_product.get("referencia")
        or matched_product.get("codigo_articulo")
        or matched_product.get("codigo")
        or ""
    )
    description = get_exact_product_description(matched_product)
    unit_value = requested_unit or infer_product_presentation_from_row(matched_product) or "unidad"
    return {
        "status": "matched",
        "source": source,
        "autogenerated": source != "manual",
        "relationship_type": relationship_type,
        "auto_note": auto_note or "",
        "original_text": description,
        "descripcion_comercial": description,
        "referencia": reference,
        "cantidad": requested_quantity,
        "unidad_medida": unit_value,
        "matched_product": matched_product,
        "product_request": {
            "requested_quantity": requested_quantity,
            "requested_unit": unit_value,
        },
    }


def _resolve_inventory_row_for_commercial_term(
    term: str,
    store_filters: list[str],
    conversation_context: Optional[dict],
    preferred_reference: Optional[str] = None,
) -> Optional[dict]:
    if not term:
        return None
    request = {"store_filters": list(store_filters or [])}
    if preferred_reference:
        preferred_rows = lookup_product_context(preferred_reference, {"product_codes": [preferred_reference], **request})
        preferred_match = next(
            (
                row for row in preferred_rows
                if normalize_reference_value(
                    row.get("referencia") or row.get("codigo_articulo") or row.get("producto_codigo")
                ) == normalize_reference_value(preferred_reference)
            ),
            None,
        )
        if preferred_match:
            return preferred_match
    rows = lookup_product_context(term, request)
    if not rows:
        return None
    if preferred_reference:
        preferred_match = next(
            (
                row for row in rows
                if normalize_reference_value(
                    row.get("referencia") or row.get("codigo_articulo") or row.get("producto_codigo")
                ) == normalize_reference_value(preferred_reference)
            ),
            None,
        )
        if preferred_match:
            return preferred_match
    return rows[0]


def _has_color_signal(text: str) -> bool:
    normalized = normalize_text_value(text)
    if any(keyword in normalized for keyword in COMMERCIAL_COLOR_KEYWORDS):
        return True
    return bool(re.search(r"\bral\s*\d{3,4}\b", normalized))


def _build_quote_reasoning_text(confirmed_items: list[dict], resumen_asesoria: str) -> str:
    joined_text = " ".join(_extract_confirmed_item_summary(item)["normalized_text"] for item in confirmed_items)
    resumen_limpio = " ".join((resumen_asesoria or "").split())
    reasons: list[str] = []
    if "interthane" in joined_text and any(token in joined_text for token in ["interseal", "intergard", "epoxi", "epox", "pintucoat"]):
        reasons.append(
            "Se recomendó un sistema de alto desempeño porque la base epóxica aporta anclaje y barrera anticorrosiva, "
            "mientras el acabado poliuretano protege frente a radiación UV, conserva color y mejora la durabilidad del proyecto."
        )
    elif "interthane" in joined_text:
        reasons.append(
            "Se recomendó Interthane como acabado final porque entrega mejor estabilidad de color, resistencia a intemperie y presentación visual que un acabado convencional."
        )
    elif "pintucoat" in joined_text:
        reasons.append(
            "Se recomendó Pintucoat porque es un recubrimiento epóxico de mejor barrera química y mecánica que un sistema alquídico de mantenimiento."
        )
    elif any(token in joined_text for token in ["interseal", "primer epoxico", "epoxi", "epox"]):
        reasons.append(
            "Se priorizó un sistema epóxico porque ofrece mejor adherencia, sellado y resistencia que una pintura decorativa cuando la exigencia técnica es mayor."
        )
    elif any(token in joined_text for token in ["trafico", "trafico acrilico", "pintutrafico"]):
        reasons.append(
            "Se recomendó este sistema de demarcación porque necesita viscosidad y secado controlados para que el trazo quede uniforme y durable."
        )
    if resumen_limpio:
        if not any(cue in normalize_text_value(resumen_limpio) for cue in ["porque", "ya que", "debido", "por eso"]):
            reasons.insert(0, resumen_limpio)
        else:
            return resumen_limpio
    return " ".join(dict.fromkeys(reason.strip() for reason in reasons if reason.strip()))


def _build_quote_completion_metadata(
    confirmed_items: list[dict],
    store_filters: list[str],
    conversation_context: Optional[dict],
    resumen_asesoria: str,
) -> dict:
    enriched_items = list(confirmed_items)
    existing_refs = {
        normalize_reference_value(_extract_confirmed_item_summary(item)["reference"])
        for item in confirmed_items
        if _extract_confirmed_item_summary(item)["reference"]
    }
    joined_text = " ".join(_extract_confirmed_item_summary(item)["normalized_text"] for item in confirmed_items)
    included_components: list[str] = []
    pending_components: list[str] = []
    auto_added_labels: list[str] = []

    def add_auto_item(term: str, *, preferred_reference: Optional[str] = None, quantity=None, unit: Optional[str] = None, relationship_type: str, note: str):
        row = _resolve_inventory_row_for_commercial_term(term, store_filters, conversation_context, preferred_reference)
        if not row:
            pending_components.append(note)
            return
        reference = normalize_reference_value(row.get("referencia") or row.get("codigo_articulo") or row.get("producto_codigo"))
        if reference in existing_refs:
            included_components.append(note)
            return
        item = _build_confirmed_item_from_row(
            row,
            quantity,
            unit,
            source="auto_complemento",
            relationship_type=relationship_type,
            auto_note=note,
        )
        enriched_items.append(item)
        existing_refs.add(reference)
        included_components.append(note)
        auto_added_labels.append(f"{item.get('descripcion_comercial')} ({item.get('referencia')})")

    for item in confirmed_items:
        summary = _extract_confirmed_item_summary(item)
        normalized_text = summary["normalized_text"]
        requested_quantity = summary["requested_quantity"] or 1
        requested_unit = normalize_text_value(summary["requested_unit"])
        bicomp_info = get_bicomponent_info(normalized_text)
        if bicomp_info and bicomp_info.get("componente_b_codigo"):
            component_b_code = str(bicomp_info.get("componente_b_codigo") or "").strip()
            if component_b_code and normalize_reference_value(component_b_code) not in existing_refs:
                add_auto_item(
                    component_b_code,
                    preferred_reference=component_b_code,
                    quantity=requested_quantity,
                    unit="unidad",
                    relationship_type="catalizador_obligatorio",
                    note=f"Catalizador obligatorio del sistema {bicomp_info.get('producto_base', '').title()}.",
                )

        if "interthane" in normalized_text:
            add_auto_item(
                "UFA151",
                preferred_reference="21050",
                quantity=None,
                unit="unidad",
                relationship_type="diluyente_obligatorio",
                note="Thinner UFA151 requerido para ajuste y aplicación del sistema de poliuretano.",
            )

        if any(token in normalized_text for token in ["pintucoat", "epoxi", "epox", "interseal", "intergard", "primer epoxico"]):
            add_auto_item(
                "thinner epoxico pintuco",
                quantity=None,
                unit="unidad",
                relationship_type="diluyente_recomendado",
                note="Thinner epóxico recomendado para mezcla, ajuste o limpieza del sistema epóxico.",
            )

        if any(token in normalized_text for token in ["trafico", "pintutrafico"]):
            thinner_qty = requested_quantity * 5
            if requested_unit in {"cunete", "cuñete", "caneca", "cubeta"}:
                thinner_qty = requested_quantity * 25
            add_auto_item(
                "21204",
                preferred_reference="21204",
                quantity=thinner_qty,
                unit="botella",
                relationship_type="diluyente_obligatorio",
                note="Thinner 21204 obligatorio para Pintura de Tráfico con cálculo por botellas.",
            )

    tool_terms = []
    if any(token in joined_text for token in ["interthane", "pintucoat", "interseal", "intergard", "epoxi", "epox"]):
        tool_terms.extend(["brocha goya profesional", "rodillo epoxico goya", "lija abracol"])
    elif any(token in joined_text for token in ["trafico", "pintutrafico"]):
        tool_terms.extend(["rodillo goya", "cinta de enmascarar abracol"])

    tool_suggestions: list[str] = []
    seen_tools: set[str] = set()
    for term in tool_terms:
        row = _resolve_inventory_row_for_commercial_term(term, store_filters, conversation_context)
        if not row:
            continue
        description = get_exact_product_description(row)
        reference = row.get("referencia") or row.get("codigo_articulo") or row.get("producto_codigo") or ""
        label = f"{description} ({reference})" if reference else description
        normalized_label = normalize_text_value(label)
        if normalized_label in seen_tools:
            continue
        seen_tools.add(normalized_label)
        tool_suggestions.append(label)

    color_note = ""
    if "interthane" in joined_text and not _has_color_signal(f"{joined_text} {resumen_asesoria}"):
        color_note = (
            "Queda pendiente confirmar color o código final del acabado. "
            "Si el cliente aún no lo define, puede revisar Cartas de Colores en www.ferreinox.co antes del cierre definitivo."
        )

    reasoning_text = _build_quote_reasoning_text(enriched_items, resumen_asesoria)
    if included_components:
        included_sentence = "Se incluyeron o verificaron como parte crítica del sistema: " + "; ".join(dict.fromkeys(included_components))
        reasoning_text = f"{reasoning_text} {included_sentence}".strip()
    if color_note:
        reasoning_text = f"{reasoning_text} {color_note}".strip()

    return {
        "items": enriched_items,
        "justificacion_comercial_pdf": reasoning_text,
        "sistema_completo_pdf": list(dict.fromkeys(included_components)),
        "componentes_pendientes_pdf": list(dict.fromkeys(pending_components)),
        "herramientas_sugeridas_pdf": tool_suggestions,
        "nota_color_pdf": color_note,
        "items_auto_agregados_pdf": auto_added_labels,
        "resumen_asesoria_enriquecido": reasoning_text,
    }


def _resolve_pdf_line_pricing(reference: str) -> dict:
    empty = {"unit_price": 0.0, "price_includes_iva": False, "price_source": None}
    if not reference:
        return empty
    price_info = fetch_product_price(str(reference))
    if price_info:
        if price_info.get("precio_mejor"):
            return {
                "unit_price": float(price_info["precio_mejor"]),
                "price_includes_iva": False,
                "price_source": "agent_precios",
            }
        if price_info.get("pvp_sap"):
            return {
                "unit_price": float(price_info["pvp_sap"]),
                "price_includes_iva": False,
                "price_source": "agent_precios",
            }
        if price_info.get("pvp_franquicia"):
            return {
                "unit_price": float(price_info["pvp_franquicia"]),
                "price_includes_iva": False,
                "price_source": "agent_precios",
            }
    intl_entry = _INTERNATIONAL_PRODUCTS_BY_CODE.get(str(reference).strip())
    if intl_entry:
        for key in ["precio_galon", "precio_cat_galon", "precio_cunete", "precio_cat_cunete"]:
            if intl_entry.get(key):
                return {
                    "unit_price": float(intl_entry[key]),
                    "price_includes_iva": True,
                    "price_source": "international",
                }
    return empty


def generate_commercial_pdf(
    conversation_id: int,
    request_type: str,
    profile_name: Optional[str],
    cliente_contexto: Optional[dict],
    detail: dict,
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    items = detail.get("items") or []
    matched_items = [item for item in items if item.get("status") == "matched"]
    compact_mode = True

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=(12 if compact_mode else 20) * mm,
        bottomMargin=(12 if compact_mode else 20) * mm,
        leftMargin=(14 if compact_mode else 20) * mm,
        rightMargin=(14 if compact_mode else 20) * mm,
    )
    styles = getSampleStyleSheet()

    brand_dark = colors.HexColor(CORPORATE_BRAND["brand_dark"])
    brand_accent = colors.HexColor(CORPORATE_BRAND["brand_accent"])
    brand_light_bg = colors.HexColor(CORPORATE_BRAND["brand_light"])
    brand_border = colors.HexColor(CORPORATE_BRAND["brand_border"])
    white = colors.white

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18 if compact_mode else 22, textColor=white, alignment=TA_LEFT, spaceAfter=3 if compact_mode else 4)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], fontSize=8.5 if compact_mode else 10, textColor=colors.HexColor("#D1D5DB"), alignment=TA_LEFT, leading=10 if compact_mode else 12)
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], fontSize=11 if compact_mode else 13, textColor=brand_dark, spaceBefore=8 if compact_mode else 14, spaceAfter=4 if compact_mode else 6)
    normal_style = ParagraphStyle("Body", parent=styles["Normal"], fontSize=8.6 if compact_mode else 10, textColor=brand_dark, leading=10.5 if compact_mode else 14)
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7 if compact_mode else 8, textColor=colors.HexColor("#6B7280"), leading=9 if compact_mode else 11)
    right_style = ParagraphStyle("Right", parent=styles["Normal"], fontSize=8.6 if compact_mode else 10, textColor=brand_dark, alignment=TA_RIGHT, leading=10.5 if compact_mode else 14)

    request_label = "Pedido" if request_type == "pedido" else "Cotización"
    document_version = "Formato CRM 2026.3"
    case_ref = f"CRM-{conversation_id}"
    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    time_str = now.strftime("%I:%M %p")
    commercial_customer_context = dict(detail.get("customer_context") or cliente_contexto or {})
    cliente_label = commercial_customer_context.get("nombre_cliente") or profile_name or "Cliente Ferreinox"
    cliente_codigo = commercial_customer_context.get("cliente_codigo") or ""
    cliente_nit = commercial_customer_context.get("nit") or commercial_customer_context.get("documento") or ""
    store_filters = detail.get("store_filters") or []
    store_name = STORE_CODE_LABELS.get(store_filters[0]) if len(store_filters) == 1 else (", ".join(STORE_CODE_LABELS.get(c, c) for c in store_filters) if store_filters else "Por definir")
    delivery_channel = detail.get("delivery_channel") or "chat"
    contact_email = detail.get("contact_email") or commercial_customer_context.get("email") or ""
    dispatch_name = detail.get("nombre_despacho") or cliente_label
    observations = detail.get("facturador_notes") or detail.get("observaciones") or ""
    justificacion_comercial = (detail.get("justificacion_comercial_pdf") or "").strip()
    sistema_completo = detail.get("sistema_completo_pdf") or []
    componentes_pendientes = detail.get("componentes_pendientes_pdf") or []
    herramientas_sugeridas = detail.get("herramientas_sugeridas_pdf") or []
    nota_color = (detail.get("nota_color_pdf") or "").strip()

    elements = []

    logo_cell = ""
    if CORPORATE_LOGO_PATH.exists():
        logo = Image(str(CORPORATE_LOGO_PATH))
        logo.drawHeight = 18 * mm
        logo.drawWidth = 42 * mm
        logo_cell = logo

    header_data = [
        [
            logo_cell or Paragraph(f"<b>FERREINOX S.A.S. BIC</b>", title_style),
            Paragraph(f"<b>{request_label}</b><br/><font size='10'>{document_version}</font>", ParagraphStyle("RightTitle", parent=title_style, alignment=TA_RIGHT)),
        ],
        [
            Paragraph(f"NIT {CORPORATE_BRAND['nit']} | {CORPORATE_BRAND['address']}", subtitle_style),
            Paragraph(f"Ref: {case_ref} | {date_str}", ParagraphStyle("RightSub", parent=subtitle_style, alignment=TA_RIGHT)),
        ],
    ]
    header_table = Table(header_data, colWidths=[doc.width * 0.55, doc.width * 0.45])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), brand_dark),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 12 if compact_mode else 16),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 10 if compact_mode else 14),
        ("LEFTPADDING", (0, 0), (-1, -1), 12 if compact_mode else 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12 if compact_mode else 16),
        ("ROUNDEDCORNERS", [8, 8, 0, 0]),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, (2 if compact_mode else 4) * mm))

    banner_table = Table(
        [[
            Paragraph(
                f"<b>Solicitud Comercial Digital</b><br/>{request_label} generado por el CRM Ferreinox para seguimiento operativo inmediato.",
                ParagraphStyle("BannerBody", parent=normal_style),
            ),
            Paragraph(
                f"<b>{document_version}</b><br/>Referencia {case_ref}",
                ParagraphStyle("BannerRight", parent=right_style),
            ),
        ]],
        colWidths=[doc.width * 0.64, doc.width * 0.36],
    )
    banner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FEF3C7")),
        ("BOX", (0, 0), (-1, -1), 0.6, brand_accent),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7 if compact_mode else 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7 if compact_mode else 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 10 if compact_mode else 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10 if compact_mode else 12),
    ]))
    elements.append(banner_table)
    elements.append(Spacer(1, (3 if compact_mode else 5) * mm))

    if compact_mode:
        info_data = [
            [Paragraph("<b>Cliente</b>", normal_style), Paragraph(str(cliente_label), normal_style), Paragraph("<b>Cód. Cliente</b>", normal_style), Paragraph(str(cliente_codigo) if cliente_codigo else "—", normal_style)],
            [Paragraph("<b>Solicita</b>", normal_style), Paragraph(str(dispatch_name), normal_style), Paragraph("<b>NIT / Cédula</b>", normal_style), Paragraph(str(cliente_nit) if cliente_nit else "—", normal_style)],
            [Paragraph("<b>Tienda / Ciudad</b>", normal_style), Paragraph(str(store_name), normal_style), Paragraph("<b>Canal</b>", normal_style), Paragraph(str(delivery_channel).title(), normal_style)],
            [Paragraph("<b>Correo</b>", normal_style), Paragraph(str(contact_email) if contact_email else "—", normal_style), Paragraph("<b>Fecha</b>", normal_style), Paragraph(f"{date_str} - {time_str}", normal_style)],
        ]
        info_table = Table(info_data, colWidths=[doc.width * 0.14, doc.width * 0.36, doc.width * 0.16, doc.width * 0.34])
    else:
        info_data = [
            [Paragraph("<b>Cliente</b>", normal_style), Paragraph(str(cliente_label), normal_style)],
            [Paragraph("<b>Solicita</b>", normal_style), Paragraph(str(dispatch_name), normal_style)],
            [Paragraph("<b>Cód. Cliente</b>", normal_style), Paragraph(str(cliente_codigo) if cliente_codigo else "—", normal_style)],
            [Paragraph("<b>NIT / Cédula</b>", normal_style), Paragraph(str(cliente_nit) if cliente_nit else "—", normal_style)],
            [Paragraph("<b>Tienda / Ciudad</b>", normal_style), Paragraph(str(store_name), normal_style)],
            [Paragraph("<b>Canal</b>", normal_style), Paragraph(str(delivery_channel).title(), normal_style)],
            [Paragraph("<b>Correo</b>", normal_style), Paragraph(str(contact_email) if contact_email else "—", normal_style)],
            [Paragraph("<b>Fecha</b>", normal_style), Paragraph(f"{date_str} - {time_str}", normal_style)],
        ]
        info_table = Table(info_data, colWidths=[doc.width * 0.28, doc.width * 0.72])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), brand_light_bg),
        ("BOX", (0, 0), (-1, -1), 0.5, brand_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, brand_border),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4 if compact_mode else 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4 if compact_mode else 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 7 if compact_mode else 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7 if compact_mode else 10),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, (3 if compact_mode else 6) * mm))

    # ── Resumen de la Asesoría (sustento de la recomendación) ──
    resumen_asesoria = detail.get("resumen_asesoria", "").strip()
    if resumen_asesoria:
        elements.append(Paragraph("Resumen de la Asesoría", heading_style))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=brand_border, spaceBefore=1, spaceAfter=3))
        resumen_style = ParagraphStyle(
            "Resumen", parent=normal_style, fontSize=8.2 if compact_mode else 9.5,
            textColor=colors.HexColor("#374151"), leading=11 if compact_mode else 13,
            spaceBefore=2, spaceAfter=2,
        )
        elements.append(Paragraph(escape(resumen_asesoria), resumen_style))
        elements.append(Spacer(1, (3 if compact_mode else 5) * mm))

    if justificacion_comercial:
        elements.append(Paragraph("Por Qué Este Sistema", heading_style))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=brand_border, spaceBefore=1, spaceAfter=3))
        justificacion_table = Table(
            [[Paragraph(escape(justificacion_comercial), normal_style)]],
            colWidths=[doc.width],
        )
        justificacion_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ECFDF5")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#10B981")),
            ("TOPPADDING", (0, 0), (-1, -1), 8 if compact_mode else 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8 if compact_mode else 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(justificacion_table)
        elements.append(Spacer(1, (3 if compact_mode else 5) * mm))

    if sistema_completo or componentes_pendientes or herramientas_sugeridas or nota_color:
        elements.append(Paragraph("Sistema Completo y Cierre", heading_style))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=brand_border, spaceBefore=1, spaceAfter=3))
        checklist_lines: list[str] = []
        checklist_lines.extend(f"• {escape(str(line))}" for line in sistema_completo if line)
        checklist_lines.extend(f"• Pendiente por definir: {escape(str(line))}" for line in componentes_pendientes if line)
        checklist_lines.extend(f"• Herramienta sugerida: {escape(str(line))}" for line in herramientas_sugeridas if line)
        if nota_color:
            checklist_lines.append(f"• {escape(nota_color)}")
        checklist_html = "<br/>".join(checklist_lines)
        checklist_table = Table([[Paragraph(checklist_html, normal_style)]], colWidths=[doc.width])
        checklist_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#60A5FA")),
            ("TOPPADDING", (0, 0), (-1, -1), 8 if compact_mode else 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8 if compact_mode else 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(checklist_table)
        elements.append(Spacer(1, (3 if compact_mode else 5) * mm))

    elements.append(Paragraph(f"Detalle del {request_label}", heading_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=brand_accent, spaceBefore=2, spaceAfter=4))

    table_header = [
        Paragraph("<b>#</b>", ParagraphStyle("TH", parent=normal_style, textColor=white, alignment=TA_CENTER)),
        Paragraph("<b>Producto</b>", ParagraphStyle("TH", parent=normal_style, textColor=white)),
        Paragraph("<b>Ref.</b>", ParagraphStyle("TH", parent=normal_style, textColor=white, alignment=TA_CENTER)),
        Paragraph("<b>Cant.</b>", ParagraphStyle("TH", parent=normal_style, textColor=white, alignment=TA_CENTER)),
        Paragraph("<b>Precio Unit.</b>", ParagraphStyle("TH", parent=normal_style, textColor=white, alignment=TA_CENTER)),
        Paragraph("<b>Subtotal</b>", ParagraphStyle("TH", parent=normal_style, textColor=white, alignment=TA_CENTER)),
    ]
    table_data = [table_header]

    grand_subtotal = 0
    subtotal_iva_incluido = 0
    for idx, item in enumerate(matched_items, start=1):
        matched_product = item.get("matched_product") or {}
        raw_desc = matched_product.get("descripcion") or matched_product.get("nombre_articulo") or item.get("original_text") or "Producto"
        presentation = infer_product_presentation_from_row(matched_product)
        brand = infer_product_brand_from_row(matched_product)
        commercial_name = translate_product_to_commercial(raw_desc, presentation, brand)
        ref_code = matched_product.get("referencia") or matched_product.get("codigo_articulo") or "—"
        req = item.get("product_request") or {}
        qty_val = req.get("requested_quantity")
        qty_unit = req.get("requested_unit")
        if qty_val and qty_unit:
            qty_label = f"{format_quantity(qty_val)} {qty_unit}"
        elif qty_val:
            qty_label = format_quantity(qty_val)
        else:
            qty_label = "Por confirmar"

        # --- Price lookup ---
        price_snapshot = _resolve_pdf_line_pricing(str(ref_code))
        unit_price = float(price_snapshot.get("unit_price") or 0)
        qty_numeric = float(qty_val) if qty_val else 0
        line_subtotal = unit_price * qty_numeric
        if price_snapshot.get("price_includes_iva"):
            subtotal_iva_incluido += line_subtotal
        else:
            grand_subtotal += line_subtotal

        price_label = f"${unit_price:,.0f}".replace(",", ".") if unit_price > 0 else "Pendiente"
        subtotal_label = f"${line_subtotal:,.0f}".replace(",", ".") if line_subtotal > 0 else "—"
        role_label = item.get("auto_note") or ""
        product_label = escape(commercial_name)
        if role_label:
            product_label = f"{product_label}<br/><font size='7' color='#6B7280'>{escape(str(role_label))}</font>"

        row_bg = white if idx % 2 == 1 else brand_light_bg
        table_data.append([
            Paragraph(str(idx), ParagraphStyle("Cell", parent=normal_style, alignment=TA_CENTER)),
            Paragraph(product_label, normal_style),
            Paragraph(str(ref_code), ParagraphStyle("Cell", parent=normal_style, alignment=TA_CENTER)),
            Paragraph(qty_label, ParagraphStyle("Cell", parent=normal_style, alignment=TA_CENTER)),
            Paragraph(price_label, ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
            Paragraph(subtotal_label, ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
        ])

    # --- Totals row ---
    iva_amount = round(grand_subtotal * 0.19)
    grand_total = round(grand_subtotal + iva_amount + subtotal_iva_incluido)
    if grand_subtotal > 0 or subtotal_iva_incluido > 0:
        fmt_sub = f"${grand_subtotal:,.0f}".replace(",", ".")
        fmt_iva = f"${iva_amount:,.0f}".replace(",", ".")
        fmt_total = f"${grand_total:,.0f}".replace(",", ".")
        table_data.append([
            Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style),
            Paragraph("<b>Subtotal</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
            Paragraph(f"<b>{fmt_sub}</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
        ])
        table_data.append([
            Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style),
            Paragraph("<b>IVA (19%)</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
            Paragraph(f"<b>{fmt_iva}</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
        ])
        if subtotal_iva_incluido > 0:
            fmt_iva_incluido = f"${subtotal_iva_incluido:,.0f}".replace(",", ".")
            table_data.append([
                Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style),
                Paragraph("<b>Items con IVA incluido</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
                Paragraph(f"<b>{fmt_iva_incluido}</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT)),
            ])
        table_data.append([
            Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style), Paragraph("", normal_style),
            Paragraph("<b>TOTAL</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT, textColor=brand_dark)),
            Paragraph(f"<b>{fmt_total}</b>", ParagraphStyle("Cell", parent=normal_style, alignment=TA_RIGHT, textColor=brand_dark)),
        ])

    if not matched_items:
        table_data.append([Paragraph("—", normal_style)] * 6)

    col_widths = [doc.width * 0.04, doc.width * 0.32, doc.width * 0.13, doc.width * 0.11, doc.width * 0.18, doc.width * 0.22] if compact_mode else [doc.width * 0.05, doc.width * 0.30, doc.width * 0.14, doc.width * 0.13, doc.width * 0.18, doc.width * 0.20]
    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), brand_dark),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BOX", (0, 0), (-1, -1), 0.5, brand_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, brand_border),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4 if compact_mode else 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4 if compact_mode else 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 4 if compact_mode else 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4 if compact_mode else 6),
    ]
    for row_idx in range(1, len(table_data)):
        bg = white if row_idx % 2 == 1 else brand_light_bg
        table_style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), bg))
    items_table.setStyle(TableStyle(table_style_cmds))
    elements.append(items_table)
    elements.append(Spacer(1, (3 if compact_mode else 6) * mm))

    total_items = len(matched_items)
    pending_items = len([i for i in items if i.get("status") != "matched"])
    total_label = f"${grand_total:,.0f}".replace(",", ".") if grand_total > 0 else "Pendiente"
    metrics_table = Table(
        [[
            Paragraph(f"<b>Productos</b><br/>{total_items}", ParagraphStyle("Metric", parent=normal_style, alignment=TA_CENTER)),
            Paragraph(f"<b>Sede</b><br/>{escape(str(store_name))}", ParagraphStyle("Metric", parent=normal_style, alignment=TA_CENTER)),
            Paragraph(f"<b>Total IVA inc.</b><br/>{total_label}", ParagraphStyle("Metric", parent=normal_style, alignment=TA_CENTER)),
        ]],
        colWidths=[doc.width / 3.0, doc.width / 3.0, doc.width / 3.0],
    )
    metrics_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), brand_light_bg),
        ("BOX", (0, 0), (-1, -1), 0.5, brand_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, brand_border),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5 if compact_mode else 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5 if compact_mode else 8),
    ]))
    elements.append(metrics_table)
    elements.append(Spacer(1, (2 if compact_mode else 5) * mm))
    summary_text = f"<b>Total productos:</b> {total_items}"
    if pending_items > 0:
        summary_text += f" — <i>{pending_items} pendiente(s) por precisar</i>"
    if grand_total > 0:
        summary_text += f" | <b>Total con IVA:</b> {total_label}"
    elements.append(Paragraph(summary_text, normal_style))
    if observations:
        elements.append(Spacer(1, (2 if compact_mode else 3) * mm))
        elements.append(Paragraph(f"<b>Observaciones operativas:</b> {escape(str(observations))}", normal_style))
    if subtotal_iva_incluido > 0:
        elements.append(Spacer(1, (2 if compact_mode else 3) * mm))
        elements.append(Paragraph("<b>Nota de precios:</b> Algunos productos especializados o de línea International ya vienen con IVA incluido y se consolidaron así en el total.", normal_style))
    elements.append(Spacer(1, (3 if compact_mode else 8) * mm))

    elements.append(HRFlowable(width="100%", thickness=0.5, color=brand_border, spaceBefore=4, spaceAfter=4))
    elements.append(Paragraph(
        f"Este documento resume una solicitud comercial generada desde el CRM Ferreinox. "
        f"Los precios incluidos son antes de IVA (19%). Precios sujetos a disponibilidad y cambio sin previo aviso. "
        f"No constituye factura y está sujeto a validación operativa por Ferreinox SAS BIC.",
        small_style,
    ))
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph(
        f"{CORPORATE_BRAND['company_name']} | {CORPORATE_BRAND['service_email']} | {CORPORATE_BRAND['phone_landline']} | {CORPORATE_BRAND['website']} | {document_version} | {date_str}",
        ParagraphStyle("Footer", parent=small_style, alignment=TA_CENTER),
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def store_commercial_pdf(conversation_id: int, request_type: str, profile_name: Optional[str], cliente_contexto: Optional[dict], detail: dict):
    pdf_buffer = generate_commercial_pdf(conversation_id, request_type, profile_name, cliente_contexto, detail)
    pdf_id = uuid.uuid4().hex[:12]
    request_label = "Pedido" if request_type == "pedido" else "Cotizacion"
    filename = f"Ferreinox_{request_label}_CRM-{conversation_id}_{pdf_id}.pdf"
    PDF_STORAGE[pdf_id] = {
        "buffer": pdf_buffer.getvalue(),
        "filename": filename,
        "created_at": datetime.now().isoformat(),
        "conversation_id": conversation_id,
    }
    return pdf_id, filename


    return pdf_id, filename


# ── Memoria Técnica / Technical Advisory PDF ──────────────────────────────
def generate_technical_advisory_pdf(
    conversation_id: int,
    cliente_nombre: str,
    diagnostico_resumen: list[dict],
    sistema_recomendado: list[dict],
    productos_tabla: list[dict],
    notas_experto: str = "",
):
    """Generate a Technical Advisory PDF ('Memoria Técnica') with diagnostic + recommendation."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=14 * mm, bottomMargin=14 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    brand_dark = colors.HexColor(CORPORATE_BRAND["brand_dark"])
    brand_accent = colors.HexColor(CORPORATE_BRAND["brand_accent"])
    brand_light_bg = colors.HexColor(CORPORATE_BRAND["brand_light"])
    white = colors.white

    title_style = ParagraphStyle("TATtitle", parent=styles["Title"], fontSize=17, textColor=white, alignment=TA_LEFT, spaceAfter=3)
    subtitle_style = ParagraphStyle("TATsub", parent=styles["Normal"], fontSize=8.5, textColor=colors.HexColor("#D1D5DB"), alignment=TA_LEFT, leading=10)
    heading_style = ParagraphStyle("TATheading", parent=styles["Heading2"], fontSize=11, textColor=brand_dark, spaceBefore=10, spaceAfter=4)
    normal_style = ParagraphStyle("TATbody", parent=styles["Normal"], fontSize=9, textColor=brand_dark, leading=11)
    small_style = ParagraphStyle("TATsmall", parent=styles["Normal"], fontSize=7.5, textColor=colors.HexColor("#6B7280"), leading=9)
    right_style = ParagraphStyle("TATright", parent=styles["Normal"], fontSize=9, textColor=brand_dark, alignment=TA_RIGHT, leading=11)

    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    time_str = now.strftime("%I:%M %p")
    case_ref = f"MEM-{conversation_id}"

    elements = []

    # ── Header ──
    logo_cell = ""
    if CORPORATE_LOGO_PATH.exists():
        logo = Image(str(CORPORATE_LOGO_PATH))
        logo.drawHeight = 18 * mm
        logo.drawWidth = 42 * mm
        logo_cell = logo

    header_data = [
        [logo_cell, Paragraph("Recomendación Técnica Especializada", title_style)],
        ["", Paragraph(f"Ref: {case_ref} | Fecha: {date_str} {time_str}", subtitle_style)],
    ]
    header_table = Table(header_data, colWidths=[50 * mm, 130 * mm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), brand_dark),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 6 * mm))

    # ── Client info ──
    elements.append(Paragraph("Datos del Cliente", heading_style))
    elements.append(Paragraph(f"<b>Cliente:</b> {cliente_nombre}", normal_style))
    elements.append(Spacer(1, 3 * mm))

    # ── Diagnostic summary ──
    elements.append(Paragraph("1. Diagnóstico Realizado", heading_style))
    for item in diagnostico_resumen:
        q = item.get("pregunta", "")
        a = item.get("respuesta", "")
        elements.append(Paragraph(f"<b>P:</b> {q}", normal_style))
        elements.append(Paragraph(f"<b>R:</b> {a}", normal_style))
        elements.append(Spacer(1, 1.5 * mm))
    elements.append(Spacer(1, 3 * mm))

    # ── Recommended system ──
    elements.append(Paragraph("2. Sistema Recomendado (Paso a Paso)", heading_style))
    for i, paso in enumerate(sistema_recomendado, 1):
        desc = paso.get("descripcion", "")
        prod = paso.get("producto", "")
        elements.append(Paragraph(f"<b>Paso {i}:</b> {desc}", normal_style))
        if prod:
            elements.append(Paragraph(f"&nbsp;&nbsp;&nbsp;Producto: <b>{prod}</b>", normal_style))
        elements.append(Spacer(1, 1.5 * mm))
    elements.append(Spacer(1, 3 * mm))

    # ── Product table ──
    if productos_tabla:
        elements.append(Paragraph("3. Tabla de Productos y Referencias", heading_style))
        table_data = [["Producto", "Referencia", "Presentación", "Precio Unit.", "Cantidad", "Subtotal"]]
        for p in productos_tabla:
            table_data.append([
                Paragraph(str(p.get("nombre", "")), normal_style),
                str(p.get("referencia", "")),
                str(p.get("presentacion", "")),
                str(p.get("precio", "")),
                str(p.get("cantidad", "")),
                str(p.get("subtotal", "")),
            ])
        prod_table = Table(table_data, colWidths=[55 * mm, 28 * mm, 28 * mm, 22 * mm, 18 * mm, 25 * mm])
        prod_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand_dark),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, brand_light_bg]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(prod_table)
        elements.append(Spacer(1, 4 * mm))

    # ── Expert notes ──
    if notas_experto:
        elements.append(Paragraph("4. Notas del Asesor Experto", heading_style))
        elements.append(Paragraph(notas_experto, normal_style))
        elements.append(Spacer(1, 3 * mm))

    # ── Disclaimer footer ──
    elements.append(HRFlowable(width="100%", color=colors.HexColor("#D1D5DB"), thickness=0.5))
    elements.append(Spacer(1, 2 * mm))
    disclaimer = (
        "⚠️ IMPORTANTE: Esta recomendación técnica se basa en la información proporcionada por el cliente "
        "y el conocimiento técnico de Ferreinox. Los rendimientos, tiempos de curado y resultados finales "
        "pueden variar según las condiciones reales de la superficie, el clima y la aplicación. "
        "Ferreinox recomienda realizar pruebas de adherencia en un área pequeña antes de la aplicación total. "
        "Consulte siempre la ficha técnica del fabricante para datos exactos. "
        "Esta memoria técnica NO constituye una garantía sobre el resultado final."
    )
    elements.append(Paragraph(disclaimer, small_style))
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph(
        f"{CORPORATE_BRAND['company_name']} | {CORPORATE_BRAND['service_email']} | {CORPORATE_BRAND['phone_landline']} | {CORPORATE_BRAND['website']} | {date_str}",
        ParagraphStyle("TATfooter", parent=small_style, alignment=TA_CENTER),
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def store_technical_advisory_pdf(conversation_id: int, cliente_nombre: str, diagnostico_resumen: list[dict],
                                  sistema_recomendado: list[dict], productos_tabla: list[dict], notas_experto: str = ""):
    """Store a Technical Advisory PDF and return its ID and filename."""
    pdf_buffer = generate_technical_advisory_pdf(
        conversation_id, cliente_nombre, diagnostico_resumen, sistema_recomendado, productos_tabla, notas_experto
    )
    pdf_id = uuid.uuid4().hex[:12]
    filename = f"Ferreinox_MemoriaTecnica_MEM-{conversation_id}_{pdf_id}.pdf"
    PDF_STORAGE[pdf_id] = {
        "buffer": pdf_buffer.getvalue(),
        "filename": filename,
        "created_at": datetime.now().isoformat(),
        "conversation_id": conversation_id,
    }
    return pdf_id, filename


def infer_confirmed_order_store_filters(commercial_draft: dict, context: dict, conversation_context: dict, internal_user: Optional[dict] = None):
    existing_filters = [normalize_store_code(value) for value in (commercial_draft.get("store_filters") or [])]
    existing_filters = [value for value in existing_filters if value]
    if existing_filters:
        return list(dict.fromkeys(existing_filters))

    candidate_texts = []
    for item in commercial_draft.get("items") or []:
        original_text = item.get("original_text")
        if original_text:
            candidate_texts.append(str(original_text))

    recent_messages = load_recent_conversation_messages(context["conversation_id"], limit=20)
    for message in reversed(recent_messages):
        if message.get("direction") != "inbound":
            continue
        content = message.get("contenido") or ""
        if content:
            candidate_texts.append(content)

    for text_value in candidate_texts:
        inferred = [normalize_store_code(value) for value in extract_store_filters(text_value)]
        inferred = [value for value in inferred if value]
        if inferred:
            return list(dict.fromkeys(inferred))

    metadata = dict((internal_user or {}).get("metadata") or {})
    fallback_store = normalize_store_code(metadata.get("store_code"))
    return [fallback_store] if fallback_store else []


def fetch_last_year_purchase_summary(cliente_codigo: Optional[str]):
    if not cliente_codigo:
        return None

    engine = get_db_engine()
    with engine.connect() as connection:
        totals = connection.execute(
            text(
                """
                SELECT
                    COUNT(*) AS lineas,
                    COALESCE(SUM(valor_venta_neto), 0) AS valor_total,
                    COALESCE(SUM(unidades_vendidas_netas), 0) AS unidades_totales,
                    MAX(fecha_venta) AS ultima_compra
                FROM public.vw_ventas_netas
                WHERE cliente_id = :cliente_codigo
                  AND fecha_venta >= CURRENT_DATE - INTERVAL '365 days'
                """
            ),
            {"cliente_codigo": cliente_codigo},
        ).mappings().one()

        top_products = connection.execute(
            text(
                """
                SELECT nombre_articulo, codigo_articulo,
                       COALESCE(SUM(unidades_vendidas_netas), 0) AS unidades,
                       COALESCE(SUM(valor_venta_neto), 0) AS valor
                FROM public.vw_ventas_netas
                WHERE cliente_id = :cliente_codigo
                  AND fecha_venta >= CURRENT_DATE - INTERVAL '365 days'
                GROUP BY 1, 2
                ORDER BY valor DESC NULLS LAST
                LIMIT 5
                """
            ),
            {"cliente_codigo": cliente_codigo},
        ).mappings().all()

    return {"totals": dict(totals), "top_products": [dict(row) for row in top_products]}


def fetch_purchase_summary(
    cliente_codigo: Optional[str],
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    if not cliente_codigo:
        return None

    if start_date and end_date:
        where_clause = "fecha_venta BETWEEN :start_date AND :end_date"
        params = {"cliente_codigo": cliente_codigo, "start_date": start_date, "end_date": end_date}
    else:
        where_clause = "fecha_venta >= CURRENT_DATE - INTERVAL '365 days'"
        params = {"cliente_codigo": cliente_codigo}

    engine = get_db_engine()
    with engine.connect() as connection:
        totals = connection.execute(
            text(
                f"""
                SELECT
                    COUNT(*) AS lineas,
                    COALESCE(SUM(valor_venta_neto), 0) AS valor_total,
                    COALESCE(SUM(unidades_vendidas_netas), 0) AS unidades_totales,
                    MIN(fecha_venta) AS primera_compra,
                    MAX(fecha_venta) AS ultima_compra
                FROM public.vw_ventas_netas
                WHERE cliente_id = :cliente_codigo
                                    AND {where_clause}
                                    AND {PURCHASE_LINE_FILTER}
                """
            ),
            params,
        ).mappings().one()

        product_rows = connection.execute(
            text(
                f"""
                SELECT
                    fecha_venta,
                    codigo_articulo,
                    nombre_articulo,
                    COALESCE(SUM(unidades_vendidas_netas), 0) AS unidades,
                    COALESCE(SUM(valor_venta_neto), 0) AS valor
                FROM public.vw_ventas_netas
                WHERE cliente_id = :cliente_codigo
                                    AND {where_clause}
                                    AND {PURCHASE_LINE_FILTER}
                GROUP BY 1, 2, 3
                ORDER BY fecha_venta DESC, valor DESC NULLS LAST
                LIMIT 12
                """
            ),
            params,
        ).mappings().all()

    return {"totals": dict(totals), "products": [dict(row) for row in product_rows]}


def fetch_latest_purchase_detail(cliente_codigo: Optional[str]):
    if not cliente_codigo:
        return None

    engine = get_db_engine()
    with engine.connect() as connection:
        latest_row = connection.execute(
            text(
                f"""
                SELECT MAX(fecha_venta) AS fecha_venta
                FROM public.vw_ventas_netas
                WHERE cliente_id = :cliente_codigo
                  AND {PURCHASE_LINE_FILTER}
                """
            ),
            {"cliente_codigo": cliente_codigo},
        ).mappings().one()

        latest_date = latest_row.get("fecha_venta")
        if not latest_date:
            return None

        totals = connection.execute(
            text(
                f"""
                SELECT
                    COUNT(*) AS lineas,
                    COALESCE(SUM(valor_venta_neto), 0) AS valor_total,
                    COALESCE(SUM(unidades_vendidas_netas), 0) AS unidades_totales
                FROM public.vw_ventas_netas
                WHERE cliente_id = :cliente_codigo
                  AND fecha_venta = :latest_date
                                    AND {PURCHASE_LINE_FILTER}
                """
            ),
            {"cliente_codigo": cliente_codigo, "latest_date": latest_date},
        ).mappings().one()

        products = connection.execute(
            text(
                f"""
                SELECT codigo_articulo, nombre_articulo,
                       COALESCE(SUM(unidades_vendidas_netas), 0) AS unidades,
                       COALESCE(SUM(valor_venta_neto), 0) AS valor
                FROM public.vw_ventas_netas
                WHERE cliente_id = :cliente_codigo
                  AND fecha_venta = :latest_date
                                    AND {PURCHASE_LINE_FILTER}
                GROUP BY 1, 2
                ORDER BY valor DESC NULLS LAST
                LIMIT 10
                """
            ),
            {"cliente_codigo": cliente_codigo, "latest_date": latest_date},
        ).mappings().all()

    return {"fecha_venta": latest_date, "totals": dict(totals), "products": [dict(row) for row in products]}


def build_purchase_summary_response_text(cliente_label: str, purchase_query: dict, purchases: Optional[dict]):
    totals = (purchases or {}).get("totals") or {}
    product_rows = (purchases or {}).get("products") or []
    if not totals or not totals.get("ultima_compra"):
        if purchase_query.get("has_time_filter"):
            return f"No encontré compras registradas para {cliente_label} en {purchase_query.get('label')}."
        return f"No encontré compras registradas en los últimos 12 meses para {cliente_label}."

    top_summary = "; ".join(
        f"{row['nombre_articulo']} ({format_currency(row['valor'])}, {int(float(row['unidades'] or 0))} unidades)"
        for row in product_rows[:5]
    ) or "sin productos destacados"
    if purchase_query.get("has_time_filter"):
        return (
            f"{cliente_label}\n"
            f"En {purchase_query.get('label')} compró {format_currency(totals.get('valor_total'))}.\n"
            f"Fueron {int(totals.get('lineas') or 0)} líneas y {int(float(totals.get('unidades_totales') or 0))} unidades.\n"
            f"Primera compra del periodo: {totals.get('primera_compra') or 'sin fecha'} | última compra: {totals.get('ultima_compra') or 'sin fecha'}.\n"
            f"Productos principales: {top_summary}"
        )
    return (
        f"{cliente_label}\n"
        f"En los últimos 12 meses compró {format_currency(totals.get('valor_total'))}.\n"
        f"Acumula {int(totals.get('lineas') or 0)} líneas y {int(float(totals.get('unidades_totales') or 0))} unidades.\n"
        f"Última compra: {totals.get('ultima_compra') or 'sin fecha'}.\n"
        f"Productos principales: {top_summary}"
    )


def fetch_overdue_documents(cliente_codigo: Optional[str]):
    if not cliente_codigo:
        return None

    engine = get_db_engine()
    with engine.connect() as connection:
        totals = connection.execute(
            text(
                """
                SELECT
                    COALESCE(SUM(importe_normalizado), 0) AS saldo_vencido,
                    COUNT(*) AS documentos_vencidos,
                    COALESCE(MAX(dias_vencido), 0) AS max_dias_vencido
                FROM public.vw_estado_cartera
                WHERE cod_cliente = :cliente_codigo
                  AND COALESCE(dias_vencido, 0) > 0
                  AND COALESCE(importe_normalizado, 0) > 0
                """
            ),
            {"cliente_codigo": cliente_codigo},
        ).mappings().one()

        documents = connection.execute(
            text(
                """
                SELECT numero_documento, fecha_documento, fecha_vencimiento, importe_normalizado, dias_vencido
                FROM public.vw_estado_cartera
                WHERE cod_cliente = :cliente_codigo
                  AND COALESCE(dias_vencido, 0) > 0
                  AND COALESCE(importe_normalizado, 0) > 0
                ORDER BY dias_vencido DESC NULLS LAST, fecha_vencimiento ASC NULLS LAST
                LIMIT 8
                """
            ),
            {"cliente_codigo": cliente_codigo},
        ).mappings().all()

    return {"totals": dict(totals), "documents": [dict(row) for row in documents]}


def build_direct_reply(
    intent: str,
    cliente_contexto: Optional[dict],
    product_context: list[dict],
    profile_name: Optional[str],
    product_request: Optional[dict] = None,
    user_message: Optional[str] = None,
    conversation_context: Optional[dict] = None,
):
    product_context_updates = {}
    if intent == "consulta_productos" and product_context:
        product_context_updates = {
            "last_product_request": product_request or {},
            "last_product_query": user_message or "",
            "last_product_context": [
                {
                    "referencia": row.get("referencia") or row.get("codigo_articulo"),
                    "descripcion": get_exact_product_description(row),
                    "presentacion": infer_product_presentation_from_row(row),
                    "stock_por_tienda": row.get("stock_por_tienda"),
                }
                for row in product_context[:5]
            ],
        }

    if intent == "consulta_cartera":
        if not cliente_contexto:
            return None
        if cliente_contexto.get("verified_source") == "raw_sales" and not any(
            cliente_contexto.get(field_name) is not None for field_name in ["saldo_cartera", "documentos_vencidos", "max_dias_vencido"]
        ):
            return {
                "tono": "informativo",
                "intent": intent,
                "priority": "media",
                "summary": f"Cliente identificado sin cartera consolidada para {cliente_contexto.get('cliente_codigo')}",
                "response_text": (
                    f"Ya te identifiqué como {cliente_contexto.get('nombre_cliente') or cliente_contexto.get('cliente_codigo')}, "
                    "pero en la base actual no veo cartera consolidada para ese cliente. "
                    "Si quieres, te reviso compras recientes o dejo el caso escalado a contabilidad para validarlo."
                ),
                "should_create_task": True,
                "task_type": "validacion_cartera",
                "task_summary": "Validar cliente sin cartera consolidada en CRM",
                "task_detail": cliente_contexto,
            }
        cartera_query = extract_cartera_query(user_message)
        saldo = format_currency(cliente_contexto.get("saldo_cartera"))
        dias = cliente_contexto.get("max_dias_vencido") or 0
        vencidos = cliente_contexto.get("documentos_vencidos") or 0
        vendedor = cliente_contexto.get("vendedor") or "tu asesor comercial"
        overdue_info = None
        if cartera_query.get("wants_overdue_only") or cartera_query.get("wants_invoice_list"):
            overdue_info = fetch_overdue_documents(cliente_contexto.get("cliente_codigo"))

        if overdue_info and (cartera_query.get("wants_overdue_only") or cartera_query.get("wants_invoice_list")):
            overdue_total = format_currency(overdue_info["totals"].get("saldo_vencido"))
            documents = overdue_info.get("documents") or []
            if cartera_query.get("wants_invoice_list") and documents:
                doc_lines = "; ".join(
                    f"factura {row['numero_documento']} por {format_currency(row['importe_normalizado'])}, vence {row['fecha_vencimiento']}, {format_days(row['dias_vencido'])} vencida"
                    for row in documents[:5]
                )
                response_text = (
                    f"Tienes {int(overdue_info['totals'].get('documentos_vencidos') or 0)} facturas vencidas por {overdue_total}. "
                    f"Estas son las principales: {doc_lines}."
                )
            else:
                response_text = (
                    f"Tu cartera vencida es {overdue_total}. "
                    f"Tienes {int(overdue_info['totals'].get('documentos_vencidos') or 0)} documentos vencidos y el mayor atraso es de {format_days(overdue_info['totals'].get('max_dias_vencido'))}."
                )
            return {
                "tono": "informativo",
                "intent": intent,
                "priority": "alta",
                "summary": f"Consulta de cartera vencida de {cliente_contexto.get('cliente_codigo')}",
                "response_text": response_text,
                "should_create_task": bool(int(overdue_info['totals'].get('max_dias_vencido') or 0) > 30),
                "task_type": "seguimiento_cartera",
                "task_summary": "Revisar cliente con cartera vencida",
                "task_detail": overdue_info,
            }

        return {
            "tono": "informativo",
            "intent": intent,
            "priority": "alta" if dias and int(dias) > 0 else "media",
            "summary": f"Consulta de cartera de {cliente_contexto.get('cliente_codigo')}",
            "response_text": (
                f"Tu saldo de cartera actual es {saldo}. "
                f"Tienes {vencidos} documentos vencidos y el mayor atraso es de {format_days(dias)}. "
                f"Tu asesor asignado es {vendedor}."
            ),
            "should_create_task": bool(dias and int(dias) > 30),
            "task_type": "seguimiento_cartera" if dias and int(dias) > 30 else "seguimiento_cliente",
            "task_summary": "Revisar cliente con cartera vencida" if dias and int(dias) > 30 else "Seguimiento a consulta de cartera",
            "task_detail": cliente_contexto,
        }

    if intent == "consulta_compras":
        if not cliente_contexto or not cliente_contexto.get("cliente_codigo"):
            return None
        purchase_query = extract_purchase_query(user_message)
        if not purchase_query.get("has_time_filter") and purchase_query.get("wants_products"):
            context_purchase_date = (conversation_context or {}).get("last_purchase_date")
            if context_purchase_date:
                purchase_query["start_date"] = context_purchase_date
                purchase_query["end_date"] = context_purchase_date
                purchase_query["label"] = str(context_purchase_date)
                purchase_query["has_time_filter"] = True
        if purchase_query.get("wants_last_purchase"):
            latest_purchase = fetch_latest_purchase_detail(cliente_contexto.get("cliente_codigo"))
            if not latest_purchase or not latest_purchase.get("fecha_venta"):
                return {
                    "tono": "informativo",
                    "intent": intent,
                    "priority": "media",
                    "summary": f"Consulta de ultima compra de {cliente_contexto.get('cliente_codigo')}",
                    "response_text": "No encontré una compra registrada para este cliente.",
                    "should_create_task": False,
                    "task_type": "seguimiento_cliente",
                    "task_summary": "Consulta de ultima compra",
                    "task_detail": {},
                }

            product_summary = "; ".join(
                f"{row['nombre_articulo']} ({format_currency(row['valor'])}, {int(float(row['unidades'] or 0))} unidades)"
                for row in latest_purchase.get("products", [])[:6]
            ) or "sin detalle de productos"
            totals = latest_purchase.get("totals") or {}
            return {
                "tono": "informativo",
                "intent": intent,
                "priority": "media",
                "summary": f"Consulta de ultima compra de {cliente_contexto.get('cliente_codigo')}",
                "response_text": (
                    f"Tu última compra fue el {latest_purchase.get('fecha_venta')} por {format_currency(totals.get('valor_total'))}. "
                    f"Incluyó {int(totals.get('lineas') or 0)} líneas y {int(float(totals.get('unidades_totales') or 0))} unidades. "
                    f"Productos principales: {product_summary}."
                ),
                "should_create_task": False,
                "task_type": "seguimiento_cliente",
                "task_summary": "Consulta de ultima compra",
                "task_detail": latest_purchase,
            }

        purchases = fetch_purchase_summary(
            cliente_contexto.get("cliente_codigo"),
            purchase_query.get("start_date"),
            purchase_query.get("end_date"),
        )
        totals = purchases["totals"] if purchases else {}
        product_rows = purchases["products"] if purchases else []
        if not totals or not totals.get("ultima_compra"):
            response_text = "No encontré compras registradas en los últimos 12 meses para este cliente."
        else:
            top_summary = "; ".join(
                f"{row['nombre_articulo']} ({format_currency(row['valor'])}, {int(float(row['unidades'] or 0))} unidades)"
                for row in product_rows[:5]
            ) or "sin productos destacados"
            if purchase_query.get("has_time_filter"):
                response_text = (
                    f"En {purchase_query.get('label')} compraste {format_currency(totals.get('valor_total'))}. "
                    f"Fueron {int(totals.get('lineas') or 0)} líneas y {int(float(totals.get('unidades_totales') or 0))} unidades. "
                    f"Productos principales: {top_summary}."
                )
            else:
                response_text = (
                    f"En los últimos 12 meses registras compras por {format_currency(totals.get('valor_total'))}. "
                    f"Acumulas {int(totals.get('lineas') or 0)} líneas y {int(float(totals.get('unidades_totales') or 0))} unidades. "
                    f"Tu última compra fue el {totals.get('ultima_compra')}. Productos destacados: {top_summary}."
                )
        return {
            "tono": "informativo",
            "intent": intent,
            "priority": "media",
            "summary": f"Consulta de compras de {purchase_query.get('label')} para {cliente_contexto.get('cliente_codigo')}",
            "response_text": response_text,
            "should_create_task": False,
            "task_type": "seguimiento_cliente",
            "task_summary": "Consulta de compras recientes",
            "task_detail": purchases or {},
        }

    if intent == "reclamo_servicio":
        claim_case = extract_claim_case_details(user_message, conversation_context, product_request)
        return build_claim_reply(profile_name, claim_case, cliente_contexto)

    if intent == "cotizacion":
        return build_commercial_flow_reply(intent, profile_name, user_message, conversation_context)

    if intent == "pedido":
        return build_commercial_flow_reply(intent, profile_name, user_message, conversation_context)

    if intent == "consulta_productos":
        if not product_context:
            referencia_solicitada = ", ".join((product_request or {}).get("core_terms") or [])
            return {
                "tono": "informativo",
                "intent": intent,
                "priority": "media",
                "summary": "Consulta de productos sin coincidencia exacta",
                "response_text": (
                    f"No encontré algo claro con {referencia_solicitada or 'esa referencia'}. "
                    "Dame la referencia, el código, la marca o la presentación y te ubico lo que necesitas."
                ),
                "should_create_task": False,
                "task_type": "seguimiento_cliente",
                "task_summary": "Consulta de productos sin match exacto",
                "task_detail": {"product_request": product_request or {}},
            }

        if len(product_context) == 1:
            top_row = product_context[0]
            audit_label = build_product_audit_label(top_row)
            top_stock = top_row.get("stock_total") if top_row.get("stock_total") is not None else top_row.get("stock")
            requested_store_codes = (product_request or {}).get("store_filters") or []
            requested_store_label = STORE_CODE_LABELS.get(requested_store_codes[0]) if len(requested_store_codes) == 1 else None
            if top_stock is not None and parse_numeric_value(top_stock) and parse_numeric_value(top_stock) > 0:
                if requested_store_label:
                    direct_response = f"Sí tenemos {audit_label} en {requested_store_label}. ¿Te separo alguna cantidad?"
                else:
                    direct_response = f"Sí tenemos {audit_label} disponible. ¿En qué tienda lo necesitas?"
            else:
                direct_response = f"El producto {audit_label} lo veo agotado en este momento. ¿Quieres que te revise otra presentación o alternativa?"
            return {
                "tono": "informativo",
                "intent": intent,
                "priority": "media",
                "summary": "Consulta de producto con coincidencia directa",
                "response_text": direct_response,
                "context_updates": product_context_updates,
                "should_create_task": False,
                "task_type": "seguimiento_cliente",
                "task_summary": "Consulta de producto resuelta",
                "task_detail": {"products": product_context, "product_request": product_request or {}},
            }

        if should_ask_product_clarification(product_request, product_context):
            clarification_options = []
            clarification_lines = []
            for index, row in enumerate(product_context[:4], start=1):
                option_payload = {
                    "referencia": row.get("referencia") or row.get("codigo_articulo"),
                    "descripcion": row.get("descripcion") or row.get("nombre_articulo"),
                    "marca": infer_product_brand_from_row(row),
                    "presentacion": infer_product_presentation_from_row(row),
                    "departamentos": row.get("departamentos") or row.get("categoria_producto"),
                    "stock_total": row.get("stock_total") if row.get("stock_total") is not None else row.get("stock"),
                    "stock_por_tienda": row.get("stock_por_tienda"),
                }
                clarification_options.append(option_payload)
                commercial_label = build_product_audit_label(row)
                stock_val = parse_numeric_value(row.get("stock_total") if row.get("stock_total") is not None else row.get("stock"))
                stock_note = f" | stock {format_quantity(stock_val)}" if stock_val and stock_val > 0 else " | agotado"
                clarification_lines.append(f"{index}. {commercial_label}{stock_note}")

            return {
                "tono": "consultivo",
                "intent": intent,
                "priority": "media",
                "summary": "Consulta de productos con necesidad de aclaracion",
                "response_text": build_best_product_clarification_question(product_request, product_context)
                + "\n"
                + "\n".join(clarification_lines),
                "context_updates": product_context_updates,
                "should_create_task": False,
                "task_type": "seguimiento_cliente",
                "task_summary": "Aclaracion de producto",
                "task_detail": {
                    "product_request": product_request or {},
                    "clarification_options": clarification_options,
                },
                "awaiting_product_clarification": True,
                "clarification_options": clarification_options,
            }

        product_lines = []
        quantity_note = None
        requested_store_codes = (product_request or {}).get("store_filters") or []
        requested_store_label = STORE_CODE_LABELS.get(requested_store_codes[0]) if len(requested_store_codes) == 1 else None
        if product_request:
            requested_quantity = product_request.get("requested_quantity")
            requested_unit = product_request.get("requested_unit")
            quantity_expression = product_request.get("quantity_expression")
            if requested_quantity and requested_unit:
                quantity_note = f"Entendí una solicitud de {requested_quantity:g} {get_presentation_label(requested_unit, requested_quantity)}"
            elif quantity_expression:
                quantity_note = f"Tomé la referencia de cantidad {quantity_expression} para orientarte mejor"

        if requested_store_label and product_context:
            top_row = product_context[0]
            top_stock = top_row.get("stock_total") if top_row.get("stock_total") is not None else top_row.get("stock")
            audit_label = build_product_audit_label(top_row)
            stock_value = parse_numeric_value(top_stock) or 0
            if stock_value > 0:
                store_response = f"Sí, en {requested_store_label} tenemos {audit_label}."
                if product_request and product_request.get("requested_quantity"):
                    req_qty = float(product_request["requested_quantity"])
                    if stock_value >= req_qty:
                        store_response += " Te alcanza perfecto para lo que necesitas."
                    else:
                        store_response += " Pero no alcanza para toda la cantidad que pides."
                store_response += " ¿Te separo alguna cantidad o te reviso otra presentación?"
            else:
                store_response = f"El producto {audit_label} no lo veo disponible en {requested_store_label} en este momento. ¿Quieres que revise en otra sede?"
            return {
                "tono": "informativo",
                "intent": intent,
                "priority": "media",
                "summary": "Consulta de producto con tienda especifica",
                "response_text": store_response,
                "context_updates": product_context_updates,
                "should_create_task": False,
                "task_type": "seguimiento_cliente",
                "task_summary": "Consulta de producto por tienda",
                "task_detail": {"products": product_context, "product_request": product_request or {}},
            }

        for row in product_context[:3]:
            commercial_name = build_product_audit_label(row)
            stock = row.get("stock_total") if row.get("stock_total") is not None else row.get("stock")
            stock_value = parse_numeric_value(stock)
            if stock_value and stock_value > 0:
                line = f"{commercial_name} — disponible"
            else:
                line = f"{commercial_name} — agotado"
            product_lines.append(line)
        return {
            "tono": "informativo",
            "intent": intent,
            "priority": "media",
            "summary": "Consulta de productos",
            "context_updates": product_context_updates,
            "response_text": (
                f"{quantity_note + '. ' if quantity_note else ''}"
                f"Encontré estas opciones: {'; '.join(product_lines)}. "
                "¿Cuál es la que buscas o en qué tienda lo necesitas?"
            ),
            "should_create_task": False,
            "task_type": "seguimiento_cliente",
            "task_summary": "Consulta de productos",
            "task_detail": {"products": product_context, "product_request": product_request or {}},
        }

    return None


def build_verification_success_reply(profile_name: Optional[str], cliente_contexto: Optional[dict]):
    cliente_nombre = (cliente_contexto or {}).get("nombre_cliente")
    return (
        f"¡Listo, ya te ubiqué{', ' + str(cliente_nombre) if cliente_nombre else ''}! "
        "Ya puedo ayudarte con cartera, compras y todo tu historial comercial."
    )


def fetch_products_from_catalog(connection, where_clause: str, params: dict, match_score_sql: str, limit: int = 25):
    return connection.execute(
        text(
            f"""
            SELECT producto_codigo, referencia, descripcion, marca, departamentos, stock_total, costo_promedio_und, stock_por_tienda,
                   linea_clasificacion, marca_clasificacion, familia_clasificacion, aplicacion_clasificacion, cat_producto, descripcion_ebs, tipo_articulo,
                   nombre_comercial_abracol, familia_abracol, descripcion_larga_abracol, portafolio_abracol,
                   ({match_score_sql}) AS match_score
            FROM mv_productos
            WHERE {where_clause}
            ORDER BY match_score DESC, stock_total DESC NULLS LAST, descripcion ASC NULLS LAST
            LIMIT {int(limit)}
            """
        ),
        params,
    ).mappings().all()


def fetch_products_from_store_inventory(connection, where_clause: str, params: dict, match_score_sql: str, limit: int = 25):
    return connection.execute(
        text(
            f"""
            SELECT referencia, descripcion, marca, departamentos, stock_total, costo_promedio_und, stock_por_tienda,
                   linea_clasificacion, marca_clasificacion, familia_clasificacion, aplicacion_clasificacion, cat_producto, descripcion_ebs, tipo_articulo,
                   ({match_score_sql}) AS match_score
            FROM (
                SELECT
                    referencia,
                    descripcion,
                    marca,
                    STRING_AGG(DISTINCT departamento, ', ' ORDER BY departamento) AS departamentos,
                    COALESCE(SUM(stock_disponible), 0) AS stock_total,
                    AVG(costo_promedio_und) AS costo_promedio_und,
                    STRING_AGG(
                        almacen_nombre || ': ' || COALESCE(stock_disponible::text, '0'),
                        '; '
                        ORDER BY almacen_nombre
                    ) FILTER (WHERE COALESCE(stock_disponible, 0) > 0) AS stock_por_tienda,
                    MAX(search_blob) AS search_blob,
                    public.fn_keep_alnum(
                        COALESCE(MAX(descripcion), '') || ' ' ||
                        COALESCE(MAX(referencia), '') || ' ' ||
                        COALESCE(MAX(marca), '')
                    ) AS search_compact,
                    MAX(referencia_normalizada) AS referencia_normalizada,
                    MAX(linea_clasificacion) AS linea_clasificacion,
                    MAX(marca_clasificacion) AS marca_clasificacion,
                    MAX(familia_clasificacion) AS familia_clasificacion,
                    MAX(aplicacion_clasificacion) AS aplicacion_clasificacion,
                    MAX(cat_producto) AS cat_producto,
                    MAX(descripcion_ebs) AS descripcion_ebs,
                    MAX(tipo_articulo) AS tipo_articulo
                FROM public.vw_inventario_agente
                WHERE {where_clause}
                GROUP BY referencia, descripcion, marca
            ) inventory
            ORDER BY match_score DESC, stock_total DESC NULLS LAST, descripcion ASC NULLS LAST
            LIMIT {int(limit)}
            """
        ),
        params,
    ).mappings().all()


def fetch_reference_product_rows(connection, references: list[str], store_filters: list[str], match_score: int):
    if not references:
        return []

    params = {}
    catalog_reference_filters = []
    inventory_reference_filters = []
    for index, reference_value in enumerate(references[:5]):
        params[f"reference_{index}"] = normalize_reference_value(reference_value)
        catalog_reference_filters.append(f"producto_codigo = :reference_{index}")
        inventory_reference_filters.append(f"referencia_normalizada = :reference_{index}")

    if store_filters:
        store_filters_sql = []
        for store_index, store_code in enumerate(store_filters):
            params[f"store_{store_index}"] = store_code
            store_filters_sql.append(f"cod_almacen = :store_{store_index}")
        return fetch_products_from_store_inventory(
            connection,
            f"({' OR '.join(inventory_reference_filters)}) AND ({' OR '.join(store_filters_sql)})",
            params,
            str(match_score),
            limit=5,
        )

    return fetch_products_from_catalog(
        connection,
        f"({' OR '.join(catalog_reference_filters)})",
        params,
        str(match_score),
        limit=5,
    )


def fetch_code_product_rows(connection, product_codes: list[str], store_filters: list[str]):
    if not product_codes:
        return []

    params = {}
    code_filters = []
    score_terms = []
    for index, code in enumerate(product_codes[:3]):
        is_numeric_code = bool(re.fullmatch(r"\d{4,10}", str(code or "")))
        params[f"code_like_{index}"] = f"%{code}%"
        params[f"code_exact_{index}"] = str(code)
        params[f"code_compact_{index}"] = f"%{normalize_reference_value(code)}%"
        code_filters.append(f"producto_codigo = :code_exact_{index}")
        code_filters.append(f"referencia = :code_exact_{index}")
        code_filters.append(f"producto_codigo LIKE :code_like_{index}")
        code_filters.append(f"search_blob ILIKE :code_like_{index}")
        if not is_numeric_code:
            code_filters.append(f"search_compact LIKE :code_compact_{index}")
            score_terms.append(
                f"CASE WHEN producto_codigo = :code_exact_{index} OR referencia = :code_exact_{index} THEN 100"
                f" WHEN producto_codigo LIKE :code_like_{index} OR search_blob ILIKE :code_like_{index} OR search_compact LIKE :code_compact_{index} THEN 1 ELSE 0 END"
            )
        else:
            score_terms.append(
                f"CASE WHEN producto_codigo = :code_exact_{index} OR referencia = :code_exact_{index} THEN 100"
                f" WHEN producto_codigo LIKE :code_like_{index} OR search_blob ILIKE :code_like_{index} THEN 1 ELSE 0 END"
            )

    if store_filters:
        store_code_filters = []
        store_score_terms = []
        for index, code in enumerate(product_codes[:3]):
            store_code_filters.append(f"referencia_normalizada = :code_exact_{index}")
            store_code_filters.append(f"referencia_normalizada LIKE :code_like_{index}")
            store_code_filters.append(f"search_blob ILIKE :code_like_{index}")
            store_score_terms.append(
                f"CASE WHEN referencia_normalizada = :code_exact_{index} THEN 100"
                f" WHEN referencia_normalizada LIKE :code_like_{index} OR search_blob ILIKE :code_like_{index} THEN 1 ELSE 0 END"
            )
        store_filters_sql = []
        for store_index, store_code in enumerate(store_filters):
            params[f"store_{store_index}"] = store_code
            store_filters_sql.append(f"cod_almacen = :store_{store_index}")
        return fetch_products_from_store_inventory(
            connection,
            f"({' OR '.join(store_code_filters)}) AND ({' OR '.join(store_filters_sql)})",
            params,
            " + ".join(store_score_terms) if store_score_terms else "0",
            limit=15,
        )

    where_clause = f"({' OR '.join(code_filters)})"
    match_score_sql = " + ".join(score_terms) if score_terms else "0"
    return fetch_products_from_catalog(connection, where_clause, params, match_score_sql, limit=15)


def fetch_term_product_rows(connection, query_terms: list[str], store_filters: list[str]):
    if not query_terms:
        return []

    params = {}
    search_filters = []
    score_terms = []
    var_idx = 500
    for index, term in enumerate(query_terms[:5]):
        params[f"pattern_{index}"] = f"%{term}%"
        compact_term = normalize_reference_value(term)
        params[f"compact_{index}"] = f"%{compact_term}%"
        search_filters.append(f"search_blob ILIKE :pattern_{index}")
        if compact_term:
            search_filters.append(f"search_compact LIKE :compact_{index}")

        # Collect variant ILIKE conditions for WHERE clause
        variant_conditions = []
        variants = _SEARCH_TERM_VARIANTS.get(term.lower(), [])
        for variant in variants[:3]:
            vk = f"tvar_{var_idx}"
            var_idx += 1
            params[vk] = f"%{variant.upper()}%"
            search_filters.append(f"search_blob ILIKE :{vk}")
            variant_conditions.append(f"search_blob ILIKE :{vk}")

        # Combined score: 1 point if original OR compact OR any variant matches
        all_conditions = [f"search_blob ILIKE :pattern_{index}"]
        if compact_term:
            all_conditions.append(f"search_compact LIKE :compact_{index}")
        all_conditions.extend(variant_conditions)
        score_terms.append(
            f"CASE WHEN {' OR '.join(all_conditions)} THEN 1 ELSE 0 END"
        )

    # ── Abbreviation prefix matching ──────────────────────────────────────
    # ERP often truncates multi-word product names (e.g. "PINTUTRAF" for
    # "PINTURA TRAFICO").  For each pair of adjacent terms, concatenate them
    # and generate progressively shorter prefixes (min 6 chars) so that
    # e.g. "pintutrafico" → also tries "pintutrafic", "pintutraf", etc.
    abbrev_index = len(query_terms[:5]) * 10  # offset to avoid param name collisions
    for i in range(len(query_terms[:5]) - 1):
        concat_compact = normalize_reference_value(query_terms[i]) + normalize_reference_value(query_terms[i + 1])
        if len(concat_compact) < 8:
            continue
        # Try progressively shorter prefixes down to 6 chars
        for trim in range(0, min(len(concat_compact) - 6, 5)):
            prefix = concat_compact[: len(concat_compact) - trim]
            param_key = f"abbrev_{abbrev_index}"
            abbrev_index += 1
            params[param_key] = f"%{prefix}%"
            search_filters.append(f"search_compact LIKE :{param_key}")
            score_terms.append(f"CASE WHEN search_compact LIKE :{param_key} THEN 1 ELSE 0 END")

    if store_filters:
        store_search_filters = []
        store_score_terms = []
        for index, term in enumerate(query_terms[:5]):
            store_search_filters.append(f"search_blob ILIKE :pattern_{index}")
            store_score_terms.append(f"CASE WHEN search_blob ILIKE :pattern_{index} THEN 1 ELSE 0 END")
        store_filters_sql = []
        for store_index, store_code in enumerate(store_filters):
            params[f"store_{store_index}"] = store_code
            store_filters_sql.append(f"cod_almacen = :store_{store_index}")
        return fetch_products_from_store_inventory(
            connection,
            f"({' OR '.join(store_search_filters)}) AND ({' OR '.join(store_filters_sql)})",
            params,
            " + ".join(store_score_terms) if store_score_terms else "0",
            limit=25,
        )

    where_clause = f"({' OR '.join(search_filters)})"
    match_score_sql = " + ".join(score_terms) if score_terms else "0"
    return fetch_products_from_catalog(connection, where_clause, params, match_score_sql, limit=25)


def build_curated_catalog_search_terms(text_value: Optional[str], product_request: Optional[dict]):
    request = product_request or {}
    search_terms = []

    def add_term(value: Optional[str]):
        normalized = normalize_text_value(value)
        if normalized and normalized not in search_terms:
            search_terms.append(normalized)

    add_term(request.get("canonical_product"))
    for term in get_specific_product_terms(request):
        add_term(term)
    for term in request.get("brand_filters") or []:
        add_term(term)
    for term in request.get("color_filters") or []:
        add_term(term)
    for term in request.get("finish_filters") or []:
        add_term(term)
    for code in request.get("product_codes") or []:
        add_term(code)
    if not search_terms:
        for term in request.get("core_terms") or []:
            add_term(term)
    add_term(text_value)
    return search_terms[:8]


def fetch_curated_catalog_product_rows(connection, text_value: Optional[str], product_request: Optional[dict], limit: int = 12):
    request = product_request or {}
    search_terms = build_curated_catalog_search_terms(text_value, request)
    if not search_terms:
        return []

    params = {}
    search_filters = []
    score_terms = []
    for index, term in enumerate(search_terms):
        params[f"catalog_term_{index}"] = f"%{term}%"
        search_filters.append(
            "(" 
            f"p.search_blob ILIKE :catalog_term_{index} OR "
            f"a.alias_normalizado ILIKE :catalog_term_{index} OR "
            f"public.fn_normalize_text(COALESCE(a.producto_padre_busqueda, '')) ILIKE :catalog_term_{index} OR "
            f"public.fn_normalize_text(COALESCE(a.familia_consulta, '')) ILIKE :catalog_term_{index} OR "
            f"public.fn_normalize_text(COALESCE(p.producto_padre_busqueda_sugerido, '')) ILIKE :catalog_term_{index} OR "
            f"public.fn_normalize_text(COALESCE(p.familia_consulta_sugerida, '')) ILIKE :catalog_term_{index})"
        )
        score_terms.append(
            f"MAX(CASE WHEN p.search_blob ILIKE :catalog_term_{index} OR a.alias_normalizado ILIKE :catalog_term_{index} THEN 1 ELSE 0 END)"
        )

    base_exact = normalize_text_value(
        request.get("canonical_product")
        or " ".join(get_specific_product_terms(request)[:4])
        or " ".join((request.get("core_terms") or [])[:2])
    )
    color_exact = normalize_text_value(" ".join(request.get("color_filters") or []))
    finish_exact = normalize_text_value(" ".join(request.get("finish_filters") or []))
    params.update(
        {
            "base_exact": base_exact,
            "color_exact": color_exact,
            "color_like": f"%{color_exact}%" if color_exact else "",
            "finish_exact": finish_exact,
            "finish_like": f"%{finish_exact}%" if finish_exact else "",
            "presentation_exact": normalize_text_value(request.get("requested_unit")),
        }
    )
    where_clause = " OR ".join(search_filters)
    brand_filters = request.get("brand_filters") or []
    if brand_filters:
        brand_clauses = []
        for index, brand_name in enumerate(brand_filters[:4]):
            params[f"brand_term_{index}"] = f"%{normalize_text_value(brand_name)}%"
            brand_clauses.append(
                f"p.search_blob ILIKE :brand_term_{index} OR "
                f"public.fn_normalize_text(COALESCE(p.marca, '')) ILIKE :brand_term_{index} OR "
                f"public.fn_normalize_text(COALESCE(p.producto_padre_busqueda_sugerido, '')) ILIKE :brand_term_{index} OR "
                f"public.fn_normalize_text(COALESCE(p.familia_consulta_sugerida, '')) ILIKE :brand_term_{index}"
            )
        where_clause = f"({where_clause}) AND ({' OR '.join(brand_clauses)})"
    requested_colors = request.get("color_filters") or []
    if requested_colors:
        color_groups = []
        for color_index, color_value in enumerate(requested_colors[:3]):
            token_clauses = []
            for token_index, token in enumerate(tokenize_search_phrase(color_value) or [color_value]):
                token_prefix = normalize_text_value(token)[:4]
                if not token_prefix:
                    continue
                param_name = f"color_term_{color_index}_{token_index}"
                params[param_name] = f"%{token_prefix}%"
                token_clauses.append(
                    f"p.search_blob ILIKE :{param_name} OR "
                    f"public.fn_normalize_text(COALESCE(p.color_detectado, '')) ILIKE :{param_name} OR "
                    f"public.fn_normalize_text(COALESCE(p.color_raiz, '')) ILIKE :{param_name}"
                )
            if token_clauses:
                color_groups.append("(" + " AND ".join(token_clauses) + ")")
        if color_groups:
            where_clause = f"({where_clause}) AND ({' OR '.join(color_groups)})"
    score_clause = " + ".join(score_terms) if score_terms else "0"

    return connection.execute(
        text(
            f"""
            SELECT
                p.producto_codigo,
                p.referencia,
                COALESCE(p.descripcion_inventario, p.descripcion_base) AS descripcion,
                p.marca,
                p.departamentos,
                p.stock_total,
                p.stock_por_tienda,
                p.costo_promedio_und,
                p.ventas_unidades_total,
                p.ventas_valor_total,
                p.ultima_venta,
                p.presentacion_canonica,
                p.color_detectado,
                p.color_raiz,
                p.acabado_detectado,
                COALESCE(MAX(NULLIF(a.familia_consulta, '')), p.familia_consulta_sugerida) AS familia_consulta,
                COALESCE(MAX(NULLIF(a.producto_padre_busqueda, '')), p.producto_padre_busqueda_sugerido) AS producto_padre_busqueda,
                MAX(a.pregunta_desambiguacion) AS pregunta_desambiguacion,
                MAX(a.terminos_excluir) AS terminos_excluir,
                ({score_clause}) AS match_score,
                CASE
                    WHEN :base_exact <> '' AND (
                        public.fn_normalize_text(COALESCE(p.producto_padre_busqueda_sugerido, '')) = :base_exact
                        OR MAX(CASE WHEN public.fn_normalize_text(COALESCE(a.producto_padre_busqueda, '')) = :base_exact THEN 1 ELSE 0 END) = 1
                        OR MAX(CASE WHEN a.alias_normalizado = :base_exact THEN 1 ELSE 0 END) = 1
                    ) THEN 2 ELSE 0
                END AS base_exact_score,
                CASE
                    WHEN :presentation_exact <> '' AND public.fn_normalize_text(COALESCE(p.presentacion_canonica, '')) = :presentation_exact THEN 1 ELSE 0
                END AS presentation_score,
                CASE
                    WHEN :color_exact <> '' AND (
                        public.fn_normalize_text(COALESCE(p.color_detectado, '')) = :color_exact
                        OR public.fn_normalize_text(COALESCE(p.color_raiz, '')) = :color_exact
                        OR MAX(CASE WHEN a.alias_type = 'color' AND a.alias_normalizado = :color_exact THEN 1 ELSE 0 END) = 1
                        OR MAX(CASE WHEN p.search_blob ILIKE :color_like THEN 1 ELSE 0 END) = 1
                    ) THEN 1 ELSE 0
                END AS color_score,
                CASE
                    WHEN :finish_exact <> '' AND (
                        public.fn_normalize_text(COALESCE(p.acabado_detectado, '')) = :finish_exact
                        OR MAX(CASE WHEN p.search_blob ILIKE :finish_like THEN 1 ELSE 0 END) = 1
                    ) THEN 1 ELSE 0
                END AS finish_score,
                CASE WHEN COALESCE(p.stock_total, 0) > 0 THEN 1 ELSE 0 END AS stock_score
            FROM public.vw_agent_catalog_product_search p
            LEFT JOIN public.vw_agent_catalog_alias_active a
                ON a.producto_codigo = p.producto_codigo
            WHERE {where_clause}
            GROUP BY
                p.producto_codigo,
                p.referencia,
                COALESCE(p.descripcion_inventario, p.descripcion_base),
                p.marca,
                p.departamentos,
                p.stock_total,
                p.stock_por_tienda,
                p.costo_promedio_und,
                p.ventas_unidades_total,
                p.ventas_valor_total,
                p.ultima_venta,
                p.presentacion_canonica,
                p.color_detectado,
                p.color_raiz,
                p.acabado_detectado,
                p.familia_consulta_sugerida,
                p.producto_padre_busqueda_sugerido
            ORDER BY
                base_exact_score DESC,
                presentation_score DESC,
                color_score DESC,
                finish_score DESC,
                stock_score DESC,
                COALESCE(p.ventas_unidades_total, 0) DESC,
                COALESCE(p.ultima_venta, DATE '1900-01-01') DESC,
                COALESCE(p.stock_total, 0) DESC,
                match_score DESC,
                descripcion ASC NULLS LAST
            LIMIT {int(limit)}
            """
        ),
        params,
    ).mappings().all()


def hydrate_curated_rows_with_store_inventory(connection, curated_rows: list[dict], store_filters: list[str]):
    if not curated_rows or not store_filters:
        return curated_rows

    reference_values = [str(row.get("referencia")) for row in curated_rows if row.get("referencia")]
    if not reference_values:
        return curated_rows

    inventory_rows = fetch_reference_product_rows(connection, reference_values, store_filters, 100)
    if not inventory_rows:
        return curated_rows

    inventory_map = {
        normalize_reference_value(row.get("referencia") or row.get("codigo_articulo")): dict(row)
        for row in inventory_rows
    }
    hydrated_rows = []
    for curated_row in curated_rows:
        reference_key = normalize_reference_value(curated_row.get("referencia") or curated_row.get("producto_codigo"))
        inventory_row = inventory_map.get(reference_key)
        if not inventory_row:
            continue
        merged_row = dict(inventory_row)
        for key, value in curated_row.items():
            if value is not None or key not in merged_row:
                merged_row[key] = value
        hydrated_rows.append(merged_row)
    return hydrated_rows or curated_rows


def rank_product_match_rows(product_rows: list[dict], product_request: Optional[dict], normalized_query: Optional[str], rotation_cache: Optional[dict] = None, query_text: Optional[str] = None):
    if not product_rows:
        return []

    request = product_request or {}
    brand_filters = request.get("brand_filters") or []
    core_terms = request.get("core_terms") or []
    preferred_family_terms = expand_product_terms([normalize_reference_value(core_terms[0])]) if core_terms else []
    specific_terms = get_specific_product_terms(request)
    code_terms = []
    seen_code_terms = set()
    for raw_code in request.get("product_codes") or []:
        normalized_code = normalize_reference_value(raw_code)
        if normalized_code and normalized_code not in seen_code_terms:
            seen_code_terms.add(normalized_code)
            code_terms.append(normalized_code)

    ranked_rows = []
    for row in product_rows:
        candidate = dict(row)
        candidate_text = " ".join(
            str(value)
            for value in [
                candidate.get("descripcion") or candidate.get("nombre_articulo"),
                candidate.get("referencia") or candidate.get("codigo_articulo"),
                candidate.get("producto_codigo"),
                candidate.get("marca") or candidate.get("marca_producto"),
                candidate.get("familia_consulta"),
                candidate.get("producto_padre_busqueda"),
            ]
            if value
        )
        normalized_candidate_text = normalize_text_value(candidate_text)
        compact_candidate_text = normalize_reference_value(candidate_text)
        candidate_reference = normalize_reference_value(
            candidate.get("referencia") or candidate.get("producto_codigo") or candidate.get("codigo_articulo")
        )
        candidate_presentation = infer_product_presentation_from_row(candidate)
        candidate_brand = infer_product_brand_from_row(candidate)
        candidate_size = infer_product_size_from_row(candidate)
        candidate_direction = infer_product_direction_from_row(candidate)
        candidate_color = infer_product_color_from_row(candidate)
        candidate_finish = infer_product_finish_from_row(candidate)

        # ── Kit/Promo deprioritization ──
        # ERP has KIT, PAGUE, NO INV bundles that match product terms but are NOT
        # the actual stock product.  Penalize them so real products rank higher.
        _kit_promo_penalty = 0
        _desc_upper = (candidate.get("descripcion") or candidate.get("nombre_articulo") or "").upper()
        if any(kw in _desc_upper for kw in ("KIT ", "PAGUE ", "PAGU ", "NO INV", "GRATIS", "GTIS", "LLEVE")):
            _kit_promo_penalty = -10

        specific_matches = 0
        for term in specific_terms:
            normalized_term = normalize_text_value(term)
            compact_term = normalize_reference_value(term)
            if (
                normalized_term and normalized_term in normalized_candidate_text
            ) or (
                compact_term and len(compact_term) >= 4 and compact_term in compact_candidate_text
            ):
                specific_matches += 1

        exact_code_matches = 0
        for code_term in code_terms:
            if not code_term:
                continue
            if candidate_reference == code_term:
                exact_code_matches += 10
            elif code_term in compact_candidate_text:
                exact_code_matches += 1

        candidate["exact_code_score"] = exact_code_matches
        candidate["fuzzy_score"] = round(sequence_similarity(normalized_query, candidate_text), 4)
        candidate["family_score"] = 1 if any(term and term in normalized_candidate_text for term in preferred_family_terms[:5]) else 0
        candidate["specific_score"] = specific_matches
        candidate["presentation_score"] = 1 if request.get("requested_unit") and candidate_presentation == request.get("requested_unit") else 0
        candidate["brand_score"] = 1 if brand_filters and candidate_brand in brand_filters else 0
        candidate["size_score"] = 1 if (request.get("size_filters") or []) and candidate_size in (request.get("size_filters") or []) else 0
        candidate["direction_score"] = 1 if (request.get("direction_filters") or []) and candidate_direction in (request.get("direction_filters") or []) else 0
        candidate["color_score"] = 1 if (request.get("color_filters") or []) and candidate_color in (request.get("color_filters") or []) else 0
        candidate["finish_score"] = 1 if (request.get("finish_filters") or []) and candidate_finish in (request.get("finish_filters") or []) else 0
        candidate["kit_promo_penalty"] = _kit_promo_penalty
        # ── Smart Score (unified 0-1 scoring) ──
        _smart_query = query_text or normalized_query or ""
        candidate["smart_score"] = smart_score_product(candidate, _smart_query, request, rotation_cache)
        # Use rotation_score from DB if present, else from cache
        candidate["rotation_score"] = float(candidate.get("rotation_score") or (rotation_cache or {}).get(
            str(candidate.get("producto_codigo") or candidate.get("referencia") or ""), 0
        ))
        ranked_rows.append(candidate)

    ranked_rows.sort(
        key=lambda item: (
            item.get("kit_promo_penalty") or 0,  # Negative for kits → sorts them to the bottom
            item.get("exact_code_score") or 0,
            item.get("specific_score") or 0,       # Product-specific term matches (moved up for accuracy)
            item.get("match_score") or 0,
            item.get("smart_score") or 0,         # Unified 0-1 smart score
            item.get("rotation_score") or 0,       # Historical sales rotation
            item.get("direction_score") or 0,
            item.get("size_score") or 0,
            item.get("presentation_score") or 0,
            item.get("finish_score") or 0,
            item.get("color_score") or 0,
            item.get("brand_score") or 0,
            item.get("base_exact_score") or 0,
            item.get("family_score") or 0,
            item.get("fuzzy_score") or 0,
            parse_numeric_value(item.get("stock_total")) or 0,
        ),
        reverse=True,
    )

    top_exact_code_score = ranked_rows[0].get("exact_code_score") or 0 if ranked_rows else 0
    if top_exact_code_score > 0:
        ranked_rows = [item for item in ranked_rows if (item.get("exact_code_score") or 0) == top_exact_code_score]

    max_specific_score = max((item.get("specific_score") or 0 for item in ranked_rows), default=0) if ranked_rows else 0
    if max_specific_score >= 2:
        ranked_rows = [item for item in ranked_rows if (item.get("specific_score") or 0) == max_specific_score]
    elif max_specific_score > 0 and len(specific_terms) == 1:
        ranked_rows = [item for item in ranked_rows if (item.get("specific_score") or 0) > 0]

    top_match_score = ranked_rows[0].get("match_score") or 0 if ranked_rows else 0
    if top_match_score >= 2:
        ranked_rows = [
            item for item in ranked_rows
            if (item.get("match_score") or 0) >= max(2, top_match_score - 1)
            or (item.get("size_score") or 0) > 0
            or (item.get("brand_score") or 0) > 0
            or (item.get("family_score") or 0) > 0
            or (item.get("exact_code_score") or 0) > 0
        ]

    if request.get("requested_unit"):
        exact_presentation_rows = [
            item for item in ranked_rows
            if infer_product_presentation_from_row(item) == request.get("requested_unit")
        ]
        if exact_presentation_rows:
            ranked_rows = exact_presentation_rows
    if request.get("color_filters"):
        exact_color_rows = [
            item for item in ranked_rows
            if row_matches_requested_colors(item, request.get("color_filters") or [])
        ]
        if exact_color_rows:
            ranked_rows = exact_color_rows
    if request.get("finish_filters"):
        exact_finish_rows = [
            item for item in ranked_rows
            if infer_product_finish_from_row(item) in (request.get("finish_filters") or [])
        ]
        if exact_finish_rows:
            ranked_rows = exact_finish_rows
    if brand_filters:
        exact_brand_rows = [
            item for item in ranked_rows
            if infer_product_brand_from_row(item) in brand_filters
        ]
        if exact_brand_rows:
            ranked_rows = exact_brand_rows
    ranked_rows = filter_rows_by_requested_size(ranked_rows, request)
    if any((parse_numeric_value(item.get("stock_total")) or 0) > 0 for item in ranked_rows):
        ranked_rows = [item for item in ranked_rows if (parse_numeric_value(item.get("stock_total")) or 0) > 0]
    return ranked_rows


def _merge_product_rows(primary: list[dict], secondary: list[dict]) -> list[dict]:
    """Merge two ranked product lists, deduplicating by product code. Primary rows take priority."""
    seen: set[str] = set()
    merged: list[dict] = []
    for row in primary:
        code = str(row.get("producto_codigo") or row.get("referencia") or "")
        if code and code not in seen:
            seen.add(code)
            merged.append(row)
    for row in secondary:
        code = str(row.get("producto_codigo") or row.get("referencia") or "")
        if code and code not in seen:
            seen.add(code)
            merged.append(row)
    return merged


def lookup_product_context(text_value: Optional[str], product_request: Optional[dict] = None):
    product_request = prepare_product_request_for_search(text_value, product_request)
    core_terms = product_request.get("core_terms") or []
    terms = product_request.get("search_terms") or []
    product_codes = product_request.get("product_codes") or []
    learned_references = fetch_learned_product_references(product_request)
    store_filters = product_request.get("store_filters") or []
    brand_filters = product_request.get("brand_filters") or []
    normalized_query = normalize_text_value(text_value)

    if not terms and not product_codes and not learned_references:
        return []

    try:
        engine = get_db_engine()
        with engine.connect() as connection:
            # ── Load rotation cache once for the full search session ──
            rotation_cache = fetch_rotation_cache(connection)

            if learned_references:
                learned_rows = fetch_reference_product_rows(connection, learned_references, store_filters, 90)
                if learned_rows:
                    return filter_rows_by_requested_presentation([dict(row) for row in learned_rows], product_request)

            if product_codes:
                code_rows = fetch_code_product_rows(connection, product_codes, store_filters)
                if code_rows:
                    ranked_code_rows = rank_product_match_rows([dict(row) for row in code_rows], product_request, normalized_query, rotation_cache, text_value)
                    ranked_code_rows = filter_rows_by_requested_presentation(ranked_code_rows, product_request)
                    return ranked_code_rows[:10]
                # Fuzzy near-code fallback: if no results, try digit-deletion and digit-transposition variants
                # (e.g. user types "17174" when correct code is "117474" or vice-versa)
                fuzzy_codes: list[str] = []
                for code in product_codes[:2]:
                    if re.fullmatch(r"\d{5,10}", code):
                        for pos in range(len(code)):
                            fuzzy_codes.append(code[:pos] + code[pos + 1:])
                        for pos in range(len(code) - 1):
                            fuzzy_codes.append(code[:pos] + code[pos + 1] + code[pos] + code[pos + 2:])
                fuzzy_codes = list(dict.fromkeys(c for c in fuzzy_codes if len(c) >= 4))[:20]
                if fuzzy_codes:
                    fuzzy_rows = fetch_code_product_rows(connection, fuzzy_codes[:3], store_filters)
                    if fuzzy_rows:
                        ranked_fuzzy = rank_product_match_rows([dict(row) for row in fuzzy_rows], product_request, normalized_query, rotation_cache, text_value)
                        ranked_fuzzy = filter_rows_by_requested_presentation(ranked_fuzzy, product_request)
                        return ranked_fuzzy[:10]

            if not terms:
                return []

            # ── Build query_terms once (used by both curated and full-catalog) ──
            query_terms = []
            for term in list(core_terms) + list(terms):
                if term not in query_terms:
                    query_terms.append(term)
                if len(query_terms) == 6:
                    break

            # ── Stage 3+4 combined: curated catalog + smart full-catalog search ──
            curated_rows = fetch_curated_catalog_product_rows(connection, text_value, product_request, limit=100)
            ranked_curated_rows: list[dict] = []
            if curated_rows:
                ranked_curated_rows = hydrate_curated_rows_with_store_inventory(
                    connection,
                    [dict(row) for row in curated_rows],
                    store_filters,
                )
                ranked_curated_rows = rank_product_match_rows(ranked_curated_rows, product_request, normalized_query, rotation_cache, text_value)

            # Smart full-catalog search with trigram + phonetic + rotation
            smart_rows = fetch_smart_product_rows(connection, text_value or "", query_terms, product_request, store_filters, limit=30)
            ranked_term_rows: list[dict] = []
            if smart_rows:
                ranked_term_rows = rank_product_match_rows(smart_rows, product_request, normalized_query, rotation_cache, text_value)

            # Also run legacy term search as fallback ONLY if smart search has weak results
            # (skip it when smart search already found strong matches to save a DB roundtrip)
            good_smart_count = sum(1 for r in ranked_term_rows if (r.get("match_score") or 0) >= 3)
            if good_smart_count < 3:
                legacy_rows = fetch_term_product_rows(connection, query_terms, store_filters)
                if legacy_rows:
                    legacy_ranked = rank_product_match_rows([dict(row) for row in legacy_rows], product_request, normalized_query, rotation_cache, text_value)
                    ranked_term_rows = _merge_product_rows(ranked_term_rows, legacy_ranked)

            # Merge curated + full-catalog, curated rows take priority, then re-sort by smart_score
            if ranked_curated_rows or ranked_term_rows:
                merged = _merge_product_rows(ranked_curated_rows, ranked_term_rows)
                if merged:
                    # Final re-sort: specific_score before smart_score so relevant matches win
                    merged.sort(
                        key=lambda item: (
                            item.get("kit_promo_penalty") or 0,
                            item.get("exact_code_score") or 0,
                            item.get("specific_score") or 0,
                            item.get("smart_score") or 0,
                            item.get("rotation_score") or 0,
                            item.get("match_score") or 0,
                            parse_numeric_value(item.get("stock_total")) or 0,
                        ),
                        reverse=True,
                    )
                    return merged[:10]

            sales_filters = []
            sales_scores = []
            sales_params = {}
            for index, term in enumerate(query_terms):
                sales_params[f"pattern_{index}"] = f"%{term}%"
                sales_filters.append(f"search_blob ILIKE :pattern_{index}")
                sales_scores.append(f"CASE WHEN search_blob ILIKE :pattern_{index} THEN 1 ELSE 0 END")

            sales_rows = connection.execute(
                text(
                    f"""
                    SELECT codigo_articulo, nombre_articulo, marca_producto, categoria_producto,
                           SUM(unidades_vendidas_netas) AS unidades_vendidas,
                           SUM(valor_venta_neto) AS valor_vendido,
                           MAX(match_score) AS match_score
                    FROM (
                        SELECT
                            codigo_articulo,
                            nombre_articulo,
                            marca_producto,
                            categoria_producto,
                            unidades_vendidas_netas,
                            valor_venta_neto,
                            ({' + '.join(sales_scores)}) AS match_score,
                            translate(lower(
                                COALESCE(nombre_articulo, '') || ' ' ||
                                COALESCE(codigo_articulo, '') || ' ' ||
                                COALESCE(marca_producto, '') || ' ' ||
                                COALESCE(categoria_producto, '')
                            ), 'áéíóúàèìòùâêîôûäëïöüñ', 'aeiouaeiouaeiouaeioun') AS search_blob
                        FROM public.vw_ventas_netas
                    ) sales
                    WHERE {' OR '.join(sales_filters)}
                    GROUP BY 1, 2, 3, 4
                    ORDER BY match_score DESC, valor_vendido DESC NULLS LAST
                    LIMIT 10
                    """
                ),
                sales_params,
            ).mappings().all()

            return [dict(row) for row in sales_rows]
    except Exception:
        return []


def build_verification_challenge():
    return (
        "Para darte esa info por tu seguridad, ¿me regalas tu número de cédula o NIT por favor? 🔒"
    )


def build_name_confirmation_challenge(cliente_nombre: str):
    return (
        f"Por seguridad, encontré una cuenta asociada. ¿Me confirmas si el titular es *{cliente_nombre}*? "
        "Respóndeme sí o no."
    )


def is_name_confirmation_response(text_value: Optional[str]):
    """Check if the message is a yes/no confirmation to the name challenge."""
    lowered = normalize_text_value(text_value)
    if not lowered:
        return None
    if re.match(r"^(si|sí|eso es|asi es|as[ií] es|correcto|exacto|dale|listo|de una|ok|okay|perfecto|confirmado|ese soy|soy yo|es[ae]? soy|es[ae]? es|ese soy yo|si se[ñn]or|si claro|claro que si|afirmativo|efectivamente)[.!?,\s]*.*$", lowered):
        return True
    if re.match(r"^(no|nop|negativo|no soy|no es|ese no|esa no|no es esa?|no soy yo|ese no es|esa no es|no ese no|para nada)[.!?,\s]*.*$", lowered):
        return False
    return None




def build_fallback_agent_result(user_message: str, error_message: str):
    # Detect if this looks like a bulk order to give a more actionable fallback
    lines = (user_message or "").strip().splitlines()
    looks_like_order = len(lines) >= 3 and sum(
        1 for ln in lines
        if any(
            kw in ln.lower()
            for kw in ["galon", "galón", "cuarto", "cuñete", "cuñetes", "brocha", "balde", "kilo"]
        )
    ) >= 2
    if looks_like_order:
        fallback_text = (
            "Recibí tu pedido pero tuve un problema técnico al procesarlo. "
            "Por favor vuelve a enviarlo de nuevo y lo proceso de inmediato. "
            "Si el problema persiste, un asesor te ayuda pronto."
        )
    else:
        fallback_text = "Recibimos tu mensaje. Un asesor te contactará pronto."
    return {
        "tono": "neutral",
        "intent": "consulta_general",
        "priority": "media",
        "summary": user_message[:200] if user_message else "Consulta entrante",
        "response_text": fallback_text,
        "should_create_task": True,
        "task_type": "revision_manual",
        "task_summary": "Revisar conversacion con falla en respuesta automatica",
        "task_detail": {"mensaje": user_message, "error": error_message},
    }


def send_whatsapp_text_message(to_phone: str, body: str):
    response = requests.post(
        f"https://graph.facebook.com/v22.0/{get_whatsapp_phone_number_id()}/messages",
        headers={
            "Authorization": f"Bearer {get_whatsapp_access_token()}",
            "Content-Type": "application/json",
        },
        json={
            "messaging_product": "whatsapp",
            "to": to_phone.lstrip("+"),
            "type": "text",
            "text": {"preview_url": False, "body": body},
        },
        timeout=20,
    )
    if response.status_code >= 400:
        try:
            error_payload = response.json()
        except Exception:
            error_payload = {"raw": response.text}
        raise RuntimeError(
            f"WhatsApp Cloud API devolvió {response.status_code}: {safe_json_dumps(error_payload)}"
        )
    return response.json()


def send_whatsapp_document_message(to_phone: str, document_link: str, filename: str, caption: Optional[str] = None):
    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone.lstrip("+"),
        "type": "document",
        "document": {
            "link": document_link,
            "filename": filename,
        },
    }
    if caption:
        payload["document"]["caption"] = caption

    response = requests.post(
        f"https://graph.facebook.com/v22.0/{get_whatsapp_phone_number_id()}/messages",
        headers={
            "Authorization": f"Bearer {get_whatsapp_access_token()}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=30,
    )
    if response.status_code >= 400:
        try:
            error_payload = response.json()
        except Exception:
            error_payload = {"raw": response.text}
        raise RuntimeError(
            f"WhatsApp Cloud API devolvió {response.status_code}: {safe_json_dumps(error_payload)}"
        )
    return response.json()


def send_whatsapp_document_bytes(to_phone: str, document_bytes: bytes, filename: str, caption: Optional[str] = None):
    upload_response = requests.post(
        f"https://graph.facebook.com/v22.0/{get_whatsapp_phone_number_id()}/media",
        headers={"Authorization": f"Bearer {get_whatsapp_access_token()}"},
        data={"messaging_product": "whatsapp", "type": "application/pdf"},
        files={"file": (filename, document_bytes, "application/pdf")},
        timeout=30,
    )
    if upload_response.status_code >= 400:
        try:
            error_payload = upload_response.json()
        except Exception:
            error_payload = {"raw": upload_response.text}
        raise RuntimeError(
            f"WhatsApp media upload devolvió {upload_response.status_code}: {safe_json_dumps(error_payload)}"
        )

    media_payload = upload_response.json()
    media_id = media_payload.get("id")
    if not media_id:
        raise RuntimeError(f"WhatsApp media upload no devolvió id: {safe_json_dumps(media_payload)}")

    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone.lstrip("+"),
        "type": "document",
        "document": {
            "id": media_id,
            "filename": filename,
        },
    }
    if caption:
        payload["document"]["caption"] = caption

    response = requests.post(
        f"https://graph.facebook.com/v22.0/{get_whatsapp_phone_number_id()}/messages",
        headers={
            "Authorization": f"Bearer {get_whatsapp_access_token()}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=30,
    )
    if response.status_code >= 400:
        try:
            error_payload = response.json()
        except Exception:
            error_payload = {"raw": response.text}
        raise RuntimeError(
            f"WhatsApp Cloud API devolvió {response.status_code}: {safe_json_dumps(error_payload)}"
        )
    return response.json()




def _handle_tool_consultar_inventario(args, conversation_context):
    # ── Nuevo schema V3.1: nombre_base + variante_o_color ──
    nombre_base = (args.get("nombre_base") or "").strip()
    variante_o_color = (args.get("variante_o_color") or "").strip()
    producto_legacy = (args.get("producto") or "").strip()

    # Construir query de búsqueda priorizando el schema nuevo
    if nombre_base:
        # Schema nuevo: buscar por nombre_base, luego filtrar por variante
        producto_busqueda = nombre_base
        if variante_o_color:
            producto_busqueda = f"{nombre_base} {variante_o_color}"
    elif producto_legacy:
        # Fallback: schema antiguo (backward compat)
        producto_busqueda = producto_legacy
    else:
        return json.dumps(
            {"encontrados": 0, "mensaje": "Se requiere nombre_base o producto."},
            ensure_ascii=False,
        )
    producto_raw = producto_busqueda
    # ── Phase 20: Traducir jerga coloquial del cliente a términos de catálogo en Python ──
    producto = translate_customer_jargon(producto_raw)
    # Skip NLU (OpenAI) call — the main LLM already parsed the product query via tool call
    base_request = extract_product_request(producto)
    base_request["nlu_processed"] = True
    base_request = apply_deterministic_product_alias_rules(producto, base_request)
    product_request = build_followup_inventory_request(
        producto,
        base_request,
        conversation_context,
    )
    product_request["nlu_processed"] = True
    rows = lookup_product_context(producto, product_request)

    # ── Broader fallback: si la búsqueda combinada (base+variante) no encontró nada,
    # reintenta solo con nombre_base para dar más resultados ──
    if not rows and variante_o_color and nombre_base:
        producto_solo_base = translate_customer_jargon(nombre_base)
        base_only_request = extract_product_request(producto_solo_base)
        base_only_request["nlu_processed"] = True
        base_only_request = apply_deterministic_product_alias_rules(producto_solo_base, base_only_request)
        base_only_product_req = build_followup_inventory_request(
            producto_solo_base, base_only_request, conversation_context,
        )
        base_only_product_req["nlu_processed"] = True
        rows = lookup_product_context(producto_solo_base, base_only_product_req)

    # ── Post-variante re-ranking (Rec 2): si se proporcionó variante_o_color,
    # re-ordena resultados priorizando los que contienen la variante en la descripción ──
    if variante_o_color and rows and len(rows) > 1:
        _var_lower = variante_o_color.lower()
        _var_terms = [t for t in _var_lower.split() if len(t) > 2]

        def _variante_score(row):
            desc = (
                (row.get("descripcion_comercial") or "")
                + " " + (row.get("descripcion") or "")
                + " " + (row.get("descripcion_ebs") or "")
            ).lower()
            score = 0
            for term in _var_terms:
                if term in desc:
                    score += 1
            # Exact substring match gets bonus
            if _var_lower in desc:
                score += 5
            return score

        rows = sorted(rows, key=_variante_score, reverse=True)

    requested_store_codes = product_request.get("store_filters") or []
    requested_store_code = requested_store_codes[0] if len(requested_store_codes) == 1 else None
    if not rows and product_request.get("followup_from_previous_product"):
        rows = filter_previous_product_context(conversation_context, product_request)
    if not rows:
        # Check known portfolio gaps (products Ferreinox doesn't carry)
        producto_key = producto.strip().lower()
        if producto_key in PORTFOLIO_GAPS:
            return json.dumps(
                {
                    "encontrados": 0,
                    "gap_portafolio": True,
                    "mensaje": PORTFOLIO_GAPS[producto_key],
                    "estrategia_ranking": "catalogo_curado_postgresql",
                },
                ensure_ascii=False,
            )
        return json.dumps(
            {
                "encontrados": 0,
                "mensaje": "No se encontraron productos con esa descripción.",
                "nlu_extraccion": product_request.get("nlu_extraction") or {},
                "estrategia_ranking": "catalogo_curado_postgresql",
                "requiere_aclaracion": False,
            },
            ensure_ascii=False,
        )
    results = []
    for row in rows[:10]:
        item = {
            "codigo": row.get("codigo_articulo") or row.get("referencia") or row.get("codigo"),
            "descripcion": get_exact_product_description(row),
            "descripcion_exacta": get_exact_product_description(row),
            "etiqueta_auditable": build_product_audit_label(row),
            "marca": row.get("marca") or row.get("marca_producto"),
            "presentacion": infer_product_presentation_from_row(row),
            "familia_consulta": row.get("familia_consulta"),
            "producto_padre_busqueda": row.get("producto_padre_busqueda"),
            "pregunta_desambiguacion": row.get("pregunta_desambiguacion"),
        }
        # Clasificación enriquecida desde articulos_maestro
        _lc = row.get("linea_clasificacion")
        _mc = row.get("marca_clasificacion")
        _fc = row.get("familia_clasificacion")
        _ac = row.get("aplicacion_clasificacion")
        _cp = row.get("cat_producto")
        _de = row.get("descripcion_ebs")
        _ta = row.get("tipo_articulo")
        clasificacion = {}
        if _lc:
            clasificacion["linea"] = _lc
        if _mc:
            clasificacion["tipo_marca"] = _mc
        if _fc:
            clasificacion["familia"] = _fc
        if _ac:
            clasificacion["aplicacion"] = _ac
        if _cp:
            clasificacion["categoria"] = _cp
        if _de:
            clasificacion["descripcion_larga"] = _de
        if _ta:
            clasificacion["tipo"] = _ta
        if clasificacion:
            item["clasificacion"] = clasificacion
        # Abracol enrichment: nombre comercial, familia, descripcion larga, portafolio
        _ab_nc = row.get("nombre_comercial_abracol")
        _ab_fa = row.get("familia_abracol")
        _ab_dl = row.get("descripcion_larga_abracol")
        _ab_pf = row.get("portafolio_abracol")
        if _ab_nc or _ab_fa or _ab_dl:
            abracol = {}
            if _ab_nc:
                abracol["nombre_comercial"] = _ab_nc
            if _ab_fa:
                abracol["familia"] = _ab_fa
            if _ab_dl:
                abracol["descripcion_detallada"] = _ab_dl
            if _ab_pf:
                abracol["portafolio"] = _ab_pf
            item["info_catalogo_complementario"] = abracol
        stock = parse_numeric_value(row.get("stock_total"))
        # Solo indicar disponible/no disponible, NUNCA cantidades exactas
        item["disponible"] = (stock or 0) > 0
        requested_store_stock = row.get("stock_en_tienda_solicitada")
        if requested_store_stock is None and requested_store_code:
            requested_store_stock = extract_store_stock_from_summary(row.get("stock_por_tienda"), requested_store_code)
        if requested_store_stock is not None:
            item["disponible_tienda_solicitada"] = requested_store_stock > 0
        if requested_store_code:
            item["visibilidad_tienda_exacta"] = bool(row.get("visibilidad_tienda_exacta") or requested_store_stock is not None)
            item["tienda_solicitada"] = STORE_CODE_LABELS.get(requested_store_code) or requested_store_code
        stock_189 = parse_numeric_value(row.get("stock_189"))
        if stock_189 is not None:
            item["disponible_pereira"] = stock_189 > 0
        precio = row.get("precio_venta")
        if precio is not None:
            item["precio"] = precio
        # --- Price lookup from agent_precios ---
        ref_code = item.get("codigo") or ""
        price_info = fetch_product_price(str(ref_code))
        if price_info and price_info.get("precio_mejor"):
            pvp = float(price_info["precio_mejor"])
            item["precio_unitario"] = round(pvp)
            item["nota_precio"] = "Este precio es ANTES DE IVA. Para el total al cliente: Subtotal (precio × cantidad) + IVA 19% = Total a Pagar."
            if price_info.get("pvp_franquicia") and not price_info.get("pvp_sap"):
                item["lista_precio"] = "franquicia"
        elif not precio:
            # --- Fallback: International price enrichment from JSON ---
            _intl_entry = _INTERNATIONAL_PRODUCTS_BY_CODE.get(str(ref_code).strip())
            if _intl_entry:
                # Determine which role this code plays in the entry
                _is_base_gal = str(ref_code) == (_intl_entry.get("codigo_base_galon") or "").strip()
                _is_cat_gal = str(ref_code) == (_intl_entry.get("codigo_cat_galon") or "").strip()
                _is_base_cun = str(ref_code) == (_intl_entry.get("codigo_cunete") or "").strip()
                _is_cat_cun = str(ref_code) == (_intl_entry.get("codigo_cat_cunete") or "").strip()
                _is_acr_gal = str(ref_code) == (_intl_entry.get("codigo_galon") or "").strip()
                if _is_base_gal and _intl_entry.get("precio_galon"):
                    item["precio_unitario"] = int(_intl_entry["precio_galon"])
                    item["nota_precio"] = "Precio International IVA INCLUIDO. NO sumes IVA de nuevo."
                    item["precio_iva_incluido"] = True
                elif _is_cat_gal and _intl_entry.get("precio_cat_galon"):
                    item["precio_unitario"] = int(_intl_entry["precio_cat_galon"])
                    item["nota_precio"] = "Precio International IVA INCLUIDO. NO sumes IVA de nuevo."
                    item["precio_iva_incluido"] = True
                elif _is_base_cun and _intl_entry.get("precio_cunete"):
                    item["precio_unitario"] = int(_intl_entry["precio_cunete"])
                    item["nota_precio"] = "Precio International IVA INCLUIDO. NO sumes IVA de nuevo."
                    item["precio_iva_incluido"] = True
                elif _is_cat_cun and _intl_entry.get("precio_cat_cunete"):
                    item["precio_unitario"] = int(_intl_entry["precio_cat_cunete"])
                    item["nota_precio"] = "Precio International IVA INCLUIDO. NO sumes IVA de nuevo."
                    item["precio_iva_incluido"] = True
                elif _is_acr_gal and _intl_entry.get("precio_galon"):
                    item["precio_unitario"] = int(_intl_entry["precio_galon"])
                    item["nota_precio"] = "Precio International IVA INCLUIDO. NO sumes IVA de nuevo."
                    item["precio_iva_incluido"] = True
                # Enrich with International product info
                item["producto_international"] = _intl_entry.get("producto", "")
                item["ral"] = _intl_entry.get("ral", "")
                item["base"] = _intl_entry.get("base", "")
                if _intl_entry.get("kit_galon"):
                    item["precio_kit_galon_iva_inc"] = int(_intl_entry["kit_galon"])
                if _intl_entry.get("kit_cunete"):
                    item["precio_kit_cunete_iva_inc"] = int(_intl_entry["kit_cunete"])
                # Inject catalyst as companion if not already present
                if (_is_base_gal or _is_base_cun) and _intl_entry.get("codigo_cat_galon"):
                    item.setdefault("productos_complementarios", [])
                    item["productos_complementarios"].append({
                        "referencia": _intl_entry.get("codigo_cat_galon", ""),
                        "descripcion": f"Catalizador {_intl_entry.get('producto', '')}",
                        "tipo": "catalizador",
                        "precio_iva_inc": int(_intl_entry.get("precio_cat_galon", 0) or 0),
                    })
            else:
                item["precio_unitario"] = None
                item["precio_nota"] = "Precio pendiente de confirmación"
        # --- Companion/complementary products ---
        ref_for_companion = item.get("codigo") or ""
        companions = fetch_product_companions(ref_for_companion)
        if companions:
            item["productos_complementarios"] = [
                {
                    "referencia": c.get("companion_referencia"),
                    "descripcion": c.get("companion_descripcion") or c.get("descripcion_inventario"),
                    "tipo": c.get("tipo_relacion"),
                    "proporcion": c.get("proporcion"),
                    "notas": c.get("notas"),
                    "stock_total": c.get("stock_total"),
                }
                for c in companions
            ]

        # --- BICOMPONENTE AUTO-INJECTION: si el producto es bicomponente y no
        # se encontró catalizador en companion DB, inyectar desde BICOMPONENT_CATALOG ---
        desc_lower = (item.get("descripcion") or "").lower()
        nombre_lower = (item.get("nombre") or "").lower()
        producto_text_lower = f"{desc_lower} {nombre_lower}"
        has_catalyst_companion = any(
            "catalizador" in (c.get("tipo") or "").lower() or
            "comp b" in (c.get("tipo") or "").lower() or
            "hardener" in (c.get("tipo") or "").lower()
            for c in (companions or [])
        )
        if not has_catalyst_companion:
            for bicomp_key, bicomp_info in BICOMPONENT_CATALOG.items():
                if bicomp_key in producto_text_lower:
                    cat_code = bicomp_info.get("componente_b_codigo", "")
                    cat_desc = bicomp_info.get("componente_b_descripcion", "")
                    proporcion = bicomp_info.get("proporcion_galon", bicomp_info.get("proporcion", ""))
                    item["⚠️_BICOMPONENTE_OBLIGATORIO"] = {
                        "advertencia": f"ESTE PRODUCTO ES BICOMPONENTE. NUNCA cotizar sin catalizador.",
                        "catalizador_codigo": cat_code,
                        "catalizador_descripcion": cat_desc,
                        "proporcion": proporcion,
                        "nota": bicomp_info.get("nota", ""),
                    }
                    # Auto-lookup catalyst price
                    if cat_code:
                        cat_price_info = fetch_product_price(cat_code)
                        if cat_price_info and cat_price_info.get("precio_mejor"):
                            cat_pvp = float(cat_price_info["precio_mejor"])
                            item["⚠️_BICOMPONENTE_OBLIGATORIO"]["catalizador_precio_unitario"] = round(cat_pvp)
                    break

        results.append(item)
    if rows:
        conversation_context["last_product_request"] = product_request
        conversation_context["last_product_query"] = producto
        conversation_context["last_product_context"] = results[:10]
    clarification_required = should_ask_product_clarification(product_request, rows)
    clarification_question = build_best_product_clarification_question(product_request, rows) if clarification_required else None
    # ── Conocimiento experto como señal adicional en inventario ──
    _expert_query_parts = [producto]
    for _r in results[:3]:
        _expert_query_parts.append(_r.get("descripcion") or "")
    _expert_inv = fetch_expert_knowledge(" ".join(_expert_query_parts), limit=4)
    response_payload = {
        "encontrados": len(results),
        "productos": results,
        "seguimiento_producto_previo": bool(product_request.get("followup_from_previous_product")),
        "nlu_extraccion": product_request.get("nlu_extraction") or {},
        "estrategia_ranking": "catalogo_curado_postgresql",
        "requiere_aclaracion": clarification_required,
        "pregunta_desambiguacion": clarification_question,
        "mensaje": (
            "Se recuperó el producto del mensaje anterior para resolver este seguimiento."
            if product_request.get("followup_from_previous_product") else None
        ),
    }
    if _expert_inv:
        response_payload["conocimiento_experto_producto"] = [
            {
                "tipo": n["tipo"],
                "contexto": n["contexto_tags"],
                "recomendar": n.get("producto_recomendado"),
                "evitar": n.get("producto_desestimado"),
                "nota": n["nota_comercial"],
            }
            for n in _expert_inv
        ]
        response_payload["instruccion_conocimiento_experto"] = (
            "💡 CONOCIMIENTO EXPERTO FERREINOX DISPONIBLE para este producto. "
            "Lee 'conocimiento_experto_producto' ANTES de presentar resultados al cliente. "
            "Si hay un 'evitar', NO lo recomiendes para este contexto. "
            "Si hay reglas comerciales (sobre pedido, obligatorio cuarzo, preguntar m², etc.), APLÍCALAS. "
            "Presenta las notas relevantes como '💡 Experiencia Ferreinox: [nota]'."
        )
    return json.dumps(response_payload, ensure_ascii=False, default=str)


def _handle_tool_consultar_inventario_lote(args, conversation_context):
    """Batch inventory lookup — processes multiple products in one call for speed."""
    productos_raw = args.get("productos") or []
    if not productos_raw:
        return json.dumps({"encontrados": 0, "mensaje": "No se enviaron productos."}, ensure_ascii=False)

    # ── Normalizar: soportar tanto schema nuevo (objetos) como antiguo (strings) ──
    productos: list[str] = []
    for item in productos_raw[:15]:
        if isinstance(item, dict):
            # Schema nuevo V3.1: {"nombre_base": "Koraza", "variante_o_color": "blanco galón"}
            nb = (item.get("nombre_base") or "").strip()
            vc = (item.get("variante_o_color") or "").strip()
            if nb:
                productos.append(f"{nb} {vc}".strip() if vc else nb)
        elif isinstance(item, str):
            # Schema antiguo: "Koraza blanco galon"
            productos.append(item.strip())
    if not productos:
        return json.dumps({"encontrados": 0, "mensaje": "No se enviaron productos válidos."}, ensure_ascii=False)

    all_results = []
    # Pre-warm rotation cache once for the entire batch
    try:
        engine = get_db_engine()
        with engine.connect() as connection:
            fetch_rotation_cache(connection)  # Warms global cache
    except Exception:
        pass

    for producto_text in productos[:15]:  # Cap at 15 items max
        producto_text = str(producto_text).strip()
        if not producto_text:
            continue
        try:
            # ── Phase 20: Traducir jerga coloquial antes de procesar ──
            producto_text = translate_customer_jargon(producto_text)
            # Skip NLU (OpenAI) call for batch items — the main LLM already parsed them
            base_request = extract_product_request(producto_text)
            base_request["nlu_processed"] = True
            base_request = apply_deterministic_product_alias_rules(producto_text, base_request)
            product_request = build_followup_inventory_request(
                producto_text,
                base_request,
                conversation_context,
            )
            product_request["nlu_processed"] = True  # Ensure it stays set
            rows = lookup_product_context(producto_text, product_request)
            if not rows:
                producto_key = producto_text.strip().lower()
                if producto_key in PORTFOLIO_GAPS:
                    all_results.append({
                        "busqueda": producto_text,
                        "encontrados": 0,
                        "gap_portafolio": True,
                        "mensaje": PORTFOLIO_GAPS[producto_key],
                    })
                    continue
                all_results.append({
                    "busqueda": producto_text,
                    "encontrados": 0,
                    "productos": [],
                    "mensaje": "No se encontraron productos con esa descripción.",
                })
                continue

            items = []
            for row in rows[:5]:  # Top 5 per product (less than single to save tokens)
                item = {
                    "codigo": row.get("codigo_articulo") or row.get("referencia") or row.get("codigo"),
                    "descripcion": get_exact_product_description(row),
                    "descripcion_exacta": get_exact_product_description(row),
                    "etiqueta_auditable": build_product_audit_label(row),
                    "marca": row.get("marca") or row.get("marca_producto"),
                    "presentacion": infer_product_presentation_from_row(row),
                }
                stock = parse_numeric_value(row.get("stock_total"))
                item["disponible"] = (stock or 0) > 0
                # Price lookup
                ref_code = item.get("codigo") or ""
                price_info = fetch_product_price(str(ref_code))
                if price_info and price_info.get("precio_mejor"):
                    pvp = float(price_info["precio_mejor"])
                    item["precio_unitario"] = round(pvp)
                    item["nota_precio"] = "Este precio es ANTES DE IVA. Para el total al cliente: Subtotal (precio × cantidad) + IVA 19% = Total a Pagar."
                elif not item.get("precio"):
                    item["precio_unitario"] = None
                    item["precio_nota"] = "Precio pendiente de confirmación"
                items.append(item)

            clarification_required = should_ask_product_clarification(product_request, rows)
            clarification_question = build_best_product_clarification_question(product_request, rows) if clarification_required else None
            all_results.append({
                "busqueda": producto_text,
                "encontrados": len(items),
                "productos": items,
                "requiere_aclaracion": clarification_required,
                "pregunta_desambiguacion": clarification_question,
                "nlu_extraccion": product_request.get("nlu_extraction") or {},
            })

            # Store last product context
            if rows:
                conversation_context["last_product_request"] = product_request
                conversation_context["last_product_query"] = producto_text
                conversation_context["last_product_context"] = items[:10]

        except Exception as exc:
            logger.warning("Batch lookup error for '%s': %s", producto_text, exc)
            all_results.append({
                "busqueda": producto_text,
                "encontrados": 0,
                "error": str(exc),
            })

    # ── Conocimiento experto como señal adicional en lote ──
    _lote_query_parts = [str(p) for p in productos[:8]]
    _expert_lote = fetch_expert_knowledge(" ".join(_lote_query_parts), limit=6)
    lote_payload = {
        "total_buscados": len(all_results),
        "resultados": all_results,
        "estrategia_ranking": "catalogo_curado_postgresql",
    }
    if _expert_lote:
        lote_payload["conocimiento_experto_producto"] = [
            {
                "tipo": n["tipo"],
                "contexto": n["contexto_tags"],
                "recomendar": n.get("producto_recomendado"),
                "evitar": n.get("producto_desestimado"),
                "nota": n["nota_comercial"],
            }
            for n in _expert_lote
        ]
        lote_payload["instruccion_conocimiento_experto"] = (
            "💡 CONOCIMIENTO EXPERTO FERREINOX para este lote. "
            "Lee 'conocimiento_experto_producto' ANTES de presentar resultados. "
            "Aplica reglas comerciales (sobre pedido, obligatorio cuarzo, preguntar m², etc.). "
            "Si hay un 'evitar', NO lo recomiendes."
        )
    return json.dumps(lote_payload, ensure_ascii=False, default=str)


def _handle_tool_verificar_identidad(args, context, conversation_context):
    criterio = args.get("criterio_busqueda", "").strip()
    if not criterio:
        return json.dumps({"verificado": False, "mensaje": "No se proporcionó criterio de búsqueda."}, ensure_ascii=False)

    is_numeric = bool(re.fullmatch(r"[\d\-\.]+", criterio.replace(" ", "")))

    verified_context = None
    verified_by = None

    if is_numeric:
        identity_candidate = {"type": "numeric_lookup", "value": criterio}
        try:
            verified_context, verified_by = resolve_identity_candidate(
                identity_candidate, context.get("telefono_e164", "")
            )
        except Exception:
            verified_context, verified_by = None, None
    else:
        try:
            name_result = find_cliente_contexto_by_name(criterio)
            if name_result:
                verified_context = name_result
                verified_by = "name"
        except Exception:
            verified_context, verified_by = None, None

    if verified_context:
        cliente_codigo = verified_context.get("cliente_codigo")
        try:
            cliente_id = update_contact_cliente(context["contact_id"], cliente_codigo)
            context["cliente_id"] = cliente_id
        except Exception:
            pass
        update_conversation_context(
            context["conversation_id"],
            {
                "verified": True,
                "verified_document": criterio if is_numeric else None,
                "verified_by": verified_by,
                "verified_cliente_codigo": cliente_codigo,
                "awaiting_verification": False,
                "awaiting_name_confirmation": False,
            },
        )
        conversation_context.update(
            {
                "verified": True,
                "verified_document": criterio if is_numeric else None,
                "verified_by": verified_by,
                "verified_cliente_codigo": cliente_codigo,
            }
        )
        result = {
                "verificado": True,
                "nombre_cliente": verified_context.get("nombre_cliente"),
                "cliente_codigo": cliente_codigo,
                "ciudad": verified_context.get("ciudad"),
                "nit": verified_context.get("nit"),
            }
        # Enrich with agent_clientes data (by NIF, cliente_codigo, or criterio)
        enrich_keys = [verified_context.get("nit"), cliente_codigo, criterio if is_numeric else None]
        client_extra = None
        for ek in enrich_keys:
            if ek and not client_extra:
                client_extra = fetch_client_by_nif_or_codigo(str(ek))
        if client_extra:
            if client_extra.get("categoria"):
                result["categoria_cliente"] = client_extra["categoria"]
            if client_extra.get("ciudad"):
                result["ciudad"] = result.get("ciudad") or client_extra["ciudad"]
            if client_extra.get("email"):
                result["email"] = client_extra["email"]
            if client_extra.get("segmento"):
                result["segmento"] = client_extra["segmento"]
            if client_extra.get("direccion"):
                result["direccion"] = client_extra["direccion"]
            if client_extra.get("razon_social"):
                result["razon_social"] = client_extra["razon_social"]
        return json.dumps(result, ensure_ascii=False, default=str)
    else:
        tipo = "documento" if is_numeric else "nombre"
        return json.dumps(
            {
                "verificado": False,
                "mensaje": f"No se encontró un cliente con ese {tipo}: {criterio}. "
                "Puede estar incorrecto o no estar registrado.",
            },
            ensure_ascii=False,
        )


# ---------------------------------------------------------------------------
# TOOL: registrar_cliente_nuevo — Registra un cliente nuevo en agent_clientes
# y lo vincula al contacto WhatsApp actual.
# ---------------------------------------------------------------------------
def _handle_tool_registrar_cliente_nuevo(args, context, conversation_context):
    """Register a new client in agent_clientes and link to WhatsApp contact."""
    modo_registro = (args.get("modo_registro") or args.get("tipo_documento") or "").strip().lower()
    if modo_registro not in {"cotizacion", "pedido"}:
        modo_registro = (conversation_context.get("commercial_draft") or {}).get("tipo_documento") or "pedido"
    if modo_registro not in {"cotizacion", "pedido"}:
        modo_registro = "pedido"

    nombre = (args.get("nombre_completo") or "").strip()
    cedula_nit = (args.get("cedula_nit") or "").strip()
    telefono = (args.get("telefono") or context.get("telefono_e164") or "").strip()
    direccion = (args.get("direccion_entrega") or "").strip()
    ciudad = (args.get("ciudad") or "").strip()
    email = (args.get("email") or "").strip()

    if not nombre:
        return json.dumps({"registrado": False, "mensaje": "Falta el nombre completo del cliente."}, ensure_ascii=False)
    if not cedula_nit:
        return json.dumps({"registrado": False, "mensaje": "Falta la cédula o NIT del cliente."}, ensure_ascii=False)
    if modo_registro == "pedido" and not direccion:
        return json.dumps({"registrado": False, "mensaje": "Falta la dirección de entrega."}, ensure_ascii=False)
    if modo_registro == "pedido" and not ciudad:
        return json.dumps({"registrado": False, "mensaje": "Falta la ciudad de entrega."}, ensure_ascii=False)

    cedula_clean = cedula_nit.replace(".", "").replace("-", "").replace(" ", "")
    telefono_normalizado = normalize_phone_e164(telefono) or telefono
    document_type = "NIT" if len(cedula_clean) > 10 else "CC"

    def _split_customer_name(full_name: str):
        parts = [part for part in full_name.split() if part]
        return {
            "nombre_1": parts[0] if len(parts) > 0 else "",
            "otros_nombres": " ".join(parts[1:-2]) if len(parts) > 3 else (parts[1] if len(parts) > 1 else ""),
            "apellido_1": parts[-2] if len(parts) >= 3 else "",
            "apellido_2": parts[-1] if len(parts) >= 4 else "",
        }

    def _next_digital_customer_code(connection):
        max_agent = connection.execute(text(
            "SELECT COALESCE(MAX(codigo), 900000) FROM public.agent_clientes WHERE codigo >= 900000"
        )).scalar() or 900000
        max_public = connection.execute(text(
            "SELECT COALESCE(MAX(CASE WHEN codigo ~ '^[0-9]+$' THEN codigo::bigint END), 900000) FROM public.cliente"
        )).scalar() or 900000
        return int(max(max_agent, max_public)) + 1

    existing_context = None
    try:
        existing_context = find_cliente_contexto_by_document(cedula_clean)
    except Exception:
        existing_context = None

    ciudad_lower = ciudad.lower().strip()
    ciudades_eje_cafetero = ["pereira", "manizales", "armenia", "dosquebradas", "santa rosa", "chinchiná", "chinchina"]
    ciudad_despacho = ciudad if any(c in ciudad_lower for c in ciudades_eje_cafetero) else "Pereira"
    nota_logistica = ""
    if modo_registro == "pedido" and ciudad_despacho == "Pereira" and not any(c in ciudad_lower for c in ciudades_eje_cafetero):
        nota_logistica = (
            f"⚠️ La ciudad '{ciudad}' está fuera de las zonas principales del Eje Cafetero. "
            "El despacho se centraliza desde Pereira. El equipo de logística se comunicará "
            "con el cliente lo antes posible para coordinar la entrega."
        )

    try:
        engine = get_db_engine()
        with engine.begin() as conn:
            existing_codigo = None
            if existing_context:
                existing_codigo = existing_context.get("cliente_codigo") or existing_context.get("verified_cliente_codigo")
            codigo_cliente = int(existing_codigo) if str(existing_codigo or "").isdigit() else _next_digital_customer_code(conn)

            conn.execute(text(
                """
                INSERT INTO public.cliente (
                    codigo, tipo_documento, numero_documento, nombre_legal, nombre_comercial,
                    email, telefono, celular, direccion, ciudad, segmento, updated_at
                ) VALUES (
                    :codigo, :tipo_documento, :numero_documento, :nombre_legal, :nombre_comercial,
                    :email, :telefono, :celular, :direccion, :ciudad, :segmento, now()
                )
                ON CONFLICT (tipo_documento, numero_documento)
                DO UPDATE SET
                    codigo = COALESCE(public.cliente.codigo, EXCLUDED.codigo),
                    nombre_legal = EXCLUDED.nombre_legal,
                    nombre_comercial = COALESCE(NULLIF(EXCLUDED.nombre_comercial, ''), public.cliente.nombre_comercial),
                    email = COALESCE(NULLIF(EXCLUDED.email, ''), public.cliente.email),
                    telefono = COALESCE(NULLIF(EXCLUDED.telefono, ''), public.cliente.telefono),
                    celular = COALESCE(NULLIF(EXCLUDED.celular, ''), public.cliente.celular),
                    direccion = COALESCE(NULLIF(EXCLUDED.direccion, ''), public.cliente.direccion),
                    ciudad = COALESCE(NULLIF(EXCLUDED.ciudad, ''), public.cliente.ciudad),
                    segmento = COALESCE(NULLIF(EXCLUDED.segmento, ''), public.cliente.segmento),
                    updated_at = now()
                """
            ), {
                "codigo": str(codigo_cliente),
                "tipo_documento": document_type,
                "numero_documento": cedula_clean,
                "nombre_legal": nombre.upper(),
                "nombre_comercial": nombre.upper(),
                "email": email,
                "telefono": telefono_normalizado,
                "celular": telefono_normalizado,
                "direccion": direccion,
                "ciudad": ciudad.upper(),
                "segmento": "WHATSAPP",
            })

            name_parts = _split_customer_name(nombre)
            name_payload = {key: value.upper() for key, value in name_parts.items()}
            existing_agent = conn.execute(text(
                """
                SELECT id
                FROM public.agent_clientes
                WHERE REPLACE(REPLACE(COALESCE(nif, ''), '.', ''), '-', '') = :nif
                   OR codigo = :codigo
                LIMIT 1
                """
            ), {"nif": cedula_clean, "codigo": codigo_cliente}).mappings().first()

            if existing_agent:
                conn.execute(text(
                    """
                    UPDATE public.agent_clientes
                    SET codigo = :codigo,
                        nombre = :nombre,
                        nif = :nif,
                        direccion = COALESCE(NULLIF(:direccion, ''), direccion),
                        telefono = COALESCE(NULLIF(:telefono, ''), telefono),
                        ciudad = COALESCE(NULLIF(:ciudad, ''), ciudad),
                        email = COALESCE(NULLIF(:email, ''), email),
                        nombre_1 = :nombre_1,
                        otros_nombres = :otros_nombres,
                        apellido_1 = :apellido_1,
                        apellido_2 = :apellido_2,
                        categoria = COALESCE(categoria, 'NUEVO'),
                        segmento = 'WHATSAPP',
                        clasificacion = 'CLIENTE_DIGITAL'
                    WHERE id = :id
                    """
                ), {
                    "id": existing_agent["id"],
                    "codigo": codigo_cliente,
                    "nombre": nombre.upper(),
                    "nif": cedula_clean,
                    "direccion": direccion,
                    "telefono": telefono_normalizado,
                    "ciudad": ciudad.upper(),
                    "email": email,
                    **name_payload,
                })
            else:
                conn.execute(text(
                    """
                    INSERT INTO public.agent_clientes
                        (codigo, nombre, nif, direccion, telefono, ciudad, email,
                         nombre_1, otros_nombres, apellido_1, apellido_2,
                         categoria, segmento, clasificacion)
                    VALUES
                        (:codigo, :nombre, :nif, :direccion, :telefono, :ciudad, :email,
                         :nombre_1, :otros_nombres, :apellido_1, :apellido_2,
                         'NUEVO', 'WHATSAPP', 'CLIENTE_DIGITAL')
                    """
                ), {
                    "codigo": codigo_cliente,
                    "nombre": nombre.upper(),
                    "nif": cedula_clean,
                    "direccion": direccion,
                    "telefono": telefono_normalizado,
                    "ciudad": ciudad.upper(),
                    "email": email,
                    **name_payload,
                })

            contact_id = context.get("contact_id")
            if contact_id:
                try:
                    conn.execute(text(
                        """
                        UPDATE public.whatsapp_contacto
                        SET nombre_visible = :nombre,
                            metadata = COALESCE(metadata, '{}'::jsonb) || CAST(:meta AS jsonb),
                            updated_at = NOW()
                        WHERE id = :cid
                        """
                    ), {
                        "nombre": nombre,
                        "meta": json.dumps({"cliente_codigo": codigo_cliente, "cedula": cedula_clean}),
                        "cid": contact_id,
                    })
                except Exception:
                    pass

            try:
                cliente_id = update_contact_cliente(contact_id, str(codigo_cliente)) if contact_id else None
                if cliente_id:
                    context["cliente_id"] = cliente_id
            except Exception:
                pass

            update_conversation_context(
                context["conversation_id"],
                {
                    "verified": True,
                    "verified_document": cedula_clean,
                    "verified_by": "registration",
                    "verified_cliente_codigo": codigo_cliente,
                    "client_registered_now": True,
                },
            )
            conversation_context.update({
                "verified": True,
                "verified_document": cedula_clean,
                "verified_by": "registration",
                "verified_cliente_codigo": codigo_cliente,
            })

            result = {
                "registrado": True,
                "codigo_cliente": codigo_cliente,
                "nombre": nombre.upper(),
                "cedula_nit": cedula_clean,
                "telefono": telefono_normalizado,
                "direccion": direccion,
                "ciudad": ciudad.upper(),
                "ciudad_despacho": ciudad_despacho.upper(),
                "modo_registro": modo_registro,
                "mensaje": (
                    f"Cliente {nombre} validado exitosamente con código {codigo_cliente}."
                    if existing_context else
                    f"Cliente {nombre} registrado exitosamente con código {codigo_cliente}."
                ),
            }
            if modo_registro == "cotizacion":
                result["datos_pendientes_para_pedido"] = ["direccion_entrega", "ciudad"]
            if nota_logistica:
                result["nota_logistica"] = nota_logistica
            return json.dumps(result, ensure_ascii=False)

    except Exception as exc:
        logger.error("Error registrando cliente nuevo: %s", exc)
        return json.dumps(
            {"registrado": False, "mensaje": f"Error técnico al registrar: {exc}"},
            ensure_ascii=False,
        )


def _handle_tool_consultar_cartera(args, conversation_context):
    cliente_codigo = conversation_context.get("verified_cliente_codigo")

    # If not verified, try nombre_o_nit for internal employees
    nombre_o_nit = (args or {}).get("nombre_o_nit", "").strip() if args else ""
    if not cliente_codigo and nombre_o_nit:
        is_internal = bool((conversation_context or {}).get("internal_auth"))
        if is_internal:
            is_numeric = bool(re.fullmatch(r"[\d\-\.]+", nombre_o_nit.replace(" ", "")))
            if is_numeric:
                client_data = fetch_client_by_nif_or_codigo(nombre_o_nit)
                if client_data:
                    cliente_codigo = client_data.get("codigo") or client_data.get("cliente_codigo")
            if not cliente_codigo:
                try:
                    name_result = find_cliente_contexto_by_name(nombre_o_nit)
                    if name_result:
                        cliente_codigo = name_result.get("cliente_codigo")
                except Exception:
                    pass
            if not cliente_codigo:
                return json.dumps(
                    {"error": f"No se encontró cliente '{nombre_o_nit}'. Intenta con el NIT exacto."},
                    ensure_ascii=False,
                )
        else:
            return json.dumps(
                {"error": "Cliente no verificado. Pide la cédula o NIT primero."},
                ensure_ascii=False,
            )

    if not cliente_codigo:
        return json.dumps(
            {"error": "Cliente no verificado. Pide la cédula o NIT primero."},
            ensure_ascii=False,
        )

    result = {}
    try:
        contexto = get_cliente_contexto(cliente_codigo)
        result["nombre_cliente"] = contexto.get("nombre_cliente")
        result["saldo_cartera"] = contexto.get("saldo_cartera")
    except Exception:
        pass

    try:
        overdue = fetch_overdue_documents(cliente_codigo)
        if overdue:
            totals = overdue.get("totals", {})
            result["documentos_vencidos"] = totals.get("documentos_vencidos", 0)
            result["saldo_vencido"] = totals.get("saldo_vencido", 0)
            result["max_dias_vencido"] = totals.get("max_dias_vencido", 0)
            if overdue.get("documents"):
                result["detalle_documentos"] = overdue["documents"][:5]
    except Exception:
        pass

    if not result:
        return json.dumps({"error": "No se pudo consultar la cartera."}, ensure_ascii=False)
    return json.dumps(result, ensure_ascii=False, default=str)


def _handle_tool_consultar_compras(args, conversation_context):
    cliente_codigo = conversation_context.get("verified_cliente_codigo")
    if not cliente_codigo:
        return json.dumps(
            {"error": "Cliente no verificado. Pide la cédula o NIT primero."},
            ensure_ascii=False,
        )

    periodo = args.get("periodo", "")
    purchase_query = extract_purchase_query(periodo) if periodo else {}

    if purchase_query.get("wants_last_purchase"):
        summary = fetch_latest_purchase_detail(cliente_codigo)
    else:
        summary = fetch_purchase_summary(
            cliente_codigo,
            purchase_query.get("start_date"),
            purchase_query.get("end_date"),
        )

    if not summary:
        return json.dumps(
            {"encontrados": 0, "mensaje": "No se encontraron compras en ese periodo."},
            ensure_ascii=False,
        )
    return json.dumps(summary, ensure_ascii=False, default=str)


# ── Traslados internos (v2 agent tool) ────────────────────────────────────────

def _handle_tool_solicitar_traslado_interno(args: dict, context: dict, conversation_context: dict) -> str:
    """Registra solicitud de traslado entre sedes y envía correo a la tienda origen para despacho."""
    internal_auth = conversation_context.get("internal_auth") or {}
    if not internal_auth:
        return json.dumps(
            {"error": "No hay sesión interna activa. Autentícate primero con tu cédula de empleado."},
            ensure_ascii=False,
        )

    employee_ctx = dict(internal_auth.get("employee_context") or {})
    role = internal_auth.get("role") or "empleado"
    full_name = employee_ctx.get("full_name") or "Colaborador"
    cargo = employee_ctx.get("cargo") or role
    employee_store_code = normalize_store_code(employee_ctx.get("store_code") or employee_ctx.get("sede"))

    # Validate access
    if role not in {"administrador", "gerente", "operador", "lider"}:
        return json.dumps(
            {"error": "Tu perfil no tiene permisos para crear solicitudes de traslado."},
            ensure_ascii=False,
        )

    # Resolve stores
    tienda_destino_raw = args.get("tienda_destino") or ""
    tienda_origen_raw = args.get("tienda_origen") or ""
    destino_code = normalize_store_code(tienda_destino_raw)
    origen_code = normalize_store_code(tienda_origen_raw) or employee_store_code

    if not destino_code:
        tiendas = ", ".join(_VENTAS_STORE_SERIES.keys())
        return json.dumps(
            {"error": f"No reconozco la tienda destino '{tienda_destino_raw}'. Disponibles: {tiendas}."},
            ensure_ascii=False,
        )

    destino_label = get_store_short_label(destino_code) or STORE_CODE_LABELS.get(destino_code) or destino_code
    origen_label = get_store_short_label(origen_code) or STORE_CODE_LABELS.get(origen_code) or origen_code or "Sin definir"

    producto_desc = (args.get("producto_descripcion") or "").strip()
    producto_ref = (args.get("producto_referencia") or "").strip() or None
    cantidad = args.get("cantidad") or 0
    notas = (args.get("notas") or "").strip() or None

    # ── Auto-fill from conversation context when product is vague/missing ──
    if (not producto_desc or len(producto_desc) < 5) and conversation_context.get("last_product_context"):
        last_products = conversation_context["last_product_context"]
        if last_products and len(last_products) >= 1:
            best = last_products[0]
            if not producto_desc:
                producto_desc = best.get("descripcion") or best.get("descripcion_exacta") or ""
            if not producto_ref:
                producto_ref = best.get("codigo") or None
    # Also try to fill reference from context if we have description but no reference
    if producto_desc and not producto_ref and conversation_context.get("last_product_context"):
        desc_lower = producto_desc.lower()
        for ctx_product in conversation_context.get("last_product_context") or []:
            ctx_desc = (ctx_product.get("descripcion") or ctx_product.get("descripcion_exacta") or "").lower()
            if ctx_desc and (desc_lower in ctx_desc or ctx_desc in desc_lower):
                producto_ref = ctx_product.get("codigo") or producto_ref
                break

    if not producto_desc:
        return json.dumps({"error": "Falta la descripción del producto."}, ensure_ascii=False)
    if not cantidad or float(cantidad) <= 0:
        return json.dumps({"error": "La cantidad debe ser mayor a cero."}, ensure_ascii=False)

    # Build transfer row compatible with notify_transfer_requests_by_email
    transfer_row = {
        "source_store_code": origen_code,
        "source_store_name": origen_label,
        "destination_store_code": destino_code,
        "destination_store_name": destino_label,
        "referencia": producto_ref or "",
        "reference": producto_ref or "",
        "descripcion": producto_desc,
        "description": producto_desc,
        "suggested_qty": float(cantidad),
        "quantity_requested": float(cantidad),
        "origin_store_code": origen_code,
        "origin_store_name": origen_label,
    }

    # Save to DB
    try:
        engine = get_db_engine()
        with engine.begin() as conn:
            transfer_id = conn.execute(
                text(
                    """
                    INSERT INTO public.agent_transfer_request (
                        requested_by_user_id, requested_via,
                        source_store_code, source_store_name,
                        destination_store_code, destination_store_name,
                        referencia, descripcion, quantity_requested,
                        status, summary, notes, metadata,
                        created_at, updated_at
                    ) VALUES (
                        :uid, 'whatsapp_interno',
                        :src_code, :src_name,
                        :dst_code, :dst_name,
                        :ref, :desc_v, :qty,
                        'pendiente',
                        :summary, :notes_v, CAST(:meta AS jsonb),
                        now(), now()
                    ) RETURNING id
                    """
                ),
                {
                    "uid": (internal_auth.get("employee_context") or {}).get("id"),
                    "src_code": origen_code,
                    "src_name": origen_label,
                    "dst_code": destino_code,
                    "dst_name": destino_label,
                    "ref": producto_ref or "",
                    "desc_v": producto_desc,
                    "qty": float(cantidad),
                    "summary": f"Traslado {producto_desc} {origen_label} -> {destino_label}",
                    "notes_v": notas,
                    "meta": safe_json_dumps({"solicitante": full_name, "cargo": cargo, "args": args}),
                },
            ).scalar_one()
        transfer_row["id"] = transfer_id
    except Exception as exc:
        logger.warning("solicitar_traslado_interno DB insert error: %s", exc)
        transfer_row["id"] = None

    # Send email to ORIGIN store (the one that will dispatch the product)
    email_map = get_transfer_destination_email_map()
    notification: dict = {"sent": [], "errors": []}
    to_email = email_map.get(origen_code) if origen_code else None
    if to_email:
        cc_emails = get_transfer_cc_emails()
        requested_by_name = full_name
        body_html = (
            "<p style='margin:0 0 14px 0;font-size:15px;'>Se generó una solicitud de traslado desde el agente de WhatsApp.</p>"
            f"<div style='background:#ffffff;border:1px solid {CORPORATE_BRAND['brand_border']};border-radius:14px;padding:18px 20px;margin-bottom:18px;'>"
            f"<p style='margin:0 0 8px 0;'><strong>Solicitado por:</strong> {escape(requested_by_name)}</p>"
            f"<p style='margin:0 0 8px 0;'><strong>Cargo:</strong> {escape(str(cargo))}</p>"
            f"<p style='margin:0 0 8px 0;'><strong>Sede origen (despacha):</strong> {escape(origen_label)}</p>"
            f"<p style='margin:0 0 8px 0;'><strong>Sede destino (recibe):</strong> {escape(destino_label)}</p>"
            f"<p style='margin:0;'><strong>Fecha solicitud:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
            "</div>"
            "<table style='width:100%;border-collapse:collapse;background:#ffffff;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;'>"
            "<thead><tr style='background:#111827;color:#ffffff;'>"
            "<th style='padding:12px;text-align:left;'>Referencia</th>"
            "<th style='padding:12px;text-align:left;'>Descripción</th>"
            "<th style='padding:12px;text-align:center;'>Cantidad</th>"
            "</tr></thead><tbody>"
            f"<tr><td style='padding:10px 12px;border-top:1px solid #e5e7eb;'>{escape(producto_ref or 'N/A')}</td>"
            f"<td style='padding:10px 12px;border-top:1px solid #e5e7eb;'>{escape(producto_desc)}</td>"
            f"<td style='padding:10px 12px;border-top:1px solid #e5e7eb;text-align:center;'>{escape(str(format_quantity(float(cantidad))))}</td></tr>"
            "</tbody></table>"
            + (f"<p style='margin-top:18px;'><strong>Observaciones:</strong> {escape(notas)}</p>" if notas else "")
        )
        html_content = build_brand_email_shell("Solicitud de traslado entre sedes", body_html)
        text_content = (
            f"Solicitud de traslado Ferreinox\n"
            f"Solicitado por: {requested_by_name}\n"
            f"Cargo: {cargo}\n"
            f"Origen (despacha): {origen_label}\n"
            f"Destino (recibe): {destino_label}\n"
            f"Producto: {producto_desc}\n"
            f"Referencia: {producto_ref or 'N/A'}\n"
            f"Cantidad: {cantidad}\n"
            f"Observaciones: {notas or 'Sin observaciones'}"
        )
        try:
            attachment_bytes = build_transfer_request_excel_bytes([transfer_row])
            attachment_name = (
                f"traslado_{sanitize_filename_segment(origen_label, 'Origen')}"
                f"_{sanitize_filename_segment(destino_label, 'Destino')}"
                f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            send_sendgrid_email(
                to_email,
                f"Solicitud de traslado | {origen_label} → {destino_label}",
                html_content,
                text_content,
                cc_emails=cc_emails,
                attachments=[
                    {
                        "content": base64.b64encode(attachment_bytes).decode("ascii"),
                        "filename": attachment_name,
                        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "disposition": "attachment",
                    }
                ],
            )
            notification["sent"].append({
                "to_email": to_email,
                "cc_emails": cc_emails,
                "attachment": attachment_name,
            })
        except Exception as exc:
            logger.warning("solicitar_traslado_interno email error: %s", exc)
            notification["errors"].append(str(exc))
    else:
        notification["errors"].append(f"No hay correo configurado para la tienda origen {origen_label}.")

    result: dict = {
        "traslado_registrado": True,
        "id": transfer_row.get("id"),
        "producto": producto_desc,
        "referencia": producto_ref,
        "cantidad": float(cantidad),
        "origen": origen_label,
        "destino": destino_label,
        "ruta": f"{origen_label} → {destino_label}",
        "correo_enviado": bool(notification["sent"]),
        "correo_destino": to_email or "No configurado",
        "cc": get_transfer_cc_emails() if notification["sent"] else [],
        "estado": "REGISTRADO",
        "notas": notas,
        "instruccion_agente": (
            "Confirma al empleado: ✅ Producto, ✅ Cantidad, ✅ Ruta origen→destino, ✅ Estado REGISTRADO, "
            "✅ Correo enviado a (email) con CC a compras@ferreinox.co. "
            "Si hubo error en el correo, indica que el traslado quedó registrado en el sistema aunque el correo no llegó."
        ),
    }
    if notification["errors"]:
        result["errores_correo"] = notification["errors"]
    return json.dumps(result, ensure_ascii=False, default=str)


# ── BI de ventas internas ──────────────────────────────────────────────────────

_VENTAS_STORE_SERIES: dict = {
    "pereira":   {"prefix": "189", "credito": "189G", "contado": "189W"},
    "manizales": {"prefix": "157", "credito": "157G", "contado": "157W"},
    "armenia":   {"prefix": "156"},
    "laureles":  {"prefix": "238"},
    "opalo":     {"prefix": "158"},
    "ferrebox":  {"prefix": "439"},
    "cerritos":  {"prefix": "463"},
}

# Vendedores de mostrador por sede (del sistema de BI Ferreinox Ventas)
# Names must match fn_normalize_text() output exactly
_VENDEDORES_MOSTRADOR: dict = {
    "pereira":   ["ALEJANDRO CARBALLO MARQUEZ", "GEORGINA A. GALVIS HERRERA"],
    "armenia":   ["CRISTIAN CAMILO RENDON MONTES", "FANDRY JOHANA ABRIL PENHA", "JAVIER ORLANDO PATINO HURTADO"],
    "manizales": ["DAVID FELIPE MARTINEZ RIOS", "JHON JAIRO CASTANO MONTES"],
    "laureles":  ["MAURICIO RIOS MORALES"],
    "opalo":     ["MARIA PAULA DEL JESUS GALVIS HERRERA"],
}


def _build_series_condition(store_key: str, tipo_venta: str, param_idx: int = 0):
    """Returns (sql_fragment, params_dict) for serie filter.

    Uses :bind_param notation (value contains %) so the % stays in the
    parameter dict, not in the SQL string — avoids psycopg2 format errors.
    """
    info = _VENTAS_STORE_SERIES.get(store_key)
    if not info:
        return None, {}
    pkey = f"serie_p{param_idx}"
    prefix = info["prefix"]
    if tipo_venta == "credito":
        val = info.get("credito") or f"{prefix}G"
    elif tipo_venta == "contado":
        val = info.get("contado") or f"{prefix}W"
    else:
        val = prefix
    return f"serie ILIKE :{pkey}", {pkey: f"{val}%"}


def _build_series_condition_raw(store_key: str, tipo_venta: str, param_idx: int = 0):
    """Like _build_series_condition but for raw_ventas_detalle (text columns, no fn_ wrappers on serie)."""
    info = _VENTAS_STORE_SERIES.get(store_key)
    if not info:
        return None, {}
    pkey = f"serie_p{param_idx}"
    prefix = info["prefix"]
    if tipo_venta == "credito":
        val = info.get("credito") or f"{prefix}G"
    elif tipo_venta == "contado":
        val = info.get("contado") or f"{prefix}W"
    else:
        val = prefix
    return f"serie ILIKE :{pkey}", {pkey: f"{val}%"}


def _build_canal_case_sql() -> str:
    """Build a SQL CASE expression that classifies a vendedor as 'Mostrador' or 'Comercial'."""
    when_clauses = []
    for names in _VENDEDORES_MOSTRADOR.values():
        for name in names:
            when_clauses.append(f"WHEN fn_normalize_text(nom_vendedor) = '{name}' THEN 'Mostrador'")
    return "CASE " + " ".join(when_clauses) + " ELSE 'Comercial' END"


_MESES_NOMBRES: dict = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def _parse_periodo_ventas(periodo_str: Optional[str]):
    """Parse a natural-language period string.

    Returns (date_from, date_to, label, period_filter) where period_filter is:
      {"type": "month",  "anio": int, "mes": int}  – for reliable anio+mes SQL filter
      {"type": "year",   "anio": int}               – full year
      {"type": "date_range"}                         – fall-back to fecha_venta BETWEEN
    Using anio+mes avoids fn_parse_date format ambiguity (DD/MM/YYYY vs YYYY-MM-DD in CSV).
    """
    import calendar as _cal
    today = date.today()
    normalized = normalize_text_value(periodo_str or "")

    if "hoy" in normalized:
        return today, today, "hoy", {"type": "day", "anio": today.year, "mes": today.month, "dia": today.day}
    if "esta semana" in normalized or "semana actual" in normalized:
        start = today - timedelta(days=today.weekday())
        return start, today, "esta semana", {"type": "date_range"}
    if "este mes" in normalized or "el mes actual" in normalized:
        return today.replace(day=1), today, "este mes", {"type": "month", "anio": today.year, "mes": today.month}
    if any(x in normalized for x in ["este ano", "este año", "año actual", "ano actual"]):
        return today.replace(month=1, day=1), today, "este año", {"type": "year", "anio": today.year}

    # Detect specific month name (e.g. "abril", "enero 2025")
    for mes_name, mes_num in _MESES_NOMBRES.items():
        if mes_name in normalized:
            anio_match = re.search(r'\b(20\d{2})\b', normalized)
            anio = int(anio_match.group(1)) if anio_match else today.year
            last_day = _cal.monthrange(anio, mes_num)[1]
            start = date(anio, mes_num, 1)
            end = min(date(anio, mes_num, last_day), today)
            label = f"{mes_name.capitalize()} {anio}"
            return start, end, label, {"type": "month", "anio": anio, "mes": mes_num}

    query = extract_purchase_query(periodo_str or "")
    if query.get("start_date"):
        return query["start_date"], query["end_date"] or today, query.get("label", periodo_str or ""), {"type": "date_range"}

    return today.replace(day=1), today, "este mes", {"type": "month", "anio": today.year, "mes": today.month}


def _handle_tool_consultar_ventas_internas(args: dict, conversation_context: dict) -> str:
    """Consultas de BI de ventas consolidadas desde raw_ventas_detalle para empleados internos con control de acceso por rol.

    Queries raw_ventas_detalle directly (not vw_ventas_netas) to include ALL channels:
    mostradores, comerciales, credito, contado. Uses fn_parse_* for safe type casting.
    """
    internal_auth = conversation_context.get("internal_auth") or {}
    if not internal_auth:
        return json.dumps(
            {"error": "No hay sesión interna activa. Autentícate primero con tu cédula de empleado."},
            ensure_ascii=False,
        )

    employee_ctx = dict(internal_auth.get("employee_context") or {})
    role = internal_auth.get("role") or "empleado"
    full_name = employee_ctx.get("full_name") or ""
    sede = normalize_text_value(employee_ctx.get("sede") or "")

    # Parse args
    periodo_raw = args.get("periodo") or "este mes"
    tienda_arg = normalize_text_value(args.get("tienda") or "")
    vendedor_arg = normalize_text_value(args.get("vendedor_nombre") or "")
    vendedor_codigo_arg = (args.get("vendedor_codigo") or "").strip()
    canal_arg = (args.get("canal") or "empresa").lower().strip()
    tipo_venta = (args.get("tipo_venta") or "todos").lower().strip()
    desglose = (args.get("desglose") or "total").lower().strip()

    date_from, date_to, period_label, period_filter = _parse_periodo_ventas(periodo_raw)

    # ── Access control ─────────────────────────────────────────────────────────
    if role in {"administrador", "gerente"}:
        tienda_final = tienda_arg
        # Explicit vendor code parameter takes priority
        if vendedor_codigo_arg:
            _clean_code = re.sub(r'[^a-zA-Z0-9]', '', vendedor_codigo_arg)
            vendedor_filter = {"type": "codigo", "value": _clean_code}
            vendedor_filter_label = vendedor_codigo_arg
        # Detect if vendedor_arg looks like a vendor code (digits/dots)
        elif vendedor_arg and re.match(r'^[\d\.]+$', vendedor_arg):
            _clean_code = re.sub(r'[^a-zA-Z0-9]', '', vendedor_arg)
            vendedor_filter = {"type": "codigo", "value": _clean_code}
            vendedor_filter_label = vendedor_arg
        else:
            vendedor_filter = vendedor_arg or None
            vendedor_filter_label = vendedor_arg or None
    elif role == "operador":
        if tienda_arg and tienda_arg != sede:
            return json.dumps(
                {"acceso_denegado": True,
                 "mensaje": f"Tu perfil solo tiene acceso a los datos de {sede or 'tu sede'}. No puedes consultar otras tiendas."},
                ensure_ascii=False,
            )
        tienda_final = tienda_arg or sede
        vendedor_filter = None
        vendedor_filter_label = None
    elif role == "vendedor":
        if vendedor_arg and vendedor_arg not in normalize_text_value(full_name):
            return json.dumps(
                {"acceso_denegado": True,
                 "mensaje": "Solo puedes consultar tus propias ventas."},
                ensure_ascii=False,
            )
        tienda_final = tienda_arg
        codigo_vendedor_erp = employee_ctx.get("codigo_vendedor") or None
        if codigo_vendedor_erp:
            vendedor_filter = {"type": "codigo", "value": str(codigo_vendedor_erp)}
        else:
            name_tokens = [t for t in normalize_text_value(full_name).split() if len(t) >= 4]
            vendedor_filter = {"type": "nombre_tokens", "tokens": name_tokens[:2] if len(name_tokens) >= 2 else name_tokens[:1]}
        vendedor_filter_label = full_name
    else:
        tienda_final = tienda_arg
        vendedor_filter = None
        vendedor_filter_label = None
        desglose = "total"

    # ── Build WHERE clause — now queries raw_ventas_detalle directly ──────────
    # Always filter to FACTURA + NOTA_CREDITO (exclude ALBARAN_PENDIENTE)
    conditions = [
        "(fn_normalize_text(tipo_documento) LIKE '%FACTURA%' OR fn_normalize_text(tipo_documento) LIKE '%NOTA%CREDITO%')"
    ]
    params: dict = {}

    # Period filter
    pf_type = period_filter.get("type", "date_range")
    if pf_type == "day":
        conditions.append("fn_parse_integer(anio) = :pf_anio")
        conditions.append("fn_parse_integer(mes) = :pf_mes")
        conditions.append("EXTRACT(DAY FROM fn_parse_date(fecha_venta))::int = :pf_dia")
        params["pf_anio"] = period_filter["anio"]
        params["pf_mes"] = period_filter["mes"]
        params["pf_dia"] = period_filter["dia"]
    elif pf_type == "month":
        conditions.append("fn_parse_integer(anio) = :pf_anio")
        conditions.append("fn_parse_integer(mes) = :pf_mes")
        params["pf_anio"] = period_filter["anio"]
        params["pf_mes"] = period_filter["mes"]
    elif pf_type == "year":
        conditions.append("fn_parse_integer(anio) = :pf_anio")
        params["pf_anio"] = period_filter["anio"]
    else:
        conditions.append("fn_parse_date(fecha_venta) BETWEEN :date_from AND :date_to")
        params["date_from"] = date_from
        params["date_to"] = date_to

    # Tienda (store) filter by serie prefix
    if tienda_final:
        store_sql, store_params = _build_series_condition_raw(tienda_final, tipo_venta)
        if not store_sql:
            tiendas_disponibles = ", ".join(_VENTAS_STORE_SERIES.keys())
            return json.dumps(
                {"error": f"No reconozco la tienda '{tienda_final}'. Disponibles: {tiendas_disponibles}."},
                ensure_ascii=False,
            )
        conditions.append(store_sql)
        params.update(store_params)
    elif tipo_venta == "credito":
        conditions.append("RIGHT(serie, 1) = 'G'")
    elif tipo_venta == "contado":
        conditions.append("RIGHT(serie, 1) = 'W'")

    # Canal filter (mostradores vs comerciales)
    _ALL_MOSTRADOR_NAMES = []
    for names in _VENDEDORES_MOSTRADOR.values():
        _ALL_MOSTRADOR_NAMES.extend(names)

    if canal_arg == "mostradores":
        if _ALL_MOSTRADOR_NAMES:
            mostrador_placeholders = ", ".join(f":most_{i}" for i in range(len(_ALL_MOSTRADOR_NAMES)))
            conditions.append(f"fn_normalize_text(nom_vendedor) IN ({mostrador_placeholders})")
            for i, name in enumerate(_ALL_MOSTRADOR_NAMES):
                params[f"most_{i}"] = name
    elif canal_arg == "comerciales":
        if _ALL_MOSTRADOR_NAMES:
            mostrador_placeholders = ", ".join(f":most_{i}" for i in range(len(_ALL_MOSTRADOR_NAMES)))
            conditions.append(f"fn_normalize_text(nom_vendedor) NOT IN ({mostrador_placeholders})")
            for i, name in enumerate(_ALL_MOSTRADOR_NAMES):
                params[f"most_{i}"] = name
    # canal_arg == "empresa" → no filter, all channels

    # Vendedor filter
    _MOSTRADOR_GROUPS = {
        "mostrador pereira": "MOSTRADOR PEREIRA",
        "mostrador armenia": "MOSTRADOR ARMENIA",
        "mostrador manizales": "MOSTRADOR MANIZALES",
        "mostrador laureles": "MOSTRADOR LAURELES",
        "mostrador opalo": "MOSTRADOR OPALO",
    }
    if vendedor_filter:
        if isinstance(vendedor_filter, dict):
            if vendedor_filter.get("type") == "codigo":
                conditions.append("fn_keep_alnum(codigo_vendedor) = :vendedor_codigo")
                params["vendedor_codigo"] = re.sub(r'[^a-zA-Z0-9]', '', str(vendedor_filter["value"]))
            elif vendedor_filter.get("type") == "grupo":
                # For grupo filter, match the list of vendedores in that mostrador group
                vf_norm = normalize_text_value(str(vendedor_filter.get("value", "")))
                for tienda_key, mostrador_group_name in _MOSTRADOR_GROUPS.items():
                    if vf_norm == mostrador_group_name or vf_norm == tienda_key:
                        mostrador_names = _VENDEDORES_MOSTRADOR.get(tienda_key.replace("mostrador ", ""), [])
                        if mostrador_names:
                            g_placeholders = ", ".join(f":grp_{i}" for i in range(len(mostrador_names)))
                            conditions.append(f"fn_normalize_text(nom_vendedor) IN ({g_placeholders})")
                            for i, n in enumerate(mostrador_names):
                                params[f"grp_{i}"] = n
                        break
            else:
                for i, tok in enumerate(vendedor_filter.get("tokens", [])):
                    conditions.append(f"fn_normalize_text(nom_vendedor) ILIKE :vendedor_tok{i}")
                    params[f"vendedor_tok{i}"] = f"%{tok}%"
        elif isinstance(vendedor_filter, list):
            for i, tok in enumerate(vendedor_filter):
                conditions.append(f"fn_normalize_text(nom_vendedor) ILIKE :vendedor_tok{i}")
                params[f"vendedor_tok{i}"] = f"%{tok}%"
        elif isinstance(vendedor_filter, str):
            vf_norm = normalize_text_value(vendedor_filter)
            matched_group = _MOSTRADOR_GROUPS.get(vf_norm)
            if matched_group:
                sede_key = vf_norm.replace("mostrador ", "")
                mostrador_names = _VENDEDORES_MOSTRADOR.get(sede_key, [])
                if mostrador_names:
                    g_placeholders = ", ".join(f":grp_{i}" for i in range(len(mostrador_names)))
                    conditions.append(f"fn_normalize_text(nom_vendedor) IN ({g_placeholders})")
                    for i, n in enumerate(mostrador_names):
                        params[f"grp_{i}"] = n
            else:
                conditions.append("fn_normalize_text(nom_vendedor) ILIKE :vendedor_nombre")
                params["vendedor_nombre"] = f"%{vendedor_filter}%"

    where_clause = " AND ".join(conditions)

    # ── Helper: raw table name ─────────────────────────────────────────────────
    _RAW = "public.raw_ventas_detalle"

    # ── Nombre de marca mapping (inline CASE) ─────────────────────────────────
    _MARCA_CASE = """
        CASE fn_parse_integer(marca_producto)
            WHEN 50 THEN 'P8-ASC-MEGA'
            WHEN 54 THEN 'MPY-International'
            WHEN 55 THEN 'DPP-AN COLORANTS LATAM'
            WHEN 56 THEN 'DPP-Pintuco Profesional'
            WHEN 57 THEN 'ASC-Mega'
            WHEN 58 THEN 'DPP-Pintuco'
            WHEN 59 THEN 'DPP-Madetec'
            WHEN 60 THEN 'POW-Interpon'
            WHEN 61 THEN 'various'
            WHEN 62 THEN 'DPP-ICO'
            WHEN 63 THEN 'DPP-Terinsa'
            WHEN 64 THEN 'MPY-Pintuco'
            WHEN 65 THEN 'non-AN Third Party'
            WHEN 66 THEN 'ICO-AN Packaging'
            WHEN 67 THEN 'ASC-Automotive OEM'
            WHEN 68 THEN 'POW-Resicoat'
            WHEN 73 THEN 'DPP-Coral'
            WHEN 91 THEN 'DPP-Sikkens'
            ELSE 'No Especificada'
        END
    """

    # ── Tienda label from serie prefix ─────────────────────────────────────────
    _TIENDA_CASE = """
        CASE LEFT(serie, 3)
            WHEN '189' THEN 'Pereira'
            WHEN '157' THEN 'Manizales'
            WHEN '156' THEN 'Armenia'
            WHEN '238' THEN 'Laureles'
            WHEN '158' THEN 'Opalo'
            WHEN '439' THEN 'Ferrebox'
            WHEN '463' THEN 'Cerritos'
            ELSE 'Otra (' || LEFT(serie, 3) || ')'
        END
    """

    # ── Canal classification (mostrador vs comercial) ──────────────────────────
    _CANAL_CASE = _build_canal_case_sql()

    engine = get_db_engine()

    # ── Sync freshness ────────────────────────────────────────────────────────
    ultima_sincronizacion: Optional[str] = None
    alerta_datos_desactualizados: bool = False
    try:
        with engine.connect() as conn:
            sync_row = conn.execute(
                text("""
                    SELECT executed_at
                    FROM public.sync_run_log
                    WHERE target_table = 'raw_ventas_detalle' AND status = 'success'
                    ORDER BY executed_at DESC
                    LIMIT 1
                """)
            ).mappings().one_or_none()
        if sync_row:
            ts = sync_row["executed_at"]
            ultima_sincronizacion = str(ts)[:19]
            from datetime import datetime as _dt, timezone as _tz
            if hasattr(ts, "tzinfo") and ts.tzinfo:
                age_hours = (_dt.now(_tz.utc) - ts).total_seconds() / 3600
            else:
                age_hours = (_dt.utcnow() - ts).total_seconds() / 3600 if hasattr(ts, "hour") else 9999
            alerta_datos_desactualizados = age_hours > 8
    except Exception:
        pass

    try:
        with engine.connect() as conn:
            total_row = conn.execute(
                text(f"""
                    SELECT
                        COUNT(*) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS num_lineas,
                        COUNT(DISTINCT fn_keep_alnum(cliente_id)) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS num_clientes,
                        SUM(CASE WHEN fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                                 THEN COALESCE(fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturas_bruto,
                        SUM(CASE WHEN fn_normalize_text(tipo_documento) LIKE '%%NOTA%%'
                                 THEN ABS(COALESCE(fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones,
                        SUM(COALESCE(fn_parse_numeric(valor_venta), 0)) AS ventas_netas_directas,
                        COUNT(DISTINCT fn_normalize_text(nom_vendedor)) AS num_vendedores
                    FROM {_RAW}
                    WHERE {where_clause}
                """),
                params,
            ).mappings().one_or_none()
    except Exception as exc:
        logger.error("consultar_ventas_internas DB error: %s", exc, exc_info=True)
        return json.dumps(
            {"error": f"Error consultando base de datos: {str(exc)[:200]}"},
            ensure_ascii=False,
        )

    if not total_row:
        return json.dumps(
            {"encontrado": False, "mensaje": f"No se encontraron datos de ventas para {period_label}."},
            ensure_ascii=False,
        )

    facturas_bruto = float(total_row.get("facturas_bruto") or 0)
    devoluciones = float(total_row.get("devoluciones") or 0)
    neto = facturas_bruto - devoluciones
    ventas_netas_directas = float(total_row.get("ventas_netas_directas") or 0)
    num_lineas = int(total_row.get("num_lineas") or 0)
    num_clientes = int(total_row.get("num_clientes") or 0)
    num_vendedores = int(total_row.get("num_vendedores") or 0)

    canal_label = {"empresa": "consolidado (todos los canales)", "mostradores": "solo mostradores", "comerciales": "solo comerciales"}.get(canal_arg, canal_arg)

    result: dict = {
        "periodo": period_label,
        "tienda": tienda_final or "todas las tiendas",
        "canal": canal_label,
        "vendedor_consultado": vendedor_filter_label or ("todos" if role in {"administrador", "gerente", "operador"} else full_name),
        "tipo_venta": tipo_venta,
        "ventas": {
            "facturas_bruto": round(facturas_bruto, 2),
            "devoluciones_notas_credito": round(devoluciones, 2),
            "ventas_netas": round(neto, 2),
            "ventas_totales_equivalente_app": round(ventas_netas_directas, 2),
            "num_lineas_factura": num_lineas,
            "num_clientes_distintos": num_clientes,
            "num_vendedores": num_vendedores,
        },
    }

    if ultima_sincronizacion:
        result["datos_db"] = {
            "ultima_sincronizacion": ultima_sincronizacion,
            "alerta_datos_desactualizados": alerta_datos_desactualizados,
        }
        if alerta_datos_desactualizados:
            result["datos_db"]["aviso"] = (
                "⚠️ Los datos en la base de datos pueden estar desactualizados (última sincronización "
                f"hace más de 8 h: {ultima_sincronizacion}). "
                "Presiona 'Sincronizar Dropbox' en el panel del frontend para actualizar antes de comparar cifras."
            )

    # ── Año anterior — comparativa ────────────────────────────────────────────
    try:
        prev_anio = period_filter.get("anio", date_from.year) - 1
        prev_from = date_from.replace(year=date_from.year - 1)
        prev_to = date_to.replace(year=date_to.year - 1)
        if pf_type == "month":
            prev_params = {k: v for k, v in params.items() if not k.startswith("pf_")}
            prev_params["pf_anio"] = prev_anio
            prev_params["pf_mes"] = period_filter["mes"]
        elif pf_type == "year":
            prev_params = {k: v for k, v in params.items() if not k.startswith("pf_")}
            prev_params["pf_anio"] = prev_anio
        else:
            prev_params = {k: v for k, v in params.items() if k not in ("date_from", "date_to")}
            prev_params["date_from"] = prev_from
            prev_params["date_to"] = prev_to
        with engine.connect() as conn:
            prev_row = conn.execute(
                text(f"""
                    SELECT
                        SUM(CASE WHEN fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                                THEN COALESCE(fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturas_bruto,
                        SUM(CASE WHEN fn_normalize_text(tipo_documento) LIKE '%%NOTA%%'
                                THEN ABS(COALESCE(fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones
                    FROM {_RAW}
                    WHERE {where_clause}
                """),
                prev_params,
            ).mappings().one_or_none()
        prev_bruto = float((prev_row.get("facturas_bruto") or 0) if prev_row else 0)
        prev_dev = float((prev_row.get("devoluciones") or 0) if prev_row else 0)
        prev_neto = prev_bruto - prev_dev
        if prev_neto > 0:
            variacion_pct = ((neto - prev_neto) / prev_neto) * 100
            result["vs_anio_anterior"] = {
                "anio": prev_anio,
                "ventas_netas": round(prev_neto, 2),
                "variacion_pct": round(variacion_pct, 1),
            }
    except Exception as exc:
        logger.debug("consultar_ventas_internas comparativa año anterior error: %s", exc)

    # ── Desglose por vendedor ─────────────────────────────────────────────────
    if desglose == "por_vendedor" and role in {"administrador", "gerente", "operador"}:
        try:
            with engine.connect() as conn:
                rows = conn.execute(
                    text(f"""
                        SELECT fn_normalize_text(nom_vendedor) AS vendedor_label,
                               fn_keep_alnum(codigo_vendedor) AS codigo,
                               {_CANAL_CASE} AS canal_vendedor,
                               SUM(CASE WHEN fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                                        THEN COALESCE(fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturado,
                               SUM(CASE WHEN fn_normalize_text(tipo_documento) LIKE '%%NOTA%%'
                                        THEN ABS(COALESCE(fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones,
                               COUNT(*) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS lineas,
                               COUNT(DISTINCT fn_keep_alnum(cliente_id)) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS clientes
                        FROM {_RAW}
                        WHERE {where_clause}
                        GROUP BY vendedor_label, codigo, canal_vendedor
                        ORDER BY facturado DESC
                        LIMIT 30
                    """),
                    params,
                ).mappings().all()
            result["desglose_vendedores"] = [
                {
                    "vendedor": r["vendedor_label"],
                    "codigo": r["codigo"],
                    "canal": r["canal_vendedor"],
                    "facturado": round(float(r["facturado"] or 0), 2),
                    "devoluciones": round(float(r["devoluciones"] or 0), 2),
                    "neto": round(float(r["facturado"] or 0) - float(r["devoluciones"] or 0), 2),
                    "lineas": int(r["lineas"] or 0),
                    "clientes": int(r["clientes"] or 0),
                }
                for r in rows
            ]
        except Exception as exc:
            logger.warning("consultar_ventas_internas desglose_vendedores error: %s", exc)

    # ── Desglose por tienda ───────────────────────────────────────────────────
    elif desglose == "por_tienda" and role in {"administrador", "gerente"}:
        try:
            with engine.connect() as conn:
                rows = conn.execute(
                    text(f"""
                        SELECT {_TIENDA_CASE} AS tienda_label,
                               LEFT(serie, 3) AS serie_prefix,
                               SUM(CASE WHEN fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                                        THEN COALESCE(fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturado,
                               SUM(CASE WHEN fn_normalize_text(tipo_documento) LIKE '%%NOTA%%'
                                        THEN ABS(COALESCE(fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones,
                               COUNT(*) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS lineas,
                               COUNT(DISTINCT fn_keep_alnum(cliente_id)) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS clientes,
                               COUNT(DISTINCT fn_normalize_text(nom_vendedor)) AS vendedores
                        FROM {_RAW}
                        WHERE {where_clause}
                        GROUP BY tienda_label, serie_prefix
                        ORDER BY facturado DESC
                    """),
                    params,
                ).mappings().all()
            result["desglose_tiendas"] = [
                {
                    "tienda": r["tienda_label"],
                    "facturado": round(float(r["facturado"] or 0), 2),
                    "devoluciones": round(float(r["devoluciones"] or 0), 2),
                    "neto": round(float(r["facturado"] or 0) - float(r["devoluciones"] or 0), 2),
                    "lineas": int(r["lineas"] or 0),
                    "clientes": int(r["clientes"] or 0),
                    "vendedores": int(r["vendedores"] or 0),
                }
                for r in rows
            ]
        except Exception as exc:
            logger.warning("consultar_ventas_internas desglose_tiendas error: %s", exc)

    # ── Desglose por canal (mostradores vs comerciales) ───────────────────────
    elif desglose == "por_canal" and role in {"administrador", "gerente", "operador"}:
        try:
            with engine.connect() as conn:
                rows = conn.execute(
                    text(f"""
                        SELECT {_CANAL_CASE} AS canal_label,
                               SUM(CASE WHEN fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                                        THEN COALESCE(fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturado,
                               SUM(CASE WHEN fn_normalize_text(tipo_documento) LIKE '%%NOTA%%'
                                        THEN ABS(COALESCE(fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones,
                               COUNT(*) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS lineas,
                               COUNT(DISTINCT fn_keep_alnum(cliente_id)) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS clientes,
                               COUNT(DISTINCT fn_normalize_text(nom_vendedor)) AS vendedores
                        FROM {_RAW}
                        WHERE {where_clause}
                        GROUP BY canal_label
                        ORDER BY facturado DESC
                    """),
                    params,
                ).mappings().all()
            result["desglose_canales"] = [
                {
                    "canal": r["canal_label"],
                    "facturado": round(float(r["facturado"] or 0), 2),
                    "devoluciones": round(float(r["devoluciones"] or 0), 2),
                    "neto": round(float(r["facturado"] or 0) - float(r["devoluciones"] or 0), 2),
                    "lineas": int(r["lineas"] or 0),
                    "clientes": int(r["clientes"] or 0),
                    "vendedores": int(r["vendedores"] or 0),
                }
                for r in rows
            ]
        except Exception as exc:
            logger.warning("consultar_ventas_internas desglose_canales error: %s", exc)

    # ── Desglose por producto ─────────────────────────────────────────────────
    elif desglose == "por_producto":
        try:
            with engine.connect() as conn:
                rows = conn.execute(
                    text(f"""
                        SELECT fn_normalize_text(nombre_articulo) AS nombre_articulo,
                               fn_normalize_text(linea_producto) AS linea_producto,
                               {_MARCA_CASE} AS nombre_marca,
                               SUM(COALESCE(fn_parse_numeric(valor_venta), 0)) AS total,
                               SUM(COALESCE(fn_parse_numeric(unidades_vendidas), 0)) AS unidades
                        FROM {_RAW}
                        WHERE {where_clause}
                          AND fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                        GROUP BY 1, 2, 3
                        ORDER BY total DESC
                        LIMIT 15
                    """),
                    params,
                ).mappings().all()
            result["top_productos"] = [
                {
                    "producto": r["nombre_articulo"],
                    "linea": r["linea_producto"],
                    "marca": r["nombre_marca"],
                    "total": round(float(r["total"] or 0), 2),
                    "unidades": float(r["unidades"] or 0),
                }
                for r in rows
            ]
        except Exception as exc:
            logger.warning("consultar_ventas_internas top_productos error: %s", exc)

    # ── Desglose por cliente ──────────────────────────────────────────────────
    elif desglose == "por_cliente" and role in {"administrador", "gerente", "operador", "vendedor"}:
        try:
            with engine.connect() as conn:
                rows = conn.execute(
                    text(f"""
                        SELECT fn_normalize_text(nombre_cliente) AS nombre_cliente,
                               fn_keep_alnum(cliente_id) AS cliente_id,
                               SUM(COALESCE(fn_parse_numeric(valor_venta), 0)) AS total
                        FROM {_RAW}
                        WHERE {where_clause}
                          AND fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                        GROUP BY 1, 2
                        ORDER BY total DESC
                        LIMIT 20
                    """),
                    params,
                ).mappings().all()
            result["top_clientes"] = [
                {"cliente": r["nombre_cliente"], "total": round(float(r["total"] or 0), 2)}
                for r in rows
            ]
        except Exception as exc:
            logger.warning("consultar_ventas_internas top_clientes error: %s", exc)

    # ── Desglose por día ──────────────────────────────────────────────────────
    elif desglose == "por_dia":
        try:
            with engine.connect() as conn:
                rows = conn.execute(
                    text(f"""
                        SELECT fn_parse_date(fecha_venta) AS fecha,
                               SUM(CASE WHEN fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%'
                                        THEN COALESCE(fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturado,
                               COUNT(*) FILTER (WHERE fn_normalize_text(tipo_documento) NOT LIKE '%%NOTA%%') AS lineas
                        FROM {_RAW}
                        WHERE {where_clause}
                        GROUP BY fecha
                        ORDER BY fecha
                    """),
                    params,
                ).mappings().all()
            result["desglose_dias"] = [
                {
                    "fecha": str(r["fecha"]),
                    "facturado": round(float(r["facturado"] or 0), 2),
                    "lineas": int(r["lineas"] or 0),
                }
                for r in rows
            ]
        except Exception as exc:
            logger.warning("consultar_ventas_internas por_dia error: %s", exc)

    return json.dumps(result, ensure_ascii=False, default=str)


def _send_document_and_respond(doc, context):
    """Helper: send a single document via WhatsApp and return success JSON."""
    filename = doc.get("name") or "documento.pdf"
    path_lower = doc.get("path_lower")
    try:
        temporary_link = get_dropbox_temporary_link(path_lower)
        send_whatsapp_document_message(
            context["telefono_e164"],
            temporary_link,
            filename,
            caption=f"Aquí tienes: {filename}",
        )
        store_outbound_message(
            context["conversation_id"],
            None,
            "document",
            f"Documento técnico enviado: {filename}",
            {"filename": filename, "path": path_lower},
            intent_detectado="consulta_documentacion",
        )
        return json.dumps(
            {"status": "exito", "encontrado": True, "enviado": True, "archivo": filename,
             "mensaje": f"El archivo '{filename}' fue enviado exitosamente por WhatsApp."},
            ensure_ascii=False,
        )
    except Exception as exc:
        return json.dumps(
            {"encontrado": True, "enviado": False, "archivo": filename,
             "mensaje": f"Encontré el archivo '{filename}' pero no pude enviarlo: {exc}"},
            ensure_ascii=False,
        )


def _handle_tool_buscar_documento_tecnico(args, context, conversation_context):
    termino = args.get("termino_busqueda", "")
    es_hds = args.get("es_hoja_de_seguridad", False)
    es_seleccion_final = args.get("es_seleccion_final", False)
    if not termino:
        return json.dumps({"encontrado": False, "mensaje": "No se indicó qué producto buscar."}, ensure_ascii=False)

    product_request = extract_product_request(termino)
    document_request = extract_technical_document_request(
        termino, product_request, conversation_context
    )
    if es_hds:
        document_request["wants_safety_sheet"] = True
        document_request["wants_technical_sheet"] = False
    else:
        document_request["wants_technical_sheet"] = True

    documents = search_technical_documents(document_request)
    if not documents:
        return json.dumps(
            {"encontrado": False, "mensaje": f"No encontré documentos técnicos para '{termino}'."}, ensure_ascii=False
        )

    # --- Exact match: if termino matches a filename exactly, send it immediately ---
    termino_lower = termino.lower().strip()
    for doc in documents:
        doc_name = (doc.get("name") or "").lower().strip()
        if doc_name == termino_lower or doc_name == termino_lower + ".pdf":
            return _send_document_and_respond(doc, context)

    # --- Final selection mode: client already chose, force-send best match ---
    if es_seleccion_final:
        return _send_document_and_respond(documents[0], context)

    # --- Multiple results: ask client to choose ---
    if len(documents) > 1:
        opciones = [d.get("name", "documento.pdf") for d in documents]
        return json.dumps(
            {"status": "multiples_opciones", "opciones": opciones,
             "mensaje": f"Se encontraron {len(opciones)} documentos para '{termino}'. Pregúntale al cliente cuál necesita."},
            ensure_ascii=False,
        )

    # --- Single result: send directly ---
    return _send_document_and_respond(documents[0], context)


def _handle_tool_radicar_reclamo(args, context, conversation_context):
    producto_reclamado = args.get("producto_reclamado", "")
    descripcion_problema = args.get("descripcion_problema", "")
    diagnostico_previo = args.get("diagnostico_previo", "")
    correo_cliente = args.get("correo_cliente", "")
    evidencia = args.get("evidencia", "Pendiente")

    existing_claim = dict(conversation_context.get("claim_case") or {})
    existing_product = normalize_text_value(existing_claim.get("product_label"))
    incoming_product = normalize_text_value(producto_reclamado)
    if existing_claim.get("submitted") and existing_product and incoming_product and existing_product == incoming_product:
        return json.dumps(
            {
                "status": "ya_radicado",
                "numero_caso": existing_claim.get("case_reference") or f"CRM-{context['conversation_id']}",
                "producto": producto_reclamado,
                "mensaje": "Ese reclamo ya estaba radicado en esta conversación. No lo reenvié para evitar duplicados.",
            },
            ensure_ascii=False,
        )

    if not producto_reclamado or not descripcion_problema:
        return json.dumps(
            {"status": "error", "mensaje": "Faltan datos: producto y descripción del problema son requeridos."},
            ensure_ascii=False,
        )

    conversation_id = context["conversation_id"]
    numero_caso = f"CRM-{conversation_id}"

    verified_cliente = conversation_context.get("verified_cliente_codigo")
    cliente_contexto = None
    if verified_cliente:
        try:
            cliente_contexto = get_cliente_contexto(verified_cliente)
        except Exception:
            pass

    recent_messages = load_recent_conversation_messages(conversation_id)

    claim_detail = {
        "product_label": producto_reclamado,
        "issue_summary": descripcion_problema,
        "diagnostico_previo": diagnostico_previo,
        "evidence_note": evidencia,
        "contact_email": correo_cliente,
        "case_reference": numero_caso,
        "store_name": (cliente_contexto or {}).get("ciudad") or "Pendiente",
    }

    # Save claim case in conversation context
    try:
        update_conversation_context(
            conversation_id,
            {
                "claim_case": {
                    "submitted": True,
                    "case_reference": numero_caso,
                    "product_label": producto_reclamado,
                    "issue_summary": descripcion_problema,
                    "contact_email": correo_cliente,
                },
            },
        )
    except Exception:
        pass
    conversation_context["claim_case"] = claim_detail

    # Create agent task for tracking
    try:
        upsert_agent_task(
            conversation_id,
            context.get("cliente_id"),
            "reclamo_servicio",
            f"Reclamo radicado: {producto_reclamado}",
            claim_detail,
            "alta",
        )
    except Exception:
        pass

    correos_enviados = []

    # 1. Internal email to claims department
    try:
        internal_payload = build_operational_email_payload(
            "reclamos",
            context.get("nombre_visible"),
            cliente_contexto,
            claim_detail,
            recent_messages,
        )
        if internal_payload:
            send_sendgrid_email(
                internal_payload["to_email"],
                internal_payload["subject"],
                internal_payload["html_content"],
                internal_payload["text_content"],
                reply_to=correo_cliente,
            )
            correos_enviados.append(f"Área técnica ({internal_payload['to_email']})")
            store_outbound_message(
                conversation_id, None, "system",
                f"Correo reclamo interno enviado a {internal_payload['to_email']}",
                {"email_to": internal_payload["to_email"], "case": numero_caso},
                intent_detectado="correo_reclamo_interno",
            )
    except Exception as exc:
        try:
            store_outbound_message(
                conversation_id, None, "system",
                f"Error enviando correo interno de reclamo: {exc}",
                {"error": str(exc)},
                intent_detectado="correo_reclamo_interno_error",
            )
        except Exception:
            pass

    # 2. Confirmation email to customer
    if correo_cliente:
        try:
            customer_payload = build_customer_claim_confirmation_email(
                conversation_id,
                context.get("nombre_visible"),
                cliente_contexto,
                claim_detail,
            )
            if customer_payload:
                send_sendgrid_email(
                    customer_payload["to_email"],
                    customer_payload["subject"],
                    customer_payload["html_content"],
                    customer_payload["text_content"],
                )
                correos_enviados.append(f"Cliente ({correo_cliente})")
                store_outbound_message(
                    conversation_id, None, "system",
                    f"Correo constancia reclamo enviado a {correo_cliente}",
                    {"email_to": correo_cliente, "case": numero_caso},
                    intent_detectado="correo_reclamo_cliente",
                )
        except Exception as exc:
            try:
                store_outbound_message(
                    conversation_id, None, "system",
                    f"Error enviando constancia al cliente: {exc}",
                    {"error": str(exc)},
                    intent_detectado="correo_reclamo_cliente_error",
                )
            except Exception:
                pass

    return json.dumps(
        {
            "status": "exito",
            "numero_caso": numero_caso,
            "producto": producto_reclamado,
            "correos_enviados": correos_enviados,
            "mensaje": f"Reclamo radicado exitosamente con número {numero_caso}. "
            f"Correos enviados a: {', '.join(correos_enviados) if correos_enviados else 'ninguno (verificar configuración SendGrid)'}.",
        },
        ensure_ascii=False,
    )


def _handle_tool_confirmar_pedido(args, context, conversation_context):
    t_confirm_start = time.time()
    nombre_despacho = args.get("nombre_despacho", "")
    canal_envio = args.get("canal_envio", "whatsapp")
    correo_cliente = args.get("correo_cliente", "")
    items_pedido = args.get("items_pedido") or []
    tipo_documento = args.get("tipo_documento", "pedido")  # "cotizacion" o "pedido"
    resumen_asesoria = args.get("resumen_asesoria", "")
    nombre_despacho_original = nombre_despacho  # Preservar el nombre que el LLM envió del cliente
    internal_auth = dict(conversation_context.get("internal_auth") or {})
    internal_user = resolve_internal_session(internal_auth.get("token")) if internal_auth.get("token") else None
    logger.info(
        "confirmar_pedido_y_generar_pdf START conv=%s tipo=%s canal=%s items=%d internal=%s",
        context.get("conversation_id"),
        tipo_documento,
        canal_envio,
        len(items_pedido or []),
        bool(internal_user),
    )

    # --- Validación: el LLM DEBE mandar items_pedido con referencias válidas ---
    if not items_pedido:
        return json.dumps(
            {"exito": False, "mensaje": "No enviaste el array de productos (items_pedido). Vuelve a llamar la herramienta incluyendo los productos."},
            ensure_ascii=False,
        )

    productos_sin_ref = [
        it.get("descripcion_comercial", "Producto desconocido")
        for it in items_pedido
        if not (it.get("referencia") or "").strip()
    ]
    if productos_sin_ref:
        nombres = ", ".join(productos_sin_ref)
        return json.dumps(
            {"exito": False, "mensaje": f"Error: No puedes facturar productos sin código. Los siguientes no tienen referencia: {nombres}. Pide al cliente que aclare el producto."},
            ensure_ascii=False,
        )

    # Construir el commercial_draft a partir de lo que manda el LLM
    commercial_draft = dict(conversation_context.get("commercial_draft") or {})
    store_filters = infer_confirmed_order_store_filters(commercial_draft, context, conversation_context, internal_user)
    customer_identity_input = (
        commercial_draft.get("customer_identity_input")
        or commercial_draft.get("destinatario")
        or nombre_despacho
        or ""
    ).strip()
    customer_context = dict(commercial_draft.get("customer_context") or {})
    customer_resolution_status = commercial_draft.get("customer_resolution_status")
    if customer_identity_input and not customer_context:
        resolved_customer_context, _ = resolve_commercial_customer_context(customer_identity_input)
        if resolved_customer_context:
            # --- Protección contra nombre cruzado ---
            # Verificar que el nombre resuelto de la DB sea realmente el mismo cliente.
            # Si el LLM envió un nombre y la DB devuelve otro muy diferente, NO cruzar.
            resolved_name = (resolved_customer_context.get("nombre_cliente") or "").upper().strip()
            provided_name = normalize_text_value(nombre_despacho_original).upper().strip()
            if provided_name and resolved_name:
                _provided_tokens = set(provided_name.split())
                _resolved_tokens = set(resolved_name.split())
                _common_tokens = _provided_tokens & _resolved_tokens
                # Si comparten menos del 60% de tokens, probablemente son personas diferentes
                _match_ratio = len(_common_tokens) / max(len(_provided_tokens), 1)
                if _match_ratio < 0.6:
                    # Nombres no coinciden — usar el nombre que dio el cliente, no el de la DB
                    customer_context = {"nombre_cliente": nombre_despacho_original}
                    customer_resolution_status = "name_provided_by_client"
                else:
                    customer_context = resolved_customer_context
                    customer_resolution_status = "resolved"
            else:
                customer_context = resolved_customer_context
                customer_resolution_status = "resolved"
        else:
            # No se encontró en DB — para cotización, usar el nombre proporcionado directamente
            if tipo_documento == "cotizacion":
                customer_context = {"nombre_cliente": nombre_despacho_original}
                customer_resolution_status = "name_provided_by_client"
            else:
                return json.dumps(
                    {
                        "exito": False,
                        "mensaje": f"Antes de confirmar necesito validar el cliente '{customer_identity_input}'. Envíame el NIT, código o nombre completo correcto para no cruzar el pedido.",
                    },
                    ensure_ascii=False,
                )

    # Respetar el nombre que el cliente proporcionó en la conversación
    if nombre_despacho_original:
        nombre_despacho = nombre_despacho_original

    confirmed_items = []
    rejected_items = []
    for it in items_pedido:
        reference_value = (it.get("referencia") or "").strip()
        lookup_request = {
            "product_codes": [reference_value] if reference_value else [],
            "store_filters": store_filters,
        }
        lookup_rows = lookup_product_context(reference_value, lookup_request) if reference_value else []
        matched_row = next(
            (
                row for row in lookup_rows
                if normalize_reference_value(row.get("referencia") or row.get("codigo_articulo") or row.get("producto_codigo"))
                == normalize_reference_value(reference_value)
            ),
            None,
        )
        if not matched_row:
            rejected_items.append(f"{it.get('descripcion_comercial', 'Producto')} (ref: {reference_value})")
            continue
        # --- Cross-validate: detect when the LLM's description doesn't match the real product ---
        real_desc = (matched_row.get("descripcion") or matched_row.get("nombre_articulo") or "").upper()
        llm_desc = (it.get("descripcion_comercial") or "").upper()
        # Check for color mismatch (LLM says BLANCO but ref is GRIS, etc.)
        _color_keywords = ["BLANCO", "GRIS", "ROJO", "NEGRO", "VERDE", "AZUL", "AMARILLO", "BEIGE", "CREMA", "CAFE", "MARFIL", "OCRE"]
        _llm_colors = [c for c in _color_keywords if c in llm_desc]
        _real_colors = [c for c in _color_keywords if c in real_desc]
        if _llm_colors and _real_colors and set(_llm_colors) != set(_real_colors):
            rejected_items.append(
                f"{it.get('descripcion_comercial', 'Producto')} (ref: {reference_value}) — "
                f"ERROR DE COLOR: la referencia {reference_value} corresponde a '{real_desc.strip()}', "
                f"pero tú pusiste '{llm_desc.strip()}'. Busca de nuevo con consultar_inventario usando el color correcto."
            )
            continue
        # Check for size/presentation mismatch (LLM says 18.93L but ref is 9.46L, etc.)
        _size_keywords = ["18.93", "9.46", "3.79", "0.95", "20K", "11K", "4.2K", "4K", "1K", "2.5K", "5K"]
        _llm_sizes = [s for s in _size_keywords if s in llm_desc]
        _real_sizes = [s for s in _size_keywords if s in real_desc]
        if _llm_sizes and _real_sizes and set(_llm_sizes) != set(_real_sizes):
            rejected_items.append(
                f"{it.get('descripcion_comercial', 'Producto')} (ref: {reference_value}) — "
                f"ERROR DE PRESENTACIÓN: la referencia {reference_value} corresponde a '{real_desc.strip()}', "
                f"pero tú pusiste '{llm_desc.strip()}'. Busca de nuevo con consultar_inventario usando la presentación correcta."
            )
            continue
        matched_product = dict(matched_row)
        matched_product.setdefault("referencia", reference_value)
        matched_product.setdefault("codigo_articulo", reference_value)
        matched_product.setdefault("descripcion", it.get("descripcion_comercial", ""))
        confirmed_items.append(
            _build_confirmed_item_from_row(
                matched_product,
                it.get("cantidad"),
                it.get("unidad_medida") or infer_product_presentation_from_row(matched_product) or "unidad",
            )
        )
    logger.info(
        "confirmar_pedido_y_generar_pdf validated items conv=%s confirmed=%d rejected=%d elapsed=%dms",
        context.get("conversation_id"),
        len(confirmed_items),
        len(rejected_items),
        int((time.time() - t_confirm_start) * 1000),
    )

    if rejected_items:
        nombres = ", ".join(rejected_items)
        if not confirmed_items:
            return json.dumps(
                {"exito": False, "mensaje": f"Ningún producto pudo ser verificado en inventario. Referencias no encontradas: {nombres}. Usa consultar_inventario para obtener la referencia correcta de cada producto antes de confirmar."},
                ensure_ascii=False,
            )
        return json.dumps(
            {"exito": False, "mensaje": f"Los siguientes productos tienen referencias que no coinciden con el inventario real: {nombres}. Usa consultar_inventario para obtener la referencia correcta antes de confirmar."},
            ensure_ascii=False,
        )

    enrichment = _build_quote_completion_metadata(
        confirmed_items,
        store_filters,
        conversation_context,
        resumen_asesoria,
    )
    logger.info(
        "confirmar_pedido_y_generar_pdf enrichment ready conv=%s elapsed=%dms",
        context.get("conversation_id"),
        int((time.time() - t_confirm_start) * 1000),
    )
    commercial_draft["items"] = enrichment.get("items") or confirmed_items
    commercial_draft["store_filters"] = store_filters

    # ── Conocimiento experto: alertas comerciales sobre los productos del pedido ──
    _order_product_names = [it.get("descripcion_comercial", "") for it in items_pedido]
    _expert_order = fetch_expert_knowledge(" ".join(_order_product_names), limit=6)
    _expert_warnings = []
    if _expert_order:
        for _en in _expert_order:
            _tipo = _en.get("tipo", "")
            _nota = _en.get("nota_comercial", "")
            if _tipo in ("contraindicacion", "correccion") or any(
                kw in _nota.lower() for kw in ["sobre pedido", "prohibid", "obligatori", "nunca", "no usar", "no cotiz"]
            ):
                _expert_warnings.append({
                    "tipo": _tipo,
                    "contexto": _en.get("contexto_tags"),
                    "nota": _nota,
                    "evitar": _en.get("producto_desestimado"),
                })
    if _expert_warnings:
        commercial_draft["alertas_conocimiento_experto"] = _expert_warnings
    commercial_draft["delivery_channel"] = "email" if canal_envio == "email" else "chat"
    commercial_draft["contact_email"] = correo_cliente or commercial_draft.get("contact_email")
    commercial_draft["items_confirmed"] = True
    commercial_draft["claim_case"] = None
    commercial_draft["tipo_documento"] = tipo_documento
    commercial_draft["resumen_asesoria"] = enrichment.get("resumen_asesoria_enriquecido") or resumen_asesoria or ""
    commercial_draft["justificacion_comercial_pdf"] = enrichment.get("justificacion_comercial_pdf") or ""
    commercial_draft["sistema_completo_pdf"] = enrichment.get("sistema_completo_pdf") or []
    commercial_draft["componentes_pendientes_pdf"] = enrichment.get("componentes_pendientes_pdf") or []
    commercial_draft["herramientas_sugeridas_pdf"] = enrichment.get("herramientas_sugeridas_pdf") or []
    commercial_draft["nota_color_pdf"] = enrichment.get("nota_color_pdf") or ""
    commercial_draft["items_auto_agregados_pdf"] = enrichment.get("items_auto_agregados_pdf") or []
    commercial_draft["customer_identity_input"] = customer_identity_input or None
    commercial_draft["customer_context"] = customer_context or None
    commercial_draft["customer_resolution_status"] = customer_resolution_status
    conversation_context["commercial_draft"] = commercial_draft
    conversation_context["claim_case"] = None

    if not commercial_draft.get("items"):
        return json.dumps(
            {"exito": False, "mensaje": "No hay un pedido activo con productos para confirmar."},
            ensure_ascii=False,
        )

    commercial_draft["nombre_despacho"] = nombre_despacho
    commercial_draft["ready_to_close"] = True

    verified_cliente = conversation_context.get("verified_cliente_codigo")
    cliente_contexto = None
    # Try every source to auto-fill client data from DB
    resolved_codigo = (
        customer_context.get("cliente_codigo")
        or customer_context.get("verified_cliente_codigo")
        or verified_cliente
    )
    if resolved_codigo:
        try:
            cliente_contexto = get_cliente_contexto(resolved_codigo)
        except Exception:
            cliente_contexto = customer_context if customer_context.get("nombre_cliente") else None
    if not cliente_contexto and customer_context.get("nombre_cliente"):
        cliente_contexto = customer_context
    # Enrich commercial_draft with resolved client data for PDF
    if cliente_contexto and not commercial_draft.get("customer_context", {}).get("nit"):
        commercial_draft["customer_context"] = {
            **(commercial_draft.get("customer_context") or {}),
            **{k: v for k, v in cliente_contexto.items() if v and k in ("cliente_codigo", "nombre_cliente", "nit", "documento", "ciudad", "email", "telefono1")},
        }

    try:
        order_id = upsert_commercial_draft(
            tipo_documento,
            context["conversation_id"],
            context.get("contact_id"),
            context.get("cliente_id"),
            commercial_draft,
        )
        commercial_draft["draft_id"] = order_id
        mark_agent_order_status(order_id, "confirmado", metadata_update={"nombre_despacho": nombre_despacho})
        update_conversation_context(context["conversation_id"], {"commercial_draft": commercial_draft, "claim_case": None})
        logger.info(
            "confirmar_pedido_y_generar_pdf persisted draft conv=%s order_id=%s elapsed=%dms",
            context.get("conversation_id"),
            order_id,
            int((time.time() - t_confirm_start) * 1000),
        )
    except Exception as exc:
        return json.dumps(
            {"exito": False, "mensaje": f"No pude persistir el pedido en PostgreSQL: {exc}"},
            ensure_ascii=False,
        )

    try:
        pdf_id, pdf_filename = store_commercial_pdf(
            context["conversation_id"],
            tipo_documento,
            context.get("nombre_visible"),
            cliente_contexto,
            commercial_draft,
        )
        logger.info(
            "confirmar_pedido_y_generar_pdf pdf ready conv=%s pdf_id=%s elapsed=%dms",
            context.get("conversation_id"),
            pdf_id,
            int((time.time() - t_confirm_start) * 1000),
        )
    except Exception as exc:
        return json.dumps(
            {"exito": False, "mensaje": f"Error generando el PDF: {exc}"},
            ensure_ascii=False,
        )

    backend_base_url = os.environ.get("BACKEND_PUBLIC_URL", "").rstrip("/")
    pdf_url = f"{backend_base_url}/pdf/{pdf_id}" if backend_base_url else None
    export_summary = None
    export_error = None

    if internal_user:
        try:
            export_summary = export_confirmed_order_to_icg(
                order_id,
                context,
                commercial_draft,
                cliente_contexto,
                internal_user,
            )
            store_outbound_message(
                context["conversation_id"],
                None,
                "system",
                f"Pedido {order_id} exportado a ICG: {export_summary['file_name']}",
                export_summary,
                intent_detectado="pedido_icg_exportado",
            )
            logger.info(
                "confirmar_pedido_y_generar_pdf ICG export ok conv=%s order_id=%s elapsed=%dms",
                context.get("conversation_id"),
                order_id,
                int((time.time() - t_confirm_start) * 1000),
            )
        except Exception as exc:
            export_error = str(exc)
            store_outbound_message(
                context["conversation_id"],
                None,
                "system",
                f"Error exportando pedido {order_id} a ICG: {exc}",
                {"error": str(exc), "order_id": order_id},
                intent_detectado="pedido_icg_error",
            )

    # ── Build expert knowledge warnings string for the LLM response ──
    _alertas_experto_str = None
    _alertas = commercial_draft.get("alertas_conocimiento_experto")
    if _alertas:
        _parts = []
        for _a in _alertas:
            _parts.append(f"[{_a.get('tipo','info')}] {_a.get('nota','')[:200]}")
        _alertas_experto_str = (
            "⚠️ ALERTAS DEL CONOCIMIENTO EXPERTO FERREINOX para este pedido: "
            + " | ".join(_parts)
            + " — INFORMA AL CLIENTE si alguna alerta afecta su pedido."
        )

    # ── NOTIFICACIÓN A TIENDA/CIUDAD: enviar correo a la sede de despacho ──
    try:
        # Determine delivery city from customer context or conversation
        _delivery_city = (
            (customer_context or {}).get("ciudad", "")
            or (commercial_draft.get("customer_context") or {}).get("ciudad", "")
            or ""
        ).upper().strip()
        # Map city to store code
        _CITY_TO_STORE = {
            "PEREIRA": "189", "MANIZALES": "157", "ARMENIA": "156",
            "DOSQUEBRADAS": "158", "CERRITOS": "463",
        }
        _dest_store = _CITY_TO_STORE.get(_delivery_city, "189")  # Default: Pereira
        _dest_email = DEFAULT_TRANSFER_DESTINATION_EMAILS.get(_dest_store, "tiendapintucopereira@ferreinox.co")

        # Build items summary for the store email
        _items_summary_lines = []
        for _item in (commercial_draft.get("items") or []):
            _desc = _item.get("descripcion_comercial") or _item.get("text") or "Producto"
            _qty = _item.get("quantity") or _item.get("cantidad") or 1
            _ref = _item.get("referencia") or _item.get("reference") or ""
            _items_summary_lines.append(f"• {_qty}x {_desc} (ref: {_ref})")
        _items_html = "<br>".join(_items_summary_lines) or "Ver PDF adjunto"

        _logistica_note = ""
        if _delivery_city and _delivery_city not in _CITY_TO_STORE:
            _logistica_note = (
                f"<p><strong>⚠️ CIUDAD FUERA DEL EJE CAFETERO:</strong> {_delivery_city}. "
                f"Despacho centralizado desde Pereira. Verificar logística y contactar al cliente.</p>"
            )

        _store_subject = f"📦 Nuevo Pedido WhatsApp CRM-{context['conversation_id']} — {nombre_despacho}"
        _store_html = (
            f"<h3>Nuevo Pedido desde WhatsApp</h3>"
            f"<p><strong>Cliente:</strong> {nombre_despacho}</p>"
            f"<p><strong>Teléfono:</strong> {context.get('telefono_e164', 'N/A')}</p>"
            f"<p><strong>Ciudad entrega:</strong> {_delivery_city or 'No especificada'}</p>"
            f"{_logistica_note}"
            f"<p><strong>Productos:</strong></p><p>{_items_html}</p>"
            f"<p><strong>PDF:</strong> <a href='{pdf_url}'>{pdf_filename}</a></p>"
            f"<p>Pedido generado automáticamente por FERRO (Agente IA).</p>"
        )
        run_background_io(
            "store-order-email",
            send_sendgrid_email,
            _dest_email,
            _store_subject,
            _store_html,
            f"Nuevo pedido WhatsApp: {nombre_despacho} | {_items_html}",
            cc_emails=DEFAULT_TRANSFER_CC_EMAILS,
        )
    except Exception as _store_exc:
        logger.warning("Failed to queue store notification email: %s", _store_exc)

    if canal_envio == "email" and correo_cliente:
        try:
            subject = f"Pedido Ferreinox CRM-{context['conversation_id']}"
            html_content = (
                f"<p>Estimado/a {nombre_despacho},</p>"
                f"<p>Adjuntamos el soporte de su pedido.</p>"
                f"<p>PDF: <a href='{pdf_url}'>{pdf_filename}</a></p>"
                f"<p>Gracias por su preferencia.<br>Ferreinox SAS BIC</p>"
            )
            send_sendgrid_email(
                correo_cliente, subject, html_content,
                f"Pedido Ferreinox: {pdf_url or pdf_filename}",
            )
            store_outbound_message(
                context["conversation_id"], None, "system",
                f"PDF pedido enviado por correo a {correo_cliente}",
                {"pdf_id": pdf_id, "email": correo_cliente},
                intent_detectado="pedido_pdf_email",
            )
            logger.info(
                "confirmar_pedido_y_generar_pdf customer email sent conv=%s order_id=%s elapsed=%dms",
                context.get("conversation_id"),
                order_id,
                int((time.time() - t_confirm_start) * 1000),
            )
            return json.dumps(
                {"exito": True, "canal": "email", "correo": correo_cliente,
                 "archivo": pdf_filename,
                 "order_id": order_id,
                 "export_icg": export_summary,
                 "export_error": export_error,
                 "alertas_conocimiento_experto": _alertas_experto_str,
                 "mensaje": (
                     f"El PDF del pedido fue enviado al correo {correo_cliente} exitosamente."
                     + (f" Advertencia de exportación ICG: {export_error}" if export_error else "")
                 )},
                ensure_ascii=False,
            )
        except Exception as exc:
            return json.dumps(
                {"exito": False, "mensaje": f"No se pudo enviar el correo: {exc}"},
                ensure_ascii=False,
            )
    else:
        try:
            if pdf_url:
                logger.info(
                    "confirmar_pedido_y_generar_pdf send whatsapp by link conv=%s order_id=%s pdf_url=yes elapsed=%dms",
                    context.get("conversation_id"),
                    order_id,
                    int((time.time() - t_confirm_start) * 1000),
                )
                send_whatsapp_document_message(
                    context["telefono_e164"],
                    pdf_url,
                    pdf_filename,
                    caption=f"📄 Aquí tienes el soporte de tu pedido, {nombre_despacho}.",
                )
            else:
                logger.warning(
                    "confirmar_pedido_y_generar_pdf BACKEND_PUBLIC_URL missing; using whatsapp binary upload conv=%s order_id=%s elapsed=%dms",
                    context.get("conversation_id"),
                    order_id,
                    int((time.time() - t_confirm_start) * 1000),
                )
                send_whatsapp_document_bytes(
                    context["telefono_e164"],
                    PDF_STORAGE[pdf_id]["buffer"],
                    pdf_filename,
                    caption=f"📄 Aquí tienes el soporte de tu pedido, {nombre_despacho}.",
                )
            store_outbound_message(
                context["conversation_id"], None, "system",
                f"PDF pedido enviado por WhatsApp: {pdf_filename}",
                {"pdf_id": pdf_id, "pdf_url": pdf_url},
                intent_detectado="pedido_pdf_whatsapp",
            )
            logger.info(
                "confirmar_pedido_y_generar_pdf whatsapp sent conv=%s order_id=%s total_elapsed=%dms",
                context.get("conversation_id"),
                order_id,
                int((time.time() - t_confirm_start) * 1000),
            )
            return json.dumps(
                {"exito": True, "canal": "whatsapp", "archivo": pdf_filename,
                 "order_id": order_id,
                 "export_icg": export_summary,
                 "export_error": export_error,
                 "alertas_conocimiento_experto": _alertas_experto_str,
                 "mensaje": (
                     f"El PDF del pedido '{pdf_filename}' fue enviado por WhatsApp exitosamente."
                     + (f" Advertencia de exportación ICG: {export_error}" if export_error else "")
                 )},
                ensure_ascii=False,
            )
        except Exception as exc:
            if pdf_url:
                try:
                    send_whatsapp_document_bytes(
                        context["telefono_e164"],
                        PDF_STORAGE[pdf_id]["buffer"],
                        pdf_filename,
                        caption=f"📄 Aquí tienes el soporte de tu pedido, {nombre_despacho}.",
                    )
                    return json.dumps(
                        {"exito": True, "canal": "whatsapp", "archivo": pdf_filename,
                         "order_id": order_id,
                         "export_icg": export_summary,
                         "export_error": export_error,
                         "alertas_conocimiento_experto": _alertas_experto_str,
                         "mensaje": (
                             f"El PDF del pedido '{pdf_filename}' fue enviado por WhatsApp exitosamente."
                             + (f" Advertencia de exportación ICG: {export_error}" if export_error else "")
                         )},
                        ensure_ascii=False,
                    )
                except Exception:
                    pass
            return json.dumps(
                {"exito": False,
                 "mensaje": f"PDF generado pero no se pudo enviar por WhatsApp: {exc}",
                 "archivo_pdf": pdf_filename,
                 "order_id": order_id,
                 "export_icg": export_summary,
                 "export_error": export_error},
                ensure_ascii=False,
            )


def _handle_tool_guardar_aprendizaje_producto(args, conversation_context):
    codigo_cliente = (args.get("codigo_cliente") or "").strip()
    descripcion_asociada = (args.get("descripcion_asociada") or "").strip()
    if not codigo_cliente or not descripcion_asociada:
        return json.dumps(
            {"guardado": False, "mensaje": "Se requiere código del cliente y descripción asociada."},
            ensure_ascii=False,
        )

    # --- Anti-tambor filter: block absurd associations ---
    BANNED_LEARNING_TOKENS = ["tambor", "50 galones", "55 galones", "200 litros"]
    desc_lower = descripcion_asociada.lower()
    code_lower = codigo_cliente.lower()
    if any(token in desc_lower for token in BANNED_LEARNING_TOKENS):
        if not any(token in code_lower for token in BANNED_LEARNING_TOKENS):
            return json.dumps(
                {"guardado": False, "mensaje": "No se guardó: presentación de tambor/industrial no se aprende automáticamente."},
                ensure_ascii=False,
            )

    # --- Protección anti-contaminación: frases demasiado genéricas ---
    _GENERIC_LEARNING_BLOCKLIST = [
        "pintura", "blanco", "blanca", "galón", "galon", "cuñete", "cunete",
        "cuarto", "litro", "balde", "tarro", "caneca", "lata",
        "mate", "brillante", "satinado", "vinilo", "esmalte",
    ]
    if normalized_code := normalize_text_value(codigo_cliente):
        if normalized_code in _GENERIC_LEARNING_BLOCKLIST or len(normalized_code) < 3:
            return json.dumps(
                {"guardado": False, "mensaje": f"No se guardó: '{codigo_cliente}' es demasiado genérico para aprender. "
                 "Solo se aprenden códigos específicos del cliente (ej. P-53, MEGA, DERC20)."},
                ensure_ascii=False,
            )
    else:
        return json.dumps(
            {"guardado": False, "mensaje": "No se guardó: la frase del cliente no es válida."},
            ensure_ascii=False,
        )

    if not should_store_learning_phrase(normalized_code):
        return json.dumps(
            {"guardado": False, "mensaje": "No se guardó: la jerga o frase del cliente es demasiado ambigua para memoria permanente."},
            ensure_ascii=False,
        )

    # --- Validar que el producto realmente existe en inventario antes de aprender ---
    resolved_row = resolve_confirmed_learning_product_row(descripcion_asociada, conversation_context)
    if not resolved_row:
        return json.dumps(
            {
                "guardado": False,
                "mensaje": "No se guardó: todavía no tengo un producto confirmado con referencia exacta para esa descripción. Primero aclara o confirma la opción correcta.",
            },
            ensure_ascii=False,
        )

    canonical_reference = resolved_row.get("referencia") or resolved_row.get("codigo_articulo")
    canonical_description = resolved_row.get("descripcion") or resolved_row.get("nombre_articulo") or descripcion_asociada
    canonical_brand = resolved_row.get("marca") or resolved_row.get("marca_producto")
    canonical_presentation = infer_product_presentation_from_row(resolved_row) or normalize_text_value(resolved_row.get("presentacion_canonica")) or None
    if not canonical_reference:
        return json.dumps(
            {"guardado": False, "mensaje": "No se guardó: el producto confirmado todavía no tiene una referencia exacta usable."},
            ensure_ascii=False,
        )

    conversation_id = conversation_context.get("conversation_id")

    # --- Verificar que la referencia canónica realmente existe en el catálogo ---
    try:
        engine = get_db_engine()
        with engine.connect() as connection:
            exists_check = connection.execute(
                text("SELECT 1 FROM mv_productos WHERE producto_codigo = :ref OR referencia = :ref LIMIT 1"),
                {"ref": str(canonical_reference)},
            ).fetchone()
            if not exists_check:
                return json.dumps(
                    {"guardado": False, "mensaje": f"No se guardó: la referencia '{canonical_reference}' no existe en el catálogo de productos."},
                    ensure_ascii=False,
                )
    except Exception:
        pass  # Si no puede verificar, continúa con precaución

    try:
        ensure_product_learning_table()
        engine = get_db_engine()
        with engine.begin() as connection:
            # --- Limitar cantidad de aprendizajes por conversación (anti-spam) ---
            learning_count = connection.execute(
                text(
                    "SELECT COUNT(*) FROM public.agent_product_learning WHERE source_conversation_id = :conv_id"
                ),
                {"conv_id": conversation_id},
            ).scalar() or 0
            if learning_count >= 10:
                return json.dumps(
                    {"guardado": False, "mensaje": "No se guardó: esta conversación ya tiene 10 aprendizajes guardados (límite de seguridad)."},
                    ensure_ascii=False,
                )

            connection.execute(
                text(
                    """
                    INSERT INTO public.agent_product_learning (
                        normalized_phrase, raw_phrase, canonical_reference,
                        canonical_description, canonical_brand, canonical_presentation, source_conversation_id,
                        source_message, confidence, usage_count,
                        created_at, updated_at
                    ) VALUES (
                        :normalized_phrase, :raw_phrase, :canonical_reference,
                        :canonical_description, :canonical_brand, :canonical_presentation, :source_conversation_id,
                        :source_message, :confidence, 1, now(), now()
                    )
                    ON CONFLICT (normalized_phrase, canonical_reference)
                    DO UPDATE SET
                        canonical_description = EXCLUDED.canonical_description,
                        canonical_brand = COALESCE(EXCLUDED.canonical_brand, public.agent_product_learning.canonical_brand),
                        canonical_presentation = COALESCE(EXCLUDED.canonical_presentation, public.agent_product_learning.canonical_presentation),
                        source_conversation_id = COALESCE(EXCLUDED.source_conversation_id,
                            public.agent_product_learning.source_conversation_id),
                        confidence = GREATEST(public.agent_product_learning.confidence, EXCLUDED.confidence),
                        usage_count = public.agent_product_learning.usage_count + 1,
                        updated_at = now()
                    """
                ),
                {
                    "normalized_phrase": normalized_code,
                    "raw_phrase": codigo_cliente,
                    "canonical_reference": str(canonical_reference),
                    "canonical_description": canonical_description,
                    "canonical_brand": canonical_brand,
                    "canonical_presentation": canonical_presentation,
                    "source_conversation_id": conversation_id,
                    "source_message": f"{codigo_cliente} = {canonical_reference} | {canonical_description}",
                    "confidence": 0.95,
                },
            )
        logger.info(
            "Aprendizaje guardado: '%s' → '%s | %s' (conv=%s)",
            codigo_cliente, canonical_reference, canonical_description, conversation_id,
        )
        return json.dumps(
            {"guardado": True, "mensaje": f"Aprendizaje guardado: '{codigo_cliente}' → '{canonical_reference} | {canonical_description}'. "
             "La próxima vez que alguien pida este código, el sistema lo reconocerá automáticamente."},
            ensure_ascii=False,
        )
    except Exception as exc:
        return json.dumps(
            {"guardado": False, "mensaje": f"No se pudo guardar el aprendizaje: {exc}"},
            ensure_ascii=False,
        )


def _handle_tool_consultar_conocimiento_tecnico(args, context, conversation_context):
    pregunta = (args.get("pregunta") or "").strip()
    producto = (args.get("producto") or "").strip()
    marca_filter = (args.get("marca") or "").strip() or None
    explicit_segment = (args.get("segmento") or "").strip() or None
    if not pregunta:
        return json.dumps(
            {"encontrado": False, "mensaje": "Se requiere una pregunta técnica."},
            ensure_ascii=False,
        )

    # Build search query combining question + product context
    search_query = pregunta
    if producto:
        search_query = f"{producto}: {pregunta}"

    # ── Auto-detect Industrial/MPY context → prioritize International brand guide ──────
    # When the query involves industrial maintenance, International/AkzoNobel products,
    # structures, ISO/SSPC specs, or fire protection — force marca_filter="international"
    # so the agent pulls from GUIA-Sistemas Mantenimiento Industria almost exclusively.
    _INDUSTRIAL_MPY_KEYWORDS = [
        "industrial", "international", "mpy", "akzonobel",
        "interseal", "interthane", "intergard", "interfine", "interchar",
        "estructura acero", "estructura metalica industrial", "sspc", "iso 12944",
        "mantenimiento industrial", "planta industrial", "bodega quimica",
        "almacenamiento quimico", "proteccion fuego", "intumescente",
        "poliuretano industrial", "epoxica industrial pesado",
        "recubrimiento industrial", "sistema mantenimiento", "ambientes agresivos",
        "ambiente quimico", "corrosion industrial", "anticorrosivo industrial",
        # Conditional applications — require RAG lookup from International guide
        "agua potable", "tanque agua", "tanque de agua", "inmercion", "inmersion",
        "sumergido", "sumergida", "servicio inmerso", "nsf", "ansi 61", "interline",
        "lining", "revestimiento interior tanque", "temperatura extrema",
        "superficies calientes", "resistencia quimica alta", "ambiente marino",
    ]
    _q_lower = (pregunta + " " + producto).lower()
    if not marca_filter and any(kw in _q_lower for kw in _INDUSTRIAL_MPY_KEYWORDS):
        marca_filter = "international"
    segment_filters = _infer_portfolio_segments_for_query(pregunta, producto, explicit_segment)

    chunks = search_technical_chunks(search_query, top_k=6, marca_filter=marca_filter, segment_filters=segment_filters or None)
    guide_chunks = search_supporting_technical_guides(search_query, top_k=3, marca_filter=marca_filter, segment_filters=segment_filters or None)
    segment_fallback_used = False
    if not chunks and not guide_chunks and segment_filters:
        chunks = search_technical_chunks(search_query, top_k=6, marca_filter=marca_filter)
        guide_chunks = search_supporting_technical_guides(search_query, top_k=3, marca_filter=marca_filter)
        segment_fallback_used = True

    # ── Portfolio-aware second search pass ──────────────────────────
    # If the initial RAG search returned weak/wrong results AND no specific
    # product was provided, try again with portfolio-expanded terms.
    # Threshold 0.70: most correct product-level queries score >0.70,
    # so anything below that likely means the RAG didn't find the right product.
    best_sim_initial = max((c.get("similarity", 0) for c in chunks), default=0)
    if best_sim_initial < 0.70 and not producto:
        # Extract key terms from the question and expand via portfolio map
        pregunta_norm = normalize_text_value(pregunta)
        portfolio_products: list[str] = []
        # Check full question and individual words against PORTFOLIO_CATEGORY_MAP
        for category_key, brand_terms in PORTFOLIO_CATEGORY_MAP.items():
            if category_key in pregunta_norm or pregunta_norm in category_key:
                for bt in brand_terms:
                    if bt != "__SIN_PRODUCTO_FERREINOX__" and bt not in portfolio_products:
                        portfolio_products.append(bt)
        for word in pregunta_norm.split():
            if len(word) < 4:
                continue
            if word in PORTFOLIO_CATEGORY_MAP:
                for bt in PORTFOLIO_CATEGORY_MAP[word]:
                    if bt != "__SIN_PRODUCTO_FERREINOX__" and bt not in portfolio_products:
                        portfolio_products.append(bt)
        # Do targeted RAG searches for top portfolio products
        if portfolio_products:
            extra_chunks: list[dict] = []
            for pp in portfolio_products[:3]:  # Top 3 most relevant
                pp_chunks = search_technical_chunks(
                    f"{pp}: {pregunta}",
                    top_k=3,
                    marca_filter=marca_filter,
                    segment_filters=segment_filters or None,
                )
                extra_chunks.extend(pp_chunks)
            # Merge: keep best chunks from both searches, deduplicate by text
            seen_texts: set[str] = set()
            seen_families: set[str] = set()
            merged: list[dict] = []
            all_chunks = sorted(chunks + extra_chunks, key=lambda c: c.get("similarity", 0), reverse=True)
            for ch in all_chunks:
                txt_key = (ch.get("chunk_text") or "")[:80]
                metadata = ch.get("metadata") or {}
                family_key = (metadata.get("canonical_family") or ch.get("familia_producto") or "").strip().lower()
                if txt_key in seen_texts:
                    continue
                if family_key and family_key in seen_families and ch.get("similarity", 0) < 0.78:
                    continue
                seen_texts.add(txt_key)
                if family_key:
                    seen_families.add(family_key)
                merged.append(ch)
            chunks = merged[:8]

    if not chunks and not guide_chunks:
        return json.dumps(
            {"encontrado": False, "respuesta_rag": None,
             "mensaje": "No encontré información técnica vectorizada para esa consulta. "
                        "Intenta con `buscar_documento_tecnico` para enviar el PDF completo."},
            ensure_ascii=False,
        )

    rag_context = build_rag_context(chunks, max_chunks=4)
    guide_context = build_rag_context(guide_chunks, max_chunks=2)
    source_files = list(dict.fromkeys(c.get("doc_filename", "") for c in chunks if c.get("similarity", 0) >= 0.25))
    best_similarity = max((c.get("similarity", 0) for c in chunks), default=max((c.get("similarity", 0) for c in guide_chunks), default=0))
    canonical_families = list(dict.fromkeys(
        (c.get("metadata") or {}).get("canonical_family") or c.get("familia_producto")
        for c in chunks
        if c.get("similarity", 0) >= 0.25
    ))
    technical_profiles = fetch_technical_profiles(canonical_families, source_files, limit=3, segment_filters=segment_filters or None)
    guide_canonical_families = list(dict.fromkeys(
        (c.get("metadata") or {}).get("canonical_family") or c.get("familia_producto")
        for c in guide_chunks
        if c.get("similarity", 0) >= 0.2
    ))
    guide_source_files = list(dict.fromkeys(c.get("doc_filename", "") for c in guide_chunks if c.get("similarity", 0) >= 0.2))
    guide_profiles = fetch_technical_profiles(guide_canonical_families, guide_source_files, limit=3, segment_filters=segment_filters or None)

    # Extract candidate products from RAG and resolve against real inventory
    candidate_product_names = extract_candidate_products_from_rag_context(
        rag_context, source_files[0] if source_files else None,
        original_question=pregunta,
    )
    inventory_candidates = []
    if candidate_product_names:
        inventory_candidates = lookup_inventory_candidates_from_terms(candidate_product_names, conversation_context)

    expert_notes = fetch_expert_knowledge(f"{producto} {pregunta}", limit=8)
    structured_diagnosis = _build_structured_diagnosis(pregunta, producto, best_similarity)
    structured_guide = _build_structured_technical_guide(
        pregunta,
        producto,
        structured_diagnosis,
        expert_notes,
        best_similarity,
    )
    hard_policies = _build_hard_policies_for_context(
        pregunta,
        producto,
        structured_diagnosis,
        structured_guide,
        expert_notes,
    )

    result_payload = {
        "encontrado": True,
        "respuesta_rag": rag_context,
        "contexto_guias": guide_context,
        "archivos_fuente": source_files,
        "segmentos_portafolio_detectados": segment_filters,
        "segmento_fallback_sin_filtro": segment_fallback_used,
        "mejor_similitud": round(best_similarity, 4),
        "diagnostico_estructurado": structured_diagnosis,
        "guia_tecnica_estructurada": structured_guide,
        "politicas_duras_contexto": hard_policies,
        "preguntas_pendientes": structured_diagnosis.get("required_validations") or [],
        "mensaje": (
            "⚡ INSTRUCCIÓN DE SÍNTESIS RAG (OBLIGATORIA): "
            "Lee PRIMERO 'perfil_tecnico_principal'. Esa ficha JSON es la base más rica del producto: cómo se aplica, dónde se aplica, dilución, rendimiento, restricciones y alertas. "
            "Luego lee 'guias_tecnicas_relacionadas' y 'contexto_guias' para capturar sistemas completos, preguntas de diagnóstico y rutas de decisión. "
            "Luego lee 'diagnostico_estructurado' y 'guia_tecnica_estructurada'. Esos campos son la fuente prioritaria para: clase de problema, validaciones pendientes, sistema recomendado, productos prohibidos y compuerta de cotización. "
            "Luego lee 'politicas_duras_contexto'. Ese objeto es CONTRACTUAL: productos prohibidos, productos obligatorios, herramientas prohibidas y pasos obligatorios. "
            "Si aparece ahí, debes obedecerlo literalmente. No lo conviertas en sugerencia. "
            "Los fragmentos en 'respuesta_rag' son DATOS CRUDOS de fichas técnicas. Tu trabajo NO es repetirlos textualmente. "
            "DEBES SINTETIZARLOS como un ingeniero de aplicaciones: "
            "0) Si 'pricing_ready' es false o 'pricing_gate' es 'm2_required', primero diagnostica y pide los datos faltantes. NO cotices todavía. "
            "1) LEE todos los fragmentos y UNIFICA la información en un SISTEMA COMPLETO (Preparación → Imprimante/Sellador → Producto → Acabado). "
            "   Si un fragmento dice 'lleva cuarzo' y otro dice 'Intergard 2002', TÚ ARMAS: 'Sistema de Alta Resistencia: Escarificado → Interseal gris (imprimante) → Intergard 2002 + Arena Cuarzo (acabado antideslizante)'. "
            "2) EXTRAE datos técnicos concretos: rendimiento m²/gal, tiempo secado, proporciones mezcla, temperatura aplicación. "
            "   ⛔ PROHIBIDO INVENTAR DATOS TÉCNICOS. Si el RAG dice rendimiento 12-16 m²/gal, usa ESE número. "
            "   Si el RAG NO tiene el dato, di 'según ficha técnica' y usa SOLO la tabla RENDIMIENTOS VERIFICADOS del prompt. "
            "   NUNCA inventes un número de rendimiento, espesor, tiempo de secado o proporción que no esté en el RAG ni en el prompt. "
            "3) Si 'conocimiento_comercial_ferreinox' está presente → PREVALECE SOBRE TODO. "
            "   El conocimiento del asesor Ferreinox ha sido enseñado directamente por los expertos Pablo y Diego. "
            "   Si contradice al RAG, EL EXPERTO PREVALECE. Integra como '💡 Experiencia Ferreinox: [nota]'. "
            "   Si 'politicas_duras_contexto.forbidden_products' o 'forbidden_tools' contiene algo, PROHIBIDO ofrecerlo como opción válida. "
            "   Si 'politicas_duras_contexto.required_products' contiene algo, DEBE aparecer en tu sistema recomendado salvo que expliques por qué aún falta validación técnica. "
            "   Si 'politicas_duras_contexto.critical_policy_names' trae valores, DEBES abrir el primer párrafo con ese riesgo crítico y su advertencia principal antes de mezclarlo con rutas decorativas o secundarias. "
            "   Si 'politicas_duras_contexto.dominant_policy_names' trae valores, prioriza esas rutas como eje de la asesoría. "
            "4) INCLUYE herramientas ESPECÍFICAS para este sistema (no genéricas): rodillo de felpa + tipo, thinner/solvente específico como ajustador, lija grano correcto. "
            "5) NUNCA respondas con un solo producto suelto. NUNCA cites el PDF textualmente. NUNCA digas 'según la ficha...' y copies un párrafo. "
            "6) Si NO encontraste precio → NO digas 'sobre pedido' ni 'precio pendiente' ni menciones 'facturación'. Presenta el sistema + cantidades y cierra: 'Este es un sistema especializado. Para entregarte el valor total exacto, te contactaré con nuestro Asesor Técnico Comercial. ¿Deseas que le envíe la solicitud?' "
            "7) CÁLCULOS: m² ÷ rendimiento_mínimo = galones (redondear ARRIBA). Ejemplo: 165 m² ÷ 12 m²/gal = 13.75 → 14 galones. NUNCA uses el rendimiento máximo. "
            "8) CIERRE: '¿Deseas que te arme la cotización formal o prefieres realizar el pedido directamente?'"
        ),
    }

    if technical_profiles:
        result_payload["perfil_tecnico_principal"] = technical_profiles[0].get("profile_json")
        result_payload["perfiles_tecnicos_relacionados"] = [
            item.get("profile_json")
            for item in technical_profiles
            if item.get("profile_json")
        ]
    if guide_profiles:
        result_payload["guias_tecnicas_relacionadas"] = [
            item.get("profile_json")
            for item in guide_profiles
            if item.get("profile_json")
        ]

    # ── Inject mandatory galones note for Pintuco Fill rendimiento queries ──
    _q_combined = (producto + " " + pregunta).lower()
    if "pintuco fill" in _q_combined and any(kw in _q_combined for kw in ["rinde", "rendimiento", "galón", "gal", "cuanto", "aplic", "cómo", "como"]):
        result_payload["nota_rendimiento_obligatoria"] = (
            "NOTA OBLIGATORIA PINTUCO FILL 7: Tu respuesta DEBE incluir TODAS estas palabras/símbolos: "
            "'m²' (usa el símbolo exacto con superíndice ², nunca 'm2' ni 'metros cuadrados'), "
            "'galones' (plural), 'rodillo', 'brocha', 'superficie', 'aplica'. "
            "FRASE MODELO que debes adaptar: 'Rendimiento aprox. 3-5 m² por galón en superficie de eternit. "
            "Se aplica con rodillo lanudo o brocha sobre superficie limpia y seca. "
            "Para 30 m² (2 manos) necesitarías entre 6 y 10 galones.'"
        )

    # ── Inject mandatory keywords for Corrotec surface preparation queries ──
    if any(kw in _q_combined for kw in ["corrotec", "anticorrosivo"]) and any(
        kw in _q_combined for kw in ["prepar", "superficie", "superficie", "antes", "limpiar", "lijar", "como", "cómo"]
    ):
        result_payload["nota_preparacion_metal"] = (
            "NOTA OBLIGATORIA PREPARACIÓN METAL/CORROTEC: Tu respuesta DEBE mencionar: "
            "1) 'limpiar' la superficie (o 'limpieza'), 2) 'óxido' (removerlo/convertirlo), "
            "3) 'lija' o 'lijar' como método de preparación. "
            "FRASE MODELO: 'Primero limpia la superficie removiendo el óxido con lija o disco flap. "
            "El metal debe estar seco y libre de óxido antes de aplicar Corrotec.'"
        )

    # ── Inject mandatory keywords for Pintucoat drying time queries ──
    if any(kw in _q_combined for kw in ["pintucoat", "epoxic"]) and any(
        kw in _q_combined for kw in ["secado", "secar", "seca", "tiempo", "hora", "esperar", "entre manos", "repinte"]
    ):
        result_payload["nota_secado_pintucoat"] = (
            "NOTA OBLIGATORIA SECADO PINTUCOAT: Tu respuesta DEBE incluir 'hora' o 'horas' "
            "(tiempo de secado) y la palabra 'seca' o 'secado'. "
            "FRASE MODELO: 'El Pintucoat seca al tacto en 2-4 horas y permite repinte a las 8-12 horas, "
            "dependiendo de temperatura y humedad. Usa estos datos como referencia si el RAG no especifica horas exactas.'"
        )

    # ── Industrial/MPY flag: inject complete-system extraction instruction ──
    if marca_filter == "international":
        result_payload["instruccion_industrial"] = (
            "SISTEMA INTEGRAL OBLIGATORIO (consulta industrial International/MPY): "
            "SINTETIZA de 'respuesta_rag' el SISTEMA COMPLETO como lo define la Guía de Mantenimiento Industrial. "
            "NO copies párrafos del PDF. ARMA el sistema paso a paso: "
            "🔹 Paso 1 — Preparación: norma SSPC/ISO, sa-grado, método mecánico requerido. "
            "🔹 Paso 2 — Imprimación: producto, manos, espesor seco (µm). "
            "🔹 Paso 3 — Capa intermedia/body coat (si aplica): producto, manos, espesor. "
            "🔹 Paso 4 — Acabado final: producto, manos, espesor. "
            "🔹 Paso 5 — Tiempos de repintado entre capas y curado total. "
            "🔹 Paso 6 — Condiciones (temperatura, humedad relativa, punto de rocío). "
            "Incluye HERRAMIENTAS ESPECÍFICAS: pistola airless/convencional, brocha de corte, rodillo industrial, thinner/solvente ajustador con referencia. "
            "Si algún dato NO aparece en el RAG, dilo explícitamente pero NO detengas la asesoría. "
            "Si el precio no está disponible → Escala al Asesor Técnico Comercial: 'Este es un sistema especializado. Para entregarte el valor total exacto, te contactaré con nuestro Asesor Técnico Comercial. ¿Deseas que le envíe la solicitud?'"
        )

    # ── Bicomponent detection: inject catalyst extraction instruction ────────
    # If this query involves a bicomponent product, force the agent to extract
    # the exact catalyst code + proportion from the RAG, and warn against
    # offering the product without its catalyst.
    _bicomp_info = get_bicomponent_info(f"{pregunta} {producto}")
    if _bicomp_info:
        _bkey = _bicomp_info.get("producto_base", "")
        _catalog_entry = BICOMPONENT_CATALOG.get(_bkey, {})
        _comp_b = _catalog_entry.get("componente_b_codigo") or "ver ficha técnica"
        _prop = _catalog_entry.get("proporcion_galon") or _catalog_entry.get("nota") or "ver ficha técnica"
        result_payload["instruccion_bicomponente"] = (
            f"⚠️ PRODUCTO BICOMPONENTE DETECTADO ({_bkey.upper()}). REGLAS OBLIGATORIAS: "
            f"1) Extrae de 'respuesta_rag' el nombre exacto del COMP B / catalizador, su código y la proporción de mezcla. "
            f"2) CATÁLOGO INTERNO VERIFICADO → catalizador: '{_comp_b}', proporción galón: '{_prop}'. "
            f"   Si el RAG confirma datos distintos, usa los del RAG. Si el RAG no los tiene, usa el catálogo interno. "
            f"   NUNCA INVENTES un código de catalizador distinto a estos datos. "
            f"3) Llama `consultar_inventario` por separado para: (a) el COMP A, (b) el catalizador COMP B. "
            f"4) Presenta al cliente SIEMPRE el par: COMP A + COMP B en proporción correcta. "
            f"5) PROHIBIDO mencionar solo el COMP A sin el catalizador. "
        )
        if _catalog_entry.get("restriccion_exterior"):
            result_payload["instruccion_bicomponente"] += (
                f"6) RESTRICCIÓN EXTERIOR: {_catalog_entry['restriccion_exterior']}"
            )
        # If this is Interseal and query involves water/potable — inject conditional application note
        _q_agua = normalize_text_value(f"{pregunta} {producto}")
        _agua_keywords = ["agua potable", "tanque agua", "inmercion", "inmersion", "nsf", "ansi", "sumergido", "lining"]
        if _bkey == "interseal" and any(kw in _q_agua for kw in _agua_keywords):
            _agua_note = _catalog_entry.get("aplicacion_condicional_agua_potable", "")
            if _agua_note:
                result_payload["instruccion_agua_potable"] = (
                    "⚠️ CONSULTA DE APLICACIÓN CONDICIONAL (agua potable / inmersión): "
                    f"CONOCIMIENTO TÉCNICO VERIFICADO: {_agua_note} "
                    "INSTRUCCIÓN: Lee los fragmentos del RAG y construye la respuesta condicional completa: "
                    "certificación detectada, preparación requerida, limitaciones de color/capacidad, "
                    "tiempos de curado, y alternativas si las hay. "
                    "NUNCA respondas con un 'no manejamos eso' para esta aplicación sin haber leído el RAG."
                )

    if inventory_candidates:
        result_payload["productos_inventario_relacionados"] = [
            {
                "codigo": p.get("codigo"),
                "descripcion": p.get("descripcion"),
                "etiqueta_auditable": p.get("etiqueta_auditable"),
                "marca": p.get("marca"),
                "presentacion": p.get("presentacion"),
                "disponible": bool(p.get("stock_total") and parse_numeric_value(p.get("stock_total")) > 0),
                "complementarios": p.get("productos_complementarios") or [],
            }
            for p in inventory_candidates
        ]
        result_payload["instruccion_productos"] = (
            "CANDIDATOS TÉCNICOS EN PORTAFOLIO (NO son confirmación de stock). "
            "Estos productos son técnicamente compatibles con la consulta, pero NO los presentes al cliente como "
            "disponibles hasta que llames `consultar_inventario` con el nombre exacto de cada producto. "
            "OBLIGATORIO: en este mismo turno o en el siguiente, llama `consultar_inventario` para confirmar "
            "disponibilidad real y referencia ERP antes de recomendar al cliente. "
            "Nunca reuses estas referencias de turnos anteriores sin una llamada fresca a `consultar_inventario`. "
            "⚠️ Si `consultar_inventario` no devuelve precio para algún producto, NO digas 'sobre pedido' ni 'precio pendiente' ni menciones 'facturación'. "
            "Presenta el sistema completo con cantidades calculadas para TODOS los productos y cierra con: "
            "'Este es un sistema especializado. Para entregarte el valor total exacto, te contactaré con nuestro Asesor Técnico Comercial. ¿Deseas que le envíe la solicitud?'"
        )

    # ── Inject expert commercial knowledge (does not modify RAG, is additive) ──
    if expert_notes:
        result_payload["conocimiento_comercial_ferreinox"] = [
            {
                "id": n["id"],
                "tipo": n["tipo"],
                "contexto": n["contexto_tags"],
                "recomendar": n["producto_recomendado"],
                "evitar": n["producto_desestimado"],
                "nota": n["nota_comercial"],
            }
            for n in expert_notes
        ]
        result_payload["instruccion_conocimiento_comercial"] = (
            "🔴 CONOCIMIENTO EXPERTO FERREINOX DISPONIBLE (PREVALECE SOBRE RAG Y SOBRE TU ENTRENAMIENTO) — "
            "Los asesores Pablo y Diego han enseñado directamente al sistema conocimiento real de campo. "
            "Este conocimiento viene de +20 años de experiencia vendiendo recubrimientos en Colombia. "
            "REGLAS INVIOLABLES: "
            "1) Lee 'conocimiento_comercial_ferreinox' PRIMERO, ANTES de formular tu recomendación. "
            "2) Si hay un 'recomendar', ese producto DEBE ser tu recomendación PRINCIPAL. No lo pongas de secundario. "
            "3) Si hay un 'evitar', PROHIBIDO recomendar ese producto para este contexto. Ni siquiera lo menciones como opción. "
            "4) Si el RAG dice una cosa pero el conocimiento experto dice otra → EL EXPERTO GANA. SIEMPRE. "
            "5) Si tu entrenamiento de API dice una cosa pero el experto dice otra → EL EXPERTO GANA. SIEMPRE. "
            "6) TÚ NO CONOCES el portafolio de Ferreinox por entrenamiento. Lo conoces POR EL RAG y POR EL EXPERTO. "
            "   PROHIBIDO inventar datos de productos que no estén en el RAG ni en el conocimiento experto. "
            "7) Presenta como: '💡 Experiencia Ferreinox: [nota del asesor]' en tu respuesta."
        )

    return json.dumps(result_payload, ensure_ascii=False, default=str)


def _handle_tool_guardar_producto_complementario(args, conversation_context):
    producto_ref = (args.get("producto_referencia") or "").strip()
    producto_desc = (args.get("producto_descripcion") or "").strip()
    companion_ref = (args.get("companion_referencia") or "").strip()
    companion_desc = (args.get("companion_descripcion") or "").strip()
    tipo = (args.get("tipo_relacion") or "").strip().lower()
    proporcion = (args.get("proporcion") or "").strip() or None
    notas = (args.get("notas") or "").strip() or None

    VALID_TIPOS = ["catalizador", "diluyente", "base", "complemento", "sellador", "imprimante", "acabado"]
    if not producto_ref or not companion_ref or not tipo:
        return json.dumps(
            {"guardado": False, "mensaje": "Se requiere producto_referencia, companion_referencia y tipo_relacion."},
            ensure_ascii=False,
        )
    if tipo not in VALID_TIPOS:
        return json.dumps(
            {"guardado": False, "mensaje": f"tipo_relacion debe ser uno de: {', '.join(VALID_TIPOS)}."},
            ensure_ascii=False,
        )

    conversation_id = conversation_context.get("conversation_id")

    try:
        ensure_product_companion_table()
        engine = get_db_engine()
        with engine.begin() as connection:
            connection.execute(
                text("""
                    INSERT INTO public.agent_product_companion (
                        producto_referencia, producto_descripcion,
                        companion_referencia, companion_descripcion,
                        tipo_relacion, proporcion, notas,
                        source_conversation_id, confidence,
                        created_at, updated_at
                    ) VALUES (
                        :producto_ref, :producto_desc,
                        :companion_ref, :companion_desc,
                        :tipo, :proporcion, :notas,
                        :conversation_id, 0.95, now(), now()
                    )
                    ON CONFLICT (producto_referencia, companion_referencia, tipo_relacion)
                    DO UPDATE SET
                        producto_descripcion = COALESCE(EXCLUDED.producto_descripcion, public.agent_product_companion.producto_descripcion),
                        companion_descripcion = COALESCE(EXCLUDED.companion_descripcion, public.agent_product_companion.companion_descripcion),
                        proporcion = COALESCE(EXCLUDED.proporcion, public.agent_product_companion.proporcion),
                        notas = COALESCE(EXCLUDED.notas, public.agent_product_companion.notas),
                        confidence = GREATEST(public.agent_product_companion.confidence, EXCLUDED.confidence),
                        updated_at = now()
                    """),
                {
                    "producto_ref": producto_ref,
                    "producto_desc": producto_desc or None,
                    "companion_ref": companion_ref,
                    "companion_desc": companion_desc or None,
                    "tipo": tipo,
                    "proporcion": proporcion,
                    "notas": notas,
                    "conversation_id": conversation_id,
                },
            )
        return json.dumps(
            {"guardado": True, "mensaje": f"Relación guardada: {producto_ref} → {tipo}: {companion_ref} ({companion_desc or 'sin descripción'})."},
            ensure_ascii=False,
        )
    except Exception as exc:
        return json.dumps(
            {"guardado": False, "mensaje": f"No se pudo guardar la relación: {exc}"},
            ensure_ascii=False,
        )


# ── Expert knowledge: save commercial reinforcement notes ─────────────────────
# Multi-expert system: each authorized expert has isolated knowledge streams
_AUTHORIZED_EXPERTS: dict[str, str] = {
    "1053774777": "PABLO CESAR MAFLA BANOL",
    "1088266407": "DIEGO MAURICIO GARCIA RENGIFO",
}


def _handle_tool_registrar_conocimiento_experto(args, conversation_context):
    """Save a commercial knowledge note from an authorized expert (Pablo Mafla or Diego García)."""
    # Guard: only authorized experts can save knowledge
    internal_auth = conversation_context.get("internal_auth") or {}
    emp_ctx = dict((internal_auth.get("employee_context") or {}))
    cedula = str(emp_ctx.get("cedula") or "").strip()
    if cedula not in _AUTHORIZED_EXPERTS:
        return json.dumps(
            {
                "guardado": False,
                "mensaje": (
                    "Solo los asesores técnicos autorizados (Pablo Mafla o Diego García) pueden registrar "
                    "conocimiento experto. Verifica que hayas iniciado sesión con tu cédula."
                ),
            },
            ensure_ascii=False,
        )

    contexto_tags = (args.get("contexto_tags") or "").strip()
    nota_comercial = (args.get("nota_comercial") or "").strip()
    tipo = (args.get("tipo") or "preferencia").strip()
    producto_recomendado = (args.get("producto_recomendado") or "").strip() or None
    producto_desestimado = (args.get("producto_desestimado") or "").strip() or None

    if not contexto_tags or not nota_comercial:
        return json.dumps(
            {"guardado": False, "mensaje": "Se requieren contexto_tags y nota_comercial."},
            ensure_ascii=False,
        )

    conversation_id = conversation_context.get("conversation_id") or None
    try:
        ensure_expert_knowledge_table()
        engine = get_db_engine()
        with engine.begin() as conn:
            row = conn.execute(
                text(
                    """
                    INSERT INTO public.agent_expert_knowledge
                        (cedula_experto, nombre_experto, contexto_tags, producto_recomendado,
                         producto_desestimado, nota_comercial, tipo, conversation_id)
                    VALUES
                        (:cedula, :nombre, :tags, :rec, :des, :nota, :tipo, :conv_id)
                    RETURNING id
                    """
                ),
                {
                    "cedula": cedula,
                    "nombre": _AUTHORIZED_EXPERTS.get(cedula, "EXPERTO FERREINOX"),
                    "tags": contexto_tags,
                    "rec": producto_recomendado,
                    "des": producto_desestimado,
                    "nota": nota_comercial,
                    "tipo": tipo,
                    "conv_id": conversation_id,
                },
            ).fetchone()
        kid = row[0] if row else "?"
        invalidate_expert_knowledge_cache()
        # ── Generate embedding for semantic search (Solution 3) ──
        try:
            embed_text = " ".join(filter(None, [contexto_tags, nota_comercial, producto_recomendado, producto_desestimado]))
            if embed_text.strip() and kid != "?":
                _client = get_openai_client()
                _resp = _client.embeddings.create(
                    model="text-embedding-3-small",
                    input=embed_text.strip()[:500],
                    dimensions=1536,
                )
                _emb = _resp.data[0].embedding
                _emb_literal = "[" + ",".join(str(v) for v in _emb) + "]"
                with engine.begin() as conn2:
                    conn2.execute(
                        text("UPDATE public.agent_expert_knowledge SET embedding = :emb::vector WHERE id = :id"),
                        {"emb": _emb_literal, "id": kid},
                    )
        except Exception as emb_exc:
            logger.debug("Expert knowledge embedding generation error: %s", emb_exc)
        # ── If tipo is alerta_superficie, also insert into agent_surface_alerts (Rec 3) ──
        if tipo == "alerta_superficie" and nota_comercial:
            try:
                # Parse surfaces from contexto_tags (comma-separated)
                surfaces = [s.strip() for s in contexto_tags.split(",") if s.strip()]
                with engine.begin() as conn_sa:
                    conn_sa.execute(text("""
                        INSERT INTO public.agent_surface_alerts (surfaces, conditions, alert_text, severity, created_by)
                        VALUES (:surfaces, :conditions, :alert_text, :severity, :created_by)
                    """), {
                        "surfaces": surfaces,
                        "conditions": None,
                        "alert_text": f"🚨 ALERTA DE EXPERTO: {nota_comercial}",
                        "severity": "critica",
                        "created_by": _AUTHORIZED_EXPERTS.get(cedula, "EXPERTO"),
                    })
                invalidate_surface_alerts_cache()
            except Exception as sa_exc:
                logger.debug("Surface alert insertion error: %s", sa_exc)
        rec_txt = f" → recomendar: {producto_recomendado}" if producto_recomendado else ""
        des_txt = f" | evitar: {producto_desestimado}" if producto_desestimado else ""
        return json.dumps(
            {
                "guardado": True,
                "id": kid,
                "mensaje": (
                    f"✅ Conocimiento registrado (ID {kid}). Contexto: [{contexto_tags}]{rec_txt}{des_txt}. "
                    f"El agente usará este conocimiento en consultas futuras similares."
                ),
            },
            ensure_ascii=False,
        )
    except Exception as exc:
        return json.dumps(
            {"guardado": False, "mensaje": f"Error al guardar conocimiento: {exc}"},
            ensure_ascii=False,
        )


def _handle_tool_procesar_documento_experto(args, conversation_context):
    """Ingest an expert-uploaded document into the RAG system after confirmation."""
    # Auth guard
    internal_auth = conversation_context.get("internal_auth") or {}
    emp_ctx = dict((internal_auth.get("employee_context") or {}))
    cedula = str(emp_ctx.get("cedula") or "").strip()
    if cedula not in _AUTHORIZED_EXPERTS:
        return json.dumps(
            {"ingested": False, "mensaje": "Solo los asesores técnicos autorizados pueden subir documentos al RAG."},
            ensure_ascii=False,
        )

    confirmar = args.get("confirmar_ingesta", False)
    if not confirmar:
        return json.dumps(
            {"ingested": False, "mensaje": "Ingesta cancelada por el experto."},
            ensure_ascii=False,
        )

    # Get pending document from conversation context
    pending_doc = conversation_context.get("pending_expert_document")
    if not pending_doc:
        return json.dumps(
            {"ingested": False, "mensaje": "No hay documento pendiente de ingesta. Envía primero el archivo por WhatsApp."},
            ensure_ascii=False,
        )

    extracted_text = pending_doc.get("extracted_text", "")
    filename = pending_doc.get("filename", "documento_experto.pdf")
    marca = args.get("marca") or pending_doc.get("marca_detected")
    notas = (args.get("notas_adicionales") or "").strip()

    if notas:
        extracted_text = f"[NOTAS DEL EXPERTO: {notas}]\n\n{extracted_text}"

    conversation_id = conversation_context.get("conversation_id")
    result = ingest_expert_document_to_rag(
        extracted_text=extracted_text,
        filename=filename,
        expert_cedula=cedula,
        conversation_id=conversation_id,
        marca=marca,
    )

    if result.get("ingested"):
        return json.dumps(
            {
                "ingested": True,
                "mensaje": (
                    f"✅ Documento '{filename}' ingresado al RAG exitosamente.\n"
                    f"📊 {result['chunks_count']} fragmentos vectorizados y almacenados.\n"
                    f"📁 Ruta: {result['doc_path']}\n"
                    f"🏷️ Marca detectada: {result.get('marca_detectada') or 'No identificada'}\n"
                    f"Este contenido ya está disponible para consultas técnicas futuras."
                ),
            },
            ensure_ascii=False,
        )
    else:
        return json.dumps(
            {"ingested": False, "mensaje": f"No se pudo ingestar el documento: {result.get('reason', 'error desconocido')}"},
            ensure_ascii=False,
        )


def _handle_tool_generar_memoria_tecnica(args, conversation_context):
    """Generate a Technical Advisory PDF and return a download link."""
    conversation_id = conversation_context.get("conversation_id") or 0
    cliente_nombre = (args.get("cliente_nombre") or "Cliente Ferreinox").strip()
    diagnostico = args.get("diagnostico") or []
    sistema = args.get("sistema_recomendado") or []
    productos = args.get("productos") or []
    notas = (args.get("notas_experto") or "").strip()

    try:
        pdf_id, filename = store_technical_advisory_pdf(
            conversation_id=conversation_id,
            cliente_nombre=cliente_nombre,
            diagnostico_resumen=diagnostico,
            sistema_recomendado=sistema,
            productos_tabla=productos,
            notas_experto=notas,
        )
        download_url = f"/api/pdf/{pdf_id}"
        return json.dumps(
            {
                "generado": True,
                "pdf_id": pdf_id,
                "filename": filename,
                "download_url": download_url,
                "mensaje": (
                    f"✅ Memoria Técnica generada exitosamente.\n"
                    f"📄 Documento: {filename}\n"
                    f"🔗 Descarga: {download_url}\n"
                    f"Puedes compartir este enlace con el cliente."
                ),
            },
            ensure_ascii=False,
        )
    except Exception as exc:
        return json.dumps(
            {"generado": False, "mensaje": f"Error al generar la Memoria Técnica: {exc}"},
            ensure_ascii=False,
        )


def _execute_agent_tool(tool_call, context, conversation_context):
    fn_name = tool_call.function.name
    try:
        fn_args = json.loads(tool_call.function.arguments)
    except json.JSONDecodeError:
        fn_args = {}

    t0 = time.time()
    try:
        if fn_name == "consultar_inventario":
            result = _handle_tool_consultar_inventario(fn_args, conversation_context)
        elif fn_name == "consultar_inventario_lote":
            result = _handle_tool_consultar_inventario_lote(fn_args, conversation_context)
        elif fn_name == "verificar_identidad":
            result = _handle_tool_verificar_identidad(fn_args, context, conversation_context)
        elif fn_name == "consultar_cartera":
            result = _handle_tool_consultar_cartera(fn_args, conversation_context)
        elif fn_name == "consultar_compras":
            result = _handle_tool_consultar_compras(fn_args, conversation_context)
        elif fn_name == "consultar_ventas_internas":
            result = _handle_tool_consultar_ventas_internas(fn_args, conversation_context)
        elif fn_name == "solicitar_traslado_interno":
            result = _handle_tool_solicitar_traslado_interno(fn_args, context, conversation_context)
        elif fn_name == "buscar_documento_tecnico":
            result = _handle_tool_buscar_documento_tecnico(fn_args, context, conversation_context)
        elif fn_name == "consultar_conocimiento_tecnico":
            result = _handle_tool_consultar_conocimiento_tecnico(fn_args, context, conversation_context)
        elif fn_name == "radicar_reclamo":
            result = _handle_tool_radicar_reclamo(fn_args, context, conversation_context)
        elif fn_name == "confirmar_pedido_y_generar_pdf":
            result = _handle_tool_confirmar_pedido(fn_args, context, conversation_context)
        elif fn_name == "registrar_cliente_nuevo":
            result = _handle_tool_registrar_cliente_nuevo(fn_args, context, conversation_context)
        elif fn_name == "guardar_aprendizaje_producto":
            result = _handle_tool_guardar_aprendizaje_producto(fn_args, conversation_context)
        elif fn_name == "guardar_producto_complementario":
            result = _handle_tool_guardar_producto_complementario(fn_args, conversation_context)
        elif fn_name == "registrar_conocimiento_experto":
            result = _handle_tool_registrar_conocimiento_experto(fn_args, conversation_context)
        elif fn_name == "procesar_documento_experto":
            result = _handle_tool_procesar_documento_experto(fn_args, conversation_context)
        elif fn_name == "generar_memoria_tecnica":
            result = _handle_tool_generar_memoria_tecnica(fn_args, conversation_context)
        elif fn_name == "consultar_base_color":
            result = _handle_tool_consultar_base_color(fn_args)
        elif fn_name == "consultar_referencia_international":
            result = _handle_tool_consultar_referencia_international(fn_args)
        else:
            result = json.dumps({"error": f"Herramienta desconocida: {fn_name}"}, ensure_ascii=False)
    except Exception as tool_exc:
        elapsed_ms = int((time.time() - t0) * 1000)
        logger.error("Tool %s FAILED after %dms: %s", fn_name, elapsed_ms, tool_exc, exc_info=True)
        result = json.dumps(
            {"error": True, "mensaje": f"Error temporal al ejecutar {fn_name}: {str(tool_exc)[:200]}. Infórmale al cliente que hubo un problema técnico momentáneo y que puede intentar de nuevo en unos minutos."},
            ensure_ascii=False,
        )

    elapsed_ms = int((time.time() - t0) * 1000)
    logger.info("Tool %s completed in %dms | args=%s", fn_name, elapsed_ms, json.dumps(fn_args, ensure_ascii=False)[:200])
    return fn_name, fn_args, result


# ── Handler: consultar_base_color ──
def _handle_tool_consultar_base_color(fn_args: dict) -> str:
    color = fn_args.get("color", "")
    producto = fn_args.get("producto", "")
    results = lookup_color_base(color, producto)
    if not results:
        return json.dumps({
            "encontrado": False,
            "mensaje": f"No encontré el color '{color}' en el catálogo de fórmulas. El cliente puede visitar www.ferreinox.co sección Cartas de Colores para ver la gama completa. Si tiene el código del color (ej: 1502), búscalo por código.",
        }, ensure_ascii=False)
    items = []
    for r in results:
        items.append({
            "codigo": r["codigo"],
            "nombre_color": r["nombre"],
            "base_requerida": r["base"],
            "linea_producto": r["producto"],
        })
    return json.dumps({
        "encontrado": True,
        "colores": items,
        "instruccion": (
            "Usa la BASE indicada para buscar en inventario. "
            "Ej: si base='Base Deep' y producto='Viniltex' → buscar 'Viniltex Base Deep galón'. "
            "La tintometría se realiza en tienda con la fórmula del color. "
            "⚠️ REGLA DOMÉSTICO: Doméstico NO viene en Base Accent. Si el color requiere Base Accent "
            "y el cliente pide Doméstico → decir: 'Ese color no viene en Doméstico porque requiere "
            "Base Accent. Te lo puedo ofrecer en Pintulux 3en1 o Pintulux Máxima Protección.' "
            "Los colores de Doméstico y Pintulux usan las mismas bases que Viniltex (excepto Accent en Doméstico). "
            "Menciona al cliente: 'Puedes ver toda nuestra gama de colores en www.ferreinox.co sección Cartas de Colores'."
        ),
    }, ensure_ascii=False)


# ── Handler: consultar_referencia_international ──
def _handle_tool_consultar_referencia_international(fn_args: dict) -> str:
    producto = fn_args.get("producto", "")
    base = fn_args.get("base", "")
    ral = fn_args.get("ral", "")

    # ── Intergard 2002 is SOBRE PEDIDO — block lookup, force escalation ──
    if "2002" in producto:
        return json.dumps({
            "encontrado": False,
            "sobre_pedido": True,
            "producto": "Intergard 2002",
            "mensaje": (
                "⚠️ Intergard 2002 es un PRODUCTO SOBRE PEDIDO. NO cotices precio. "
                "Este sistema de alto desempeño para pisos de tráfico pesado requiere "
                "asesoría técnica personalizada. Pregunta al cliente: "
                "'¿Deseas que te contacte con nuestro Asesor Técnico Comercial para "
                "estructurar tu proyecto?' Si acepta → escalar a tiendapintucopereira@ferreinox.co."
            ),
        }, ensure_ascii=False)

    # ── Default RAL 7038 if no RAL specified ──
    _used_default_ral = False
    if not ral:
        ral = "7038"
        _used_default_ral = True

    results = lookup_international_product(producto, base, ral)
    if not results:
        return json.dumps({
            "encontrado": False,
            "mensaje": f"No encontré '{producto}' en la tabla de referencia International. Usa consultar_inventario para buscar por nombre comercial.",
        }, ensure_ascii=False)
    items = []
    for r in results:
        entry = {"producto": r.get("producto", ""), "base": r.get("base", ""), "ral": r.get("ral", "")}
        if r.get("kit_galon"):
            entry["precio_kit_galon_iva_inc"] = r["kit_galon"]
        if r.get("precio_galon"):
            entry["precio_base_galon_iva_inc"] = r["precio_galon"]
        if r.get("codigo_base_galon"):
            entry["codigo_base_galon"] = r["codigo_base_galon"]
        if r.get("codigo_cat_galon"):
            entry["codigo_catalizador_galon"] = r["codigo_cat_galon"]
        if r.get("precio_cat_galon"):
            entry["precio_catalizador_galon_iva_inc"] = r["precio_cat_galon"]
        if r.get("kit_cunete"):
            entry["precio_kit_cunete_iva_inc"] = r["kit_cunete"]
        if r.get("precio_cunete"):
            entry["precio_base_cunete_iva_inc"] = r["precio_cunete"]
        if r.get("codigo_cunete"):
            entry["codigo_base_cunete"] = r["codigo_cunete"]
        if r.get("codigo_cat_cunete"):
            entry["codigo_catalizador_cunete"] = r["codigo_cat_cunete"]
        if r.get("precio_cat_cunete"):
            entry["precio_catalizador_cunete_iva_inc"] = r["precio_cat_cunete"]
        # Acrilica Mantenimiento fields
        if r.get("codigo_galon"):
            entry["codigo_galon"] = r["codigo_galon"]
            entry["precio_galon_iva_inc"] = r.get("precio_galon", "")
        if r.get("codigo_cunete"):
            entry["codigo_cunete"] = r["codigo_cunete"]
            entry["precio_cunete_iva_inc"] = r.get("precio_cunete", "")
        if r.get("tonalidad"):
            entry["tonalidad"] = r["tonalidad"]
        items.append(entry)
    return json.dumps({
        "encontrado": True,
        "productos": items,
        "total_resultados": len(items),
        "ral_usado": ral,
        "ral_default_aplicado": _used_default_ral,
        "instruccion": (
            "⚠️ IMPORTANTE: Los precios de esta tabla YA INCLUYEN IVA. "
            "NO sumes IVA de nuevo. El precio KIT galón = base + catalizador ya con IVA. "
            "Para cotizar: precio_kit × cantidad = subtotal. El total YA es con IVA incluido. "
            "Usa los CÓDIGOS de referencia para buscar disponibilidad con consultar_inventario."
            + (" Se usó RAL 7038 (gris claro) por defecto porque el cliente no especificó color. "
               "SIEMPRE agrega: 'Para más colores RAL disponibles, visita www.ferreinox.co'" if _used_default_ral else "")
        ),
    }, ensure_ascii=False)



def get_internal_bootstrap_token():
    return os.getenv("INTERNAL_AUTH_BOOTSTRAP_TOKEN")


@app.post("/agent/auth/bootstrap-user")
def bootstrap_internal_user(
    payload: InternalBootstrapUserRequest,
    x_bootstrap_token: Optional[str] = Header(default=None),
):
    configured_token = get_internal_bootstrap_token()
    if not configured_token:
        raise HTTPException(status_code=503, detail="No se configuró INTERNAL_AUTH_BOOTSTRAP_TOKEN en el backend.")
    if x_bootstrap_token != configured_token:
        raise HTTPException(status_code=403, detail="Bootstrap token inválido.")
    user_payload = upsert_internal_user(payload)
    return {
        "status": "ok",
        "user": {
            "id": user_payload.get("id"),
            "username": user_payload.get("username"),
            "full_name": user_payload.get("full_name"),
            "role": user_payload.get("role"),
            "phone_e164": user_payload.get("phone_e164"),
            "email": user_payload.get("email"),
            "scopes": user_payload.get("scopes") or [],
        },
    }


@app.post("/agent/auth/login-internal")
def login_internal_agent(payload: InternalLoginRequest):
    user_payload = authenticate_internal_user(payload.username, payload.password)
    if not user_payload:
        raise HTTPException(status_code=401, detail="Usuario o contraseña interna inválidos.")
    session_payload = create_internal_session(user_payload, channel="api")
    return {
        "status": "ok",
        "access_token": session_payload["token"],
        "token_type": "bearer",
        "expires_at": session_payload["expires_at"],
        "user": {
            "id": user_payload.get("id"),
            "username": user_payload.get("username"),
            "full_name": user_payload.get("full_name"),
            "role": user_payload.get("role"),
            "phone_e164": user_payload.get("phone_e164"),
            "email": user_payload.get("email"),
            "scopes": user_payload.get("scopes") or [],
        },
    }


@app.post("/agent/auth/logout")
def logout_internal_agent(authorization: Optional[str] = Header(default=None)):
    token = extract_bearer_token(authorization)
    if token:
        revoke_internal_session(token)
    return {"status": "ok"}


@app.get("/agent/auth/me")
def get_internal_agent_me(authorization: Optional[str] = Header(default=None)):
    internal_user = require_internal_user(authorization)
    return {
        "id": internal_user.get("id"),
        "username": internal_user.get("username"),
        "full_name": internal_user.get("full_name"),
        "role": internal_user.get("role"),
        "email": internal_user.get("email"),
        "phone_e164": internal_user.get("phone_e164"),
        "session_expires_at": internal_user.get("session_expires_at"),
        "scopes": internal_user.get("scopes") or [],
    }


@app.get("/agent/internal/clientes/buscar")
def search_internal_customers(q: str, authorization: Optional[str] = Header(default=None)):
    internal_user = require_internal_user(authorization)
    customer_rows = search_customer_lookup_rows(q, limit=10)
    visible_rows = [row for row in customer_rows if internal_user_can_access_customer(internal_user, row)]
    return {"items": visible_rows}


@app.get("/agent/internal/clientes/{cliente_codigo}/contexto")
def get_internal_customer_context(cliente_codigo: str, authorization: Optional[str] = Header(default=None)):
    internal_user = require_internal_user(authorization)
    customer_row = fetch_customer_lookup_row(cliente_codigo)
    if not customer_row:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
    if not internal_user_can_access_customer(internal_user, customer_row):
        raise HTTPException(status_code=403, detail="No tienes permisos para consultar ese cliente.")
    contexto = get_cliente_contexto(cliente_codigo)
    return {"lookup": customer_row, "contexto": contexto}


@app.get("/agent/internal/clientes/{cliente_codigo}/cartera")
def get_internal_customer_portfolio(cliente_codigo: str, authorization: Optional[str] = Header(default=None)):
    internal_user = require_internal_user(authorization)
    customer_row = fetch_customer_lookup_row(cliente_codigo)
    if not customer_row:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
    if not internal_user_can_access_customer(internal_user, customer_row):
        raise HTTPException(status_code=403, detail="No tienes permisos para consultar ese cliente.")
    contexto = get_cliente_contexto(cliente_codigo)
    overdue_info = fetch_overdue_documents(cliente_codigo)
    return {
        "lookup": customer_row,
        "contexto": contexto,
        "vencidos": overdue_info,
    }


@app.get("/agent/internal/clientes/{cliente_codigo}/compras")
def get_internal_customer_purchases(cliente_codigo: str, authorization: Optional[str] = Header(default=None)):
    internal_user = require_internal_user(authorization)
    customer_row = fetch_customer_lookup_row(cliente_codigo)
    if not customer_row:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
    if not internal_user_can_access_customer(internal_user, customer_row):
        raise HTTPException(status_code=403, detail="No tienes permisos para consultar ese cliente.")
    latest_purchase = fetch_latest_purchase_detail(cliente_codigo)
    purchase_summary = fetch_purchase_summary(cliente_codigo)
    return {
        "lookup": customer_row,
        "ultima_compra": latest_purchase,
        "resumen": purchase_summary,
    }


@app.get("/agent/internal/productos/buscar")
def search_internal_products(q: str, store: Optional[str] = None, authorization: Optional[str] = Header(default=None)):
    require_internal_user(authorization)
    product_request = prepare_product_request_for_search(q)
    rows = lookup_product_context(q, product_request)
    response_rows = []
    for row in rows[:10]:
        if store and normalize_text_value(store) not in normalize_text_value(row.get("stock_by_store") or row.get("stock_por_tienda") or ""):
            continue
        response_rows.append(row)
    return {"items": response_rows, "nlu_extraccion": product_request.get("nlu_extraction") or {}}


@app.get("/")
def read_root():
    return {
        "estado": "Sistema CRM Ferreinox Activo",
        "version": "2026.3",
        "postgrest_url": get_postgrest_url(),
        "endpoints": [
            "/health",
            "/agent/clientes/{cliente_codigo}/contexto",
            "/agent/auth/login-internal",
            "/agent/auth/me",
            "/agent/internal/clientes/buscar",
            "/webhooks/whatsapp",
        ],
    }


@app.get("/health")
def health_check():
    postgrest_url = get_postgrest_url()
    try:
        response = requests.get(f"{postgrest_url}/", timeout=5)
        response.raise_for_status()
        return {"backend": "ok", "postgrest": "ok", "postgrest_url": postgrest_url}
    except Exception as exc:
        return {"backend": "ok", "postgrest": "error", "postgrest_url": postgrest_url, "detail": str(exc)}


@app.get("/admin/rag-buscar")
def admin_rag_buscar(
    q: str = "",
    top_k: int = 6,
    producto: str = "",
    admin_key: str = Header(None, alias="x-admin-key"),
):
    """Test de búsqueda semántica RAG con expansión de portafolio (simula el agente real)."""
    expected = os.getenv("ADMIN_API_KEY", "ferreinox_admin_2024")
    if admin_key != expected:
        raise HTTPException(status_code=403, detail="Admin key inválida")
    if not q.strip():
        return {"error": "Parámetro q requerido"}
    try:
        query = q.strip()
        search_q = f"{producto.strip()}: {query}" if producto.strip() else query
        chunks = search_technical_chunks(search_q, top_k=top_k)

        # Portfolio-aware expansion (same logic as the real agent handler)
        best_sim = max((c.get("similarity", 0) for c in chunks), default=0)
        if best_sim < 0.70 and not producto.strip():
            q_norm = normalize_text_value(query)
            portfolio_prods: list[str] = []
            for cat_key, brand_terms in PORTFOLIO_CATEGORY_MAP.items():
                if cat_key in q_norm or q_norm in cat_key:
                    for bt in brand_terms:
                        if bt != "__SIN_PRODUCTO_FERREINOX__" and bt not in portfolio_prods:
                            portfolio_prods.append(bt)
            for word in q_norm.split():
                if len(word) < 4 or word in ("para", "como", "esto", "esta", "esos", "esas", "unos", "unas", "tiene"):
                    continue
                if word in PORTFOLIO_CATEGORY_MAP:
                    for bt in PORTFOLIO_CATEGORY_MAP[word]:
                        if bt != "__SIN_PRODUCTO_FERREINOX__" and bt not in portfolio_prods:
                            portfolio_prods.append(bt)
            if portfolio_prods:
                extra: list[dict] = []
                for pp in portfolio_prods[:3]:
                    extra.extend(search_technical_chunks(f"{pp}: {query}", top_k=3))
                seen_t: set[str] = set()
                merged: list[dict] = []
                for ch in sorted(chunks + extra, key=lambda c: c.get("similarity", 0), reverse=True):
                    tk = (ch.get("chunk_text") or "")[:80]
                    if tk not in seen_t:
                        seen_t.add(tk)
                        merged.append(ch)
                chunks = merged[:top_k + 2]

        results = []
        for c in chunks:
            results.append({
                "archivo": c.get("doc_filename"),
                "familia": c.get("familia_producto"),
                "similitud": round(c.get("similarity", 0), 4),
                "texto": (c.get("chunk_text") or "")[:500],
            })
        rag_ctx = build_rag_context(chunks, max_chunks=4)
        candidates = extract_candidate_products_from_rag_context(
            rag_ctx, original_question=query,
        )
        return {
            "query": query,
            "resultados": results,
            "productos_candidatos": candidates,
        }
    except Exception as exc:
        return {"error": str(exc)}



@app.get("/admin/alertas-agente")
def admin_alertas_agente(
    limit: int = 20,
    estado: str = "pendiente",
    admin_key: str = Header(None, alias="x-admin-key"),
):
    """Consulta alertas generadas por respuestas de baja confianza del agente."""
    expected = os.getenv("ADMIN_API_KEY", "ferreinox_admin_2024")
    if admin_key != expected:
        raise HTTPException(status_code=403, detail="Admin key inválida")
    try:
        engine = get_db_engine()
        with engine.connect() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT t.id, t.conversation_id, t.tipo_tarea, t.prioridad, t.estado,
                           t.resumen, t.detalle, t.created_at,
                           c.telefono_e164, c.nombre_visible
                    FROM public.agent_task t
                    LEFT JOIN public.agent_conversation ac ON ac.id = t.conversation_id
                    LEFT JOIN public.whatsapp_contacto c ON c.id = ac.contact_id
                    WHERE t.tipo_tarea IN ('respuesta_baja_confianza', 'multiples_señales_problema', 'revision_manual')
                      AND (:estado = 'todos' OR t.estado = :estado)
                    ORDER BY t.created_at DESC
                    LIMIT :limit
                    """
                ),
                {"estado": estado, "limit": limit},
            ).mappings().all()
            alertas = []
            for r in rows:
                detalle = r["detalle"]
                if isinstance(detalle, str):
                    try:
                        detalle = json.loads(detalle)
                    except Exception:
                        pass
                alertas.append({
                    "id": r["id"],
                    "conversation_id": r["conversation_id"],
                    "tipo": r["tipo_tarea"],
                    "prioridad": r["prioridad"],
                    "estado": r["estado"],
                    "resumen": r["resumen"],
                    "detalle": detalle,
                    "created_at": str(r["created_at"]),
                    "telefono": r.get("telefono_e164"),
                    "nombre": r.get("nombre_visible"),
                })
            return {"total": len(alertas), "alertas": alertas}
    except Exception as exc:
        return {"error": str(exc)}


@app.get("/admin/calidad-agente")
def admin_calidad_agente(
    horas: int = 24,
    admin_key: str = Header(None, alias="x-admin-key"),
):
    """Resumen de calidad del agente: conversaciones cerradas, alertas, confianza promedio."""
    expected = os.getenv("ADMIN_API_KEY", "ferreinox_admin_2024")
    if admin_key != expected:
        raise HTTPException(status_code=403, detail="Admin key inválida")
    try:
        engine = get_db_engine()
        with engine.connect() as connection:
            # Conversaciones en las últimas N horas
            conv_stats = connection.execute(
                text(
                    """
                    SELECT
                        COUNT(*) AS total,
                        COUNT(*) FILTER (WHERE estado = 'cerrada') AS cerradas,
                        COUNT(*) FILTER (WHERE estado = 'abierta') AS abiertas,
                        COUNT(*) FILTER (WHERE contexto->>'final_status' = 'gestionado') AS gestionadas,
                        COUNT(*) FILTER (WHERE contexto->>'close_reason' = 'farewell_detected') AS cerradas_despedida
                    FROM public.agent_conversation
                    WHERE last_message_at >= now() - make_interval(hours => :horas)
                    """
                ),
                {"horas": horas},
            ).mappings().one()

            # Alertas en las últimas N horas
            alert_stats = connection.execute(
                text(
                    """
                    SELECT
                        COUNT(*) AS total_alertas,
                        COUNT(*) FILTER (WHERE tipo_tarea = 'respuesta_baja_confianza') AS baja_confianza,
                        COUNT(*) FILTER (WHERE tipo_tarea = 'multiples_señales_problema') AS multiples_señales,
                        COUNT(*) FILTER (WHERE tipo_tarea = 'revision_manual') AS revision_manual,
                        COUNT(*) FILTER (WHERE estado = 'pendiente') AS pendientes
                    FROM public.agent_task
                    WHERE created_at >= now() - make_interval(hours => :horas)
                      AND tipo_tarea IN ('respuesta_baja_confianza', 'multiples_señales_problema', 'revision_manual')
                    """
                ),
                {"horas": horas},
            ).mappings().one()

            # Aprendizajes recientes
            learning_stats = connection.execute(
                text(
                    """
                    SELECT COUNT(*) AS total_aprendizajes
                    FROM public.agent_product_learning
                    WHERE updated_at >= now() - make_interval(hours => :horas)
                    """
                ),
                {"horas": horas},
            ).mappings().one()

            return {
                "periodo_horas": horas,
                "conversaciones": dict(conv_stats),
                "alertas": dict(alert_stats),
                "aprendizajes_recientes": learning_stats["total_aprendizajes"],
            }
    except Exception as exc:
        return {"error": str(exc)}


@app.get("/admin/rag-diagnostico")
def admin_rag_diagnostico(admin_key: str = Header(None, alias="x-admin-key")):
    """Diagnóstico del RAG: qué fichas técnicas hay indexadas."""
    expected = os.getenv("ADMIN_API_KEY", "ferreinox_admin_2024")
    if admin_key != expected:
        raise HTTPException(status_code=403, detail="Admin key inválida")
    try:
        engine = get_db_engine()
        with engine.connect() as conn:
            total = conn.execute(text(
                "SELECT COUNT(*) FROM public.agent_technical_doc_chunk WHERE tipo_documento IN ('ficha_tecnica', 'ficha_tecnica_experto')"
            )).scalar() or 0
            total_profiles = conn.execute(text(
                "SELECT COUNT(*) FROM public.agent_technical_profile WHERE extraction_status = 'ready'"
            )).scalar() or 0
            avg_profile_score = conn.execute(text(
                "SELECT AVG(completeness_score) FROM public.agent_technical_profile WHERE extraction_status = 'ready'"
            )).scalar()
            docs = conn.execute(text(
                "SELECT doc_filename, marca, familia_producto, tipo_documento, "
                "       COALESCE(metadata ->> 'canonical_family', familia_producto) AS canonical_family, "
                "       COALESCE(metadata ->> 'document_scope', 'primary') AS document_scope, "
                "       COUNT(*) as chunks "
                "FROM public.agent_technical_doc_chunk "
                "WHERE tipo_documento IN ('ficha_tecnica', 'ficha_tecnica_experto') "
                "GROUP BY doc_filename, marca, familia_producto, tipo_documento, "
                "         COALESCE(metadata ->> 'canonical_family', familia_producto), "
                "         COALESCE(metadata ->> 'document_scope', 'primary') "
                "ORDER BY COALESCE(metadata ->> 'canonical_family', familia_producto), doc_filename"
            )).fetchall()
        return {
            "total_chunks": total,
            "total_perfiles": total_profiles,
            "promedio_completitud_perfiles": round(float(avg_profile_score or 0), 4),
            "documentos": [
                {
                    "archivo": r[0],
                    "marca": r[1],
                    "familia": r[2],
                    "tipo": r[3],
                    "familia_canonica": r[4],
                    "scope": r[5],
                    "chunks": r[6],
                }
                for r in docs
            ],
        }
    except Exception as exc:
        return {"error": str(exc)}


@app.post("/admin/agent-test")
async def admin_agent_test(request: Request, admin_key: str = Header(None, alias="x-admin-key")):
    """Endpoint de testing: ejecuta generate_agent_reply_v3 sin pasar por WhatsApp.
    Útil para pruebas automáticas y validación de aserciones por turno.
    Body JSON esperado: { profile_name, conversation_context, recent_messages, user_message, context }
    """
    expected = os.getenv("ADMIN_API_KEY", "ferreinox_admin_2024")
    if admin_key != expected:
        raise HTTPException(status_code=403, detail="Admin key inválida")
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    profile_name = payload.get("profile_name")
    conversation_context = payload.get("conversation_context") or {}
    recent_messages = payload.get("recent_messages") or []
    user_message = payload.get("user_message") or ""
    context = payload.get("context") or {}

    try:
        result = generate_agent_reply_v3(profile_name, conversation_context, recent_messages, user_message, context)
        return {"ok": True, "result": result}
    except Exception as exc:
        return {"error": str(exc)}

# ---------------------------------------------------------------------------
# ADMIN: Importar articulos_maestro desde Excel (upload)
# ---------------------------------------------------------------------------
@app.post("/admin/importar-articulos-maestro")
async def admin_importar_articulos_maestro(
    archivo: UploadFile = File(...),
    admin_key: str = Header(None, alias="x-admin-key"),
):
    """
    Sube articulos.xlsx y ejecuta:
      1. Crea tabla articulos_maestro (si no existe)
      2. Importa/actualiza 20,000+ artículos con clasificación ERP
      3. Actualiza las vistas de inventario (search_blob enriquecido)

    Uso con curl:
      curl -X POST https://apicrm.datovatenexuspro.com/admin/importar-articulos-maestro \
        -H "x-admin-key: TU_ADMIN_KEY" \
        -F "archivo=@articulos.xlsx"
    """
    expected_key = os.getenv("ADMIN_API_KEY", "ferreinox_admin_2024")
    if admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Admin key inválida. Envía header x-admin-key.")

    import tempfile
    import unicodedata as _unicodedata

    import pandas as _pd

    # --- Guardar archivo temporal ---
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        content = await archivo.read()
        tmp.write(content)
        tmp.close()
        excel_path = tmp.name

        # --- Leer Excel ---
        df = _pd.read_excel(excel_path, sheet_name="Hoja1", header=0, dtype=str)

        # --- Mapeo columnas ---
        HEADER_TO_DB = {
            "Descripción": "descripcion", "Cód. Barras": "codigo_barras",
            "Código Artículo": "codigo_articulo", "Referencia": "referencia",
            "Descripción Adicional": "descripcion_adicional", "Departamento": "departamento",
            "Seccion": "seccion", "Família": "familia", "SubFamilia": "subfamilia",
            "Marca": "marca_erp", "Linea": "linea_erp", "PROVEEDOR": "proveedor",
            "DESCRIPCION_EBS": "descripcion_ebs", "UDM": "udm",
            "CAT_PRODUCTO": "cat_producto", "APLICACION": "aplicacion",
            "LINEA": "linea_clasificacion", "SUBLINEA": "sublinea",
            "MARCA": "marca_clasificacion", "FAMILIA": "familia_clasificacion",
            "SUBFAMILIA": "subfamilia_clasificacion", "TIPO": "tipo",
        }
        rename_map = {}
        for header, db_col in HEADER_TO_DB.items():
            matches = [h for h in df.columns if h.strip() == header.strip()]
            if matches:
                rename_map[matches[-1]] = db_col

        df_sel = df[list(rename_map.keys())].copy()
        df_sel.columns = [rename_map[c] for c in df_sel.columns]

        def _clean(val):
            if val is None or _pd.isna(val):
                return None
            s = str(val).strip()
            return None if s in ("", "None", "0", " ", "nan") else s

        for col in df_sel.columns:
            df_sel[col] = df_sel[col].apply(_clean)

        df_sel = df_sel[df_sel["referencia"].notna()].copy()
        df_sel = df_sel.drop_duplicates(subset="referencia", keep="first")

        def _keep_alnum(tv):
            if not tv or str(tv).strip() in ("", "None"):
                return None
            s = _unicodedata.normalize("NFD", str(tv).strip().upper())
            s = "".join(c for c in s if _unicodedata.category(c) != "Mn")
            s = re.sub(r"[^A-Z0-9]", "", s)
            return s or None

        df_sel["referencia_normalizada"] = df_sel["referencia"].apply(_keep_alnum)

        DB_COLUMNS = [
            "codigo_articulo", "referencia", "referencia_normalizada", "codigo_barras",
            "descripcion", "descripcion_adicional", "descripcion_ebs",
            "departamento", "seccion", "familia", "subfamilia", "marca_erp", "linea_erp",
            "proveedor", "udm", "cat_producto", "aplicacion",
            "linea_clasificacion", "sublinea", "marca_clasificacion",
            "familia_clasificacion", "subfamilia_clasificacion", "tipo",
        ]
        for col in DB_COLUMNS:
            if col not in df_sel.columns:
                df_sel[col] = None

        # --- Conectar y ejecutar ---
        engine = get_db_engine()
        with engine.begin() as conn:
            # 1. Crear tabla
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS public.articulos_maestro (
                    id bigserial PRIMARY KEY, codigo_articulo text, referencia text NOT NULL,
                    referencia_normalizada text, codigo_barras text, descripcion text,
                    descripcion_adicional text, descripcion_ebs text, departamento text,
                    seccion text, familia text, subfamilia text, marca_erp text, linea_erp text,
                    proveedor text, udm text, cat_producto text, aplicacion text,
                    linea_clasificacion text, sublinea text, marca_clasificacion text,
                    familia_clasificacion text, subfamilia_clasificacion text, tipo text,
                    activo boolean DEFAULT true,
                    created_at timestamptz DEFAULT now(), updated_at timestamptz DEFAULT now()
                )
            """))
            conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_articulos_maestro_ref ON public.articulos_maestro (referencia)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS idx_articulos_maestro_ref_norm ON public.articulos_maestro (referencia_normalizada)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS idx_articulos_maestro_codigo ON public.articulos_maestro (codigo_articulo)"))

            # 2. Truncar e importar
            conn.execute(text("TRUNCATE public.articulos_maestro RESTART IDENTITY"))

            upsert_sql = """
                INSERT INTO public.articulos_maestro (
                    codigo_articulo, referencia, referencia_normalizada, codigo_barras,
                    descripcion, descripcion_adicional, descripcion_ebs,
                    departamento, seccion, familia, subfamilia, marca_erp, linea_erp,
                    proveedor, udm, cat_producto, aplicacion,
                    linea_clasificacion, sublinea, marca_clasificacion,
                    familia_clasificacion, subfamilia_clasificacion, tipo
                ) VALUES (
                    :codigo_articulo, :referencia, :referencia_normalizada, :codigo_barras,
                    :descripcion, :descripcion_adicional, :descripcion_ebs,
                    :departamento, :seccion, :familia, :subfamilia, :marca_erp, :linea_erp,
                    :proveedor, :udm, :cat_producto, :aplicacion,
                    :linea_clasificacion, :sublinea, :marca_clasificacion,
                    :familia_clasificacion, :subfamilia_clasificacion, :tipo
                )
            """
            records = df_sel[DB_COLUMNS].to_dict("records")
            batch_size = 500
            for i in range(0, len(records), batch_size):
                conn.execute(text(upsert_sql), records[i: i + batch_size])

            total_imported = conn.execute(text("SELECT COUNT(*) FROM public.articulos_maestro")).scalar()

            # 3. Actualizar vistas: DROP cascade y recrear con clasificación enriquecida
            conn.execute(text("DROP VIEW IF EXISTS public.vw_agente_producto_disponibilidad CASCADE"))
            conn.execute(text("DROP VIEW IF EXISTS public.productos CASCADE"))
            conn.execute(text("DROP VIEW IF EXISTS public.vw_inventario_agente CASCADE"))

            sql_migration_path = Path(__file__).resolve().parent / "articulos_maestro_setup.sql"
            if sql_migration_path.exists():
                sql_content = sql_migration_path.read_text(encoding="utf-8")
                # Quitar comentarios SQL (-- ...) antes de parsear
                clean_sql = re.sub(r"--[^\n]*", "", sql_content)
                # Regex que respeta strings SQL entre comillas simples
                for stmt_match in re.finditer(
                    r"(CREATE\s+(?:OR\s+REPLACE\s+)?VIEW\s+(?:[^;']|'[^']*')*;)",
                    clean_sql,
                    re.DOTALL | re.IGNORECASE,
                ):
                    conn.execute(text(stmt_match.group(1)))

            # Refresh materialized views used by product search
            try:
                conn.execute(text("REFRESH MATERIALIZED VIEW mv_productos"))
                conn.execute(text("REFRESH MATERIALIZED VIEW mv_product_rotation"))
            except Exception:
                pass  # May not exist yet on first setup

        engine.dispose()
        return {
            "exito": True,
            "articulos_importados": total_imported,
            "columnas_detectadas": list(rename_map.values()),
            "mensaje": f"Se importaron {total_imported:,} artículos y se actualizaron las vistas de búsqueda.",
        }

    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error importando artículos: {exc}") from exc
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


# ADMIN: Importar catálogo Abracol desde Dropbox
@app.post("/admin/importar-catalogo-abracol")
async def admin_importar_catalogo_abracol(
    request: Request,
    x_admin_key: str = Header(None, alias="x-admin-key"),
):
    """
    Importa el catálogo enriquecido de Abracol desde Dropbox (hoja Productos).
    Enriquece search_blob de mv_productos con nombre_comercial, familia y descripcion_larga.

    curl -X POST https://apicrm.datovatenexuspro.com/admin/importar-catalogo-abracol \\
         -H 'x-admin-key: ferreinox_admin_2024'
    """
    if x_admin_key != os.getenv("ADMIN_API_KEY", "ferreinox_admin_2024"):
        raise HTTPException(status_code=403, detail="Admin key inválida")

    try:
        from io import BytesIO
        import pandas as pd
        from frontend.config import get_dropbox_sources
        from frontend.dropbox_sync_service import get_dropbox_client

        # 1. Download from Dropbox
        sources = get_dropbox_sources()
        rotacion = None
        for key, cfg in sources.items():
            if "rotaci" in key.lower():
                rotacion = cfg
                break
        if not rotacion:
            raise HTTPException(status_code=500, detail="No se encontró config dropbox_rotacion")

        dbx = get_dropbox_client(rotacion)
        folder = rotacion.get("folder", "/data")
        result = dbx.files_list_folder(folder)
        abracol_path = None
        for entry in result.entries:
            if "abracol" in entry.name.lower() and entry.name.lower().endswith((".xlsx", ".xls")):
                abracol_path = entry.path_lower
                break
        if not abracol_path:
            raise HTTPException(status_code=404, detail=f"No se encontró archivo Abracol en {folder}")

        _, response = dbx.files_download(abracol_path)
        df = pd.read_excel(BytesIO(response.content), sheet_name="Productos", dtype=str)

        # 2. Upsert into DB
        engine = get_engine()
        upsert_sql = """
        INSERT INTO public.abracol_productos (
            codigo, nombre_comercial, descripcion, grano, medida,
            familia, empaque, portafolio, descripcion_larga, search_keywords
        ) VALUES (
            :codigo, :nombre_comercial, :descripcion, :grano, :medida,
            :familia, :empaque, :portafolio, :descripcion_larga, :search_keywords
        )
        ON CONFLICT (codigo) DO UPDATE SET
            nombre_comercial = EXCLUDED.nombre_comercial,
            descripcion = EXCLUDED.descripcion,
            grano = EXCLUDED.grano, medida = EXCLUDED.medida,
            familia = EXCLUDED.familia, empaque = EXCLUDED.empaque,
            portafolio = EXCLUDED.portafolio,
            descripcion_larga = EXCLUDED.descripcion_larga,
            search_keywords = EXCLUDED.search_keywords,
            updated_at = now();
        """

        rows = []
        for _, row in df.iterrows():
            codigo = (row.get("CODIGO") or "").strip()
            if not codigo:
                continue
            sk_parts = [row.get(c) or "" for c in ["NOMBRE COMERCIAL", "DESCRIPCION", "FAMILIA", "PORTAFOLIO", "GRANO", "MEDIDA", "DESCRIPCION_LARGA"]]
            rows.append({
                "codigo": codigo,
                "nombre_comercial": (row.get("NOMBRE COMERCIAL") or "").strip() or None,
                "descripcion": (row.get("DESCRIPCION") or "").strip() or None,
                "grano": (row.get("GRANO") or "").strip() or None,
                "medida": (row.get("MEDIDA") or "").strip() or None,
                "familia": (row.get("FAMILIA") or "").strip() or None,
                "empaque": (row.get("EMPAQUE") or "").strip() or None,
                "portafolio": (row.get("PORTAFOLIO") or "").strip() or None,
                "descripcion_larga": (row.get("DESCRIPCION_LARGA") or "").strip() or None,
                "search_keywords": " ".join(p.strip() for p in sk_parts if p.strip()).lower(),
            })

        with engine.begin() as conn:
            # Ensure table exists
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS public.abracol_productos (
                    codigo varchar(20) PRIMARY KEY,
                    nombre_comercial text, descripcion text, grano varchar(60),
                    medida varchar(120), familia varchar(200), empaque varchar(20),
                    portafolio varchar(60), descripcion_larga text, search_keywords text,
                    created_at timestamptz DEFAULT now(), updated_at timestamptz DEFAULT now()
                )
            """))
            # Batch upsert
            for i in range(0, len(rows), 200):
                conn.execute(text(upsert_sql), rows[i:i+200])
            # Refresh search matview to include new Abracol data
            try:
                conn.execute(text("REFRESH MATERIALIZED VIEW mv_productos"))
            except Exception:
                pass

        portafolios = {}
        for r in rows:
            p = r.get("portafolio") or "SIN_PORTAFOLIO"
            portafolios[p] = portafolios.get(p, 0) + 1

        return {
            "exito": True,
            "productos_importados": len(rows),
            "portafolios": portafolios,
            "mensaje": f"Se importaron {len(rows)} productos Abracol y se refrescó mv_productos.",
        }

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error importando catálogo Abracol: {exc}") from exc


@app.get("/agent/clientes/{cliente_codigo}/contexto")
def get_cliente_contexto(cliente_codigo: str):
    postgrest_url = get_postgrest_url()
    try:
        response = requests.get(
            f"{postgrest_url}/vw_cliente_contexto_agente",
            params={"cliente_codigo": f"eq.{cliente_codigo}", "select": "*", "limit": 1},
            timeout=10,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible consultar PostgREST: {exc}") from exc

    if not payload:
        raise HTTPException(status_code=404, detail=f"No se encontró contexto para el cliente {cliente_codigo}")

    return payload[0]


@app.get("/webhooks/whatsapp")
def verify_whatsapp_webhook(
    hub_mode: str = Query(alias="hub.mode"),
    hub_verify_token: str = Query(alias="hub.verify_token"),
    hub_challenge: str = Query(alias="hub.challenge"),
):
    if hub_mode == "subscribe" and hub_verify_token == get_whatsapp_verify_token():
        return int(hub_challenge)
    raise HTTPException(status_code=403, detail="Token de verificación inválido")


@app.get("/pdf/{pdf_id}")
def serve_commercial_pdf(pdf_id: str):
    entry = PDF_STORAGE.get(pdf_id)
    if not entry:
        raise HTTPException(status_code=404, detail="PDF no encontrado o expirado")
    return StreamingResponse(
        io.BytesIO(entry["buffer"]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=\"{entry['filename']}\""},
    )


@app.post("/conversations/{conversation_id}/reset")
def reset_conversation_context(conversation_id: int):
    """Clear all conversation context so the agent starts fresh."""
    engine = get_db_engine()
    with engine.begin() as connection:
        row = connection.execute(
            text(
                """
                SELECT id
                FROM public.agent_conversation
                WHERE id = :conversation_id
                """
            ),
            {"conversation_id": conversation_id},
        ).mappings().one_or_none()
        if not row:
            raise HTTPException(status_code=404, detail="Conversación no encontrada")
        connection.execute(
            text(
                """
                UPDATE public.agent_conversation
                SET contexto = '{}'::jsonb,
                    resumen = 'Contexto reiniciado manualmente',
                    estado = 'abierta',
                    updated_at = now(),
                    last_message_at = now()
                WHERE id = :conversation_id
                """
            ),
            {"conversation_id": conversation_id},
        )
    return {"status": "ok", "conversation_id": conversation_id, "message": "Contexto limpiado"}


@app.post("/admin/cleanup-phone")
def admin_cleanup_phone(request: Request, payload: dict = Body(...)):
    """
    Limpiar COMPLETAMENTE el historial de conversación de un teléfono.
    Borra mensajes, resetea contexto, o borra toda la conversación.
    Requiere x-admin-key.
    
    Body: {"phone": "3205046277", "mode": "full"}
    Modes:
      - "context": Solo resetear contexto (mantiene mensajes)
      - "messages": Borrar mensajes (mantiene conversación)  
      - "full": Borrar mensajes + resetear contexto (fresh start)
    """
    admin_key = request.headers.get("x-admin-key", "")
    if admin_key != os.getenv("ADMIN_KEY", "ferreinox_admin_2024"):
        raise HTTPException(status_code=403, detail="Admin key inválida")

    phone_raw = str(payload.get("phone", "")).strip()
    mode = str(payload.get("mode", "full")).strip().lower()
    if not phone_raw:
        raise HTTPException(status_code=400, detail="Falta 'phone'")

    # Normalizar teléfono a E.164
    phone_e164 = phone_raw if phone_raw.startswith("+") else f"+57{phone_raw}"

    engine = get_db_engine()
    result = {"phone": phone_e164, "mode": mode}

    with engine.begin() as conn:
        # Buscar contacto
        contact = conn.execute(
            text("SELECT id FROM public.whatsapp_contacto WHERE telefono_e164 = :phone"),
            {"phone": phone_e164},
        ).mappings().one_or_none()

        if not contact:
            return {"status": "ok", "message": f"No se encontró contacto para {phone_e164}", **result}

        contacto_id = contact["id"]

        # Buscar conversaciones
        convs = conn.execute(
            text("SELECT id FROM public.agent_conversation WHERE contacto_id = :cid"),
            {"cid": contacto_id},
        ).mappings().all()
        conv_ids = [c["id"] for c in convs]
        result["conversations_found"] = len(conv_ids)

        if not conv_ids:
            return {"status": "ok", "message": "No tiene conversaciones", **result}

        if mode in ("messages", "full"):
            del_msgs = conn.execute(
                text("DELETE FROM public.agent_message WHERE conversation_id = ANY(:ids)"),
                {"ids": conv_ids},
            )
            result["messages_deleted"] = del_msgs.rowcount

        if mode in ("context", "full"):
            conn.execute(
                text("""
                    UPDATE public.agent_conversation
                    SET contexto = '{}'::jsonb,
                        resumen = 'Limpieza admin manual',
                        estado = 'abierta',
                        updated_at = now(),
                        last_message_at = now() - interval '4 hours'
                    WHERE contacto_id = :cid
                """),
                {"cid": contacto_id},
            )
            result["contexts_reset"] = len(conv_ids)

    result["status"] = "ok"
    result["message"] = f"Limpieza '{mode}' completada para {phone_e164}"
    return result


# ══════════════════════════════════════════════════════════════════════════════
# DEBOUNCE / BUFFER DE MENSAJES WHATSAPP
# Cuando un usuario envía varios mensajes rápidos ("60 mts", "blanco"), los
# acumula durante DEBOUNCE_WINDOW_SECONDS y los concatena en un solo mensaje
# antes de enviarlo al LLM. Evita "choques de trenes" y saludos falsos.
# ══════════════════════════════════════════════════════════════════════════════
DEBOUNCE_WINDOW_SECONDS = float(os.getenv("WA_DEBOUNCE_SECONDS", "4.0"))

# In-memory buffer: {phone_number: {"messages": [...], "timer_task": asyncio.Task, "context": ..., "meta": [...]}}
_wa_message_buffer: dict[str, dict] = {}
_wa_buffer_lock = asyncio.Lock()


async def _flush_debounce_buffer(phone_number: str):
    """Called after the debounce window expires. Concatenates buffered messages
    and processes them as a single unified message."""
    await asyncio.sleep(DEBOUNCE_WINDOW_SECONDS)

    async with _wa_buffer_lock:
        buf = _wa_message_buffer.pop(phone_number, None)
    if not buf or not buf["messages"]:
        return

    # Concatenate all buffered text messages into one
    unified_content = " ".join(buf["messages"])
    first_meta = buf["meta"][0]  # Use context/meta from the first message

    logger.info(
        "DEBOUNCE FLUSH: %s → %d mensajes unificados: '%s'",
        phone_number, len(buf["messages"]), unified_content[:200],
    )

    # Process the unified message through the normal pipeline
    try:
        context = first_meta["context"]
        conversation_context = first_meta["conversation_context"]
        recent_messages = first_meta["recent_messages"]
        watchdog_key = None

        if should_send_processing_ack(unified_content, conversation_context):
            watchdog_key = start_processing_watchdog(
                context,
                "⏳ Un momento, estoy procesando tu solicitud para enviártela completa.",
            )

        try:
            ai_result = handle_internal_whatsapp_message(unified_content, context, conversation_context)
            if ai_result is None:
                ai_result = generate_agent_reply_v3(
                    context.get("nombre_visible"),
                    conversation_context,
                    recent_messages,
                    unified_content,
                    context,
                )
        finally:
            stop_processing_watchdog(watchdog_key)

        response_text = ai_result.get("response_text") or "Gracias por escribirnos. ¿En qué te puedo ayudar?"

        try:
            outbound_payload = send_whatsapp_text_message(context["telefono_e164"], response_text)
            provider_message_id = None
            if outbound_payload.get("messages"):
                provider_message_id = outbound_payload["messages"][0].get("id")
            store_outbound_message(
                context["conversation_id"],
                provider_message_id,
                "text",
                response_text,
                outbound_payload,
                intent_detectado=ai_result.get("intent"),
            )
        except Exception as exc:
            store_outbound_message(
                context["conversation_id"],
                None,
                "system",
                f"No fue posible enviar respuesta: {exc}",
                {"error": str(exc), "response_text": response_text},
                intent_detectado=ai_result.get("intent"),
            )

        # Send technical document if applicable
        source_filename = ai_result.get("technical_source_filename") if isinstance(ai_result, dict) else None
        if source_filename:
            try:
                doc_entry = find_technical_document_entry_by_name(source_filename)
                if doc_entry:
                    _send_document_and_respond(doc_entry, context)
            except Exception:
                pass

        # Update conversation context
        context_updates = {
            "intent": ai_result.get("intent"),
            "last_direct_intent": ai_result.get("intent"),
            "verified": conversation_context.get("verified", False),
            "verified_document": conversation_context.get("verified_document"),
            "verified_cliente_codigo": conversation_context.get("verified_cliente_codigo"),
            "awaiting_verification": False,
        }
        extra_context_updates = ai_result.get("context_updates") or {}
        if extra_context_updates:
            context_updates.update(extra_context_updates)

        confidence = ai_result.get("confidence") or {}
        if confidence:
            context_updates["last_confidence"] = confidence

        if confidence and confidence.get("level") in ("baja", "media"):
            evaluate_and_create_alert(
                context["conversation_id"],
                context.get("cliente_id"),
                unified_content,
                ai_result,
                confidence,
            )

        if ai_result.get("is_farewell"):
            context_updates["conversation_closed"] = True
            context_updates["close_reason"] = "farewell_detected"
            try:
                close_conversation(
                    context["conversation_id"],
                    context_updates,
                    summary=f"Conversación cerrada por despedida del cliente. Último intent: {ai_result.get('intent')}",
                    final_status="gestionado",
                )
            except Exception:
                update_conversation_context(
                    context["conversation_id"],
                    context_updates,
                    summary=unified_content[:200],
                )
        else:
            update_conversation_context(
                context["conversation_id"],
                context_updates,
                summary=unified_content[:200],
            )

    except Exception as exc:
        logger.error("DEBOUNCE FLUSH ERROR for %s: %s", phone_number, exc, exc_info=True)


@app.post("/webhooks/whatsapp")
async def receive_whatsapp_webhook(request: Request):
    payload = await request.json()
    processed_messages = []

    for entry in payload.get("entry", []):
        for change in entry.get("changes", []):
            value = change.get("value", {})
            contacts = value.get("contacts", [])
            messages = value.get("messages", [])

            profile_name = None
            wa_id = None
            if contacts:
                contact = contacts[0]
                profile_name = contact.get("profile", {}).get("name")
                wa_id = contact.get("wa_id")

            for message in messages:
                from_number = message.get("from") or wa_id
                context = ensure_contact_and_conversation(from_number, profile_name)
                message_type = message.get("type", "text")
                if inbound_message_already_processed(message.get("id")):
                    processed_messages.append(
                        {
                            "conversation_id": context["conversation_id"],
                            "telefono": context["telefono_e164"],
                            "message_type": message_type,
                            "provider_message_id": message.get("id"),
                            "duplicate_skipped": True,
                        }
                    )
                    continue

                content = None
                media_extracted_text = None
                media_filename = None
                media_doc_type = None
                if message_type == "text":
                    content = message.get("text", {}).get("body")
                elif message_type == "button":
                    content = message.get("button", {}).get("text")
                elif message_type == "interactive":
                    content = __import__("json").dumps(message.get("interactive", {}), ensure_ascii=False)
                elif message_type in ("document", "image"):
                    # Handle document/image uploads
                    media_obj = message.get(message_type, {})
                    media_id = media_obj.get("id")
                    media_filename = media_obj.get("filename") or f"archivo.{message_type}"
                    caption = media_obj.get("caption", "")
                    if media_id:
                        try:
                            media_bytes, mime_type = download_whatsapp_media(media_id)
                            media_extracted_text, media_doc_type = extract_text_from_media(
                                media_bytes, mime_type, media_filename
                            )
                            # Build content description for the AI
                            text_preview = media_extracted_text[:1500] if media_extracted_text else "(sin texto)"
                            content = (
                                f"[DOCUMENTO RECIBIDO: {media_filename} | tipo: {media_doc_type}]\n"
                                f"{f'Mensaje adjunto: {caption}' if caption else ''}\n"
                                f"--- CONTENIDO EXTRAÍDO ---\n{text_preview}\n"
                                f"--- FIN CONTENIDO ---\n"
                                f"{'[Documento truncado, hay más contenido...]' if len(media_extracted_text or '') > 1500 else ''}"
                            )
                        except Exception as exc:
                            logger.error("Media download/extraction failed for %s: %s", media_id, exc)
                            content = (
                                f"[El cliente envió un archivo ({media_filename}) pero no pude procesarlo. "
                                f"Error: {str(exc)[:200]}. Infórmale que hubo un problema procesando su archivo.]"
                            )

                stored_content = content
                if content:
                    login_match = INTERNAL_LOGIN_PATTERN.match(content)
                    if login_match:
                        stored_content = f"login {login_match.group(1)} ******"
                    else:
                        employee_by_phone = find_employee_record_by_phone(context.get("telefono_e164"))
                        cedula_match = extract_internal_cedula_candidate(content)
                        if employee_by_phone and cedula_match:
                            stored_content = content.replace(cedula_match, f"{cedula_match[:2]}******{cedula_match[-2:]}")

                store_inbound_message(
                    context["conversation_id"],
                    message.get("id"),
                    message_type,
                    stored_content,
                    message,
                )

                recent_messages = load_recent_conversation_messages(context["conversation_id"])
                conversation_snapshot = get_conversation_snapshot(context["conversation_id"])
                conversation_context = dict(conversation_snapshot.get("contexto") or {})

                # ── Auto-reset conversation context after 3 hours of inactivity ──
                last_msg_at = conversation_snapshot.get("last_message_at")
                if last_msg_at:
                    if hasattr(last_msg_at, "tzinfo") and last_msg_at.tzinfo is not None:
                        from datetime import timezone
                        now_aware = datetime.now(timezone.utc)
                        elapsed = now_aware - last_msg_at
                    else:
                        elapsed = datetime.utcnow() - last_msg_at
                    if elapsed > timedelta(hours=3):
                        conversation_context = {}
                        update_conversation_context(
                            context["conversation_id"],
                            {
                                "verified": None,
                                "verified_document": None,
                                "verified_by": None,
                                "verified_cliente_codigo": None,
                                "awaiting_verification": None,
                                "awaiting_name_confirmation": None,
                                "pending_verified_context": None,
                                "pending_intent": None,
                                "commercial_draft": None,
                                "last_direct_intent": None,
                                "claim_case": None,
                                "pending_product_clarification": None,
                                "pending_document_options": None,
                                "last_product_request": None,
                                "technical_advisory_case": None,
                                "internal_auth": None,
                                "internal_last_cliente_codigo": None,
                                "awaiting_internal_auth_cedula": None,
                                "internal_transfer_flow": None,
                            },
                            summary="Contexto reiniciado por inactividad (3h+)",
                        )

                # ── Function Calling routing (v2) ──
                # Load client context if already verified
                cliente_contexto = None
                verified_cliente_codigo = conversation_context.get("verified_cliente_codigo")
                if verified_cliente_codigo:
                    try:
                        cliente_contexto = get_cliente_contexto(verified_cliente_codigo)
                    except HTTPException:
                        cliente_contexto = None
                if cliente_contexto is None:
                    cliente_contexto = find_cliente_contexto_by_phone(context["telefono_e164"])
                    if cliente_contexto:
                        try:
                            cliente_id = update_contact_cliente(context["contact_id"], cliente_contexto.get("cliente_codigo"))
                            context["cliente_id"] = cliente_id
                        except Exception:
                            pass

                # Generate response using LLM with function calling
                ai_result = None
                outbound_payload = None

                # ── If expert (Pablo/Diego) uploaded a document, store in context for RAG ingestion ──
                if media_extracted_text and message_type in ("document", "image"):
                    internal_auth = conversation_context.get("internal_auth") or {}
                    expert_cedula = str((internal_auth.get("user") or {}).get("cedula") or "").strip()
                    if expert_cedula in _AUTHORIZED_EXPERTS:
                        # Store full extracted text in conversation context for the tool
                        conversation_context["pending_expert_document"] = {
                            "extracted_text": media_extracted_text,
                            "filename": media_filename or "documento_experto",
                            "doc_type": media_doc_type,
                            "marca_detected": None,
                        }
                        update_conversation_context(
                            context["conversation_id"],
                            {"pending_expert_document": conversation_context["pending_expert_document"]},
                        )

                if content and message_type in {"text", "button", "interactive", "document", "image"}:
                    # ── DEBOUNCE: buffer plain text messages to concatenate rapid-fire inputs ──
                    if message_type == "text" and DEBOUNCE_WINDOW_SECONDS > 0:
                        async with _wa_buffer_lock:
                            phone_key = from_number or context.get("telefono_e164", "unknown")
                            if phone_key in _wa_message_buffer:
                                # Append to existing buffer
                                _wa_message_buffer[phone_key]["messages"].append(content)
                                # Cancel the previous timer and restart
                                old_task = _wa_message_buffer[phone_key].get("timer_task")
                                if old_task and not old_task.done():
                                    old_task.cancel()
                                # Update recent_messages to latest for best context
                                _wa_message_buffer[phone_key]["meta"][0]["recent_messages"] = recent_messages
                                _wa_message_buffer[phone_key]["meta"][0]["conversation_context"] = conversation_context
                                _wa_message_buffer[phone_key]["timer_task"] = asyncio.create_task(
                                    _flush_debounce_buffer(phone_key)
                                )
                                logger.info("DEBOUNCE BUFFER: +1 msg for %s (total: %d)", phone_key, len(_wa_message_buffer[phone_key]["messages"]))
                            else:
                                # First message — start the buffer
                                buf_meta = {
                                    "context": context,
                                    "conversation_context": conversation_context,
                                    "recent_messages": recent_messages,
                                }
                                _wa_message_buffer[phone_key] = {
                                    "messages": [content],
                                    "meta": [buf_meta],
                                    "timer_task": asyncio.create_task(
                                        _flush_debounce_buffer(phone_key)
                                    ),
                                }
                                logger.info("DEBOUNCE BUFFER: started for %s", phone_key)

                        # Store inbound but DON'T generate AI response yet — debounce will handle it
                        processed_messages.append({
                            "conversation_id": context["conversation_id"],
                            "telefono": context["telefono_e164"],
                            "message_type": message_type,
                            "provider_message_id": message.get("id"),
                            "debounce_buffered": True,
                        })
                        continue

                    # ── Non-debounced path: documents, images, buttons, interactive ──
                    watchdog_key = None
                    try:
                        if should_send_processing_ack(content, conversation_context):
                            watchdog_key = start_processing_watchdog(
                                context,
                                "⏳ Un momento, estoy procesando tu solicitud para enviártela completa.",
                            )
                        ai_result = handle_internal_whatsapp_message(content, context, conversation_context)
                        if ai_result is None:
                            ai_result = generate_agent_reply_v3(
                                context.get("nombre_visible"),
                                conversation_context,
                                recent_messages,
                                content,
                                context,
                            )
                    except Exception as exc:
                        logger.error("Agent reply FAILED for conversation %s: %s", context.get("conversation_id"), exc, exc_info=True)
                        ai_result = build_fallback_agent_result(content, str(exc))
                    finally:
                        stop_processing_watchdog(watchdog_key)

                    response_text = ai_result.get("response_text") or "Gracias por escribirnos. ¿En qué te puedo ayudar?"

                    try:
                        outbound_payload = send_whatsapp_text_message(context["telefono_e164"], response_text)
                        provider_message_id = None
                        if outbound_payload.get("messages"):
                            provider_message_id = outbound_payload["messages"][0].get("id")
                        store_outbound_message(
                            context["conversation_id"],
                            provider_message_id,
                            "text",
                            response_text,
                            outbound_payload,
                            intent_detectado=ai_result.get("intent"),
                        )
                    except Exception as exc:
                        store_outbound_message(
                            context["conversation_id"],
                            None,
                            "system",
                            f"No fue posible enviar respuesta: {exc}",
                            {"error": str(exc), "response_text": response_text},
                            intent_detectado=ai_result.get("intent"),
                        )

                    source_filename = ai_result.get("technical_source_filename") if isinstance(ai_result, dict) else None
                    if source_filename:
                        try:
                            doc_entry = find_technical_document_entry_by_name(source_filename)
                            if doc_entry:
                                _send_document_and_respond(doc_entry, context)
                        except Exception as exc:
                            store_outbound_message(
                                context["conversation_id"],
                                None,
                                "system",
                                f"No fue posible enviar ficha técnica de respaldo: {exc}",
                                {"error": str(exc), "filename": source_filename},
                                intent_detectado="consulta_documentacion",
                            )

                    # Update conversation context
                    context_updates = {
                        "intent": ai_result.get("intent"),
                        "last_direct_intent": ai_result.get("intent"),
                        "verified": conversation_context.get("verified", False),
                        "verified_document": conversation_context.get("verified_document"),
                        "verified_cliente_codigo": conversation_context.get("verified_cliente_codigo"),
                        "awaiting_verification": False,
                    }
                    extra_context_updates = ai_result.get("context_updates") or {}
                    if extra_context_updates:
                        context_updates.update(extra_context_updates)

                    # ── Confianza y alertas ──
                    confidence = ai_result.get("confidence") or {}
                    if confidence:
                        context_updates["last_confidence"] = confidence
                        logger.info(
                            "Confianza respuesta conv=%d: score=%.2f level=%s signals=%s",
                            context["conversation_id"],
                            confidence.get("score", 0),
                            confidence.get("level", "?"),
                            confidence.get("signals", []),
                        )

                    # ── Evaluar si necesita alerta para el administrador ──
                    if confidence and confidence.get("level") in ("baja", "media"):
                        evaluate_and_create_alert(
                            context["conversation_id"],
                            context.get("cliente_id"),
                            content,
                            ai_result,
                            confidence,
                        )

                    # ── Cierre automático por despedida ──
                    if ai_result.get("is_farewell"):
                        context_updates["conversation_closed"] = True
                        context_updates["close_reason"] = "farewell_detected"
                        try:
                            close_conversation(
                                context["conversation_id"],
                                context_updates,
                                summary=f"Conversación cerrada por despedida del cliente. Último intent: {ai_result.get('intent')}",
                                final_status="gestionado",
                            )
                            logger.info("Conversación %d cerrada por despedida del cliente", context["conversation_id"])
                        except Exception as exc:
                            logger.error("No se pudo cerrar conversación %d: %s", context["conversation_id"], exc)
                            update_conversation_context(
                                context["conversation_id"],
                                context_updates,
                                summary=content[:200] if content else "Mensaje procesado",
                            )
                    else:
                        update_conversation_context(
                            context["conversation_id"],
                            context_updates,
                            summary=content[:200] if content else "Mensaje procesado",
                        )

                    if ai_result.get("should_create_task"):
                        upsert_agent_task(
                            context["conversation_id"],
                            context.get("cliente_id"),
                            ai_result.get("task_type") or "seguimiento_cliente",
                            ai_result.get("task_summary") or "Revisar conversacion de WhatsApp",
                            ai_result.get("task_detail") or {"mensaje": content},
                            ai_result.get("priority") or "media",
                        )

                processed_messages.append(
                    {
                        "conversation_id": context["conversation_id"],
                        "telefono": context["telefono_e164"],
                        "message_type": message_type,
                        "provider_message_id": message.get("id"),
                        "ai_response_sent": bool(outbound_payload),
                    }
                )

    return {"status": "ok", "processed_messages": processed_messages}
