"""Capa de Herramientas — extracción de texto desde medios (PDF/Excel/Imagen/WhatsApp).

Extraído de ``backend.main`` durante la Fase C2 (Modularización), Paso 4.

Contiene:
  - Descarga de medios desde WhatsApp Cloud API
  - Extracción de texto desde PDF (PyMuPDF), Excel (openpyxl) e imágenes
    (OpenAI vision)
  - Dispatcher ``extract_text_from_media`` por MIME type

Reglas:
  - Lógica intacta (Move & Wire). Sin cambios de comportamiento.
  - ``get_whatsapp_access_token`` y ``get_openai_client`` se acceden vía lazy
    import desde ``backend.main`` para evitar ciclo y respetar configuración
    runtime del logger ``ferreinox_agent``.
  - Las funciones se re-exportan desde ``backend.main`` para preservar la API.
"""

from __future__ import annotations

import base64
import io
import logging

import requests

logger = logging.getLogger("ferreinox_agent")


def _get_whatsapp_access_token() -> str:
    try:
        from backend.main import get_whatsapp_access_token
    except ImportError:
        from main import get_whatsapp_access_token  # type: ignore
    return get_whatsapp_access_token()


def _get_openai_client():
    try:
        from backend.main import get_openai_client
    except ImportError:
        from main import get_openai_client  # type: ignore
    return get_openai_client()


def download_whatsapp_media(media_id: str) -> tuple[bytes, str]:
    """Download media bytes from WhatsApp Cloud API. Returns (bytes, mime_type)."""
    # Step 1: Get media URL
    url_resp = requests.get(
        f"https://graph.facebook.com/v22.0/{media_id}",
        headers={"Authorization": f"Bearer {_get_whatsapp_access_token()}"},
        timeout=15,
    )
    if url_resp.status_code >= 400:
        raise RuntimeError(f"WhatsApp media metadata error {url_resp.status_code}: {url_resp.text[:300]}")
    media_info = url_resp.json()
    media_url = media_info.get("url")
    mime_type = media_info.get("mime_type", "application/octet-stream")
    if not media_url:
        raise RuntimeError(f"No URL in media response: {media_info}")

    # Step 2: Download actual bytes
    dl_resp = requests.get(
        media_url,
        headers={"Authorization": f"Bearer {_get_whatsapp_access_token()}"},
        timeout=60,
    )
    if dl_resp.status_code >= 400:
        raise RuntimeError(f"WhatsApp media download error {dl_resp.status_code}")
    return dl_resp.content, mime_type


def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    """Extract text from PDF using PyMuPDF (mirrors ingest_technical_sheets.py)."""
    import fitz
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    for page in doc:
        page_parts = []
        try:
            tables = page.find_tables()
            if tables and tables.tables:
                for table in tables:
                    table_data = table.extract()
                    if table_data:
                        formatted_rows = []
                        for row in table_data:
                            clean_cells = [str(cell).strip() if cell else "" for cell in row]
                            if len(clean_cells) == 2 and clean_cells[0] and clean_cells[1]:
                                formatted_rows.append(f"{clean_cells[0]}: {clean_cells[1]}")
                            elif any(c for c in clean_cells):
                                formatted_rows.append(" | ".join(c for c in clean_cells if c))
                        if formatted_rows:
                            page_parts.append("\n".join(formatted_rows))
        except Exception:
            pass
        text_content = page.get_text("text")
        if text_content and text_content.strip():
            if page_parts:
                blocks = page.get_text("blocks")
                non_table_text = []
                for block in blocks:
                    if block[6] == 0:
                        block_text = block[4].strip()
                        if block_text:
                            non_table_text.append(block_text)
                if non_table_text:
                    page_parts.insert(0, "\n".join(non_table_text))
                pages.append("\n\n".join(page_parts))
            else:
                pages.append(text_content.strip())
    doc.close()
    return "\n\n".join(pages)


def extract_text_from_excel_bytes(excel_bytes: bytes) -> str:
    """Extract text from Excel file (all sheets) using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows_text.append(" | ".join(c for c in cells if c))
        if rows_text:
            parts.append(f"[HOJA: {sheet_name}]\n" + "\n".join(rows_text))
    wb.close()
    return "\n\n".join(parts)


def extract_text_from_image_bytes(image_bytes: bytes, mime_type: str) -> str:
    """Use OpenAI GPT-4o vision to extract text/content from an image."""
    try:
        b64 = base64.b64encode(image_bytes).decode("utf-8")
        client = _get_openai_client()
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": (
                                "Extrae TODO el texto visible en esta imagen. "
                                "Si es una ficha técnica, tabla de datos, cálculo o documento técnico, "
                                "transcríbelo completo preservando la estructura (tablas, listas, secciones). "
                                "Si es una foto de un producto o superficie, describe exactamente lo que ves. "
                                "Responde SOLO con el contenido extraído, sin comentarios adicionales."
                            ),
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{mime_type};base64,{b64}"},
                        },
                    ],
                }
            ],
            max_tokens=4000,
            temperature=0,
        )
        return response.choices[0].message.content.strip()
    except Exception as exc:
        logger.warning("Image text extraction failed: %s", exc)
        return f"[Error extrayendo texto de imagen: {exc}]"


def extract_text_from_media(media_bytes: bytes, mime_type: str, filename: str = "") -> tuple[str, str]:
    """Extract text from media based on MIME type. Returns (text, doc_type)."""
    mime_lower = (mime_type or "").lower()
    fname_lower = (filename or "").lower()

    if "pdf" in mime_lower or fname_lower.endswith(".pdf"):
        return extract_text_from_pdf_bytes(media_bytes), "pdf"
    elif any(x in mime_lower for x in ["spreadsheet", "excel", "xlsx", "xls"]) or \
         fname_lower.endswith((".xlsx", ".xls")):
        return extract_text_from_excel_bytes(media_bytes), "excel"
    elif any(x in mime_lower for x in ["image/", "png", "jpeg", "jpg", "webp", "gif"]):
        return extract_text_from_image_bytes(media_bytes, mime_type), "image"
    elif "text" in mime_lower or fname_lower.endswith((".txt", ".csv", ".md")):
        try:
            return media_bytes.decode("utf-8"), "text"
        except UnicodeDecodeError:
            return media_bytes.decode("latin-1"), "text"
    else:
        # Try as text first, fall back to binary description
        try:
            text = media_bytes.decode("utf-8")
            if text.strip():
                return text, "text"
        except UnicodeDecodeError:
            pass
        return f"[Archivo binario no soportado: {mime_type}, {len(media_bytes)} bytes]", "unsupported"


__all__ = [
    "download_whatsapp_media",
    "extract_text_from_pdf_bytes",
    "extract_text_from_excel_bytes",
    "extract_text_from_image_bytes",
    "extract_text_from_media",
]
