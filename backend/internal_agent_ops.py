import base64
import calendar
import io
import json
import logging
from numbers import Number
import re
from datetime import date, datetime, timedelta
from html import escape
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError


logger = logging.getLogger("internal_agent_ops")

_XL_DARK_FILL = PatternFill("solid", fgColor="111827")
_XL_ACCENT_FILL = PatternFill("solid", fgColor="F59E0B")
_XL_LIGHT_FILL = PatternFill("solid", fgColor="F9FAFB")
_XL_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


_ROUTINE_ALIAS_MAP = {
    "/rutina_diaria_gerencia": "rutina_diaria_gerencia",
    "/rutina_cartera": "rutina_cartera",
    "/rutina_bodega": "rutina_bodega",
    "/rutina_compras": "rutina_compras",
    "/rutina_comercial": "rutina_comercial",
    "rutina gerencia": "rutina_diaria_gerencia",
    "cierre gerencia": "rutina_diaria_gerencia",
    "rutina cartera": "rutina_cartera",
    "seguimiento cartera": "rutina_cartera",
    "rutina bodega": "rutina_bodega",
    "reposicion bodega": "rutina_bodega",
    "rutina compras": "rutina_compras",
    "rutina comercial": "rutina_comercial",
}

_ROLE_FALLBACKS = {
    "administrador": ("administracion", "Administracion", "control"),
    "gerente": ("gerencia_general", "Gerencia General", "ejecutivo"),
    "operador": ("bodega", "Bodega", "operativo"),
    "vendedor": ("comercial", "Comercial", "ventas"),
    "empleado": ("empleado_operativo", "Empleado Operativo", "soporte"),
}

_MESES_NOMBRES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

_STORE_ALIAS_TO_CODE = {
    "armenia": "156",
    "san francisco": "156",
    "manizales": "157",
    "san antonio": "157",
    "opalo": "158",
    "ópalo": "158",
    "dosquebradas": "158",
    "pereira": "189",
    "parque olaya": "189",
    "olaya": "189",
    "laureles": "238",
    "laures": "238",
    "ferrebox": "439",
    "ferre box": "439",
    "cerritos": "463",
}

_STORE_CODE_LABELS = {
    "156": "Armenia",
    "157": "Manizales",
    "158": "Opalo",
    "189": "Pereira",
    "238": "Laureles",
    "439": "FerreBOX",
    "463": "Cerritos",
}

_UNIVERSAL_BI_DIMENSIONS = {
    "tienda": ["tienda", "sede", "almacen", "almacén"],
    "vendedor": ["vendedor", "asesor", "comercial"],
    "cliente": ["cliente", "clientes"],
    "producto": ["producto", "productos", "referencia", "referencias"],
    "linea": ["linea", "línea", "categoria", "categoría", "familia"],
    "zona": ["zona", "regional", "region", "región"],
}


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _format_currency(value: Any) -> str:
    try:
        numeric = float(value or 0)
    except (TypeError, ValueError):
        numeric = 0.0
    return f"${numeric:,.0f}"


def _format_number(value: Any) -> str:
    try:
        numeric = float(value or 0)
    except (TypeError, ValueError):
        numeric = 0.0
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:,.2f}"


def _format_percent(value: Any) -> str:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return "N/D"
    return f"{numeric:,.1f}%"


def _parse_bi_period(periodo_raw: Optional[str]) -> tuple[date, date, str]:
    normalized = _normalize_text(periodo_raw or "")
    today = date.today()
    if "hoy" in normalized:
        return today, today, "hoy"
    if "ayer" in normalized:
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday, "ayer"
    if "esta semana" in normalized or "semana actual" in normalized:
        start = today - timedelta(days=today.weekday())
        return start, today, "esta semana"
    if "semana pasada" in normalized:
        end = today - timedelta(days=today.weekday() + 1)
        start = end - timedelta(days=6)
        return start, end, "semana pasada"
    if not normalized or "este ano" in normalized or "este año" in normalized or "ano actual" in normalized or "año actual" in normalized:
        return date(today.year, 1, 1), today, "este año"
    if "este mes" in normalized or "mes actual" in normalized:
        return date(today.year, today.month, 1), today, "este mes"
    if "mes pasado" in normalized:
        ref = date(today.year - 1, 12, 1) if today.month == 1 else date(today.year, today.month - 1, 1)
        end = date(ref.year, ref.month, calendar.monthrange(ref.year, ref.month)[1])
        return ref, min(end, today), f"{list(_MESES_NOMBRES.keys())[ref.month - 1]} {ref.year}"
    for month_name, month_num in _MESES_NOMBRES.items():
        if month_name in normalized:
            year_match = re.search(r"\b(20\d{2})\b", normalized)
            year_value = int(year_match.group(1)) if year_match else today.year
            start = date(year_value, month_num, 1)
            end = date(year_value, month_num, calendar.monthrange(year_value, month_num)[1])
            return start, min(end, today), f"{month_name} {year_value}"
    return date(today.year, 1, 1), today, "este año"


def _format_preview_value(key: str, value: Any) -> str:
    if value is None:
        return ""
    lowered = str(key or "").lower()
    if isinstance(value, Number):
        if any(token in lowered for token in ["valor", "saldo", "neto", "facturado", "devoluciones", "total", "precio", "ventas", "promedio", "brecha"]):
            return _format_currency(value)
        if any(token in lowered for token in ["variacion", "participacion"]):
            return _format_percent(value)
        return _format_number(value)
    return str(value)


def detect_requested_routine(user_message: Optional[str]) -> Optional[str]:
    normalized = _normalize_text(user_message)
    if not normalized:
        return None
    for alias, routine_key in _ROUTINE_ALIAS_MAP.items():
        if alias in normalized:
            return routine_key
    return None


def _infer_role_profile_from_auth(internal_auth: Optional[dict]) -> dict:
    internal_auth = internal_auth or {}
    employee_context = internal_auth.get("employee_context") or {}
    cargo = _normalize_text(employee_context.get("cargo"))
    base_role = _normalize_text(internal_auth.get("role")) or "empleado"

    if any(token in cargo for token in ("geren", "direccion")):
        role_key, display_name, prompt_mode = ("gerencia_general", "Gerencia General", "ejecutivo")
    elif "cartera" in cargo:
        role_key, display_name, prompt_mode = ("cartera", "Cartera", "cobranza")
    elif any(token in cargo for token in ("compra", "abastecimiento")):
        role_key, display_name, prompt_mode = ("compras", "Compras", "abastecimiento")
    elif any(token in cargo for token in ("bodega", "almacen", "despacho")):
        role_key, display_name, prompt_mode = ("bodega", "Bodega", "operativo")
    elif any(token in cargo for token in ("admin", "contab", "tesorer")):
        role_key, display_name, prompt_mode = ("administracion", "Administracion", "control")
    else:
        role_key, display_name, prompt_mode = _ROLE_FALLBACKS.get(
            base_role,
            ("empleado_operativo", "Empleado Operativo", "soporte"),
        )

    return {
        "role_key": role_key,
        "display_name": display_name,
        "prompt_mode": prompt_mode,
        "base_role": base_role or "empleado",
        "priority_focus": [],
        "allowed_kpis": [],
        "allowed_tools": [],
        "guidance_template": None,
    }


def _fetch_role_profile(engine, internal_auth: Optional[dict]) -> dict:
    fallback = _infer_role_profile_from_auth(internal_auth)
    if not engine or not (internal_auth or {}).get("user_id"):
        return fallback

    try:
        with engine.begin() as connection:
            row = connection.execute(
                text(
                    """
                    SELECT
                        rp.role_key,
                        rp.display_name,
                        rp.base_role,
                        rp.prompt_mode,
                        rp.priority_focus,
                        rp.allowed_kpis,
                        rp.allowed_tools,
                        rp.guidance_template
                    FROM public.agent_user_role_profile urp
                    JOIN public.agent_role_profile rp ON rp.role_key = urp.role_key
                    WHERE urp.user_id = :user_id
                      AND urp.is_active = true
                    ORDER BY urp.updated_at DESC
                    LIMIT 1
                    """
                ),
                {"user_id": internal_auth.get("user_id")},
            ).mappings().one_or_none()
    except SQLAlchemyError as exc:
        logger.debug("role profile lookup unavailable: %s", exc)
        row = None

    if not row:
        return fallback

    return {
        "role_key": row.get("role_key") or fallback["role_key"],
        "display_name": row.get("display_name") or fallback["display_name"],
        "prompt_mode": row.get("prompt_mode") or fallback["prompt_mode"],
        "base_role": row.get("base_role") or fallback["base_role"],
        "priority_focus": list(row.get("priority_focus") or []),
        "allowed_kpis": list(row.get("allowed_kpis") or []),
        "allowed_tools": list(row.get("allowed_tools") or []),
        "guidance_template": row.get("guidance_template"),
    }


def _fetch_routine_definition(engine, routine_key: Optional[str]) -> Optional[dict]:
    if not routine_key:
        return None
    if not engine:
        return {"routine_key": routine_key, "display_name": routine_key.replace("_", " ").title()}

    try:
        with engine.begin() as connection:
            row = connection.execute(
                text(
                    """
                    SELECT routine_key, display_name, command_token, role_key, description, prompt_hint, default_filters
                    FROM public.agent_internal_routine
                    WHERE routine_key = :routine_key
                      AND is_active = true
                    LIMIT 1
                    """
                ),
                {"routine_key": routine_key},
            ).mappings().one_or_none()
            return dict(row) if row else None
    except SQLAlchemyError as exc:
        logger.debug("routine definition unavailable: %s", exc)
        return None


def _fetch_alerts(engine, role_key: str, store_code: Optional[str], limit: int = 3) -> list[dict]:
    if not engine:
        return []

    store_code = re.sub(r"\D", "", str(store_code or "")) or None
    try:
        with engine.begin() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT alert_type, severity, title, detail, payload
                    FROM public.vw_internal_alert_candidates
                    WHERE assigned_role_key = :role_key
                       OR :role_key = 'gerencia_general'
                       OR (:store_code IS NOT NULL AND COALESCE(payload->>'cod_almacen', '') = :store_code)
                    ORDER BY CASE severity WHEN 'alta' THEN 1 WHEN 'media' THEN 2 ELSE 3 END,
                             title ASC
                    LIMIT :limit
                    """
                ),
                {"role_key": role_key, "store_code": store_code, "limit": limit},
            ).mappings().all()
            return [dict(row) for row in rows]
    except SQLAlchemyError as exc:
        logger.debug("alert snapshot unavailable: %s", exc)
        return []


def _fetch_pending_items(engine, user_id: Optional[int], role_key: str, limit: int = 5) -> list[dict]:
    if not engine:
        return []

    try:
        with engine.begin() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT item_type, item_id, title, summary, due_at, priority, owner_user_id, role_key
                    FROM public.vw_internal_pending_queue
                    WHERE (:user_id IS NOT NULL AND owner_user_id = :user_id)
                       OR role_key = :role_key
                    ORDER BY priority_rank ASC, due_at NULLS LAST, created_at DESC
                    LIMIT :limit
                    """
                ),
                {"user_id": user_id, "role_key": role_key, "limit": limit},
            ).mappings().all()
            return [dict(row) for row in rows]
    except SQLAlchemyError as exc:
        logger.debug("pending items unavailable: %s", exc)
        return []


def _fetch_role_kpi_lines(engine, role_key: str, store_code: Optional[str], routine_key: Optional[str]) -> list[str]:
    if not engine:
        return []

    lines: list[str] = []
    params = {"store_code": re.sub(r"\D", "", str(store_code or "")) or None}
    try:
        with engine.begin() as connection:
            if role_key == "gerencia_general" or routine_key == "rutina_diaria_gerencia":
                ventas = connection.execute(
                    text(
                        """
                        SELECT
                            COALESCE(SUM(CASE WHEN sales_date = CURRENT_DATE THEN net_sales END), 0) AS ventas_hoy,
                            COALESCE(SUM(CASE WHEN sales_date >= date_trunc('month', CURRENT_DATE)::date THEN net_sales END), 0) AS ventas_mes
                        FROM public.mv_internal_sales_daily
                        """
                    )
                ).mappings().one()
                cartera = connection.execute(
                    text(
                        """
                        SELECT
                            COALESCE(SUM(balance_total), 0) AS cartera_total,
                            COALESCE(SUM(balance_61_90 + balance_91_plus), 0) AS cartera_vencida
                        FROM public.mv_internal_cartera_cliente
                        """
                    )
                ).mappings().one()
                inventario = connection.execute(
                    text(
                        """
                        SELECT
                            COUNT(*) FILTER (WHERE health_status = 'quiebre_critico') AS quiebres,
                            COUNT(*) FILTER (WHERE health_status = 'sobrestock') AS sobrestock
                        FROM public.mv_internal_inventory_health
                        """
                    )
                ).mappings().one()
                lines.append(
                    f"KPI rapido: ventas hoy {_format_currency(ventas.get('ventas_hoy'))}, ventas mes {_format_currency(ventas.get('ventas_mes'))}."
                )
                if float(ventas.get("ventas_mes") or 0) > 0:
                    today = date.today()
                    proyeccion_mes = (float(ventas.get("ventas_mes") or 0) / max(today.day, 1)) * calendar.monthrange(today.year, today.month)[1]
                    lines.append(f"KPI rapido: proyección cierre mes {_format_currency(proyeccion_mes)}.")
                lines.append(
                    f"KPI rapido: cartera vencida {_format_currency(cartera.get('cartera_vencida'))} sobre total {_format_currency(cartera.get('cartera_total'))}."
                )
                lines.append(
                    f"KPI rapido: quiebres criticos {inventario.get('quiebres') or 0}, sobrestock {inventario.get('sobrestock') or 0}."
                )

            elif role_key in {"cartera", "administracion"} or routine_key == "rutina_cartera":
                cartera = connection.execute(
                    text(
                        """
                        SELECT
                            COUNT(*) FILTER (WHERE (balance_61_90 + balance_91_plus) > 0) AS clientes_vencidos,
                            COALESCE(SUM(balance_61_90), 0) AS bucket_61_90,
                            COALESCE(SUM(balance_91_plus), 0) AS bucket_91_plus
                        FROM public.mv_internal_cartera_cliente
                        """
                    )
                ).mappings().one()
                lines.append(
                    f"KPI rapido: clientes vencidos {cartera.get('clientes_vencidos') or 0}, 61-90 {_format_currency(cartera.get('bucket_61_90'))}, >90 {_format_currency(cartera.get('bucket_91_plus'))}."
                )

            elif role_key in {"bodega", "compras"} or routine_key in {"rutina_bodega", "rutina_compras"}:
                inventario = connection.execute(
                    text(
                        """
                        SELECT
                            COUNT(*) FILTER (WHERE health_status = 'quiebre_critico') AS quiebres,
                            COUNT(*) FILTER (WHERE health_status = 'reposicion_recomendada') AS reposicion,
                            COUNT(*) FILTER (WHERE health_status = 'sobrestock') AS sobrestock,
                            COUNT(*) FILTER (WHERE health_status = 'sin_movimiento') AS sin_movimiento
                        FROM public.mv_internal_inventory_health
                        WHERE :store_code IS NULL OR cod_almacen = :store_code
                        """
                    ),
                    params,
                ).mappings().one()
                lines.append(
                    f"KPI rapido: quiebres {inventario.get('quiebres') or 0}, reposicion {inventario.get('reposicion') or 0}, sobrestock {inventario.get('sobrestock') or 0}, sin movimiento {inventario.get('sin_movimiento') or 0}."
                )

            elif role_key == "comercial" or routine_key == "rutina_comercial":
                ventas = connection.execute(
                    text(
                        """
                        SELECT
                            COALESCE(SUM(CASE WHEN sales_date = CURRENT_DATE THEN net_sales END), 0) AS ventas_hoy,
                            COALESCE(SUM(CASE WHEN sales_date >= date_trunc('month', CURRENT_DATE)::date THEN net_sales END), 0) AS ventas_mes,
                            COUNT(DISTINCT CASE WHEN sales_date >= date_trunc('month', CURRENT_DATE)::date THEN codigo_vendedor END) AS vendedores_activos
                        FROM public.mv_internal_sales_daily
                        WHERE :store_code IS NULL OR regexp_replace(COALESCE(serie, ''), '[^0-9]', '', 'g') = :store_code
                        """
                    ),
                    params,
                ).mappings().one()
                lines.append(
                    f"KPI rapido: ventas hoy {_format_currency(ventas.get('ventas_hoy'))}, ventas mes {_format_currency(ventas.get('ventas_mes'))}, vendedores activos {ventas.get('vendedores_activos') or 0}."
                )
                if float(ventas.get("ventas_mes") or 0) > 0:
                    today = date.today()
                    proyeccion_mes = (float(ventas.get("ventas_mes") or 0) / max(today.day, 1)) * calendar.monthrange(today.year, today.month)[1]
                    lines.append(f"KPI rapido: proyección cierre mes {_format_currency(proyeccion_mes)}.")
    except SQLAlchemyError as exc:
        logger.debug("role KPI snapshot unavailable: %s", exc)

    return lines


def build_internal_operational_context(engine, internal_auth: Optional[dict], user_message: Optional[str], conversation_context: Optional[dict]) -> str:
    internal_auth = internal_auth or {}
    role_profile = _fetch_role_profile(engine, internal_auth)
    routine_key = detect_requested_routine(user_message)
    routine = _fetch_routine_definition(engine, routine_key)
    employee_context = internal_auth.get("employee_context") or {}

    lines = []
    lines.append("")
    lines.append("═══ CAPA OPERATIVA INTERNA ═══")
    lines.append(
        f"Perfil operativo activo: {role_profile.get('display_name', 'Empleado Operativo')} ({role_profile.get('role_key', 'empleado_operativo')})."
    )
    lines.append(f"Modo de respuesta esperado: {role_profile.get('prompt_mode', 'soporte')}.")
    if role_profile.get("guidance_template"):
        lines.append(f"Directriz: {role_profile['guidance_template']}")
    if routine:
        lines.append(f"Rutina detectada: {routine.get('display_name')} ({routine.get('command_token')}).")
        if routine.get("prompt_hint"):
            lines.append(f"Objetivo de la rutina: {routine.get('prompt_hint')}")

    lines.extend(
        _fetch_role_kpi_lines(
            engine,
            role_profile.get("role_key") or "empleado_operativo",
            employee_context.get("store_code"),
            routine_key,
        )
    )

    alerts = _fetch_alerts(engine, role_profile.get("role_key") or "empleado_operativo", employee_context.get("store_code"))
    if alerts:
        lines.append("Alertas activas:")
        for alert in alerts[:3]:
            lines.append(f"  • [{str(alert.get('severity') or '').upper()}] {alert.get('title')}: {alert.get('detail')}")

    lines.append("Si el pedido del colaborador es gerencial o administrativo, prioriza KPIs, alertas y accion concreta.")
    return "\n".join(lines)


def _clamp_limit(raw_value: Any, default: int, minimum: int, maximum: int) -> int:
    try:
        value = int(raw_value or default)
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(value, maximum))


def _resolve_store_code(raw_value: Any) -> Optional[str]:
    normalized = re.sub(r"\D", "", str(raw_value or "")).strip()
    return normalized or None


def _resolve_vendor_code(raw_value: Any, internal_auth: Optional[dict]) -> Optional[str]:
    internal_auth = internal_auth or {}
    employee_context = internal_auth.get("employee_context") or {}
    role = _normalize_text(internal_auth.get("role") or "")
    candidate = raw_value
    if role == "vendedor":
        candidate = employee_context.get("codigo_vendedor") or raw_value
    normalized = re.sub(r"[^A-Za-z0-9]", "", str(candidate or "")).strip()
    return normalized or None


def _extract_store_code_from_question(question: str) -> Optional[str]:
    normalized = _normalize_text(question)
    direct_code = re.search(r"\b(156|157|158|189|238|439|463)\b", normalized)
    if direct_code:
        return direct_code.group(1)
    for alias, code in _STORE_ALIAS_TO_CODE.items():
        if alias in normalized:
            return code
    return None


def _extract_vendor_code_from_question(question: str) -> Optional[str]:
    normalized = _normalize_text(question)
    match = re.search(r"\b(\d{2,4}[\.\-]?\d{2,4})\b", normalized)
    if not match:
        return None
    return re.sub(r"[^A-Za-z0-9]", "", match.group(1))


def _extract_limit_from_question(question: str, default: int, minimum: int, maximum: int) -> int:
    normalized = _normalize_text(question)
    patterns = [
        r"(?:top|primeros|primeras|ultimos|ultimas|listado de|dame|muestrame|muéstrame)\s+(\d{1,3})\b",
        r"\b(\d{1,3})\s+(?:productos|clientes|vendedores|lineas|líneas|sedes|tiendas|referencias)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized)
        if match:
            return _clamp_limit(int(match.group(1)), default=default, minimum=minimum, maximum=maximum)
    return default


def _infer_sales_dimension(question: str) -> Optional[str]:
    normalized = _normalize_text(question)
    for dimension, aliases in _UNIVERSAL_BI_DIMENSIONS.items():
        if any(alias in normalized for alias in aliases):
            return dimension
    return None


def _infer_sort_direction(question: str) -> str:
    normalized = _normalize_text(question)
    if any(token in normalized for token in ["menor", "menores", "peor", "peores", "menos", "bajo", "bajos"]):
        return "asc"
    return "desc"


def _infer_comparison_mode(question: str) -> str:
    normalized = _normalize_text(question)
    if any(token in normalized for token in ["vs mes pasado", "versus mes pasado", "comparado con mes pasado", "contra mes pasado", "periodo anterior", "período anterior"]):
        return "vs_periodo_anterior"
    return "vs_anio_anterior"


def _resolve_semantic_dimension(question: str, fallback: str) -> str:
    return _infer_sales_dimension(question) or fallback


def _infer_universal_bi_plan(question: str, explicit_period: Optional[str], explicit_limit: Any) -> dict:
    normalized = _normalize_text(question)
    period_value = explicit_period or question or "este mes"
    limit = _clamp_limit(explicit_limit, default=_extract_limit_from_question(question, default=10, minimum=3, maximum=50), minimum=3, maximum=50)

    if any(token in normalized for token in ["participacion", "participación", "mix", "% del total", "porcentaje del total"]):
        return {
            "kind": "semantic",
            "analysis": "participacion",
            "periodo": period_value,
            "dimension": _resolve_semantic_dimension(question, "linea"),
            "direction": _infer_sort_direction(question),
            "limite": limit,
        }
    if any(token in normalized for token in ["crecimiento", "crecer", "creció", "crecio", "creciendo", "crecen", "crecieron", "variacion por", "variación por", "cayendo", "cayeron", "vienen creciendo", "vienen cayendo"]):
        detected_direction = _infer_sort_direction(question)
        if any(w in normalized for w in ["cayendo", "cayeron", "caen"]):
            detected_direction = "asc"
        return {
            "kind": "semantic",
            "analysis": "crecimiento",
            "periodo": period_value,
            "dimension": _resolve_semantic_dimension(question, "linea"),
            "direction": detected_direction,
            "comparison": _infer_comparison_mode(question),
            "limite": limit,
        }
    if any(token in normalized for token in ["caida de frecuencia", "caída de frecuencia", "frecuencia de compra", "menos frecuencia", "menos visitas", "menos compras", "frecuencia"]):
        return {
            "kind": "semantic",
            "analysis": "caida_frecuencia",
            "periodo": period_value,
            "direction": "asc",
            "comparison": _infer_comparison_mode(question),
            "limite": limit,
        }
    if any(token in normalized for token in ["concentracion de cartera", "concentración de cartera", "cartera concentrada", "concentracion cartera", "concentración cartera", "concentracion de la cartera", "concentración de la cartera"]):
        return {
            "kind": "semantic",
            "analysis": "concentracion_cartera",
            "periodo": period_value,
            "dimension": _resolve_semantic_dimension(question, "cliente"),
            "direction": "desc",
            "limite": limit,
        }
    if any(token in normalized for token in ["oportunidades por sede", "oportunidades por vendedor", "oportunidad por sede", "oportunidad por vendedor", "oportunidades", "donde estan las oportunidades", "dónde están las oportunidades"]):
        fallback_dimension = "vendedor" if "vendedor" in normalized else "tienda"
        return {
            "kind": "semantic",
            "analysis": "oportunidades_dimension",
            "periodo": period_value,
            "dimension": _resolve_semantic_dimension(question, fallback_dimension),
            "direction": "desc",
            "comparison": _infer_comparison_mode(question),
            "limite": limit,
        }

    if any(token in normalized for token in ["reactivar", "visitar", "volver a comprar", "recuperar cliente"]):
        return {"kind": "indicator", "tipo_consulta": "clientes_a_reactivar", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["no me han comprado", "no compraron", "sin compra", "no han comprado"]):
        return {"kind": "indicator", "tipo_consulta": "clientes_sin_compra_periodo", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["no he vendido", "no se ha vendido", "sin venta", "dejaron de venderse"]):
        return {"kind": "indicator", "tipo_consulta": "productos_no_vendidos_periodo", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["impulsar", "mover", "oportunidad", "reactivar producto"]):
        return {"kind": "indicator", "tipo_consulta": "productos_a_impulsar", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["decrecimiento", "caida", "caída", "decrece"]):
        return {"kind": "indicator", "tipo_consulta": "clientes_mayor_decrecimiento", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["cartera", "vencido", "vencida", "saldo"]):
        return {"kind": "indicator", "tipo_consulta": "cartera_vencida_resumen", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["quiebre", "agotado", "sin stock"]):
        return {"kind": "indicator", "tipo_consulta": "quiebres_stock", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["sobrestock", "sobre stock", "exceso inventario"]):
        return {"kind": "indicator", "tipo_consulta": "sobrestock", "periodo": period_value, "limite": limit}
    if any(token in normalized for token in ["baja rotacion", "baja rotación", "quedado", "sin movimiento"]):
        return {"kind": "indicator", "tipo_consulta": "inventario_baja_rotacion", "periodo": period_value, "limite": limit}
    if "proyeccion" in normalized or "proyección" in normalized:
        return {"kind": "indicator", "tipo_consulta": "proyeccion_ventas_mes", "periodo": period_value, "limite": limit}

    return {
        "kind": "sales",
        "periodo": period_value,
        "dimension": _infer_sales_dimension(question),
        "direction": _infer_sort_direction(question),
        "limite": limit,
    }


def _is_valid_email(value: Optional[str]) -> bool:
    if not value:
        return False
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", value.strip()))


def _fetch_sales_projection(engine, store_code: Optional[str]) -> dict:
    today = date.today()
    days_in_month = calendar.monthrange(today.year, today.month)[1]
    params = {"start_month": today.replace(day=1), "today": today, "store_code": store_code}
    with engine.begin() as connection:
        row = connection.execute(
            text(
                """
                SELECT
                    COALESCE(SUM(net_sales), 0) AS ventas_mes_actual,
                    COUNT(DISTINCT sales_date) AS dias_con_venta,
                    COALESCE(AVG(NULLIF(net_sales, 0)), 0) AS promedio_dia_con_venta
                FROM public.mv_internal_sales_daily
                WHERE sales_date BETWEEN :start_month AND :today
                  AND (:store_code IS NULL OR regexp_replace(COALESCE(serie, ''), '[^0-9]', '', 'g') = :store_code)
                """
            ),
            params,
        ).mappings().one()

        prev_start = (today.replace(day=1).replace(month=today.month - 1) if today.month > 1 else date(today.year - 1, 12, 1))
        prev_end = prev_start.replace(day=min(today.day, calendar.monthrange(prev_start.year, prev_start.month)[1]))
        prev_row = connection.execute(
            text(
                """
                SELECT COALESCE(SUM(net_sales), 0) AS ventas_mes_anterior_mismo_corte
                FROM public.mv_internal_sales_daily
                WHERE sales_date BETWEEN :prev_start AND :prev_end
                  AND (:store_code IS NULL OR regexp_replace(COALESCE(serie, ''), '[^0-9]', '', 'g') = :store_code)
                """
            ),
            {"prev_start": prev_start, "prev_end": prev_end, "store_code": store_code},
        ).mappings().one()

    ventas_mes = float(row.get("ventas_mes_actual") or 0)
    elapsed_days = max(today.day, 1)
    proyeccion = (ventas_mes / elapsed_days) * days_in_month if ventas_mes > 0 else 0
    ventas_prev = float(prev_row.get("ventas_mes_anterior_mismo_corte") or 0)
    variacion = ((ventas_mes - ventas_prev) / ventas_prev * 100.0) if ventas_prev > 0 else None
    return {
        "ventas_mes_actual": ventas_mes,
        "dias_transcurridos": today.day,
        "dias_mes": days_in_month,
        "proyeccion_cierre_mes": proyeccion,
        "promedio_dia_con_venta": float(row.get("promedio_dia_con_venta") or 0),
        "ventas_mes_anterior_mismo_corte": ventas_prev,
        "variacion_pct": round(variacion, 1) if variacion is not None else None,
    }


def _fetch_inventory_rows(engine, health_statuses: list[str], store_code: Optional[str], limit: int) -> list[dict]:
    query_params = {"statuses": health_statuses, "store_code": store_code, "limit": limit}

    primary_sql = text(
        """
        WITH last_sales AS (
            SELECT
                am.referencia_normalizada,
                LEFT(COALESCE(rv.serie, ''), 3) AS store_prefix,
                MAX(public.fn_parse_date(rv.fecha_venta)) AS last_sale_date
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE UPPER(COALESCE(rv.tipo_documento, '')) LIKE '%FACTURA%'
            GROUP BY am.referencia_normalizada, LEFT(COALESCE(rv.serie, ''), 3)
        )
        SELECT
            inv.cod_almacen,
            inv.almacen_nombre,
            inv.referencia,
            inv.descripcion,
            inv.stock_total,
            inv.historial_ventas_metric,
            inv.reorder_point,
            inv.reorder_qty_recommended,
            inv.inventory_value,
            inv.health_status,
            ls.last_sale_date AS fecha_ultima_venta,
            CASE
                WHEN ls.last_sale_date IS NULL THEN NULL
                ELSE (CURRENT_DATE - ls.last_sale_date)
            END AS dias_sin_venta
        FROM public.mv_internal_inventory_health
        LEFT JOIN last_sales ls
          ON ls.referencia_normalizada = inv.referencia_normalizada
         AND ls.store_prefix = inv.cod_almacen
        WHERE inv.health_status = ANY(:statuses)
          AND (:store_code IS NULL OR inv.cod_almacen = :store_code)
        ORDER BY inv.inventory_value DESC, inv.reorder_qty_recommended DESC, inv.historial_ventas_metric ASC
        LIMIT :limit
        """
    )

    fallback_sql = text(
        """
        WITH last_sales AS (
            SELECT
                am.referencia_normalizada,
                LEFT(COALESCE(rv.serie, ''), 3) AS store_prefix,
                MAX(public.fn_parse_date(rv.fecha_venta)) AS last_sale_date
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE UPPER(COALESCE(rv.tipo_documento, '')) LIKE '%FACTURA%'
            GROUP BY am.referencia_normalizada, LEFT(COALESCE(rv.serie, ''), 3)
        ),
        base AS (
            SELECT
                cod_almacen,
                almacen_nombre,
                referencia_normalizada,
                MAX(referencia) AS referencia,
                MAX(descripcion) AS descripcion,
                COALESCE(SUM(stock_disponible), 0) AS stock_total,
                COALESCE(AVG(costo_promedio_und), 0) AS costo_promedio_und,
                COALESCE(SUM(unidades_vendidas), 0) AS unidades_vendidas_total,
                COALESCE(MAX(historial_ventas), 0) AS historial_ventas_metric,
                COALESCE(MAX(lead_time_proveedor), 0) AS lead_time_proveedor_dias
            FROM public.vw_inventario_agente
            GROUP BY cod_almacen, almacen_nombre, referencia_normalizada
        ),
        health AS (
            SELECT
                cod_almacen,
                almacen_nombre,
                referencia,
                descripcion,
                stock_total,
                historial_ventas_metric,
                stock_total * costo_promedio_und AS inventory_value,
                GREATEST(
                    CEIL(
                        ((GREATEST(historial_ventas_metric, 0) / 30.0) * GREATEST(NULLIF(lead_time_proveedor_dias, 0), 7))
                        + GREATEST(historial_ventas_metric * 0.25, 1)
                    ),
                    1
                )::numeric(18,2) AS reorder_point,
                GREATEST(
                    CEIL(
                        (((GREATEST(historial_ventas_metric, 0) / 30.0) * GREATEST(NULLIF(lead_time_proveedor_dias, 0), 7))
                        + GREATEST(historial_ventas_metric * 0.25, 1)) - stock_total
                    ),
                    0
                )::numeric(18,2) AS reorder_qty_recommended,
                CASE
                    WHEN stock_total <= 0 AND GREATEST(historial_ventas_metric, 0) > 0 THEN 'quiebre_critico'
                    WHEN stock_total < GREATEST(
                        CEIL(
                            ((GREATEST(historial_ventas_metric, 0) / 30.0) * GREATEST(NULLIF(lead_time_proveedor_dias, 0), 7))
                            + GREATEST(historial_ventas_metric * 0.25, 1)
                        ),
                        1
                    ) THEN 'reposicion_recomendada'
                    WHEN stock_total > GREATEST(historial_ventas_metric * 4, 12) AND GREATEST(historial_ventas_metric, 0) > 0 THEN 'sobrestock'
                    WHEN stock_total > 0 AND GREATEST(historial_ventas_metric, 0) = 0 AND GREATEST(unidades_vendidas_total, 0) = 0 THEN 'sin_movimiento'
                    ELSE 'saludable'
                END AS health_status
            FROM base
        )
        SELECT
                        h.cod_almacen,
                        h.almacen_nombre,
                        h.referencia,
                        h.descripcion,
                        h.stock_total,
                        h.historial_ventas_metric,
                        h.reorder_point,
                        h.reorder_qty_recommended,
                        h.inventory_value,
                        h.health_status,
                        ls.last_sale_date AS fecha_ultima_venta,
                        CASE
                                WHEN ls.last_sale_date IS NULL THEN NULL
                                ELSE (CURRENT_DATE - ls.last_sale_date)
                        END AS dias_sin_venta
                FROM health h
                LEFT JOIN last_sales ls
                    ON ls.referencia_normalizada = public.fn_keep_alnum(h.referencia)
                 AND ls.store_prefix = h.cod_almacen
                WHERE h.health_status = ANY(:statuses)
                    AND (:store_code IS NULL OR h.cod_almacen = :store_code)
                ORDER BY h.inventory_value DESC, h.reorder_qty_recommended DESC, h.historial_ventas_metric ASC
        LIMIT :limit
        """
    )

    try:
        with engine.begin() as connection:
            rows = connection.execute(primary_sql, query_params).mappings().all()
    except SQLAlchemyError as exc:
        logger.warning("inventory health MV unavailable, using direct fallback query: %s", exc)
        with engine.begin() as connection:
            rows = connection.execute(fallback_sql, query_params).mappings().all()
    return [dict(row) for row in rows]


def _fetch_cartera_rows(engine, limit: int) -> list[dict]:
    with engine.begin() as connection:
        rows = connection.execute(
            text(
                """
                SELECT
                    cod_cliente,
                    nombre_cliente,
                    nom_vendedor,
                    zona,
                    balance_total,
                    balance_31_60,
                    balance_61_90,
                    balance_91_plus,
                    max_dias_vencido
                FROM public.mv_internal_cartera_cliente
                WHERE (balance_31_60 + balance_61_90 + balance_91_plus) > 0
                ORDER BY balance_91_plus DESC, balance_61_90 DESC, balance_31_60 DESC
                LIMIT :limit
                """
            ),
            {"limit": limit},
        ).mappings().all()
    return [dict(row) for row in rows]


def _build_sales_projection_summary(projection: dict, store_code: Optional[str]) -> str:
    scope = f" para sede {store_code}" if store_code else ""
    variation_line = ""
    if projection.get("variacion_pct") is not None:
        variation_line = f" | variación vs corte mes anterior {projection['variacion_pct']}%"
    return (
        f"Proyección del mes{scope}: llevamos {_format_currency(projection.get('ventas_mes_actual'))} en {projection.get('dias_transcurridos')} días. "
        f"A este ritmo cerraríamos en {_format_currency(projection.get('proyeccion_cierre_mes'))}.{variation_line}"
    )


def _build_inventory_indicator_summary(title: str, rows: list[dict], limit: int, include_suggested: bool = False) -> str:
    metrics = _build_inventory_summary_metrics(rows)
    lines = [f"{title}: {metrics[0][1]} referencias y {metrics[2][1]} comprometidos en inventario."]
    for row in rows[: min(limit, 5)]:
        suffix = f" | sugerido {_format_number(row.get('reorder_qty_recommended'))}" if include_suggested else ""
        dias = row.get("dias_sin_venta")
        dias_text = f" | {int(dias)} días sin venta" if isinstance(dias, Number) else " | sin venta reciente registrada"
        lines.append(
            f"- {row.get('referencia')} | {row.get('descripcion')} | {row.get('almacen_nombre')} | stock {_format_number(row.get('stock_total'))} | valor {_format_currency(row.get('inventory_value'))}{dias_text}{suffix}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _build_cartera_indicator_summary(rows: list[dict], limit: int) -> str:
    metrics = _build_cartera_summary_metrics(rows)
    lines = [f"Cartera vencida: {metrics[0][1]} clientes con {metrics[1][1]} en saldo total y {metrics[2][1]} en >90 días."]
    for row in rows[: min(limit, 5)]:
        vencido = float(row.get("balance_31_60") or 0) + float(row.get("balance_61_90") or 0) + float(row.get("balance_91_plus") or 0)
        lines.append(
            f"- {row.get('nombre_cliente')} | vencido {_format_currency(vencido)} | >90 {_format_currency(row.get('balance_91_plus'))} | vendedor {row.get('nom_vendedor') or 'N/A'}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_clients_without_purchase_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str], limit: int) -> tuple[list[dict], str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    history_end = current_start - timedelta(days=1)
    history_start = history_end - timedelta(days=180)

    sql = text(
        """
        WITH history AS (
            SELECT
                public.fn_keep_alnum(rv.cliente_id) AS cod_cliente,
                INITCAP(MAX(public.fn_normalize_text(rv.nombre_cliente))) AS nombre_cliente,
                INITCAP(MAX(public.fn_normalize_text(rv.nom_vendedor))) AS nom_vendedor,
                COALESCE(SUM(COALESCE(public.fn_parse_numeric(rv.valor_venta), 0)), 0) AS ventas_historicas,
                MAX(public.fn_parse_date(rv.fecha_venta)) AS ultima_compra,
                COUNT(DISTINCT date_trunc('month', public.fn_parse_date(rv.fecha_venta))) AS meses_activos
            FROM public.raw_ventas_detalle rv
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :history_start AND :history_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
            GROUP BY public.fn_keep_alnum(rv.cliente_id)
        ),
        current_period AS (
            SELECT DISTINCT public.fn_keep_alnum(rv.cliente_id) AS cod_cliente
            FROM public.raw_ventas_detalle rv
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
        )
        SELECT
            h.cod_cliente,
            h.nombre_cliente,
            h.nom_vendedor,
            h.ventas_historicas,
            h.ultima_compra,
            h.meses_activos,
            (CURRENT_DATE - h.ultima_compra) AS dias_sin_compra
        FROM history h
        LEFT JOIN current_period cp ON cp.cod_cliente = h.cod_cliente
        WHERE cp.cod_cliente IS NULL
          AND h.ventas_historicas > 0
          AND h.meses_activos >= 2
        ORDER BY h.ventas_historicas DESC, h.ultima_compra DESC
        LIMIT :limit
        """
    )

    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "history_start": history_start,
                "history_end": history_end,
                "current_start": current_start,
                "current_end": current_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label


def _build_clients_without_purchase_summary(rows: list[dict], period_label: str, limit: int, title: str) -> str:
    total_sales = sum(float(row.get("ventas_historicas") or 0) for row in rows)
    lines = [f"{title} en {period_label}: {len(rows)} clientes activos dejaron de comprar y representan {_format_currency(total_sales)} en base histórica reciente."]
    for row in rows[: min(limit, 5)]:
        dias = row.get("dias_sin_compra")
        dias_text = f"{int(dias)} días" if isinstance(dias, Number) else "sin fecha clara"
        lines.append(
            f"- {row.get('nombre_cliente') or row.get('cod_cliente')} | última compra {row.get('ultima_compra') or 'N/D'} | {dias_text} sin compra | histórico {_format_currency(row.get('ventas_historicas'))} | vendedor {row.get('nom_vendedor') or 'N/A'}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_products_without_sale_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str], limit: int) -> tuple[list[dict], str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    history_end = current_start - timedelta(days=1)
    history_start = history_end - timedelta(days=180)

    sql = text(
        """
        WITH inventory AS (
            SELECT
                inv.referencia_normalizada,
                CASE WHEN :store_code IS NULL THEN 'Consolidado' ELSE MAX(inv.almacen_nombre) END AS almacen_nombre,
                MAX(inv.referencia) AS referencia,
                MAX(inv.descripcion) AS descripcion,
                COALESCE(SUM(inv.stock_disponible), 0) AS stock_total,
                COALESCE(SUM(inv.stock_disponible * COALESCE(inv.costo_promedio_und, 0)), 0) AS inventory_value
            FROM public.vw_inventario_agente inv
            WHERE :store_code IS NULL OR inv.cod_almacen = :store_code
            GROUP BY inv.referencia_normalizada
        ),
        history AS (
            SELECT
                am.referencia_normalizada,
                INITCAP(MAX(public.fn_normalize_text(rv.nombre_articulo))) AS descripcion_venta,
                INITCAP(MAX(public.fn_normalize_text(rv.linea_producto))) AS linea_producto,
                COALESCE(SUM(COALESCE(public.fn_parse_numeric(rv.valor_venta), 0)), 0) AS ventas_historicas,
                COALESCE(SUM(COALESCE(public.fn_parse_numeric(rv.unidades_vendidas), 0)), 0) AS unidades_historicas,
                MAX(public.fn_parse_date(rv.fecha_venta)) AS ultima_venta,
                COUNT(DISTINCT date_trunc('month', public.fn_parse_date(rv.fecha_venta))) AS meses_activos
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :history_start AND :history_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
            GROUP BY am.referencia_normalizada
        ),
        current_period AS (
            SELECT
                am.referencia_normalizada,
                COALESCE(SUM(COALESCE(public.fn_parse_numeric(rv.valor_venta), 0)), 0) AS ventas_actuales
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
            GROUP BY am.referencia_normalizada
        )
        SELECT
            COALESCE(inv.almacen_nombre, 'Consolidado') AS almacen_nombre,
            COALESCE(inv.referencia, h.referencia_normalizada) AS referencia,
            COALESCE(inv.descripcion, h.descripcion_venta) AS descripcion,
            h.linea_producto,
            COALESCE(inv.stock_total, 0) AS stock_total,
            COALESCE(inv.inventory_value, 0) AS inventory_value,
            h.ventas_historicas,
            h.unidades_historicas,
            h.ultima_venta AS fecha_ultima_venta,
            (CURRENT_DATE - h.ultima_venta) AS dias_sin_venta,
            h.meses_activos
        FROM history h
        LEFT JOIN current_period cp ON cp.referencia_normalizada = h.referencia_normalizada
        LEFT JOIN inventory inv ON inv.referencia_normalizada = h.referencia_normalizada
        WHERE COALESCE(cp.ventas_actuales, 0) <= 0
          AND h.ventas_historicas > 0
          AND h.meses_activos >= 2
          AND COALESCE(inv.stock_total, 0) > 0
        ORDER BY h.ventas_historicas DESC, COALESCE(inv.stock_total, 0) DESC, h.ultima_venta DESC
        LIMIT :limit
        """
    )

    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "history_start": history_start,
                "history_end": history_end,
                "current_start": current_start,
                "current_end": current_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label


def _build_products_without_sale_summary(rows: list[dict], period_label: str, limit: int) -> str:
    total_value = sum(float(row.get("ventas_historicas") or 0) for row in rows)
    lines = [f"Productos que normalmente se venden y no salieron en {period_label}: {len(rows)} referencias con una base histórica de {_format_currency(total_value)} y stock disponible para mover."]
    for row in rows[: min(limit, 5)]:
        dias = row.get("dias_sin_venta")
        dias_text = f"{int(dias)} días" if isinstance(dias, Number) else "sin dato"
        lines.append(
            f"- {row.get('referencia')} | {row.get('descripcion')} | stock {_format_number(row.get('stock_total'))} | histórico {_format_currency(row.get('ventas_historicas'))} | {dias_text} sin venta"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_products_to_push_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str], limit: int) -> tuple[list[dict], str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    history_end = current_start - timedelta(days=1)
    history_start = history_end - timedelta(days=180)

    sql = text(
        """
        WITH inventory AS (
            SELECT
                inv.referencia_normalizada,
                CASE WHEN :store_code IS NULL THEN 'Consolidado' ELSE MAX(inv.almacen_nombre) END AS almacen_nombre,
                MAX(inv.referencia) AS referencia,
                MAX(inv.descripcion) AS descripcion,
                COALESCE(SUM(inv.stock_disponible), 0) AS stock_total,
                COALESCE(SUM(inv.stock_disponible * COALESCE(inv.costo_promedio_und, 0)), 0) AS inventory_value
            FROM public.vw_inventario_agente inv
            WHERE :store_code IS NULL OR inv.cod_almacen = :store_code
            GROUP BY inv.referencia_normalizada
        ),
        product_history AS (
            SELECT
                am.referencia_normalizada,
                INITCAP(MAX(public.fn_normalize_text(rv.nombre_articulo))) AS descripcion_venta,
                COALESCE(SUM(COALESCE(public.fn_parse_numeric(rv.valor_venta), 0)), 0) AS ventas_historicas,
                MAX(public.fn_parse_date(rv.fecha_venta)) AS ultima_venta,
                COUNT(DISTINCT date_trunc('month', public.fn_parse_date(rv.fecha_venta))) AS meses_activos
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :history_start AND :history_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
            GROUP BY am.referencia_normalizada
        ),
        product_current AS (
            SELECT
                am.referencia_normalizada,
                COALESCE(SUM(COALESCE(public.fn_parse_numeric(rv.valor_venta), 0)), 0) AS ventas_actuales
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
            GROUP BY am.referencia_normalizada
        ),
        ranked_products AS (
            SELECT
                ph.referencia_normalizada,
                COALESCE(inv.almacen_nombre, 'Consolidado') AS almacen_nombre,
                COALESCE(inv.referencia, ph.referencia_normalizada) AS referencia,
                COALESCE(inv.descripcion, ph.descripcion_venta) AS descripcion,
                COALESCE(inv.stock_total, 0) AS stock_total,
                COALESCE(inv.inventory_value, 0) AS inventory_value,
                ph.ventas_historicas,
                COALESCE(pc.ventas_actuales, 0) AS ventas_actuales,
                ROUND(ph.ventas_historicas / GREATEST(ph.meses_activos, 1), 2) AS promedio_base,
                GREATEST(ROUND(ph.ventas_historicas / GREATEST(ph.meses_activos, 1), 2) - COALESCE(pc.ventas_actuales, 0), 0) AS brecha_oportunidad,
                ph.ultima_venta,
                (CURRENT_DATE - ph.ultima_venta) AS dias_sin_venta
            FROM product_history ph
            LEFT JOIN product_current pc ON pc.referencia_normalizada = ph.referencia_normalizada
            LEFT JOIN inventory inv ON inv.referencia_normalizada = ph.referencia_normalizada
            WHERE ph.ventas_historicas > 0
              AND ph.meses_activos >= 2
              AND COALESCE(inv.stock_total, 0) > 0
        ),
        top_products AS (
            SELECT *
            FROM ranked_products
            WHERE brecha_oportunidad > 0
            ORDER BY brecha_oportunidad DESC, stock_total DESC, ventas_historicas DESC
            LIMIT :limit
        ),
        client_history AS (
            SELECT
                am.referencia_normalizada,
                public.fn_keep_alnum(rv.cliente_id) AS cod_cliente,
                INITCAP(MAX(public.fn_normalize_text(rv.nombre_cliente))) AS nombre_cliente,
                COALESCE(SUM(COALESCE(public.fn_parse_numeric(rv.valor_venta), 0)), 0) AS ventas_historicas_cliente
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :history_start AND :history_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
            GROUP BY am.referencia_normalizada, public.fn_keep_alnum(rv.cliente_id)
        ),
        client_current AS (
            SELECT DISTINCT
                am.referencia_normalizada,
                public.fn_keep_alnum(rv.cliente_id) AS cod_cliente
            FROM public.raw_ventas_detalle rv
            JOIN public.articulos_maestro am ON am.codigo_articulo = rv.codigo_articulo
            WHERE public.fn_normalize_text(rv.tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(rv.fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(rv.serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(rv.codigo_vendedor) = :vendor_code)
        ),
        target_ranked AS (
            SELECT
                ch.referencia_normalizada,
                ch.nombre_cliente,
                ch.ventas_historicas_cliente,
                ROW_NUMBER() OVER (PARTITION BY ch.referencia_normalizada ORDER BY ch.ventas_historicas_cliente DESC) AS rn
            FROM client_history ch
            LEFT JOIN client_current cc
              ON cc.referencia_normalizada = ch.referencia_normalizada
             AND cc.cod_cliente = ch.cod_cliente
            WHERE cc.cod_cliente IS NULL
        ),
        target_clients AS (
            SELECT
                referencia_normalizada,
                string_agg(nombre_cliente, ', ' ORDER BY ventas_historicas_cliente DESC) AS clientes_objetivo
            FROM target_ranked
            WHERE rn <= 3
            GROUP BY referencia_normalizada
        )
        SELECT
            tp.almacen_nombre,
            tp.referencia,
            tp.descripcion,
            tp.stock_total,
            tp.inventory_value,
            tp.ventas_historicas,
            tp.ventas_actuales,
            tp.promedio_base,
            tp.brecha_oportunidad,
            tp.ultima_venta AS fecha_ultima_venta,
            tp.dias_sin_venta,
            COALESCE(tc.clientes_objetivo, 'Sin clientes objetivo claros') AS clientes_objetivo
        FROM top_products tp
        LEFT JOIN target_clients tc ON tc.referencia_normalizada = tp.referencia_normalizada
        ORDER BY tp.brecha_oportunidad DESC, tp.stock_total DESC, tp.ventas_historicas DESC
        """
    )

    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "history_start": history_start,
                "history_end": history_end,
                "current_start": current_start,
                "current_end": current_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label


def _build_products_to_push_summary(rows: list[dict], period_label: str, limit: int) -> str:
    total_gap = sum(float(row.get("brecha_oportunidad") or 0) for row in rows)
    lines = [f"Productos para impulsar en {period_label}: {len(rows)} referencias con brecha comercial estimada de {_format_currency(total_gap)} frente a su comportamiento reciente."]
    for row in rows[: min(limit, 5)]:
        dias = row.get("dias_sin_venta")
        dias_text = f"{int(dias)} días" if isinstance(dias, Number) else "sin dato"
        lines.append(
            f"- {row.get('referencia')} | {row.get('descripcion')} | stock {_format_number(row.get('stock_total'))} | actual {_format_currency(row.get('ventas_actuales'))} vs base {_format_currency(row.get('promedio_base'))} | brecha {_format_currency(row.get('brecha_oportunidad'))} | clientes {row.get('clientes_objetivo')} | {dias_text} sin venta"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _shift_year_safe(value: date, delta_years: int) -> date:
    target_year = value.year + delta_years
    target_day = min(value.day, calendar.monthrange(target_year, value.month)[1])
    return date(target_year, value.month, target_day)


def _resolve_previous_period(periodo_raw: Optional[str], comparison_mode: str) -> tuple[date, date, str, str]:
    current_start, current_end, current_label = _parse_bi_period(periodo_raw)
    if comparison_mode == "vs_periodo_anterior":
        span_days = max((current_end - current_start).days, 0) + 1
        previous_end = current_start - timedelta(days=1)
        previous_start = previous_end - timedelta(days=span_days - 1)
        return previous_start, previous_end, current_label, "periodo anterior equivalente"
    previous_start = _shift_year_safe(current_start, -1)
    previous_end = _shift_year_safe(current_end, -1)
    return previous_start, previous_end, current_label, "mismo corte del año anterior"


def _get_sales_dimension_sql_parts(dimension: str) -> tuple[str, str, str]:
    if dimension == "tienda":
        return (
            "LEFT(COALESCE(serie, ''), 3)",
            "CASE LEFT(COALESCE(serie, ''), 3) WHEN '156' THEN 'Armenia' WHEN '157' THEN 'Manizales' WHEN '158' THEN 'Opalo' WHEN '189' THEN 'Pereira' WHEN '238' THEN 'Laureles' WHEN '439' THEN 'FerreBOX' WHEN '463' THEN 'Cerritos' ELSE 'Otra sede' END",
            "NULL::text AS codigo_aux, NULL::text AS detalle_aux",
        )
    if dimension == "vendedor":
        return (
            "public.fn_keep_alnum(codigo_vendedor)",
            "INITCAP(MAX(public.fn_normalize_text(nom_vendedor)))",
            "public.fn_keep_alnum(codigo_vendedor) AS codigo_aux, NULL::text AS detalle_aux",
        )
    if dimension == "cliente":
        return (
            "public.fn_keep_alnum(cliente_id)",
            "INITCAP(MAX(public.fn_normalize_text(nombre_cliente)))",
            "public.fn_keep_alnum(cliente_id) AS codigo_aux, NULL::text AS detalle_aux",
        )
    if dimension == "producto":
        return (
            "public.fn_keep_alnum(codigo_articulo)",
            "INITCAP(MAX(public.fn_normalize_text(nombre_articulo)))",
            "public.fn_keep_alnum(codigo_articulo) AS codigo_aux, INITCAP(MAX(public.fn_normalize_text(linea_producto))) AS detalle_aux",
        )
    if dimension == "linea":
        return (
            "public.fn_normalize_text(linea_producto)",
            "INITCAP(MAX(public.fn_normalize_text(linea_producto)))",
            "NULL::text AS codigo_aux, NULL::text AS detalle_aux",
        )
    if dimension == "zona":
        return (
            "COALESCE(public.fn_normalize_text(zona), 'sin zona')",
            "INITCAP(MAX(COALESCE(public.fn_normalize_text(zona), 'sin zona')))",
            "NULL::text AS codigo_aux, NULL::text AS detalle_aux",
        )
    raise ValueError("dimension de ventas no soportada")


def _build_sales_scope_filters(question: str, args: dict, internal_auth: Optional[dict]) -> tuple[Optional[str], Optional[str], Optional[str]]:
    internal_auth = internal_auth or {}
    employee_context = internal_auth.get("employee_context") or {}
    role = _normalize_text(internal_auth.get("role") or "")

    requested_store = _resolve_store_code(args.get("almacen")) or _extract_store_code_from_question(question)
    employee_store = _resolve_store_code(employee_context.get("store_code"))
    if role == "operador":
        if requested_store and employee_store and requested_store != employee_store:
            return None, None, f"Tu perfil solo tiene acceso a la sede {employee_store}."
        store_code = employee_store
    else:
        store_code = requested_store or employee_store if role == "empleado" else requested_store

    requested_vendor = _resolve_vendor_code(args.get("vendedor_codigo") or _extract_vendor_code_from_question(question), internal_auth)
    employee_vendor = _resolve_vendor_code((employee_context or {}).get("codigo_vendedor"), internal_auth)
    if role == "vendedor":
        if requested_vendor and employee_vendor and requested_vendor != employee_vendor:
            return None, None, "Solo puedes consultar tu propio código de vendedor."
        vendor_code = employee_vendor or requested_vendor
    else:
        vendor_code = requested_vendor

    return store_code, vendor_code, None


def _fetch_sales_total_snapshot(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str]) -> dict:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    previous_start = _shift_year_safe(current_start, -1)
    previous_end = _shift_year_safe(current_end, -1)
    sql = text(
        """
        WITH current_period AS (
            SELECT
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturado,
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones,
                COUNT(*) FILTER (WHERE public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%') AS lineas,
                COUNT(DISTINCT public.fn_keep_alnum(cliente_id)) FILTER (WHERE public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%') AS clientes
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
        ),
        previous_period AS (
            SELECT
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturado,
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :previous_start AND :previous_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
        )
        SELECT
            cp.facturado,
            cp.devoluciones,
            cp.lineas,
            cp.clientes,
            pp.facturado AS facturado_prev,
            pp.devoluciones AS devoluciones_prev
        FROM current_period cp
        CROSS JOIN previous_period pp
        """
    )
    with engine.begin() as connection:
        row = connection.execute(
            sql,
            {
                "current_start": current_start,
                "current_end": current_end,
                "previous_start": previous_start,
                "previous_end": previous_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
            },
        ).mappings().one()
    facturado = float(row.get("facturado") or 0)
    devoluciones = float(row.get("devoluciones") or 0)
    neto = facturado - devoluciones
    prev_neto = float(row.get("facturado_prev") or 0) - float(row.get("devoluciones_prev") or 0)
    variacion_pct = ((neto - prev_neto) / prev_neto * 100.0) if prev_neto > 0 else None
    return {
        "period_label": period_label,
        "facturado": facturado,
        "devoluciones": devoluciones,
        "neto": neto,
        "clientes": int(row.get("clientes") or 0),
        "lineas": int(row.get("lineas") or 0),
        "prev_neto": prev_neto,
        "variacion_pct": round(variacion_pct, 1) if variacion_pct is not None else None,
    }


def _fetch_sales_dimension_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str], dimension: str, limit: int, direction: str) -> tuple[list[dict], str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    order_direction = "ASC" if direction == "asc" else "DESC"
    group_key, label_expr, extra_select = _get_sales_dimension_sql_parts(dimension)

    sql = text(
        f"""
        SELECT
            {group_key} AS group_key,
            {label_expr} AS group_label,
            {extra_select},
            SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END) AS facturado,
            SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS devoluciones,
            COUNT(*) FILTER (WHERE public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%') AS lineas,
            COUNT(DISTINCT public.fn_keep_alnum(cliente_id)) FILTER (WHERE public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%') AS clientes,
            SUM(COALESCE(public.fn_parse_numeric(unidades_vendidas), 0)) AS unidades
        FROM public.raw_ventas_detalle
        WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
          AND public.fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
          AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
          AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
        GROUP BY 1
        ORDER BY (SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
             - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END)) {order_direction}
        LIMIT :limit
        """
    )
    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "current_start": current_start,
                "current_end": current_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    mapped_rows = []
    for row in rows:
        neto = float(row.get("facturado") or 0) - float(row.get("devoluciones") or 0)
        mapped_rows.append({
            "label": row.get("group_label") or row.get("group_key") or "N/D",
            "codigo": row.get("codigo_aux"),
            "detalle": row.get("detalle_aux"),
            "neto": neto,
            "facturado": float(row.get("facturado") or 0),
            "devoluciones": float(row.get("devoluciones") or 0),
            "clientes": int(row.get("clientes") or 0),
            "lineas": int(row.get("lineas") or 0),
            "unidades": float(row.get("unidades") or 0),
        })
    return mapped_rows, period_label


def _build_sales_total_summary(snapshot: dict, store_code: Optional[str], vendor_code: Optional[str]) -> str:
    scope_parts = []
    if store_code:
        scope_parts.append(_STORE_CODE_LABELS.get(store_code, store_code))
    if vendor_code:
        scope_parts.append(f"vendedor {vendor_code}")
    scope_text = " / ".join(scope_parts) if scope_parts else "toda la empresa"
    summary = (
        f"Ventas de {snapshot.get('period_label')} en {scope_text}: neto {_format_currency(snapshot.get('neto'))}, "
        f"facturado {_format_currency(snapshot.get('facturado'))}, devoluciones {_format_currency(snapshot.get('devoluciones'))}, "
        f"clientes {snapshot.get('clientes')} y líneas {snapshot.get('lineas')}."
    )
    if snapshot.get("variacion_pct") is not None:
        summary += f" Variación vs mismo corte del año anterior: {_format_percent(snapshot.get('variacion_pct'))}."
    return summary


def _build_sales_dimension_summary(question: str, rows: list[dict], period_label: str, dimension: str, limit: int, direction: str) -> str:
    if not rows:
        return f"No encontré datos de ventas por {dimension} para {period_label}."
    total_neto = sum(float(row.get("neto") or 0) for row in rows)
    direction_text = "menor" if direction == "asc" else "mayor"
    lines = [f"Ventas por {dimension} en {period_label}: top {min(limit, len(rows))} con {direction_text} desempeño dentro del corte, por {_format_currency(total_neto)} acumulados en esta vista."]
    for row in rows[: min(limit, 5)]:
        detail = f" | código {row.get('codigo')}" if row.get("codigo") and dimension in {"vendedor", "cliente", "producto"} else ""
        extra = f" | línea {row.get('detalle')}" if row.get("detalle") else ""
        units = f" | unidades {_format_number(row.get('unidades'))}" if dimension in {"producto", "linea"} else ""
        lines.append(
            f"- {row.get('label')}{detail}{extra} | neto {_format_currency(row.get('neto'))} | clientes {_format_number(row.get('clientes'))} | líneas {_format_number(row.get('lineas'))}{units}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_sales_share_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str], dimension: str, limit: int, direction: str) -> tuple[list[dict], str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    order_direction = "ASC" if direction == "asc" else "DESC"
    group_key, label_expr, extra_select = _get_sales_dimension_sql_parts(dimension)
    sql = text(
        f"""
        WITH grouped AS (
            SELECT
                {group_key} AS group_key,
                {label_expr} AS group_label,
                {extra_select},
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
                  - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS neto
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
            GROUP BY 1
        )
        SELECT
            group_label,
            codigo_aux,
            detalle_aux,
            neto,
            CASE WHEN SUM(neto) OVER () = 0 THEN NULL ELSE ROUND((neto / SUM(neto) OVER ()) * 100.0, 1) END AS participacion_pct
        FROM grouped
        ORDER BY neto {order_direction}
        LIMIT :limit
        """
    )
    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "current_start": current_start,
                "current_end": current_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label


def _build_sales_share_summary(rows: list[dict], period_label: str, dimension: str, limit: int) -> str:
    if not rows:
        return f"No encontré participación por {dimension} para {period_label}."
    total_neto = sum(float(row.get("neto") or 0) for row in rows)
    lines = [f"Participación por {dimension} en {period_label}: las {min(limit, len(rows))} principales posiciones explican {_format_currency(total_neto)} dentro de esta vista analítica."]
    for row in rows[: min(limit, 5)]:
        extra = f" | código {row.get('codigo_aux')}" if row.get("codigo_aux") else ""
        detail = f" | detalle {row.get('detalle_aux')}" if row.get("detalle_aux") else ""
        lines.append(
            f"- {row.get('group_label')}{extra}{detail} | neto {_format_currency(row.get('neto'))} | participación {_format_percent(row.get('participacion_pct'))}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_sales_growth_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str], dimension: str, limit: int, direction: str, comparison_mode: str) -> tuple[list[dict], str, str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    previous_start, previous_end, _, comparison_label = _resolve_previous_period(periodo_raw, comparison_mode)
    order_direction = "ASC" if direction == "asc" else "DESC"
    group_key, label_expr, extra_select = _get_sales_dimension_sql_parts(dimension)
    sql = text(
        f"""
        WITH current_period AS (
            SELECT
                {group_key} AS group_key,
                {label_expr} AS group_label,
                {extra_select},
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
                  - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS neto_actual
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
            GROUP BY 1
        ),
        previous_period AS (
            SELECT
                {group_key} AS group_key,
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
                  - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS neto_previo
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :previous_start AND :previous_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
            GROUP BY 1
        )
        SELECT
            cp.group_label,
            cp.codigo_aux,
            cp.detalle_aux,
            cp.neto_actual,
            COALESCE(pp.neto_previo, 0) AS neto_previo,
            cp.neto_actual - COALESCE(pp.neto_previo, 0) AS variacion_absoluta,
            CASE WHEN COALESCE(pp.neto_previo, 0) = 0 THEN NULL ELSE ROUND(((cp.neto_actual - pp.neto_previo) / pp.neto_previo) * 100.0, 1) END AS variacion_pct
        FROM current_period cp
        LEFT JOIN previous_period pp ON pp.group_key = cp.group_key
        ORDER BY COALESCE(CASE WHEN COALESCE(pp.neto_previo, 0) = 0 THEN NULL ELSE ((cp.neto_actual - pp.neto_previo) / pp.neto_previo) * 100.0 END, 0) {order_direction}, cp.neto_actual {order_direction}
        LIMIT :limit
        """
    )
    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "current_start": current_start,
                "current_end": current_end,
                "previous_start": previous_start,
                "previous_end": previous_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label, comparison_label


def _build_sales_growth_summary(rows: list[dict], period_label: str, comparison_label: str, dimension: str, limit: int, direction: str) -> str:
    if not rows:
        return f"No encontré crecimiento por {dimension} para {period_label}."
    focus = "mayor crecimiento" if direction == "desc" else "mayor caída"
    lines = [f"Crecimiento por {dimension} en {period_label}: top {min(limit, len(rows))} con enfoque en {focus} frente a {comparison_label}."]
    for row in rows[: min(limit, 5)]:
        lines.append(
            f"- {row.get('group_label')} | actual {_format_currency(row.get('neto_actual'))} | previo {_format_currency(row.get('neto_previo'))} | variación {_format_currency(row.get('variacion_absoluta'))} | crecimiento {_format_percent(row.get('variacion_pct'))}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_client_frequency_drop_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], vendor_code: Optional[str], limit: int, comparison_mode: str) -> tuple[list[dict], str, str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    previous_start, previous_end, _, comparison_label = _resolve_previous_period(periodo_raw, comparison_mode)
    sql = text(
        """
        WITH current_period AS (
            SELECT
                public.fn_keep_alnum(cliente_id) AS cod_cliente,
                INITCAP(MAX(public.fn_normalize_text(nombre_cliente))) AS nombre_cliente,
                INITCAP(MAX(public.fn_normalize_text(nom_vendedor))) AS nom_vendedor,
                COUNT(DISTINCT public.fn_parse_date(fecha_venta)) AS dias_compra_actual,
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
                  - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS neto_actual
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
            GROUP BY 1
        ),
        previous_period AS (
            SELECT
                public.fn_keep_alnum(cliente_id) AS cod_cliente,
                COUNT(DISTINCT public.fn_parse_date(fecha_venta)) AS dias_compra_prev,
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
                  - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS neto_previo
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :previous_start AND :previous_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
            GROUP BY 1
        )
        SELECT
            pp.cod_cliente,
            COALESCE(cp.nombre_cliente, pp.cod_cliente) AS nombre_cliente,
            cp.nom_vendedor,
            COALESCE(cp.dias_compra_actual, 0) AS dias_compra_actual,
            pp.dias_compra_prev,
            COALESCE(cp.neto_actual, 0) AS neto_actual,
            pp.neto_previo,
            COALESCE(cp.dias_compra_actual, 0) - pp.dias_compra_prev AS delta_frecuencia,
            COALESCE(cp.neto_actual, 0) - pp.neto_previo AS delta_neto
        FROM previous_period pp
        LEFT JOIN current_period cp ON cp.cod_cliente = pp.cod_cliente
        WHERE pp.dias_compra_prev > COALESCE(cp.dias_compra_actual, 0)
        ORDER BY delta_frecuencia ASC, delta_neto ASC
        LIMIT :limit
        """
    )
    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "current_start": current_start,
                "current_end": current_end,
                "previous_start": previous_start,
                "previous_end": previous_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label, comparison_label


def _build_client_frequency_drop_summary(rows: list[dict], period_label: str, comparison_label: str, limit: int) -> str:
    if not rows:
        return f"No encontré caída de frecuencia para {period_label}."
    lines = [f"Caída de frecuencia en {period_label}: {len(rows)} clientes compraron en menos días frente a {comparison_label}."]
    for row in rows[: min(limit, 5)]:
        lines.append(
            f"- {row.get('nombre_cliente')} | días compra actual {_format_number(row.get('dias_compra_actual'))} vs previo {_format_number(row.get('dias_compra_prev'))} | neto actual {_format_currency(row.get('neto_actual'))} vs previo {_format_currency(row.get('neto_previo'))} | vendedor {row.get('nom_vendedor') or 'N/A'}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_cartera_concentration_rows(engine, dimension: str, limit: int) -> list[dict]:
    if dimension == "vendedor":
        group_expr = "COALESCE(public.fn_normalize_text(nom_vendedor), 'sin vendedor')"
        label_expr = "INITCAP(COALESCE(public.fn_normalize_text(nom_vendedor), 'sin vendedor'))"
        code_expr = "NULL::text"
    elif dimension == "zona":
        group_expr = "COALESCE(public.fn_normalize_text(zona), 'sin zona')"
        label_expr = "INITCAP(COALESCE(public.fn_normalize_text(zona), 'sin zona'))"
        code_expr = "NULL::text"
    else:
        group_expr = "COALESCE(cod_cliente, 'sin cliente')"
        label_expr = "INITCAP(COALESCE(public.fn_normalize_text(nombre_cliente), cod_cliente, 'sin cliente'))"
        code_expr = "COALESCE(cod_cliente, 'sin cliente')"

    sql = text(
        f"""
        WITH grouped AS (
            SELECT
                {group_expr} AS group_key,
                MAX({label_expr}) AS group_label,
                MAX({code_expr}) AS codigo_aux,
                SUM(COALESCE(balance_61_90, 0) + COALESCE(balance_91_plus, 0)) AS saldo_vencido,
                SUM(COALESCE(balance_total, 0)) AS saldo_total,
                MAX(COALESCE(max_dias_vencido, 0)) AS max_dias_vencido
            FROM public.mv_internal_cartera_cliente
            GROUP BY 1
        )
        SELECT
            group_label,
            codigo_aux,
            saldo_vencido,
            saldo_total,
            max_dias_vencido,
            CASE WHEN SUM(saldo_vencido) OVER () = 0 THEN NULL ELSE ROUND((saldo_vencido / SUM(saldo_vencido) OVER ()) * 100.0, 1) END AS participacion_pct
        FROM grouped
        WHERE saldo_vencido > 0
        ORDER BY saldo_vencido DESC
        LIMIT :limit
        """
    )
    with engine.begin() as connection:
        rows = connection.execute(sql, {"limit": limit}).mappings().all()
    return [dict(row) for row in rows]


def _build_cartera_concentration_summary(rows: list[dict], dimension: str, limit: int) -> str:
    if not rows:
        return "No encontré concentración de cartera para ese filtro."
    total = sum(float(row.get("saldo_vencido") or 0) for row in rows)
    lines = [f"Concentración de cartera por {dimension}: las {min(limit, len(rows))} principales posiciones acumulan {_format_currency(total)} en saldo vencido dentro de esta vista."]
    for row in rows[: min(limit, 5)]:
        code = f" | código {row.get('codigo_aux')}" if row.get("codigo_aux") else ""
        lines.append(
            f"- {row.get('group_label')}{code} | vencido {_format_currency(row.get('saldo_vencido'))} | participación {_format_percent(row.get('participacion_pct'))} | saldo total {_format_currency(row.get('saldo_total'))} | max {_format_number(row.get('max_dias_vencido'))} días"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_opportunity_dimension_rows(engine, periodo_raw: Optional[str], dimension: str, store_code: Optional[str], vendor_code: Optional[str], limit: int, comparison_mode: str) -> tuple[list[dict], str, str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    previous_start, previous_end, _, comparison_label = _resolve_previous_period(periodo_raw, comparison_mode)
    if dimension not in {"tienda", "vendedor"}:
        dimension = "tienda"
    group_key, label_expr, extra_select = _get_sales_dimension_sql_parts(dimension)
    sql = text(
        f"""
        WITH current_sales AS (
            SELECT
                {group_key} AS group_key,
                {label_expr} AS group_label,
                {extra_select},
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
                  - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS neto_actual
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
            GROUP BY 1
        ),
        previous_sales AS (
            SELECT
                {group_key} AS group_key,
                SUM(CASE WHEN public.fn_normalize_text(tipo_documento) NOT LIKE '%nota%' THEN COALESCE(public.fn_parse_numeric(valor_venta), 0) ELSE 0 END)
                  - SUM(CASE WHEN public.fn_normalize_text(tipo_documento) LIKE '%nota%' THEN ABS(COALESCE(public.fn_parse_numeric(valor_venta), 0)) ELSE 0 END) AS neto_previo
            FROM public.raw_ventas_detalle
            WHERE (public.fn_normalize_text(tipo_documento) LIKE '%factura%' OR public.fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND public.fn_parse_date(fecha_venta) BETWEEN :previous_start AND :previous_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
            GROUP BY 1
        ),
        previous_clients AS (
            SELECT DISTINCT
                {group_key} AS group_key,
                public.fn_keep_alnum(cliente_id) AS cod_cliente
            FROM public.raw_ventas_detalle
            WHERE public.fn_normalize_text(tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(fecha_venta) BETWEEN :previous_start AND :previous_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
        ),
        current_clients AS (
            SELECT DISTINCT
                {group_key} AS group_key,
                public.fn_keep_alnum(cliente_id) AS cod_cliente
            FROM public.raw_ventas_detalle
            WHERE public.fn_normalize_text(tipo_documento) LIKE '%factura%'
              AND public.fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
              AND (:vendor_code IS NULL OR public.fn_keep_alnum(codigo_vendedor) = :vendor_code)
        ),
        inactive_clients AS (
            SELECT
                pc.group_key,
                COUNT(*) AS clientes_reactivables
            FROM previous_clients pc
            LEFT JOIN current_clients cc ON cc.group_key = pc.group_key AND cc.cod_cliente = pc.cod_cliente
            WHERE cc.cod_cliente IS NULL
            GROUP BY pc.group_key
        )
        SELECT
            cs.group_label,
            cs.codigo_aux,
            COALESCE(cs.neto_actual, 0) AS neto_actual,
            COALESCE(ps.neto_previo, 0) AS neto_previo,
            GREATEST(COALESCE(ps.neto_previo, 0) - COALESCE(cs.neto_actual, 0), 0) AS brecha_oportunidad,
            COALESCE(ic.clientes_reactivables, 0) AS clientes_reactivables
        FROM current_sales cs
        LEFT JOIN previous_sales ps ON ps.group_key = cs.group_key
        LEFT JOIN inactive_clients ic ON ic.group_key = cs.group_key
        ORDER BY brecha_oportunidad DESC, clientes_reactivables DESC, neto_actual DESC
        LIMIT :limit
        """
    )
    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "current_start": current_start,
                "current_end": current_end,
                "previous_start": previous_start,
                "previous_end": previous_end,
                "store_code": store_code,
                "vendor_code": vendor_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label, comparison_label


def _build_opportunity_dimension_summary(rows: list[dict], period_label: str, comparison_label: str, dimension: str, limit: int) -> str:
    if not rows:
        return f"No encontré oportunidades por {dimension} para {period_label}."
    total_gap = sum(float(row.get("brecha_oportunidad") or 0) for row in rows)
    lines = [f"Oportunidades por {dimension} en {period_label}: la brecha acumulada frente a {comparison_label} es de {_format_currency(total_gap)} en esta vista, con foco en recuperación comercial y clientes reactivables."]
    for row in rows[: min(limit, 5)]:
        lines.append(
            f"- {row.get('group_label')} | actual {_format_currency(row.get('neto_actual'))} | previo {_format_currency(row.get('neto_previo'))} | brecha {_format_currency(row.get('brecha_oportunidad'))} | clientes reactivables {_format_number(row.get('clientes_reactivables'))}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _fetch_client_decline_rows(engine, periodo_raw: Optional[str], store_code: Optional[str], limit: int) -> tuple[list[dict], str]:
    current_start, current_end, period_label = _parse_bi_period(periodo_raw)
    previous_start = date(current_start.year - 1, current_start.month, current_start.day)
    previous_end_candidate = current_end - timedelta(days=365)
    previous_end = date(previous_start.year, previous_end_candidate.month, min(previous_end_candidate.day, calendar.monthrange(previous_end_candidate.year, previous_end_candidate.month)[1]))

    sql = text(
        """
        WITH current_period AS (
            SELECT
                fn_keep_alnum(cliente_id) AS cod_cliente,
                MAX(fn_normalize_text(nombre_cliente)) AS nombre_cliente,
                COALESCE(SUM(COALESCE(fn_parse_numeric(valor_venta), 0)), 0) AS ventas_actuales
            FROM public.raw_ventas_detalle
            WHERE (fn_normalize_text(tipo_documento) LIKE '%factura%' OR fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND fn_parse_date(fecha_venta) BETWEEN :current_start AND :current_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
            GROUP BY fn_keep_alnum(cliente_id)
        ),
        previous_period AS (
            SELECT
                fn_keep_alnum(cliente_id) AS cod_cliente,
                COALESCE(SUM(COALESCE(fn_parse_numeric(valor_venta), 0)), 0) AS ventas_previas
            FROM public.raw_ventas_detalle
            WHERE (fn_normalize_text(tipo_documento) LIKE '%factura%' OR fn_normalize_text(tipo_documento) LIKE '%nota%credito%')
              AND fn_parse_date(fecha_venta) BETWEEN :previous_start AND :previous_end
              AND (:store_code IS NULL OR LEFT(COALESCE(serie, ''), 3) = :store_code)
            GROUP BY fn_keep_alnum(cliente_id)
        )
        SELECT
            c.cod_cliente,
            c.nombre_cliente,
            c.ventas_actuales,
            COALESCE(p.ventas_previas, 0) AS ventas_previas,
            COALESCE(c.ventas_actuales, 0) - COALESCE(p.ventas_previas, 0) AS variacion_absoluta,
            CASE
                WHEN COALESCE(p.ventas_previas, 0) <= 0 THEN NULL
                ELSE ROUND(((COALESCE(c.ventas_actuales, 0) - COALESCE(p.ventas_previas, 0)) / p.ventas_previas) * 100.0, 1)
            END AS variacion_pct
        FROM current_period c
        JOIN previous_period p ON p.cod_cliente = c.cod_cliente
        WHERE COALESCE(c.ventas_actuales, 0) < COALESCE(p.ventas_previas, 0)
        ORDER BY variacion_absoluta ASC, variacion_pct ASC NULLS LAST
        LIMIT :limit
        """
    )

    with engine.begin() as connection:
        rows = connection.execute(
            sql,
            {
                "current_start": current_start,
                "current_end": current_end,
                "previous_start": previous_start,
                "previous_end": previous_end,
                "store_code": store_code,
                "limit": limit,
            },
        ).mappings().all()
    return [dict(row) for row in rows], period_label


def _build_client_decline_summary(rows: list[dict], period_label: str, limit: int) -> str:
    total_drop = sum(abs(float(row.get("variacion_absoluta") or 0)) for row in rows)
    lines = [f"Mayor decrecimiento de clientes en {period_label}: {len(rows)} clientes concentran una caída acumulada de {_format_currency(total_drop)} frente al mismo corte del año anterior."]
    for row in rows[: min(limit, 5)]:
        lines.append(
            f"- {str(row.get('nombre_cliente') or row.get('cod_cliente') or 'Cliente').title()} | actual {_format_currency(row.get('ventas_actuales'))} | previo {_format_currency(row.get('ventas_previas'))} | caída {_format_currency(abs(float(row.get('variacion_absoluta') or 0)))} | variación {_format_percent(row.get('variacion_pct'))}"
        )
    lines.append("Si quieres el detalle completo, te lo envío por correo con Excel.")
    return "\n".join(lines)


def _humanize_health_status(status: Any) -> str:
    mapping = {
        "sin_movimiento": "Sin movimiento",
        "sobrestock": "Sobrestock",
        "quiebre_critico": "Quiebre crítico",
        "reposicion_recomendada": "Reposición recomendada",
    }
    return mapping.get(str(status or "").strip(), str(status or "").replace("_", " ").strip().title())


def handle_consultar_indicadores_internos(engine, args: dict, conversation_context: Optional[dict]) -> str:
    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    if not internal_auth.get("user_id"):
        return "No hay sesión interna válida para consultar indicadores."

    employee_context = internal_auth.get("employee_context") or {}
    store_code = _resolve_store_code(args.get("almacen") or employee_context.get("store_code"))
    vendor_code = _resolve_vendor_code(args.get("vendedor_codigo"), internal_auth)
    query_type = _normalize_text(args.get("tipo_consulta") or "")
    limit = _clamp_limit(args.get("limite"), default=5, minimum=3, maximum=20)

    try:
        if query_type == "proyeccion_ventas_mes":
            projection = _fetch_sales_projection(engine, store_code)
            return _build_sales_projection_summary(projection, store_code)

        if query_type == "inventario_baja_rotacion":
            rows = _fetch_inventory_rows(engine, ["sin_movimiento", "sobrestock"], store_code, limit)
            if not rows:
                return "No encontré referencias quedadas o de baja rotación para ese filtro."
            return _build_inventory_indicator_summary("Baja rotación", rows, limit)

        if query_type == "cartera_vencida_resumen":
            rows = _fetch_cartera_rows(engine, limit)
            if not rows:
                return "No encontré clientes vencidos para ese corte."
            return _build_cartera_indicator_summary(rows, limit)

        if query_type == "quiebres_stock":
            rows = _fetch_inventory_rows(engine, ["quiebre_critico"], store_code, limit)
            if not rows:
                return "No encontré quiebres críticos para ese filtro."
            return _build_inventory_indicator_summary("Quiebres críticos", rows, limit, include_suggested=True)

        if query_type == "sobrestock":
            rows = _fetch_inventory_rows(engine, ["sobrestock"], store_code, limit)
            if not rows:
                return "No encontré referencias en sobrestock para ese filtro."
            return _build_inventory_indicator_summary("Sobrestock", rows, limit)

        if query_type == "clientes_mayor_decrecimiento":
            rows, period_label = _fetch_client_decline_rows(engine, args.get("periodo"), store_code, limit)
            if not rows:
                return f"No encontré clientes con decrecimiento para {period_label}."
            return _build_client_decline_summary(rows, period_label, limit)

        if query_type == "clientes_a_reactivar":
            rows, period_label = _fetch_clients_without_purchase_rows(engine, args.get("periodo") or "este mes", store_code, vendor_code, limit)
            if not rows:
                return f"No encontré clientes claros para reactivar en {period_label}."
            return _build_clients_without_purchase_summary(rows, period_label, limit, "Clientes a reactivar")

        if query_type == "clientes_sin_compra_periodo":
            rows, period_label = _fetch_clients_without_purchase_rows(engine, args.get("periodo") or "este mes", store_code, vendor_code, limit)
            if not rows:
                return f"No encontré clientes sin compra para {period_label}."
            return _build_clients_without_purchase_summary(rows, period_label, limit, "Clientes sin compra")

        if query_type == "productos_no_vendidos_periodo":
            rows, period_label = _fetch_products_without_sale_rows(engine, args.get("periodo") or "este mes", store_code, vendor_code, limit)
            if not rows:
                return f"No encontré productos con historial que hayan dejado de venderse en {period_label}."
            return _build_products_without_sale_summary(rows, period_label, limit)

        if query_type == "productos_a_impulsar":
            rows, period_label = _fetch_products_to_push_rows(engine, args.get("periodo") or "este mes", store_code, vendor_code, limit)
            if not rows:
                return f"No encontré productos claros para impulsar en {period_label}."
            return _build_products_to_push_summary(rows, period_label, limit)
    except SQLAlchemyError as exc:
        logger.warning("consultar_indicadores_internos failed: %s", exc)
        return "No pude consultar los indicadores internos. Verifica que esté aplicado backend/internal_agent_ops.sql."

    return "No reconozco ese indicador interno. Usa proyección, baja rotación, cartera vencida, quiebres, sobrestock, clientes con decrecimiento, clientes a reactivar, clientes sin compra, productos no vendidos o productos a impulsar."


def handle_consultar_bi_universal(engine, args: dict, conversation_context: Optional[dict]) -> str:
    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    if not internal_auth.get("user_id"):
        return "No hay sesión interna válida para consultar BI."

    question = str(args.get("pregunta") or args.get("consulta") or "").strip()
    if not question:
        return "Falta la pregunta BI a analizar."

    plan = _infer_universal_bi_plan(question, args.get("periodo"), args.get("limite"))
    logger.info("consultar_bi_universal: question=%r plan=%s store=%s vendor=%s", question[:120], json.dumps(plan, ensure_ascii=False, default=str)[:300], store_code, vendor_code)
    store_code, vendor_code, scope_error = _build_sales_scope_filters(question, args, internal_auth)
    if scope_error:
        return scope_error

    if plan.get("kind") == "indicator":
        indicator_args = dict(args or {})
        indicator_args.update(
            {
                "tipo_consulta": plan.get("tipo_consulta"),
                "periodo": args.get("periodo") or plan.get("periodo"),
                "limite": args.get("limite") or plan.get("limite"),
            }
        )
        if store_code:
            indicator_args["almacen"] = store_code
        if vendor_code:
            indicator_args["vendedor_codigo"] = vendor_code
        return handle_consultar_indicadores_internos(engine, indicator_args, conversation_context)

    if plan.get("kind") == "semantic":
        analysis = str(plan.get("analysis") or "")
        logger.info("consultar_bi_universal semantic plan: analysis=%s dimension=%s period=%s", analysis, plan.get("dimension"), plan.get("periodo"))
        try:
            if analysis == "participacion":
                rows, period_label = _fetch_sales_share_rows(
                    engine,
                    args.get("periodo") or plan.get("periodo"),
                    store_code,
                    vendor_code,
                    str(plan.get("dimension") or "linea"),
                    int(plan.get("limite") or 10),
                    str(plan.get("direction") or "desc"),
                )
                return _build_sales_share_summary(rows, period_label, str(plan.get("dimension") or "linea"), int(plan.get("limite") or 10))

            if analysis == "crecimiento":
                rows, period_label, comparison_label = _fetch_sales_growth_rows(
                    engine,
                    args.get("periodo") or plan.get("periodo"),
                    store_code,
                    vendor_code,
                    str(plan.get("dimension") or "linea"),
                    int(plan.get("limite") or 10),
                    str(plan.get("direction") or "desc"),
                    str(plan.get("comparison") or "vs_anio_anterior"),
                )
                return _build_sales_growth_summary(rows, period_label, comparison_label, str(plan.get("dimension") or "linea"), int(plan.get("limite") or 10), str(plan.get("direction") or "desc"))

            if analysis == "caida_frecuencia":
                rows, period_label, comparison_label = _fetch_client_frequency_drop_rows(
                    engine,
                    args.get("periodo") or plan.get("periodo"),
                    store_code,
                    vendor_code,
                    int(plan.get("limite") or 10),
                    str(plan.get("comparison") or "vs_anio_anterior"),
                )
                return _build_client_frequency_drop_summary(rows, period_label, comparison_label, int(plan.get("limite") or 10))

            if analysis == "concentracion_cartera":
                rows = _fetch_cartera_concentration_rows(
                    engine,
                    str(plan.get("dimension") or "cliente"),
                    int(plan.get("limite") or 10),
                )
                return _build_cartera_concentration_summary(rows, str(plan.get("dimension") or "cliente"), int(plan.get("limite") or 10))

            if analysis == "oportunidades_dimension":
                rows, period_label, comparison_label = _fetch_opportunity_dimension_rows(
                    engine,
                    args.get("periodo") or plan.get("periodo"),
                    str(plan.get("dimension") or "tienda"),
                    store_code,
                    vendor_code,
                    int(plan.get("limite") or 10),
                    str(plan.get("comparison") or "vs_anio_anterior"),
                )
                return _build_opportunity_dimension_summary(rows, period_label, comparison_label, str(plan.get("dimension") or "tienda"), int(plan.get("limite") or 10))
        except SQLAlchemyError as exc:
            logger.error("consultar_bi_universal semantic analysis=%s FAILED: %s", analysis, exc, exc_info=True)
            return f"No pude resolver el análisis '{analysis}' en este momento. Error de consulta a base de datos."

    dimension = plan.get("dimension")
    try:
        if dimension:
            rows, period_label = _fetch_sales_dimension_rows(
                engine,
                args.get("periodo") or plan.get("periodo"),
                store_code,
                vendor_code,
                dimension,
                int(plan.get("limite") or 10),
                str(plan.get("direction") or "desc"),
            )
            return _build_sales_dimension_summary(
                question,
                rows,
                period_label,
                dimension,
                int(plan.get("limite") or 10),
                str(plan.get("direction") or "desc"),
            )

        snapshot = _fetch_sales_total_snapshot(
            engine,
            args.get("periodo") or plan.get("periodo"),
            store_code,
            vendor_code,
        )
        return _build_sales_total_summary(snapshot, store_code, vendor_code)
    except SQLAlchemyError as exc:
        logger.warning("consultar_bi_universal failed: %s", exc)
        return "No pude consultar BI universal en este momento. Revisa la estructura de datos interna y vuelve a intentarlo."


def _build_inventory_summary_metrics(rows: list[dict]) -> list[tuple[str, str]]:
    total_rows = len(rows)
    total_value = sum(float(row.get("inventory_value") or 0) for row in rows)
    total_stock = sum(float(row.get("stock_total") or 0) for row in rows)
    return [
        ("Referencias incluidas", _format_number(total_rows)),
        ("Unidades acumuladas", _format_number(total_stock)),
        ("Valor inventario", _format_currency(total_value)),
    ]


def _build_cartera_summary_metrics(rows: list[dict]) -> list[tuple[str, str]]:
    saldo_total = sum(float(row.get("balance_total") or 0) for row in rows)
    saldo_90 = sum(float(row.get("balance_91_plus") or 0) for row in rows)
    return [
        ("Clientes incluidos", _format_number(len(rows))),
        ("Saldo total", _format_currency(saldo_total)),
        ("Saldo >90 días", _format_currency(saldo_90)),
    ]


def _extract_sales_report_payload(report_type: str, args: dict, conversation_context: Optional[dict], sales_query_fn) -> dict:
    if sales_query_fn is None:
        raise ValueError("sales query function not available")

    report_to_desglose = {
        "ventas_detalladas": "total",
        "ventas_por_tienda": "por_tienda",
        "ventas_por_vendedor": "por_vendedor",
        "ventas_por_producto": "por_producto",
        "ventas_por_cliente": "por_cliente",
        "ventas_por_dia": "por_dia",
        "ventas_por_canal": "por_canal",
    }
    desglose = report_to_desglose.get(report_type)
    if not desglose:
        raise ValueError("sales report type not supported")

    raw_payload = sales_query_fn(
        {
            "periodo": args.get("periodo") or "este mes",
            "tienda": args.get("almacen"),
            "desglose": desglose,
            "canal": args.get("canal") or "empresa",
            "tipo_venta": args.get("tipo_venta") or "todos",
            "vendedor_codigo": args.get("vendedor_codigo"),
        },
        dict(conversation_context or {}),
    )
    return json.loads(raw_payload or "{}")


def _build_sales_report_dataset(report_type: str, payload: dict) -> tuple[str, list[tuple[str, str]], list[str], list[dict], list[str], str]:
    ventas = dict(payload.get("ventas") or {})
    metrics = [
        ("Periodo", str(payload.get("periodo") or "N/D")),
        ("Tienda", str(payload.get("tienda") or "Todas")),
        ("Canal", str(payload.get("canal") or "Todos")),
        ("Ventas netas", _format_currency(ventas.get("ventas_netas"))),
        ("Facturación bruta", _format_currency(ventas.get("facturas_bruto"))),
        ("Devoluciones", _format_currency(ventas.get("devoluciones_notas_credito"))),
        ("Clientes distintos", _format_number(ventas.get("num_clientes_distintos"))),
        ("Vendedores", _format_number(ventas.get("num_vendedores"))),
    ]
    vs_prev = payload.get("vs_anio_anterior") or {}
    if vs_prev:
        metrics.append(("Variación vs año anterior", _format_percent(vs_prev.get("variacion_pct"))))

    if report_type == "ventas_detalladas":
        title = "Reporte Ejecutivo de Ventas"
        headers = ["Indicador", "Valor"]
        rows = [{"indicador": label, "valor": value} for label, value in metrics]
        keys = ["indicador", "valor"]
        return title, metrics, headers, rows, keys, "Resumen ventas"

    mapping = {
        "ventas_por_tienda": (
            "Reporte de Ventas por Tienda",
            ["Tienda", "Facturado", "Devoluciones", "Neto", "Líneas", "Clientes", "Vendedores"],
            payload.get("desglose_tiendas") or [],
            ["tienda", "facturado", "devoluciones", "neto", "lineas", "clientes", "vendedores"],
            "Detalle tiendas",
        ),
        "ventas_por_vendedor": (
            "Reporte de Ventas por Vendedor",
            ["Vendedor", "Código", "Canal", "Facturado", "Devoluciones", "Neto", "Líneas", "Clientes"],
            payload.get("desglose_vendedores") or [],
            ["vendedor", "codigo", "canal", "facturado", "devoluciones", "neto", "lineas", "clientes"],
            "Detalle vendedores",
        ),
        "ventas_por_producto": (
            "Reporte de Ventas por Producto",
            ["Producto", "Línea", "Marca", "Total", "Unidades"],
            payload.get("top_productos") or [],
            ["producto", "linea", "marca", "total", "unidades"],
            "Detalle productos",
        ),
        "ventas_por_cliente": (
            "Reporte de Ventas por Cliente",
            ["Cliente", "Total"],
            payload.get("top_clientes") or [],
            ["cliente", "total"],
            "Detalle clientes",
        ),
        "ventas_por_dia": (
            "Reporte de Ventas por Día",
            ["Fecha", "Facturado", "Líneas"],
            payload.get("desglose_dias") or [],
            ["fecha", "facturado", "lineas"],
            "Detalle diario",
        ),
        "ventas_por_canal": (
            "Reporte de Ventas por Canal",
            ["Canal", "Facturado", "Devoluciones", "Neto", "Líneas", "Clientes", "Vendedores"],
            payload.get("desglose_canales") or [],
            ["canal", "facturado", "devoluciones", "neto", "lineas", "clientes", "vendedores"],
            "Detalle canales",
        ),
    }
    title, headers, rows, keys, detail_sheet = mapping[report_type]
    return title, metrics, headers, list(rows), keys, detail_sheet


def _report_rows_and_headers(report_type: str, engine, store_code: Optional[str], limit: int, conversation_context: Optional[dict], sales_query_fn=None) -> tuple[str, list[tuple[str, str]], list[str], list[dict], list[str], str]:
    pending_args = (conversation_context or {}).get("pending_tool_args") or {}
    internal_auth = (conversation_context or {}).get("internal_auth") or {}
    vendor_code = _resolve_vendor_code(pending_args.get("vendedor_codigo"), internal_auth)
    if report_type == "inventario_baja_rotacion":
        rows = _fetch_inventory_rows(engine, ["sin_movimiento", "sobrestock"], store_code, limit)
        headers = ["Almacen", "Referencia", "Descripcion", "Estado", "Stock", "Historial ventas", "Dias sin venta", "Valor inventario"]
        keys = ["almacen_nombre", "referencia", "descripcion", "health_status", "stock_total", "historial_ventas_metric", "dias_sin_venta", "inventory_value"]
        title = "Reporte de Baja Rotación y Sobrestock"
        for row in rows:
            row["health_status"] = _humanize_health_status(row.get("health_status"))
        metrics = _build_inventory_summary_metrics(rows)
        detail_sheet = "Detalle inventario"
    elif report_type == "cartera_vencida":
        rows = _fetch_cartera_rows(engine, limit)
        headers = ["Cliente", "Codigo", "Vendedor", "Zona", "31-60", "61-90", ">90", "Saldo total", "Max días vencido"]
        keys = ["nombre_cliente", "cod_cliente", "nom_vendedor", "zona", "balance_31_60", "balance_61_90", "balance_91_plus", "balance_total", "max_dias_vencido"]
        title = "Reporte de Cartera Vencida"
        metrics = _build_cartera_summary_metrics(rows)
        detail_sheet = "Detalle cartera"
    elif report_type == "quiebres_stock":
        rows = _fetch_inventory_rows(engine, ["quiebre_critico"], store_code, limit)
        headers = ["Almacen", "Referencia", "Descripcion", "Stock", "Punto de reposición", "Sugerido", "Valor inventario"]
        keys = ["almacen_nombre", "referencia", "descripcion", "stock_total", "reorder_point", "reorder_qty_recommended", "inventory_value"]
        title = "Reporte de Quiebres Críticos"
        metrics = _build_inventory_summary_metrics(rows)
        detail_sheet = "Detalle quiebres"
    elif report_type == "sobrestock":
        rows = _fetch_inventory_rows(engine, ["sobrestock"], store_code, limit)
        headers = ["Almacen", "Referencia", "Descripcion", "Stock", "Historial ventas", "Dias sin venta", "Valor inventario"]
        keys = ["almacen_nombre", "referencia", "descripcion", "stock_total", "historial_ventas_metric", "dias_sin_venta", "inventory_value"]
        title = "Reporte de Sobrestock"
        metrics = _build_inventory_summary_metrics(rows)
        detail_sheet = "Detalle sobrestock"
    elif report_type == "clientes_mayor_decrecimiento":
        rows, period_label = _fetch_client_decline_rows(engine, pending_args.get("periodo"), store_code, limit)
        headers = ["Cliente", "Codigo", "Ventas actuales", "Ventas previas", "Caida absoluta", "Variacion %"]
        keys = ["nombre_cliente", "cod_cliente", "ventas_actuales", "ventas_previas", "variacion_absoluta", "variacion_pct"]
        title = "Reporte de Clientes con Mayor Decrecimiento"
        metrics = [
            ("Periodo analizado", period_label),
            ("Clientes incluidos", _format_number(len(rows))),
            ("Caída acumulada", _format_currency(sum(abs(float(row.get("variacion_absoluta") or 0)) for row in rows))),
        ]
        detail_sheet = "Detalle decrecimiento"
    elif report_type == "clientes_a_reactivar":
        rows, period_label = _fetch_clients_without_purchase_rows(engine, pending_args.get("periodo") or "este mes", store_code, vendor_code, limit)
        headers = ["Cliente", "Codigo", "Vendedor", "Ultima compra", "Dias sin compra", "Ventas historicas", "Meses activos"]
        keys = ["nombre_cliente", "cod_cliente", "nom_vendedor", "ultima_compra", "dias_sin_compra", "ventas_historicas", "meses_activos"]
        title = "Reporte de Clientes a Reactivar"
        metrics = [
            ("Periodo analizado", period_label),
            ("Clientes incluidos", _format_number(len(rows))),
            ("Base histórica comprometida", _format_currency(sum(float(row.get("ventas_historicas") or 0) for row in rows))),
        ]
        detail_sheet = "Detalle reactivacion"
    elif report_type == "clientes_sin_compra_periodo":
        rows, period_label = _fetch_clients_without_purchase_rows(engine, pending_args.get("periodo") or "este mes", store_code, vendor_code, limit)
        headers = ["Cliente", "Codigo", "Vendedor", "Ultima compra", "Dias sin compra", "Ventas historicas", "Meses activos"]
        keys = ["nombre_cliente", "cod_cliente", "nom_vendedor", "ultima_compra", "dias_sin_compra", "ventas_historicas", "meses_activos"]
        title = "Reporte de Clientes sin Compra"
        metrics = [
            ("Periodo analizado", period_label),
            ("Clientes incluidos", _format_number(len(rows))),
            ("Base histórica comprometida", _format_currency(sum(float(row.get("ventas_historicas") or 0) for row in rows))),
        ]
        detail_sheet = "Detalle sin compra"
    elif report_type == "productos_no_vendidos_periodo":
        rows, period_label = _fetch_products_without_sale_rows(engine, pending_args.get("periodo") or "este mes", store_code, vendor_code, limit)
        headers = ["Almacen", "Referencia", "Descripcion", "Stock", "Dias sin venta", "Ventas historicas", "Valor inventario"]
        keys = ["almacen_nombre", "referencia", "descripcion", "stock_total", "dias_sin_venta", "ventas_historicas", "inventory_value"]
        title = "Reporte de Productos sin Venta en el Periodo"
        metrics = [
            ("Periodo analizado", period_label),
            ("Referencias incluidas", _format_number(len(rows))),
            ("Base histórica comprometida", _format_currency(sum(float(row.get("ventas_historicas") or 0) for row in rows))),
        ]
        detail_sheet = "Detalle productos"
    elif report_type == "productos_a_impulsar":
        rows, period_label = _fetch_products_to_push_rows(engine, pending_args.get("periodo") or "este mes", store_code, vendor_code, limit)
        headers = ["Almacen", "Referencia", "Descripcion", "Stock", "Ventas actuales", "Promedio base", "Brecha oportunidad", "Clientes objetivo"]
        keys = ["almacen_nombre", "referencia", "descripcion", "stock_total", "ventas_actuales", "promedio_base", "brecha_oportunidad", "clientes_objetivo"]
        title = "Reporte de Productos a Impulsar"
        metrics = [
            ("Periodo analizado", period_label),
            ("Referencias incluidas", _format_number(len(rows))),
            ("Brecha comercial estimada", _format_currency(sum(float(row.get("brecha_oportunidad") or 0) for row in rows))),
        ]
        detail_sheet = "Detalle impulso"
    elif report_type in {
        "ventas_detalladas",
        "ventas_por_tienda",
        "ventas_por_vendedor",
        "ventas_por_producto",
        "ventas_por_cliente",
        "ventas_por_dia",
        "ventas_por_canal",
    }:
        sales_payload = _extract_sales_report_payload(report_type, conversation_context.get("pending_tool_args") or {}, conversation_context, sales_query_fn)
        title, metrics, headers, rows, keys, detail_sheet = _build_sales_report_dataset(report_type, sales_payload)
    else:
        raise ValueError("tipo de reporte no soportado")
    return title, metrics, headers, rows, keys, detail_sheet


def _autosize_sheet_columns(sheet):
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 38)


def _style_summary_sheet(sheet, title: str, metrics: list[tuple[str, str]], subtitle_lines: list[str]):
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:D1")
    hero = sheet["A1"]
    hero.value = title
    hero.fill = _XL_DARK_FILL
    hero.font = Font(color="FFFFFF", size=18, bold=True)
    hero.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    row_idx = 3
    for line in subtitle_lines:
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        cell = sheet.cell(row=row_idx, column=1, value=line)
        cell.font = Font(color="374151", size=11)
        cell.alignment = Alignment(wrap_text=True)
        row_idx += 1

    row_idx += 1
    for label, value in metrics:
        label_cell = sheet.cell(row=row_idx, column=1, value=label)
        value_cell = sheet.cell(row=row_idx, column=2, value=value)
        label_cell.fill = _XL_LIGHT_FILL
        value_cell.fill = _XL_LIGHT_FILL
        label_cell.font = Font(bold=True, color="111827")
        value_cell.font = Font(color="111827")
        label_cell.border = _XL_BORDER
        value_cell.border = _XL_BORDER
        row_idx += 1
    sheet.freeze_panes = "A3"
    _autosize_sheet_columns(sheet)


def _infer_number_format(key: str) -> Optional[str]:
    lowered = str(key or "").lower()
    if any(token in lowered for token in ["valor", "saldo", "neto", "facturado", "devoluciones", "total", "precio", "ventas", "promedio", "brecha"]):
        return '$#,##0'
    if "variacion" in lowered:
        return '0.0%'
    if any(token in lowered for token in ["stock", "unidades", "lineas", "clientes", "vendedores", "dias", "sugerido", "punto"]):
        return '#,##0'
    return None


def _style_detail_sheet(sheet, headers: list[str], keys: list[str], rows: list[dict]):
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = _XL_DARK_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _XL_BORDER

    for row_index, row in enumerate(rows, start=2):
        fill = _XL_LIGHT_FILL if row_index % 2 == 0 else None
        for col_index, key in enumerate(keys, start=1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _XL_BORDER
            if fill:
                cell.fill = fill
            number_format = _infer_number_format(key)
            if number_format and isinstance(cell.value, Number):
                if number_format == '0.0%' and abs(float(cell.value or 0)) > 1:
                    cell.value = float(cell.value) / 100.0
                cell.number_format = number_format

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"
    _autosize_sheet_columns(sheet)


def _build_excel_attachment(title: str, metrics: list[tuple[str, str]], headers: list[str], rows: list[dict], keys: list[str], detail_sheet_name: str, subtitle_lines: list[str]) -> tuple[str, bytes]:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    _style_summary_sheet(summary_sheet, title, metrics, subtitle_lines)

    sheet = workbook.create_sheet(detail_sheet_name[:31])
    sheet.append(headers)
    for row in rows:
        values = [row.get(key) for key in keys]
        sheet.append(values)
    _style_detail_sheet(sheet, headers, keys, rows)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"{re.sub(r'[^a-z0-9]+', '_', title.lower()).strip('_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return filename, buffer.getvalue()


def _build_html_preview_rows(headers: list[str], rows: list[dict], keys: list[str]) -> str:
    preview_rows = rows[:5]
    if not preview_rows:
        return ""
    header_html = "".join(f"<th style='padding:10px 12px;text-align:left;border-bottom:1px solid #e5e7eb;'>{escape(str(header))}</th>" for header in headers)
    body_html = []
    for row in preview_rows:
        cols = []
        for key in keys:
            value = row.get(key)
            value_text = _format_preview_value(key, value)
            cols.append(f"<td style='padding:10px 12px;border-bottom:1px solid #f3f4f6;'>{escape(value_text)}</td>")
        body_html.append(f"<tr>{''.join(cols)}</tr>")
    return (
        "<table style='width:100%;border-collapse:collapse;background:#ffffff;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;'>"
        f"<thead><tr style='background:#111827;color:#ffffff;'>{header_html}</tr></thead>"
        f"<tbody>{''.join(body_html)}</tbody></table>"
    )


def handle_enviar_reporte_interno_correo(engine, args: dict, conversation_context: Optional[dict], build_brand_email_shell_fn, send_email_fn, sales_query_fn=None) -> str:
    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    if not internal_auth.get("user_id"):
        return "No hay sesión interna válida para enviar reportes por correo."

    report_type = _normalize_text(args.get("tipo_reporte") or "")
    employee_context = internal_auth.get("employee_context") or {}
    destination_email = str(args.get("email_destino") or internal_auth.get("email") or "").strip().lower()
    if not _is_valid_email(destination_email):
        return "No tengo un correo destino válido. Pídele al colaborador el correo y luego reintenta el envío."
    store_code = _resolve_store_code(args.get("almacen") or employee_context.get("store_code"))
    limit = _clamp_limit(args.get("limite"), default=100, minimum=10, maximum=500)
    context_with_args = dict(conversation_context or {})
    context_with_args["pending_tool_args"] = dict(args or {})

    try:
        title, metrics, headers, rows, keys, detail_sheet_name = _report_rows_and_headers(
            report_type,
            engine,
            store_code,
            limit,
            context_with_args,
            sales_query_fn=sales_query_fn,
        )
    except ValueError:
        return "No reconozco ese tipo de reporte interno para correo."
    except SQLAlchemyError as exc:
        logger.warning("email report query failed: %s", exc)
        return "No pude armar el reporte interno. Verifica que esté aplicado backend/internal_agent_ops.sql."

    if not rows:
        return "No encontré datos para ese reporte."

    subtitle_lines = [
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Sede consultada: {requested_by.get('sede') if (requested_by := internal_auth.get('employee_context') or {}).get('sede') else store_code or 'Consolidado'}",
        f"Filas incluidas en detalle: {len(rows)}",
    ]
    attachment_name, attachment_bytes = _build_excel_attachment(title, metrics, headers, rows, keys, detail_sheet_name, subtitle_lines)
    requested_by = internal_auth.get("employee_context") or {}
    preview_table = _build_html_preview_rows(headers, rows, keys)
    metrics_html = "".join(
        f"<tr><td style='padding:10px 12px;border-bottom:1px solid #f3f4f6;font-weight:600;color:#111827;'>{escape(label)}</td>"
        f"<td style='padding:10px 12px;border-bottom:1px solid #f3f4f6;color:#111827;'>{escape(value)}</td></tr>"
        for label, value in metrics[:8]
    )
    body_html = (
        "<p style='margin:0 0 14px 0;font-size:15px;'>Se generó un reporte interno solicitado desde WhatsApp.</p>"
        + "<div style='background:#ffffff;border:1px solid #e5e7eb;border-radius:14px;padding:18px 20px;margin-bottom:18px;'>"
        + f"<p style='margin:0 0 8px 0;'><strong>Solicitante:</strong> {escape(str(requested_by.get('full_name') or internal_auth.get('username') or 'Colaborador Ferreinox'))}</p>"
        + f"<p style='margin:0 0 8px 0;'><strong>Cargo:</strong> {escape(str(requested_by.get('cargo') or internal_auth.get('role') or 'Perfil interno'))}</p>"
        + f"<p style='margin:0 0 8px 0;'><strong>Sede:</strong> {escape(str(requested_by.get('sede') or store_code or 'Consolidado'))}</p>"
        + f"<p style='margin:0;'><strong>Filas incluidas:</strong> {len(rows)}</p>"
        + "</div>"
        + "<div style='background:#ffffff;border:1px solid #e5e7eb;border-radius:14px;padding:0;margin-bottom:18px;overflow:hidden;'>"
        + "<div style='background:#f59e0b;color:#111827;padding:12px 16px;font-weight:700;'>Resumen ejecutivo</div>"
        + f"<table style='width:100%;border-collapse:collapse;'>{metrics_html}</table>"
        + "</div>"
        + (preview_table if preview_table else "")
        + "<p style='margin:18px 0 14px 0;'>Adjunto encontrarás el Excel con portada ejecutiva, resumen y detalle estructurado del reporte.</p>"
    )
    html_content = build_brand_email_shell_fn(title, body_html)
    text_content = (
        f"{title}\n"
        f"Solicitante: {requested_by.get('full_name') or internal_auth.get('username') or 'Colaborador Ferreinox'}\n"
        f"Cargo: {requested_by.get('cargo') or internal_auth.get('role') or 'Perfil interno'}\n"
        f"Sede: {requested_by.get('sede') or store_code or 'Consolidado'}\n"
        f"Filas incluidas: {len(rows)}\n"
        + "\n".join(f"{label}: {value}" for label, value in metrics[:6])
        + "\nSe adjunta Excel con resumen ejecutivo y detalle completo."
    )

    try:
        send_email_fn(
            destination_email,
            title,
            html_content,
            text_content,
            attachments=[
                {
                    "content": base64.b64encode(attachment_bytes).decode("ascii"),
                    "filename": attachment_name,
                    "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "disposition": "attachment",
                }
            ],
        )
    except Exception as exc:
        logger.warning("send internal report email failed: %s", exc)
        return f"No pude enviar el correo a {destination_email}."

    return f"Reporte enviado a {destination_email} con adjunto {attachment_name}."


def _log_action(engine, internal_auth: Optional[dict], role_key: str, tool_name: str, status: str, summary: str, payload: dict):
    if not engine:
        return
    try:
        with engine.begin() as connection:
            connection.execute(
                text(
                    """
                    INSERT INTO public.agent_internal_action_log (
                        user_id, role_key, tool_name, action_status, action_summary, payload
                    ) VALUES (
                        :user_id, :role_key, :tool_name, :action_status, :action_summary, CAST(:payload AS jsonb)
                    )
                    """
                ),
                {
                    "user_id": (internal_auth or {}).get("user_id"),
                    "role_key": role_key,
                    "tool_name": tool_name,
                    "action_status": status,
                    "action_summary": summary,
                    "payload": json.dumps(payload or {}, ensure_ascii=False, default=str),
                },
            )
    except SQLAlchemyError as exc:
        logger.debug("action log skipped for %s: %s", tool_name, exc)


def handle_crear_recordatorio_interno(engine, args: dict, conversation_context: Optional[dict]) -> str:
    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    if not internal_auth.get("user_id"):
        return "No hay sesion interna valida para crear recordatorios."

    title = str(args.get("titulo") or "").strip()
    if not title:
        return "Falta el titulo del recordatorio interno."

    role_profile = _fetch_role_profile(engine, internal_auth)
    role_key = str(args.get("role_profile_key") or role_profile.get("role_key") or "empleado_operativo").strip()
    priority = _normalize_text(args.get("prioridad") or "media")
    if priority not in {"alta", "media", "baja"}:
        priority = "media"
    due_at = str(args.get("fecha_hora") or "").strip() or None
    detail = str(args.get("detalle") or "").strip() or None

    try:
        with engine.begin() as connection:
            reminder_id = connection.execute(
                text(
                    """
                    INSERT INTO public.agent_internal_reminder (
                        owner_user_id, role_key, title, detail, due_at, priority, status, metadata
                    ) VALUES (
                        :owner_user_id,
                        :role_key,
                        :title,
                        :detail,
                        CASE WHEN :due_at IS NULL OR :due_at = '' THEN NULL ELSE CAST(:due_at AS timestamptz) END,
                        :priority,
                        'pendiente',
                        CAST(:metadata AS jsonb)
                    )
                    RETURNING id
                    """
                ),
                {
                    "owner_user_id": internal_auth.get("user_id"),
                    "role_key": role_key,
                    "title": title,
                    "detail": detail,
                    "due_at": due_at,
                    "priority": priority,
                    "metadata": "{}",
                },
            ).scalar_one()
        _log_action(engine, internal_auth, role_key, "crear_recordatorio_interno", "ok", f"Recordatorio #{reminder_id}: {title}", args)
        due_note = f" | vence {due_at}" if due_at else ""
        return f"Recordatorio interno creado: #{reminder_id} | {title} [{priority}]{due_note}"
    except SQLAlchemyError as exc:
        logger.warning("crear_recordatorio_interno failed: %s", exc)
        return "No pude crear el recordatorio. Verifica que este aplicado backend/internal_agent_ops.sql y que la fecha tenga formato ISO."


def handle_generar_lista_pendientes(engine, args: dict, conversation_context: Optional[dict]) -> str:
    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    role_profile = _fetch_role_profile(engine, internal_auth)
    scope = _normalize_text(args.get("alcance") or "mixto")
    if scope not in {"personal", "equipo", "mixto"}:
        scope = "mixto"
    try:
        limit = int(args.get("limite") or 8)
    except (TypeError, ValueError):
        limit = 8
    limit = max(1, min(limit, 20))

    if not engine:
        return "No hay acceso a la base para listar pendientes."

    try:
        with engine.begin() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT item_type, item_id, title, summary, due_at, priority, owner_user_id, role_key
                    FROM public.vw_internal_pending_queue
                    WHERE (
                        :scope = 'personal' AND owner_user_id = :user_id
                    ) OR (
                        :scope = 'equipo' AND role_key = :role_key
                    ) OR (
                        :scope = 'mixto' AND (owner_user_id = :user_id OR role_key = :role_key)
                    )
                    ORDER BY priority_rank ASC, due_at NULLS LAST, created_at DESC
                    LIMIT :limit
                    """
                ),
                {
                    "scope": scope,
                    "user_id": internal_auth.get("user_id"),
                    "role_key": role_profile.get("role_key") or "empleado_operativo",
                    "limit": limit,
                },
            ).mappings().all()
    except SQLAlchemyError as exc:
        logger.warning("generar_lista_pendientes failed: %s", exc)
        return "No pude consultar la lista de pendientes. Verifica que este aplicado backend/internal_agent_ops.sql."

    if not rows:
        _log_action(engine, internal_auth, role_profile.get("role_key") or "empleado_operativo", "generar_lista_pendientes", "ok", "Sin pendientes", args)
        return "No hay pendientes abiertos para ese alcance."

    lines = [f"Pendientes ({scope}):"]
    for row in rows:
        due_at = row.get("due_at")
        due_note = f" | vence {due_at:%Y-%m-%d %H:%M}" if due_at else ""
        lines.append(
            f"- {row.get('title')} [{row.get('priority')}] ({row.get('item_type')}){due_note}"
        )
    _log_action(
        engine,
        internal_auth,
        role_profile.get("role_key") or "empleado_operativo",
        "generar_lista_pendientes",
        "ok",
        f"{len(rows)} pendientes listados",
        args,
    )
    return "\n".join(lines)


def handle_sugerir_reposicion_bodega(engine, args: dict, conversation_context: Optional[dict]) -> str:
    internal_auth = dict((conversation_context or {}).get("internal_auth") or {})
    role_profile = _fetch_role_profile(engine, internal_auth)
    employee_context = internal_auth.get("employee_context") or {}
    store_code = re.sub(r"\D", "", str(args.get("almacen") or employee_context.get("store_code") or "")).strip() or None
    health_status = _normalize_text(args.get("estado") or "reposicion_recomendada")
    if health_status not in {"quiebre_critico", "reposicion_recomendada", "sobrestock", "sin_movimiento"}:
        health_status = "reposicion_recomendada"
    try:
        limit = int(args.get("limite") or 10)
    except (TypeError, ValueError):
        limit = 10
    limit = max(1, min(limit, 20))

    if not engine:
        return "No hay acceso a la base para sugerir reposicion."

    try:
        with engine.begin() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT
                        cod_almacen,
                        almacen_nombre,
                        referencia,
                        descripcion,
                        stock_total,
                        historial_ventas_metric,
                        reorder_point,
                        reorder_qty_recommended,
                        inventory_value,
                        health_status
                    FROM public.mv_internal_inventory_health
                    WHERE health_status = :health_status
                      AND (:store_code IS NULL OR cod_almacen = :store_code)
                    ORDER BY reorder_qty_recommended DESC, historial_ventas_metric DESC, stock_total ASC
                    LIMIT :limit
                    """
                ),
                {"health_status": health_status, "store_code": store_code, "limit": limit},
            ).mappings().all()
    except SQLAlchemyError as exc:
        logger.warning("sugerir_reposicion_bodega failed: %s", exc)
        return "No pude calcular reposicion. Verifica que este aplicado backend/internal_agent_ops.sql."

    if not rows:
        return "No encontré referencias para ese criterio de reposición."

    title = f"Sugerencias de {health_status}"
    if store_code:
        title += f" para almacen {store_code}"
    lines = [title + ":"]
    for row in rows:
        qty = _format_number(row.get("reorder_qty_recommended"))
        stock = _format_number(row.get("stock_total"))
        demand = _format_number(row.get("historial_ventas_metric"))
        lines.append(
            f"- {row.get('referencia')} | {row.get('descripcion')} | stock {stock} | hist {demand} | sugerido {qty}"
        )

    _log_action(
        engine,
        internal_auth,
        role_profile.get("role_key") or "empleado_operativo",
        "sugerir_reposicion_bodega",
        "ok",
        f"{len(rows)} referencias sugeridas",
        args,
    )
    return "\n".join(lines)