#!/usr/bin/env python3
"""
Import precios.xlsx and CLIENTES.xlsx into PostgreSQL tables.
Tables: agent_precios, agent_clientes
Uses psycopg2 execute_values for fast bulk insert.
"""
import os
import sys
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:o5S3X9VIYcbBWqd525hqT24UhYAc8AdjtevyHtlZHhGxJkfMQVZXReCTxkcjSOAX@192.81.216.49:3000/postgres",
)

DDL_PRECIOS = """
CREATE TABLE IF NOT EXISTS public.agent_precios (
    id SERIAL PRIMARY KEY,
    codigo INTEGER,
    descripcion_adicional TEXT,
    descripcion TEXT,
    referencia TEXT,
    codigo_barras TEXT,
    familia TEXT,
    subfamilia TEXT,
    marca TEXT,
    linea TEXT,
    cat_producto TEXT,
    aplicacion TEXT,
    departamento TEXT,
    seccion TEXT,
    sublinea TEXT,
    peso_articulo NUMERIC,
    pvp_sap NUMERIC DEFAULT 0,
    pvp_franquicia NUMERIC DEFAULT 0,
    created_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_agent_precios_referencia ON public.agent_precios(referencia);
CREATE INDEX IF NOT EXISTS idx_agent_precios_codigo ON public.agent_precios(codigo);
"""

DDL_CLIENTES = """
CREATE TABLE IF NOT EXISTS public.agent_clientes (
    id SERIAL PRIMARY KEY,
    codigo INTEGER,
    nombre TEXT,
    nif TEXT,
    direccion TEXT,
    telefono TEXT,
    poblacion TEXT,
    codigo_postal TEXT,
    provincia TEXT,
    riesgo_concedido NUMERIC,
    telefono_2 TEXT,
    tipo_documento TEXT,
    email TEXT,
    persona_contacto TEXT,
    ciudad TEXT,
    categoria TEXT,
    segmento TEXT,
    negocio TEXT,
    tipocliente2 TEXT,
    tipo_de_documento TEXT,
    nombre_1 TEXT,
    otros_nombres TEXT,
    apellido_1 TEXT,
    apellido_2 TEXT,
    razon_social TEXT,
    dv TEXT,
    clasificacion TEXT,
    created_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_agent_clientes_nif ON public.agent_clientes(nif);
CREATE INDEX IF NOT EXISTS idx_agent_clientes_codigo ON public.agent_clientes(codigo);
CREATE INDEX IF NOT EXISTS idx_agent_clientes_telefono ON public.agent_clientes(telefono);
"""


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.upper() != "NULL" else None


def safe_num(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0


def import_precios(conn, filepath):
    print(f"[PRECIOS] Loading {filepath}...", flush=True)
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    tuples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = row[0]
        if not codigo:
            continue
        tuples.append((
            int(codigo) if codigo else None,
            safe_str(row[1]),
            safe_str(row[3]),
            safe_str(row[6]),
            safe_str(row[5]),
            safe_str(row[7]) or safe_str(row[18]),
            safe_str(row[8]) or safe_str(row[19]),
            safe_str(row[9]) or safe_str(row[17]),
            safe_str(row[10]) or safe_str(row[15]),
            safe_str(row[11]),
            safe_str(row[12]),
            safe_str(row[2]) or safe_str(row[13]),
            safe_str(row[4]) or safe_str(row[14]),
            safe_str(row[16]),
            safe_num(row[20]) if row[20] else None,
            safe_num(row[21]),
            safe_num(row[23]),
        ))
    wb.close()
    print(f"[PRECIOS] Parsed {len(tuples)} rows. Inserting...", flush=True)

    cur = conn.cursor()
    cur.execute("DELETE FROM public.agent_precios")
    print(f"  Cleared existing rows.", flush=True)

    execute_values(cur, """
        INSERT INTO public.agent_precios
        (codigo, descripcion_adicional, descripcion, referencia, codigo_barras,
         familia, subfamilia, marca, linea, cat_producto, aplicacion,
         departamento, seccion, sublinea, peso_articulo, pvp_sap, pvp_franquicia)
        VALUES %s
    """, tuples, page_size=2000)

    conn.commit()
    print(f"[PRECIOS] Done. {len(tuples)} rows imported.", flush=True)
    return len(tuples)


def import_clientes(conn, filepath):
    print(f"[CLIENTES] Loading {filepath}...", flush=True)
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    tuples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = row[0]
        if not codigo:
            continue
        nombre = safe_str(row[1])
        if not nombre:
            continue
        tuples.append((
            int(codigo) if codigo else None,
            nombre,
            safe_str(row[2]),
            safe_str(row[3]),
            safe_str(row[4]),
            safe_str(row[5]),
            safe_str(row[6]),
            safe_str(row[7]),
            safe_num(row[8]) if row[8] else None,
            safe_str(row[9]),
            safe_str(row[10]),
            safe_str(row[13]),
            safe_str(row[20]),
            safe_str(row[21]),
            safe_str(row[22]),
            safe_str(row[25]) if len(row) > 25 else None,
            safe_str(row[26]) if len(row) > 26 else None,
            safe_str(row[27]) if len(row) > 27 else None,
            safe_str(row[28]) if len(row) > 28 else None,
            safe_str(row[29]) if len(row) > 29 else None,
            safe_str(row[30]) if len(row) > 30 else None,
            safe_str(row[31]) if len(row) > 31 else None,
            safe_str(row[32]) if len(row) > 32 else None,
            safe_str(row[33]) if len(row) > 33 else None,
            safe_str(row[34]) if len(row) > 34 else None,
            safe_str(row[36]) if len(row) > 36 else None,
        ))
    wb.close()
    print(f"[CLIENTES] Parsed {len(tuples)} rows. Inserting...", flush=True)

    cur = conn.cursor()
    cur.execute("DELETE FROM public.agent_clientes")
    print(f"  Cleared existing rows.", flush=True)

    execute_values(cur, """
        INSERT INTO public.agent_clientes
        (codigo, nombre, nif, direccion, telefono, poblacion, codigo_postal,
         provincia, riesgo_concedido, telefono_2, tipo_documento, email,
         persona_contacto, ciudad, categoria, segmento, negocio, tipocliente2,
         tipo_de_documento, nombre_1, otros_nombres, apellido_1, apellido_2,
         razon_social, dv, clasificacion)
        VALUES %s
    """, tuples, page_size=2000)

    conn.commit()
    print(f"[CLIENTES] Done. {len(tuples)} rows imported.", flush=True)
    return len(tuples)


if __name__ == "__main__":
    conn = psycopg2.connect(DATABASE_URL)

    # Create tables
    print("Creating tables...", flush=True)
    cur = conn.cursor()
    cur.execute(DDL_PRECIOS)
    cur.execute(DDL_CLIENTES)
    conn.commit()
    print("Tables created.", flush=True)

    # Import
    precios_file = os.path.join(os.path.dirname(__file__), "..", "precios.xlsx")
    clientes_file = os.path.join(os.path.dirname(__file__), "..", "CLIENTES.xlsx")

    if os.path.exists(precios_file):
        import_precios(conn, precios_file)
    else:
        print(f"[SKIP] {precios_file} not found")

    if os.path.exists(clientes_file):
        import_clientes(conn, clientes_file)
    else:
        print(f"[SKIP] {clientes_file} not found")

    # Verify
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM public.agent_precios")
    p_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM public.agent_clientes")
    c_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM public.agent_precios WHERE pvp_sap > 0")
    p_with_price = cur.fetchone()[0]
    print(f"\n=== RESUMEN ===")
    print(f"agent_precios: {p_count} rows ({p_with_price} con precio PVP SAP > 0)")
    print(f"agent_clientes: {c_count} rows")
    conn.close()
