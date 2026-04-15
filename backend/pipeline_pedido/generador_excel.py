"""
generador_excel.py — Generación de Excel para pedidos ICG + Dropbox
===================================================================

Genera archivos Excel profesionales con:
  - Hoja PedidoICG: REFERENCIA, CANTIDAD, PRECIO, DESCUENTO (formato ICG)
  - Hoja Detalle: información extendida para control interno
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .matcher_inventario import ResultadoMatchPedido

logger = logging.getLogger("pipeline_pedido.generador_excel")

# ============================================================================
# ESTILOS
# ============================================================================
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
_ACCENT_FILL = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_MONEY_FMT = '#,##0'
_THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


def _safe_filename(text: str, fallback: str = "SinDato") -> str:
    if not text:
        return fallback
    cleaned = re.sub(r"[^\w\s-]", "", text).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned[:40] or fallback


# ============================================================================
# GENERADOR EXCEL ICG
# ============================================================================

def generar_excel_pedido(
    match_result: ResultadoMatchPedido,
    cliente_nombre: str = "",
    notas: str = "",
) -> tuple[bytes, list[dict]]:
    """
    Genera Excel con hoja ICG + hoja Detalle.

    Retorna: (excel_bytes, filas_icg_dict)
    """
    wb = Workbook()

    # ── Hoja 1: PedidoICG (formato para importar al sistema) ──
    ws_icg = wb.active
    ws_icg.title = "PedidoICG"
    icg_headers = ["REFERENCIA", "CANTIDAD", "PRECIO", "DESCUENTO"]
    for col_idx, header in enumerate(icg_headers, 1):
        cell = ws_icg.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    icg_rows = []
    row_num = 2
    for prod in match_result.productos_resueltos:
        if not prod.codigo_encontrado:
            continue
        row_data = {
            "REFERENCIA": prod.codigo_encontrado,
            "CANTIDAD": prod.cantidad,
            "PRECIO": prod.precio_unitario,
            "DESCUENTO": prod.descuento_pct,
        }
        icg_rows.append(row_data)
        ws_icg.cell(row=row_num, column=1, value=prod.codigo_encontrado).border = _THIN_BORDER
        ws_icg.cell(row=row_num, column=2, value=prod.cantidad).border = _THIN_BORDER
        c_precio = ws_icg.cell(row=row_num, column=3, value=prod.precio_unitario)
        c_precio.number_format = _MONEY_FMT
        c_precio.border = _THIN_BORDER
        ws_icg.cell(row=row_num, column=4, value=prod.descuento_pct).border = _THIN_BORDER
        row_num += 1

    # Bicomponentes inyectados
    for bico in match_result.bicomponentes_inyectados:
        if not bico.codigo_encontrado or not bico.disponible:
            continue
        row_data = {
            "REFERENCIA": bico.codigo_encontrado,
            "CANTIDAD": bico.cantidad_sugerida,
            "PRECIO": bico.precio_unitario,
            "DESCUENTO": 0,
        }
        icg_rows.append(row_data)
        ws_icg.cell(row=row_num, column=1, value=bico.codigo_encontrado).border = _THIN_BORDER
        ws_icg.cell(row=row_num, column=2, value=bico.cantidad_sugerida).border = _THIN_BORDER
        c_precio = ws_icg.cell(row=row_num, column=3, value=bico.precio_unitario)
        c_precio.number_format = _MONEY_FMT
        c_precio.border = _THIN_BORDER
        ws_icg.cell(row=row_num, column=4, value=0).border = _THIN_BORDER
        row_num += 1

    # Ajustar ancho columnas ICG
    ws_icg.column_dimensions["A"].width = 18
    ws_icg.column_dimensions["B"].width = 12
    ws_icg.column_dimensions["C"].width = 16
    ws_icg.column_dimensions["D"].width = 14

    # ── Hoja 2: Detalle (información extendida) ──
    ws_det = wb.create_sheet("Detalle")
    det_headers = [
        "Referencia", "Descripcion", "Marca", "Presentacion",
        "Cantidad", "Precio Unit.", "Descuento %", "Subtotal",
        "Stock", "Disponible", "Tipo Match", "RAL", "Observaciones",
    ]
    for col_idx, header in enumerate(det_headers, 1):
        cell = ws_det.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # Info del pedido en fila 2 (merge)
    ws_det.merge_cells("A2:M2")
    info_cell = ws_det.cell(row=2, column=1)
    info_cell.value = (
        f"Cliente: {cliente_nombre or 'N/A'} | "
        f"Tienda: {match_result.tienda_nombre or 'N/A'} ({match_result.tienda_codigo}) | "
        f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"Productos: {match_result.total_resueltos}"
    )
    info_cell.font = Font(name="Calibri", bold=True, size=10, color="111827")
    info_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    row_num = 3
    for prod in match_result.productos_resueltos:
        subtotal = prod.precio_unitario * prod.cantidad
        if prod.descuento_pct:
            subtotal *= (1 - prod.descuento_pct / 100)
        obs = []
        if prod.es_bicomponente:
            obs.append("BICOMPONENTE")
        if prod.ral_detectado:
            obs.append(f"RAL {prod.ral_detectado}")
        if prod.linea_international:
            obs.append(f"Linea: {prod.linea_international}")

        values = [
            prod.codigo_encontrado, prod.descripcion_real, prod.marca,
            prod.presentacion_real, prod.cantidad, prod.precio_unitario,
            prod.descuento_pct, subtotal, prod.stock_disponible,
            "SI" if prod.disponible else "NO", prod.tipo_match,
            prod.ral_detectado, " | ".join(obs),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_det.cell(row=row_num, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            if col_idx in (6, 8):  # Precio, Subtotal
                cell.number_format = _MONEY_FMT
        row_num += 1

    # Bicomponentes inyectados en detalle
    for bico in match_result.bicomponentes_inyectados:
        if not bico.codigo_encontrado:
            continue
        values = [
            bico.codigo_encontrado, bico.descripcion_real, "",
            "", bico.cantidad_sugerida, bico.precio_unitario,
            0, bico.precio_unitario * bico.cantidad_sugerida,
            bico.stock_disponible,
            "SI" if bico.disponible else "NO", "auto_inyectado",
            "", f"{bico.tipo.upper()} para {bico.para_producto}",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_det.cell(row=row_num, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(name="Calibri", italic=True, color="6B7280")
            if col_idx in (6, 8):
                cell.number_format = _MONEY_FMT
        row_num += 1

    # Productos fallidos
    for fallido in match_result.productos_fallidos:
        ws_det.cell(row=row_num, column=1, value="???").border = _THIN_BORDER
        ws_det.cell(row=row_num, column=2, value=fallido.producto_solicitado).border = _THIN_BORDER
        cell_obs = ws_det.cell(row=row_num, column=13, value=f"NO ENCONTRADO: {fallido.razon}")
        cell_obs.border = _THIN_BORDER
        cell_obs.font = Font(name="Calibri", bold=True, color="DC2626")
        row_num += 1

    # Ajustar anchos detalle
    widths = [18, 40, 16, 14, 10, 16, 12, 16, 10, 10, 14, 10, 30]
    for i, w in enumerate(widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 3: Notas (si hay) ──
    if notas or match_result.descuentos_aplicados:
        ws_notas = wb.create_sheet("Notas")
        ws_notas.cell(row=1, column=1, value="Observaciones del Pedido").font = Font(bold=True, size=12)
        if notas:
            ws_notas.cell(row=3, column=1, value=notas)
        if match_result.descuentos_aplicados:
            ws_notas.cell(row=5, column=1, value="Descuentos aplicados:").font = Font(bold=True)
            for i, d in enumerate(match_result.descuentos_aplicados, 6):
                ws_notas.cell(row=i, column=1, value=f"  Marca: {d.get('marca', 'General')} — {d.get('porcentaje', 0)}%")
        ws_notas.column_dimensions["A"].width = 60

    # ── Escribir a bytes ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), icg_rows


def build_nombre_archivo_pedido(
    cliente: str = "",
    tienda_codigo: str = "",
    tienda_nombre: str = "",
    pedido_id: int | str = 0,
) -> str:
    """Genera nombre de archivo sanitizado para el Excel del pedido."""
    fecha = datetime.now().strftime("%Y-%m-%d")
    cliente_safe = _safe_filename(cliente, "SinCliente")
    tienda_safe = _safe_filename(tienda_nombre or tienda_codigo, "SinTienda")
    return f"pedido_{cliente_safe}_{fecha}_{tienda_safe}_{pedido_id}.xlsx"
